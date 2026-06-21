# -*- coding: utf-8 -*-
"""
BOM 展开查询工具 v10 - 树形展开 + 数量计算 + 外采物料ABC分类安全库存
功能：
  ✓ 正确列索引：母件col1 / 子件col4 / 品名col5 / 规格col6 / 用量col7
  ✓ 选项卡1「🌲 BOM 结构展开」：递归展开，支持 + 懒加载
  ✓ 选项卡2「📦 数量计算」：读取选项卡1的展开树，按上层用量递归计算
  ✓ 选项卡3「📊 外采物料管理」：上传排产计划+外采清单，ABC分类+安全库存
  ✓ 上传 BOM 文件 / 模糊搜索 / 导出 Excel
"""
import os, gzip, pickle, csv
import calendar
import hashlib
import hmac
import logging
import math
import re
import shutil
import subprocess
import sys
import tempfile
import threading
from collections import defaultdict, deque
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd


U9_DB_CONN_ENV = 'MRP_U9_CONN'
IS_WINDOWS = sys.platform.startswith('win')
IS_MACOS = sys.platform == 'darwin'
APP_DIR = Path(__file__).resolve().parent
CONFIG_DIR = Path(os.environ.get('APPDATA', Path.home() / '.config')) / 'bom_tool'
U9_DB_CONN_FILE = str(CONFIG_DIR / 'u9_db_conn.txt')
UI_FONT_FAMILY = 'PingFang SC' if IS_MACOS else 'Microsoft YaHei'
UI_FONT_FAMILY_ALT = 'PingFang SC' if IS_MACOS else 'Microsoft YaHei UI'
MONO_FONT_FAMILY = 'Menlo' if IS_MACOS else 'Consolas'


def bundled_resource_path(*parts: str) -> Path:
    base = Path(getattr(sys, '_MEIPASS', APP_DIR))
    return base.joinpath(*parts)


def open_local_path(path: str | os.PathLike) -> None:
    target = str(Path(path).expanduser().resolve())
    if IS_WINDOWS:
        os.startfile(target)
    elif IS_MACOS:
        subprocess.run(['open', target], check=True)
    else:
        subprocess.run(['xdg-open', target], check=True)


def fmt_qty(v):
    """把数量格式化成易读字符串：整数就不带小数点，小数去掉尾零，不用科学计数法"""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if v == int(v) and abs(v) < 1e15:
        return str(int(v))
    # 最多6位小数，去掉尾零
    s = f'{v:.6f}'.rstrip('0').rstrip('.')
    return s


def normalize_sheet_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for column in df.columns:
        if isinstance(column, str):
            renamed[column] = column.strip().replace('\n', '').replace('\r', '')
    return df.rename(columns=renamed)


def find_first_matching_column(df: pd.DataFrame, candidates, *, contains=False, required=True):
    columns = [str(column).strip() for column in df.columns]
    for candidate in candidates:
        if contains:
            for original, current in zip(df.columns, columns):
                if candidate in current:
                    return str(original)
        else:
            for original, current in zip(df.columns, columns):
                if current == candidate:
                    return str(original)
    if required:
        raise ValueError(f'缺少字段: {", ".join(candidates)}')
    return None


def normalize_material_code(value) -> str:
    if value is None or pd.isna(value):
        return ''
    text = str(value).strip()
    match = re.search(r'[A-Za-z0-9]+-\d+', text)
    return match.group(0) if match else text


def latest_existing_file(folders, patterns, *, exclude_keywords=None):
    exclude_keywords = [str(item).lower() for item in (exclude_keywords or [])]
    candidates = []
    for folder in folders:
        if not folder:
            continue
        base = Path(folder).expanduser()
        if not base.exists():
            continue
        for pattern in patterns:
            for path in base.glob(pattern):
                if not path.is_file():
                    continue
                name_lower = path.name.lower()
                if path.name.startswith('~$'):
                    continue
                if any(keyword in name_lower for keyword in exclude_keywords):
                    continue
                candidates.append(path)
    if not candidates:
        return ''
    return str(max(candidates, key=lambda item: item.stat().st_mtime))


def default_balance_template_path():
    candidates = [
        bundled_resource_path('assets', '静态平衡表模板.xlsx'),
        APP_DIR / 'assets' / '静态平衡表模板.xlsx',
        Path(WORKSPACE) / '静态平衡表模板.xlsx' if 'WORKSPACE' in globals() else None,
    ]
    for path in candidates:
        if path and path.exists():
            return str(path)
    return ''


MATERIAL_CODE_PATTERN = re.compile(r'^[A-Za-z0-9]+(?:-[A-Za-z0-9]+)?$')


def maybe_material_code(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if MATERIAL_CODE_PATTERN.fullmatch(text):
        return text
    match = re.search(r'[A-Za-z0-9]+-\d+', text)
    if match:
        return match.group(0)
    if text.isdigit():
        return text
    return None

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from rapidfuzz import fuzz, process
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

# ── 配置 ──────────────────────────────────────────────────────
WORKSPACE   = os.path.dirname(os.path.abspath(__file__))


def _candidate_balance_src_dirs() -> list[str]:
    """尝试多处定位 静态平衡表/src：支持源码目录结构与 PyInstaller 打包后的临时解压目录。"""
    candidates = [
        os.path.join(WORKSPACE, '..', '静态平衡表', 'src'),
        os.path.join(WORKSPACE, '静态平衡表', 'src'),
    ]
    meipass = getattr(sys, '_MEIPASS', None)
    if meipass:
        candidates.extend([
            os.path.join(meipass, '静态平衡表', 'src'),
            os.path.join(meipass, 'src'),
            meipass,
        ])
    return [os.path.abspath(path) for path in candidates]


BALANCE_SRC_DIR = next(
    (path for path in _candidate_balance_src_dirs() if os.path.isdir(path)),
    None,
)
if BALANCE_SRC_DIR and BALANCE_SRC_DIR not in sys.path:
    sys.path.insert(0, BALANCE_SRC_DIR)

BALANCE_PIPELINE_IMPORT_ERROR = None
try:
    from mrp_balance_tool.pipeline import (
        analyze_arrival_status_records as run_arrival_status_analysis,
        analyze_external_purchase_readiness as run_purchase_readiness_analysis,
        fill_work_order_shortage_replies as run_work_order_reply_fill,
        read_substitute_rules as read_balance_substitute_rules,
        read_workbook_tables as read_balance_workbook_tables,
        run_pipeline as run_balance_pipeline,
        validate_input_workbook as run_input_precheck,
    )
except Exception as exc:
    run_arrival_status_analysis = None
    run_work_order_reply_fill = None
    read_balance_substitute_rules = None
    read_balance_workbook_tables = None
    run_balance_pipeline = None
    run_purchase_readiness_analysis = None
    run_input_precheck = None
    BALANCE_PIPELINE_IMPORT_ERROR = str(exc)

def _resolve_cache_file() -> str:
    """缓存文件放到用户可写目录（避免 PyInstaller 打包后的只读 _MEIxxx 目录）。"""
    # 首选：APPDATA\BOMTool（Windows）或 ~/.config/bom_tool（类 Unix）
    base = os.environ.get('APPDATA') or os.path.join(os.path.expanduser('~'), '.config')
    target_dir = os.path.join(base, 'bom_tool')
    try:
        os.makedirs(target_dir, exist_ok=True)
        return os.path.join(target_dir, '.bom_cache.pkl')
    except OSError:
        return os.path.join(tempfile.gettempdir(), 'bom_tool_cache.pkl')


CACHE_FILE  = _resolve_cache_file()
CACHE_MAGIC = b'BOMCACHEv2'
# 简单完整性校验的本地密钥：防止缓存被替换后直接反序列化任意代码。
# 这不是加密强度，只是阻挡非本机产生的 pickle 载荷。
_CACHE_KEY = hashlib.sha256(
    (os.environ.get('COMPUTERNAME', '') + os.environ.get('USERNAME', '') + 'bom_tool').encode('utf-8')
).digest()
MAX_DEPTH   = 50
MAX_ROWS    = 100000

HEADERS_TREE = ['层级', '母器件料号', '母器件品名', '母器件规格',
                '子器件料号', '子器件品名', '子器件规格', '用量']
BOM_SUPPLY_HEADERS = [
    '查询母件', '查询母件名称', '查询母件规格', '输入数量', '层级', '母件料号', '母件名称', '母件规格',
    '子件料号', '料品名称', '规格', '单套用量', '累计用量', '需求数量',
    '库存', '未转单数量', '未到货数量', '采购周期', '供应商', '采购', '缺口', '是否末级', '路径'
]
BOM_SUPPLY_WIDTHS = [145, 220, 180, 90, 60, 145, 220, 180, 145, 220, 180, 90, 90, 90, 90, 100, 100, 90, 180, 90, 90, 80, 520]

ABC_HEADERS = ['分类', '物料编码', '物料分类', '物料名称', '规格型号', '供应商',
               '交期(天)', 'SPQ', 'MOQ',
               '月均用量', '年用量',
               '安全库存', '再订货点(ROP)', '当前库存', '未清PO', '未转PR', '库存位置',
               '是否触发采购', '理论采购量', '建议采购量', '采购建议说明',
               '安全库存覆盖天数', '安全库存判断',
               '使用项目(点击展开)']
ABC_WIDTHS  = [50, 130, 90, 160, 140, 100,
               60, 50, 50,
               75, 75,
               65, 75, 80, 80, 80, 90, 95, 90, 90, 260, 75, 105,
               200]
COL_WIDTHS   = [70, 145, 210, 160, 145, 210, 160, 60]
HEADERS_CALC = ['层级', '料号', '品名', '规格', '路径数', '汇总用量']
COL_WIDTHS_C = [60, 145, 210, 160, 60, 90]
HEADERS_SUM = ['最浅层级', '料号', '品名', '规格', '汇总用量', '来源数量']
COL_WIDTHS_S = [70, 145, 210, 160, 100, 80]
HEADERS_DIFF_GRID = ['差异类型', 'A物料编码', 'A品名', 'A规格', 'A总用量',
                     'B物料编码', 'B品名', 'B规格', 'B总用量', '差值(A-B)']
COL_WIDTHS_D_GRID = [90, 135, 180, 170, 85, 135, 180, 170, 85, 90]
BOM_DOC_COMPARE_HEADERS = [
    '来源BOM', '母件料号', '母件名称', '母件型号', '差异类型', '物料编码',
    'DOC名称', 'BOM名称', 'DOC规格', 'BOM规格', 'DOC用量', 'BOM用量',
    '差值(DOC-BOM)', '说明'
]
BOM_DOC_COMPARE_WIDTHS = [240, 145, 150, 180, 120, 145, 180, 180, 160, 160, 95, 95, 110, 320]
BALANCE_SUGGEST_HEADERS = ['客户', '母件料号', '母件品名', '母件规格', '上线日期', '上线数量', '直接上层数量', '直接上层摘要', '顶层来源数量', '顶层来源摘要', '说明']
COL_WIDTHS_BALANCE = [110, 150, 220, 200, 110, 110, 110, 280, 110, 280, 140]
READINESS_ROOT_HEADERS = ['母料号', '料品名称', '输入数量', '外购物料数', '问题物料数', '未识别物料数', '外购齐套日期', '结论']
READINESS_ROOT_WIDTHS = [150, 220, 100, 100, 100, 110, 120, 160]
READINESS_ISSUE_HEADERS = ['母料号', '输入数量', '问题料号', '本母件需求', '料品名称', '规格', '上层物料编码', '供应商', '采购', '当前可用', '当前缺口', '当前可用缺口', '替代料清单', '采购答复累计', '最早齐套日期', '问题原因']
READINESS_ISSUE_WIDTHS = [150, 90, 150, 100, 220, 180, 160, 160, 90, 90, 90, 100, 220, 100, 120, 260]
READINESS_MATERIAL_HEADERS = ['料号', '料品名称', '规格', '供应商', '采购', '总需求', '当前库存', '未清PO', '未转', '替代料库存', '替代料未清PO', '替代料未转', '当前库存+替代库存', '替代料清单', '当前可用', '当前缺口', '当前可用缺口', '采购答复累计', '最早齐套日期', '状态', '来源母料号', '上层物料编码', '问题原因']
READINESS_MATERIAL_WIDTHS = [150, 220, 180, 160, 90, 90, 90, 90, 90, 90, 110, 100, 130, 220, 90, 90, 100, 100, 120, 110, 220, 180, 260]
READINESS_UNKNOWN_HEADERS = ['母料号', '输入数量', '料号', '料品名称', '规格', '上层物料编码', '需求数量', '问题原因']
READINESS_UNKNOWN_WIDTHS = [150, 90, 150, 220, 180, 160, 90, 260]
READINESS_RECOMMEND_HEADERS = ['母料号', '料品名称', '规格', '外购物料数', '现货可生产', '库存位置可生产', '做1台缺料数', '做1台缺口', '短板物料', '未识别下层', '结论']
READINESS_RECOMMEND_WIDTHS = [150, 220, 180, 90, 100, 120, 100, 90, 360, 360, 160]
READINESS_PRODUCIBLE_HEADERS = ['母料号', '料品名称', '输入数量', '排产顺序', '排产日期', '日期', '可生产数量', '当前库存可生成', '距离输入缺口', '瓶颈物料', '瓶颈物料名称', '瓶颈单台用量', '瓶颈可用量', '瓶颈原因']
READINESS_PRODUCIBLE_WIDTHS = [150, 220, 90, 80, 110, 110, 100, 120, 110, 240, 260, 110, 110, 360]
READINESS_HORIZONTAL_HEADERS = ['排产顺序', '排产日期', '母料号', '料品名称', '输入数量', '料号', '物料名称', '单台用量', '本行需求', '库存分配前', '库存扣减', '库存缺口', '库存分配后', '当前可用分配前', '当前可用扣减', '当前可用缺口', '当前可用分配后', '替代料清单', 'BOM差异标识', '共用母料数', 'BOM差异说明']
READINESS_HORIZONTAL_WIDTHS = [80, 110, 150, 220, 90, 150, 220, 90, 90, 100, 90, 90, 100, 120, 110, 110, 120, 220, 130, 90, 360]
READINESS_DIFF_CAPACITY_HEADERS = ['排产顺序', '排产日期', '母料号', '料品名称', '输入数量', '差异物料数', '通用物料数', '通用物料清单', '当前库存差异可生产', '当前库存差异缺口', '当前可用差异可生产', '当前可用差异缺口', '瓶颈差异物料', '瓶颈物料名称', '瓶颈单台差异用量', '瓶颈库存可用', '瓶颈当前可用', 'BOM差异说明']
READINESS_DIFF_CAPACITY_WIDTHS = [80, 110, 150, 220, 90, 100, 100, 420, 140, 130, 140, 130, 220, 240, 130, 120, 120, 420]
READINESS_ROLLING_MATRIX_HEADERS = ['共用状态', '物料编码', '品名', '规格', '出现BOM数']
READINESS_ROLLING_MATRIX_WIDTHS = [120, 150, 220, 180, 90]
READINESS_COMMON_HEADERS = ['排名', '物料编码', '物料名称', '规格', '采购', '供应商', '使用母料数', '使用母料清单', '总需求', '当前库存', '未清PO', '未转PR', '库存位置', '通用风险等级']
READINESS_COMMON_WIDTHS = [60, 145, 220, 180, 90, 180, 90, 420, 90, 90, 90, 90, 90, 160]
ARRIVAL_STATUS_HEADERS = ['收货状态', '物料编码', '物料', '供应商', '仓库', '到货数量', '可催数量', '当前库存', '未来3天需求', '未来3天缺口', '未来总缺口', '最早需求日期', '是否缺料', '影响项目', '到货日期', '批次入库时间', '备注', '跟催建议']
ARRIVAL_STATUS_WIDTHS = [90, 145, 220, 180, 100, 90, 90, 90, 95, 95, 95, 120, 80, 260, 110, 150, 180, 220]
ARRIVAL_PURCHASE_PENDING_HEADERS = ['供应商', '物料编码', '物料', '规格', '计划到货日期', '计划数量', '到货数量', '未到货数量', '外部单据编号', '编码', '采购订单', '计划状态', '当前库存', '未来3天需求', '未来3天缺口', '未来总缺口', '最早需求日期', '是否缺料', '备注', '跟催建议']
ARRIVAL_PURCHASE_PENDING_WIDTHS = [180, 145, 220, 160, 110, 90, 90, 90, 140, 140, 140, 90, 90, 95, 95, 95, 120, 80, 180, 260]
MATERIAL_BUY_PREFIX_HEADERS = [
    '判定', '物料编码', '物料名称', '规格', '状态', '用途', '使用部门', '查询区间', '输入/请购数量',
]
MATERIAL_BUY_PREFIX_WIDTHS = [90, 145, 220, 180, 90, 220, 180, 120, 110]
MATERIAL_BUY_SUFFIX_HEADERS = [
    '查询期领用合计', '实际月均领用', '表内月均用量', '当前库存', 'PR未转PO', 'PO欠交',
    '可用供给', '判断需求量', '建议新增购买量', '采购员', '供应商', '处理意见',
]
MATERIAL_BUY_SUFFIX_WIDTHS = [110, 100, 100, 90, 90, 90, 95, 100, 120, 100, 180, 520]

MOTHER_PN_COL  = 1   # 母件料号
CHILD_PN_COL   = 11  # BOM子项_子件料号_料号
CHILD_NAME_COL = 13  # BOM子项_子件料号_自制名称
CHILD_SPEC_COL = 14  # BOM子项_子件料号_规格
CHILD_QTY_COL  = 15  # BOM子项_子件料号_数量


def _is_bom_data_row(row) -> bool:
    min_cols = max(MOTHER_PN_COL, CHILD_PN_COL) + 1
    if not row or len(row) < min_cols:
        return False
    parent_raw = '' if row[MOTHER_PN_COL] is None else str(row[MOTHER_PN_COL]).strip()
    child_raw = '' if row[CHILD_PN_COL] is None else str(row[CHILD_PN_COL]).strip()
    parent = normalize_material_code(row[MOTHER_PN_COL])
    child = normalize_material_code(row[CHILD_PN_COL])
    if not parent or not child:
        return False
    marker_text = f'{parent_raw} {child_raw}'
    header_markers = ('母件料号', '母件编码', '子件料号', 'BOM子项', '料品.料号', 'BOM')
    return not any(marker in marker_text for marker in header_markers)


def _find_bom_data_start(rows) -> int:
    for idx, row in enumerate(rows):
        if _is_bom_data_row(row):
            return idx
    return 1 if len(rows) > 1 else 0


# ── BOM 加载 ──────────────────────────────────────────────────
def load_bom_from_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif ext == '.csv':
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            rows = list(csv.reader(f))
    else:
        raise ValueError(f"不支持的格式: {ext}")
    if not rows:
        raise ValueError("文件为空")
    data_start_idx = _find_bom_data_start(rows)
    bom_index = {}
    for r in rows[data_start_idx:]:
        if len(r) < 16:
            continue
        mpn = str(r[MOTHER_PN_COL]).strip()
        if not mpn:
            continue
        bom_index.setdefault(mpn, []).append(r)
    return bom_index, rows[0]


def load_bom_supply_maps(workbook_path):
    """读取 MRP 计算表里的库存、未转、未到货和采购主数据。"""
    result = {
        'inventory': {},
        'po': {},
        'pr': {},
        'purchase': {},
        'warnings': [],
    }
    if not workbook_path:
        return result

    try:
        workbook = pd.ExcelFile(workbook_path)
    except Exception as exc:
        raise ValueError(f'无法打开 MRP 计算表：{exc}') from exc

    def _read_sheet(sheet_name):
        if sheet_name not in workbook.sheet_names:
            result['warnings'].append(f'缺少工作表: {sheet_name}')
            return None
        try:
            return normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
        except Exception as exc:
            result['warnings'].append(f'{sheet_name} 读取失败: {exc}')
            return None

    def _sum_map(sheet_name, code_candidates, qty_candidates, *, qty_contains=False):
        df = _read_sheet(sheet_name)
        if df is None or df.empty:
            return {}
        try:
            code_col = find_first_matching_column(df, code_candidates, contains=False)
            qty_col = find_first_matching_column(df, qty_candidates, contains=qty_contains)
        except ValueError as exc:
            result['warnings'].append(f'{sheet_name} {exc}')
            return {}
        temp = pd.DataFrame({
            'code': df[code_col].map(normalize_material_code),
            'qty': pd.to_numeric(df[qty_col], errors='coerce').fillna(0),
        })
        temp = temp[temp['code'].astype(str).str.strip() != '']
        if temp.empty:
            return {}
        return temp.groupby('code', dropna=False)['qty'].sum().to_dict()

    result['inventory'] = _sum_map(
        '期初库存',
        ['物料编码', '料号', '物料号', '物料码', '编码'],
        ['库存量', '实时库存', '当前库存', '数量'],
        qty_contains=True,
    )
    result['po'] = _sum_map(
        '在途采购',
        ['料号', '物料编码', '物料号', '物料码', '编码'],
        ['欠交数量', '未交数量', '未清数量', '未到货数量', '数量'],
        qty_contains=True,
    )
    result['pr'] = _sum_map(
        '在途请购',
        ['料号', '物料编码', '物料号', '物料码', '编码'],
        ['未转PO数量', '未转数量', '未转单数量', '数量'],
        qty_contains=True,
    )

    purchase = _read_sheet('采购数据')
    if purchase is not None and not purchase.empty:
        try:
            code_col = find_first_matching_column(
                purchase,
                ['物料号', '料号', '物料编码', '物料码', '编码'],
                required=True,
            )
        except ValueError as exc:
            result['warnings'].append(f'采购数据 {exc}')
            code_col = None
        if code_col:
            supplier_col = find_first_matching_column(purchase, ['供应商', '供应商名称'], required=False)
            buyer_col = find_first_matching_column(purchase, ['采购', '采购员'], required=False)
            name_col = find_first_matching_column(purchase, ['名称', '料品名称', '物料名称', '物料'], required=False)
            spec_col = find_first_matching_column(purchase, ['规格', '规格型号', '型号'], required=False)
            cycle_col = find_first_matching_column(
                purchase,
                ['采购周期', '提前期', '交期', '天数', 'LT', 'L/T'],
                contains=True,
                required=False,
            )
            for record in purchase.to_dict('records'):
                code = normalize_material_code(record.get(code_col))
                if not code:
                    continue
                info = result['purchase'].setdefault(code, {})
                if supplier_col and not info.get('supplier'):
                    info['supplier'] = str(record.get(supplier_col, '') or '').strip()
                if buyer_col and not info.get('buyer'):
                    info['buyer'] = str(record.get(buyer_col, '') or '').strip()
                if name_col and not info.get('name'):
                    info['name'] = str(record.get(name_col, '') or '').strip()
                if spec_col and not info.get('spec'):
                    info['spec'] = str(record.get(spec_col, '') or '').strip()
                if cycle_col and not info.get('cycle'):
                    info['cycle'] = str(record.get(cycle_col, '') or '').strip()
    return result


def build_bom_supply_rows(bom_index, root_code, root_qty, max_depth, supply_maps):
    """按母件展开 BOM，并横向匹配库存、未转、未到货和采购周期。"""
    root_code = str(root_code or '').strip()
    try:
        root_qty = float(root_qty)
    except (TypeError, ValueError):
        root_qty = 1.0
    if root_qty <= 0:
        root_qty = 1.0
    try:
        max_depth = int(max_depth)
    except (TypeError, ValueError):
        max_depth = MAX_DEPTH
    max_depth = max(1, min(max_depth, MAX_DEPTH))

    normalized_parent_lookup = {}
    for parent in bom_index.keys():
        normalized_parent_lookup.setdefault(normalize_material_code(parent), parent)

    root_key = root_code if root_code in bom_index else normalized_parent_lookup.get(normalize_material_code(root_code), root_code)
    if root_key not in bom_index:
        raise ValueError(f'BOM 中找不到母件: {root_code}')

    material_info = {}
    for parent, rows in bom_index.items():
        parent_norm = normalize_material_code(parent)
        if not rows:
            continue
        first = rows[0]
        parent_name = str(first[2]).strip() if len(first) > 2 and first[2] else ''
        parent_spec = str(first[3]).strip() if len(first) > 3 and first[3] else ''
        if parent_norm:
            material_info.setdefault(parent_norm, {})
            if parent_name:
                material_info[parent_norm].setdefault('name', parent_name)
            if parent_spec:
                material_info[parent_norm].setdefault('spec', parent_spec)
        for row in rows:
            if len(row) <= CHILD_PN_COL:
                continue
            child_code = normalize_material_code(row[CHILD_PN_COL])
            if not child_code:
                continue
            child_name = str(row[CHILD_NAME_COL]).strip() if len(row) > CHILD_NAME_COL and row[CHILD_NAME_COL] else ''
            child_spec = str(row[CHILD_SPEC_COL]).strip() if len(row) > CHILD_SPEC_COL and row[CHILD_SPEC_COL] else ''
            material_info.setdefault(child_code, {})
            if child_name:
                material_info[child_code].setdefault('name', child_name)
            if child_spec:
                material_info[child_code].setdefault('spec', child_spec)

    for code, info in supply_maps.get('purchase', {}).items():
        material_info.setdefault(code, {})
        for key in ('name', 'spec'):
            if info.get(key):
                material_info[code].setdefault(key, info.get(key, ''))

    rows = []
    truncated = False

    def _qty(value):
        try:
            if value in (None, ''):
                return 1.0
            number = float(value)
            return number if number > 0 else 0.0
        except (TypeError, ValueError):
            return 1.0

    def _fmt_number(value):
        try:
            return fmt_qty(value)
        except Exception:
            return str(value or '')

    def _walk(parent_key, parent_code, level, cumulative_qty, trail, path_codes):
        nonlocal truncated
        if truncated or level >= max_depth:
            return
        for bom_row in bom_index.get(parent_key, []):
            if len(rows) >= MAX_ROWS:
                truncated = True
                return
            if len(bom_row) <= max(CHILD_PN_COL, CHILD_QTY_COL):
                continue
            child_code = normalize_material_code(bom_row[CHILD_PN_COL])
            if not child_code:
                continue
            unit_qty = _qty(bom_row[CHILD_QTY_COL])
            child_cumulative = cumulative_qty * unit_qty
            demand_qty = child_cumulative * root_qty
            child_key = normalized_parent_lookup.get(child_code)
            is_cycle = child_code in trail
            has_children = bool(child_key and not is_cycle and level + 1 < max_depth)
            purchase_info = supply_maps.get('purchase', {}).get(child_code, {})
            stock = float(supply_maps.get('inventory', {}).get(child_code, 0) or 0)
            pr_qty = float(supply_maps.get('pr', {}).get(child_code, 0) or 0)
            po_qty = float(supply_maps.get('po', {}).get(child_code, 0) or 0)
            shortage = max(demand_qty - stock - pr_qty - po_qty, 0)
            root_info = material_info.get(normalize_material_code(root_key), {})
            parent_info = material_info.get(normalize_material_code(parent_code), {})
            info = material_info.get(child_code, {})
            child_path = [*path_codes, child_code]
            rows.append({
                '查询母件': normalize_material_code(root_key),
                '查询母件名称': root_info.get('name', ''),
                '查询母件规格': root_info.get('spec', ''),
                '输入数量': _fmt_number(root_qty),
                '层级': level + 1,
                '母件料号': normalize_material_code(parent_code),
                '母件名称': parent_info.get('name', ''),
                '母件规格': parent_info.get('spec', ''),
                '子件料号': child_code,
                '料品名称': info.get('name', ''),
                '规格': info.get('spec', ''),
                '单套用量': _fmt_number(unit_qty),
                '累计用量': _fmt_number(child_cumulative),
                '需求数量': _fmt_number(demand_qty),
                '库存': _fmt_number(stock),
                '未转单数量': _fmt_number(pr_qty),
                '未到货数量': _fmt_number(po_qty),
                '采购周期': purchase_info.get('cycle', ''),
                '供应商': purchase_info.get('supplier', ''),
                '采购': purchase_info.get('buyer', ''),
                '缺口': _fmt_number(shortage),
                '是否末级': '否' if has_children else '是',
                '路径': ' > '.join(child_path),
            })
            if has_children:
                _walk(child_key, child_code, level + 1, child_cumulative, trail | {child_code}, child_path)

    _walk(root_key, normalize_material_code(root_key), 1, 1.0, {normalize_material_code(root_key)}, [normalize_material_code(root_key)])
    return rows, truncated


def _find_approved_status(status_vals, df, col):
    for status in status_vals:
        text = str(status).strip()
        if text == '已核准':
            return status
    counts = {status: int((df[col] == status).sum()) for status in status_vals}
    return max(counts, key=counts.get) if counts else ''


def load_mrp_need_pool_keys(path: str) -> set[tuple[str, str]]:
    df = pd.read_excel(path, header=None)
    if df.shape[1] <= 35:
        raise ValueError('需求池报表列数不足，无法读取来源单据和物料编码')
    data = df.iloc[2:, :].copy()
    src_series = data.iloc[:, 35].astype(str).str.strip()
    mat_series = data.iloc[:, 3].astype(str).str.strip()
    keys = set()
    for src, mat in zip(src_series, mat_series):
        if src and src.lower() != 'nan' and mat and mat.lower() != 'nan':
            keys.add((src, mat))
    return keys


def merge_mrp_pr_files(pr_files: list[str], need_pool_path: str) -> tuple[pd.DataFrame, dict]:
    frames = []
    for file_path in pr_files:
        if file_path and os.path.exists(file_path):
            raw = pd.read_excel(file_path, header=None)
            if raw.shape[1] < 17:
                raise ValueError(f'PR 文件列数不足: {os.path.basename(file_path)}')
            frames.append(raw.iloc[1:, :].copy())
    if not frames:
        return pd.DataFrame(columns=['料号', '数量:未转PO数量']), {
            'total': 0, 'approved': 0, 'deleted': 0, 'deleted_qty': 0, 'items': 0, 'sum': 0
        }
    pr_all = pd.concat(frames, ignore_index=True)
    pr_all.columns = range(pr_all.shape[1])
    for col in [2, 4, 16]:
        pr_all[col] = pr_all[col].astype(str).str.strip()
    pr_all[11] = pd.to_numeric(pr_all[11], errors='coerce').fillna(0)

    approved = _find_approved_status(pr_all[16].unique(), pr_all, 16)
    need_keys = load_mrp_need_pool_keys(need_pool_path)
    in_need = pr_all.apply(lambda row: (str(row[2]).strip(), str(row[4]).strip()) in need_keys, axis=1)
    delete_mask = (pr_all[16] == approved) & (~in_need)
    deleted_qty = float(pr_all.loc[delete_mask, 11].sum())
    pr_filt = pr_all.loc[~delete_mask].copy()

    agg = pr_filt.groupby(4, dropna=False)[11].sum().reset_index()
    agg.columns = ['料号', '数量:未转PO数量']
    agg['数量:未转PO数量'] = agg['数量:未转PO数量'].round(2)
    agg = agg[agg['料号'].astype(str).str.strip().ne('')]
    stats = {
        'total': int(len(pr_all)),
        'approved': int((pr_all[16] == approved).sum()),
        'deleted': int(delete_mask.sum()),
        'deleted_qty': deleted_qty,
        'items': int(len(agg)),
        'sum': float(agg['数量:未转PO数量'].sum()) if not agg.empty else 0.0,
    }
    return agg, stats


def merge_mrp_po_files(po_files: list[str]) -> tuple[pd.DataFrame, dict]:
    frames = []
    for file_path in po_files:
        if file_path and os.path.exists(file_path):
            raw = pd.read_excel(file_path, header=None)
            if raw.shape[1] < 19:
                raise ValueError(f'PO 文件列数不足: {os.path.basename(file_path)}')
            frames.append(raw.iloc[2:, :].copy())
    if not frames:
        return pd.DataFrame(columns=['料号', '数量:欠交数量']), {
            'total': 0, 'approved': 0, 'deleted': 0, 'items': 0, 'sum': 0
        }
    po_all = pd.concat(frames, ignore_index=True)
    po_all.columns = range(po_all.shape[1])
    for col in [1, 3, 8]:
        po_all[col] = po_all[col].astype(str).str.strip()
    po_all[18] = pd.to_numeric(po_all[18], errors='coerce').fillna(0)

    approved = _find_approved_status(po_all[1].unique(), po_all, 1)
    agg = po_all.groupby(8, dropna=False)[18].sum().reset_index()
    agg.columns = ['料号', '数量:欠交数量']
    agg['数量:欠交数量'] = agg['数量:欠交数量'].round(2)
    agg = agg[agg['料号'].astype(str).str.strip().ne('')]
    stats = {
        'total': int(len(po_all)),
        'approved': int((po_all[1] == approved).sum()),
        'deleted': 0,
        'items': int(len(agg)),
        'sum': float(agg['数量:欠交数量'].sum()) if not agg.empty else 0.0,
    }
    return agg, stats


def read_u9_connection_string() -> str:
    conn = os.environ.get(U9_DB_CONN_ENV, '').strip()
    if conn:
        return conn

    candidates = [
        Path(U9_DB_CONN_FILE),
        Path(sys.executable).with_name('u9_db_conn.txt'),
        Path(__file__).with_name('u9_db_conn.txt'),
    ]
    for path in candidates:
        try:
            if not path.exists():
                continue
            for line in path.read_text(encoding='utf-8-sig', errors='replace').splitlines():
                text = line.strip()
                if text and not text.startswith('#'):
                    return text
        except OSError:
            continue
    raise ValueError(
        f'未配置数据库连接。请设置环境变量 {U9_DB_CONN_ENV}，'
        f'或在 {U9_DB_CONN_FILE} 写入一行 SQL Server 连接串；也可以取消勾选“PR/PO使用数据库”改用文件夹。'
    )


def query_u9_sql_to_dataframe(sql: str, *, timeout: int = 180) -> pd.DataFrame:
    conn_str = read_u9_connection_string()
    if not IS_WINDOWS:
        try:
            import pyodbc
        except ImportError as exc:
            raise ValueError(
                'macOS 数据库模式需要安装 Microsoft ODBC Driver 18 for SQL Server 和 pyodbc；'
                '也可以取消勾选“PR/PO使用数据库”改用文件上传。'
            ) from exc
        normalized_conn = conn_str
        if 'DRIVER=' not in conn_str.upper():
            parsed = {}
            for item in conn_str.split(';'):
                if '=' not in item:
                    continue
                key, value = item.split('=', 1)
                parsed[key.strip().lower()] = value.strip()
            server = parsed.get('server') or parsed.get('data source') or ''
            database = parsed.get('database') or parsed.get('initial catalog') or ''
            user = parsed.get('user id') or parsed.get('userid') or parsed.get('uid') or ''
            password = parsed.get('password') or parsed.get('pwd') or ''
            trust = parsed.get('trustservercertificate', 'yes')
            normalized_conn = ';'.join(
                [
                    'DRIVER={ODBC Driver 18 for SQL Server}',
                    f'SERVER={server}',
                    f'DATABASE={database}',
                    f'UID={user}',
                    f'PWD={password}',
                    f'TrustServerCertificate={trust}',
                ]
            )
        try:
            with pyodbc.connect(normalized_conn, timeout=timeout) as connection:
                cursor = connection.cursor()
                cursor.timeout = timeout
                cursor.execute(sql)
                columns = [item[0] for item in cursor.description or []]
                rows = cursor.fetchall()
                return pd.DataFrame.from_records(rows, columns=columns)
        except Exception as exc:
            raise ValueError(f'macOS 数据库查询失败: {exc}') from exc

    script_path = ''
    csv_path = ''
    try:
        with tempfile.NamedTemporaryFile('w', delete=False, suffix='.ps1', encoding='utf-8-sig') as script_file:
            script_path = script_file.name
            script_file.write(f"""$ErrorActionPreference = 'Stop'
$connStr = $env:MRP_U9_CONN_RUNTIME
$outPath = $env:MRP_U9_OUT_CSV
$query = @'
{sql}
'@
$conn = New-Object System.Data.SqlClient.SqlConnection
$conn.ConnectionString = $connStr
$cmd = $conn.CreateCommand()
$cmd.CommandTimeout = {timeout}
$cmd.CommandText = $query
$adapter = New-Object System.Data.SqlClient.SqlDataAdapter $cmd
$table = New-Object System.Data.DataTable
[void]$adapter.Fill($table)
$conn.Close()
$table | Export-Csv -LiteralPath $outPath -NoTypeInformation -Encoding UTF8
""")
        csv_fd, csv_path = tempfile.mkstemp(suffix='.csv')
        os.close(csv_fd)
        env = os.environ.copy()
        env['MRP_U9_CONN_RUNTIME'] = conn_str
        env['MRP_U9_OUT_CSV'] = csv_path
        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-File', script_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=timeout + 30,
            env=env,
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or '').strip()
            raise ValueError(f'数据库查询失败: {detail}')
        if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
            return pd.DataFrame()
        return pd.read_csv(csv_path, dtype=object, encoding='utf-8-sig')
    finally:
        for path in (script_path, csv_path):
            if path:
                try:
                    os.remove(path)
                except OSError:
                    pass


U9_PR_OPEN_SQL = """
SELECT
    pp.DocNo AS 请购单号,
    ppl.ItemInfo_ItemCode AS 物料编码,
    ppl.ItemInfo_ItemName AS 物料名称,
    ppl.Status AS 行状态,
    ppl.ApprovedQtyReqUOM AS 核准数量,
    ppl.TotalToPOQtyTU AS 转PO数量,
    ppl.ApprovedQtyReqUOM - ppl.TotalToPOQtyTU AS 未转PO数量
FROM PR_PR pp
LEFT JOIN PR_PRLine ppl ON pp.id = ppl.pr
WHERE ppl.Status IN ('1','2')
  AND pp.PRDocType IN ('1002409080122637','1002409090110038')
  AND pp.Org = '1002409010000051'
  AND ppl.ApprovedQtyReqUOM - ppl.TotalToPOQtyTU > 0;
"""


U9_PO_OPEN_SQL = """
WITH open_po AS (
    SELECT
        ppo.DocNo AS 采购单号,
        ppo.DocumentType AS 单据类型,
        ppl.DocLineNo AS 行号,
        ppl.Status AS 行状态,
        ppl.ItemInfo_ItemCode AS 物料编码,
        ppl.ItemInfo_ItemName AS 物料名称,
        ISNULL(ppl.SupplierConfirmQtyTU, 0)
          - ISNULL(ppl.TotalRecievedQtyTU, 0)
          - ISNULL(ppl.TotalRtnDeductQtyTU, 0)
          + ISNULL(ppl.TotalRtnFillQtyTU, 0) AS 欠交数量
    FROM PM_PurchaseOrder ppo
    LEFT JOIN PM_POLine ppl ON ppo.id = ppl.PurchaseOrder
    WHERE ppl.Status NOT IN ('0','3','4')
      AND ppo.Org = '1002409010000051'
      AND ppo.DocNo LIKE 'PO%'
      AND ppo.DocumentType IN ('1002409080122437','1002410180110083','1002410180110100')
)
SELECT *
FROM open_po
WHERE 欠交数量 > 0;
"""


def merge_mrp_pr_database(need_pool_path: str) -> tuple[pd.DataFrame, dict]:
    pr_all = query_u9_sql_to_dataframe(U9_PR_OPEN_SQL)
    if pr_all.empty:
        return pd.DataFrame(columns=['料号', '数量:未转PO数量']), {
            'total': 0, 'approved': 0, 'deleted': 0, 'deleted_qty': 0, 'items': 0, 'sum': 0
        }
    for col in ['请购单号', '物料编码', '行状态']:
        if col in pr_all.columns:
            pr_all[col] = pr_all[col].astype(str).str.strip()
    pr_all['未转PO数量'] = pd.to_numeric(pr_all.get('未转PO数量'), errors='coerce').fillna(0)

    need_keys = load_mrp_need_pool_keys(need_pool_path)
    in_need = pr_all.apply(lambda row: (str(row.get('请购单号', '')).strip(), str(row.get('物料编码', '')).strip()) in need_keys, axis=1)
    approved_mask = pr_all.get('行状态', pd.Series('', index=pr_all.index)).astype(str).str.strip().isin(['2', '已核准'])
    delete_mask = approved_mask & (~in_need)
    deleted_qty = float(pr_all.loc[delete_mask, '未转PO数量'].sum())
    pr_filt = pr_all.loc[~delete_mask].copy()

    agg = pr_filt.groupby('物料编码', dropna=False)['未转PO数量'].sum().reset_index()
    agg.columns = ['料号', '数量:未转PO数量']
    agg['数量:未转PO数量'] = agg['数量:未转PO数量'].round(2)
    agg = agg[agg['料号'].astype(str).str.strip().ne('')]
    stats = {
        'total': int(len(pr_all)),
        'approved': int(approved_mask.sum()),
        'deleted': int(delete_mask.sum()),
        'deleted_qty': deleted_qty,
        'items': int(len(agg)),
        'sum': float(agg['数量:未转PO数量'].sum()) if not agg.empty else 0.0,
    }
    return agg, stats


def merge_mrp_po_database() -> tuple[pd.DataFrame, dict]:
    po_all = query_u9_sql_to_dataframe(U9_PO_OPEN_SQL)
    if po_all.empty:
        return pd.DataFrame(columns=['料号', '数量:欠交数量']), {
            'total': 0, 'approved': 0, 'deleted': 0, 'items': 0, 'sum': 0
        }
    po_all['物料编码'] = po_all.get('物料编码', '').astype(str).str.strip()
    po_all['欠交数量'] = pd.to_numeric(po_all.get('欠交数量'), errors='coerce').fillna(0)
    po_all = po_all[po_all['物料编码'].astype(str).str.strip().ne('')]

    agg = po_all.groupby('物料编码', dropna=False)['欠交数量'].sum().reset_index()
    agg.columns = ['料号', '数量:欠交数量']
    agg['数量:欠交数量'] = agg['数量:欠交数量'].round(2)
    stats = {
        'total': int(len(po_all)),
        'approved': int(len(po_all)),
        'deleted': 0,
        'items': int(len(agg)),
        'sum': float(agg['数量:欠交数量'].sum()) if not agg.empty else 0.0,
    }
    return agg, stats


def merge_mrp_inventory_file(inventory_path: str) -> tuple[pd.DataFrame, dict]:
    raw = pd.read_excel(inventory_path, header=None)
    if raw.shape[1] < 6:
        raise ValueError(f'库存明细列数不足: {os.path.basename(inventory_path)}')
    data = raw.iloc[1:, :].copy()
    data.columns = range(data.shape[1])
    data[1] = data[1].astype(str).str.strip()
    data[2] = data[2].astype(str).str.strip()
    data[5] = pd.to_numeric(data[5], errors='coerce').fillna(0)
    data = data[data[2].astype(str).str.strip().ne('')]
    agg = data.groupby(2, dropna=False).agg(
        仓库名称=pd.NamedAgg(column=1, aggfunc='first'),
        库存量=pd.NamedAgg(column=5, aggfunc='sum'),
    ).reset_index()
    result = pd.DataFrame({
        '仓库编码': '',
        '仓库名称': agg['仓库名称'],
        '物料编码': agg[2],
        'Unnamed: 3': '',
        'Unnamed: 4': '',
        '库存量': agg['库存量'].round(2),
    })
    stats = {
        'total': int(len(data)),
        'items': int(len(result)),
        'sum': float(result['库存量'].sum()) if not result.empty else 0.0,
    }
    return result, stats


def update_mrp_workbook_sheets(
    mrp_path: str,
    pr_df: pd.DataFrame | None,
    po_df: pd.DataFrame | None,
    inventory_df: pd.DataFrame | None,
) -> str:
    if not os.path.exists(mrp_path):
        raise ValueError(f'MRP计算表不存在: {mrp_path}')
    sheets = {}
    with pd.ExcelFile(mrp_path) as workbook:
        for sheet_name in workbook.sheet_names:
            sheets[sheet_name] = pd.read_excel(workbook, sheet_name=sheet_name)
    if pr_df is not None:
        sheets['在途请购'] = pr_df
    if po_df is not None:
        sheets['在途采购'] = po_df
    if inventory_df is not None:
        sheets['期初库存'] = inventory_df
    backup_path = str(Path(mrp_path).with_name(
        f'{Path(mrp_path).stem}_自动更新前备份_{datetime.now().strftime("%Y%m%d_%H%M%S")}{Path(mrp_path).suffix}'
    ))
    shutil.copy2(mrp_path, backup_path)
    with pd.ExcelWriter(mrp_path, engine='openpyxl') as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return backup_path


def find_latest_mrp_source_files(folder_path: str, *, require_pr_po: bool = True) -> dict:
    folder = Path(folder_path)
    if not folder.exists():
        raise ValueError(f'源文件夹不存在: {folder_path}')

    def latest(patterns, count):
        files = []
        for pattern in patterns:
            files.extend(folder.glob(pattern))
        files = [
            path for path in files
            if path.is_file() and not path.name.startswith('~$')
        ]
        unique = {str(path.resolve()).lower(): path for path in files}
        files = sorted(unique.values(), key=lambda path: path.stat().st_mtime, reverse=True)
        return files[:count]

    pr_files = latest(['PR*.xlsx', 'PR*.xls'], 2)
    po_files = latest(['PurchaseOrder*.xlsx', 'PurchaseOrder*.xls'], 2)
    need_files = latest(['需求池报表*.xls', '需求池报表*.xlsx'], 1)
    inventory_files = latest(['库存明细*.xlsx', '库存明细*.xls'], 1)

    missing = []
    if require_pr_po and len(pr_files) < 2:
        missing.append(f'PR 文件不足 2 个（找到 {len(pr_files)} 个）')
    if require_pr_po and len(po_files) < 2:
        missing.append(f'PO 文件不足 2 个（找到 {len(po_files)} 个）')
    if not need_files:
        missing.append('未找到需求池报表')
    if not inventory_files:
        missing.append('未找到库存明细')
    if missing:
        raise ValueError('；'.join(missing))
    return {
        'pr_files': [str(path) for path in pr_files],
        'po_files': [str(path) for path in po_files],
        'need_pool': str(need_files[0]),
        'inventory': str(inventory_files[0]),
    }


def load_root_qtys_from_file(path):
    """读取母料号+数量清单，支持 xlsx/xls/csv/txt；默认前两列为：母料号、数量"""
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif ext == '.csv':
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            rows = list(csv.reader(f))
    elif ext == '.txt':
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            rows = [line.rstrip('\n') for line in f]
    else:
        raise ValueError(f"不支持的格式: {ext}")

    result = []

    if ext == '.txt':
        for idx, line in enumerate(rows, 1):
            raw = str(line).strip()
            if not raw:
                continue
            parts = [p.strip() for p in raw.replace('，', ',').replace('\t', ',').split(',') if p.strip()]
            if len(parts) < 2:
                parts = raw.split()
            if len(parts) < 2:
                continue
            pn = parts[0]
            qty_raw = parts[1]
            try:
                qty = float(str(qty_raw).replace(',', '').strip())
            except ValueError:
                if idx == 1:
                    continue
                raise ValueError(f'第 {idx} 行数量无效: {qty_raw}')
            if not pn or qty <= 0:
                continue
            result.append((pn, qty))
    else:
        for idx, row in enumerate(rows, 1):
            if not row or len(row) < 2:
                continue
            pn = '' if row[0] is None else str(row[0]).strip()
            qty_raw = row[1]
            if not pn:
                continue
            try:
                qty = float(str(qty_raw).replace(',', '').strip())
            except ValueError:
                if idx == 1:
                    continue
                raise ValueError(f'第 {idx} 行数量无效: {qty_raw}')
            if qty <= 0:
                continue
            result.append((pn, qty))

    if not result:
        raise ValueError('文件中未读取到有效数据，请按"母料号, 数量"两列提供')

    merged = {}
    order = []
    for pn, qty in result:
        if pn not in merged:
            merged[pn] = 0.0
            order.append(pn)
        merged[pn] += qty
    return [(pn, merged[pn]) for pn in order]


def load_production_plan_from_file(path, manual_plan_days=''):
    """读取排产计划，返回合并后的母件数量、原始行和排产周期。"""
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in ('.xlsx', '.xls'):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    elif ext == '.csv':
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            rows = list(csv.reader(f))
    else:
        raise ValueError(f'不支持的排产计划格式: {ext}')

    if not rows:
        raise ValueError('排产计划文件为空')

    header_row_idx = -1
    for idx, row in enumerate(rows):
        if row and len(row) >= 2:
            first = str(row[0]).strip() if row[0] else ''
            row_text = str(row)
            if '母件料号' in row_text or '母件编码' in row_text or first == 'NO.':
                header_row_idx = idx
                break
    if header_row_idx < 0:
        raise ValueError('未找到排产计划表头（需包含“母件料号/母件编码”和数量列）')

    headers = [str(h).strip() if h else '' for h in rows[header_row_idx]]
    col_map = {}
    for ci, h in enumerate(headers):
        h2 = h.replace('\n', '').replace('\r', '').strip()
        if '母件料号' in h2 or '母件编码' in h2 or h2 in ('料号', '物料编码', '物料号'):
            col_map['code'] = ci
        elif '上线数量' in h2 or '计划数量' in h2 or h2 == '数量':
            col_map['qty'] = ci
        elif '上线日期' in h2 or '开工日期' in h2 or h2 == '日期':
            col_map['date'] = ci

    if 'code' not in col_map or 'qty' not in col_map:
        raise ValueError(f'未找到母件料号或数量列，表头：{headers[:8]}')

    result = []
    raw_rows = []
    all_dates = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) < 2:
            continue
        pn = normalize_material_code(row[col_map['code']]) if col_map['code'] < len(row) else ''
        qty_raw = row[col_map['qty']] if col_map['qty'] < len(row) else None
        if not pn or '合计' in pn or '总计' in pn:
            continue
        try:
            qty = float(str(qty_raw).replace(',', '').strip()) if qty_raw not in (None, '') else 0
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        d = None
        if 'date' in col_map and col_map['date'] < len(row):
            d_raw = row[col_map['date']]
            if isinstance(d_raw, datetime):
                d = d_raw
            elif isinstance(d_raw, date):
                d = datetime(d_raw.year, d_raw.month, d_raw.day)
            elif isinstance(d_raw, str) and d_raw.strip():
                for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y.%m.%d'):
                    try:
                        d = datetime.strptime(d_raw.strip(), fmt)
                        break
                    except ValueError:
                        pass

        result.append((pn, qty))
        raw_rows.append({'pn': pn, 'qty': qty, 'date': d})
        if d:
            all_dates.append(d)

    if not result:
        raise ValueError('排产计划中未读取到有效母件和数量')

    manual_val = str(manual_plan_days or '').strip()
    if manual_val:
        try:
            plan_days = max(1, int(manual_val))
        except ValueError:
            plan_days = 30
    elif len(all_dates) >= 2:
        plan_days = (max(all_dates) - min(all_dates)).days + 1
    else:
        d = all_dates[0] if all_dates else None
        if d and d.day == 1:
            plan_days = calendar.monthrange(d.year, d.month)[1]
        else:
            plan_days = 30

    merged = {}
    order = []
    for pn, qty in result:
        if pn not in merged:
            merged[pn] = 0.0
            order.append(pn)
        merged[pn] += qty

    return {
        'items': [(pn, merged[pn]) for pn in order],
        'raw_rows': raw_rows,
        'plan_days': plan_days,
        'dates': all_dates,
        'headers': headers,
    }


def clean_bom_by_production_plan(bom_path, plan_path, output_path, plan_items=None, plan_source=None):
    """按排产母件递归保留下层BOM，删除排产链路外的父项行。"""
    if plan_items is None:
        if not plan_path:
            raise ValueError('缺少排产计划来源')
        plan = load_production_plan_from_file(plan_path)
        plan_items = plan['items']
        plan_source = plan_source or os.path.basename(plan_path)
    else:
        normalized_items = []
        for pn, qty in plan_items:
            code = normalize_material_code(pn)
            try:
                qty_value = float(qty or 0)
            except (TypeError, ValueError):
                qty_value = 0
            if code and qty_value > 0:
                normalized_items.append((code, qty_value))
        plan_items = normalized_items
        plan_source = plan_source or 'MRP计算已上传排产计划'
    if not plan_items:
        raise ValueError('排产计划中未读取到有效母件和数量')

    root_qty_map = defaultdict(float)
    root_codes = []
    for pn, qty in plan_items:
        code = normalize_material_code(pn)
        if not code:
            continue
        try:
            qty_value = float(qty or 0)
        except (TypeError, ValueError):
            qty_value = 0
        if code not in root_qty_map:
            root_codes.append(code)
        root_qty_map[code] += qty_value

    ext = os.path.splitext(bom_path)[1].lower()
    source_wb = None
    source_ws = None
    if ext in ('.xlsx', '.xls'):
        source_wb = openpyxl.load_workbook(bom_path, read_only=False, data_only=True)
        source_ws = source_wb[source_wb.sheetnames[0]]
        rows = list(source_ws.iter_rows(values_only=True))
        source_title = source_ws.title
    elif ext == '.csv':
        source_title = 'BOM'
        with open(bom_path, 'r', encoding='utf-8-sig', errors='replace', newline='') as f:
            rows = list(csv.reader(f))
    else:
        raise ValueError(f'不支持的 BOM 文件格式: {ext}')

    if not rows:
        if source_wb:
            source_wb.close()
        raise ValueError('BOM 文件为空')

    data_start_idx = _find_bom_data_start(rows)
    header_rows = [tuple(row) for row in rows[:data_start_idx]]
    if not header_rows:
        header_rows = [tuple(rows[0])]
        data_start_idx = 1
    data_rows = [tuple(row) for row in rows[data_start_idx:]]

    min_cols = max(MOTHER_PN_COL, CHILD_PN_COL) + 1
    parent_to_children = defaultdict(set)
    parent_to_rows = defaultdict(list)
    skipped_short_rows = 0

    for offset, row in enumerate(data_rows):
        idx = data_start_idx + offset + 1
        if len(row) < min_cols:
            skipped_short_rows += 1
            continue
        parent = normalize_material_code(row[MOTHER_PN_COL])
        child = normalize_material_code(row[CHILD_PN_COL])
        if not parent:
            continue
        parent_to_rows[parent].append((idx, row))
        if child:
            parent_to_children[parent].add(child)

    if not parent_to_rows:
        if source_wb:
            source_wb.close()
        raise ValueError('BOM 文件中未识别到母件料号列，请确认文件结构是否为当前工具支持的 BOM 格式')

    queue = list(root_codes)
    reachable_parents = set()
    traversal_order = []
    while queue:
        parent = queue.pop(0)
        if parent in reachable_parents:
            continue
        reachable_parents.add(parent)
        if parent not in parent_to_rows:
            continue
        traversal_order.append(parent)
        for child in sorted(parent_to_children.get(parent, set())):
            if child in parent_to_rows and child not in reachable_parents:
                queue.append(child)

    matched_roots = [code for code in root_codes if code in parent_to_rows]
    missing_roots = [code for code in root_codes if code not in parent_to_rows]
    missing_root_rows = [
        {
            'code': code,
            'qty': root_qty_map.get(code, 0.0),
            'reason': '排产计划中存在，但BOM母件料号中没有对应父项记录',
        }
        for code in missing_roots
    ]
    kept_rows = []
    kept_row_numbers = set()
    for parent in traversal_order:
        for row_number, row in parent_to_rows.get(parent, []):
            kept_rows.append(row)
            kept_row_numbers.add(row_number)

    total_rows = len(data_rows)
    output_ext = os.path.splitext(output_path)[1].lower()
    if output_ext == '.csv':
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(header_rows)
            writer.writerows(kept_rows)
        if source_wb:
            source_wb.close()
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = (source_title or 'BOM')[:31]
        for row in header_rows:
            ws_out.append(list(row))
        for row in kept_rows:
            ws_out.append(list(row))

        if source_ws is not None:
            for ri in range(1, len(header_rows) + 1):
                ws_out.row_dimensions[ri].height = source_ws.row_dimensions[ri].height
                for ci in range(1, source_ws.max_column + 1):
                    src_cell = source_ws.cell(row=ri, column=ci)
                    dst_cell = ws_out.cell(row=ri, column=ci)
                    if src_cell.has_style:
                        dst_cell._style = copy(src_cell._style)
                    if src_cell.number_format:
                        dst_cell.number_format = src_cell.number_format
                    if src_cell.alignment:
                        dst_cell.alignment = copy(src_cell.alignment)
            for ci in range(1, source_ws.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(ci)
                src_dim = source_ws.column_dimensions[col_letter]
                if src_dim.width:
                    ws_out.column_dimensions[col_letter].width = src_dim.width
            for merged_range in source_ws.merged_cells.ranges:
                if merged_range.max_row <= len(header_rows):
                    ws_out.merge_cells(str(merged_range))
        else:
            hdr_font = Font(bold=True, color='000000')
            hdr_fill = PatternFill('solid', fgColor='D9EAF7')
            thin = Side(style='thin', color='D9E2EC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws_out.iter_rows(min_row=1, max_row=len(header_rows)):
                for cell in row:
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for ci in range(1, min(ws_out.max_column, 30) + 1):
                values = [ws_out.cell(row=ri, column=ci).value for ri in range(1, min(ws_out.max_row, 80) + 1)]
                width = min(max(len(str(v)) if v is not None else 0 for v in values) + 2, 28)
                ws_out.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(width, 10)

        if ws_out.max_row >= 1 and ws_out.max_column >= 1:
            filter_row = max(1, len(header_rows))
            ws_out.auto_filter.ref = f"A{filter_row}:{openpyxl.utils.get_column_letter(ws_out.max_column)}{ws_out.max_row}"
            ws_out.freeze_panes = f'A{filter_row + 1}'

        hdr_font = Font(bold=True, color='000000')
        hdr_fill = PatternFill('solid', fgColor='D9EAF7')
        thin = Side(style='thin', color='D9E2EC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        info_ws = wb_out.create_sheet('清洗说明')
        info_rows = [
            ['项目', '结果', '说明'],
            ['原始BOM表头行', len(header_rows), '已从原文件完整保留到输出文件顶部'],
            ['原始BOM数据行', total_rows, '不含表头区'],
            ['保留BOM数据行', len(kept_rows), '排产母件及其递归下层父项对应的BOM行'],
            ['删除BOM数据行', total_rows - len(kept_rows), '父项不在排产展开链路内的BOM行'],
            ['排产母件数', len(root_codes), '排产计划中数量大于0的母件'],
            ['排产来源', plan_source or '', '清洗使用的排产计划来源'],
            ['匹配到BOM的排产母件数', len(matched_roots), '这些母件作为递归展开起点'],
            ['BOM中缺失的排产母件数', len(missing_roots), '排产有但BOM中没有父项记录'],
            ['递归保留父项数', len(traversal_order), '包含排产母件和所有下层仍可继续展开的半成品父项'],
            ['列数不足跳过行数', skipped_short_rows, '这些行列数小于当前工具识别BOM所需列数'],
            ['输出文件', output_path, ''],
        ]
        for row in info_rows:
            info_ws.append(row)
        if missing_roots:
            info_ws.append([])
            info_ws.append(['BOM缺失排产母件', '排产数量', '说明'])
            for item in missing_root_rows[:1000]:
                info_ws.append([item['code'], fmt_qty(item['qty']), item['reason']])
        for cell in info_ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
        info_ws.column_dimensions['A'].width = 24
        info_ws.column_dimensions['B'].width = 50
        info_ws.column_dimensions['C'].width = 72
        info_ws.freeze_panes = 'A2'

        missing_ws = wb_out.create_sheet('排产有但BOM缺失')
        missing_headers = ['母件料号', '排产数量', '问题说明']
        missing_ws.append(missing_headers)
        for item in missing_root_rows:
            missing_ws.append([item['code'], item['qty'], item['reason']])
        for cell in missing_ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in missing_ws.iter_rows(min_row=2, max_row=max(missing_ws.max_row, 2), max_col=3):
            for cell in row:
                cell.border = border
        missing_ws.column_dimensions['A'].width = 22
        missing_ws.column_dimensions['B'].width = 12
        missing_ws.column_dimensions['C'].width = 56
        missing_ws.auto_filter.ref = f"A1:C{max(missing_ws.max_row, 1)}"
        missing_ws.freeze_panes = 'A2'
        wb_out.save(output_path)
        if source_wb:
            source_wb.close()

    return {
        'output_path': output_path,
        'header_rows': len(header_rows),
        'total_rows': total_rows,
        'kept_rows': len(kept_rows),
        'deleted_rows': total_rows - len(kept_rows),
        'root_count': len(root_codes),
        'matched_root_count': len(matched_roots),
        'missing_roots': missing_roots,
        'missing_root_rows': missing_root_rows,
        'reachable_parent_count': len(traversal_order),
        'skipped_short_rows': skipped_short_rows,
        'plan_items': plan_items,
        'plan_source': plan_source or '',
    }

def save_cache(bom_index, cache_path):
    """保存 BOM 缓存，写入 magic 头 + HMAC 校验 + gzip pickle 载荷。"""
    payload = pickle.dumps({'index': bom_index}, protocol=pickle.HIGHEST_PROTOCOL)
    signature = hmac.new(_CACHE_KEY, payload, hashlib.sha256).digest()
    tmp_path = cache_path + '.tmp'
    with open(tmp_path, 'wb') as f:
        f.write(CACHE_MAGIC)
        f.write(signature)
        f.write(gzip.compress(payload))
    os.replace(tmp_path, cache_path)


def load_cache(cache_path):
    """加载 BOM 缓存；校验失败或文件损坏时抛 ValueError 以便调用方忽略缓存。"""
    with open(cache_path, 'rb') as f:
        header = f.read(len(CACHE_MAGIC))
        if header != CACHE_MAGIC:
            raise ValueError('缓存格式无法识别，已忽略')
        signature = f.read(hashlib.sha256().digest_size)
        compressed = f.read()
    try:
        payload = gzip.decompress(compressed)
    except OSError as exc:
        raise ValueError(f'缓存解压失败：{exc}') from exc
    expected = hmac.new(_CACHE_KEY, payload, hashlib.sha256).digest()
    if not hmac.compare_digest(signature, expected):
        raise ValueError('缓存签名不匹配（可能被外部替换），已忽略')
    data = pickle.loads(payload)
    if not isinstance(data, dict) or 'index' not in data or not isinstance(data['index'], dict):
        raise ValueError('缓存内容结构异常，已忽略')
    return data


def show_data_error(parent, title: str, summary: str, detail: str = '', fix_hint: str = ''):
    """展示结构化错误对话框：出了什么 + 建议怎么改 + 技术详情（可复制）。

    适合处理数据/IO 级错误，代替裸 `messagebox.showerror(str(e))`。
    """
    win = tk.Toplevel(parent)
    win.title(title or '出错了')
    win.transient(parent)
    win.grab_set()
    win.resizable(True, True)
    try:
        win.geometry('560x360')
    except tk.TclError:
        pass

    frm = ttk.Frame(win, padding=12)
    frm.pack(fill='both', expand=True)
    frm.columnconfigure(0, weight=1)
    frm.rowconfigure(2, weight=1)

    ttk.Label(frm, text='❌ ' + (summary or '操作失败'),
              font=('Microsoft YaHei UI', 11, 'bold'),
              foreground='#B91C1C', wraplength=520, justify='left').grid(
        row=0, column=0, sticky='ew')
    if fix_hint:
        ttk.Label(frm, text='💡 建议：' + fix_hint, wraplength=520,
                  justify='left', foreground='#1F2937').grid(
            row=1, column=0, sticky='ew', pady=(8, 4))

    detail_frame = ttk.LabelFrame(frm, text='技术详情', padding=6)
    detail_frame.grid(row=2, column=0, sticky='nsew', pady=(8, 0))
    detail_frame.columnconfigure(0, weight=1)
    detail_frame.rowconfigure(0, weight=1)

    text = tk.Text(detail_frame, wrap='word', height=8, font=('Consolas', 9))
    text.insert('1.0', detail or '(无更多信息)')
    text.configure(state='disabled')
    vsb = ttk.Scrollbar(detail_frame, orient='vertical', command=text.yview)
    text.configure(yscrollcommand=vsb.set)
    text.grid(row=0, column=0, sticky='nsew')
    vsb.grid(row=0, column=1, sticky='ns')

    btn_row = ttk.Frame(frm)
    btn_row.grid(row=3, column=0, sticky='e', pady=(10, 0))

    def _copy_detail():
        parent.clipboard_clear()
        parent.clipboard_append(detail or '')

    ttk.Button(btn_row, text='复制详情', command=_copy_detail).pack(side='right', padx=(0, 6))
    ttk.Button(btn_row, text='关闭', command=win.destroy).pack(side='right')
    win.bind('<Escape>', lambda _e: win.destroy())
    win.wait_window(win)


def _suggest_fix_for_exception(exc: BaseException) -> str:
    """根据常见异常类型返回建议的修复方向。"""
    msg = str(exc)
    cls_name = exc.__class__.__name__
    if '缺少工作表' in msg or '缺少' in msg and 'sheet' in msg.lower():
        return '请检查输入 Excel 是否包含所有必需工作表，或在标题行使用标准列名（如"物料编码"、"库存量"）。'
    if '未找到' in msg and ('列' in msg or '表头' in msg):
        return '输入 Excel 的表头可能被改名或缺失某列，请按模板比对列名（含中文全角/半角差异）。'
    if isinstance(exc, PermissionError) or '拒绝访问' in msg or 'Permission denied' in msg:
        return '输出文件可能在 Excel 中被打开，请先关闭后重试；或确认目标目录有写权限。'
    if isinstance(exc, FileNotFoundError) or 'No such file' in msg or '未找到文件' in msg:
        return '文件路径不存在，可能已被移动或删除，请重新选择文件。'
    if 'openpyxl' in msg or 'bad zipfile' in msg.lower() or '不是 zip' in msg:
        return '文件可能不是有效的 .xlsx（比如 WPS 老版 .xls 或内容损坏），请另存为 .xlsx 后重试。'
    if cls_name in ('ValueError', 'TypeError') and '数量' in msg:
        return '输入的数量列存在非数字内容，请删除或修正后重试。'
    return '请根据下方技术详情定位问题，或查看日志获取更多上下文。'


def attach_tooltip(widget, text_source):
    """给 widget 绑定鼠标悬停 tooltip。text_source 可以是字符串或 StringVar 或 callable。"""
    state = {'tip': None}

    def _resolve():
        if callable(text_source):
            return str(text_source() or '')
        if isinstance(text_source, tk.StringVar):
            return text_source.get() or ''
        return str(text_source or '')

    def _show(event):
        _hide()
        text = _resolve()
        if not text:
            return
        tip = tk.Toplevel(widget)
        tip.wm_overrideredirect(True)
        try:
            tip.attributes('-topmost', True)
        except tk.TclError:
            pass
        x = widget.winfo_rootx() + 16
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        tip.wm_geometry(f'+{x}+{y}')
        tk.Label(
            tip,
            text=text,
            justify='left',
            background='#1E293B',
            foreground='#F8FAFC',
            padx=8,
            pady=4,
            font=('Microsoft YaHei', 9),
            wraplength=500,
        ).pack()
        state['tip'] = tip

    def _hide(_event=None):
        tip = state.get('tip')
        if tip is not None:
            try:
                tip.destroy()
            except tk.TclError:
                pass
            state['tip'] = None

    widget.bind('<Enter>', _show)
    widget.bind('<Leave>', _hide)
    widget.bind('<ButtonPress>', _hide)


def make_path_label(parent, path_var, placeholder='未选择'):
    """创建一个"短路径显示 + 完整路径 tooltip"的只读组合。

    path_var: 存放完整路径的 StringVar。返回 (frame, display_var) 供外部继续绑定。
    """
    display_var = tk.StringVar()

    def _truncate(full: str) -> str:
        if not full:
            return placeholder
        # 超过 60 字符只显示盘符 + 省略 + 最后两层
        if len(full) <= 60:
            return full
        parts = full.replace('\\', '/').split('/')
        if len(parts) <= 3:
            return full
        return parts[0] + '/.../' + '/'.join(parts[-2:])

    def _sync(*_args):
        display_var.set(_truncate(path_var.get()))

    path_var.trace_add('write', _sync)
    _sync()

    label = ttk.Label(parent, textvariable=display_var, foreground='#334155',
                      background='#FFFFFF', padding=(6, 4), borderwidth=1,
                      relief='solid', anchor='w')
    attach_tooltip(label, path_var)
    return label, display_var


def enable_treeview_copy(tree, columns):
    """给 Treeview 绑定 Ctrl+C：把选中行复制为 TSV，可直接粘到 Excel。"""
    def _copy_rows(_event=None):
        sels = tree.selection()
        if not sels:
            return 'break'
        lines = []
        for iid in sels:
            row = []
            for col in columns:
                value = tree.set(iid, col)
                if value == '':
                    # 兼容 tree 列（#0）或没有值的情况
                    try:
                        value = tree.item(iid, 'text') or ''
                    except tk.TclError:
                        value = ''
                row.append(str(value))
            lines.append('\t'.join(row))
        payload = '\n'.join(lines)
        try:
            tree.clipboard_clear()
            tree.clipboard_append(payload)
        except tk.TclError:
            pass
        return 'break'

    tree.bind('<Control-c>', _copy_rows)
    tree.bind('<Control-C>', _copy_rows)


def enable_treeview_sort(tree, columns, numeric_columns=None):
    """让 Treeview 的列头支持点击排序。只对扁平层级的 Treeview 生效。

    - columns: 所有列 id 列表
    - numeric_columns: 其中按数字排序的列 id 集合（其它按文本排序）
    """
    numeric_columns = set(numeric_columns or [])
    state = {'col': None, 'reverse': False, 'original_text': {c: tree.heading(c)['text'] for c in columns}}

    def _as_number(value):
        if value is None or value == '':
            return float('-inf')
        try:
            return float(str(value).replace(',', '').replace('%', '').strip())
        except ValueError:
            return float('-inf')

    def _sort(col):
        # 只排顶层行；若 Treeview 当前有子层级（懒加载/分组）则跳过
        top_items = tree.get_children('')
        if any(tree.get_children(iid) for iid in top_items):
            return  # 含有子节点，放弃排序避免打乱结构
        if state['col'] == col:
            state['reverse'] = not state['reverse']
        else:
            state['col'] = col
            state['reverse'] = False
        if col in numeric_columns:
            key = lambda iid: _as_number(tree.set(iid, col))
        else:
            key = lambda iid: str(tree.set(iid, col) or '')
        ordered = sorted(top_items, key=key, reverse=state['reverse'])
        for idx, iid in enumerate(ordered):
            tree.move(iid, '', idx)
        # 给当前排序列的表头加箭头
        for c in columns:
            base = state['original_text'][c]
            if c == col:
                arrow = ' ▼' if state['reverse'] else ' ▲'
                tree.heading(c, text=base + arrow)
            else:
                tree.heading(c, text=base)

    for c in columns:
        tree.heading(c, command=lambda _c=c: _sort(_c))


def fuzzy_search(pn_list, query, limit=50):
    if not query:
        return pn_list[:limit]
    if not HAS_RAPIDFUZZ:
        q = query.lower()
        return sorted(pn_list, key=lambda x: -len(q) if q in x.lower() else 0)[:limit]
    results = process.extract(query, pn_list, scorer=fuzz.QRatio, limit=limit)
    return [r[0] for r in results if r[1] > 25]


# ── GUI ───────────────────────────────────────────────────────
class BOMApp:
    def __init__(self):
        self.bom_index = {}
        self.current_file = None
        self._opened = set()    # 已触发过 TreeviewOpen 的节点
        self._calc_result = None  # {'levels': {}, 'flat': {}, 'mother_pn': str, 'qty': float}
        self._diff_compare_running = False
        self._diff_pending_rows = []
        self._diff_meta_cache = None
        self._bom_doc_compare_rows = []
        self._bom_doc_compare_summary = []
        self._bom_doc_compare_missing_roots = []
        self._bom_doc_compare_source = ''
        self._bom_doc_compare_results = []
        self._bom_edges_cache = None
        self._balance_thread = None
        self._balance_running = False
        self.balance_output_path = ''
        self.balance_reply_source_paths = []
        self._readiness_thread = None
        self._readiness_running = False
        self.readiness_reply_source_paths = []
        self._arrival_thread = None
        self._arrival_running = False
        self._arrival_last_result = None
        self._bom_clean_thread = None
        self._bom_clean_running = False
        self._bom_clean_last_result = None
        self._bom_supply_result_rows = []
        self._mrp_update_running = False
        self._mrp_update_thread = None

        # ── 外采物料ABC分类数据 ───────────────────────────────
        self._po_data   = []   # 排产计划：[(母件编码, 数量), ...]
        self._po_raw_rows = []  # 排产原始行（含日期）
        self._plan_days = 30    # 排产天数（从上线日期范围计算）
        self._ext_data  = {}   # 外采清单
        self._abc_result = []  # ABC分类结果
        self._abc_custom_diff_rows = []  # 定制机型与其他排产项目的BOM差异物料
        self._abc_plan_expansion_rows = []  # ABC计算实际采用的排产行（已应用去重/定制控制）
        self._abc_mrp_path = ''

        self.root = tk.Tk()
        self.calc_summary_var = tk.StringVar(value='📌 当前BOM：0颗物料')
        self.root.title('物料供需协同工具 V3.0 - 细分进度版')
        self.root.geometry('1460x940')
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        # 样式在 _build_ui 里统一配置，这里不再重复写入以避免相互覆盖
        self._build_menu()
        self._build_ui()
        self._try_load_cache()
        self._update_balance_bom_badge()
        self._update_readiness_bom_badge()
        self._update_arrival_bom_badge()

    def _build_menu(self):
        """顶部菜单：文件 / 视图 / 帮助。"""
        menubar = tk.Menu(self.root)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label='上传 BOM...', accelerator='Ctrl+O',
                              command=self._upload_bom)
        file_menu.add_command(label='清除 BOM 缓存', command=self._clear_bom_cache)
        file_menu.add_separator()
        file_menu.add_command(label='退出', accelerator='Alt+F4',
                              command=self.root.destroy)
        menubar.add_cascade(label='文件', menu=file_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label='查看运行日志', accelerator='F9',
                              command=self._show_any_log_window)
        view_menu.add_command(label='重置所有表格列宽', command=self._reset_all_column_widths)
        menubar.add_cascade(label='视图', menu=view_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label='快速入门', command=self._show_quick_start)
        help_menu.add_command(label='快捷键一览', command=self._show_shortcuts)
        help_menu.add_separator()
        help_menu.add_command(label='关于 V3.0', command=self._show_about)
        menubar.add_cascade(label='帮助', menu=help_menu)

        self.root.configure(menu=menubar)

        # 全局快捷键
        self.root.bind_all('<Control-o>', lambda _e: self._upload_bom())
        self.root.bind_all('<Control-O>', lambda _e: self._upload_bom())
        self.root.bind_all('<F9>', lambda _e: self._show_any_log_window())

    def _clear_bom_cache(self):
        if not messagebox.askyesno('清除缓存', '确定要清除本地 BOM 缓存吗？\n缓存清除后需要重新上传 BOM 文件。'):
            return
        try:
            if os.path.exists(CACHE_FILE):
                os.remove(CACHE_FILE)
            self.bom_index = {}
            self._diff_meta_cache = None
            self._bom_edges_cache = None
            self.file_var.set('未加载')
            self.status_var.set('缓存已清除，请重新上传 BOM 文件')
            self._update_balance_bom_badge()
            self._update_readiness_bom_badge()
            self._update_arrival_bom_badge()
            self._clear_bom_supply()
            self._refresh_calc_hint()
        except OSError as exc:
            messagebox.showerror('清除失败', f'无法删除缓存文件：{exc}')

    def _show_any_log_window(self):
        """唤起日志窗口（含平衡/齐套/到货多个标签页）。"""
        try:
            win = self._ensure_log_window()
            win.deiconify()
            win.lift()
        except Exception as exc:
            messagebox.showinfo('日志', f'日志窗口不可用：{exc}')

    def _reset_all_column_widths(self):
        """把主要 Treeview 的列宽重置为默认值。"""
        pairs = [
            (getattr(self, 'tree', None), HEADERS_TREE, None),
            (getattr(self, 'abc_tree', None), ABC_HEADERS, ABC_WIDTHS),
            (getattr(self, 'bom_supply_tree', None), BOM_SUPPLY_HEADERS, BOM_SUPPLY_WIDTHS),
            (getattr(self, 'calc_tree', None), HEADERS_CALC, COL_WIDTHS_C),
            (getattr(self, 'diff_tree', None), HEADERS_DIFF_GRID, COL_WIDTHS_D_GRID),
            (getattr(self, 'balance_tree', None), BALANCE_SUGGEST_HEADERS, COL_WIDTHS_BALANCE),
        ]
        for tree, cols, widths in pairs:
            if tree is None or widths is None:
                continue
            for col, width in zip(cols, widths):
                try:
                    tree.column(col, width=width)
                except tk.TclError:
                    pass
        self.status_var.set('已重置所有表格列宽')

    def _show_quick_start(self):
        text = (
            '🚀 快速入门\n\n'
            '1) 通过 "文件 > 上传 BOM" 或主界面按钮导入 BOM 文件。\n'
            '2) "BOM 结构展开" 页：输入母料号查询层级结构，双击复制单元格。\n'
            '3) "数量计算" 页：支持批量母料号+数量的聚合展开。\n'
            '4) "外采物料管理" 页：配合排产计划做 ABC 分类和安全库存建议。\n'
            '5) "平衡表/齐套/到货" 页：调用 mrp_balance_tool 做 MRP 计算。\n\n'
            '任意数据表头都支持点击排序；选中数据行后按 Ctrl+C 可以直接复制到剪贴板。'
        )
        messagebox.showinfo('快速入门', text)

    def _show_shortcuts(self):
        text = (
            '⌨ 快捷键一览\n\n'
            'Ctrl+O      上传 BOM 文件\n'
            'F9          打开运行日志窗口\n'
            'Ctrl+C      在任何数据表格中复制选中行（TSV 格式，可直接贴 Excel）\n'
            'F5          刷新 / 重新查询（在 BOM 结构展开页）\n'
            'Esc         关闭当前对话框 / 错误窗口\n'
            'Enter       BOM 展开页：触发查询'
        )
        messagebox.showinfo('快捷键', text)

    def _show_about(self):
        text = (
            '物料供需协同工具 V3.0 - 细分进度版\n\n'
            '集成：BOM 结构展开 / 批量数量计算 / 外采物料 ABC / 差异对比 /\n'
            '         MRP 平衡表 / 外购齐套 / 到货跟催\n\n'
            '本版本强化：表头排序、阶段式进度条、写入 Excel 细分进度、\n'
            '            BOM 环路检测、缓存签名校验、打包路径兼容。'
        )
        messagebox.showinfo('关于', text)

    def _try_load_cache(self):
        """启动时尝试加载本地缓存"""
        if not os.path.exists(CACHE_FILE):
            self.status_var.set('就绪  |  未找到缓存，请上传 BOM 文件')
            return
        try:
            data = load_cache(CACHE_FILE)
            self.bom_index = data['index']
            self._diff_meta_cache = None
            self._bom_edges_cache = None
            self.file_var.set('缓存: .bom_cache.pkl')
            self.status_var.set(f'缓存加载成功  |  母器件: {len(self.bom_index)} 个')
            self._update_balance_bom_badge()
            self._update_readiness_bom_badge()
            self._update_arrival_bom_badge()
            self._refresh_calc_hint()
        except Exception as e:
            self.status_var.set(f'缓存加载失败: {e}')

    # ── UI ────────────────────────────────────────────────────
    def _build_ui(self):
        style = ttk.Style()
        style.theme_use('clam')

        bg_main = '#F8FAFC'
        bg_card = '#FFFFFF'
        bg_muted = '#EDF2F7'
        accent = '#2563EB'
        accent_hover = '#1D4ED8'
        accent_soft = '#DBEAFE'
        cta = '#F97316'
        cta_hover = '#EA580C'
        text_primary = '#1E293B'
        text_secondary = '#64748B'
        border = '#E2E8F0'
        success = '#0F766E'
        dark_header = '#0F172A'
        dark_header_soft = '#94A3B8'
        body_font = ('Microsoft YaHei', 9)
        body_font_sm = ('Microsoft YaHei', 9)
        body_font_lg = ('Microsoft YaHei', 10)
        heading_font = ('Microsoft YaHei', 10, 'bold')
        title_font = ('Microsoft YaHei', 16, 'bold')

        self.root.configure(bg=bg_main)

        style.configure('.', background=bg_main, foreground=text_primary, font=body_font)
        style.configure('Shell.TFrame', background=bg_main)
        style.configure('Card.TFrame', background=bg_card, relief='flat')
        style.configure('Toolbar.TFrame', background=bg_card)
        style.configure('TLabel', background=bg_main, foreground=text_primary)
        style.configure('Card.TLabelframe', background=bg_card, borderwidth=1, relief='solid',
                        bordercolor=border, lightcolor=border, darkcolor=border)
        style.configure('Card.TLabelframe.Label', background=bg_card, foreground=accent,
                        font=heading_font)
        style.configure('Subtle.TLabel', background=bg_main, foreground=text_secondary,
                        font=body_font_sm)
        style.configure('Section.TLabel', background=bg_main, foreground=text_primary,
                        font=heading_font)
        style.configure('Metric.TLabel', background=bg_main, foreground=accent,
                        font=heading_font)
        style.configure('Badge.TLabel', background=accent_soft, foreground=accent,
                        font=('Microsoft YaHei', 9, 'bold'), padding=(8, 3))
        style.configure('Status.TLabel', background=bg_main, foreground=success,
                        font=('Microsoft YaHei', 10, 'bold'))

        style.configure('TButton', padding=(10, 5), font=body_font,
                        foreground=text_primary, background=bg_muted, borderwidth=1)
        style.map('TButton',
                  background=[('active', '#E2E8F0'), ('pressed', '#CBD5E1')],
                  relief=[('pressed', 'sunken')])
        style.configure('Accent.TButton', padding=(12, 6), font=('Microsoft YaHei', 10, 'bold'),
                        foreground='white', background=cta, borderwidth=0)
        style.map('Accent.TButton',
                  background=[('active', cta_hover), ('pressed', '#C2410C')],
                  foreground=[('active', 'white')])
        style.configure('Quiet.TButton', padding=(10, 5), font=body_font,
                        foreground=text_secondary, background=bg_card, borderwidth=1)

        style.configure('TNotebook', background=bg_main, borderwidth=0, tabmargins=(0, 0, 0, 0))
        style.configure('TNotebook.Tab', padding=(14, 7), font=('Microsoft YaHei', 10, 'bold'),
                        background=bg_muted, foreground=text_secondary, borderwidth=0)
        style.map('TNotebook.Tab',
                  background=[('selected', bg_card), ('active', '#E2E8F0')],
                  foreground=[('selected', accent), ('active', text_primary)],
                  expand=[('selected', [0, 0, 0, 0])])

        style.configure('Treeview', rowheight=28, font=body_font_sm,
                        background=bg_card, fieldbackground=bg_card,
                        foreground=text_primary, borderwidth=0)
        style.map('Treeview',
                  background=[('selected', accent)],
                  foreground=[('selected', 'white')])
        style.configure('Treeview.Heading', font=('Microsoft YaHei', 10, 'bold'),
                        background='#1E3A5F', foreground='white', relief='flat', borderwidth=0)
        style.map('Treeview.Heading', background=[('active', '#27476F')])

        style.configure('TEntry', padding=(7, 5), font=body_font_lg,
                        fieldbackground='white', borderwidth=1)
        style.configure('TSpinbox', padding=(5, 4))
        style.configure('TProgressbar', troughcolor=bg_muted, background=accent, bordercolor=bg_muted)

        top_shell = ttk.Frame(self.root, style='Shell.TFrame')
        top_shell.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 6))
        top_shell.columnconfigure(0, weight=1)

        header = tk.Frame(top_shell, bg=dark_header, padx=20, pady=14)
        header.grid(row=0, column=0, sticky='ew')
        header.grid_columnconfigure(0, weight=1)
        tk.Label(header, text='物料供需协同工具', bg=dark_header, fg='white',
                 font=title_font).grid(row=0, column=0, sticky='w')
        tk.Label(header, text='BOM展开、平衡表生成、采购协同与差异对比一体化工具',
                 bg=dark_header, fg=dark_header_soft, font=body_font).grid(
            row=1, column=0, sticky='w', pady=(3, 0)
        )

        toolbar = ttk.Frame(top_shell, style='Toolbar.TFrame', padding=(14, 8))
        toolbar.grid(row=1, column=0, sticky='ew', pady=(6, 0))
        toolbar.columnconfigure(2, weight=1)

        ttk.Label(toolbar, text='当前数据源', style='Section.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 8))
        self.file_var = tk.StringVar(value='未加载 BOM 文件')
        ttk.Label(toolbar, textvariable=self.file_var, style='Badge.TLabel').grid(row=0, column=1, sticky='w', padx=(0, 14))
        ttk.Button(toolbar, text='上传BOM文件', command=self._upload_bom, style='Accent.TButton').grid(row=0, column=2, sticky='w')

        search_frame = ttk.Frame(toolbar, style='Toolbar.TFrame')
        search_frame.grid(row=0, column=3, sticky='e')
        ttk.Label(search_frame, text='全局搜索', style='Subtle.TLabel').pack(side='left', padx=(0, 8))
        self.search_var = tk.StringVar()
        se = ttk.Entry(search_frame, textvariable=self.search_var, width=20)
        se.pack(side='left', padx=(0, 6))
        se.bind('<Return>', lambda e: self._do_quick_search())
        ttk.Button(search_frame, text='搜索', width=6, command=self._do_quick_search).pack(side='left', padx=(0, 4))
        ttk.Button(search_frame, text='上一个', style='Quiet.TButton', command=self._search_prev_match).pack(side='left', padx=(0, 4))
        ttk.Button(search_frame, text='下一个', style='Quiet.TButton', command=self._search_next_match).pack(side='left')

        self._search_matches = []
        self._search_current_index = -1
        self.status_var = tk.StringVar(value='就绪')
        self._log_window = None
        self._log_notebook = None
        self._log_text_widgets = {}
        self._log_buffers = {'balance': [], 'readiness': [], 'arrival': [], 'material_buy': []}

        # 顶层分组 Notebook：把 7 个子 tab 按业务场景分成 3 组
        self.nb = ttk.Notebook(self.root)
        self.nb.grid(row=1, column=0, sticky='nsew', padx=12, pady=(0, 6))
        self.nb.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        # 每组下的"内层" Notebook
        group_query = ttk.Frame(self.nb)
        group_mrp = ttk.Frame(self.nb)
        group_followup = ttk.Frame(self.nb)
        for group in (group_query, group_mrp, group_followup):
            group.columnconfigure(0, weight=1)
            group.rowconfigure(0, weight=1)

        self.nb.add(group_query, text='🧬 BOM 查询')
        self.nb.add(group_mrp, text='📊 MRP 计算')
        self.nb.add(group_followup, text='🚚 协同跟催')

        self._inner_query = ttk.Notebook(group_query)
        self._inner_query.grid(row=0, column=0, sticky='nsew')
        self._inner_query.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        self._inner_mrp = ttk.Notebook(group_mrp)
        self._inner_mrp.grid(row=0, column=0, sticky='nsew')
        self._inner_mrp.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        self._inner_followup = ttk.Notebook(group_followup)
        self._inner_followup.grid(row=0, column=0, sticky='nsew')
        self._inner_followup.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        # tab_frame -> 所属的 inner notebook，供 _switch_to_tab / _get_current_tab 使用
        self._tab_parent: dict[tk.Widget, ttk.Notebook] = {}
        self._tab_group: dict[tk.Widget, ttk.Frame] = {}

        def _add_inner(frame_attr: str, builder, inner_nb: ttk.Notebook, group_frame: ttk.Frame, text: str):
            tab = ttk.Frame(inner_nb)
            inner_nb.add(tab, text=text)
            setattr(self, frame_attr, tab)
            self._tab_parent[tab] = inner_nb
            self._tab_group[tab] = group_frame
            builder()

        # BOM 查询组：结构展开 / 数量计算 / 差异对比 / 文件清洗
        _add_inner('tab_tree', self._build_tree_tab, self._inner_query, group_query, 'BOM 结构展开')
        _add_inner('tab_bom_supply', self._build_bom_supply_tab, self._inner_query, group_query, 'BOM供需展开')
        _add_inner('tab_calc', self._build_calc_tab, self._inner_query, group_query, '数量计算')
        _add_inner('tab_diff', self._build_diff_tab, self._inner_query, group_query, 'BOM 差异对比')
        _add_inner('tab_bom_clean', self._build_bom_clean_tab, self._inner_query, group_query, 'BOM 文件清洗')

        # MRP 计算组：外采 ABC / 平衡表
        _add_inner('tab_abc', self._build_abc_tab, self._inner_mrp, group_mrp, '外采物料管理')
        _add_inner('tab_balance', self._build_balance_tab, self._inner_mrp, group_mrp, '平衡表生成')
        _add_inner('tab_mrp_update', self._build_mrp_update_tab, self._inner_mrp, group_mrp, 'MRP数据更新')
        _add_inner('tab_material_buy', self._build_material_buy_tab, self._inner_mrp, group_mrp, '单料采购判断')

        # 协同跟催组：外购齐套 / 到货跟催
        _add_inner('tab_readiness', self._build_readiness_tab, self._inner_followup, group_followup, '外购齐套分析')
        _add_inner('tab_arrival', self._build_arrival_tab, self._inner_followup, group_followup, '到货跟催分析')
        _add_inner('tab_work_order_reply', self._build_work_order_reply_tab, self._inner_followup, group_followup, '工单缺料回复')

        footer = ttk.Frame(self.root, style='Toolbar.TFrame', padding=(14, 10))
        footer.grid(row=2, column=0, sticky='ew', padx=12, pady=(0, 12))
        footer.columnconfigure(0, weight=1)
        ttk.Label(footer, textvariable=self.status_var, anchor='w',
                  style='Status.TLabel').grid(row=0, column=0, sticky='ew')

    # ── 嵌套 Notebook 辅助 ───────────────────────────────────
    def _switch_to_tab(self, tab_frame):
        """切换到指定 tab（自动处理外层分组 notebook 和内层 notebook）。"""
        inner = self._tab_parent.get(tab_frame)
        group = self._tab_group.get(tab_frame)
        if inner is None or group is None:
            # 兼容外部传入的旧索引或直接 tab frame
            try:
                self.nb.select(tab_frame)
            except tk.TclError:
                pass
            return
        try:
            self.nb.select(group)
            inner.select(tab_frame)
        except tk.TclError:
            pass

    def _get_current_tab(self):
        """返回当前被激活的内层 tab Frame（没有则返回 None）。"""
        try:
            current_group = self.nb.nametowidget(self.nb.select())
        except tk.TclError:
            return None
        for inner in (self._inner_query, self._inner_mrp, self._inner_followup):
            try:
                master = inner.master
            except tk.TclError:
                continue
            if master is current_group:
                try:
                    return inner.nametowidget(inner.select())
                except tk.TclError:
                    return None
        return None

    def _is_current_tab(self, tab_frame) -> bool:
        return self._get_current_tab() is tab_frame

    # ── 选项卡1：树形展开 ──────────────────────────────────────
    def _build_tree_tab(self):
        f = self.tab_tree
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='结构查询控制台', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(1, weight=4)
        control.columnconfigure(3, weight=2)
        control.columnconfigure(9, weight=1)

        ttk.Label(control, text='输入母料号', style='Section.TLabel').grid(row=0, column=0, padx=(0, 6), sticky='w')
        self.pn_entry = ttk.Entry(control, font=('Microsoft YaHei', 12))
        self.pn_entry.grid(row=0, column=1, sticky='ew', padx=(0, 10))
        self.pn_entry.bind('<Return>', lambda e: self._query())

        ttk.Label(control, text='最大层数', style='Subtle.TLabel').grid(row=0, column=2, padx=(0, 6))
        self.depth_var = tk.IntVar(value=50)
        ttk.Spinbox(control, from_=1, to=MAX_DEPTH, textvariable=self.depth_var, width=6).grid(
            row=0, column=3, padx=(0, 10), sticky='w')

        ttk.Button(control, text='开始查询', style='Accent.TButton', command=self._query).grid(
            row=0, column=4, padx=(0, 6))
        ttk.Button(control, text='全部母器件', command=self._list_all_pns).grid(
            row=0, column=5, padx=(0, 6))
        ttk.Button(control, text='导入料号', command=self._import_pns_file).grid(
            row=0, column=6, padx=(0, 6))
        ttk.Button(control, text='导出Excel', command=self._export_tree).grid(
            row=0, column=7, padx=(0, 6))
        ttk.Button(control, text='清空', style='Quiet.TButton', command=self._clear_tree).grid(
            row=0, column=8)

        ttk.Label(control, text='输入子料号', style='Section.TLabel').grid(row=1, column=0, padx=(0, 6), pady=(8, 0), sticky='w')
        self.reverse_pn_entry = ttk.Entry(control, font=('Microsoft YaHei', 11))
        self.reverse_pn_entry.grid(row=1, column=1, sticky='ew', padx=(0, 10), pady=(8, 0))
        self.reverse_pn_entry.bind('<Return>', lambda e: self._reverse_lookup_from_entry())
        ttk.Label(control, text='反查所有直接/间接上层母料号', style='Subtle.TLabel').grid(
            row=1, column=2, columnspan=2, sticky='w', padx=(0, 8), pady=(8, 0)
        )
        ttk.Button(control, text='反查上层母件', command=self._reverse_lookup_from_entry).grid(
            row=1, column=4, padx=(0, 6), pady=(8, 0), sticky='w'
        )

        tree_frame = ttk.LabelFrame(f, text='BOM结构明细', style='Card.TLabelframe', padding=6)
        tree_frame.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        self.tree = ttk.Treeview(tree_frame, columns=HEADERS_TREE,
                                 show='tree headings', selectmode='extended')
        for h, w in zip(HEADERS_TREE, COL_WIDTHS):
            self.tree.column(h, width=w, anchor='w')
            self.tree.heading(h, text=h)
        self.tree.column('#0', width=160, stretch=False)

        style = ttk.Style()
        style.configure('Treeview', rowheight=32)
        style.map('Treeview', background=[('selected', '#2563EB')])

        self.tree.bind('<<TreeviewOpen>>', self._on_node_open)
        self.tree.bind('<Double-Button-1>', self._copy_tree_cell)
        self.tree.bind('<Button-3>', self._show_tree_menu)
        enable_treeview_copy(self.tree, HEADERS_TREE)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        self._ctx_menu = tk.Menu(self.tree, tearoff=0)
        self._ctx_menu.add_command(label='复制整行', command=self._copy_tree_row)
        self._ctx_menu.add_command(label='复制子件料号', command=self._copy_tree_pn)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label='展开全部节点', command=self._expand_all)
        self._ctx_menu.add_command(label='折叠全部节点', command=self._collapse_all)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label='📦 计算此节点用量', command=self._calc_from_selected)
        self._ctx_menu.add_command(label='🔎 反查此料上层母件', command=self._reverse_lookup_selected_tree_item)

    # ── BOM供需展开 ───────────────────────────────────────────
    def _build_bom_supply_tab(self):
        f = self.tab_bom_supply
        f.columnconfigure(0, weight=1)
        f.rowconfigure(3, weight=1)

        control = ttk.LabelFrame(f, text='BOM供需展开参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(1, weight=3)
        control.columnconfigure(5, weight=2)
        control.columnconfigure(10, weight=1)

        ttk.Label(control, text='母件料号/规格/名称', style='Section.TLabel').grid(
            row=0, column=0, sticky='w', padx=(0, 6)
        )
        self.bom_supply_root_var = tk.StringVar(value='')
        root_entry = ttk.Entry(control, textvariable=self.bom_supply_root_var, font=('Microsoft YaHei', 11))
        root_entry.grid(row=0, column=1, sticky='ew', padx=(0, 10))
        root_entry.bind('<Return>', lambda _e: self._run_bom_supply_expand())

        ttk.Label(control, text='数量').grid(row=0, column=2, sticky='e', padx=(0, 4))
        self.bom_supply_qty_var = tk.StringVar(value='1')
        ttk.Entry(control, textvariable=self.bom_supply_qty_var, width=8).grid(
            row=0, column=3, sticky='w', padx=(0, 10)
        )

        ttk.Label(control, text='最大层数').grid(row=0, column=4, sticky='e', padx=(0, 4))
        self.bom_supply_depth_var = tk.IntVar(value=50)
        ttk.Spinbox(control, from_=1, to=MAX_DEPTH, textvariable=self.bom_supply_depth_var, width=6).grid(
            row=0, column=5, sticky='w', padx=(0, 10)
        )

        ttk.Button(control, text='生成供需展开', style='Accent.TButton',
                   command=self._run_bom_supply_expand).grid(row=0, column=6, padx=(0, 6))
        ttk.Button(control, text='导出Excel',
                   command=self._export_bom_supply_result).grid(row=0, column=7, padx=(0, 6))
        ttk.Button(control, text='清空', style='Quiet.TButton',
                   command=self._clear_bom_supply).grid(row=0, column=8)

        ttk.Label(control, text='MRP计算表', style='Section.TLabel').grid(
            row=1, column=0, sticky='w', padx=(0, 6), pady=(8, 0)
        )
        self.bom_supply_mrp_var = tk.StringVar(value='')
        mrp_label, _ = make_path_label(control, self.bom_supply_mrp_var, placeholder='未选择')
        mrp_label.grid(row=1, column=1, columnspan=5, sticky='ew', padx=(0, 10), pady=(8, 0))
        ttk.Button(control, text='选择MRP计算表',
                   command=self._choose_bom_supply_mrp_file).grid(row=1, column=6, padx=(0, 6), pady=(8, 0))

        self.bom_supply_summary_var = tk.StringVar(value='先选择或确认 MRP 计算表，再输入母件料号/规格/名称。')
        ttk.Label(control, textvariable=self.bom_supply_summary_var, foreground='#2563EB',
                  font=('Microsoft YaHei', 10, 'bold')).grid(
            row=1, column=7, columnspan=4, sticky='w', padx=(6, 0), pady=(8, 0)
        )

        batch_frame = ttk.LabelFrame(f, text='批量母件+数量（可从 Excel 粘贴两列）', style='Card.TLabelframe', padding=6)
        batch_frame.grid(row=1, column=0, sticky='ew', padx=6, pady=(0, 3))
        batch_frame.columnconfigure(1, weight=1)
        ttk.Label(batch_frame, text='每行格式', style='Subtle.TLabel').grid(row=0, column=0, sticky='nw', padx=(0, 6))
        ttk.Label(
            batch_frame,
            text='母件料号/规格/名称<Tab>数量；数量空着则用上方默认数量。批量框有内容时，优先按批量框计算。',
            foreground='#64748B',
        ).grid(row=0, column=1, sticky='w', pady=(0, 3))
        self.bom_supply_batch_text = tk.Text(batch_frame, height=4, wrap='none', font=('Microsoft YaHei', 10))
        self.bom_supply_batch_text.grid(row=1, column=0, columnspan=2, sticky='ew')
        batch_vsb = ttk.Scrollbar(batch_frame, orient='vertical', command=self.bom_supply_batch_text.yview)
        batch_vsb.grid(row=1, column=2, sticky='ns')
        self.bom_supply_batch_text.configure(yscrollcommand=batch_vsb.set)

        note = ttk.Label(
            f,
            text='供给口径：库存=期初库存；未转单数量=在途请购；未到货数量=在途采购欠交。缺口按 需求-库存-未转-未到货 计算。',
            foreground='#64748B',
        )
        note.grid(row=2, column=0, sticky='w', padx=10, pady=(0, 3))

        result_frame = ttk.LabelFrame(f, text='BOM供需展开明细', style='Card.TLabelframe', padding=6)
        result_frame.grid(row=3, column=0, sticky='nsew', padx=6, pady=(0, 6))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)

        self.bom_supply_tree = ttk.Treeview(
            result_frame,
            columns=BOM_SUPPLY_HEADERS,
            show='headings',
            selectmode='extended',
        )
        numeric_cols = {
            '输入数量', '层级', '单套用量', '累计用量', '需求数量',
            '库存', '未转单数量', '未到货数量', '采购周期', '缺口'
        }
        for h, w in zip(BOM_SUPPLY_HEADERS, BOM_SUPPLY_WIDTHS):
            self.bom_supply_tree.column(h, width=w, anchor='w')
            self.bom_supply_tree.heading(h, text=h)
        enable_treeview_sort(self.bom_supply_tree, BOM_SUPPLY_HEADERS, numeric_columns=numeric_cols)
        enable_treeview_copy(self.bom_supply_tree, BOM_SUPPLY_HEADERS)

        vsb = ttk.Scrollbar(result_frame, orient='vertical', command=self.bom_supply_tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient='horizontal', command=self.bom_supply_tree.xview)
        self.bom_supply_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.bom_supply_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

    def _choose_bom_supply_mrp_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP 计算表',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.bom_supply_mrp_var.set(path)
        self.status_var.set(f'BOM供需展开已选择 MRP 计算表：{os.path.basename(path)}')

    def _clear_bom_supply(self):
        self._bom_supply_result_rows = []
        if hasattr(self, 'bom_supply_tree'):
            self.bom_supply_tree.delete(*self.bom_supply_tree.get_children(''))
        if hasattr(self, 'bom_supply_batch_text'):
            self.bom_supply_batch_text.delete('1.0', tk.END)
        if hasattr(self, 'bom_supply_summary_var'):
            self.bom_supply_summary_var.set('已清空')
        self.status_var.set('BOM供需展开结果已清空')

    def _parse_bom_supply_inputs(self):
        default_qty_text = str(self.bom_supply_qty_var.get() or '1').replace(',', '').strip()
        try:
            default_qty = float(default_qty_text)
            if default_qty <= 0:
                raise ValueError
        except ValueError:
            raise ValueError('默认数量必须是大于 0 的数字')

        batch_text = ''
        if hasattr(self, 'bom_supply_batch_text'):
            batch_text = self.bom_supply_batch_text.get('1.0', tk.END).strip()

        entries = []
        if batch_text:
            for line_no, raw_line in enumerate(batch_text.splitlines(), 1):
                line = str(raw_line or '').strip()
                if not line:
                    continue
                lower = line.lower()
                if line_no == 1 and any(token in lower for token in ('母件', '料号', '规格', '名称', '数量', 'qty')):
                    continue
                if '\t' in line:
                    parts = [part.strip() for part in line.split('\t')]
                elif '，' in line or ',' in line:
                    parts = [part.strip() for part in line.replace('，', ',').split(',')]
                else:
                    parts = line.split()
                term = parts[0].strip() if parts else ''
                qty_text = parts[1].strip() if len(parts) >= 2 else ''
                if not term:
                    continue
                qty = default_qty
                if qty_text:
                    try:
                        qty = float(qty_text.replace(',', ''))
                    except ValueError as exc:
                        raise ValueError(f'批量输入第 {line_no} 行数量无效: {qty_text}') from exc
                    if qty <= 0:
                        raise ValueError(f'批量输入第 {line_no} 行数量必须大于 0')
                entries.append((term, qty))
        else:
            query = self.bom_supply_root_var.get().strip()
            if not query:
                raise ValueError('请输入母件料号、规格或名称，或在批量框粘贴母件+数量')
            for part in [p.strip() for p in query.replace('，', ',').replace('\n', ',').split(',') if p.strip()]:
                entries.append((part, default_qty))

        merged = {}
        order = []
        for term, qty in entries:
            key = term.strip()
            if not key:
                continue
            if key not in merged:
                merged[key] = 0.0
                order.append(key)
            merged[key] += float(qty)
        return [(term, merged[term]) for term in order]

    def _run_bom_supply_expand(self):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        mrp_path = self.bom_supply_mrp_var.get().strip()
        if not mrp_path:
            messagebox.showwarning('缺少 MRP 计算表', '请先选择 MRP 计算表')
            return
        if not os.path.exists(mrp_path):
            messagebox.showwarning('文件不存在', f'MRP 计算表不存在：\n{mrp_path}')
            return
        try:
            input_entries = self._parse_bom_supply_inputs()
        except ValueError as exc:
            messagebox.showwarning('输入无效', str(exc))
            return

        matched_entries = []
        unresolved_terms = []
        for term, qty in input_entries:
            matches = self._match_bom_roots(term, limit=20)
            if not matches:
                unresolved_terms.append(term)
                continue
            for matched in matches:
                matched_entries.append((matched, qty, term))
        if not matched_entries:
            messagebox.showwarning('未匹配', f'BOM 中找不到：{", ".join(unresolved_terms[:8])}')
            return

        self.status_var.set('正在生成 BOM供需展开...')
        self.bom_supply_summary_var.set('正在读取 MRP 并展开 BOM...')
        self.root.update_idletasks()
        try:
            supply_maps = load_bom_supply_maps(mrp_path)
            all_rows = []
            truncated_roots = []
            for root_code, root_qty, source_term in matched_entries:
                rows, truncated = build_bom_supply_rows(
                    self.bom_index,
                    root_code,
                    root_qty,
                    self.bom_supply_depth_var.get(),
                    supply_maps,
                )
                for row in rows:
                    if source_term != root_code:
                        row['查询母件'] = f'{source_term} -> {root_code}'
                all_rows.extend(rows)
                if truncated:
                    truncated_roots.append(root_code)

            self._bom_supply_result_rows = all_rows
            self.bom_supply_tree.delete(*self.bom_supply_tree.get_children(''))
            for row in all_rows:
                self.bom_supply_tree.insert('', 'end', values=[row.get(h, '') for h in BOM_SUPPLY_HEADERS])

            warning_text = ''
            if supply_maps.get('warnings'):
                warning_text = '；提示: ' + '；'.join(supply_maps['warnings'][:3])
            if truncated_roots:
                warning_text += f'；{len(truncated_roots)} 个母件达到行数上限已截断'
            if unresolved_terms:
                warning_text += f'；未匹配 {len(unresolved_terms)} 个'
            self.bom_supply_summary_var.set(
                f'输入 {len(input_entries)} 行 | 匹配母件 {len(matched_entries)} 个 | 明细 {len(all_rows)} 行 | MRP: {os.path.basename(mrp_path)}{warning_text}'
            )
            self.status_var.set(f'BOM供需展开完成：{len(all_rows)} 行')
        except Exception as exc:
            messagebox.showerror('生成失败', str(exc))
            self.bom_supply_summary_var.set('生成失败')
            self.status_var.set(f'BOM供需展开失败: {exc}')

    def _export_bom_supply_result(self):
        if not self._bom_supply_result_rows:
            messagebox.showwarning('无数据', '请先生成 BOM供需展开')
            return
        path = filedialog.asksaveasfilename(
            title='导出 BOM供需展开',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'BOM供需展开'
            hdr_font = Font(bold=True, color='FFFFFF', size=10)
            hdr_fill = PatternFill('solid', fgColor='2563EB')
            even_fill = PatternFill('solid', fgColor='EEF4FF')
            thin = Side(style='thin', color='D0D7DE')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            numeric_headers = {
                '输入数量', '层级', '单套用量', '累计用量', '需求数量',
                '库存', '未转单数量', '未到货数量', '采购周期', '缺口'
            }

            def _to_excel_number(value):
                if value is None or value == '':
                    return ''
                try:
                    number = float(str(value).replace(',', '').strip())
                except (TypeError, ValueError):
                    return value
                if math.isnan(number) or math.isinf(number):
                    return ''
                return int(number) if number == int(number) else number

            for ci, header in enumerate(BOM_SUPPLY_HEADERS, 1):
                cell = ws.cell(row=1, column=ci, value=header)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 24

            for ri, row in enumerate(self._bom_supply_result_rows, 2):
                for ci, header in enumerate(BOM_SUPPLY_HEADERS, 1):
                    value = row.get(header, '')
                    if header in numeric_headers:
                        value = _to_excel_number(value)
                    cell = ws.cell(row=ri, column=ci, value=value)
                    cell.border = border
                    if ri % 2 == 0:
                        cell.fill = even_fill
                    if header in numeric_headers:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '0.######'
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[ri].height = 20

            for idx, width in enumerate(BOM_SUPPLY_WIDTHS, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(8, min(width // 8, 60))
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            wb.save(path)
            messagebox.showinfo('导出成功', f'已保存至：\n{path}')
            self.status_var.set(f'BOM供需展开已导出: {os.path.basename(path)}')
        except Exception as exc:
            messagebox.showerror('导出失败', str(exc))

    # ── MRP数据更新 ───────────────────────────────────────────
    def _build_mrp_update_tab(self):
        f = self.tab_mrp_update
        f.columnconfigure(0, weight=1)
        f.rowconfigure(2, weight=1)

        control = ttk.LabelFrame(f, text='自动更新 MRP 计算表', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(1, weight=1)
        control.columnconfigure(4, weight=1)

        ttk.Label(control, text='源文件夹', style='Section.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 6))
        self.mrp_update_source_var = tk.StringVar(value='')
        source_label, _ = make_path_label(control, self.mrp_update_source_var, placeholder='未选择')
        source_label.grid(row=0, column=1, columnspan=2, sticky='ew', padx=(0, 8))
        ttk.Button(control, text='选择文件夹', command=self._choose_mrp_update_source_folder).grid(
            row=0, column=3, sticky='w', padx=(0, 12)
        )

        ttk.Label(control, text='MRP计算表', style='Section.TLabel').grid(row=1, column=0, sticky='w', padx=(0, 6), pady=(8, 0))
        self.mrp_update_target_var = tk.StringVar(value='')
        target_label, _ = make_path_label(control, self.mrp_update_target_var, placeholder='未选择')
        target_label.grid(row=1, column=1, columnspan=2, sticky='ew', padx=(0, 8), pady=(8, 0))
        ttk.Button(control, text='选择MRP表', command=self._choose_mrp_update_target_file).grid(
            row=1, column=3, sticky='w', padx=(0, 12), pady=(8, 0)
        )

        self.mrp_update_use_db_var = tk.BooleanVar(value=IS_WINDOWS)
        ttk.Checkbutton(
            control,
            text='PR/PO使用数据库',
            variable=self.mrp_update_use_db_var,
        ).grid(row=2, column=0, columnspan=2, sticky='w', pady=(8, 0))

        ttk.Button(control, text='自动识别并更新', style='Accent.TButton',
                   command=self._start_mrp_auto_update).grid(row=0, column=4, rowspan=2, sticky='w', padx=(8, 8))
        ttk.Button(control, text='清空日志', style='Quiet.TButton',
                   command=self._clear_mrp_update_log).grid(row=0, column=5, rowspan=2, sticky='w')

        self.mrp_update_summary_var = tk.StringVar(
            value='规则：勾选后从数据库读取 PR/PO；不勾选则使用所选文件夹内的 PR/PO 文件。'
        )
        ttk.Label(control, textvariable=self.mrp_update_summary_var, foreground='#2563EB',
                  font=('Microsoft YaHei', 10, 'bold')).grid(
            row=3, column=0, columnspan=6, sticky='w', pady=(8, 0)
        )

        hint = ttk.Label(
            f,
            text=f'更新口径：PR 按需求池过滤；PO 直接按料号汇总欠交；库存按物料编码汇总结存。数据库连接读取环境变量 {U9_DB_CONN_ENV} 或用户配置目录。',
            foreground='#64748B',
        )
        hint.grid(row=1, column=0, sticky='w', padx=10, pady=(0, 3))

        log_frame = ttk.LabelFrame(f, text='更新日志', style='Card.TLabelframe', padding=6)
        log_frame.grid(row=2, column=0, sticky='nsew', padx=6, pady=(0, 6))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.mrp_update_log = tk.Text(log_frame, height=22, wrap='word', font=('Consolas', 10))
        self.mrp_update_log.grid(row=0, column=0, sticky='nsew')
        vsb = ttk.Scrollbar(log_frame, orient='vertical', command=self.mrp_update_log.yview)
        vsb.grid(row=0, column=1, sticky='ns')
        self.mrp_update_log.configure(yscrollcommand=vsb.set)

    def _choose_mrp_update_source_folder(self):
        path = filedialog.askdirectory(title='选择自动识别源文件夹')
        if path:
            self.mrp_update_source_var.set(path)
            self.status_var.set(f'MRP数据更新源文件夹: {path}')

    def _choose_mrp_update_target_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP计算表',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if path:
            self.mrp_update_target_var.set(path)
            self.status_var.set(f'MRP数据更新目标: {path}')

    def _append_mrp_update_log(self, message):
        if not hasattr(self, 'mrp_update_log'):
            return
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.mrp_update_log.insert(tk.END, f'[{timestamp}] {message}\n')
        self.mrp_update_log.see(tk.END)
        self.mrp_update_log.update_idletasks()

    def _clear_mrp_update_log(self):
        if hasattr(self, 'mrp_update_log'):
            self.mrp_update_log.delete('1.0', tk.END)
        self.mrp_update_summary_var.set('日志已清空')

    def _start_mrp_auto_update(self):
        if self._mrp_update_running:
            self.status_var.set('MRP数据更新正在执行中，请稍候')
            return
        source_folder = self.mrp_update_source_var.get().strip()
        target_path = self.mrp_update_target_var.get().strip()
        if not source_folder:
            messagebox.showwarning('缺少源文件夹', '请先选择源文件夹')
            return
        if not target_path:
            messagebox.showwarning('缺少MRP计算表', '请先选择 MRP计算表')
            return
        if not os.path.exists(target_path):
            messagebox.showwarning('文件不存在', f'MRP计算表不存在：\n{target_path}')
            return
        use_database = bool(self.mrp_update_use_db_var.get())

        self._mrp_update_running = True
        self.mrp_update_summary_var.set('正在自动识别并更新 MRP计算表...')
        self.status_var.set('MRP数据更新执行中...')
        self._append_mrp_update_log('=' * 64)
        self._append_mrp_update_log('开始 MRP 数据自动更新')
        self._mrp_update_thread = threading.Thread(
            target=self._run_mrp_auto_update_worker,
            args=(source_folder, target_path, use_database),
            daemon=True,
        )
        self._mrp_update_thread.start()

    def _run_mrp_auto_update_worker(self, source_folder, target_path, use_database=True):
        def emit(message):
            self.root.after(0, lambda msg=message: self._append_mrp_update_log(msg))

        try:
            emit(f'源文件夹: {source_folder}')
            emit(f'目标MRP表: {target_path}')
            emit(f'PR/PO来源: {"数据库" if use_database else "文件夹"}')
            found = find_latest_mrp_source_files(source_folder, require_pr_po=not use_database)
            emit('自动识别到以下文件：')
            if use_database:
                emit('  PR: U9 数据库')
                emit('  PO: U9 数据库')
            else:
                for idx, path in enumerate(found['pr_files'], 1):
                    emit(f'  PR{idx}: {path}')
                for idx, path in enumerate(found['po_files'], 1):
                    emit(f'  PO{idx}: {path}')
            emit(f'  需求池: {found["need_pool"]}')
            emit(f'  库存明细: {found["inventory"]}')

            if use_database:
                emit('步骤 1/4：处理 PR，从数据库读取未转PO，并按需求池过滤已核准行...')
                pr_df, pr_stats = merge_mrp_pr_database(found['need_pool'])
            else:
                emit('步骤 1/4：处理 PR，两份文件合并后按需求池过滤...')
                pr_df, pr_stats = merge_mrp_pr_files(found['pr_files'], found['need_pool'])
            emit(
                f'  PR 总行 {pr_stats["total"]} | 已核准 {pr_stats["approved"]} | '
                f'删除 {pr_stats["deleted"]} 行 / {fmt_qty(pr_stats["deleted_qty"])} | '
                f'料号 {pr_stats["items"]} | 未转PO合计 {fmt_qty(pr_stats["sum"])}'
            )

            if use_database:
                emit('步骤 2/4：处理 PO，从数据库按确认数量-实收-退扣+退补汇总欠交，不按需求池过滤...')
                po_df, po_stats = merge_mrp_po_database()
            else:
                emit('步骤 2/4：处理 PO，两份文件直接汇总欠交，不按需求池过滤...')
                po_df, po_stats = merge_mrp_po_files(found['po_files'])
            emit(
                f'  PO 总行 {po_stats["total"]} | 已核准 {po_stats["approved"]} | '
                f'删除 {po_stats["deleted"]} 行 | 料号 {po_stats["items"]} | 欠交合计 {fmt_qty(po_stats["sum"])}'
            )

            emit('步骤 3/4：处理库存明细，按物料编码汇总结存...')
            inventory_df, inv_stats = merge_mrp_inventory_file(found['inventory'])
            emit(
                f'  库存明细行 {inv_stats["total"]} | 料号 {inv_stats["items"]} | '
                f'库存合计 {fmt_qty(inv_stats["sum"])}'
            )

            emit('步骤 4/4：备份并写回 MRP计算表...')
            backup_path = update_mrp_workbook_sheets(target_path, pr_df, po_df, inventory_df)
            emit(f'  已备份: {backup_path}')
            emit('  已更新 Sheet: 在途请购 / 在途采购 / 期初库存')
            emit('MRP 数据自动更新完成')
            self.root.after(
                0,
                lambda: self._finish_mrp_auto_update(
                    f'更新完成 | PR {fmt_qty(pr_stats["sum"])} | PO {fmt_qty(po_stats["sum"])} | 库存 {fmt_qty(inv_stats["sum"])}'
                )
            )
        except Exception as exc:
            emit(f'更新失败: {exc}')
            self.root.after(0, lambda err=str(exc): self._finish_mrp_auto_update(f'更新失败: {err}', error=True))

    def _finish_mrp_auto_update(self, message, error=False):
        self._mrp_update_running = False
        self.mrp_update_summary_var.set(message)
        self.status_var.set(message)
        if error:
            messagebox.showerror('MRP数据更新失败', message)
        else:
            messagebox.showinfo('MRP数据更新', '更新完成，请查看日志。')

    # ── 选项卡3：外采物料ABC分类 + 安全库存 ──────────────────
    def _build_abc_tab(self):
        f = self.tab_abc
        f.columnconfigure(0, weight=1)
        f.rowconfigure(2, weight=1)

        row0 = ttk.LabelFrame(f, text='数据输入与计划参数', style='Card.TLabelframe', padding=6)
        row0.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        row0.columnconfigure(1, weight=1)
        row0.columnconfigure(5, weight=1)

        ttk.Label(row0, text='排产计划', font=('Microsoft YaHei', 10, 'bold')).grid(
            row=0, column=0, sticky='w', padx=(0, 4))
        self._po_count_var = tk.StringVar(value='未上传')
        ttk.Label(row0, textvariable=self._po_count_var, foreground='#6B7A8F').grid(
            row=0, column=1, sticky='w', padx=(0, 6))
        ttk.Button(row0, text='上传排产计划',
                   command=self._upload_po_plan).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(row0, text='查看/编辑排产',
                   command=self._show_po_plan).grid(row=0, column=3, padx=(0, 6))
        ttk.Button(row0, text='清除',
                   style='Quiet.TButton', command=self._clear_po).grid(row=0, column=4)

        ttk.Label(row0, text='排产天数（天）：').grid(row=0, column=5, sticky='e', padx=(0, 4))
        self._plan_days_manual = tk.StringVar(value='')
        manual_entry = ttk.Entry(row0, textvariable=self._plan_days_manual, width=6)
        manual_entry.grid(row=0, column=6, sticky='w', padx=(0, 4))
        manual_entry.bind('<KeyRelease>', lambda e: self._on_plan_days_manual_change())
        ttk.Label(row0, text='（留空=自动计算）', foreground='#6B7A8F').grid(
            row=0, column=7, sticky='w', padx=(0, 6))

        ttk.Label(row0, text='外采清单', font=('Microsoft YaHei', 10, 'bold')).grid(
            row=1, column=0, sticky='w', padx=(0, 4), pady=(4, 0))
        self._ext_count_var = tk.StringVar(value='未上传')
        ttk.Label(row0, textvariable=self._ext_count_var, foreground='#6B7A8F').grid(
            row=1, column=1, sticky='w', padx=(0, 12), pady=(4, 0))
        ttk.Button(row0, text='上传外采清单',
                   command=self._upload_ext_list).grid(row=1, column=2, padx=(0, 6), pady=(4, 0))
        ttk.Button(row0, text='清除',
                   style='Quiet.TButton', command=self._clear_ext).grid(row=1, column=3, pady=(4, 0))

        ttk.Label(row0, text='MRP计算表', font=('Microsoft YaHei', 10, 'bold')).grid(
            row=2, column=0, sticky='w', padx=(0, 4), pady=(4, 0))
        self._abc_mrp_var = tk.StringVar(value='未选择（不带库存位置/排除清单/物料分类）')
        ttk.Label(row0, textvariable=self._abc_mrp_var, foreground='#6B7A8F').grid(
            row=2, column=1, columnspan=4, sticky='w', padx=(0, 12), pady=(4, 0))
        ttk.Button(row0, text='选择MRP计算表',
                   command=self._choose_abc_mrp_file).grid(row=2, column=5, padx=(0, 6), pady=(4, 0), sticky='e')
        ttk.Button(row0, text='清除',
                   style='Quiet.TButton', command=self._clear_abc_mrp).grid(row=2, column=6, pady=(4, 0), sticky='w')

        row1 = ttk.LabelFrame(f, text='ABC计算参数与结果摘要', style='Card.TLabelframe', padding=6)
        row1.grid(row=1, column=0, sticky='ew', padx=6, pady=(0, 3))
        row1.columnconfigure(3, weight=1)

        ttk.Label(row1, text='日均用量系数：').grid(row=0, column=0, sticky='w', padx=(0, 4))
        self._daily_usage_factor = tk.DoubleVar(value=1.0)
        ttk.Entry(row1, textvariable=self._daily_usage_factor, width=8).grid(
            row=0, column=1, sticky='w', padx=(0, 12))

        ttk.Label(row1, text='滚动采购月份数：').grid(row=1, column=0, sticky='w', padx=(0, 4), pady=(4, 0))
        self._rolling_purchase_months = tk.IntVar(value=6)
        ttk.Spinbox(row1, from_=1, to=24, textvariable=self._rolling_purchase_months, width=6).grid(
            row=1, column=1, sticky='w', padx=(0, 12), pady=(4, 0))

        ttk.Label(row1, text='每月下单次数：').grid(row=1, column=8, sticky='e', padx=(8, 4), pady=(4, 0))
        self._abc_order_times_per_month = tk.IntVar(value=2)
        ttk.Spinbox(row1, from_=1, to=6, textvariable=self._abc_order_times_per_month, width=5).grid(
            row=1, column=9, sticky='w', padx=(0, 6), pady=(4, 0))
        ttk.Label(row1, text='下单日：').grid(row=1, column=10, sticky='e', padx=(0, 4), pady=(4, 0))
        self._abc_order_days = tk.StringVar(value='1,15')
        ttk.Entry(row1, textvariable=self._abc_order_days, width=12).grid(
            row=1, column=11, sticky='w', padx=(0, 4), pady=(4, 0))
        ttk.Label(row1, text='例:1,15', foreground='#6B7A8F').grid(
            row=1, column=12, sticky='w', pady=(4, 0))

        ttk.Label(row1, text='A类K值：').grid(row=2, column=0, sticky='w', padx=(0, 4), pady=(6, 0))
        self._abc_k_a = tk.DoubleVar(value=0.65)
        ttk.Entry(row1, textvariable=self._abc_k_a, width=8).grid(
            row=2, column=1, sticky='w', padx=(0, 12), pady=(6, 0))

        ttk.Label(row1, text='B类K值：').grid(row=2, column=2, sticky='w', padx=(0, 4), pady=(6, 0))
        self._abc_k_b = tk.DoubleVar(value=0.28)
        ttk.Entry(row1, textvariable=self._abc_k_b, width=8).grid(
            row=2, column=3, sticky='w', padx=(0, 12), pady=(6, 0))

        ttk.Label(row1, text='C类K值：').grid(row=2, column=4, sticky='w', padx=(0, 4), pady=(6, 0))
        self._abc_k_c = tk.DoubleVar(value=0.10)
        ttk.Entry(row1, textvariable=self._abc_k_c, width=8).grid(
            row=2, column=5, sticky='w', padx=(0, 12), pady=(6, 0))
        ttk.Label(
            row1,
            text='安全库存 = K × 平均日均用量 × 交期',
            foreground='#6B7A8F'
        ).grid(row=2, column=6, columnspan=3, sticky='w', pady=(6, 0))

        ttk.Label(
            row1,
            text='采购建议：库存位置 ≤ ROP 时触发，理论采购量只补到 ROP；MOQ/SPQ 只影响建议采购量。',
            foreground='#6B7A8F'
        ).grid(row=3, column=0, columnspan=9, sticky='w', pady=(6, 0))

        self._abc_dedupe_subplan = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            row1,
            text='排产母件去重（下层已由上层覆盖时只展开剩余补排产）',
            variable=self._abc_dedupe_subplan,
        ).grid(row=3, column=9, columnspan=3, sticky='w', padx=(8, 0), pady=(6, 0))

        # 按交期分类的阈值：A 必须 >= B，否则分类会错位
        ttk.Label(row1, text='A类交期阈值（天）≥').grid(row=0, column=2, sticky='w', padx=(0, 4))
        self._lead_a = tk.IntVar(value=60)
        self._lead_a_entry = ttk.Entry(row1, textvariable=self._lead_a, width=6)
        self._lead_a_entry.grid(row=0, column=3, sticky='w', padx=(0, 12))

        ttk.Label(row1, text='B类交期阈值（天）≥').grid(row=0, column=4, sticky='w', padx=(0, 4))
        self._lead_b = tk.IntVar(value=30)
        self._lead_b_entry = ttk.Entry(row1, textvariable=self._lead_b, width=6)
        self._lead_b_entry.grid(row=0, column=5, sticky='w', padx=(0, 12))

        self._lead_c_auto = tk.StringVar(value='30（即B类的下限）')
        ttk.Label(row1, text='C类交期（天）<').grid(row=0, column=6, sticky='w', padx=(0, 4))
        ttk.Label(row1, textvariable=self._lead_c_auto,
                  foreground='#6B7A8F').grid(row=0, column=7, sticky='w', padx=(0, 12))
        self._lead_hint_var = tk.StringVar(value='')
        self._lead_hint_label = ttk.Label(row1, textvariable=self._lead_hint_var,
                                          foreground='#B91C1C',
                                          font=('Microsoft YaHei', 9, 'bold'))
        self._lead_hint_label.grid(row=1, column=2, columnspan=6, sticky='w', pady=(2, 0))

        # 自定义一个红色边框的 Entry 风格，用于阈值非法时高亮
        ttk.Style().configure('Invalid.TEntry', fieldbackground='#FEF2F2', foreground='#B91C1C',
                              bordercolor='#DC2626', lightcolor='#DC2626', darkcolor='#DC2626')

        self._lead_a.trace_add('write', lambda *_: self._validate_lead_thresholds())
        self._lead_b.trace_add('write', lambda *_: self._validate_lead_thresholds())
        self._validate_lead_thresholds()

        ttk.Button(row1, text='执行ABC分类与安全库存计算',
                   command=self._run_abc_calc,
                   style='Accent.TButton').grid(row=0, column=8, padx=(8, 4))
        ttk.Button(row1, text='导出Excel',
                   command=self._export_abc).grid(row=0, column=9, padx=(0, 4))
        ttk.Button(row1, text='清空结果',
                   style='Quiet.TButton', command=self._clear_abc).grid(row=0, column=10)

        # ABC统计
        self._abc_summary_var = tk.StringVar(value='')
        ttk.Label(row1, textvariable=self._abc_summary_var, foreground='#2563EB',
                  font=('Microsoft YaHei', 10, 'bold')).grid(
            row=0, column=11, sticky='w', padx=(8, 0))

        res_frame = ttk.LabelFrame(f, text='外采物料分析结果', style='Card.TLabelframe', padding=6)
        res_frame.grid(row=2, column=0, sticky='nsew', padx=6, pady=(0, 6))
        res_frame.columnconfigure(0, weight=1)
        res_frame.rowconfigure(0, weight=1)

        ABC_HEADERS = ['分类', '物料编码', '物料分类', '项目', '物料名称', '规格型号', '供应商',
                       '交期(天)', 'SPQ', 'MOQ',
                       '月均用量', '年用量',
                       '安全库存', '再订货点(ROP)', '当前库存', '未清PO', '未转PR', '库存位置',
                       '是否触发采购', '理论采购量', '建议采购量', '采购建议说明',
                       '安全库存覆盖天数', '安全库存判断',
                       '使用项目(点击展开)']
        ABC_WIDTHS  = [50, 130, 90, 150, 160, 140, 100,
                       60, 50, 50,
                       75, 75,
                       65, 75, 80, 80, 80, 90, 95, 90, 90, 260, 75, 105,
                       200]

        self.abc_tree = ttk.Treeview(res_frame, columns=ABC_HEADERS,
                                      show='tree headings', selectmode='extended')
        for h, w in zip(ABC_HEADERS, ABC_WIDTHS):
            self.abc_tree.column(h, width=w, anchor='w')
            self.abc_tree.heading(h, text=h)
        self.abc_tree.column('#0', width=0, stretch=False)
        enable_treeview_sort(self.abc_tree, ABC_HEADERS,
                             numeric_columns={'交期(天)', 'SPQ', 'MOQ', '月均用量', '年用量',
                                               '安全库存', '再订货点(ROP)', '当前库存', '未清PO', '未转PR', '库存位置',
                                               '理论采购量', '建议采购量',
                                               '安全库存覆盖天数'})
        enable_treeview_copy(self.abc_tree, ABC_HEADERS)

        style2 = ttk.Style()
        style2.configure('Treeview', rowheight=32)
        style2.map('Treeview', background=[('selected', '#2563EB')])

        vsb3 = ttk.Scrollbar(res_frame, orient='vertical', command=self.abc_tree.yview)
        hsb3 = ttk.Scrollbar(res_frame, orient='horizontal', command=self.abc_tree.xview)
        self.abc_tree.configure(yscrollcommand=vsb3.set, xscrollcommand=hsb3.set)
        self.abc_tree.grid(row=0, column=0, sticky='nsew')
        vsb3.grid(row=0, column=1, sticky='ns')
        hsb3.grid(row=1, column=0, sticky='ew')

        self.abc_tree.bind('<Double-Button-1>', self._copy_abc_cell)
        self.abc_tree.bind('<Button-3>', self._show_abc_menu)

        self._abc_ctx = tk.Menu(self.abc_tree, tearoff=0)
        self._abc_ctx.add_command(label='复制整行', command=self._copy_abc_row)
        self._abc_ctx.add_command(label='复制料号', command=self._copy_abc_pn)
        self._abc_ctx.add_separator()
        self._abc_ctx.add_command(label='🌲 展开该物料BOM', command=self._bom_lookup_from_abc)

    # ── 上传排产计划 ──────────────────────────────────────────
    def _upload_po_plan(self):
        path = filedialog.askopenfilename(
            title='选择排产计划文件',
            filetypes=[('Excel', '*.xlsx *.xls'), ('CSV', '*.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            rows = []
            if ext in ('.xlsx', '.xls'):
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            elif ext == '.csv':
                with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                    rows = list(csv.reader(f))

            if not rows:
                raise ValueError('文件为空')

            # 找表头行（第一行包含 '母件料号' 或 'NO.'）
            header_row_idx = -1
            for idx, row in enumerate(rows):
                if row and len(row) >= 5:
                    first = str(row[0]).strip() if row[0] else ''
                    if '母件料号' in str(row) or first == 'NO.':
                        header_row_idx = idx
                        break
            if header_row_idx < 0:
                raise ValueError('未找到排产计划表头（需包含"母件料号"列）')

            headers = [str(h).strip() if h else '' for h in rows[header_row_idx]]

            # 建立列索引
            col_map = {}
            for ci, h in enumerate(headers):
                h2 = h.replace('\n', '').strip()
                if '母件料号' in h2 or '母件编码' in h2 or '料号' in h2:
                    col_map['code'] = ci
                elif '上线数量' in h2 or '计划数量' in h2 or '数量' in h2:
                    col_map['qty'] = ci
                elif '上线日期' in h2 or '开工日期' in h2 or '日期' in h2:
                    col_map['date'] = ci
                elif '客户' in h2 or '项目' in h2:
                    col_map['project'] = ci
                elif '母件规格' in h2 or '母料规格' in h2 or h2 == '规格' or '规格型号' in h2:
                    col_map['mother_spec'] = ci

            if 'code' not in col_map or 'qty' not in col_map:
                raise ValueError(f'未找到母件料号或数量列，表头：{headers[:8]}')

            from datetime import datetime
            result = []
            raw_rows = []   # 保存含日期的原始行，用于计算排产周期
            all_dates = []

            for row in rows[header_row_idx + 1:]:
                if not row or len(row) < 2:
                    continue
                pn = str(row[col_map['code']]).strip() if col_map['code'] < len(row) else ''
                qty_raw = row[col_map['qty']] if col_map['qty'] < len(row) else None
                if not pn or '合计' in pn or '总计' in pn:
                    continue
                try:
                    qty = float(str(qty_raw).replace(',', '').strip()) if qty_raw else 0
                except (ValueError, TypeError):
                    continue
                if qty <= 0:
                    continue
                result.append((pn, qty))
                project = ''
                if 'project' in col_map and col_map['project'] < len(row):
                    project_raw = row[col_map['project']]
                    project = '' if project_raw is None else str(project_raw).strip()
                mother_spec = ''
                if 'mother_spec' in col_map and col_map['mother_spec'] < len(row):
                    mother_spec_raw = row[col_map['mother_spec']]
                    mother_spec = '' if mother_spec_raw is None else str(mother_spec_raw).strip()

                # 收集日期
                d = None
                if 'date' in col_map and col_map['date'] < len(row):
                    d_raw = row[col_map['date']]
                    if isinstance(d_raw, datetime):
                        d = d_raw
                    elif isinstance(d_raw, str) and d_raw.strip():
                        try:
                            d = datetime.strptime(d_raw.strip(), '%Y-%m-%d')
                        except (ValueError, TypeError):
                            d = None
                raw_rows.append({'pn': pn, 'qty': qty, 'date': d, 'project': project, 'mother_spec': mother_spec})
                if d:
                    all_dates.append(d)

            if not result:
                raise ValueError('文件中未读取到有效数据')

            # 计算实际排产天数（优先使用手动设置值）
            manual_val = self._plan_days_manual.get().strip()
            if manual_val:
                try:
                    plan_days = max(1, int(manual_val))
                except ValueError:
                    plan_days = 30
            elif len(all_dates) >= 2:
                plan_days = (max(all_dates) - min(all_dates)).days + 1
            else:
                # 单日：如果是当月1日 → 按全月天数计算；否则默认30天
                d = all_dates[0] if all_dates else None
                if d and d.day == 1:
                    plan_days = calendar.monthrange(d.year, d.month)[1]
                else:
                    plan_days = 30

            # 合并同母件数量
            merged = {}
            order = []
            for pn, qty in result:
                if pn not in merged:
                    merged[pn] = 0.0
                    order.append(pn)
                merged[pn] += qty

            self._po_data = [(pn, merged[pn]) for pn in order]
            self._po_raw_rows = raw_rows
            self._plan_days = plan_days
            self._po_count_var.set(f'已上传 {len(self._po_data)} 条记录，排产周期 {plan_days} 天')
            self.status_var.set(f'排产计划已加载：{len(self._po_data)} 条，排产周期 {plan_days} 天（{min(all_dates).strftime("%m/%d") if all_dates else "?"}~{max(all_dates).strftime("%m/%d") if all_dates else "?"})')
        except Exception as e:
            messagebox.showerror('读取失败', str(e))

    # ── 查看/编辑排产计划 ────────────────────────────────────
    def _show_po_plan(self):
        if not self._po_data:
            messagebox.showinfo('提示', '请先上传排产计划')
            return
        win = tk.Toplevel(self.root)
        win.title('排产计划')
        win.geometry('500x500')

        fr = ttk.Frame(win)
        fr.pack(fill='both', expand=True, padx=6, pady=6)
        fr.columnconfigure(0, weight=1)
        fr.rowconfigure(0, weight=1)

        tv = ttk.Treeview(fr, columns=('母件编码', '计划数量'), show='headings')
        tv.column('母件编码', width=200, anchor='w')
        tv.column('计划数量', width=120, anchor='e')
        tv.heading('母件编码', text='母件编码')
        tv.heading('计划数量', text='计划数量')
        enable_treeview_sort(tv, ['母件编码', '计划数量'], numeric_columns={'计划数量'})
        enable_treeview_copy(tv, ['母件编码', '计划数量'])
        tv.grid(row=0, column=0, sticky='nsew')
        ttk.Scrollbar(fr, orient='vertical', command=tv.yview).grid(
            row=0, column=1, sticky='ns')
        for pn, qty in self._po_data:
            tv.insert('', 'end', values=(pn, fmt_qty(qty)))

        def _on_dbl(e):
            region = tv.identify('region', e.x, e.y)
            if region == 'cell':
                col = tv.identify_column(e.x)
                item = tv.identify_row(e.y)
                if item and col == '#2':
                    vals = tv.item(item, 'values')
                    win.destroy()
                    self.pn_entry.delete(0, tk.END)
                    self.pn_entry.insert(0, str(vals[0]).strip())
                    self._switch_to_tab(self.tab_tree)
                    self._query()
        tv.bind('<Double-Button-1>', _on_dbl)
        ttk.Button(fr, text='关闭', command=win.destroy).pack(pady=(4, 0))

    def _on_plan_days_manual_change(self):
        """手动输入排产天数后，立即更新标签显示和内部值"""
        val = self._plan_days_manual.get().strip()
        if val:
            try:
                days = int(val)
                if days > 0:
                    self._plan_days = days
                    self._po_count_var.set(f'已上传 {len(self._po_data)} 条，手动设置 {days} 天')
                    self.status_var.set(f'排产天数已手动设置为 {days} 天')
            except ValueError:
                pass

    def _clear_po(self):
        self._po_data = []
        self._po_raw_rows = []
        self._plan_days = 30
        self._po_count_var.set('未上传')
        self._plan_days_manual.set('')
        self.status_var.set('排产计划已清除')

    # ── 上传外采清单 ─────────────────────────────────────────
    def _upload_ext_list(self):
        path = filedialog.askopenfilename(
            title='选择外采清单文件',
            filetypes=[('Excel', '*.xlsx *.xls'), ('CSV', '*.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            rows = []
            if ext in ('.xlsx', '.xls'):
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                # 遍历所有Sheet，找到包含"物料号"的那个
                target_sheet = None
                for sname in wb.sheetnames:
                    ws_tmp = wb[sname]
                    tmp_rows = list(ws_tmp.iter_rows(values_only=True))
                    if tmp_rows and any('物料号' in str(c) for c in tmp_rows[0]):
                        target_sheet = sname
                        rows = tmp_rows
                        break
                if target_sheet is None:
                    # fallback: 用第一个sheet
                    ws = wb[wb.sheetnames[0]]
                    rows = list(ws.iter_rows(values_only=True))
                wb.close()
            elif ext == '.csv':
                with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                    rows = list(csv.reader(f))
            else:
                raise ValueError('不支持的格式')

            if not rows:
                raise ValueError('文件为空')

            # 找表头行（包含 '物料号' 或 '物料编码'）
            header_row_idx = -1
            for idx, row in enumerate(rows):
                if row and len(row) >= 3:
                    row_str = str(row)
                    if '物料号' in row_str or '物料编码' in row_str:
                        header_row_idx = idx
                        break
            if header_row_idx < 0:
                raise ValueError('未找到外采清单表头（需包含"物料号"列），请确认文件格式为Excel或CSV')

            headers = [str(h).strip() if h else '' for h in rows[header_row_idx]]
            print(f'[DEBUG] 外采清单表头: {headers}')

            # 建立列索引映射（实际文件列）
            col_map = {}
            for ci, h in enumerate(headers):
                h2 = h.replace('\n', '').strip()
                if '物料号' in h2 or '物料编码' in h2 or '件号' in h2:
                    col_map['code'] = ci
                elif '名称' in h2 and '品名' not in h2:
                    col_map['name'] = ci
                elif '规格' in h2:
                    col_map['spec'] = ci
                elif '单机' in h2:   # 单机用量
                    col_map['unit'] = ci
                elif '供应商' in h2:
                    col_map['supplier'] = ci
                elif '提前期' in h2 or '交期' in h2 or '采购周期' in h2 or '天数' in h2:
                    col_map['lead'] = ci
                elif 'SPQ' in h2 or '倍量' in h2 or '包装' in h2:
                    col_map['spq'] = ci
                elif 'MOQ' in h2 or '起订' in h2 or '最少' in h2:
                    col_map['moq'] = ci
                elif '备注' in h2:
                    col_map['remark'] = ci
                elif ('采购' in h2 or '买手' in h2) and not any(
                    keyword in h2 for keyword in ('周期', '交期', '提前', '数量', '订单', '单号', '计划', '供应商')
                ):
                    col_map['buyer'] = ci

            print(f'[DEBUG] col_map: {col_map}')
            if 'code' not in col_map:
                raise ValueError(f'未找到"物料号"列，表头：{headers[:10]}')

            self._ext_data = {}
            for row in rows[header_row_idx + 1:]:
                if not row or len(row) < 2:
                    continue
                code_idx = col_map.get('code', 0)
                code = str(row[code_idx]).strip() if code_idx < len(row) and row[code_idx] else ''
                if not code or code in ('物料号', '—', '', '序号'):
                    continue

                name = ''
                if 'name' in col_map:
                    ni = col_map['name']
                    name = str(row[ni]).strip() if ni < len(row) and row[ni] else ''

                spec = ''
                if 'spec' in col_map:
                    si = col_map['spec']
                    spec = str(row[si]).strip() if si < len(row) and row[si] else ''

                unit_qty = 0.0
                if 'unit' in col_map:
                    ui = col_map['unit']
                    try:
                        unit_qty = float(row[ui]) if row[ui] else 0.0
                    except (ValueError, TypeError):
                        unit_qty = 0.0

                supplier = ''
                if 'supplier' in col_map:
                    si = col_map['supplier']
                    supplier = str(row[si]).strip() if si < len(row) and row[si] else ''

                buyer = ''
                if 'buyer' in col_map:
                    bi = col_map['buyer']
                    buyer = str(row[bi]).strip() if bi < len(row) and row[bi] else ''

                lead_days = 0
                if 'lead' in col_map:
                    li = col_map['lead']
                    try:
                        lead_days = int(float(str(row[li]).strip()))
                    except (ValueError, TypeError):
                        lead_days = 0

                spq = ''
                if 'spq' in col_map:
                    shi = col_map['spq']
                    spq = str(row[shi]).strip() if shi < len(row) and row[shi] else ''

                moq = ''
                if 'moq' in col_map:
                    mi = col_map['moq']
                    moq = str(row[mi]).strip() if mi < len(row) and row[mi] else ''

                remark = ''
                if 'remark' in col_map:
                    ri = col_map['remark']
                    remark = str(row[ri]).strip() if ri < len(row) and row[ri] else ''

                self._ext_data[code] = {
                    'name': name, 'spec': spec,
                    'unit_qty': unit_qty, 'supplier': supplier,
                    'buyer': buyer,
                    'lead_days': lead_days, 'spq': spq, 'moq': moq,
                    'remark': remark
                }

            self._ext_count_var.set(f'已上传 {len(self._ext_data)} 条物料')
            self.status_var.set(f'外采清单已加载：{len(self._ext_data)} 条（含SPQ/MOQ/交期信息）')
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror('读取失败', str(e))

    def _clear_ext(self):
        self._ext_data = {}
        self._ext_count_var.set('未上传')
        self.status_var.set('外采清单已清除')

    # ── ABC分类 + 安全库存计算 ───────────────────────────────
    def _choose_abc_mrp_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP 计算表',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self._abc_mrp_path = path
        self._abc_mrp_var.set(os.path.basename(path))
        ensure_msgs = [
            msg for msg in [
                self._ensure_abc_model_switch_sheet(path),
                self._ensure_abc_custom_stock_sheet(path),
            ]
            if msg
        ]
        if ensure_msgs:
            self.status_var.set(f'ABC 已关联 MRP 计算表：{os.path.basename(path)}；{"；".join(ensure_msgs)}')
        else:
            self.status_var.set(f'ABC 已关联 MRP 计算表：{os.path.basename(path)}')

    def _clear_abc_mrp(self):
        self._abc_mrp_path = ''
        self._abc_mrp_var.set('未选择（不带库存位置/排除清单/物料分类）')
        self.status_var.set('ABC 的 MRP 计算表已清除')

    def _load_abc_mrp_supply_maps(self, workbook_path):
        inventory_map = {}
        po_map = {}
        pr_map = {}
        if not workbook_path:
            return inventory_map, po_map, pr_map

        required_sheets = ['期初库存', '在途采购', '在途请购']
        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表：{exc}') from exc

        missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheet_names]
        if missing_sheets:
            raise ValueError(f'MRP 计算表缺少工作表：{", ".join(missing_sheets)}')

        def _build_map(sheet_name, code_candidates, qty_candidates, *, qty_contains=False):
            df = normalize_sheet_columns(pd.read_excel(workbook_path, sheet_name=sheet_name))
            code_col = find_first_matching_column(df, code_candidates, contains=False, required=True)
            qty_col = find_first_matching_column(df, qty_candidates, contains=qty_contains, required=True)
            temp = pd.DataFrame({
                'code': df[code_col].map(normalize_material_code),
                'qty': pd.to_numeric(df[qty_col], errors='coerce').fillna(0),
            })
            temp = temp[temp['code'].astype(str).str.strip() != '']
            if temp.empty:
                return {}
            return temp.groupby('code', dropna=False)['qty'].sum().to_dict()

        inventory_map = _build_map('期初库存', ['物料编码'], ['库存量'], qty_contains=True)
        po_map = _build_map('在途采购', ['料号'], ['欠交数量'], qty_contains=True)
        pr_map = _build_map('在途请购', ['料号'], ['未转PO数量'], qty_contains=True)
        return inventory_map, po_map, pr_map

    def _load_abc_material_price_map(self, workbook_path):
        """从MRP计算表自动识别“料号/价格”页，返回 {物料编码: 单价}。"""
        price_map = {}
        if not workbook_path:
            return price_map

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception:
            return price_map

        try:
            sheet_names = list(workbook.sheet_names)
            preferred = [
                name for name in sheet_names
                if any(keyword in str(name) for keyword in ['价格', '单价', '物料价格', '采购价格'])
            ]
            candidates = preferred + [name for name in sheet_names if name not in preferred]
            for sheet_name in candidates:
                try:
                    df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
                except Exception:
                    continue
                if df.empty or df.dropna(how='all').empty:
                    continue
                code_col = find_first_matching_column(
                    df,
                    ['料号', '物料编码', '物料号', '物料代码', '编码'],
                    required=False,
                )
                price_col = find_first_matching_column(
                    df,
                    ['价格', '单价', '采购单价', '含税单价', '未税单价', '不含税单价', '最新单价'],
                    contains=True,
                    required=False,
                )
                if not code_col or not price_col:
                    continue

                temp = pd.DataFrame({
                    'code': df[code_col].map(normalize_material_code),
                    'price': (
                        df[price_col]
                        .astype(str)
                        .str.replace(',', '', regex=False)
                        .str.replace(r'[^\d\.\-]', '', regex=True)
                    ),
                })
                temp['price'] = pd.to_numeric(temp['price'], errors='coerce').fillna(0)
                temp = temp[temp['code'].astype(str).str.strip() != '']
                if temp.empty:
                    continue
                for _, row in temp.iterrows():
                    price_map[str(row['code']).strip()] = float(row['price'] or 0)
                if price_map:
                    return price_map
            return price_map
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _load_abc_mrp_po_arrival_map(self, workbook_path):
        """读取未清PO的预计到货日期明细；没有日期的未清PO不参与生产风险月度覆盖。"""
        arrival_map = defaultdict(list)
        if not workbook_path:
            return arrival_map

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception:
            return arrival_map
        try:
            if '在途采购' not in workbook.sheet_names:
                return arrival_map
            df = normalize_sheet_columns(pd.read_excel(workbook_path, sheet_name='在途采购', dtype=object))
            code_col = find_first_matching_column(df, ['料号', '物料编码', '物料号'], required=False)
            qty_col = find_first_matching_column(df, ['欠交数量', '未交数量', '未清数量', '数量'], contains=True, required=False)
            date_col = find_first_matching_column(
                df,
                ['预计到货日期', '到货日期', '交货日期', '计划到货日期', '计划交货日期', '承诺到货日期', '需求日期'],
                contains=False,
                required=False,
            )
            if not date_col:
                date_col = find_first_matching_column(
                    df,
                    ['到货', '交货', '到厂', '需求日期'],
                    contains=True,
                    required=False,
                )
            if not code_col or not qty_col or not date_col:
                return arrival_map
            for _, row in df.iterrows():
                code = normalize_material_code(row.get(code_col))
                if not code:
                    continue
                try:
                    qty = float(row.get(qty_col, 0) or 0)
                except (TypeError, ValueError):
                    qty = 0
                if qty <= 0:
                    continue
                raw_date = row.get(date_col)
                parsed = pd.to_datetime(raw_date, errors='coerce')
                if pd.isna(parsed):
                    continue
                arrival_map[code].append({
                    'arrival_date': parsed.date(),
                    'qty': qty,
                    'source': '未清PO',
                })
            for items in arrival_map.values():
                items.sort(key=lambda item: item['arrival_date'])
            return arrival_map
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _load_abc_exclusion_rules(self, workbook_path):
        excluded_codes = set()
        usage_exclusion_pairs = set()
        if not workbook_path:
            return excluded_codes, usage_exclusion_pairs

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取排除清单：{exc}') from exc

        try:
            if '物料编码排除清单' in workbook.sheet_names:
                material_df = pd.read_excel(
                    workbook,
                    sheet_name='物料编码排除清单',
                    header=None,
                    dtype=object,
                )
                for value in material_df.to_numpy().ravel():
                    material_code = maybe_material_code(value)
                    if material_code:
                        excluded_codes.add(material_code)

            if '母料用量排除清单' in workbook.sheet_names:
                usage_df = normalize_sheet_columns(
                    pd.read_excel(workbook, sheet_name='母料用量排除清单', dtype=object)
                )
                parent_col = find_first_matching_column(
                    usage_df,
                    ['母件料号', '母料号', '父项料号', '父项物料编码', '上层物料编码'],
                    required=False,
                )
                child_col = find_first_matching_column(
                    usage_df,
                    ['子件料号', '子料号', '子项料号', '子项物料编码', '下层物料编码', '物料编码'],
                    required=False,
                )

                non_empty_rows = usage_df.dropna(how='all')
                if non_empty_rows.empty:
                    return excluded_codes, usage_exclusion_pairs
                if not parent_col or not child_col:
                    raise ValueError('母料用量排除清单缺少字段：需要“母件料号”和“子件料号”两列')

                for _, row in usage_df.iterrows():
                    parent_code = maybe_material_code(row.get(parent_col))
                    child_code = maybe_material_code(row.get(child_col))
                    if parent_code and child_code:
                        usage_exclusion_pairs.add((parent_code, child_code))
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        return excluded_codes, usage_exclusion_pairs

    def _load_abc_replacement_diff_rules(self, workbook_path):
        """读取产品替换差异采购规则：替换母件只采购相对基准母件多出来的物料。"""
        if not workbook_path:
            return {}

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取产品替换差异采购清单：{exc}') from exc

        try:
            sheet_name = next(
                (
                    name for name in [
                        '产品替换差异采购清单',
                        '产品替换差异清单',
                        '替换差异采购清单',
                        '替换关系清单',
                        '产品替换清单',
                    ]
                    if name in workbook.sheet_names
                ),
                None,
            )
            if not sheet_name:
                return {}

            df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
            non_empty_rows = df.dropna(how='all')
            if non_empty_rows.empty:
                return {}

            base_col = find_first_matching_column(
                df,
                ['基准母件料号', '基准母料号', '原母件料号', '原母料号', '原产品料号', '被替换母件', 'A母件料号'],
                required=False,
            )
            replacement_col = find_first_matching_column(
                df,
                ['替换母件料号', '替换母料号', '新母件料号', '新母料号', '新产品料号', '替代母件', 'A1母件料号'],
                required=False,
            )

            if base_col and replacement_col:
                source_pairs = [(row.get(base_col), row.get(replacement_col)) for _, row in df.iterrows()]
            else:
                raw_df = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object).dropna(how='all')
                if raw_df.shape[1] < 2:
                    raise ValueError(f'{sheet_name} 缺少字段：需要“基准母件料号”和“替换母件料号”两列')
                source_pairs = list(zip(raw_df.iloc[:, 0], raw_df.iloc[:, 1]))

            rules = {}
            for raw_base, raw_replacement in source_pairs:
                base_code = maybe_material_code(raw_base)
                replacement_code = maybe_material_code(raw_replacement)
                if not base_code or not replacement_code or base_code == replacement_code:
                    continue
                rules.setdefault(replacement_code, base_code)
            return rules
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _load_abc_material_category_rules(self, workbook_path):
        if not workbook_path:
            return []

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取物料分类清单：{exc}') from exc

        try:
            sheet_name = next(
                (
                    name for name in ['物料分类清单', '物料分类规则', '物料类别清单', '物料类别规则', 'ABC物料分类']
                    if name in workbook.sheet_names
                ),
                None,
            )
            if not sheet_name:
                return []
            df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
            prefix_col = find_first_matching_column(
                df,
                ['物料编码前缀', '料号前缀', '物料前缀', '编码前缀', '前缀', '物料编码'],
                required=False,
            )
            category_col = find_first_matching_column(
                df,
                ['物料分类', '分类', '类别', '物料类别', '品类'],
                required=False,
            )
            non_empty_rows = df.dropna(how='all')
            if non_empty_rows.empty:
                return []

            source_pairs = []
            if prefix_col and category_col:
                source_pairs = [(row.get(prefix_col), row.get(category_col)) for _, row in df.iterrows()]
            else:
                raw_df = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
                raw_df = raw_df.dropna(how='all')
                if raw_df.shape[1] < 2:
                    raise ValueError(f'{sheet_name} 缺少字段：需要“物料编码前缀”和“物料分类”两列')
                source_pairs = list(zip(raw_df.iloc[:, 0], raw_df.iloc[:, 1]))
            rules = []
            for raw_prefix, raw_category in source_pairs:
                if raw_prefix is None or pd.isna(raw_prefix):
                    continue
                if isinstance(raw_prefix, float) and raw_prefix.is_integer():
                    prefix = str(int(raw_prefix))
                else:
                    prefix = str(raw_prefix).strip()
                    if re.fullmatch(r'\d+\.0+', prefix):
                        prefix = prefix.split('.', 1)[0]
                if any(keyword in prefix for keyword in ['前缀', '物料编码', '料号']):
                    continue
                category = '' if raw_category is None or pd.isna(raw_category) else str(raw_category).strip()
                if prefix and category:
                    rules.append((prefix, category))
            rules.sort(key=lambda item: len(item[0]), reverse=True)
            return rules
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _match_abc_material_category(self, material_code, category_rules):
        code = str(material_code or '').strip()
        if not code:
            return ''
        for prefix, category in category_rules:
            if code.startswith(prefix):
                return category
        return ''

    def _load_abc_clear_tail_rules(self, workbook_path):
        """读取型号切换清尾物料：命中后不补ROP，只按排产净需求下单。"""
        if not workbook_path:
            return []

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取切换清尾物料清单：{exc}') from exc

        try:
            sheet_name = next(
                (
                    name for name in [
                        '切换清尾物料清单',
                        '型号切换清尾物料清单',
                        '清尾物料清单',
                        '切换物料清单',
                    ]
                    if name in workbook.sheet_names
                ),
                None,
            )
            if not sheet_name:
                return []

            df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
            if df.dropna(how='all').empty:
                return []
            code_col = find_first_matching_column(
                df,
                ['物料编码', '物料号', '料号', '物料编码前缀', '料号前缀', '前缀'],
                required=False,
            )
            cutoff_col = find_first_matching_column(
                df,
                ['清尾截止月份', '截止月份', '停用月份', '最后需求月份', '截止月'],
                required=False,
            )
            remark_col = find_first_matching_column(df, ['备注', '说明'], required=False)

            source_rows = []
            if code_col:
                for _, row in df.iterrows():
                    source_rows.append((
                        row.get(code_col),
                        row.get(cutoff_col) if cutoff_col else '',
                        row.get(remark_col) if remark_col else '',
                    ))
            else:
                raw_df = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object).dropna(how='all')
                if raw_df.empty:
                    return []
                for _, row in raw_df.iterrows():
                    source_rows.append((
                        row.iloc[0] if len(row) > 0 else '',
                        row.iloc[1] if len(row) > 1 else '',
                        row.iloc[2] if len(row) > 2 else '',
                    ))

            def _parse_cutoff_month(value):
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    return ''
                if isinstance(value, (datetime, date)):
                    return f'{value.year:04d}-{value.month:02d}'
                text = str(value).strip()
                if not text or any(keyword in text for keyword in ['截止', '月份', '停用']):
                    return ''
                for fmt in ['%Y-%m', '%Y/%m', '%Y.%m', '%Y年%m月']:
                    try:
                        dt = datetime.strptime(text, fmt)
                        return f'{dt.year:04d}-{dt.month:02d}'
                    except ValueError:
                        pass
                match = re.search(r'(20\d{2})\D{0,2}([01]?\d)', text)
                if match:
                    year = int(match.group(1))
                    month = int(match.group(2))
                    if 1 <= month <= 12:
                        return f'{year:04d}-{month:02d}'
                return ''

            rules = []
            for raw_code, raw_cutoff, raw_remark in source_rows:
                if raw_code is None or (isinstance(raw_code, float) and pd.isna(raw_code)):
                    continue
                code_text = str(raw_code).strip()
                if not code_text or any(keyword in code_text for keyword in ['物料', '料号', '前缀']):
                    continue
                if isinstance(raw_code, float) and raw_code.is_integer():
                    code_text = str(int(raw_code))
                elif re.fullmatch(r'\d+\.0+', code_text):
                    code_text = code_text.split('.', 1)[0]
                material_code = maybe_material_code(code_text) or normalize_material_code(code_text) or code_text
                if not material_code:
                    continue
                remark = '' if raw_remark is None or (isinstance(raw_remark, float) and pd.isna(raw_remark)) else str(raw_remark).strip()
                rules.append({
                    'prefix': material_code,
                    'cutoff_month': _parse_cutoff_month(raw_cutoff),
                    'remark': remark,
                })
            rules.sort(key=lambda item: len(item['prefix']), reverse=True)
            return rules
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _ensure_abc_model_switch_sheet(self, workbook_path):
        """在MRP计算表里补一张型号切换计划空表，便于直接维护切换关系。"""
        if not workbook_path:
            return ''
        suffix = os.path.splitext(str(workbook_path))[1].lower()
        if suffix not in ('.xlsx', '.xlsm'):
            return ''
        wb = None
        try:
            wb = openpyxl.load_workbook(workbook_path)
            headers = ['原母料号', '原型号', '新母料号', '新型号', '切换日期']
            widths = [18, 24, 18, 24, 14]
            if '型号切换计划' in wb.sheetnames:
                ws = wb['型号切换计划']
                current_headers = [str(ws.cell(row=1, column=ci).value or '').strip() for ci in range(1, 6)]
                if current_headers[:5] != headers:
                    old_headers = [str(ws.cell(row=1, column=ci).value or '').strip() for ci in range(1, 5)]
                    if old_headers[:4] == ['原母料号', '新母料号', '切换日期', '备注']:
                        for row_idx in range(2, ws.max_row + 1):
                            old_new_pn = ws.cell(row=row_idx, column=2).value
                            old_switch_date = ws.cell(row=row_idx, column=3).value
                            ws.cell(row=row_idx, column=2).value = None
                            ws.cell(row=row_idx, column=3, value=old_new_pn)
                            ws.cell(row=row_idx, column=4).value = None
                            ws.cell(row=row_idx, column=5, value=old_switch_date)
                    for ci, header in enumerate(headers, start=1):
                        cell = ws.cell(row=1, column=ci, value=header)
                        cell.font = Font(bold=True, color='000000')
                        cell.fill = PatternFill('solid', fgColor='D9EAF7')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    for ci, width in enumerate(widths, start=1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
                    ws.freeze_panes = 'A2'
                    wb.save(workbook_path)
                    wb.close()
                    return '已更新“型号切换计划”表头'
                wb.close()
                return ''
            ws = wb.create_sheet('型号切换计划')
            for ci, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=ci, value=header)
                cell.font = Font(bold=True, color='000000')
                cell.fill = PatternFill('solid', fgColor='D9EAF7')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for ci, width in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
            ws.freeze_panes = 'A2'
            wb.save(workbook_path)
            wb.close()
            return '已补充“型号切换计划”sheet'
        except Exception as exc:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
            return f'“型号切换计划”sheet未自动写入：{exc}'

    def _ensure_abc_custom_stock_sheet(self, workbook_path):
        """在MRP计算表里补一张客户定制机型备货控制空表。"""
        if not workbook_path:
            return ''
        suffix = os.path.splitext(str(workbook_path))[1].lower()
        if suffix not in ('.xlsx', '.xlsm'):
            return ''
        wb = None
        try:
            wb = openpyxl.load_workbook(workbook_path)
            sheet_name = '客户定制机型备货控制'
            headers = ['料号', '型号', '数量', '客户']
            widths = [18, 24, 12, 18]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                current_headers = [str(ws.cell(row=1, column=ci).value or '').strip() for ci in range(1, 5)]
                if current_headers[:4] == headers:
                    wb.close()
                    return ''
            else:
                ws = wb.create_sheet(sheet_name)
            for ci, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=ci, value=header)
                cell.font = Font(bold=True, color='000000')
                cell.fill = PatternFill('solid', fgColor='D9EAF7')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for ci, width in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width
            ws.freeze_panes = 'A2'
            wb.save(workbook_path)
            wb.close()
            return f'已补充/更新“{sheet_name}”sheet'
        except Exception as exc:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
            return f'“客户定制机型备货控制”sheet未自动写入：{exc}'

    def _load_abc_custom_stock_rules(self, workbook_path):
        """读取客户定制机型备货控制：限制指定母料号参与采购计算的有效排产数量。"""
        if not workbook_path:
            return []
        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取客户定制机型备货控制：{exc}') from exc

        try:
            sheet_name = next(
                (
                    name for name in [
                        '客户定制机型备货控制',
                        '客户定制备货控制',
                        '定制机型备货控制',
                        '定制机型控制',
                    ]
                    if name in workbook.sheet_names
                ),
                None,
            )
            if not sheet_name:
                return []
            df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
            if df.dropna(how='all').empty:
                return []
            code_col = find_first_matching_column(df, ['料号', '母料号', '母件料号', '机型料号', '物料编码'], required=False)
            model_col = find_first_matching_column(df, ['型号', '机型', '规格型号', '名称'], required=False)
            qty_col = find_first_matching_column(df, ['数量', '备货数量', '控制数量', '上限数量'], required=False)
            customer_col = find_first_matching_column(df, ['客户', '项目', '客户/项目'], required=False)
            if not code_col or not qty_col:
                return []
            rules = []
            for _, row in df.iterrows():
                raw_code = row.get(code_col)
                if raw_code is None or (isinstance(raw_code, float) and pd.isna(raw_code)):
                    continue
                code_text = str(raw_code).strip()
                if not code_text or any(keyword in code_text for keyword in ['料号', '编码', '合计']):
                    continue
                if isinstance(raw_code, float) and raw_code.is_integer():
                    code_text = str(int(raw_code))
                elif re.fullmatch(r'\d+\.0+', code_text):
                    code_text = code_text.split('.', 1)[0]
                material_code = maybe_material_code(code_text) or normalize_material_code(code_text) or code_text
                try:
                    qty = float(str(row.get(qty_col, 0)).replace(',', '').strip())
                except (TypeError, ValueError):
                    qty = 0.0
                if not material_code or qty <= 0:
                    continue
                model = '' if not model_col or pd.isna(row.get(model_col)) else str(row.get(model_col)).strip()
                customer = '' if not customer_col or pd.isna(row.get(customer_col)) else str(row.get(customer_col)).strip()
                rules.append({
                    'pn': material_code,
                    'model': model,
                    'qty': qty,
                    'customer': customer,
                })
            return rules
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _load_abc_model_switch_rules(self, workbook_path):
        """读取型号切换计划：原母料号 -> 新母料号，切换后旧专用/降用量物料按清尾处理。"""
        if not workbook_path:
            return []

        try:
            workbook = pd.ExcelFile(workbook_path)
        except Exception as exc:
            raise ValueError(f'无法打开 MRP 计算表读取型号切换计划：{exc}') from exc

        try:
            sheet_name = next(
                (
                    name for name in [
                        '型号切换计划',
                        '配置切换计划',
                        '产品切换计划',
                        '母料切换计划',
                    ]
                    if name in workbook.sheet_names
                ),
                None,
            )
            if not sheet_name:
                return []

            df = normalize_sheet_columns(pd.read_excel(workbook, sheet_name=sheet_name, dtype=object))
            if df.dropna(how='all').empty:
                return []

            old_col = find_first_matching_column(
                df,
                ['原母料号', '旧母料号', '原母件料号', '旧母件料号', '原型号母料号', 'A母料号', '基准母料号'],
                required=False,
            )
            new_col = find_first_matching_column(
                df,
                ['新母料号', '替换母料号', '新母件料号', '替换母件料号', '新型号母料号', 'B母料号'],
                required=False,
            )
            switch_col = find_first_matching_column(
                df,
                ['切换日期', '切换月份', '生效日期', '开始日期', '停用日期'],
                required=False,
            )
            remark_col = find_first_matching_column(df, ['备注', '说明'], required=False)

            source_rows = []
            if old_col and new_col:
                for _, row in df.iterrows():
                    source_rows.append((
                        row.get(old_col),
                        row.get(new_col),
                        row.get(switch_col) if switch_col else '',
                        row.get(remark_col) if remark_col else '',
                    ))
            else:
                raw_df = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object).dropna(how='all')
                if raw_df.shape[1] < 2:
                    raise ValueError(f'{sheet_name} 缺少字段：需要“原母料号”和“新母料号”两列')
                for _, row in raw_df.iterrows():
                    source_rows.append((
                        row.iloc[0] if len(row) > 0 else '',
                        row.iloc[1] if len(row) > 1 else '',
                        row.iloc[2] if len(row) > 2 else '',
                        row.iloc[3] if len(row) > 3 else '',
                    ))

            def _parse_switch_date(value):
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    return None, ''
                if isinstance(value, datetime):
                    return value.date(), f'{value.year:04d}-{value.month:02d}'
                if isinstance(value, date):
                    return value, f'{value.year:04d}-{value.month:02d}'
                text = str(value).strip()
                if not text or any(keyword in text for keyword in ['切换', '日期', '月份', '生效']):
                    return None, ''
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日']:
                    try:
                        dt = datetime.strptime(text, fmt)
                        return dt.date(), f'{dt.year:04d}-{dt.month:02d}'
                    except ValueError:
                        pass
                for fmt in ['%Y-%m', '%Y/%m', '%Y.%m', '%Y年%m月']:
                    try:
                        dt = datetime.strptime(text, fmt)
                        return date(dt.year, dt.month, 1), f'{dt.year:04d}-{dt.month:02d}'
                    except ValueError:
                        pass
                match = re.search(r'(20\d{2})\D{0,2}([01]?\d)(?:\D{0,2}([0-3]?\d))?', text)
                if match:
                    year = int(match.group(1))
                    month = int(match.group(2))
                    day = int(match.group(3) or 1)
                    if 1 <= month <= 12:
                        day = min(max(day, 1), calendar.monthrange(year, month)[1])
                        return date(year, month, day), f'{year:04d}-{month:02d}'
                return None, ''

            rules = []
            for raw_old, raw_new, raw_date, raw_remark in source_rows:
                old_pn = maybe_material_code(raw_old) or normalize_material_code(raw_old)
                new_pn = maybe_material_code(raw_new) or normalize_material_code(raw_new)
                if not old_pn or not new_pn or old_pn == new_pn:
                    continue
                if any(keyword in old_pn for keyword in ['原母', '旧母', 'A母']) or any(keyword in new_pn for keyword in ['新母', '替换', 'B母']):
                    continue
                switch_date, switch_month = _parse_switch_date(raw_date)
                remark = '' if raw_remark is None or (isinstance(raw_remark, float) and pd.isna(raw_remark)) else str(raw_remark).strip()
                rules.append({
                    'old_pn': old_pn,
                    'new_pn': new_pn,
                    'switch_date': switch_date,
                    'switch_month': switch_month,
                    'remark': remark,
                })
            return rules
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def _run_abc_calc(self):
            if not self._po_data:
                messagebox.showwarning('缺少数据', '请先上传排产计划')
                return
            if not self._ext_data:
                messagebox.showwarning('缺少数据', '请先上传外采清单')
                return
            if not self.bom_index:
                messagebox.showwarning('缺少数据', '请先上传 BOM 文件')
                return
            try:
                a_val = self._lead_a.get()
                b_val = self._lead_b.get()
                k_a = float(self._abc_k_a.get())
                k_b = float(self._abc_k_b.get())
                k_c = float(self._abc_k_c.get())
            except tk.TclError:
                messagebox.showwarning('参数无效', 'ABC 阈值不是有效的整数')
                return
            if a_val < 0 or b_val < 0 or a_val < b_val:
                messagebox.showwarning(
                    '参数无效',
                    f'A 类阈值 ({a_val}) 必须 ≥ B 类阈值 ({b_val})，且均不能为负数。',
                )
                return
            if k_a < 0 or k_b < 0 or k_c < 0:
                messagebox.showwarning('鍙傛暟鏃犳晥', 'A / B / C 绫?K 鍊间笉鑳戒负璐熸暟')
                return

            import math
            from datetime import datetime
            self.status_var.set('正在计算外采物料需求 \u2026')
            self.root.update_idletasks()

            # ── 第一步：BOM展开，算出每个物料的汇总需求量 + 使用项目 ─────────
            flat = {}
            source_requirement_buckets = defaultdict(dict)
            # parents_map: {子件料号: {母件料号, ...}}  记录每个子件被哪些母件使用
            parents_map = {}
            monthly_demand_map = defaultdict(lambda: defaultdict(float))
            project_map = defaultdict(set)
            monthly_project_map = defaultdict(lambda: defaultdict(set))
            parent_spec_map = defaultdict(set)
            monthly_parent_spec_map = defaultdict(lambda: defaultdict(set))

            cycle_hits: set[tuple[str, str]] = set()
            excluded_codes, usage_exclusion_pairs = self._load_abc_exclusion_rules(self._abc_mrp_path)
            replacement_diff_rules = self._load_abc_replacement_diff_rules(self._abc_mrp_path)
            clear_tail_rules = self._load_abc_clear_tail_rules(self._abc_mrp_path)
            manual_clear_tail_rule_count = len(clear_tail_rules)
            model_switch_rules = self._load_abc_model_switch_rules(self._abc_mrp_path)
            custom_stock_rules = self._load_abc_custom_stock_rules(self._abc_mrp_path)

            def _code_keys(value):
                text = str(value or '').strip()
                keys = {text} if text else set()
                material_code = maybe_material_code(value)
                if material_code:
                    keys.add(material_code)
                normalized = normalize_material_code(value)
                if normalized:
                    keys.add(normalized)
                return keys

            bom_key_by_code = {}
            for bom_key in self.bom_index.keys():
                for key in _code_keys(bom_key):
                    bom_key_by_code.setdefault(key, bom_key)

            def _resolve_bom_key(value):
                for key in _code_keys(value):
                    if key in self.bom_index:
                        return key
                    if key in bom_key_by_code:
                        return bom_key_by_code[key]
                return None

            def _canonical_code(value):
                material_code = maybe_material_code(value)
                if material_code:
                    return material_code
                normalized = normalize_material_code(value)
                if normalized:
                    return normalized
                return str(value or '').strip()

            def _lookup_by_code_keys(mapping, value, default=None):
                for key in _code_keys(value):
                    if key in mapping:
                        return mapping[key]
                return default

            def _match_clear_tail_rule(value):
                code = _canonical_code(value)
                if not code:
                    return None
                for rule in clear_tail_rules:
                    prefix = str(rule.get('prefix', '') or '').strip()
                    if prefix and code.startswith(prefix):
                        return rule
                return None

            def _is_usage_excluded(parent_code, child_code):
                return any(
                    (parent_key, child_key) in usage_exclusion_pairs
                    for parent_key in _code_keys(parent_code)
                    for child_key in _code_keys(child_code)
                )

            def _month_key(value):
                if isinstance(value, datetime):
                    return value.strftime('%Y-%m')
                if isinstance(value, date):
                    return datetime(value.year, value.month, value.day).strftime('%Y-%m')
                return '未指定月份'

            def _project_display(projects):
                clean_projects = sorted({str(project).strip() for project in (projects or []) if str(project).strip()})
                return '、'.join(clean_projects)

            def _mother_spec_from_bom(mother_pn):
                resolved = _resolve_bom_key(mother_pn)
                if not resolved or resolved not in self.bom_index or not self.bom_index[resolved]:
                    return ''
                row = self.bom_index[resolved][0]
                return str(row[3]).strip() if len(row) > 3 and row[3] else ''

            custom_stock_targets = {}
            for rule in custom_stock_rules:
                target_code = _canonical_code(rule.get('pn', ''))
                if not target_code:
                    continue
                item = custom_stock_targets.setdefault(target_code, {
                    'qty': 0.0,
                    'model': '',
                    'customer': '',
                })
                item['qty'] += float(rule.get('qty', 0) or 0)
                if rule.get('model') and not item['model']:
                    item['model'] = str(rule.get('model')).strip()
                if rule.get('customer') and not item['customer']:
                    item['customer'] = str(rule.get('customer')).strip()

            plan_expansion_rows = []
            if self._po_raw_rows:
                for row in self._po_raw_rows:
                    mother_pn = row.get('pn')
                    try:
                        mother_qty = float(row.get('qty', 0) or 0)
                    except (TypeError, ValueError):
                        mother_qty = 0
                    if mother_pn and mother_qty > 0:
                        mother_spec = row.get('mother_spec', '') or _mother_spec_from_bom(mother_pn)
                        plan_expansion_rows.append((mother_pn, mother_qty, row.get('date'), row.get('project', ''), mother_spec))
            if not plan_expansion_rows:
                plan_expansion_rows = [
                    (mother_pn, mother_qty, None, '', _mother_spec_from_bom(mother_pn))
                    for mother_pn, mother_qty in self._po_data
                ]

            custom_stock_remaining = {
                code: float(item.get('qty', 0) or 0)
                for code, item in custom_stock_targets.items()
            }
            custom_stock_effective_qty = defaultdict(float)
            custom_stock_stats = {
                'rules': len(custom_stock_rules),
                'controlled_mothers': 0,
                'original_qty': 0.0,
                'effective_qty': 0.0,
                'trimmed_qty': 0.0,
            }
            capped_plan_rows = []
            for mother_pn, mother_qty, plan_date, project, mother_spec in plan_expansion_rows:
                mother_code = _canonical_code(mother_pn)
                if mother_code in custom_stock_targets:
                    custom_stock_stats['original_qty'] += float(mother_qty or 0)
                    remain = custom_stock_remaining.get(mother_code, 0.0)
                    if remain <= 0:
                        custom_stock_stats['trimmed_qty'] += float(mother_qty or 0)
                        continue
                    effective_qty = min(float(mother_qty or 0), remain)
                    custom_stock_remaining[mother_code] = remain - effective_qty
                    custom_stock_effective_qty[mother_code] += effective_qty
                    custom_stock_stats['effective_qty'] += effective_qty
                    custom_stock_stats['trimmed_qty'] += max(0.0, float(mother_qty or 0) - effective_qty)
                    target = custom_stock_targets.get(mother_code, {})
                    project = target.get('customer') or project
                    mother_spec = target.get('model') or mother_spec
                    capped_plan_rows.append((mother_pn, effective_qty, plan_date, project, mother_spec))
                else:
                    capped_plan_rows.append((mother_pn, mother_qty, plan_date, project, mother_spec))
            custom_stock_stats['controlled_mothers'] = len([qty for qty in custom_stock_effective_qty.values() if qty > 0])
            plan_expansion_rows = capped_plan_rows

            switch_clear_tail_by_source_child = {}
            planned_source_roots = sorted({
                _canonical_code(row[0])
                for row in plan_expansion_rows
                if _canonical_code(row[0])
            })

            def _get_source_bucket(material_code, source_root_code, source_mode, rule=None):
                source_root_code = str(source_root_code or '').strip()
                if source_mode == 'clear_tail':
                    rule_key = str((rule or {}).get('remark') or (rule or {}).get('prefix') or source_root_code or material_code)
                    bucket_key = f'clear_tail|{source_root_code}|{rule_key}'
                elif source_mode == 'custom':
                    bucket_key = 'custom_stock'
                else:
                    bucket_key = 'normal'
                bucket = source_requirement_buckets[material_code].get(bucket_key)
                if bucket is None:
                    bucket = {
                        'bucket_key': bucket_key,
                        'source_mode': source_mode,
                        'clear_tail_rule': rule if source_mode == 'clear_tail' else None,
                        'source_roots': set(),
                        'name': '',
                        'spec': '',
                        'total_qty': 0.0,
                        'paths': 0,
                        'parents': set(),
                        'parent_specs': set(),
                        'monthly_parent_specs': defaultdict(set),
                        'projects': set(),
                        'monthly_projects': defaultdict(set),
                        'monthly_demand': defaultdict(float),
                    }
                    source_requirement_buckets[material_code][bucket_key] = bucket
                if source_root_code:
                    bucket['source_roots'].add(source_root_code)
                return bucket

            def _source_mode_for_requirement(material_code, source_root_code):
                manual_clear_tail_rule = _match_clear_tail_rule(material_code)
                if manual_clear_tail_rule:
                    return 'clear_tail', manual_clear_tail_rule
                source_clear_tail_rule = switch_clear_tail_by_source_child.get((source_root_code, material_code))
                if source_clear_tail_rule:
                    return 'clear_tail', source_clear_tail_rule
                if source_root_code and float(custom_stock_effective_qty.get(source_root_code, 0) or 0) > 0:
                    return 'custom', None
                return 'normal', None

            def _add_flat_requirement(code, name, spec, qty, parent_label=None, month_key=None, project=None, parent_spec=None, source_root=None):
                material_code = _canonical_code(code)
                if not material_code:
                    return ''
                source_root_code = _canonical_code(source_root or parent_label or '')
                source_mode, source_rule = _source_mode_for_requirement(material_code, source_root_code)
                bucket = _get_source_bucket(material_code, source_root_code, source_mode, source_rule)
                if not bucket['name']:
                    bucket['name'] = str(name or '').strip()
                if not bucket['spec']:
                    bucket['spec'] = str(spec or '').strip()
                bucket['total_qty'] += qty
                bucket['paths'] += 1
                if parent_label:
                    bucket['parents'].add(parent_label)
                if parent_label:
                    parents_map.setdefault(material_code, set()).add(parent_label)
                if parent_spec:
                    parent_spec_text = str(parent_spec).strip()
                    if parent_spec_text:
                        bucket['parent_specs'].add(parent_spec_text)
                        parent_spec_map[material_code].add(parent_spec_text)
                        if month_key:
                            bucket['monthly_parent_specs'][month_key].add(parent_spec_text)
                            monthly_parent_spec_map[material_code][month_key].add(parent_spec_text)
                if month_key:
                    bucket['monthly_demand'][month_key] += qty
                    monthly_demand_map[material_code][month_key] += qty
                    if project:
                        bucket['monthly_projects'][month_key].add(str(project).strip())
                        monthly_project_map[material_code][month_key].add(str(project).strip())
                if project:
                    bucket['projects'].add(str(project).strip())
                    project_map[material_code].add(str(project).strip())
                if material_code in flat:
                    flat[material_code]['total_qty'] += qty
                    flat[material_code]['paths'] += 1
                else:
                    flat[material_code] = {
                        'name': str(name or '').strip(),
                        'spec': str(spec or '').strip(),
                        'total_qty': qty,
                        'paths': 1,
                    }
                return material_code

            production_parent_codes = set()
            for mother_pn, _ in self._po_data:
                production_parent_codes.update(_code_keys(mother_pn))
            abc_exclusion_stats = {
                'material': 0,
                'usage': 0,
                'production_parent': 0,
            }
            replacement_diff_stats = {
                'rules': len(replacement_diff_rules),
                'applied': 0,
                'positive_materials': 0,
                'skipped_nonpositive': 0,
                'missing_base': 0,
                'missing_replacement': 0,
            }
            model_switch_stats = {
                'rules': len(model_switch_rules),
                'applied': 0,
                'clear_tail_materials': 0,
                'inherited_sources': 0,
                'missing_old': 0,
                'missing_new': 0,
            }
            try:
                subplan_dedupe_enabled = bool(getattr(self, '_abc_dedupe_subplan').get())
            except Exception:
                subplan_dedupe_enabled = True
            subplan_dedupe_stats = {
                'enabled': subplan_dedupe_enabled,
                'original_rows': len(plan_expansion_rows),
                'effective_rows': len(plan_expansion_rows),
                'covered_rows': 0,
                'partial_rows': 0,
                'covered_qty': 0.0,
                'affected_codes': set(),
            }

            def _recurse(pn, qty, depth, parent_pn=None, trail: frozenset[str] = frozenset(), month_key=None, project=None, parent_spec=None, source_root=None):
                resolved_pn = _resolve_bom_key(pn)
                if depth > MAX_DEPTH or not resolved_pn or resolved_pn not in self.bom_index:
                    return
                for row in self.bom_index[resolved_pn]:
                    cp = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
                    if not cp:
                        continue
                    if _is_usage_excluded(resolved_pn, cp):
                        abc_exclusion_stats['usage'] += 1
                        continue
                    # BOM 环路检测：cp 如果已经在当前路径上，继续展开会死循环
                    child_code = _canonical_code(cp)
                    resolved_child = _resolve_bom_key(cp)
                    if cp in trail or child_code in trail or resolved_child in trail or cp == resolved_pn:
                        cycle_hits.add((resolved_pn, cp))
                        continue
                    cname = str(row[CHILD_NAME_COL]).strip() if row[CHILD_NAME_COL] else ''
                    cspec = str(row[CHILD_SPEC_COL]).strip() if row[CHILD_SPEC_COL] else ''
                    try:
                        cqty = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else 1.0
                    except (ValueError, TypeError):
                        cqty = 1.0
                    tqty = qty * cqty
                    flat_code = _add_flat_requirement(cp, cname, cspec, tqty, parent_pn, month_key, project, parent_spec, source_root)
                    _recurse(
                        cp,
                        tqty,
                        depth + 1,
                        parent_pn=flat_code or cp,
                        trail=trail | {resolved_pn, child_code},
                        month_key=month_key,
                        project=project,
                        parent_spec=parent_spec,
                        source_root=source_root,
                    )

            def _explode_unit_requirements(root_pn, *, count_usage_exclusions=False):
                """按 1 个母件展开BOM，返回每颗下层物料的单机用量。"""
                totals = defaultdict(float)
                info = {}
                root_key = _resolve_bom_key(root_pn)
                if not root_key:
                    return totals, info

                def _walk(pn, qty, depth, trail: frozenset[str]):
                    resolved_pn = _resolve_bom_key(pn)
                    if depth > MAX_DEPTH or not resolved_pn or resolved_pn not in self.bom_index:
                        return
                    for row in self.bom_index[resolved_pn]:
                        cp = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
                        if not cp:
                            continue
                        if _is_usage_excluded(resolved_pn, cp):
                            if count_usage_exclusions:
                                abc_exclusion_stats['usage'] += 1
                            continue
                        child_code = _canonical_code(cp)
                        resolved_child = _resolve_bom_key(cp)
                        if cp in trail or child_code in trail or resolved_child in trail or cp == resolved_pn:
                            cycle_hits.add((resolved_pn, cp))
                            continue
                        cname = str(row[CHILD_NAME_COL]).strip() if row[CHILD_NAME_COL] else ''
                        cspec = str(row[CHILD_SPEC_COL]).strip() if row[CHILD_SPEC_COL] else ''
                        try:
                            cqty = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else 1.0
                        except (ValueError, TypeError):
                            cqty = 1.0
                        tqty = qty * cqty
                        totals[child_code] += tqty
                        info.setdefault(child_code, {'name': cname, 'spec': cspec})
                        _walk(cp, tqty, depth + 1, trail | {resolved_pn, child_code})

                _walk(root_key, 1.0, 1, frozenset({root_key, _canonical_code(root_key)}))
                return totals, info

            custom_stock_child_cap = defaultdict(float)
            custom_stock_child_roots = defaultdict(set)
            for root_code, effective_qty in custom_stock_effective_qty.items():
                if effective_qty <= 0:
                    continue
                unit_totals, _ = _explode_unit_requirements(root_code, count_usage_exclusions=False)
                for child_code, unit_qty in unit_totals.items():
                    qty = float(unit_qty or 0) * float(effective_qty or 0)
                    if qty <= 0:
                        continue
                    custom_stock_child_cap[child_code] += qty
                    custom_stock_child_roots[child_code].add(root_code)

            auto_clear_tail_by_code = {}
            for rule in model_switch_rules:
                old_pn = rule.get('old_pn', '')
                new_pn = rule.get('new_pn', '')
                old_key = _resolve_bom_key(old_pn)
                new_key = _resolve_bom_key(new_pn)
                if not old_key:
                    model_switch_stats['missing_old'] += 1
                    continue
                if not new_key:
                    model_switch_stats['missing_new'] += 1
                    continue

                old_totals, _ = _explode_unit_requirements(old_pn, count_usage_exclusions=True)
                new_totals, _ = _explode_unit_requirements(new_pn, count_usage_exclusions=True)
                added_count = 0
                old_source_code = _canonical_code(old_pn)
                inherited_source_roots = set()
                for child_code, old_unit_qty in old_totals.items():
                    new_unit_qty = float(new_totals.get(child_code, 0) or 0)
                    diff_unit_qty = float(old_unit_qty or 0) - new_unit_qty
                    if diff_unit_qty <= 1e-9:
                        continue
                    if _match_clear_tail_rule(child_code):
                        continue
                    old_display = fmt_qty(old_unit_qty)
                    new_display = fmt_qty(new_unit_qty)
                    switch_month = rule.get('switch_month', '') or ''
                    remark_parts = [
                        f'型号切换:{old_pn}->{new_pn}',
                        f'原单机用量{old_display}',
                        f'新单机用量{new_display}',
                    ]
                    if rule.get('switch_date'):
                        remark_parts.append(f'切换日期{rule["switch_date"].strftime("%Y-%m-%d")}')
                    if rule.get('remark'):
                        remark_parts.append(str(rule.get('remark')))
                    auto_clear_tail_by_code[child_code] = {
                        'prefix': child_code,
                        'cutoff_month': switch_month,
                        'remark': '；'.join(remark_parts),
                    }
                    switch_clear_tail_by_source_child[(_canonical_code(old_pn), child_code)] = auto_clear_tail_by_code[child_code]
                    added_count += 1
                diff_child_codes = {
                    child_code
                    for child_code, old_unit_qty in old_totals.items()
                    if float(old_unit_qty or 0) - float(new_totals.get(child_code, 0) or 0) > 1e-9
                    and not _match_clear_tail_rule(child_code)
                }
                for source_root_code in planned_source_roots:
                    if not source_root_code or source_root_code == old_source_code:
                        continue
                    if source_root_code not in old_totals:
                        continue
                    if not _resolve_bom_key(source_root_code):
                        continue
                    source_totals, _ = _explode_unit_requirements(source_root_code, count_usage_exclusions=True)
                    inherited_count = 0
                    for child_code in diff_child_codes:
                        rule_info = auto_clear_tail_by_code.get(child_code)
                        if not rule_info or child_code not in source_totals:
                            continue
                        switch_clear_tail_by_source_child[(source_root_code, child_code)] = rule_info
                        inherited_count += 1
                    if inherited_count:
                        inherited_source_roots.add(source_root_code)
                model_switch_stats['applied'] += 1
                model_switch_stats['clear_tail_materials'] += added_count
                model_switch_stats['inherited_sources'] += len(inherited_source_roots)

            effective_clear_tail_rule_count = len(clear_tail_rules) + len(auto_clear_tail_by_code)

            def _dedupe_subplanned_rows(rows):
                if not subplan_dedupe_enabled or not rows:
                    return rows
                unit_cache = {}

                def _unit_totals(root_code):
                    root_code = _canonical_code(root_code)
                    if root_code not in unit_cache:
                        unit_cache[root_code] = _explode_unit_requirements(root_code, count_usage_exclusions=False)[0]
                    return unit_cache[root_code]

                grouped = defaultdict(list)
                for idx, row in enumerate(rows):
                    mother_pn, mother_qty, plan_date, project, mother_spec = row
                    grouped[_month_key(plan_date)].append({
                        'idx': idx,
                        'row': row,
                        'code': _canonical_code(mother_pn),
                        'qty': float(mother_qty or 0),
                    })

                effective_rows = []
                for _month, items in grouped.items():
                    planned_codes = {item['code'] for item in items if item['code']}
                    ancestor_count = defaultdict(int)
                    for item in items:
                        code = item['code']
                        if not code:
                            continue
                        for child_code in _unit_totals(code).keys():
                            if child_code in planned_codes and child_code != code:
                                ancestor_count[child_code] += 1
                    coverage_available = defaultdict(float)
                    ordered_items = sorted(items, key=lambda item: (ancestor_count.get(item['code'], 0), item['idx']))
                    for item in ordered_items:
                        code = item['code']
                        qty = float(item['qty'] or 0)
                        covered_qty = min(qty, coverage_available.get(code, 0.0))
                        if covered_qty > 0:
                            coverage_available[code] -= covered_qty
                            subplan_dedupe_stats['covered_qty'] += covered_qty
                            subplan_dedupe_stats['affected_codes'].add(code)
                        effective_qty = max(0.0, qty - covered_qty)
                        if effective_qty <= 1e-9:
                            subplan_dedupe_stats['covered_rows'] += 1
                            continue
                        if covered_qty > 0:
                            subplan_dedupe_stats['partial_rows'] += 1
                        mother_pn, _mother_qty, plan_date, project, mother_spec = item['row']
                        effective_rows.append((mother_pn, effective_qty, plan_date, project, mother_spec))
                        for child_code, unit_qty in _unit_totals(code).items():
                            if child_code in planned_codes and child_code != code:
                                coverage_available[child_code] += effective_qty * float(unit_qty or 0)

                subplan_dedupe_stats['effective_rows'] = len(effective_rows)
                return effective_rows

            plan_expansion_rows = _dedupe_subplanned_rows(plan_expansion_rows)

            for mother_pn, mother_qty, plan_date, project, mother_spec in plan_expansion_rows:
                month_key = _month_key(plan_date)
                source_root = _canonical_code(mother_pn)
                base_pn = _lookup_by_code_keys(replacement_diff_rules, mother_pn)
                if not base_pn:
                    _recurse(mother_pn, mother_qty, 1, parent_pn=mother_pn, trail=frozenset({mother_pn, _canonical_code(mother_pn)}), month_key=month_key, project=project, parent_spec=mother_spec, source_root=source_root)
                    continue

                if not _resolve_bom_key(mother_pn):
                    replacement_diff_stats['missing_replacement'] += 1
                    _recurse(mother_pn, mother_qty, 1, parent_pn=mother_pn, trail=frozenset({mother_pn, _canonical_code(mother_pn)}), month_key=month_key, project=project, parent_spec=mother_spec, source_root=source_root)
                    continue
                if not _resolve_bom_key(base_pn):
                    replacement_diff_stats['missing_base'] += 1
                    _recurse(mother_pn, mother_qty, 1, parent_pn=mother_pn, trail=frozenset({mother_pn, _canonical_code(mother_pn)}), month_key=month_key, project=project, parent_spec=mother_spec, source_root=source_root)
                    continue

                replacement_totals, replacement_info = _explode_unit_requirements(mother_pn, count_usage_exclusions=True)
                base_totals, _ = _explode_unit_requirements(base_pn, count_usage_exclusions=True)
                positive_count = 0
                parent_label = f'{mother_pn}(替换差异:{base_pn})'
                for child_code, replacement_unit_qty in replacement_totals.items():
                    diff_unit_qty = replacement_unit_qty - float(base_totals.get(child_code, 0) or 0)
                    if diff_unit_qty <= 1e-9:
                        replacement_diff_stats['skipped_nonpositive'] += 1
                        continue
                    child_info = replacement_info.get(child_code, {})
                    _add_flat_requirement(
                        child_code,
                        child_info.get('name', ''),
                        child_info.get('spec', ''),
                        diff_unit_qty * mother_qty,
                        parent_label,
                        month_key,
                        project,
                        mother_spec,
                        source_root,
                    )
                    positive_count += 1
                replacement_diff_stats['applied'] += 1
                replacement_diff_stats['positive_materials'] += positive_count

            for code in list(flat.keys()):
                code_keys = _code_keys(code)
                if production_parent_codes and any(key in production_parent_codes for key in code_keys):
                    flat.pop(code, None)
                    source_requirement_buckets.pop(code, None)
                    parents_map.pop(code, None)
                    parent_spec_map.pop(code, None)
                    monthly_demand_map.pop(code, None)
                    project_map.pop(code, None)
                    monthly_project_map.pop(code, None)
                    monthly_parent_spec_map.pop(code, None)
                    abc_exclusion_stats['production_parent'] += 1
                    continue
                if excluded_codes and any(key in excluded_codes for key in code_keys):
                    flat.pop(code, None)
                    source_requirement_buckets.pop(code, None)
                    parents_map.pop(code, None)
                    parent_spec_map.pop(code, None)
                    monthly_demand_map.pop(code, None)
                    project_map.pop(code, None)
                    monthly_project_map.pop(code, None)
                    monthly_parent_spec_map.pop(code, None)
                    abc_exclusion_stats['material'] += 1

            if cycle_hits:
                sample = '、'.join(f'{a}→{b}' for a, b in list(cycle_hits)[:3])
                try:
                    self._log_to_buffer('external', f'BOM 存在环路引用（跳过展开）：{sample}')
                except Exception:
                    pass

            # ── 第二步：从排产计划计算实际排产天数和日均总产量 ─────
            all_dates = []
            total_plan_qty = 0.0
            for row in self._po_raw_rows:
                d = row.get('date')
                q = row.get('qty', 0)
                if isinstance(d, datetime):
                    all_dates.append(d)
                if isinstance(q, (int, float)) and q > 0:
                    total_plan_qty += q

            manual_val = self._plan_days_manual.get().strip()
            if manual_val:
                try:
                    plan_days = max(1, int(manual_val))
                except ValueError:
                    plan_days = 30
            elif len(all_dates) >= 2:
                plan_days = (max(all_dates) - min(all_dates)).days + 1
            else:
                d = all_dates[0] if all_dates else None
                if d and d.day == 1:
                    plan_days = calendar.monthrange(d.year, d.month)[1]
                else:
                    plan_days = 30
            avg_daily_all = total_plan_qty / float(plan_days)

            # ── 第三步：建立完整物料清单 ──────────────────────────
            abc_rows = []
            lead_a_thresh = self._lead_a.get()
            lead_b_thresh = self._lead_b.get()
            daily_factor = self._daily_usage_factor.get()
            inventory_map, po_map, pr_map = {}, {}, {}
            po_arrival_map = defaultdict(list)
            material_category_rules = []
            if self._abc_mrp_path:
                inventory_map, po_map, pr_map = self._load_abc_mrp_supply_maps(self._abc_mrp_path)
                po_arrival_map = self._load_abc_mrp_po_arrival_map(self._abc_mrp_path)
                material_category_rules = self._load_abc_material_category_rules(self._abc_mrp_path)

            def _build_custom_diff_rows():
                custom_roots = {
                    _canonical_code(root_code): float(effective_qty or 0)
                    for root_code, effective_qty in custom_stock_effective_qty.items()
                    if _canonical_code(root_code) and float(effective_qty or 0) > 0
                }
                if not custom_roots:
                    return []

                root_meta = {}
                for mother_pn, mother_qty, _plan_date, project, mother_spec in plan_expansion_rows:
                    root_code = _canonical_code(mother_pn)
                    if not root_code:
                        continue
                    meta = root_meta.setdefault(root_code, {
                        'qty': 0.0,
                        'projects': set(),
                        'specs': set(),
                    })
                    try:
                        meta['qty'] += float(mother_qty or 0)
                    except (TypeError, ValueError):
                        pass
                    if project:
                        meta['projects'].add(str(project).strip())
                    if mother_spec:
                        meta['specs'].add(str(mother_spec).strip())

                unit_cache = {}

                def _unit_totals_with_info(root_code):
                    root_code = _canonical_code(root_code)
                    if root_code not in unit_cache:
                        totals, info = _explode_unit_requirements(root_code, count_usage_exclusions=False)
                        unit_cache[root_code] = (dict(totals), dict(info))
                    return unit_cache[root_code]

                all_root_codes = sorted(root_meta.keys())
                rows = []
                for custom_root, effective_qty in sorted(custom_roots.items()):
                    custom_totals, custom_info = _unit_totals_with_info(custom_root)
                    if not custom_totals:
                        continue

                    other_max_unit = defaultdict(float)
                    other_sources = defaultdict(list)
                    for other_root in all_root_codes:
                        if other_root == custom_root:
                            continue
                        other_totals, _other_info = _unit_totals_with_info(other_root)
                        other_meta = root_meta.get(other_root, {})
                        other_spec = _project_display(other_meta.get('specs', set()))
                        other_project = _project_display(other_meta.get('projects', set()))
                        for child_code, unit_qty in other_totals.items():
                            unit_qty = float(unit_qty or 0)
                            if unit_qty <= 0:
                                continue
                            if unit_qty > other_max_unit[child_code] + 1e-9:
                                other_max_unit[child_code] = unit_qty
                                other_sources[child_code] = [(other_root, other_spec, other_project, unit_qty)]
                            elif abs(unit_qty - other_max_unit[child_code]) <= 1e-9:
                                other_sources[child_code].append((other_root, other_spec, other_project, unit_qty))

                    custom_target = custom_stock_targets.get(custom_root, {})
                    custom_model = (
                        str(custom_target.get('model', '') or '').strip()
                        or _project_display(root_meta.get(custom_root, {}).get('specs', set()))
                    )
                    custom_project = (
                        str(custom_target.get('customer', '') or '').strip()
                        or _project_display(root_meta.get(custom_root, {}).get('projects', set()))
                    )
                    all_children = set(custom_totals.keys()) | set(other_max_unit.keys())
                    for child_code in sorted(all_children):
                        custom_unit = float(custom_totals.get(child_code, 0) or 0)
                        other_unit = float(other_max_unit.get(child_code, 0) or 0)
                        diff_unit = custom_unit - other_unit
                        if abs(diff_unit) <= 1e-9:
                            continue
                        if custom_unit > 0 and other_unit <= 0:
                            diff_type = '定制专用'
                        elif custom_unit > other_unit:
                            diff_type = '定制多用'
                        elif custom_unit <= 0 and other_unit > 0:
                            diff_type = '其他项目有/定制不用'
                        else:
                            diff_type = '定制少用'

                        ext_info = _lookup_by_code_keys(self._ext_data, child_code, {}) or {}
                        bom_info = custom_info.get(child_code, {})
                        if not bom_info:
                            for other_root, _other_spec, _other_project, _qty in other_sources.get(child_code, []):
                                other_info = _unit_totals_with_info(other_root)[1]
                                bom_info = other_info.get(child_code, {})
                                if bom_info:
                                    break
                        source_texts = []
                        for other_root, other_spec, other_project, unit_qty in other_sources.get(child_code, [])[:8]:
                            label_parts = [other_root]
                            if other_spec:
                                label_parts.append(other_spec)
                            if other_project:
                                label_parts.append(other_project)
                            source_texts.append(f'{" / ".join(label_parts)}:{fmt_qty(unit_qty)}')

                        rows.append({
                            'custom_root': custom_root,
                            'custom_model': custom_model,
                            'custom_project': custom_project,
                            'custom_effective_qty': effective_qty,
                            'code': child_code,
                            'name': ext_info.get('name') or bom_info.get('name', ''),
                            'spec': ext_info.get('spec') or bom_info.get('spec', ''),
                            'material_category': self._match_abc_material_category(child_code, material_category_rules),
                            'supplier': ext_info.get('supplier', ''),
                            'buyer': ext_info.get('buyer', ''),
                            'diff_type': diff_type,
                            'custom_unit_qty': custom_unit,
                            'other_max_unit_qty': other_unit,
                            'diff_unit_qty': diff_unit,
                            'custom_diff_qty': diff_unit * effective_qty,
                            'compare_sources': '\n'.join(source_texts),
                            'remark': '按单台用量对比；差异数量=单台差异×定制有效数量',
                        })

                priority = {'定制专用': 0, '定制多用': 1, '定制少用': 2, '其他项目有/定制不用': 3}
                rows.sort(key=lambda item: (
                    str(item.get('custom_root', '')),
                    priority.get(item.get('diff_type'), 9),
                    str(item.get('material_category', '')),
                    str(item.get('code', '')),
                ))
                return rows

            self._abc_custom_diff_rows = _build_custom_diff_rows()

            def _bucket_priority(bucket):
                mode = bucket.get('source_mode')
                if mode == 'clear_tail':
                    return 0
                if mode == 'custom':
                    return 1
                return 2

            def _allocate_supply_to_buckets(code, bucket_items):
                remaining = {
                    'current_inventory': float(_lookup_by_code_keys(inventory_map, code, 0) or 0),
                    'open_po': float(_lookup_by_code_keys(po_map, code, 0) or 0),
                    'untransfer_pr': float(_lookup_by_code_keys(pr_map, code, 0) or 0),
                }
                allocations = {
                    bucket['bucket_key']: {'current_inventory': 0.0, 'open_po': 0.0, 'untransfer_pr': 0.0}
                    for bucket in bucket_items
                }
                normal_buckets = [bucket for bucket in bucket_items if bucket.get('source_mode') == 'normal']
                special_buckets = [bucket for bucket in bucket_items if bucket.get('source_mode') != 'normal']
                for bucket in sorted(special_buckets, key=_bucket_priority):
                    need = max(0.0, float(bucket.get('total_qty', 0) or 0))
                    for component in ('current_inventory', 'open_po', 'untransfer_pr'):
                        if need <= 0:
                            break
                        take = min(remaining[component], need)
                        allocations[bucket['bucket_key']][component] += take
                        remaining[component] -= take
                        need -= take
                if normal_buckets:
                    normal_bucket = normal_buckets[0]
                    for component in ('current_inventory', 'open_po', 'untransfer_pr'):
                        allocations[normal_bucket['bucket_key']][component] += remaining[component]
                elif special_buckets and sum(remaining.values()) > 0:
                    # 没有普通来源时，剩余供应不再分摊给清尾/定制行，避免超出控制数量后继续影响采购判断。
                    pass
                return allocations

            def _bucket_monthly_texts(bucket, key):
                return {
                    month_key: _project_display(values)
                    for month_key, values in bucket.get(key, {}).items()
                }

            for code, bucket_map in source_requirement_buckets.items():
                if code not in flat:
                    continue
                material_category = self._match_abc_material_category(code, material_category_rules)
                ext_info = _lookup_by_code_keys(self._ext_data, code, {}) or {}
                aggregate_info = flat.get(code, {})
                bucket_items = sorted(bucket_map.values(), key=_bucket_priority)
                supply_allocations = _allocate_supply_to_buckets(code, bucket_items)

                for bucket in bucket_items:
                    total_qty = float(bucket.get('total_qty', 0) or 0)
                    if total_qty <= 0:
                        continue
                    current_inventory = supply_allocations.get(bucket['bucket_key'], {}).get('current_inventory', 0.0)
                    inbound_po = supply_allocations.get(bucket['bucket_key'], {}).get('open_po', 0.0)
                    untransfer_pr = supply_allocations.get(bucket['bucket_key'], {}).get('untransfer_pr', 0.0)
                    inventory_position = current_inventory + inbound_po + untransfer_pr
                    avg_daily = (total_qty / float(plan_days)) * daily_factor
                    monthly_avg = avg_daily * 30.0
                    yearly_usage = monthly_avg * 12.0
                    clear_tail_rule = bucket.get('clear_tail_rule') if bucket.get('source_mode') == 'clear_tail' else None
                    custom_stock_control = bucket.get('source_mode') == 'custom'
                    custom_stock_cap_qty = total_qty if custom_stock_control else 0.0
                    row_key = f'{code}|{bucket.get("bucket_key", "normal")}'
                    common_row = {
                        'row_key': row_key,
                        'source_mode': bucket.get('source_mode', 'normal'),
                        'code': code,
                        'material_category': material_category,
                        'total_qty': total_qty,
                        'avg_daily': avg_daily,
                        'monthly_avg': monthly_avg,
                        'yearly_usage': yearly_usage,
                        'parents': sorted(bucket.get('parents', set())),
                        'parent_specs': _project_display(bucket.get('parent_specs', set())),
                        'current_inventory': current_inventory,
                        'open_po': inbound_po,
                        'open_po_arrivals': list(_lookup_by_code_keys(po_arrival_map, code, []) or []),
                        'untransfer_pr': untransfer_pr,
                        'inventory_position': inventory_position,
                        'monthly_demand': dict(bucket.get('monthly_demand', {})),
                        'monthly_projects': _bucket_monthly_texts(bucket, 'monthly_projects'),
                        'monthly_parent_specs': _bucket_monthly_texts(bucket, 'monthly_parent_specs'),
                        'projects': _project_display(bucket.get('projects', set())),
                        'clear_tail': bool(clear_tail_rule),
                        'clear_tail_cutoff_month': clear_tail_rule.get('cutoff_month', '') if clear_tail_rule else '',
                        'clear_tail_remark': clear_tail_rule.get('remark', '') if clear_tail_rule else '',
                        'custom_stock_control': bool(custom_stock_control),
                        'custom_stock_cap_qty': custom_stock_cap_qty,
                        'custom_stock_roots': sorted(bucket.get('source_roots', set())) if custom_stock_control else [],
                    }
                    if ext_info:
                        lead_days = ext_info.get('lead_days', 0)
                        if lead_days <= 0:
                            lead_days = 30
                        try:
                            spq_val = float(str(ext_info.get('spq', '')).strip()) if ext_info.get('spq', '') not in ('', None) else 0.0
                        except (ValueError, TypeError):
                            spq_val = 0.0
                        try:
                            moq_val = float(str(ext_info.get('moq', '')).strip()) if ext_info.get('moq', '') not in ('', None) else 0.0
                        except (ValueError, TypeError):
                            moq_val = 0.0
                        common_row.update({
                            'name': ext_info.get('name', bucket.get('name') or aggregate_info.get('name', '')) or bucket.get('name') or aggregate_info.get('name', ''),
                            'spec': ext_info.get('spec', bucket.get('spec') or aggregate_info.get('spec', '')) or bucket.get('spec') or aggregate_info.get('spec', ''),
                            'supplier': ext_info.get('supplier', ''),
                            'buyer': ext_info.get('buyer', ''),
                            'lead_days': lead_days,
                            'spq': ext_info.get('spq', ''),
                            'moq': ext_info.get('moq', ''),
                            'spq_val': spq_val,
                            'moq_val': moq_val,
                            'has_ext': True,
                        })
                    else:
                        common_row.update({
                            'name': bucket.get('name') or aggregate_info.get('name', ''),
                            'spec': bucket.get('spec') or aggregate_info.get('spec', ''),
                            'supplier': '\u9700\u91c7\u8d2d\u63d0\u4f9b',
                            'buyer': '',
                            'lead_days': 0,
                            'spq': '',
                            'moq': '',
                            'spq_val': 0.0,
                            'moq_val': 0.0,
                            'has_ext': False,
                        })
                    abc_rows.append(common_row)

            # ── 第四步：ABC分类（按交期） ────────────────────────
            a_items, b_items, c_items, no_data_items = [], [], [], []
            for r in abc_rows:
                if not r['has_ext']:
                    r['abc'] = '-'
                    r['safety_stock'] = 0
                    r['raw_safety_stock'] = 0
                    r['rop'] = 0
                    r['purchase_trigger'] = ''
                    r['theoretical_purchase_qty'] = ''
                    r['suggested_purchase_qty'] = ''
                    r['purchase_advice'] = '缺采购主数据，无法计算采购建议'
                    r['ss_coverage_days'] = ''
                    r['ss_judge'] = '\u7f3a\u91c7\u8d2d\u6570\u636e'
                    no_data_items.append(r)
                    continue

                ld = r['lead_days']
                if ld >= lead_a_thresh:
                    r['abc'] = 'A'
                    a_items.append(r)
                elif ld >= lead_b_thresh:
                    r['abc'] = 'B'
                    b_items.append(r)
                else:
                    r['abc'] = 'C'
                    c_items.append(r)

            # ── 第五步：标准安全库存 + ROP计算 ───────────────────
            # 安全库存 = k × 平均日均用量 × 交期
            # k值：A类=0.65(65%覆盖) B类=0.28(28%) C类=0.10(10%)
            # 再订货点 ROP = 平均日均用量 × 交期 + 原始安全库存（不含MOQ/SPQ）
            k_map = {'A': k_a, 'B': k_b, 'C': k_c}
            for r in abc_rows:
                if not r['has_ext']:
                    continue
                abc = r['abc']
                ld = r['lead_days']
                ad = r['avg_daily']
                k = k_map.get(abc, 0.28)

                raw_ss = k * ad * ld
                r['raw_safety_stock'] = raw_ss
                theoretical_rop = ad * ld + raw_ss

                # MOQ：不低于最小起订量
                moq_val = r['moq_val']
                if moq_val > 0:
                    raw_ss = max(raw_ss, moq_val)

                # SPQ：向上取整到包装倍数（整数）
                spq_val = r['spq_val']
                if spq_val > 0:
                    raw_ss = math.ceil(raw_ss / spq_val) * spq_val

                r['safety_stock'] = max(1, int(raw_ss))

                # ROP 是补货触发线，不是建议下单量；不受MOQ/SPQ起订包装约束影响。
                r['rop'] = max(1, int(math.ceil(theoretical_rop)))

                if not self._abc_mrp_path:
                    r['purchase_trigger'] = '缺MRP'
                    r['theoretical_purchase_qty'] = ''
                    r['suggested_purchase_qty'] = ''
                    r['purchase_advice'] = '未接入MRP供给数据，无法判断本次采购量'
                elif r.get('clear_tail'):
                    theoretical_purchase_qty = max(0.0, float(r.get('total_qty', 0) or 0) - r['inventory_position'])
                    theoretical_purchase_qty_int = int(math.ceil(theoretical_purchase_qty))
                    if theoretical_purchase_qty_int <= 0:
                        r['purchase_trigger'] = '否'
                        r['theoretical_purchase_qty'] = 0
                        r['suggested_purchase_qty'] = 0
                        r['purchase_advice'] = (
                            f'切换清尾：排产总需求{fmt_qty(r.get("total_qty", 0))} <= 库存位置{fmt_qty(r["inventory_position"])}，不采购'
                        )
                    else:
                        suggested_purchase_qty = theoretical_purchase_qty
                        if moq_val > 0:
                            suggested_purchase_qty = max(suggested_purchase_qty, moq_val)
                        if spq_val > 0:
                            suggested_purchase_qty = math.ceil(suggested_purchase_qty / spq_val) * spq_val
                        else:
                            suggested_purchase_qty = math.ceil(suggested_purchase_qty)
                        r['purchase_trigger'] = '是'
                        r['theoretical_purchase_qty'] = theoretical_purchase_qty_int
                        r['suggested_purchase_qty'] = int(suggested_purchase_qty)
                        r['purchase_advice'] = (
                            f'切换清尾：只补排产净缺口，不补ROP；'
                            f'排产总需求{fmt_qty(r.get("total_qty", 0))} - 库存位置{fmt_qty(r["inventory_position"])} = {fmt_qty(theoretical_purchase_qty_int)}；'
                            f'按MOQ/SPQ建议采购{fmt_qty(suggested_purchase_qty)}'
                        )
                elif r.get('custom_stock_control'):
                    cap_qty = float(r.get('custom_stock_cap_qty', 0) or 0)
                    theoretical_purchase_qty = max(0.0, cap_qty - r['inventory_position'])
                    theoretical_purchase_qty_int = int(math.ceil(theoretical_purchase_qty))
                    if theoretical_purchase_qty_int <= 0:
                        r['purchase_trigger'] = '否'
                        r['theoretical_purchase_qty'] = 0
                        r['suggested_purchase_qty'] = 0
                        r['purchase_advice'] = (
                            f'定制备货控制：控制数量{fmt_qty(cap_qty)} <= 库存位置{fmt_qty(r["inventory_position"])}，不采购'
                        )
                    else:
                        suggested_purchase_qty = theoretical_purchase_qty
                        if moq_val > 0:
                            suggested_purchase_qty = max(suggested_purchase_qty, moq_val)
                        if spq_val > 0:
                            suggested_purchase_qty = math.ceil(suggested_purchase_qty / spq_val) * spq_val
                        else:
                            suggested_purchase_qty = math.ceil(suggested_purchase_qty)
                        r['purchase_trigger'] = '是'
                        r['theoretical_purchase_qty'] = theoretical_purchase_qty_int
                        r['suggested_purchase_qty'] = int(suggested_purchase_qty)
                        r['purchase_advice'] = (
                            f'定制备货控制：只补控制数量净缺口，不补ROP；'
                            f'控制数量{fmt_qty(cap_qty)} - 库存位置{fmt_qty(r["inventory_position"])} = {fmt_qty(theoretical_purchase_qty_int)}；'
                            f'按MOQ/SPQ建议采购{fmt_qty(suggested_purchase_qty)}'
                        )
                elif r['inventory_position'] > r['rop']:
                    r['purchase_trigger'] = '否'
                    r['theoretical_purchase_qty'] = 0
                    r['suggested_purchase_qty'] = 0
                    r['purchase_advice'] = (
                        f'库存位置{fmt_qty(r["inventory_position"])} > ROP{fmt_qty(r["rop"])}，暂不触发采购'
                    )
                else:
                    theoretical_purchase_qty = max(0.0, r['rop'] - r['inventory_position'])
                    theoretical_purchase_qty_int = int(math.ceil(theoretical_purchase_qty))
                    suggested_purchase_qty = theoretical_purchase_qty
                    if moq_val > 0:
                        suggested_purchase_qty = max(suggested_purchase_qty, moq_val)
                    if spq_val > 0:
                        suggested_purchase_qty = math.ceil(suggested_purchase_qty / spq_val) * spq_val
                    else:
                        suggested_purchase_qty = math.ceil(suggested_purchase_qty)
                    r['purchase_trigger'] = '是'
                    r['theoretical_purchase_qty'] = theoretical_purchase_qty_int
                    r['suggested_purchase_qty'] = int(suggested_purchase_qty)
                    r['purchase_advice'] = (
                        f'库存位置{fmt_qty(r["inventory_position"])} <= ROP{fmt_qty(r["rop"])}；'
                        f'补到ROP，理论采购{fmt_qty(theoretical_purchase_qty_int)}；'
                        f'按MOQ/SPQ建议采购{fmt_qty(suggested_purchase_qty)}'
                    )

                # 物料类别
                # r['category'] = self._classify_material(r['code'], r['supplier'], r['name'])

                # 安全库存覆盖天数
                if ad > 0 and ld > 0:
                    r['ss_coverage_days'] = round(r['safety_stock'] / ad, 1)
                else:
                    r['ss_coverage_days'] = ''

                # 安全库存合理性判断
                ss = r['safety_stock']
                if ss == 0:
                    r['ss_judge'] = ''
                elif r['rop'] < ld * ad:
                    r['ss_judge'] = '\u26a0 ROP\u504f\u4f4e'
                elif ss < ad * ld * 0.1:
                    r['ss_judge'] = '\u26a0 SS\u504f\u4f4e'
                elif ss > ad * ld * 1.0:
                    r['ss_judge'] = '\u26a0 SS\u504f\u9ad8'
                else:
                    r['ss_judge'] = '\u2713 \u5408\u7406'

            abc_rows.sort(key=lambda x: (x['abc'], x['code']))
            self._abc_result = abc_rows
            self._abc_plan_expansion_rows = list(plan_expansion_rows)
            self._abc_calc_meta = {
                'plan_days': plan_days,
                'daily_factor': daily_factor,
                'lead_a': lead_a_thresh,
                'lead_b': lead_b_thresh,
                'k_a': k_a,
                'k_b': k_b,
                'k_c': k_c,
                'mrp_path': self._abc_mrp_path,
                'excluded_codes_count': len(excluded_codes),
                'usage_exclusion_pairs_count': len(usage_exclusion_pairs),
                'production_parent_count': len(production_parent_codes),
                'material_category_rule_count': len(material_category_rules),
                'clear_tail_rule_count': effective_clear_tail_rule_count,
                'manual_clear_tail_rule_count': manual_clear_tail_rule_count,
                'auto_clear_tail_rule_count': len(auto_clear_tail_by_code),
                'model_switch_rule_count': model_switch_stats['rules'],
                'model_switch_applied_count': model_switch_stats['applied'],
                'model_switch_clear_tail_material_count': model_switch_stats['clear_tail_materials'],
                'model_switch_missing_old_count': model_switch_stats['missing_old'],
                'model_switch_missing_new_count': model_switch_stats['missing_new'],
                'custom_stock_rule_count': custom_stock_stats['rules'],
                'custom_stock_controlled_mother_count': custom_stock_stats['controlled_mothers'],
                'custom_stock_original_qty': custom_stock_stats['original_qty'],
                'custom_stock_effective_qty': custom_stock_stats['effective_qty'],
                'custom_stock_trimmed_qty': custom_stock_stats['trimmed_qty'],
                'custom_diff_row_count': len(self._abc_custom_diff_rows),
                'custom_diff_material_count': len({row.get('code') for row in self._abc_custom_diff_rows if row.get('code')}),
                'custom_diff_root_count': len({row.get('custom_root') for row in self._abc_custom_diff_rows if row.get('custom_root')}),
                'subplan_dedupe_enabled': subplan_dedupe_stats['enabled'],
                'subplan_dedupe_original_rows': subplan_dedupe_stats['original_rows'],
                'subplan_dedupe_effective_rows': subplan_dedupe_stats['effective_rows'],
                'subplan_dedupe_covered_rows': subplan_dedupe_stats['covered_rows'],
                'subplan_dedupe_partial_rows': subplan_dedupe_stats['partial_rows'],
                'subplan_dedupe_covered_qty': subplan_dedupe_stats['covered_qty'],
                'subplan_dedupe_affected_code_count': len(subplan_dedupe_stats['affected_codes']),
                'abc_exclusion_stats': abc_exclusion_stats,
                'replacement_diff_rule_count': replacement_diff_stats['rules'],
                'replacement_diff_applied_count': replacement_diff_stats['applied'],
                'replacement_diff_positive_material_count': replacement_diff_stats['positive_materials'],
                'replacement_diff_missing_base_count': replacement_diff_stats['missing_base'],
                'replacement_diff_missing_replacement_count': replacement_diff_stats['missing_replacement'],
            }
            self._refresh_abc_tree()

            total_yearly = sum(r['yearly_usage'] for r in abc_rows if r['has_ext'])
            exclusion_summary = (
                f' | 排除规则:物料{len(excluded_codes)}项/母料用量{len(usage_exclusion_pairs)}条/排产母件{len(production_parent_codes)}项'
                f' | 物料分类规则:{len(material_category_rules)}条'
                f' | 切换清尾:{effective_clear_tail_rule_count}条(手工{manual_clear_tail_rule_count}/切换计划{len(auto_clear_tail_by_code)})'
                f' | 型号切换:{model_switch_stats["rules"]}条/应用{model_switch_stats["applied"]}次/继承来源{model_switch_stats["inherited_sources"]}个'
                f' | 定制备货控制:{custom_stock_stats["rules"]}条/有效母件{custom_stock_stats["controlled_mothers"]}项'
                f' | 定制差异:{len(self._abc_custom_diff_rows)}行'
                f' | 排产去重:{"开" if subplan_dedupe_stats["enabled"] else "关"}/影响{len(subplan_dedupe_stats["affected_codes"])}项'
                f' | 替换差异:{replacement_diff_stats["rules"]}条/应用{replacement_diff_stats["applied"]}次/差异料{replacement_diff_stats["positive_materials"]}项'
            )
            self._abc_summary_var.set(
                f'A\u7c7b:{len(a_items)}\u79cd | B\u7c7b:{len(b_items)}\u79cd | C\u7c7b:{len(c_items)}\u79cd | '
                f'K\u503c(A/B/C):{k_a:g}/{k_b:g}/{k_c:g} | '
                f'\u9700\u91c7\u8d2d\u63d0\u4f9b:{len(no_data_items)}\u79cd | \u5e74\u7528\u91cf\u5408\u8ba1:{total_yearly:,.0f} | '
                f'\u5e93\u5b58\u4f4d\u7f6e:{"\u5df2\u63a5\u5165MRP" if self._abc_mrp_path else "\u672a\u63a5\u5165"}'
                f'{exclusion_summary}'
            )
            self.status_var.set(
                f'\u5b8c\u6210  | A\u7c7b{len(a_items)}\u79cd B\u7c7b{len(b_items)}\u79cd C\u7c7b{len(c_items)}\u79cd '
                f'\u9700\u91c7\u8d2d\u63d0\u4f9b:{len(no_data_items)}\u79cd | \u6392\u4ea7\u5468\u671f:{plan_days}\u5929 | '
                f'已剔除:物料{abc_exclusion_stats["material"]}次/母料用量{abc_exclusion_stats["usage"]}次/母件{abc_exclusion_stats["production_parent"]}项 | '
                f'替换差异应用:{replacement_diff_stats["applied"]}次 | 型号切换应用:{model_switch_stats["applied"]}次/继承来源{model_switch_stats["inherited_sources"]}个 | '
                f'定制备货控制:{custom_stock_stats["controlled_mothers"]}项 | '
                f'定制差异:{len(self._abc_custom_diff_rows)}行 | '
                f'排产去重影响:{len(subplan_dedupe_stats["affected_codes"])}项/覆盖{fmt_qty(subplan_dedupe_stats["covered_qty"])}'
            )
    def _classify_material(self, code, supplier, name):
        """根据料号/名称/供应商前缀判断物料类别"""
        name_str = (name or '').lower()
        code_str = (code or '').lower()
        supplier_str = (supplier or '').lower()
        # 常见类别关键词
        if any(k in name_str for k in ['螺丝', '螺钉', '螺栓', '螺母', '垫圈', '弹垫']):
            return '紧固件'
        elif any(k in name_str for k in ['轴承', '导轨', '丝杠', '直线', '同步带', '皮带']):
            return '传动件'
        elif any(k in name_str for k in ['电机', '马达', '伺服']):
            return '电机类'
        elif any(k in name_str for k in ['法兰', '轴', '键', '联轴']):
            return '机械件'
        elif any(k in name_str for k in ['线束', '连接器', '端子', '接插件', '排线']):
            return '电气件'
        elif any(k in name_str for k in ['壳体', '外壳', '底座', '支架', '机柜']):
            return '结构件'
        elif any(k in name_str for k in ['气缸', '电磁阀', '阀', '油缸', '液压']):
            return '气液件'
        elif any(k in name_str for k in ['传感器', '编码器', '接近', '光电']):
            return '传感器'
        elif any(k in name_str for k in ['控制板', 'PLC', '模块', 'CPU', '电路']):
            return '控制件'
        elif any(k in name_str for k in ['电缆', '电线', '电源', '开关', '断路']):
            return '电气件'
        else:
            return '标准件'

    def _refresh_abc_tree(self):
        for item in self.abc_tree.get_children(''):
            self.abc_tree.delete(item)

        colors = {'A': '#FFD700', 'B': '#87CEEB', 'C': '#90EE90', '-': '#F5C5C5'}

        for r in self._abc_result:
            abc = r['abc']
            tag = abc
            # 安全库存/ROP整数显示
            def int_disp(v):
                if v is None: return ''
                try:
                    f = float(v)
                    return str(int(f)) if f == int(f) else str(round(f, 2))
                except: return str(v)
            # 使用项目：超过5个则折叠显示前5个+省略号
            parents = r.get('parents', [])
            if len(parents) > 5:
                parents_disp = ','.join(parents[:5]) + f'... (+{len(parents)-5})'
            elif len(parents) > 0:
                parents_disp = ','.join(parents)
            else:
                parents_disp = ''
            values = (
                abc,
                r['code'],
                r.get('material_category', ''),
                r.get('projects', ''),
                r['name'],
                r['spec'],
                r['supplier'],
                str(r.get('lead_days', '')),
                str(r.get('spq', '')),
                str(r.get('moq', '')),
                fmt_qty(r.get('monthly_avg', 0)),
                fmt_qty(r.get('yearly_usage', 0)),
                int_disp(r.get('safety_stock')),
                int_disp(r.get('rop')),
                fmt_qty(r.get('current_inventory', 0)),
                fmt_qty(r.get('open_po', 0)),
                fmt_qty(r.get('untransfer_pr', 0)),
                fmt_qty(r.get('inventory_position', 0)),
                r.get('purchase_trigger', ''),
                fmt_qty(r.get('theoretical_purchase_qty', '')),
                fmt_qty(r.get('suggested_purchase_qty', '')),
                r.get('purchase_advice', ''),
                str(r.get('ss_coverage_days', '')),
                r.get('ss_judge', ''),
                parents_disp,
            )
            item = self.abc_tree.insert('', 'end', values=values, tags=(tag,))
            self.abc_tree.tag_configure(tag, background=colors.get(abc, 'white'))

    def _export_abc(self):
        if not self._abc_result:
            messagebox.showwarning('无数据', '请先执行计算')
            return
        path = filedialog.asksaveasfilename(
            title='导出ABC分类结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '外采物料ABC分类'
            monthly_ws = wb.create_sheet('T+3采购计划')
            monthly_summary_ws = wb.create_sheet('T+3下单汇总')
            rolling_ws = wb.create_sheet('月均用量滚动采购计划')
            rolling_summary_ws = wb.create_sheet('月均滚动下单汇总')
            due_arrival_ws = wb.create_sheet('应到货数量清单')
            coverage_check_ws = wb.create_sheet('排产覆盖校验')
            t3_capacity_ws = wb.create_sheet('T+3采购后可生产')
            rolling_capacity_ws = wb.create_sheet('滚动采购后可生产')
            production_risk_ws = wb.create_sheet('月度生产风险清单')
            production_risk_summary_ws = wb.create_sheet('生产风险汇总')
            production_risk_top_ws = wb.create_sheet('TOP风险分类')
            may_focus_risk_ws = wb.create_sheet('5月重点风险前20')
            custom_diff_ws = wb.create_sheet('定制机型差异物料')
            explain_ws = wb.create_sheet('ABC计算说明')

            HDRS = ['分类', '物料编码', '物料分类', '项目', '采购模式', '清尾截止月份', '物料名称', '规格型号', '供应商',
                    '交期(天)', 'SPQ', 'MOQ',
                    '月均用量', '年用量',
                    '安全库存', '再订货点(ROP)', '当前库存', '未清PO', '未转PR', '库存位置',
                    '是否触发采购', '理论采购量', '建议采购量', '采购建议说明',
                    '安全库存覆盖天数', '安全库存判断']
            hdr_font = Font(bold=True, color='000000', size=10)
            hf_a = PatternFill('solid', fgColor='B8860B')
            hf_b = PatternFill('solid', fgColor='4682B4')
            hf_c = PatternFill('solid', fgColor='2E8B57')
            thin = Side(style='thin', color='AAAAAA')
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
            title_fill = PatternFill('solid', fgColor='1F4E78')
            section_fill = PatternFill('solid', fgColor='D9EAF7')
            header_fill = PatternFill('solid', fgColor='D97706')
            soft_fill = PatternFill('solid', fgColor='F8FAFC')

            ws.row_dimensions[1].height = 22
            for ci, h in enumerate(HDRS, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            fills = {'A': PatternFill('solid', fgColor='FFFACD'),
                     'B': PatternFill('solid', fgColor='E0F0FF'),
                     'C': PatternFill('solid', fgColor='E8F5E9')}

            for ri, r in enumerate(self._abc_result, 2):
                row_data = [
                    r['abc'], r['code'], r.get('material_category', ''),
                    r.get('projects', ''),
                    '切换清尾' if r.get('clear_tail') else ('定制备货控制' if r.get('custom_stock_control') else 'ROP备货'),
                    r.get('clear_tail_cutoff_month', ''),
                    r['name'], r['spec'],
                    r['supplier'],
                    r.get('lead_days', ''),
                    r.get('spq', ''), r.get('moq', ''),
                    r.get('monthly_avg', 0),
                    r.get('yearly_usage', 0),
                    r.get('safety_stock', 0),
                    r.get('rop', 0),
                    r.get('current_inventory', 0),
                    r.get('open_po', 0),
                    r.get('untransfer_pr', 0),
                    r.get('inventory_position', 0),
                    r.get('purchase_trigger', ''),
                    r.get('theoretical_purchase_qty', ''),
                    r.get('suggested_purchase_qty', ''),
                    r.get('purchase_advice', ''),
                    r.get('ss_coverage_days', ''),
                    r.get('ss_judge', ''),
                ]
                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = bdr
                    c.fill = fills.get(r['abc'], PatternFill())
                ws.row_dimensions[ri].height = 20

            widths = [6, 16, 12, 18, 10, 12, 22, 16, 14, 8, 6, 6, 10, 10, 8, 10, 10, 10, 10, 12, 10, 10, 10, 42, 10, 14]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HDRS))}{max(ws.max_row, 1)}"

            monthly_hdrs = [
                '物料编码', '项目', '物料分类', '物料名称', '规格型号', '供应商', '分类', '采购模式', '单价', '总价', '清尾截止月份',
                '需求月份', '建议下单月份', '建议下单日期',
                '月需求', '月初库存位置', '月末预计库存位置',
                'ROP(参考)', '补货目标(参考)', '理论采购量', '建议采购量',
                'MOQ', 'SPQ', '交期(天)', '是否触发采购', '说明',
            ]
            for ci, h in enumerate(monthly_hdrs, 1):
                c = monthly_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            def _month_sort_key(month_text):
                text = str(month_text or '')
                try:
                    return datetime.strptime(text, '%Y-%m')
                except ValueError:
                    return datetime.max

            def _month_display(month_text):
                text = str(month_text or '')
                try:
                    dt = datetime.strptime(text, '%Y-%m')
                    return f'{dt.year}年{dt.month}月'
                except ValueError:
                    return text

            def _month_key_from_display_text(value):
                keys = []
                for part in str(value or '').replace(',', '、').split('、'):
                    text = part.strip()
                    if not text:
                        continue
                    match = re.search(r'(\d{4})年(\d{1,2})月', text)
                    if match:
                        keys.append(f'{int(match.group(1)):04d}-{int(match.group(2)):02d}')
                        continue
                    try:
                        dt = datetime.strptime(text, '%Y-%m')
                        keys.append(dt.strftime('%Y-%m'))
                    except ValueError:
                        continue
                return keys

            def _month_start_date(month_text):
                text = str(month_text or '')
                try:
                    return datetime.strptime(text, '%Y-%m').date()
                except ValueError:
                    return None

            def _parse_monthly_order_days():
                try:
                    count = int(self._abc_order_times_per_month.get())
                except (tk.TclError, ValueError, TypeError):
                    count = 2
                count = min(6, max(1, count))
                raw_text = str(getattr(self, '_abc_order_days', tk.StringVar(value='1,15')).get() or '')
                for sep in ('，', '、', ';', '；', ' '):
                    raw_text = raw_text.replace(sep, ',')
                days = []
                for part in raw_text.split(','):
                    part = part.strip()
                    if not part:
                        continue
                    try:
                        day = int(float(part))
                    except (TypeError, ValueError):
                        continue
                    day = min(31, max(1, day))
                    if day not in days:
                        days.append(day)

                if not days:
                    if count == 1:
                        days = [1]
                    elif count == 2:
                        days = [1, 15]
                    else:
                        step = 30 / count
                        days = [min(31, max(1, int(round(1 + step * idx)))) for idx in range(count)]
                if len(days) < count:
                    fallback = [1, 8, 15, 22, 28, 31]
                    for day in fallback:
                        if day not in days:
                            days.append(day)
                        if len(days) >= count:
                            break
                return sorted(days[:count])

            configured_order_days = _parse_monthly_order_days()
            configured_order_days_text = ','.join(str(day) for day in configured_order_days)

            def _order_month_and_date(demand_month_text, lead_days):
                text = str(demand_month_text or '')
                try:
                    demand_month_start = datetime.strptime(text, '%Y-%m')
                except ValueError:
                    return text, '', ''
                order_date = demand_month_start - timedelta(days=int(float(lead_days or 0)))
                today = datetime.today()
                original_order_date = order_date.strftime('%Y-%m-%d')
                if order_date.date() < today.date():
                    return f'{today.year}年{today.month}月', today.strftime('%Y-%m-%d'), original_order_date
                scheduled_date = _next_biweekly_order_date(order_date.date())
                scheduled_date_text = scheduled_date.strftime('%Y-%m-%d')
                return f'{scheduled_date.year}年{scheduled_date.month}月', scheduled_date_text, scheduled_date_text

            def _current_month_batch_dates():
                today = datetime.today()
                return [today.date() + timedelta(days=14 * idx) for idx in range(2)]

            def _add_months(base_date, months):
                month = base_date.month - 1 + months
                year = base_date.year + month // 12
                month = month % 12 + 1
                day = min(base_date.day, calendar.monthrange(year, month)[1])
                return date(year, month, day)

            def _month_key_from_date(value):
                if isinstance(value, datetime):
                    value = value.date()
                if isinstance(value, date):
                    return f'{value.year:04d}-{value.month:02d}'
                return ''

            def _month_display_from_date(value):
                if isinstance(value, datetime):
                    value = value.date()
                if isinstance(value, date):
                    return f'{value.year}年{value.month}月'
                return ''

            def _configured_order_slot_on_or_before(value):
                if isinstance(value, datetime):
                    value = value.date()
                last_day = calendar.monthrange(value.year, value.month)[1]
                candidates = [
                    date(value.year, value.month, min(day, last_day))
                    for day in configured_order_days
                ]
                candidates = sorted(set(candidates))
                valid_candidates = [candidate for candidate in candidates if candidate <= value]
                if valid_candidates:
                    return valid_candidates[-1]
                previous_month = _add_months(date(value.year, value.month, 1), -1)
                previous_last_day = calendar.monthrange(previous_month.year, previous_month.month)[1]
                previous_candidates = [
                    date(previous_month.year, previous_month.month, min(day, previous_last_day))
                    for day in configured_order_days
                ]
                previous_candidates = sorted(set(previous_candidates))
                return previous_candidates[-1] if previous_candidates else previous_month

            def _next_biweekly_order_date(target_date):
                if isinstance(target_date, datetime):
                    target_date = target_date.date()
                today = datetime.today().date()
                if target_date <= today:
                    return today
                scheduled_date = _configured_order_slot_on_or_before(target_date)
                if scheduled_date < today:
                    return today
                return scheduled_date

            def _estimate_month_crossing_date(month_start, start_position, rop_val, monthly_usage):
                if start_position <= rop_val:
                    return month_start
                daily_usage = float(monthly_usage or 0) / 30
                if daily_usage <= 0:
                    return month_start
                days_until_cross = int(math.floor((float(start_position) - float(rop_val)) / daily_usage))
                last_day = calendar.monthrange(month_start.year, month_start.month)[1]
                day = min(max(1, days_until_cross + 1), last_day)
                return date(month_start.year, month_start.month, day)

            def _month_display_sort_key(month_text):
                text = str(month_text or '')
                try:
                    return datetime.strptime(text, '%Y年%m月')
                except ValueError:
                    try:
                        return datetime.strptime(text, '%Y-%m')
                    except ValueError:
                        return datetime.max

            def _date_to_month_display(value):
                if isinstance(value, datetime):
                    return f'{value.year}年{value.month}月'
                if isinstance(value, date):
                    return f'{value.year}年{value.month}月'
                return ''

            def _parse_order_date(value):
                text = str(value or '').strip()
                if not text:
                    return None
                try:
                    return datetime.strptime(text, '%Y-%m-%d').date()
                except ValueError:
                    return None

            monthly_records = []
            material_price_map = self._load_abc_material_price_map(self._abc_mrp_path)

            def _material_unit_price(value):
                code = normalize_material_code(value)
                if not code:
                    return 0.0
                try:
                    return float(material_price_map.get(code, 0) or 0)
                except (TypeError, ValueError):
                    return 0.0

            def _suggest_purchase_qty(theoretical_qty, moq_val, spq_val):
                suggested_qty_raw = float(max(0, theoretical_qty))
                if moq_val > 0:
                    suggested_qty_raw = max(suggested_qty_raw, moq_val)
                if spq_val > 0:
                    suggested_qty_raw = math.ceil(suggested_qty_raw / spq_val) * spq_val
                else:
                    suggested_qty_raw = math.ceil(suggested_qty_raw)
                return int(suggested_qty_raw)

            def _join_project_texts(values):
                projects = set()
                for value in values:
                    for part in str(value or '').split('、'):
                        part = part.strip()
                        if part:
                            projects.add(part)
                return '、'.join(sorted(projects))

            def _projects_for_months(row_data, month_keys):
                monthly_projects = row_data.get('monthly_projects') or {}
                project_text = _join_project_texts(monthly_projects.get(month_key, '') for month_key in month_keys)
                return project_text or str(row_data.get('projects', '') or '')

            def _parent_specs_for_months(row_data, month_keys):
                monthly_parent_specs = row_data.get('monthly_parent_specs') or {}
                spec_text = _join_project_texts(monthly_parent_specs.get(month_key, '') for month_key in month_keys)
                return spec_text or str(row_data.get('parent_specs', '') or '')

            for r in sorted(self._abc_result, key=lambda item: str(item.get('code', ''))):
                if not r.get('has_ext'):
                    continue
                monthly_demand = r.get('monthly_demand') or {}
                if not monthly_demand:
                    continue
                rolling_position = float(r.get('inventory_position', 0) or 0)
                rop_val = float(r.get('rop', 0) or 0)
                moq_val = float(r.get('moq_val', 0) or 0)
                spq_val = float(r.get('spq_val', 0) or 0)
                lead_days = float(r.get('lead_days', 0) or 0)
                sorted_month_keys = sorted(monthly_demand.keys(), key=_month_sort_key)
                today = datetime.today().date()
                coverage_end_date = today + timedelta(days=int(float(lead_days or 0)))
                covered_month_keys = [
                    month_key for month_key in sorted_month_keys
                    if (_month_start_date(month_key) is None or _month_start_date(month_key) <= coverage_end_date)
                ]
                future_month_keys = [
                    month_key for month_key in sorted_month_keys
                    if (_month_start_date(month_key) is not None and _month_start_date(month_key) > coverage_end_date)
                ]

                covered_demand_raw = sum(float(monthly_demand.get(month_key, 0) or 0) for month_key in covered_month_keys)
                covered_plan_shortage = max(0.0, covered_demand_raw - rolling_position)
                plan_purchase_label = (
                    '切换清尾' if r.get('clear_tail')
                    else ('定制备货控制' if r.get('custom_stock_control') else '排产计划')
                )

                # T+3 采购计划只按排产净缺口购买，不再因 ROP 安全库存触发采购。
                if bool(self._abc_mrp_path) and covered_month_keys and covered_plan_shortage > 0:
                    first_month_key = covered_month_keys[0] if covered_month_keys else (sorted_month_keys[0] if sorted_month_keys else '')
                    theoretical_qty = int(math.ceil(covered_plan_shortage))
                    suggested_qty = _suggest_purchase_qty(theoretical_qty, moq_val, spq_val) if theoretical_qty > 0 else 0
                    order_month, order_date, original_order_date = _order_month_and_date(first_month_key, lead_days)
                    triggered_flag = True
                    desc = (
                        f'{plan_purchase_label}：按排产计划购买，不补ROP；覆盖期至{coverage_end_date.strftime("%Y-%m-%d")}，'
                        f'覆盖期内排产需求{fmt_qty(covered_demand_raw)}，库存位置{fmt_qty(rolling_position)}，'
                        f'排产净缺口{fmt_qty(covered_plan_shortage)}；理论采购{fmt_qty(theoretical_qty)}，'
                        f'按MOQ/SPQ建议采购{fmt_qty(suggested_qty)}'
                    )
                    covered_month_display = '、'.join(_month_display(month_key) for month_key in covered_month_keys) or _month_display(first_month_key)
                    covered_projects = _projects_for_months(r, covered_month_keys)
                    month_end_after_covered = rolling_position - covered_demand_raw
                    monthly_records.append({
                        'row_key': r.get('row_key') or r.get('code', ''),
                        'code': r.get('code', ''),
                        'projects': covered_projects,
                        'material_category': r.get('material_category', ''),
                        'name': r.get('name', ''),
                        'spec': r.get('spec', ''),
                        'supplier': r.get('supplier', ''),
                        'abc': r.get('abc', ''),
                        'purchase_mode': plan_purchase_label,
                        'no_rop_purchase': True,
                        'clear_tail': bool(r.get('clear_tail')),
                        'clear_tail_cutoff_month': r.get('clear_tail_cutoff_month', ''),
                        'demand_month': covered_month_display,
                        'order_month': order_month,
                        'order_date': order_date,
                        'month_demand': int(math.ceil(covered_demand_raw)),
                        'month_start_position': int(math.ceil(rolling_position)),
                        'month_end_position': int(math.ceil(month_end_after_covered)),
                        'rop': int(math.ceil(rop_val)),
                        'target_position': int(math.ceil(rop_val)),
                        'theoretical_qty': theoretical_qty,
                        'suggested_qty': suggested_qty,
                        'moq': r.get('moq', ''),
                        'spq': r.get('spq', ''),
                        'moq_val': moq_val,
                        'spq_val': spq_val,
                        'lead_days': r.get('lead_days', ''),
                        'triggered': triggered_flag,
                        'original_order_date': original_order_date,
                        'desc': desc,
                        'skip_merge_desc': True,
                    })
                    rolling_position = month_end_after_covered + suggested_qty
                elif covered_month_keys:
                    covered_demand_raw = sum(float(monthly_demand.get(month_key, 0) or 0) for month_key in covered_month_keys)
                    rolling_position -= covered_demand_raw

                for month_key in future_month_keys:
                    month_demand_raw = float(monthly_demand.get(month_key, 0) or 0)
                    month_demand = int(math.ceil(month_demand_raw))
                    month_start_position = rolling_position
                    month_end_before_purchase = rolling_position - month_demand_raw
                    triggered = bool(self._abc_mrp_path) and month_end_before_purchase < 0
                    theoretical_qty = 0
                    suggested_qty = 0
                    if triggered:
                        theoretical_qty = int(math.ceil(max(0.0, -month_end_before_purchase)))
                        suggested_qty = _suggest_purchase_qty(theoretical_qty, moq_val, spq_val)
                        rolling_position = month_end_before_purchase + suggested_qty
                        order_month, order_date, original_order_date = _order_month_and_date(month_key, lead_days)
                        desc = (
                            f'{plan_purchase_label}：按排产计划购买，不补ROP；'
                            f'月末预计库存位置{fmt_qty(month_end_before_purchase)} < 0；'
                            f'理论采购{fmt_qty(theoretical_qty)}；按MOQ/SPQ建议采购{fmt_qty(suggested_qty)}'
                        )
                        if original_order_date and original_order_date != order_date:
                            desc += f'；原倒推下单日期{original_order_date}已过，已归到当前月份'
                    else:
                        rolling_position = month_end_before_purchase
                        order_month, order_date = '', ''
                        original_order_date = ''
                        if not self._abc_mrp_path:
                            desc = '未接入MRP供给数据，无法滚动判断月度采购'
                        elif month_end_before_purchase >= 0:
                            desc = f'{plan_purchase_label}：月末预计库存位置{fmt_qty(month_end_before_purchase)} >= 0，排产需求已覆盖，本月不采购'
                        else:
                            desc = f'{plan_purchase_label}：本月不触发采购'

                    monthly_records.append({
                        'row_key': r.get('row_key') or r.get('code', ''),
                        'code': r.get('code', ''),
                        'projects': _projects_for_months(r, [month_key]),
                        'material_category': r.get('material_category', ''),
                        'name': r.get('name', ''),
                        'spec': r.get('spec', ''),
                        'supplier': r.get('supplier', ''),
                        'abc': r.get('abc', ''),
                        'purchase_mode': plan_purchase_label,
                        'no_rop_purchase': True,
                        'clear_tail': bool(r.get('clear_tail')),
                        'clear_tail_cutoff_month': r.get('clear_tail_cutoff_month', ''),
                        'demand_month': _month_display(month_key),
                        'order_month': order_month,
                        'order_date': order_date,
                        'month_demand': month_demand,
                        'month_start_position': int(math.ceil(month_start_position)),
                        'month_end_position': int(math.ceil(month_end_before_purchase)),
                        'rop': int(math.ceil(rop_val)),
                        'target_position': int(math.ceil(rop_val)),
                        'theoretical_qty': theoretical_qty,
                        'suggested_qty': suggested_qty,
                        'moq': r.get('moq', ''),
                        'spq': r.get('spq', ''),
                        'moq_val': moq_val,
                        'spq_val': spq_val,
                        'lead_days': r.get('lead_days', ''),
                        'triggered': triggered,
                        'original_order_date': original_order_date,
                        'desc': desc,
                    })

            merged_records = []
            merged_index = {}
            overdue_bucket_counters = defaultdict(int)
            for record in monthly_records:
                if record['triggered'] and record['order_date']:
                    if record.get('original_order_date') and record.get('original_order_date') != record.get('order_date'):
                        batch_dates = _current_month_batch_dates()
                        bucket_key = record['code']
                        batch_date = batch_dates[overdue_bucket_counters[bucket_key] % len(batch_dates)]
                        overdue_bucket_counters[bucket_key] += 1
                        record = dict(record)
                        record['order_date'] = batch_date.strftime('%Y-%m-%d')
                        record['order_month'] = f'{batch_date.year}年{batch_date.month}月'
                    key = (record.get('row_key') or record['code'], record['order_month'])
                    if key not in merged_index:
                        merged = dict(record)
                        merged['demand_months'] = [record['demand_month']]
                        merged['project_texts'] = [record.get('projects', '')] if record.get('projects') else []
                        merged['order_dates'] = [record['order_date']] if record.get('order_date') else []
                        merged['original_order_dates'] = [
                            record['original_order_date']
                        ] if record.get('original_order_date') and record.get('original_order_date') != record.get('order_date') else []
                        merged_index[key] = len(merged_records)
                        merged_records.append(merged)
                    else:
                        merged = merged_records[merged_index[key]]
                        merged['skip_merge_desc'] = False
                        if record['demand_month'] not in merged['demand_months']:
                            merged['demand_months'].append(record['demand_month'])
                        if record.get('projects') and record['projects'] not in merged.get('project_texts', []):
                            merged.setdefault('project_texts', []).append(record['projects'])
                        if record.get('order_date') and record['order_date'] not in merged.get('order_dates', []):
                            merged.setdefault('order_dates', []).append(record['order_date'])
                        if record.get('original_order_date') and record.get('original_order_date') != record.get('order_date') and record['original_order_date'] not in merged['original_order_dates']:
                            merged['original_order_dates'].append(record['original_order_date'])
                        merged['month_demand'] += record['month_demand']
                        merged['month_end_position'] = min(merged['month_end_position'], record['month_end_position'])
                        merged['theoretical_qty'] += record['theoretical_qty']
                        merged['suggested_qty'] = _suggest_purchase_qty(
                            merged['theoretical_qty'],
                            float(merged.get('moq_val', 0) or 0),
                            float(merged.get('spq_val', 0) or 0),
                        )
                else:
                    record = dict(record)
                    record['demand_months'] = [record['demand_month']]
                    record['project_texts'] = [record.get('projects', '')] if record.get('projects') else []
                    record['original_order_dates'] = []
                    merged_records.append(record)

            monthly_row = 2
            for record in merged_records:
                demand_months = '、'.join(record.get('demand_months') or [record.get('demand_month', '')])
                order_dates = '、'.join(record.get('order_dates') or ([record.get('order_date', '')] if record.get('order_date') else []))
                if record.get('skip_merge_desc'):
                    desc = record.get('desc', '')
                elif record['triggered']:
                    if record.get('no_rop_purchase'):
                        mode_label = record.get('purchase_mode') or ('切换清尾' if record.get('clear_tail') else '定制备货控制')
                        desc = (
                            f'{mode_label}合并需求月份:{demand_months}；月需求合计{fmt_qty(record["month_demand"])}；'
                            f'只补排产净缺口，不补ROP；最低月末预计库存位置{fmt_qty(record["month_end_position"])}；'
                            f'理论采购合计{fmt_qty(record["theoretical_qty"])}；'
                            f'按MOQ/SPQ合并建议采购{fmt_qty(record["suggested_qty"])}'
                        )
                    else:
                        desc = (
                            f'合并需求月份:{demand_months}；月需求合计{fmt_qty(record["month_demand"])}；'
                            f'最低月末预计库存位置{fmt_qty(record["month_end_position"])} <= ROP{fmt_qty(record["rop"])}；'
                            f'理论采购合计{fmt_qty(record["theoretical_qty"])}；'
                            f'按MOQ/SPQ合并建议采购{fmt_qty(record["suggested_qty"])}'
                        )
                    if record.get('original_order_dates'):
                        desc += f'；原倒推下单日期{",".join(record["original_order_dates"])}已过，已归到当前月份分批下单日期'
                else:
                    desc = record.get('desc', '')
                record['_output_demand_months'] = demand_months
                record['_output_projects'] = _join_project_texts(record.get('project_texts') or [record.get('projects', '')])
                record['_output_order_dates'] = order_dates
                record['_output_desc'] = desc
                unit_price = _material_unit_price(record.get('code', ''))
                total_amount = float(record.get('suggested_qty', 0) or 0) * unit_price
                row_data = [
                    record.get('code', ''),
                    record.get('_output_projects', ''),
                    record.get('material_category', ''),
                    record.get('name', ''),
                    record.get('spec', ''),
                    record.get('supplier', ''),
                    record.get('abc', ''),
                    record.get('purchase_mode', ''),
                    unit_price,
                    total_amount,
                    record.get('clear_tail_cutoff_month', ''),
                    demand_months,
                    record.get('order_month', ''),
                    order_dates,
                    record.get('month_demand', 0),
                    record.get('month_start_position', 0),
                    record.get('month_end_position', 0),
                    record.get('rop', 0),
                    record.get('target_position', 0),
                    record.get('theoretical_qty', 0),
                    record.get('suggested_qty', 0),
                    record.get('moq', ''),
                    record.get('spq', ''),
                    record.get('lead_days', ''),
                    '是' if record.get('triggered') else '否',
                    desc,
                ]
                for ci, val in enumerate(row_data, 1):
                    c = monthly_ws.cell(row=monthly_row, column=ci, value=val)
                    c.border = bdr
                    c.fill = fills.get(record.get('abc'), PatternFill())
                    if ci in (9, 10):
                        c.number_format = '#,##0.00'
                monthly_row += 1

            monthly_widths = [16, 18, 14, 22, 16, 16, 6, 10, 10, 12, 12, 12, 12, 12, 10, 12, 14, 10, 12, 12, 12, 8, 8, 9, 10, 54]
            for i, w in enumerate(monthly_widths, 1):
                monthly_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            monthly_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(monthly_hdrs))}{max(monthly_ws.max_row, 1)}"
            monthly_ws.freeze_panes = 'A2'

            summary_fill = PatternFill('solid', fgColor='E8F4FD')
            detail_fill = PatternFill('solid', fgColor='F8FAFC')
            total_fill = PatternFill('solid', fgColor='D9EAF7')
            monthly_summary_ws.merge_cells('A1:F1')
            monthly_summary_ws['A1'] = 'T+3下单汇总'
            monthly_summary_ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
            monthly_summary_ws['A1'].fill = PatternFill('solid', fgColor='1F4E78')
            monthly_summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            summary_headers = ['月份', '项目', '理论下单数量(不按MOQ/SPQ)', '建议下单数量(按MOQ/SPQ)', '下单料号数', '说明']
            for ci, h in enumerate(summary_headers, 1):
                c = monthly_summary_ws.cell(row=2, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            monthly_summary = defaultdict(lambda: {
                'theoretical_qty': 0,
                'suggested_qty': 0,
                'order_codes': set(),
                'projects': set(),
            })
            monthly_details = []

            for record in merged_records:
                if not record.get('triggered'):
                    continue
                theoretical_qty = int(math.ceil(float(record.get('theoretical_qty', 0) or 0)))
                suggested_qty = int(math.ceil(float(record.get('suggested_qty', 0) or 0)))
                if theoretical_qty <= 0 and suggested_qty <= 0:
                    continue
                code = record.get('code', '')
                name = record.get('name', '')
                supplier = record.get('supplier', '')
                order_month = record.get('order_month', '')
                projects = record.get('_output_projects') or record.get('projects', '')
                demand_months = record.get('_output_demand_months') or '、'.join(record.get('demand_months') or [record.get('demand_month', '')])
                order_dates_display = record.get('_output_order_dates') or '、'.join(record.get('order_dates') or ([record.get('order_date', '')] if record.get('order_date') else []))
                desc = record.get('_output_desc') or record.get('desc', '')

                if order_month:
                    monthly_summary[order_month]['theoretical_qty'] += theoretical_qty
                    monthly_summary[order_month]['suggested_qty'] += suggested_qty
                    if code:
                        monthly_summary[order_month]['order_codes'].add(code)
                    for project in str(projects or '').split('、'):
                        project = project.strip()
                        if project:
                            monthly_summary[order_month]['projects'].add(project)
                    monthly_details.append([
                        order_month, projects, record.get('material_category', ''), code, name, supplier, theoretical_qty, suggested_qty,
                        order_dates_display, demand_months, desc,
                    ])

            summary_row = 3
            for month_text in sorted(monthly_summary.keys(), key=_month_display_sort_key):
                item = monthly_summary[month_text]
                row_values = [
                    month_text,
                    '、'.join(sorted(item['projects'])),
                    int(item['theoretical_qty']),
                    int(item['suggested_qty']),
                    len(item['order_codes']),
                    '理论量不按MOQ/SPQ；建议量按MOQ/SPQ修正；不再推算预计到货',
                ]
                for ci, val in enumerate(row_values, 1):
                    c = monthly_summary_ws.cell(row=summary_row, column=ci, value=val)
                    c.border = bdr
                    c.fill = summary_fill
                summary_row += 1

            detail_start = summary_row + 2
            monthly_summary_ws.cell(row=detail_start - 1, column=1, value='下单明细').font = Font(bold=True)
            detail_headers = ['建议下单月份', '项目', '物料分类', '物料编码', '物料名称', '供应商', '理论下单数量(不按MOQ/SPQ)', '建议下单数量(按MOQ/SPQ)', '建议下单日期', '需求月份', '说明']
            for ci, h in enumerate(detail_headers, 1):
                c = monthly_summary_ws.cell(row=detail_start, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            for ri, row_values in enumerate(monthly_details, detail_start + 1):
                for ci, val in enumerate(row_values, 1):
                    c = monthly_summary_ws.cell(row=ri, column=ci, value=val)
                    c.border = bdr
                    c.fill = detail_fill

            summary_widths = [12, 18, 14, 20, 20, 12, 44, 16, 22, 18, 56]
            for i, w in enumerate(summary_widths, 1):
                monthly_summary_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            monthly_summary_ws.auto_filter.ref = f"A{detail_start}:{openpyxl.utils.get_column_letter(len(detail_headers))}{max(monthly_summary_ws.max_row, detail_start)}"
            monthly_summary_ws.freeze_panes = 'A3'

            try:
                rolling_months = int(self._rolling_purchase_months.get())
            except (tk.TclError, ValueError, TypeError):
                rolling_months = 6
            rolling_months = min(24, max(1, rolling_months))

            rolling_hdrs = [
                '物料编码', '项目', '物料分类', '物料名称', '规格型号', '供应商', '分类', '采购模式', '单价', '总价', '清尾截止月份',
                '计划月份', '月均用量/本月折算用量', '月初库存位置',
                '月末预计库存位置', 'ROP', '补货目标(ROP)', '理论采购量', '建议采购量',
                'MOQ', 'SPQ', '交期(天)', '建议下单月份', '建议下单日期',
                '是否触发采购', '说明',
            ]
            for ci, h in enumerate(rolling_hdrs, 1):
                c = rolling_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            rolling_summary = defaultdict(lambda: {
                'theoretical_qty': 0,
                'suggested_qty': 0,
                'order_codes': set(),
                'projects': set(),
            })
            rolling_details = []
            rolling_purchase_by_row_month = defaultdict(lambda: defaultdict(float))
            current_month_start = date(datetime.today().year, datetime.today().month, 1)
            today = datetime.today().date()

            rolling_row = 2
            for r in sorted(self._abc_result, key=lambda item: str(item.get('code', ''))):
                if not r.get('has_ext'):
                    continue
                rolling_position = float(r.get('inventory_position', 0) or 0)
                monthly_usage = float(r.get('monthly_avg', 0) or 0)
                if monthly_usage <= 0:
                    continue
                rop_val = float(r.get('rop', 0) or 0)
                moq_val = float(r.get('moq_val', 0) or 0)
                spq_val = float(r.get('spq_val', 0) or 0)
                monthly_demand = r.get('monthly_demand') or {}
                try:
                    lead_days_int = int(float(r.get('lead_days', 0) or 0))
                except (TypeError, ValueError):
                    lead_days_int = 0
                code = r.get('code', '')
                no_rop_purchase = bool(r.get('clear_tail') or r.get('custom_stock_control'))
                no_rop_label = '切换清尾' if r.get('clear_tail') else '定制备货控制'

                for month_offset in range(rolling_months):
                    month_start = _add_months(current_month_start, month_offset)
                    month_key = _month_key_from_date(month_start)
                    projects = _projects_for_months(r, [month_key]) or str(r.get('projects', '') or '')
                    if no_rop_purchase:
                        cutoff_month = str(r.get('clear_tail_cutoff_month', '') or '').strip()
                        if cutoff_month and month_key > cutoff_month:
                            continue
                        month_usage = float(monthly_demand.get(month_key, 0) or 0)
                        if month_usage <= 0:
                            continue
                        usage_desc = f'{no_rop_label}按排产计划需求'
                    elif month_offset == 0:
                        days_in_month = calendar.monthrange(month_start.year, month_start.month)[1]
                        remaining_days = max(0, days_in_month - today.day + 1)
                        month_usage_base = monthly_usage * remaining_days / 30
                        month_plan_demand = float(monthly_demand.get(month_key, 0) or 0)
                        month_usage = max(month_usage_base, month_plan_demand)
                        usage_desc = f'本月按剩余{remaining_days}天折算'
                        if month_plan_demand > month_usage_base:
                            usage_desc += f'；本月排产需求{fmt_qty(month_plan_demand)}高于折算月均{fmt_qty(month_usage_base)}，按排产需求判断'
                    else:
                        month_usage_base = monthly_usage
                        month_plan_demand = float(monthly_demand.get(month_key, 0) or 0)
                        month_usage = max(month_usage_base, month_plan_demand)
                        usage_desc = '整月月均用量'
                        if month_plan_demand > month_usage_base:
                            usage_desc += f'；本月排产需求{fmt_qty(month_plan_demand)}高于月均{fmt_qty(month_usage_base)}，按排产需求判断'

                    month_start_position = rolling_position
                    month_end_before_purchase = month_start_position - month_usage
                    plan_shortage_qty = max(0.0, month_plan_demand - month_start_position) if not no_rop_purchase else max(0.0, -month_end_before_purchase)
                    if no_rop_purchase:
                        triggered = bool(self._abc_mrp_path) and month_end_before_purchase < 0
                    else:
                        triggered = bool(self._abc_mrp_path) and (month_end_before_purchase <= rop_val or plan_shortage_qty > 0)
                    theoretical_qty = 0
                    suggested_qty = 0
                    order_month = ''
                    order_date_text = ''

                    if triggered:
                        if no_rop_purchase:
                            theoretical_qty = int(math.ceil(max(0.0, -month_end_before_purchase)))
                        else:
                            rop_shortage_qty = max(0.0, rop_val - month_end_before_purchase)
                            theoretical_qty = int(math.ceil(max(rop_shortage_qty, plan_shortage_qty)))
                        suggested_qty = _suggest_purchase_qty(theoretical_qty, moq_val, spq_val)
                        crossing_date = _estimate_month_crossing_date(month_start, month_start_position, rop_val, month_usage)
                        order_dt = _next_biweekly_order_date(crossing_date)
                        order_month = _month_display_from_date(order_dt)
                        order_date_text = order_dt.strftime('%Y-%m-%d')
                        rolling_position = month_end_before_purchase + suggested_qty
                        rolling_purchase_by_row_month[r.get('row_key') or code][month_key] += suggested_qty

                        rolling_summary[order_month]['theoretical_qty'] += theoretical_qty
                        rolling_summary[order_month]['suggested_qty'] += suggested_qty
                        if code:
                            rolling_summary[order_month]['order_codes'].add(code)
                        for project in str(projects or '').split('、'):
                            project = project.strip()
                            if project:
                                rolling_summary[order_month]['projects'].add(project)
                        detail_desc = (
                            f'月均滚动计划：{usage_desc}，按排产计划判断净缺口，不补ROP'
                            if no_rop_purchase else
                            f'月均滚动计划：{usage_desc}，排产净缺口强制触发采购'
                            if plan_shortage_qty > 0 else
                            f'月均滚动计划：{usage_desc}，库存位置低于/等于ROP后按每月{configured_order_days_text}号排单'
                        )
                        rolling_details.append([
                            order_month, projects, r.get('material_category', ''), code, r.get('name', ''), r.get('supplier', ''), theoretical_qty, suggested_qty,
                            order_date_text, _month_display_from_date(month_start),
                            detail_desc,
                        ])
                        if no_rop_purchase:
                            desc = (
                                f'{usage_desc}；只补排产净缺口，不补ROP；'
                                f'月末预计库存位置{fmt_qty(month_end_before_purchase)} < 0；'
                                f'理论采购{fmt_qty(theoretical_qty)}，按MOQ/SPQ建议采购{fmt_qty(suggested_qty)}；'
                                f'下单日期按每月{configured_order_days_text}号排到{order_date_text}'
                            )
                        else:
                            if plan_shortage_qty > 0:
                                desc = (
                                    f'{usage_desc}；排产净缺口{fmt_qty(plan_shortage_qty)}，强制触发采购；'
                                    f'理论采购{fmt_qty(theoretical_qty)}，按MOQ/SPQ建议采购{fmt_qty(suggested_qty)}；'
                                    f'下单日期按每月{configured_order_days_text}号排到{order_date_text}'
                                )
                            else:
                                desc = (
                                    f'{usage_desc}；月末预计库存位置{fmt_qty(month_end_before_purchase)} <= ROP{fmt_qty(rop_val)}；'
                                    f'理论采购{fmt_qty(theoretical_qty)}，按MOQ/SPQ建议采购{fmt_qty(suggested_qty)}；'
                                    f'下单日期按每月{configured_order_days_text}号排到{order_date_text}'
                                )
                    else:
                        rolling_position = month_end_before_purchase
                        if not self._abc_mrp_path:
                            desc = '未接入MRP供给数据，无法滚动判断月度采购'
                        elif no_rop_purchase:
                            desc = f'{usage_desc}；月末预计库存位置{fmt_qty(month_end_before_purchase)} >= 0，排产需求已覆盖，本月不采购'
                        else:
                            desc = f'{usage_desc}；月末预计库存位置{fmt_qty(month_end_before_purchase)} > ROP{fmt_qty(rop_val)}，本月不触发采购'

                    unit_price = _material_unit_price(code)
                    total_amount = float(suggested_qty or 0) * unit_price
                    row_data = [
                        code,
                        projects,
                        r.get('material_category', ''),
                        r.get('name', ''),
                        r.get('spec', ''),
                        r.get('supplier', ''),
                        r.get('abc', ''),
                        '切换清尾' if r.get('clear_tail') else ('定制备货控制' if r.get('custom_stock_control') else 'ROP备货'),
                        unit_price,
                        total_amount,
                        r.get('clear_tail_cutoff_month', ''),
                        _month_display_from_date(month_start),
                        int(math.ceil(month_usage)),
                        int(math.ceil(month_start_position)),
                        int(math.ceil(month_end_before_purchase)),
                        int(math.ceil(rop_val)),
                        int(math.ceil(rop_val)),
                        theoretical_qty,
                        suggested_qty,
                        r.get('moq', ''),
                        r.get('spq', ''),
                        r.get('lead_days', ''),
                        order_month,
                        order_date_text,
                        '是' if triggered else '否',
                        desc,
                    ]
                    for ci, val in enumerate(row_data, 1):
                        c = rolling_ws.cell(row=rolling_row, column=ci, value=val)
                        c.border = bdr
                        c.fill = fills.get(r.get('abc'), PatternFill())
                        if ci in (9, 10):
                            c.number_format = '#,##0.00'
                    rolling_row += 1

            rolling_widths = [16, 18, 14, 22, 16, 16, 6, 10, 10, 12, 12, 12, 16, 12, 14, 10, 12, 12, 12, 8, 8, 9, 12, 12, 10, 62]
            for i, w in enumerate(rolling_widths, 1):
                rolling_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            rolling_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(rolling_hdrs))}{max(rolling_ws.max_row, 1)}"
            rolling_ws.freeze_panes = 'A2'

            rolling_summary_ws.merge_cells('A1:F1')
            rolling_summary_ws['A1'] = '月均用量滚动下单汇总'
            rolling_summary_ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
            rolling_summary_ws['A1'].fill = title_fill
            rolling_summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            for ci, h in enumerate(summary_headers, 1):
                c = rolling_summary_ws.cell(row=2, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            summary_row = 3
            for month_text in sorted(rolling_summary.keys(), key=_month_display_sort_key):
                item = rolling_summary[month_text]
                row_values = [
                    month_text,
                    '、'.join(sorted(item['projects'])),
                    int(item['theoretical_qty']),
                    int(item['suggested_qty']),
                    len(item['order_codes']),
                    f'按月均用量滚动{rolling_months}个月；理论量不按MOQ/SPQ，建议量按MOQ/SPQ；未来下单日按每月{configured_order_days_text}号，过期订单仍归到当前月分批',
                ]
                for ci, val in enumerate(row_values, 1):
                    c = rolling_summary_ws.cell(row=summary_row, column=ci, value=val)
                    c.border = bdr
                    c.fill = summary_fill
                summary_row += 1

            detail_start = summary_row + 2
            rolling_summary_ws.cell(row=detail_start - 1, column=1, value='下单明细').font = Font(bold=True)
            for ci, h in enumerate(detail_headers, 1):
                c = rolling_summary_ws.cell(row=detail_start, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            for ri, row_values in enumerate(rolling_details, detail_start + 1):
                for ci, val in enumerate(row_values, 1):
                    c = rolling_summary_ws.cell(row=ri, column=ci, value=val)
                    c.border = bdr
                    c.fill = detail_fill

            for i, w in enumerate(summary_widths, 1):
                rolling_summary_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            rolling_summary_ws.auto_filter.ref = f"A{detail_start}:{openpyxl.utils.get_column_letter(len(detail_headers))}{max(rolling_summary_ws.max_row, detail_start)}"
            rolling_summary_ws.freeze_panes = 'A3'

            # 硬校验：购买建议数量必须覆盖排产计划需求。这里不再看ROP是否合理，只检查排产需求能否被库存位置+建议采购量覆盖。
            coverage_headers = [
                '计划模式', '月份', '校验结果', '物料编码', '项目', '物料分类', '物料名称', '规格型号', '供应商',
                '月需求', '月初库存位置', '本月建议采购量', '月末覆盖余额', '排产缺口', '说明',
            ]
            fail_fill = PatternFill('solid', fgColor='F8CBAD')
            ok_fill = PatternFill('solid', fgColor='D9EAD3')
            for ci, h in enumerate(coverage_headers, 1):
                c = coverage_check_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            t3_purchase_by_row_month = defaultdict(lambda: defaultdict(float))
            for record in merged_records:
                if not record.get('triggered'):
                    continue
                try:
                    suggested_qty = float(record.get('suggested_qty', 0) or 0)
                except (TypeError, ValueError):
                    suggested_qty = 0
                if suggested_qty <= 0:
                    continue
                demand_month_keys = []
                for demand_text in record.get('demand_months') or [record.get('demand_month', '')]:
                    demand_month_keys.extend(_month_key_from_display_text(demand_text))
                demand_month_keys = sorted(set(demand_month_keys), key=_month_sort_key)
                if not demand_month_keys:
                    continue
                # 合并采购先放到最早需求月，用于验证“数量够不够覆盖排产”，日期风险由生产风险页判断。
                t3_purchase_by_row_month[record.get('row_key') or record.get('code', '')][demand_month_keys[0]] += suggested_qty

            def _write_coverage_rows(mode_name, purchase_by_row_month):
                output_rows = []
                for r in sorted(self._abc_result, key=lambda item: str(item.get('code', ''))):
                    monthly_demand = r.get('monthly_demand') or {}
                    if not monthly_demand:
                        continue
                    row_key = r.get('row_key') or r.get('code', '')
                    balance = float(r.get('inventory_position', 0) or 0)
                    for month_key in sorted(monthly_demand.keys(), key=_month_sort_key):
                        try:
                            demand_qty = float(monthly_demand.get(month_key, 0) or 0)
                        except (TypeError, ValueError):
                            demand_qty = 0
                        if demand_qty <= 0:
                            continue
                        purchase_qty = float(purchase_by_row_month.get(row_key, {}).get(month_key, 0) or 0)
                        month_start_balance = balance
                        month_end_balance = month_start_balance + purchase_qty - demand_qty
                        gap_qty = max(0.0, -month_end_balance)
                        result = '不足' if gap_qty > 1e-9 else 'OK'
                        if result == '不足':
                            desc = (
                                f'{mode_name}购买建议未覆盖排产：月初库存位置{fmt_qty(month_start_balance)} + '
                                f'本月建议采购{fmt_qty(purchase_qty)} - 月需求{fmt_qty(demand_qty)} = {fmt_qty(month_end_balance)}'
                            )
                        else:
                            desc = (
                                f'{mode_name}已覆盖排产：月末覆盖余额{fmt_qty(month_end_balance)}；'
                                f'本校验只看数量覆盖，交期是否赶得上看“月度生产风险清单”。'
                            )
                        output_rows.append([
                            mode_name,
                            _month_display(month_key),
                            result,
                            r.get('code', ''),
                            _projects_for_months(r, [month_key]) or r.get('projects', ''),
                            r.get('material_category', ''),
                            r.get('name', ''),
                            r.get('spec', ''),
                            r.get('supplier', ''),
                            int(math.ceil(demand_qty)),
                            int(math.ceil(month_start_balance)),
                            int(math.ceil(purchase_qty)),
                            int(math.floor(month_end_balance)) if month_end_balance >= 0 else int(math.floor(month_end_balance)),
                            int(math.ceil(gap_qty)),
                            desc,
                        ])
                        balance = month_end_balance
                return output_rows

            coverage_rows = _write_coverage_rows('T+3采购计划', t3_purchase_by_row_month)
            coverage_rows.extend(_write_coverage_rows('月均滚动采购计划', rolling_purchase_by_row_month))
            coverage_rows.sort(key=lambda row: (0 if row[2] == '不足' else 1, row[0], _month_display_sort_key(row[1]), str(row[3])))

            coverage_row_idx = 2
            for row_values in coverage_rows:
                is_fail = row_values[2] == '不足'
                for ci, val in enumerate(row_values, 1):
                    c = coverage_check_ws.cell(row=coverage_row_idx, column=ci, value=val)
                    c.border = bdr
                    c.fill = fail_fill if is_fail else (ok_fill if ci == 3 else detail_fill)
                    c.alignment = Alignment(vertical='center', wrap_text=True)
                coverage_row_idx += 1
            if not coverage_rows:
                coverage_check_ws.cell(row=2, column=1, value='无排产需求可校验').border = bdr
            coverage_widths = [16, 12, 10, 16, 20, 14, 22, 18, 22, 10, 14, 14, 14, 12, 82]
            for i, w in enumerate(coverage_widths, 1):
                coverage_check_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            coverage_check_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(coverage_headers))}{max(coverage_check_ws.max_row, 1)}"
            coverage_check_ws.freeze_panes = 'A2'

            capacity_headers = [
                '采购模式', '计划月份', '上线日期', '项目', '母料号', '母件规格/型号', '计划数量',
                '建议采购后可生产上限', '按本行计划可生产', '未满足数量', '状态',
                '短板物料', '短板可生产台数', '最大物料缺口', '说明',
            ]

            def _capacity_code_keys(value):
                text = str(value or '').strip()
                keys = {text} if text else set()
                material_code = maybe_material_code(value)
                if material_code:
                    keys.add(material_code)
                normalized = normalize_material_code(value)
                if normalized:
                    keys.add(normalized)
                return keys

            bom_key_by_code_for_capacity = {}
            for bom_key in self.bom_index.keys():
                for key in _capacity_code_keys(bom_key):
                    bom_key_by_code_for_capacity.setdefault(key, bom_key)

            def _resolve_bom_key_for_capacity(value):
                for key in _capacity_code_keys(value):
                    if key in self.bom_index:
                        return key
                    if key in bom_key_by_code_for_capacity:
                        return bom_key_by_code_for_capacity[key]
                return None

            try:
                _capacity_excluded_codes, capacity_usage_exclusions = self._load_abc_exclusion_rules(self._abc_mrp_path)
            except Exception:
                capacity_usage_exclusions = set()

            def _is_capacity_usage_excluded(parent_code, child_code):
                return any(
                    (parent_key, child_key) in capacity_usage_exclusions
                    for parent_key in _capacity_code_keys(parent_code)
                    for child_key in _capacity_code_keys(child_code)
                )

            material_rows_by_code = {}
            row_key_to_code = {}
            for item in self._abc_result:
                code = normalize_material_code(item.get('code', ''))
                if not code:
                    continue
                material_rows_by_code.setdefault(code, item)
                row_key_to_code[item.get('row_key') or code] = code
            capacity_material_codes = set(material_rows_by_code.keys())

            def _build_capacity_start_supply():
                result = defaultdict(float)
                if self._abc_mrp_path:
                    inventory_map, po_map, pr_map = self._load_abc_mrp_supply_maps(self._abc_mrp_path)

                    def _lookup_supply(mapping, code):
                        for key in _capacity_code_keys(code):
                            if key in mapping:
                                try:
                                    return float(mapping.get(key, 0) or 0)
                                except (TypeError, ValueError):
                                    return 0.0
                        return 0.0

                    for code in capacity_material_codes:
                        result[code] = (
                            _lookup_supply(inventory_map, code)
                            + _lookup_supply(po_map, code)
                            + _lookup_supply(pr_map, code)
                        )
                else:
                    for item in self._abc_result:
                        code = normalize_material_code(item.get('code', ''))
                        if code:
                            result[code] += float(item.get('inventory_position', 0) or 0)
                return result

            capacity_unit_cache = {}

            def _capacity_unit_requirements(root_pn):
                root_code = normalize_material_code(root_pn)
                if root_code in capacity_unit_cache:
                    return capacity_unit_cache[root_code]
                totals = defaultdict(float)
                root_key = _resolve_bom_key_for_capacity(root_pn)
                if not root_key:
                    capacity_unit_cache[root_code] = {}
                    return {}

                def _walk(pn, qty, depth, trail):
                    resolved_pn = _resolve_bom_key_for_capacity(pn)
                    if depth > MAX_DEPTH or not resolved_pn or resolved_pn not in self.bom_index:
                        return
                    for row in self.bom_index[resolved_pn]:
                        child_pn = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
                        if not child_pn:
                            continue
                        if _is_capacity_usage_excluded(resolved_pn, child_pn):
                            continue
                        child_code = normalize_material_code(child_pn)
                        resolved_child = _resolve_bom_key_for_capacity(child_pn)
                        if child_pn in trail or child_code in trail or resolved_child in trail or child_pn == resolved_pn:
                            continue
                        try:
                            child_qty = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else 1.0
                        except (ValueError, TypeError):
                            child_qty = 1.0
                        total_qty = qty * child_qty
                        if child_code in capacity_material_codes:
                            totals[child_code] += total_qty
                        _walk(child_pn, total_qty, depth + 1, trail | {resolved_pn, child_code})

                _walk(root_key, 1.0, 1, {root_key, normalize_material_code(root_key)})
                capacity_unit_cache[root_code] = dict(totals)
                return capacity_unit_cache[root_code]

            def _purchase_by_code_month(purchase_by_row_month):
                by_code_month = defaultdict(lambda: defaultdict(float))
                for row_key, month_map in (purchase_by_row_month or {}).items():
                    code = row_key_to_code.get(row_key) or normalize_material_code(str(row_key).split('|')[0])
                    if not code:
                        continue
                    for month_key, qty in (month_map or {}).items():
                        try:
                            by_code_month[code][month_key] += float(qty or 0)
                        except (TypeError, ValueError):
                            pass
                return by_code_month

            def _plan_row_month_key(plan_date):
                if isinstance(plan_date, datetime):
                    return plan_date.strftime('%Y-%m')
                if isinstance(plan_date, date):
                    return f'{plan_date.year:04d}-{plan_date.month:02d}'
                return '未指定月份'

            def _plan_row_date_text(plan_date):
                if isinstance(plan_date, datetime):
                    return plan_date.strftime('%Y-%m-%d')
                if isinstance(plan_date, date):
                    return plan_date.strftime('%Y-%m-%d')
                return ''

            def _plan_sort_key(row):
                plan_date = row[2] if len(row) > 2 else None
                month_key = _plan_row_month_key(plan_date)
                date_key = plan_date.date() if isinstance(plan_date, datetime) else plan_date
                if not isinstance(date_key, date):
                    date_key = date.max
                return (_month_sort_key(month_key), date_key, str(row[0]))

            def _write_capacity_sheet(target_ws, mode_name, purchase_by_row_month):
                for ci, h in enumerate(capacity_headers, 1):
                    c = target_ws.cell(row=1, column=ci, value=h)
                    c.font = hdr_font
                    c.fill = header_fill
                    c.border = bdr
                    c.alignment = Alignment(horizontal='center', vertical='center')

                start_supply = _build_capacity_start_supply()
                available = defaultdict(float, start_supply)
                purchases = _purchase_by_code_month(purchase_by_row_month)
                purchase_months = sorted(
                    {month_key for month_map in purchases.values() for month_key in month_map.keys()},
                    key=_month_sort_key,
                )
                added_months = set()

                def _add_purchases_until(month_key):
                    for purchase_month in purchase_months:
                        if purchase_month in added_months:
                            continue
                        if month_key == '未指定月份' or _month_sort_key(purchase_month) <= _month_sort_key(month_key):
                            for code, month_map in purchases.items():
                                available[code] += float(month_map.get(purchase_month, 0) or 0)
                            added_months.add(purchase_month)

                output_row = 2
                plan_rows = list(getattr(self, '_abc_plan_expansion_rows', []) or [])
                for mother_pn, mother_qty, plan_date, project, mother_spec in sorted(plan_rows, key=_plan_sort_key):
                    try:
                        plan_qty = float(mother_qty or 0)
                    except (TypeError, ValueError):
                        plan_qty = 0.0
                    if plan_qty <= 0:
                        continue
                    month_key = _plan_row_month_key(plan_date)
                    _add_purchases_until(month_key)
                    unit_totals = _capacity_unit_requirements(mother_pn)
                    if not unit_totals:
                        row_values = [
                            mode_name, _month_display(month_key), _plan_row_date_text(plan_date),
                            project or '', mother_pn, mother_spec or '', int(math.ceil(plan_qty)),
                            '', 0, int(math.ceil(plan_qty)), '未识别下层',
                            '', '', '', '该母料号未展开到ABC外购物料，可能是BOM缺失、外采清单缺主数据或被排除规则剔除。',
                        ]
                    else:
                        supported_items = []
                        shortage_items = []
                        max_gap = 0.0
                        for code, unit_qty in unit_totals.items():
                            unit_qty = float(unit_qty or 0)
                            if unit_qty <= 0:
                                continue
                            qty_available = float(available.get(code, 0) or 0)
                            supported = math.floor(qty_available / unit_qty + 1e-9)
                            required_qty = plan_qty * unit_qty
                            gap_qty = max(0.0, required_qty - qty_available)
                            item = material_rows_by_code.get(code, {})
                            supported_items.append((supported, code, item.get('name', ''), qty_available, unit_qty, gap_qty))
                            if gap_qty > 1e-9:
                                shortage_items.append((supported, code, item.get('name', ''), gap_qty))
                                max_gap = max(max_gap, gap_qty)

                        if supported_items:
                            capacity_limit = max(0, int(min(item[0] for item in supported_items)))
                            plan_producible = min(int(math.floor(plan_qty + 1e-9)), capacity_limit)
                        else:
                            capacity_limit = 0
                            plan_producible = 0
                        unmet_qty = max(0, int(math.ceil(plan_qty - plan_producible)))
                        if unmet_qty <= 0:
                            status = '满足'
                        elif plan_producible > 0:
                            status = '部分满足'
                        else:
                            status = '不满足'

                        bottleneck_items = sorted(
                            shortage_items or supported_items,
                            key=lambda item: (item[0], str(item[1])),
                        )[:5]
                        if shortage_items:
                            bottleneck_text = '\n'.join(
                                f'{code} {name} 缺口{fmt_qty(gap)}'
                                for _supported, code, name, gap in bottleneck_items
                            )
                        else:
                            bottleneck_text = '\n'.join(
                                f'{code} {name} 可支持{fmt_qty(supported)}台'
                                for supported, code, name, _available_qty, _unit_qty, _gap in bottleneck_items
                            )
                        bottleneck_capacity = min((item[0] for item in supported_items), default='')
                        desc = (
                            f'按{mode_name}建议采购量加入库存位置后，按排产顺序逐行扣减共用物料；'
                            f'本行计划{fmt_qty(plan_qty)}台，可生产{fmt_qty(plan_producible)}台。'
                        )
                        if shortage_items:
                            desc += '短板物料按当前行需求缺口展示；共用物料已被前序排产消耗。'
                        row_values = [
                            mode_name, _month_display(month_key), _plan_row_date_text(plan_date),
                            project or '', mother_pn, mother_spec or '', int(math.ceil(plan_qty)),
                            capacity_limit, plan_producible, unmet_qty, status,
                            bottleneck_text, bottleneck_capacity, int(math.ceil(max_gap)), desc,
                        ]

                        for code, unit_qty in unit_totals.items():
                            available[code] = max(0.0, float(available.get(code, 0) or 0) - plan_producible * float(unit_qty or 0))

                    fill = ok_fill if row_values[10] == '满足' else (PatternFill('solid', fgColor='FFF2CC') if row_values[10] == '部分满足' else fail_fill)
                    for ci, val in enumerate(row_values, 1):
                        c = target_ws.cell(row=output_row, column=ci, value=val)
                        c.border = bdr
                        c.fill = fill if ci in (10, 11) else detail_fill
                        c.alignment = Alignment(vertical='center', wrap_text=True)
                    output_row += 1

                if output_row == 2:
                    target_ws.cell(row=2, column=1, value='无排产行可测算').border = bdr

                capacity_widths = [16, 12, 12, 18, 16, 24, 10, 18, 16, 12, 10, 48, 14, 14, 76]
                for idx, width in enumerate(capacity_widths, 1):
                    target_ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
                target_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(capacity_headers))}{max(target_ws.max_row, 1)}"
                target_ws.freeze_panes = 'A2'

            _write_capacity_sheet(t3_capacity_ws, 'T+3采购计划', t3_purchase_by_row_month)
            _write_capacity_sheet(rolling_capacity_ws, '月均滚动采购计划', rolling_purchase_by_row_month)

            due_arrival_headers = [
                '月份', '最晚到货日期', '物料编码', '项目', '物料分类', '物料名称', '规格型号', '供应商', '采购',
                '月需求', '当前库存覆盖', '排产净缺口', '应到货数量',
                '未清PO应到货', '未转PR应到货', '购买建议应到货', '采购建议覆盖后缺口',
                '剩余未清PO', '剩余未转PR', '交期(天)', '风险提示', '说明',
            ]
            for ci, h in enumerate(due_arrival_headers, 1):
                c = due_arrival_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            due_arrival_row = 2
            for r in sorted(self._abc_result, key=lambda item: str(item.get('code', ''))):
                monthly_demand = r.get('monthly_demand') or {}
                if not monthly_demand:
                    continue
                inventory_remaining = float(r.get('current_inventory', 0) or 0)
                open_po_remaining = float(r.get('open_po', 0) or 0)
                untransfer_pr_remaining = float(r.get('untransfer_pr', 0) or 0)
                for month_key in sorted(monthly_demand.keys(), key=_month_sort_key):
                    month_demand = float(monthly_demand.get(month_key, 0) or 0)
                    if month_demand <= 0:
                        continue
                    month_start = _month_start_date(month_key)
                    due_date_text = month_start.strftime('%Y-%m-%d') if month_start else ''
                    inventory_cover = min(inventory_remaining, month_demand)
                    need_after_inventory = max(0.0, month_demand - inventory_cover)
                    open_po_due = min(open_po_remaining, need_after_inventory)
                    need_after_po = max(0.0, need_after_inventory - open_po_due)
                    pr_due = min(untransfer_pr_remaining, need_after_po)
                    need_after_pr = max(0.0, need_after_po - pr_due)
                    purchase_suggestion_due = need_after_pr
                    uncovered_gap = max(0.0, need_after_pr - purchase_suggestion_due)
                    required_arrival = open_po_due + pr_due + purchase_suggestion_due
                    inventory_remaining = max(0.0, inventory_remaining - inventory_cover)
                    open_po_remaining = max(0.0, open_po_remaining - open_po_due)
                    untransfer_pr_remaining = max(0.0, untransfer_pr_remaining - pr_due)

                    if required_arrival <= 0:
                        continue
                    if not r.get('has_ext') and purchase_suggestion_due > 0:
                        risk_hint = '缺采购主数据，需采购提供'
                    elif purchase_suggestion_due > 0:
                        risk_hint = '依赖购买建议按期到货'
                    elif pr_due > 0:
                        risk_hint = '依赖未转PR及时转PO并到货'
                    else:
                        risk_hint = '依赖未清PO按期到货'
                    desc = (
                        f'先用当前库存覆盖{fmt_qty(inventory_cover)}，'
                        f'排产净缺口{fmt_qty(need_after_inventory)}，需到货{fmt_qty(required_arrival)}；'
                        f'其中未清PO覆盖{fmt_qty(open_po_due)}，未转PR覆盖{fmt_qty(pr_due)}，'
                        f'购买建议按生产净需求补足{fmt_qty(purchase_suggestion_due)}，'
                        f'覆盖后缺口{fmt_qty(uncovered_gap)}。'
                        f'此清单不引用T+3购买建议池，只按生产计划净需求排到货，不额外补ROP，不计算金额。'
                    )
                    if not r.get('has_ext') and purchase_suggestion_due > 0:
                        desc += '该料号缺外采主数据，数量已列入购买建议应到货，但需先补供应商/采购/交期等资料。'
                    row_values = [
                        _month_display(month_key),
                        due_date_text,
                        r.get('code', ''),
                        _projects_for_months(r, [month_key]) or r.get('projects', ''),
                        r.get('material_category', ''),
                        r.get('name', ''),
                        r.get('spec', ''),
                        r.get('supplier', ''),
                        r.get('buyer', ''),
                        int(math.ceil(month_demand)),
                        int(math.ceil(inventory_cover)),
                        int(math.ceil(need_after_inventory)),
                        int(math.ceil(required_arrival)),
                        int(math.ceil(open_po_due)),
                        int(math.ceil(pr_due)),
                        int(math.ceil(purchase_suggestion_due)),
                        int(math.ceil(uncovered_gap)),
                        int(math.ceil(open_po_remaining)),
                        int(math.ceil(untransfer_pr_remaining)),
                        r.get('lead_days', ''),
                        risk_hint,
                        desc,
                    ]
                    for ci, val in enumerate(row_values, 1):
                        c = due_arrival_ws.cell(row=due_arrival_row, column=ci, value=val)
                        c.border = bdr
                        c.fill = detail_fill
                        c.alignment = Alignment(vertical='center', wrap_text=True)
                    due_arrival_row += 1

            due_arrival_widths = [12, 14, 16, 20, 14, 22, 18, 22, 10, 10, 12, 12, 12, 12, 12, 14, 16, 12, 12, 9, 24, 86]
            for i, w in enumerate(due_arrival_widths, 1):
                due_arrival_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            due_arrival_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(due_arrival_headers))}{max(due_arrival_ws.max_row, 1)}"
            due_arrival_ws.freeze_panes = 'A2'

            # 生产风险：把T+3采购建议按下单日期+交期折算为预计到货，再逐月扣减排产需求。
            risk_fills = {
                'R': PatternFill('solid', fgColor='F8CBAD'),
                'Y': PatternFill('solid', fgColor='FFF2CC'),
                'G': PatternFill('solid', fgColor='D9EAD3'),
            }
            risk_hdrs = [
                '月份', 'RYG', '风险类别', '物料编码', '项目', '物料名称', '规格型号', '供应商',
                '物料分类', 'ABC分类', '采购模式', '月需求', '月初可用库存', '本月未清PO到货', '建议采购到货数量',
                '最晚到货日期', '预计到货日期', '交期是否满足', '需采购跟催',
                '到货后月末库存位置', 'ROP', '风险缺口',
                '当前库存', '未清PO总量', '有日期未清PO', '日期未知未清PO', '未转PR(不参与)', '交期(天)',
                '理论采购量', '建议采购量', '说明',
            ]
            for ci, h in enumerate(risk_hdrs, 1):
                c = production_risk_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            risk_records = []
            risk_summary = defaultdict(lambda: {
                'R_count': 0,
                'Y_count': 0,
                'G_count': 0,
                'R_demand': 0.0,
                'Y_demand': 0.0,
                'G_demand': 0.0,
                'R_gap': 0.0,
                'risk_codes': set(),
                'expedite_codes': set(),
                'top_items': [],
            })
            top_by_code = {}
            purchase_arrivals_by_code = defaultdict(list)
            for record in merged_records:
                if not record.get('triggered'):
                    continue
                try:
                    suggested_qty = float(record.get('suggested_qty', 0) or 0)
                except (TypeError, ValueError):
                    suggested_qty = 0
                if suggested_qty <= 0:
                    continue
                order_dates = record.get('order_dates') or ([record.get('order_date', '')] if record.get('order_date') else [])
                parsed_dates = [dt for dt in (_parse_order_date(value) for value in order_dates) if dt]
                if not parsed_dates:
                    continue
                try:
                    lead_days = int(float(record.get('lead_days', 0) or 0))
                except (TypeError, ValueError):
                    lead_days = 0
                order_dt = min(parsed_dates)
                arrival_dt = order_dt + timedelta(days=max(0, lead_days))
                purchase_arrivals_by_code[record.get('row_key') or record.get('code', '')].append({
                    'arrival_date': arrival_dt,
                    'qty': suggested_qty,
                    'order_date': order_dt,
                })
            for arrival_items in purchase_arrivals_by_code.values():
                arrival_items.sort(key=lambda item: item['arrival_date'])

            for r in sorted(self._abc_result, key=lambda item: str(item.get('code', ''))):
                monthly_demand = r.get('monthly_demand') or {}
                if not monthly_demand:
                    continue
                position = float(r.get('current_inventory', 0) or 0)
                rop_val = float(r.get('rop', 0) or 0)
                purchase_mode = '切换清尾' if r.get('clear_tail') else ('定制备货控制' if r.get('custom_stock_control') else 'ROP备货')
                dated_po_arrivals = list(r.get('open_po_arrivals') or [])
                dated_po_total = sum(float(item.get('qty', 0) or 0) for item in dated_po_arrivals)
                unknown_po_qty = max(0.0, float(r.get('open_po', 0) or 0) - dated_po_total)
                arrival_items = []
                for item in dated_po_arrivals:
                    arrival_items.append({
                        'arrival_date': item.get('arrival_date'),
                        'qty': float(item.get('qty', 0) or 0),
                        'source': '未清PO',
                    })
                for item in purchase_arrivals_by_code.get(r.get('row_key') or r.get('code', ''), []):
                    arrival_items.append({
                        'arrival_date': item.get('arrival_date'),
                        'qty': float(item.get('qty', 0) or 0),
                        'source': '建议采购',
                    })
                arrival_items = [item for item in arrival_items if item.get('arrival_date') and float(item.get('qty', 0) or 0) > 0]
                arrival_items.sort(key=lambda item: item['arrival_date'])
                arrival_index = 0
                for month_key in sorted(monthly_demand.keys(), key=_month_sort_key):
                    month_demand_raw = float(monthly_demand.get(month_key, 0) or 0)
                    if month_demand_raw <= 0:
                        continue
                    month_start_position = position
                    month_start_date = _month_start_date(month_key)
                    if month_start_date:
                        month_end_date = date(
                            month_start_date.year,
                            month_start_date.month,
                            calendar.monthrange(month_start_date.year, month_start_date.month)[1],
                        )
                    else:
                        month_end_date = date.max
                    arrival_qty = 0.0
                    po_arrival_qty = 0.0
                    suggested_arrival_qty = 0.0
                    arrival_dates = []
                    late_arrival_dates = []
                    while arrival_index < len(arrival_items) and arrival_items[arrival_index]['arrival_date'] <= month_end_date:
                        arrival_item = arrival_items[arrival_index]
                        qty = float(arrival_item.get('qty', 0) or 0)
                        arrival_qty += qty
                        if arrival_item.get('source') == '未清PO':
                            po_arrival_qty += qty
                        else:
                            suggested_arrival_qty += qty
                        arrival_date = arrival_item['arrival_date']
                        arrival_dates.append(f'{arrival_date.strftime("%Y-%m-%d")}({arrival_item.get("source", "")})')
                        if arrival_item.get('source') != '未清PO' and month_start_date and arrival_date > month_start_date:
                            late_arrival_dates.append(arrival_date.strftime('%Y-%m-%d'))
                        arrival_index += 1
                    position_after_arrival = month_start_position + arrival_qty
                    month_end_without_arrival = month_start_position - month_demand_raw
                    month_end_position = position_after_arrival - month_demand_raw
                    projects = _projects_for_months(r, [month_key]) or str(r.get('projects', '') or '')
                    parent_specs = _parent_specs_for_months(r, [month_key])
                    due_date_text = month_start_date.strftime('%Y-%m-%d') if month_start_date else ''
                    lead_time_ok = '否' if late_arrival_dates else ('是' if arrival_dates else '')
                    expedite_required = '是' if late_arrival_dates else '否'

                    if not r.get('has_ext'):
                        ryg = 'R'
                        risk_type = '缺采购主数据'
                        risk_gap = int(math.ceil(month_demand_raw))
                        desc = '外采清单缺少交期/MOQ/SPQ/供应商等主数据，无法判断供给节奏，按红色风险处理'
                    elif not self._abc_mrp_path:
                        ryg = 'Y'
                        risk_type = '缺MRP供给数据'
                        risk_gap = ''
                        desc = '未接入MRP库存/未清PO/未转PR，只能看到排产需求，无法判断库存位置'
                    elif month_end_position < 0:
                        ryg = 'R'
                        risk_type = '采购到货后仍缺口' if arrival_qty > 0 else '生产缺口'
                        risk_gap = int(math.ceil(abs(month_end_position)))
                        desc = (
                            f'当前可用库存{fmt_qty(month_start_position)}，本月未清PO到货{fmt_qty(po_arrival_qty)}，'
                            f'建议采购到货{fmt_qty(suggested_arrival_qty)}，'
                            f'扣减{_month_display(month_key)}需求{fmt_qty(month_demand_raw)}后月末预计为{fmt_qty(month_end_position)}，'
                            f'不足以覆盖该月生产，缺口{fmt_qty(risk_gap)}'
                        )
                    elif late_arrival_dates:
                        ryg = 'R'
                        risk_type = '交期不满足需跟催'
                        risk_gap = 0
                        desc = (
                            f'建议采购预计到货日期{",".join(late_arrival_dates)}晚于最晚到货日期{due_date_text}，'
                            f'数量到货后能覆盖，但按标准交期赶不上生产，需要采购协调提前/跟催'
                        )
                    elif r.get('clear_tail'):
                        if month_end_without_arrival < 0 and arrival_qty > 0:
                            ryg = 'Y'
                            risk_type = '清尾依赖采购按期到货'
                        else:
                            ryg = 'G'
                            risk_type = '清尾需求已覆盖'
                        risk_gap = 0
                        desc = (
                            f'切换清尾物料只看排产净需求；本月未清PO到货{fmt_qty(po_arrival_qty)}，'
                            f'建议采购到货{fmt_qty(suggested_arrival_qty)}，'
                            f'扣减本月需求后月末预计{fmt_qty(month_end_position)}'
                        )
                    elif month_end_position < rop_val:
                        ryg = 'Y'
                        if month_end_without_arrival < 0 and arrival_qty > 0:
                            risk_type = '依赖采购按期到货'
                        else:
                            risk_type = '低于ROP缓冲'
                        risk_gap = int(math.ceil(max(0.0, rop_val - month_end_position)))
                        desc = (
                            f'本月未清PO到货{fmt_qty(po_arrival_qty)}，建议采购到货{fmt_qty(suggested_arrival_qty)}，'
                            f'能覆盖{_month_display(month_key)}生产，'
                            f'但月末预计库存位置{fmt_qty(month_end_position)}低于ROP{fmt_qty(rop_val)}，安全缓冲不足'
                        )
                    else:
                        if month_end_without_arrival < 0 and arrival_qty > 0:
                            ryg = 'Y'
                            risk_type = '依赖采购按期到货'
                        else:
                            ryg = 'G'
                            risk_type = '健康'
                        risk_gap = 0
                        desc = (
                            f'本月未清PO到货{fmt_qty(po_arrival_qty)}，建议采购到货{fmt_qty(suggested_arrival_qty)}，扣减本月需求后月末预计库存位置'
                            f'{fmt_qty(month_end_position)} >= ROP{fmt_qty(rop_val)}'
                        )
                    if arrival_dates:
                        desc += f'；预计到货日期:{",".join(arrival_dates)}'
                    if late_arrival_dates:
                        desc += f'；需采购跟催:是'

                    record = {
                        'month_key': month_key,
                        'month': _month_display(month_key),
                        'ryg': ryg,
                        'risk_type': risk_type,
                        'code': r.get('code', ''),
                        'projects': projects,
                        'parent_specs': parent_specs,
                        'name': r.get('name', ''),
                        'spec': r.get('spec', ''),
                        'supplier': r.get('supplier', ''),
                        'buyer': r.get('buyer', ''),
                        'material_category': r.get('material_category', '') or '未分类',
                        'abc': r.get('abc', ''),
                        'purchase_mode': purchase_mode,
                        'month_demand': int(math.ceil(month_demand_raw)),
                        'start_position': int(math.ceil(month_start_position)),
                        'po_arrival_qty': int(math.ceil(po_arrival_qty)),
                        'arrival_qty': int(math.ceil(suggested_arrival_qty)),
                        'due_date': due_date_text,
                        'arrival_dates': '、'.join(arrival_dates),
                        'lead_time_ok': lead_time_ok,
                        'expedite_required': expedite_required,
                        'end_position': int(math.ceil(month_end_position)),
                        'rop': int(math.ceil(rop_val)),
                        'risk_gap': risk_gap,
                        'current_inventory': r.get('current_inventory', 0),
                        'open_po': r.get('open_po', 0),
                        'dated_po_total': int(math.ceil(dated_po_total)),
                        'unknown_po_qty': int(math.ceil(unknown_po_qty)),
                        'untransfer_pr': r.get('untransfer_pr', 0),
                        'lead_days': r.get('lead_days', ''),
                        'theoretical_purchase_qty': r.get('theoretical_purchase_qty', ''),
                        'suggested_purchase_qty': r.get('suggested_purchase_qty', ''),
                        'desc': desc,
                    }
                    risk_records.append(record)

                    summary_item = risk_summary[record['month']]
                    summary_item[f'{ryg}_count'] += 1
                    summary_item[f'{ryg}_demand'] += month_demand_raw
                    if ryg == 'R':
                        summary_item['R_gap'] += float(risk_gap or 0)
                    if ryg in ('R', 'Y') and record['code']:
                        summary_item['risk_codes'].add(record['code'])
                        summary_item['top_items'].append(record)
                    if record.get('expedite_required') == '是' and record.get('code'):
                        summary_item['expedite_codes'].add(record['code'])

                    if ryg in ('R', 'Y'):
                        aggregate = top_by_code.setdefault(record['code'], {
                            'code': record['code'],
                            'projects': set(),
                            'name': record['name'],
                            'supplier': record['supplier'],
                            'abc': record['abc'],
                            'purchase_mode': record['purchase_mode'],
                            'earliest_month_key': record['month_key'],
                            'earliest_month': record['month'],
                            'max_ryg': ryg,
                            'red_months': 0,
                            'yellow_months': 0,
                            'total_demand': 0.0,
                            'max_gap': 0.0,
                            'min_end_position': record['end_position'],
                            'lead_days': record['lead_days'],
                            'risk_types': set(),
                            'desc': record['desc'],
                        })
                        for project in str(projects or '').split('、'):
                            project = project.strip()
                            if project:
                                aggregate['projects'].add(project)
                        if _month_sort_key(record['month_key']) < _month_sort_key(aggregate['earliest_month_key']):
                            aggregate['earliest_month_key'] = record['month_key']
                            aggregate['earliest_month'] = record['month']
                        if ryg == 'R':
                            aggregate['max_ryg'] = 'R'
                            aggregate['red_months'] += 1
                        else:
                            aggregate['yellow_months'] += 1
                        aggregate['total_demand'] += month_demand_raw
                        aggregate['max_gap'] = max(aggregate['max_gap'], float(risk_gap or 0))
                        aggregate['min_end_position'] = min(aggregate['min_end_position'], record['end_position'])
                        aggregate['risk_types'].add(risk_type)
                    position = month_end_position

            category_risk_summary = defaultdict(lambda: {
                'R_count': 0,
                'Y_count': 0,
                'G_count': 0,
                'risk_codes': set(),
                'expedite_codes': set(),
                'reason_counts': defaultdict(int),
                'sample_codes': [],
                'projects': set(),
            })
            for record in risk_records:
                category = str(record.get('material_category', '') or '未分类').strip() or '未分类'
                key = (record.get('month', ''), category)
                item = category_risk_summary[key]
                ryg = record.get('ryg', '')
                if ryg in ('R', 'Y', 'G'):
                    item[f'{ryg}_count'] += 1
                if ryg in ('R', 'Y') and record.get('code'):
                    item['risk_codes'].add(record['code'])
                    if record['code'] not in item['sample_codes']:
                        item['sample_codes'].append(record['code'])
                    item['reason_counts'][record.get('risk_type', '') or '其他风险'] += 1
                if record.get('expedite_required') == '是' and record.get('code'):
                    item['expedite_codes'].add(record['code'])
                for project in str(record.get('projects', '') or '').split('、'):
                    project = project.strip()
                    if project:
                        item['projects'].add(project)

            risk_row = 2
            for record in risk_records:
                row_values = [
                    record['month'],
                    record['ryg'],
                    record['risk_type'],
                    record['code'],
                    record['projects'],
                    record['name'],
                    record['spec'],
                    record['supplier'],
                    record['material_category'],
                    record['abc'],
                    record['purchase_mode'],
                    record['month_demand'],
                    record['start_position'],
                    record['po_arrival_qty'],
                    record['arrival_qty'],
                    record['due_date'],
                    record['arrival_dates'],
                    record['lead_time_ok'],
                    record['expedite_required'],
                    record['end_position'],
                    record['rop'],
                    record['risk_gap'],
                    record['current_inventory'],
                    record['open_po'],
                    record['dated_po_total'],
                    record['unknown_po_qty'],
                    record['untransfer_pr'],
                    record['lead_days'],
                    record['theoretical_purchase_qty'],
                    record['suggested_purchase_qty'],
                    record['desc'],
                ]
                for ci, val in enumerate(row_values, 1):
                    c = production_risk_ws.cell(row=risk_row, column=ci, value=val)
                    c.border = bdr
                    c.fill = risk_fills.get(record['ryg'], detail_fill)
                risk_row += 1

            risk_widths = [12, 6, 16, 16, 20, 22, 16, 16, 12, 8, 10, 10, 12, 12, 12, 14, 24, 12, 12, 14, 10, 10, 10, 12, 12, 12, 9, 12, 12, 72]
            for i, w in enumerate(risk_widths, 1):
                production_risk_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            production_risk_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(risk_hdrs))}{max(production_risk_ws.max_row, 1)}"
            production_risk_ws.freeze_panes = 'A2'

            production_risk_summary_ws.merge_cells('A1:I1')
            production_risk_summary_ws['A1'] = '生产风险汇总'
            production_risk_summary_ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
            production_risk_summary_ws['A1'].fill = title_fill
            production_risk_summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            risk_summary_headers = [
                '月份', '最高风险物料分类', 'R物料数', 'Y物料数', 'G物料数',
                '风险料号数', '需跟催料号数', '主要风险原因', '说明',
            ]
            for ci, h in enumerate(risk_summary_headers, 1):
                c = production_risk_summary_ws.cell(row=2, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')
            summary_row = 3
            for month_text in sorted(risk_summary.keys(), key=_month_display_sort_key):
                item = risk_summary[month_text]
                month_category_items = []
                for (category_month, category), category_item in category_risk_summary.items():
                    if category_month != month_text:
                        continue
                    risk_code_count = len(category_item['risk_codes'])
                    if risk_code_count <= 0:
                        continue
                    reason_text = '、'.join(
                        f'{reason}{count}项'
                        for reason, count in sorted(category_item['reason_counts'].items(), key=lambda kv: (-kv[1], kv[0]))[:3]
                    )
                    month_category_items.append((category, category_item, reason_text))
                month_category_items.sort(
                    key=lambda item_tuple: (
                        -item_tuple[1]['R_count'],
                        -len(item_tuple[1]['expedite_codes']),
                        -item_tuple[1]['Y_count'],
                        -len(item_tuple[1]['risk_codes']),
                        item_tuple[0],
                    )
                )
                top_category = month_category_items[0][0] if month_category_items else ''
                top_reason = month_category_items[0][2] if month_category_items else ''
                row_values = [
                    month_text,
                    top_category,
                    item['R_count'],
                    item['Y_count'],
                    item['G_count'],
                    len(item['risk_codes']),
                    len(item['expedite_codes']),
                    top_reason,
                    '按风险料号数和需跟催料号数判断，不汇总不同物料的数量缺口',
                ]
                for ci, val in enumerate(row_values, 1):
                    c = production_risk_summary_ws.cell(row=summary_row, column=ci, value=val)
                    c.border = bdr
                    c.fill = summary_fill
                summary_row += 1
            for i, w in enumerate([12, 18, 10, 10, 10, 12, 14, 34, 52], 1):
                production_risk_summary_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            production_risk_summary_ws.auto_filter.ref = f"A2:I{max(production_risk_summary_ws.max_row, 2)}"
            production_risk_summary_ws.freeze_panes = 'A3'

            top_headers = [
                '月份', '排名', 'RYG', '物料分类', 'R物料数', 'Y物料数',
                '风险料号数', '需跟催料号数', '主要风险原因', '代表料号', '涉及项目', '管理建议',
            ]
            for ci, h in enumerate(top_headers, 1):
                c = production_risk_top_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')
            top_category_rows = []
            for month_text in sorted(risk_summary.keys(), key=_month_display_sort_key):
                month_category_items = []
                for (category_month, category), category_item in category_risk_summary.items():
                    if category_month != month_text or len(category_item['risk_codes']) <= 0:
                        continue
                    reason_text = '、'.join(
                        f'{reason}{count}项'
                        for reason, count in sorted(category_item['reason_counts'].items(), key=lambda kv: (-kv[1], kv[0]))[:3]
                    )
                    highest_ryg = 'R' if category_item['R_count'] > 0 else 'Y'
                    advice = (
                        '优先采购跟催，确认能否提前到货；同步检查替代/库存调拨'
                        if category_item['R_count'] > 0 else
                        '持续跟踪到货，避免延期击穿生产'
                    )
                    month_category_items.append({
                        'month': month_text,
                        'category': category,
                        'highest_ryg': highest_ryg,
                        'R_count': category_item['R_count'],
                        'Y_count': category_item['Y_count'],
                        'risk_code_count': len(category_item['risk_codes']),
                        'expedite_count': len(category_item['expedite_codes']),
                        'reason_text': reason_text,
                        'sample_codes': '、'.join(category_item['sample_codes'][:8]),
                        'projects': '、'.join(sorted(category_item['projects'])[:8]),
                        'advice': advice,
                    })
                month_category_items.sort(
                    key=lambda item: (
                        0 if item['highest_ryg'] == 'R' else 1,
                        -item['R_count'],
                        -item['expedite_count'],
                        -item['Y_count'],
                        -item['risk_code_count'],
                        item['category'],
                    )
                )
                for rank, item in enumerate(month_category_items[:10], start=1):
                    item['rank'] = rank
                    top_category_rows.append(item)
            for row_idx, item in enumerate(top_category_rows, 2):
                row_values = [
                    item['month'],
                    item['rank'],
                    item['highest_ryg'],
                    item['category'],
                    item['R_count'],
                    item['Y_count'],
                    item['risk_code_count'],
                    item['expedite_count'],
                    item['reason_text'],
                    item['sample_codes'],
                    item['projects'],
                    item['advice'],
                ]
                for ci, val in enumerate(row_values, 1):
                    c = production_risk_top_ws.cell(row=row_idx, column=ci, value=val)
                    c.border = bdr
                    c.fill = risk_fills.get(item['highest_ryg'], detail_fill)
            top_widths = [12, 8, 6, 18, 10, 10, 12, 14, 34, 42, 28, 52]
            for i, w in enumerate(top_widths, 1):
                production_risk_top_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            production_risk_top_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(top_headers))}{max(production_risk_top_ws.max_row, 1)}"
            production_risk_top_ws.freeze_panes = 'A2'

            may_focus_headers = ['采购', '物料编码', '物料描述', '供应商', '短缺', '影响项目', '风险原因']
            may_focus_fill = PatternFill('solid', fgColor='00A6D6')
            may_focus_font = Font(bold=True, color='FFFFFF', size=10)
            for ci, h in enumerate(may_focus_headers, 1):
                c = may_focus_risk_ws.cell(row=1, column=ci, value=h)
                c.font = may_focus_font
                c.fill = may_focus_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')

            def _is_may_record(record):
                month_key = str(record.get('month_key', '') or '')
                try:
                    return datetime.strptime(month_key, '%Y-%m').month == 5
                except ValueError:
                    return str(record.get('month', '') or '').endswith('5月')

            def _risk_gap_display(value):
                if value in ('', None):
                    return ''
                try:
                    qty = float(value)
                    return int(math.ceil(qty))
                except (TypeError, ValueError):
                    return value

            may_focus_categories = [
                '钣金件-研发',
                '电子板卡类',
                '电机类',
                '电子类',
                '机加件-研发',
                '减速机类',
                '驱动类',
                '轴承类',
                '线缆类',
                '铸件-研发',
            ]
            may_focus_category_set = set(may_focus_categories)
            may_focus_records = [
                record for record in risk_records
                if (
                    _is_may_record(record)
                    and record.get('ryg') in ('R', 'Y')
                    and str(record.get('material_category', '') or '').strip() in may_focus_category_set
                )
            ]
            may_focus_by_category = defaultdict(list)
            for record in may_focus_records:
                category = str(record.get('material_category', '') or '').strip()
                may_focus_by_category[category].append(record)
            for records in may_focus_by_category.values():
                records.sort(
                    key=lambda record: (
                        -float(record.get('lead_days', 0) or 0),
                        0 if record.get('ryg') == 'R' else 1,
                        str(record.get('code', '') or ''),
                    )
                )

            output_row = 2
            for category in may_focus_categories:
                records = may_focus_by_category.get(category, [])
                may_focus_risk_ws.merge_cells(start_row=output_row, start_column=1, end_row=output_row, end_column=len(may_focus_headers))
                section_cell = may_focus_risk_ws.cell(
                    row=output_row,
                    column=1,
                    value=f'{category}（5月R/Y风险，按交期降序，前20）' if records else f'{category}（5月暂无R/Y风险）',
                )
                section_cell.font = Font(bold=True, color='000000', size=10)
                section_cell.fill = section_fill
                section_cell.border = bdr
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                output_row += 1

                for record in records[:20]:
                    material_desc = ' '.join(
                        part for part in [
                            str(record.get('name', '') or '').strip(),
                            str(record.get('spec', '') or '').strip(),
                        ]
                        if part
                    )
                    impact_projects = str(record.get('projects', '') or '').strip().replace('、', '\n')
                    row_values = [
                        record.get('buyer', ''),
                        record.get('code', ''),
                        material_desc,
                        record.get('supplier', ''),
                        _risk_gap_display(record.get('risk_gap', '')),
                        impact_projects,
                        record.get('risk_type', ''),
                    ]
                    for ci, val in enumerate(row_values, 1):
                        c = may_focus_risk_ws.cell(row=output_row, column=ci, value=val)
                        c.border = bdr
                        c.fill = risk_fills.get(record.get('ryg'), detail_fill)
                        c.alignment = Alignment(vertical='center', wrap_text=True)
                    project_line_count = max(1, len([part for part in impact_projects.split('\n') if part.strip()]))
                    may_focus_risk_ws.row_dimensions[output_row].height = min(90, max(18, project_line_count * 15))
                    output_row += 1
            if len(may_focus_records) == 0:
                may_focus_risk_ws.cell(row=2, column=1, value='指定分类在5月未识别到R/Y风险物料').border = bdr
            may_focus_risk_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(may_focus_headers))}{max(may_focus_risk_ws.max_row, 1)}"
            may_focus_risk_ws.freeze_panes = 'A2'
            for i, w in enumerate([10, 16, 34, 26, 10, 28, 42], 1):
                may_focus_risk_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            may_focus_risk_ws.sheet_view.showGridLines = False
            may_focus_risk_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(may_focus_headers))}{max(may_focus_risk_ws.max_row, 1)}"
            may_focus_risk_ws.freeze_panes = 'A2'

            custom_diff_headers = [
                '定制母料号', '定制型号', '客户/项目', '定制有效数量',
                '物料编码', '物料名称', '规格型号', '物料分类', '供应商', '采购',
                '差异类型', '定制单台用量', '其他项目最大单台用量', '单台差异',
                '定制差异数量', '对比其他母料', '备注'
            ]
            for ci, h in enumerate(custom_diff_headers, 1):
                c = custom_diff_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = header_fill
                c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            custom_diff_fills = {
                '定制专用': PatternFill('solid', fgColor='FCE4D6'),
                '定制多用': PatternFill('solid', fgColor='FFF2CC'),
                '定制少用': PatternFill('solid', fgColor='E2F0D9'),
                '其他项目有/定制不用': PatternFill('solid', fgColor='D9EAF7'),
            }
            for row_idx, item in enumerate(getattr(self, '_abc_custom_diff_rows', []) or [], 2):
                row_values = [
                    item.get('custom_root', ''),
                    item.get('custom_model', ''),
                    item.get('custom_project', ''),
                    fmt_qty(item.get('custom_effective_qty', 0)),
                    item.get('code', ''),
                    item.get('name', ''),
                    item.get('spec', ''),
                    item.get('material_category', ''),
                    item.get('supplier', ''),
                    item.get('buyer', ''),
                    item.get('diff_type', ''),
                    fmt_qty(item.get('custom_unit_qty', 0)),
                    fmt_qty(item.get('other_max_unit_qty', 0)),
                    fmt_qty(item.get('diff_unit_qty', 0)),
                    fmt_qty(item.get('custom_diff_qty', 0)),
                    item.get('compare_sources', ''),
                    item.get('remark', ''),
                ]
                for ci, val in enumerate(row_values, 1):
                    c = custom_diff_ws.cell(row=row_idx, column=ci, value=val)
                    c.border = bdr
                    c.fill = custom_diff_fills.get(item.get('diff_type'), soft_fill)
                    c.alignment = Alignment(vertical='center', wrap_text=True)
                source_line_count = max(1, len([part for part in str(item.get('compare_sources', '')).split('\n') if part.strip()]))
                custom_diff_ws.row_dimensions[row_idx].height = min(90, max(18, source_line_count * 15))
            if not getattr(self, '_abc_custom_diff_rows', None):
                note = '未识别到定制机型差异物料；请确认MRP计算表是否填写“客户定制机型备货控制”，且对应母料号当前有排产和BOM。'
                custom_diff_ws.cell(row=2, column=1, value=note).border = bdr
                custom_diff_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(custom_diff_headers))
                custom_diff_ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical='center')
                custom_diff_ws.row_dimensions[2].height = 36
            for i, w in enumerate([16, 24, 18, 12, 16, 24, 22, 14, 24, 10, 16, 12, 16, 12, 14, 42, 42], 1):
                custom_diff_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            custom_diff_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(custom_diff_headers))}{max(custom_diff_ws.max_row, 1)}"
            custom_diff_ws.freeze_panes = 'A2'
            custom_diff_ws.sheet_view.showGridLines = False

            meta = getattr(self, '_abc_calc_meta', {}) or {}
            plan_days = meta.get('plan_days', '')
            daily_factor = meta.get('daily_factor', '')
            lead_a = meta.get('lead_a', '')
            lead_b = meta.get('lead_b', '')
            k_a = meta.get('k_a', '')
            k_b = meta.get('k_b', '')
            k_c = meta.get('k_c', '')
            mrp_path = meta.get('mrp_path', '')
            excluded_codes_count = meta.get('excluded_codes_count', 0)
            usage_exclusion_pairs_count = meta.get('usage_exclusion_pairs_count', 0)
            production_parent_count = meta.get('production_parent_count', 0)
            material_category_rule_count = meta.get('material_category_rule_count', 0)
            manual_clear_tail_rule_count = meta.get('manual_clear_tail_rule_count', 0)
            auto_clear_tail_rule_count = meta.get('auto_clear_tail_rule_count', 0)
            model_switch_rule_count = meta.get('model_switch_rule_count', 0)
            model_switch_applied_count = meta.get('model_switch_applied_count', 0)
            model_switch_clear_tail_material_count = meta.get('model_switch_clear_tail_material_count', 0)
            model_switch_missing_old_count = meta.get('model_switch_missing_old_count', 0)
            model_switch_missing_new_count = meta.get('model_switch_missing_new_count', 0)
            custom_stock_rule_count = meta.get('custom_stock_rule_count', 0)
            custom_stock_controlled_mother_count = meta.get('custom_stock_controlled_mother_count', 0)
            custom_stock_original_qty = meta.get('custom_stock_original_qty', 0)
            custom_stock_effective_qty = meta.get('custom_stock_effective_qty', 0)
            custom_stock_trimmed_qty = meta.get('custom_stock_trimmed_qty', 0)
            subplan_dedupe_enabled = meta.get('subplan_dedupe_enabled', True)
            subplan_dedupe_original_rows = meta.get('subplan_dedupe_original_rows', 0)
            subplan_dedupe_effective_rows = meta.get('subplan_dedupe_effective_rows', 0)
            subplan_dedupe_covered_rows = meta.get('subplan_dedupe_covered_rows', 0)
            subplan_dedupe_partial_rows = meta.get('subplan_dedupe_partial_rows', 0)
            subplan_dedupe_covered_qty = meta.get('subplan_dedupe_covered_qty', 0)
            subplan_dedupe_affected_code_count = meta.get('subplan_dedupe_affected_code_count', 0)
            replacement_diff_rule_count = meta.get('replacement_diff_rule_count', 0)
            replacement_diff_applied_count = meta.get('replacement_diff_applied_count', 0)
            replacement_diff_positive_material_count = meta.get('replacement_diff_positive_material_count', 0)
            replacement_diff_missing_base_count = meta.get('replacement_diff_missing_base_count', 0)
            replacement_diff_missing_replacement_count = meta.get('replacement_diff_missing_replacement_count', 0)
            example = next((item for item in self._abc_result if item.get('has_ext')), self._abc_result[0])
            missing_example = next((item for item in self._abc_result if not item.get('has_ext')), None)

            example_code = str(example.get('code', '') or '')
            example_name = str(example.get('name', '') or '')
            example_abc = str(example.get('abc', '') or '')
            example_total_qty = float(example.get('total_qty', 0) or 0)
            example_avg_daily = float(example.get('avg_daily', 0) or 0)
            example_monthly = float(example.get('monthly_avg', 0) or 0)
            example_yearly = float(example.get('yearly_usage', 0) or 0)
            example_lead_days = float(example.get('lead_days', 0) or 0)
            example_moq = float(example.get('moq_val', 0) or 0)
            example_spq = float(example.get('spq_val', 0) or 0)
            example_ss = float(example.get('safety_stock', 0) or 0)
            example_rop = float(example.get('rop', 0) or 0)
            example_current_inventory = float(example.get('current_inventory', 0) or 0)
            example_open_po = float(example.get('open_po', 0) or 0)
            example_untransfer_pr = float(example.get('untransfer_pr', 0) or 0)
            example_inventory_position = float(example.get('inventory_position', 0) or 0)
            example_purchase_trigger = str(example.get('purchase_trigger', '') or '')
            example_theoretical_purchase_qty = example.get('theoretical_purchase_qty', '')
            example_suggested_purchase_qty = example.get('suggested_purchase_qty', '')
            example_purchase_advice = str(example.get('purchase_advice', '') or '')
            current_k = {'A': k_a, 'B': k_b, 'C': k_c}.get(example_abc, k_b if k_b != '' else 0)
            lead_time_demand = example_avg_daily * example_lead_days
            default_a_k = 0.65
            lower_a_k_demo = 0.50
            default_a_ss = default_a_k * lead_time_demand
            lower_a_ss = lower_a_k_demo * lead_time_demand
            raw_ss = current_k * example_avg_daily * example_lead_days if isinstance(current_k, (int, float)) else 0
            ss_after_moq = max(raw_ss, example_moq) if example_moq > 0 else raw_ss
            ss_after_spq = ss_after_moq
            if example_spq > 0:
                ss_after_spq = math.ceil(ss_after_moq / example_spq) * example_spq
            raw_rop = lead_time_demand + raw_ss
            rop_after_spq = raw_rop

            explain_ws.merge_cells('A1:G1')
            explain_ws['A1'] = 'ABC分类说明（零基础版）'
            explain_ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
            explain_ws['A1'].fill = title_fill
            explain_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            explain_ws.merge_cells('A2:G2')
            explain_ws['A2'] = '一句话先讲明白：这张表不是按金额分 ABC，而是按“交期风险”分 ABC，再告诉你每颗外购物料建议备多少库存、库存掉到哪里就该下单，目的是防止停线。'
            explain_ws['A2'].font = Font(bold=True, color='000000', size=10)
            explain_ws['A2'].fill = PatternFill('solid', fgColor='E8F4FD')
            explain_ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')

            explain_rows = self._build_abc_explain_rows(
                example_code=example_code,
                example_name=example_name,
                example_abc=example_abc,
                example_total_qty=example_total_qty,
                example_monthly=example_monthly,
                example_yearly=example_yearly,
                example_lead_days=example_lead_days,
                example_moq=example_moq,
                example_spq=example_spq,
                example_ss=example_ss,
                example_rop=example_rop,
                example_current_inventory=example_current_inventory,
                example_open_po=example_open_po,
                example_untransfer_pr=example_untransfer_pr,
                example_inventory_position=example_inventory_position,
                example_purchase_trigger=example_purchase_trigger,
                example_theoretical_purchase_qty=example_theoretical_purchase_qty,
                example_suggested_purchase_qty=example_suggested_purchase_qty,
                example_purchase_advice=example_purchase_advice,
                example_avg_daily=example_avg_daily,
                example_ss_coverage_days=example.get('ss_coverage_days', ''),
                example_ss_judge=example.get('ss_judge', ''),
                missing_example=missing_example,
                plan_days=plan_days,
                daily_factor=daily_factor,
                lead_a=lead_a,
                lead_b=lead_b,
                k_a=k_a,
                k_b=k_b,
                k_c=k_c,
                current_k=current_k,
                lead_time_demand=lead_time_demand,
                default_a_ss=default_a_ss,
                lower_a_ss=lower_a_ss,
                raw_ss=raw_ss,
                ss_after_moq=ss_after_moq,
                ss_after_spq=ss_after_spq,
                raw_rop=raw_rop,
                rop_after_spq=rop_after_spq,
                mrp_path=mrp_path,
                excluded_codes_count=excluded_codes_count,
                usage_exclusion_pairs_count=usage_exclusion_pairs_count,
                production_parent_count=production_parent_count,
                material_category_rule_count=material_category_rule_count,
                manual_clear_tail_rule_count=manual_clear_tail_rule_count,
                auto_clear_tail_rule_count=auto_clear_tail_rule_count,
                model_switch_rule_count=model_switch_rule_count,
                model_switch_applied_count=model_switch_applied_count,
                model_switch_clear_tail_material_count=model_switch_clear_tail_material_count,
                model_switch_missing_old_count=model_switch_missing_old_count,
                model_switch_missing_new_count=model_switch_missing_new_count,
                custom_stock_rule_count=custom_stock_rule_count,
                custom_stock_controlled_mother_count=custom_stock_controlled_mother_count,
                custom_stock_original_qty=custom_stock_original_qty,
                custom_stock_effective_qty=custom_stock_effective_qty,
                custom_stock_trimmed_qty=custom_stock_trimmed_qty,
                subplan_dedupe_enabled=subplan_dedupe_enabled,
                subplan_dedupe_original_rows=subplan_dedupe_original_rows,
                subplan_dedupe_effective_rows=subplan_dedupe_effective_rows,
                subplan_dedupe_covered_rows=subplan_dedupe_covered_rows,
                subplan_dedupe_partial_rows=subplan_dedupe_partial_rows,
                subplan_dedupe_covered_qty=subplan_dedupe_covered_qty,
                subplan_dedupe_affected_code_count=subplan_dedupe_affected_code_count,
                replacement_diff_rule_count=replacement_diff_rule_count,
                replacement_diff_applied_count=replacement_diff_applied_count,
                replacement_diff_positive_material_count=replacement_diff_positive_material_count,
                replacement_diff_missing_base_count=replacement_diff_missing_base_count,
                replacement_diff_missing_replacement_count=replacement_diff_missing_replacement_count,
            )

            if False and missing_example:
                explain_rows.extend([
                    ['最后提醒', '', '', '', '', '', ''],
                    ['最后提醒', '这张表最怕什么错', '交期、MOQ、SPQ、供应商、K值、计划天数、日均系数任意一个错了，后面的建议都会跟着偏', '这张表不是魔法，它只是把你给的数据系统化算出来。', '先保证基础数据可靠，再相信计算结果。', example_code, '先检查主数据，再看结果异常'],
                ])

            for row_idx, row_values in enumerate(explain_rows, start=4):
                for col_idx, value in enumerate(row_values, start=1):
                    cell = explain_ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = bdr
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    if row_idx == 4:
                        cell.font = hdr_font
                        cell.fill = header_fill
                    elif row_values[1] == '' and row_values[2] == '' and row_values[3] == '':
                        cell.font = Font(bold=True)
                        cell.fill = section_fill
                    else:
                        cell.fill = soft_fill

            for col_letter, width in {'A': 12, 'B': 20, 'C': 34, 'D': 36, 'E': 34, 'F': 18, 'G': 48}.items():
                explain_ws.column_dimensions[col_letter].width = width
            explain_ws.freeze_panes = 'A5'

            wb.save(path)
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
            self.status_var.set(f'已导出: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))

    def _build_abc_explain_rows(
        self,
        *,
        example_code,
        example_name,
        example_abc,
        example_total_qty,
        example_monthly,
        example_yearly,
        example_lead_days,
        example_moq,
        example_spq,
        example_ss,
        example_rop,
        example_current_inventory,
        example_open_po,
        example_untransfer_pr,
        example_inventory_position,
        example_purchase_trigger,
        example_theoretical_purchase_qty,
        example_suggested_purchase_qty,
        example_purchase_advice,
        example_avg_daily,
        example_ss_coverage_days,
        example_ss_judge,
        missing_example,
        plan_days,
        daily_factor,
        lead_a,
        lead_b,
        k_a,
        k_b,
        k_c,
        current_k,
        lead_time_demand,
        default_a_ss,
        lower_a_ss,
        raw_ss,
        ss_after_moq,
        ss_after_spq,
        raw_rop,
        rop_after_spq,
        mrp_path,
        excluded_codes_count,
        usage_exclusion_pairs_count,
        production_parent_count,
        material_category_rule_count,
        manual_clear_tail_rule_count,
        auto_clear_tail_rule_count,
        model_switch_rule_count,
        model_switch_applied_count,
        model_switch_clear_tail_material_count,
        model_switch_missing_old_count,
        model_switch_missing_new_count,
        custom_stock_rule_count,
        custom_stock_controlled_mother_count,
        custom_stock_original_qty,
        custom_stock_effective_qty,
        custom_stock_trimmed_qty,
        subplan_dedupe_enabled,
        subplan_dedupe_original_rows,
        subplan_dedupe_effective_rows,
        subplan_dedupe_covered_rows,
        subplan_dedupe_partial_rows,
        subplan_dedupe_covered_qty,
        subplan_dedupe_affected_code_count,
        replacement_diff_rule_count,
        replacement_diff_applied_count,
        replacement_diff_positive_material_count,
        replacement_diff_missing_base_count,
        replacement_diff_missing_replacement_count,
    ):
        rows = [
            ['范围', '你看到什么', '系统怎么算', '用人话解释', '你应该怎么理解/怎么用', '示例料号', '示例说明'],
            ['先看结论', '', '', '', '', '', ''],
            ['先看结论', '这张表到底在干什么', '把外购物料按交期风险分成 A / B / C，并给出安全库存、ROP、库存位置和采购建议', '它不是在看“谁最贵”，而是在看“谁最怕断料、库存位置掉到哪里必须补货、这次建议买多少”', '先看 A 类占比、缺主数据物料、库存位置是否低于 ROP', '', ''],
            ['先看结论', 'A / B / C 是什么意思', 'A 类：交期 >= A 阈值；B 类：交期 >= B 阈值且 < A 阈值；C 类：交期 < B 阈值', '交期越长，补货越慢，越怕断料，所以越要重点管', 'A 类通常要更多保险库存；C 类相对宽松', example_code, f'示例交期 {fmt_qty(example_lead_days)} 天，阈值 A={lead_a} / B={lead_b}，归到 {example_abc} 类'],
            ['先看结论', '缺采购数据是什么意思', '外采清单里找不到该料时，分类记为 "-"，不计算正常 ABC / 安全库存 / ROP', '不是没风险，而是没资料，系统不能可靠地算', '先补供应商、交期、MOQ、SPQ，再看结果', str(missing_example.get('code', '') or '') if missing_example else '', '这类料先补主数据' if missing_example else '当前示例里没有缺主数据料号'],
            ['参数', '', '', '', '', '', ''],
            ['参数', '计划天数', '手工输入优先；否则按排产最早和最晚日期自动计算', '它决定平均日均用量的分母。计划天数越短，同样总需求摊到每天就越大', '计划天数变短，库存建议会整体变大', '', f'本次计划天数 = {plan_days}'],
            ['参数', '日均系数', '直接乘到平均日均用量上', '这是人为调节阀：觉得未来更紧张就调大，想保守一点就调小', '系数 > 1 会放大建议库存；系数 < 1 会缩小建议库存', '', f'本次日均系数 = {daily_factor}'],
            ['参数', 'K 值', 'A/B/C 三类分别使用各自 K 值', 'K 可以理解成“愿意额外准备多少保险库存”', 'K 越大，安全库存和 ROP 越高，资金占用更大，但断料风险更低', '', f'本次 K 值：A={k_a}，B={k_b}，C={k_c}'],
            ['参数', 'MRP 供给源', '从 MRP 计算表读取当前库存、未清PO、未转PR', '这部分不是理论值，而是你当前已经拿在手里或已经在供应管道里的量', '不导入 MRP，ABC 只能给出警戒线，不能判断当前是否该补货', '', os.path.basename(mrp_path) if mrp_path else '未导入 MRP 计算表'],
            ['参数', '排除规则', '物料编码排除清单只剔除ABC输出行；母料用量排除清单会剔除指定母子用量关系；排产计划母件默认只作为展开起点，不进ABC输出', '这样可以避免半成品母件或指定排除料混进外购物料ABC，同时仍然保留未被排除的下层真实外购需求', '如果想屏蔽某个母件对子件的用量，请用“母料用量排除清单”；如果只是不想在ABC结果里看到某个料号，请放“物料编码排除清单”', '', f'本次读取：物料排除{excluded_codes_count}项；母料用量排除{usage_exclusion_pairs_count}条；排产母件{production_parent_count}项'],
            ['参数', '产品替换差异采购清单', '当排产母件在清单B列时，系统先展开B列新母件和A列基准母件，再只保留“新母件单机用量 - 基准母件单机用量”大于0的下层物料', '这就是“A1只买跟A不同且多出来的物料”。共用且用量相同的料不重复买；A1用量更少的料也不买；A1新增或用量更多的料才进ABC需求', '清单建议A列填基准/原母件，B列填替换/新母件。只在ABC计算里生效，不改变BOM、不改变平衡表', '', f'本次读取替换规则{replacement_diff_rule_count}条；应用{replacement_diff_applied_count}次；计入差异物料{replacement_diff_positive_material_count}项；缺基准BOM{replacement_diff_missing_base_count}项；缺替换BOM{replacement_diff_missing_replacement_count}项'],
            ['参数', '物料分类清单', '从MRP计算表的“物料分类清单”按物料编码前缀匹配，前缀越长优先级越高', '比如 1401=电机类，料号以1401开头就显示电机类；如果同时有140101=伺服电机，会优先用更精确的140101', '这个分类只是帮助你看品类，不影响A/B/C交期风险分类、安全库存和ROP计算', example_code, f'本次读取物料分类规则 {material_category_rule_count} 条；示例分类={str(next((item for item in self._abc_result if item.get("code") == example_code), {}).get("material_category", "") or "未匹配")}'],
            ['参数', '切换清尾物料清单', 'MRP计算表里新增“切换清尾物料清单”，填物料编码或前缀；命中后不再补ROP，只按排产净需求下单', '这是手工指定清尾：未来不用或少用的专用料，不应该为了安全库存继续备货', 'T+3按排产需求净缺口算；月均滚动也优先按排产月份判断，排产计划之外不再按月均无限滚动', '', f'本次手工清尾规则{manual_clear_tail_rule_count}条；字段建议：物料编码、清尾截止月份、备注'],
            ['参数', '型号切换计划', 'MRP计算表里新增“型号切换计划”，字段为原母料号、原型号、新母料号、新型号、切换日期；系统只用原母料号/新母料号/切换日期计算，原型号和新型号只是给人看的识别说明', '系统展开两个母件BOM，自动识别“旧有新无”或“新用量更少”的下层物料。这相当于自动识别清尾料：A切B后，A来源的专用料或降用量料不再按ROP备安全库存，只按实际排产缺口采购', '现在按母料来源拆分计算：同一物料如果一部分来自切换母料、一部分来自其他共用母料，会拆成不同来源行；切换来源走清尾，其他来源继续走正常ROP。若排产里单独排了旧母料号下层半成品，该下层来源也会继承上层切换清尾规则，避免本体已切换但手腕/关节来源仍按ROP备货', '', f'本次读取型号切换{model_switch_rule_count}条；应用{model_switch_applied_count}次；自动清尾料{auto_clear_tail_rule_count}项；缺原BOM{model_switch_missing_old_count}项；缺新BOM{model_switch_missing_new_count}项'],
            ['参数', '客户定制机型备货控制', 'MRP计算表里新增“客户定制机型备货控制”，字段为料号、型号、数量、客户；命中排产母料号后，会把该母料号参与采购计算的有效数量限制到控制数量以内', '这不是新增需求，而是采购计划修正：客户定制机型只买够指定数量对应的下层物料，库存+未清PO+未转PR已覆盖就不再下单；命中的定制来源不再补ROP', '现在按母料来源拆分计算：同一物料如果既用于定制机型又用于普通机型，定制来源只买够控制数量，普通来源仍按ROP备货，不会被定制控制压掉', '', f'本次读取定制备货规则{custom_stock_rule_count}条；有效母件{custom_stock_controlled_mother_count}项；原排产{fmt_qty(custom_stock_original_qty)}，计入{fmt_qty(custom_stock_effective_qty)}，裁剪{fmt_qty(custom_stock_trimmed_qty)}'],
            ['参数', '定制机型差异物料', '导出Excel里新增“定制机型差异物料”sheet；系统按单台BOM用量，把定制母料号和当前其他排产母料号逐项对比', '它回答“定制机型相对其他当前排产项目，多了哪些料、少了哪些料、哪些是定制专用料”。差异数量=单台差异×定制有效数量，只作为识别差异用，不反向修改ABC采购计算', '用于判断定制机型到底改动了哪些下层外购物料；采购计划仍以ABC主计算结果为准', '', f'本次输出定制差异行{len(getattr(self, "_abc_custom_diff_rows", []) or [])}行'],
            ['参数', '排产母件去重', '如果排产计划里同时有整机和它下面的关节/模组，系统按月份先用上层BOM展开量抵扣下层排产量，只展开下层未被覆盖的剩余数量', '避免“整机展开一次、关节又单独展开一次”造成下层外购物料重复需求；如果下层排产量比上层覆盖量多，超出的部分仍然保留', '建议默认开启。只有当下层排产确实是独立额外需求、不能被上层整机覆盖时，才临时关闭', '', f'本次状态：{"开启" if subplan_dedupe_enabled else "关闭"}；排产行{subplan_dedupe_original_rows}->{subplan_dedupe_effective_rows}；全覆盖行{subplan_dedupe_covered_rows}；部分抵扣行{subplan_dedupe_partial_rows}；影响料号{subplan_dedupe_affected_code_count}项；抵扣数量{fmt_qty(subplan_dedupe_covered_qty)}'],
            ['参数', '为什么 A 类 K 常用 0.65', '原始安全库存 = 0.65 × 交期需求', 'A 类料交期长、补货慢、最怕断料。0.65 的意思是：在正常交期需求之外，再多准备约 65% 的保险垫', '它不是绝对标准，但常作为“资金占用”和“停线风险”之间的折中起点', example_code, f'如果交期需求 = {fmt_qty(lead_time_demand)}，按 0.65 算，原始安全库存约 = {fmt_qty(default_a_ss)}'],
            ['参数', '把 A 类 K 调低会怎样', 'K 降低 -> 安全库存降低 -> ROP 也降低', '好处是少备货、少占库存；坏处是缓冲变薄，更容易被延期、品质异常、需求放大击穿', '本质上是在调“库存占用”和“断料风险”的平衡，不是越低越好', example_code, f'如果 A 类 K 从 0.65 降到 0.50，原始安全库存会从 {fmt_qty(default_a_ss)} 降到 {fmt_qty(lower_a_ss)}'],
            ['字段说明', '', '', '', '', '', ''],
            ['字段说明', '物料编码 / 物料名称 / 规格 / 供应商', '直接取外采清单或 BOM 展开结果', '这是“这颗料是谁”的身份信息', '先确认是不是对的料，再看后面的库存建议', example_code, f'{example_code} / {example_name}'],
            ['字段说明', '交期(天)', '直接取外采清单中的交期', '从下单到拿到货，大概要等多少天', '这是 ABC 分类和库存策略最核心的输入之一', example_code, f'交期 = {fmt_qty(example_lead_days)} 天'],
            ['字段说明', 'SPQ', '直接取外采清单中的包装倍数', '供应商不一定让你想买多少就买多少，可能必须按整包整箱整卷买', '系统会把建议数量往上凑到 SPQ 的倍数', example_code, f'SPQ = {fmt_qty(example_spq)}'],
            ['字段说明', 'MOQ', '直接取外采清单中的最小起订量', '就算公式算出来只要 12，也可能因为 MOQ=50，最后至少得买 50', '看到建议值被抬高，先看是不是被 MOQ 顶上去了', example_code, f'MOQ = {fmt_qty(example_moq)}'],
            ['字段说明', '总需求（说明用）', '按排产递归展开 BOM 后汇总到子件', '表示这次排产周期内，这颗料一共预计要用多少', '这是后面所有日耗和库存建议的起点', example_code, f'总需求 = {fmt_qty(example_total_qty)}'],
            ['字段说明', '月均用量', '(总需求 / 计划天数 × 日均系数) × 30', '先把需求摊到每天，再换算成 30 天大概会用多少', '月均用量越大，说明这颗料消耗越快', example_code, f'({fmt_qty(example_total_qty)} / {plan_days} × {daily_factor}) × 30 = {fmt_qty(example_monthly)}'],
            ['字段说明', '年用量', '月均用量 × 12', '把月均放大到一年，用来判断长期量级', '年用量越大，说明长期消耗越大', example_code, f'{fmt_qty(example_monthly)} × 12 = {fmt_qty(example_yearly)}'],
            ['字段说明', '安全库存', '原始安全库存 = K × 平均日均用量 × 交期；然后再经过 MOQ 下限和 SPQ 向上取整', '它不是系统随便给的，而是“为了防断料，理论上至少要多准备多少保险量”，再套供应商规则', '如果原始值和最终值不一样，优先看是不是被 MOQ 或 SPQ 改大了', example_code, f'原始值 = {fmt_qty(current_k)} × {fmt_qty(example_avg_daily)} × {fmt_qty(example_lead_days)} = {fmt_qty(raw_ss)}；MOQ后 {fmt_qty(ss_after_moq)}；SPQ后 {fmt_qty(ss_after_spq)}；最终 {fmt_qty(example_ss)}'],
            ['字段说明', '交期需求', '平均日均用量 × 交期', '它不是保险量，而是“从现在下单到新货到厂前，正常生产本来就会吃掉多少”', '先算出这段正常消耗，再加保险库存，才能得到补货警戒线', example_code, f'{fmt_qty(example_avg_daily)} × {fmt_qty(example_lead_days)} = {fmt_qty(lead_time_demand)}'],
            ['字段说明', '当前库存', '来自 MRP 计算表中的期初库存', '这是真正在仓里、现在能直接拿来用的现货', '它只是供给的一部分，不是全部', example_code, fmt_qty(example_current_inventory)],
            ['字段说明', '未清PO', '来自 MRP 计算表中的在途采购欠交数量', '这是已经下单、但还没到厂的量', '它的确定性通常高于未转PR', example_code, fmt_qty(example_open_po)],
            ['字段说明', '未转PR', '来自 MRP 计算表中的在途请购未转数量', '这是已经提出采购需求，但还没正式变成 PO 的量', '它也算未来供给，但确定性低于未清PO', example_code, fmt_qty(example_untransfer_pr)],
            ['字段说明', '库存位置', '库存位置 = 当前库存 + 未清PO + 未转PR', '这才是要拿来和 ROP 比的数，不是只看仓里的现货', '如果库存位置还在 ROP 之上，暂时还有缓冲；如果已经到 ROP 或以下，就该补货', example_code, f'{fmt_qty(example_current_inventory)} + {fmt_qty(example_open_po)} + {fmt_qty(example_untransfer_pr)} = {fmt_qty(example_inventory_position)}'],
            ['字段说明', '再订货点(ROP)', 'ROP = 交期需求 + 原始安全库存（不含MOQ/SPQ）', 'ROP 不是库存本身，也不是下单数量，而是“什么时候该触发补货”的警戒线。起订量和包装倍数会影响一次买多少，不应该把触发线抬高', '真正要比的是“库存位置 <= ROP”是否成立。低于这条线就该触发补货；具体下单数量再按 MOQ/SPQ 另算', example_code, f'交期需求 {fmt_qty(lead_time_demand)} + 原始安全库存 {fmt_qty(raw_ss)} = {fmt_qty(raw_rop)}；最终 ROP = {fmt_qty(example_rop)}'],
            ['字段说明', '是否触发采购', '库存位置 <= ROP 时为“是”；库存位置 > ROP 时为“否”', '它回答“现在要不要买”。库存位置已经低到警戒线以下，就触发；还在警戒线以上，就先不买', '如果显示“缺MRP”，说明没有当前库存/未清PO/未转PR，系统无法判断本次采购量', example_code, f'本例触发状态：{example_purchase_trigger}'],
            ['字段说明', '理论采购量', '触发采购时：ROP - 库存位置；不考虑MOQ/SPQ', '它回答“补回到警戒线还差多少”。这个数最保守，不再额外加库存缓冲', '用它看真实补货缺口；不要直接拿它当采购订单数量，因为供应商规则还会改大', example_code, f'理论采购量 = {fmt_qty(example_theoretical_purchase_qty)}'],
            ['字段说明', '建议采购量', '先取理论采购量，再按MOQ下限和SPQ倍数向上修正', '它回答“按供应商规则这次实际建议买多少”。所以它可能比理论缺口大', '采购下单优先看这个数；如果太大，再人工判断是否拆单或调整策略', example_code, f'建议采购量 = {fmt_qty(example_suggested_purchase_qty)}；{example_purchase_advice}'],
            ['T+3采购计划', '', '', '', '', '', ''],
            ['T+3采购计划', '这张sheet做什么', '只按排产计划购买：排产需求扣完库存位置后出现净缺口才触发采购，不再因为ROP安全库存单独下单', 'ABC主表仍保留安全库存/ROP判断；T+3只回答“为了满足当前排产计划必须买多少”', '理论采购量只补排产净缺口；建议采购量仍按MOQ/SPQ修正', '', '导出Excel时自动生成'],
            ['T+3采购计划', '需求月份', '来自排产计划上线日期对应的月份；如果多个需求月落在交期覆盖期内，会合并显示', '表示这行采购建议涉及哪些月份的排产需求。注意：覆盖期内需求是展示口径，不等于本次采购量', '同一颗料同一建议下单月份会合并成一行，避免同月重复套MOQ', '', '例如：2026年5月、2026年6月、2026年7月'],
            ['T+3采购计划', '建议下单月份 / 建议下单日期', '建议下单日期 = 需求月份第一天 - 交期；如果倒推出的日期早于今天，就推到当前月份分批日期', '它回答“最晚什么时候要下单”。未来日期按界面设置的每月下单日排；过去已经错过的日期，仍按原方案从今天开始按14天间隔分批', '这样既避免所有单挤在同一天，也不会把未来订单排到最晚下单日之后', '', '例如今天是2026-04-28，则过期订单会排到2026-04-28或2026-05-12'],
            ['T+3采购计划', '月需求', '该行涉及月份的需求数量合计，向上取整', '它表示需求展示合计，不是月均用量，也不是采购量。首单行可能把交期覆盖期内多个月份合并显示', '采购量看理论采购量/建议采购量，不要把月需求当成下单量', '', '3965月均用量 × 3个月 = 11895，只表示三个月需求合计'],
            ['T+3采购计划', '月初库存位置 / 月末预计库存位置', '首单行展示覆盖期内排产需求扣减后的参考位置；后续月份按滚动位置继续计算', '它帮助解释排产计划是否已经被当前库存位置覆盖', '后续月份只有滚动扣减后库存位置小于0，才生成采购建议；ROP只作参考', '', '排产需求已覆盖时，不因为低于ROP而在T+3里额外买'],
            ['T+3采购计划', '月度理论采购量', '只按排产净缺口计算：排产需求扣完库存位置后的负数缺口；不补ROP', '它回答为了满足排产计划真实还缺多少。不考虑MOQ/SPQ，所以更接近计划缺口', '实际下单看建议采购量，因为供应商MOQ/SPQ可能把数量放大', '', '例如排产需求30、库存位置19，只补排产缺口11；即使ROP为12，也不额外补ROP'],
            ['T+3采购计划', '月度建议采购量', '先合并同物料同建议下单月份的理论采购量，再按MOQ/SPQ修正', '这样可以避免同月多行分别套MOQ，造成采购量被重复放大', '实际下单优先看这个字段', '', '例如两行理论量30和20，同月合并为50，再按MOQ/SPQ取整'],
            ['T+3采购计划', '合并规则', '同一个物料编码 + 同一个建议下单月份，会合并成一行；需求月份和建议下单日期合并显示，采购量合并计算', '如果多个需求月因为交期过期被推到当前月份，同一颗料会先按月合并，再显示该月内建议日期', '不同物料不会合并，避免料号混在一起；同料同月合并，避免重复套MOQ', '', '例如4/28和4/30都会显示，但采购量只按4月合并计算一次'],
            ['T+3下单汇总', '', '', '', '', '', ''],
            ['T+3下单汇总', '这张sheet做什么', '把T+3采购计划里触发采购的行，只按“建议下单月份”汇总', '它只回答每个月应该下单多少，不再推算预计到货', '先看上方汇总，再看下方明细追到具体料号', '', '导出Excel时自动生成'],
            ['T+3下单汇总', '理论下单数量 / 建议下单数量', '理论下单数量不按MOQ/SPQ；建议下单数量按MOQ/SPQ向上修正', '理论量回答“真实缺口是多少”；建议量回答“按供应商规则实际可能要下多少”', '如果建议量远大于理论量，优先检查MOQ/SPQ，必要时和供应商协商拆单或改包装', '', '4月下单数量会同时显示两套口径'],
            ['月均用量滚动采购计划', '', '', '', '', '', ''],
            ['月均用量滚动采购计划', '这张sheet做什么', '逐月滚动时同时看月均用量和该月排产需求；哪个更大就用哪个扣减库存位置；排产净缺口或低于ROP都会触发采购', '这样既能做常规备货节奏，也不会漏掉某个月排产集中导致的真实缺料', '适合看未来几个月的采购节奏；如果某月排产需求高于月均，会在说明里写明按排产需求判断', '', '滚动月份数由界面“滚动采购月份数”控制'],
            ['月均用量滚动采购计划', '月均用量/本月折算用量', '未来完整月份使用月均用量；当前月份只按今天到月底的剩余天数折算', '这样不会在月末还把已经过去的整个月消耗再算一遍', '例如4月28日导出，4月只按4/28-4/30剩余天数折算，5月以后按整月月均用量', '', '月均用量 = (总需求 / 计划天数 × 日均系数) × 30'],
            ['月均用量滚动采购计划', '下单节奏', '先估算库存位置在当月哪天跌到ROP，再把下单日期安排到不晚于该日期的界面设定下单日；已过期的单从今天开始每14天分批', '简单说，就是未来订单按你设定的每月几号下单，不把采购单挤在一天，也不故意排晚', '如果你想更平滑资金占用，看这张表比只看ABC主表更合适', '', '例如设置1号、15号，触发日在20号附近，通常排15号下单'],
            ['月均滚动下单汇总', '这张sheet做什么', '把月均用量滚动采购计划里触发采购的行，只按下单月份汇总，并同时显示理论量和建议量', '理论量是不按MOQ/SPQ的资金压力参考；建议量是按供应商规则后的下单参考', '不再推算预计到货，避免已有未清PO/未转PR没有到货日期导致误解', '', '适合给采购看月度下单节奏'],
            ['采购后可生产台数', '这张sheet做什么', '把当前库存、未清PO、未转PR和对应采购计划的建议采购量合在一起，再按排产顺序逐行扣减BOM外购物料', '它回答“如果按这套建议采购量去买，每个排产母料号最多能生产多少台，哪里卡住”。共用物料只会被前面的排产消耗一次，不会重复给多个机型使用', '分别输出“T+3采购后可生产”和“滚动采购后可生产”，两张表口径不同，因为采购建议数量不同', '', '看“状态/短板物料/未满足数量”即可判断哪一个机型还不够料'],
            ['采购后可生产台数', '可生产上限怎么来', '对每个排产母料号展开单台BOM；每颗下层外购物料可支持台数=当前可用数量÷单台用量；整机可生产上限取所有下层物料里的最小值', '一台机器缺一颗关键料都不能完整生产，所以取最短板物料作为上限。算完本行后，会把已生产数量消耗掉，再算下一行', '如果显示“未识别下层”，说明该母料号没有展开到ABC外购物料，需检查BOM、外采清单或排除规则', example_code, '例：某机型需要A料2个、B料1个；A可用20、B可用6，则最多只能做6台'],
            ['应到货数量清单', '这张sheet做什么', '按生产计划逐月扣减当前库存，先显示排产净缺口，再用未清PO、未转PR覆盖，剩余净需求归入购买建议应到货，并保留覆盖后缺口字段', '它回答“为了满足生产计划，每个月至少应该到多少货、原始缺口是多少、哪些由未清/未转/购买建议覆盖、覆盖后还缺不缺”', '这张表只算数量，不计算金额，也不额外补ROP；它不引用T+3购买建议池，避免5月被交期覆盖期粗略判断吞掉', '', '顺序：月需求 - 当前库存 = 排产净缺口；排产净缺口 -> 未清PO -> 未转PR -> 购买建议应到货 -> 采购建议覆盖后缺口'],
            ['月度生产风险清单', '', '', '', '', '', ''],
            ['月度生产风险清单', '这张sheet做什么', '按排产月份逐月判断生产风险，把当前库存、有日期的未清PO、建议采购预计到货一起考虑', '它回答“某个月能不能生产、是不是需要采购跟催”，不是回答长期备货量', '生产和采购可以先筛选RYG，再看风险类别和需采购跟催', '', '导出Excel时自动生成'],
            ['月度生产风险清单', '供给口径', '当前库存立即可用；未清PO只有有预计/计划/承诺到货日期才按日期参与覆盖；日期未知未清PO只展示；未转PR只展示不参与；建议采购按建议下单日期+交期推算预计到货', '这样不会把所有未清PO/未转PR都默认当成5月1日可用，风险判断更接近真实生产', '如果在途采购没有到货日期，风险会更保守；需要补齐未清PO到货日期才能降低误判', '', '字段：本月未清PO到货、日期未知未清PO、未转PR(不参与)'],
            ['月度生产风险清单', 'RYG是什么意思', 'R=生产缺口/采购后仍缺口/交期不满足需跟催/缺主数据；Y=能覆盖但依赖按期到货或低于ROP；G=覆盖生产且相对健康', 'R最高风险，Y中风险，G低风险', '先处理R，尤其是“交期不满足需跟催”；再看Y里的低于ROP和依赖到货', '', '排序优先级：R > Y > G'],
            ['月度生产风险清单', '交期不满足需跟催', '最晚到货日期按需求月份第一天；预计到货日期按建议下单日期+交期；预计到货晚于最晚到货日期时标R并显示需采购跟催=是', '数量够不代表能赶上生产。标准交期赶不上时，需要采购协调提前交付或加急', '这类不是数量缺口，而是交付时间风险，应直接给采购跟催', '', '字段：最晚到货日期、预计到货日期、交期是否满足、需采购跟催'],
            ['生产风险汇总', '', '', '', '', '', ''],
            ['生产风险汇总', '这张sheet做什么', '按月份汇总R/Y/G物料数、风险料号数、需跟催料号数，并给出当月最高风险物料分类和主要风险原因', '这是管理视角，快速回答“5月哪个类别风险最高，为什么”', '先看最高风险物料分类，再进TOP风险分类看具体分类原因和代表料号', '', '不再汇总R缺口合计/R需求合计/Y需求合计/G需求合计'],
            ['TOP风险分类', '为什么按分类而不是按缺口数量', '不同物料数量不可直接比较，1000颗螺丝和10台电机不能用数量大小判断风险高低', '所以TOP风险按风险料号数、红色料号数、需跟催料号数和主要原因排序', '这张表用于管理判断，不用于直接下采购订单', '', '排序逻辑：R物料数 > 需跟催料号数 > Y物料数 > 风险料号数'],
            ['TOP风险分类', '怎么看原因', '主要风险原因统计该分类下风险类型，例如交期不满足需跟催、生产缺口、采购到货后仍缺口、低于ROP缓冲、缺采购主数据', '它告诉你这个类别到底是数量不够、交期赶不上，还是主数据缺失', '原因不同动作不同：交期问题找采购跟催；数量问题看采购计划；主数据问题先补资料', '', '字段：主要风险原因、代表料号、涉及项目、管理建议'],
            ['5月重点风险前20', '这张sheet做什么', '只展示指定物料分类的5月份R/Y风险物料，每个分类按交期从长到短取前20条', '它用于给采购/管理层快速看5月重点分类里最该优先盯的长交期风险物料', '影响项目使用ABC分类里的项目字段，不再用产品型号；多个项目在单元格内换行显示，便于复制和阅读', '', '分类：钣金件-研发、电子板卡类、电机类、电子类、机加件-研发、减速机类、驱动类、轴承类、线缆类、铸件-研发'],
            ['字段说明', '安全库存覆盖天数', '安全库存 / 平均日均用量', '比如 59.2 天，只是在说“如果只动用安全库存这部分，理论上能扛 59.2 天”', '59.2 天之后怎么办，要看正常交期需求是否被当前库存、未清PO、未转PR接上。安全库存负责扛异常，正常消耗要靠供应管道续上', example_code, f'{fmt_qty(example_ss)} / {fmt_qty(example_avg_daily)} = {example_ss_coverage_days} 天'],
            ['字段说明', '安全库存判断', '用当前建议安全库存去对比交期需求，判断偏低 / 合理 / 偏高', '这是系统在帮你做一个快检', '大量偏低说明参数或主数据可能过于激进；大量偏高说明库存可能压得太多', example_code, str(example_ss_judge)],
        ]
        if missing_example:
            rows.extend([
                ['最后提醒', '', '', '', '', '', ''],
                ['最后提醒', '这张表最怕什么错', '交期、MOQ、SPQ、供应商、K值、计划天数、日均系数、MRP供给数据，任何一个错了，后面的建议都会跟着错', '这张表不是魔法，它只是把你给的数据系统化算出来', '先保证基础数据可靠，再相信计算结果', example_code, '先检查主数据，再看异常结果'],
            ])
        return rows

    def _refresh_calc_hint(self):
        """根据当前状态（是否加载 BOM / 是否已有计算结果）显示对应引导文案。"""
        if not hasattr(self, '_calc_hint_var'):
            return
        if not self.bom_index:
            self._calc_hint_var.set('⚠ 还没有 BOM 数据。请通过"菜单 → 文件 → 上传 BOM"或顶部按钮先导入 BOM 文件。')
            self._calc_hint_label.configure(foreground='#B91C1C')
        elif self._calc_result is None and not getattr(self, '_batch_result', None):
            self._calc_hint_var.set('提示：批量汇总可直接粘贴 "母料号+数量" 两列；单件计算请先到 "BOM 结构展开" 页查询一次。')
            self._calc_hint_label.configure(foreground='#64748B')
        else:
            self._calc_hint_var.set('已加载数据。可修改数量重新计算、导出 Excel，或返回 "BOM 结构展开" 页切换母料号。')
            self._calc_hint_label.configure(foreground='#0F766E')

    def _update_c_label(self):
        try:
            b = self._lead_b.get()
        except tk.TclError:
            return
        self._lead_c_auto.set(f'{b}（即B类下限）')

    def _validate_lead_thresholds(self):
        """A 类交期阈值必须 ≥ B 类阈值，否则输入框变红并显示提示。"""
        try:
            a = self._lead_a.get()
            b = self._lead_b.get()
        except tk.TclError:
            # 用户还在输入中（空字符串），不提示错误
            if hasattr(self, '_lead_hint_var'):
                self._lead_hint_var.set('')
            return
        self._update_c_label()
        invalid_a = a < 0
        invalid_b = b < 0
        invalid_order = a < b

        if hasattr(self, '_lead_a_entry'):
            self._lead_a_entry.configure(style='Invalid.TEntry' if invalid_a or invalid_order else 'TEntry')
        if hasattr(self, '_lead_b_entry'):
            self._lead_b_entry.configure(style='Invalid.TEntry' if invalid_b or invalid_order else 'TEntry')

        if hasattr(self, '_lead_hint_var'):
            if invalid_a or invalid_b:
                self._lead_hint_var.set('⚠ 阈值不能为负数')
            elif invalid_order:
                self._lead_hint_var.set(f'⚠ A 类阈值 {a} < B 类阈值 {b}，请保证 A ≥ B')
            else:
                self._lead_hint_var.set('')

    def _clear_abc(self):
        self._abc_result = []
        self._abc_calc_meta = {}
        self._abc_custom_diff_rows = []
        self._abc_plan_expansion_rows = []
        self._abc_summary_var.set('')
        for item in self.abc_tree.get_children(''):
            self.abc_tree.delete(item)
        self.status_var.set('外采物料管理已清空')

    def _copy_abc_cell(self, event=None):
        region = self.abc_tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        col = self.abc_tree.identify_column(event.x)
        item_id = self.abc_tree.identify_row(event.y)
        if not item_id or not col:
            return
        vals = self.abc_tree.item(item_id, 'values')
        ci = int(col[1:]) - 1
        if ci < len(vals):
            val = str(vals[ci]).strip()
            if val:
                self.root.clipboard_clear()
                self.root.clipboard_append(val)

    def _copy_abc_row(self):
        sel = self.abc_tree.selection()
        if not sel:
            return
        row = self.abc_tree.item(sel[0], 'values')
        text = '\t'.join(str(v).strip() for v in row if str(v).strip())
        self.root.clipboard_clear()
        self.root.clipboard_append(text)

    def _copy_abc_pn(self):
        sel = self.abc_tree.selection()
        if sel:
            vals = self.abc_tree.item(sel[0], 'values')
            if len(vals) > 1:
                self.root.clipboard_clear()
                self.root.clipboard_append(str(vals[1]).strip())

    def _bom_lookup_from_abc(self):
        sel = self.abc_tree.selection()
        if not sel:
            return
        vals = self.abc_tree.item(sel[0], 'values')
        if len(vals) > 1:
            code = str(vals[1]).strip()
            self.pn_entry.delete(0, tk.END)
            self.pn_entry.insert(0, code)
            self._switch_to_tab(self.tab_tree)
            self._query()

    def _show_abc_menu(self, event=None):
        self._abc_ctx.tk_popup(event.x_root, event.y_root)

    # ── 选项卡2：数量计算（读取展开树）───────────────────────
    def _build_calc_tab(self):
        f = self.tab_calc
        f.columnconfigure(0, weight=1)
        f.rowconfigure(2, weight=1)

        bf = ttk.LabelFrame(f, text='批量汇总输入区', style='Card.TLabelframe', padding=6)
        bf.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        bf.columnconfigure(0, weight=1)

        self._batch_text = tk.Text(
            bf,
            height=3,
            width=60,
            font=('Microsoft YaHei', 10),
            bg='#FFFFFF',
            fg='#1E293B',
            insertbackground='#1E293B',
            relief='flat',
            padx=12,
            pady=10
        )
        self._batch_text.grid(row=0, column=0, columnspan=4, sticky='ew', pady=(0,4))
        placeholder = ('支持直接从 Excel 复制两列（母料号 + 数量）粘贴\n'
                       '也可手输：每行一条，格式  母料号<Tab>数量  或  母料号,数量\n'
                       '（不填数量默认=1；含表头行会自动跳过）')
        self._batch_text.insert('1.0', placeholder)

        btn_row = ttk.Frame(bf)
        btn_row.grid(row=1, column=0, sticky='w', pady=(0,2))
        ttk.Button(btn_row, text='上传母料数量文件',
                   command=self._import_root_qtys_file).pack(side='left', padx=(0,6))
        ttk.Button(btn_row, text='批量汇总计算', style='Accent.TButton',
                   command=self._run_batch_calc).pack(side='left', padx=(0,6))
        ttk.Button(btn_row, text='导出汇总',
                   command=self._export_batch_calc).pack(side='left', padx=(0,6))
        ttk.Button(btn_row, text='清空批量', style='Quiet.TButton',
                   command=self._clear_batch_calc).pack(side='left')

        row = ttk.LabelFrame(f, text='计算控制与结果摘要', style='Card.TLabelframe', padding=6)
        row.grid(row=1, column=0, sticky='ew', padx=6, pady=(0, 3))
        row.columnconfigure(7, weight=1)

        self._calc_hint_var = tk.StringVar()
        self._calc_hint_label = ttk.Label(row, textvariable=self._calc_hint_var, style='Subtle.TLabel')
        self._calc_hint_label.grid(row=0, column=0, columnspan=8, sticky='w', pady=(0, 8))
        self._refresh_calc_hint()

        ttk.Label(row, text='顶层母件数量', style='Section.TLabel').grid(row=1, column=0, padx=(0, 6))
        self.calc_qty_var = tk.StringVar(value='1')
        ttk.Entry(row, textvariable=self.calc_qty_var, width=10).grid(
            row=1, column=1, sticky='w', padx=(0, 10))

        ttk.Button(row, text='重新计算', style='Accent.TButton', command=self._recalc_from_tree).grid(
            row=1, column=2, padx=(0, 6))
        ttk.Button(row, text='导出Excel', command=self._export_calc).grid(
            row=1, column=3, padx=(0, 6))
        ttk.Button(row, text='清空', style='Quiet.TButton', command=self._clear_calc).grid(
            row=1, column=4)
        ttk.Label(row, textvariable=self.calc_summary_var, foreground='#2563EB',
                  font=('Microsoft YaHei', 10, 'bold')).grid(
            row=1, column=5, columnspan=3, sticky='w', padx=(16, 0))

        res_frame = ttk.LabelFrame(f, text='数量计算结果', style='Card.TLabelframe', padding=6)
        res_frame.grid(row=2, column=0, sticky='nsew', padx=6, pady=(0, 6))
        res_frame.columnconfigure(0, weight=1)
        res_frame.rowconfigure(0, weight=1)

        self.calc_tree = ttk.Treeview(res_frame, columns=HEADERS_CALC,
                                      show='tree headings', selectmode='extended')
        for h, w in zip(HEADERS_CALC, COL_WIDTHS_C):
            self.calc_tree.column(h, width=w, anchor='w')
            self.calc_tree.heading(h, text=h)
        self.calc_tree.column('#0', width=0, stretch=False)
        enable_treeview_sort(self.calc_tree, HEADERS_CALC,
                             numeric_columns={'层级', '路径数', '汇总用量'})
        enable_treeview_copy(self.calc_tree, HEADERS_CALC)

        style = ttk.Style()
        style.configure('Treeview', rowheight=32)
        style.map('Treeview', background=[('selected', '#2563EB')])

        self.calc_tree.bind('<Double-Button-1>', self._copy_calc_cell)
        self.calc_tree.bind('<Button-3>', self._show_calc_menu)

        vsb2 = ttk.Scrollbar(res_frame, orient='vertical', command=self.calc_tree.yview)
        hsb2 = ttk.Scrollbar(res_frame, orient='horizontal', command=self.calc_tree.xview)
        self.calc_tree.configure(yscrollcommand=vsb2.set, xscrollcommand=hsb2.set)
        self.calc_tree.grid(row=0, column=0, sticky='nsew')
        vsb2.grid(row=0, column=1, sticky='ns')
        hsb2.grid(row=1, column=0, sticky='ew')

        self._calc_ctx = tk.Menu(self.calc_tree, tearoff=0)
        self._calc_ctx.add_command(label='复制整行', command=self._copy_calc_row)
        self._calc_ctx.add_command(label='复制料号', command=self._copy_calc_pn)
        self._calc_ctx.add_separator()
        self._calc_ctx.add_command(label='🔍 查询上一级物料用量', command=self._show_parent_usage)

    # ── 选项卡4：BOM差异对比 ─────────────────────────────────
    def _build_diff_tab(self):
        f = self.tab_diff
        f.columnconfigure(0, weight=1)
        f.rowconfigure(2, weight=1)

        row = ttk.LabelFrame(f, text='左右对照差异分析', style='Card.TLabelframe', padding=8)
        row.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        row.columnconfigure(1, weight=1)
        row.columnconfigure(3, weight=1)
        row.columnconfigure(8, weight=1)

        ttk.Label(row, text='物料编码 A', style='Section.TLabel').grid(row=0, column=0, padx=(0, 6), sticky='w')
        self.diff_a_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.diff_a_var).grid(row=0, column=1, sticky='ew', padx=(0, 10))

        ttk.Label(row, text='物料编码 B', style='Section.TLabel').grid(row=0, column=2, padx=(0, 6), sticky='w')
        self.diff_b_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.diff_b_var).grid(row=0, column=3, sticky='ew', padx=(0, 10))

        self.diff_run_btn = ttk.Button(row, text='开始AB对比', style='Accent.TButton', command=self._run_diff_compare)
        self.diff_run_btn.grid(row=0, column=4, padx=(0, 6))
        self.diff_export_btn = ttk.Button(row, text='导出Excel', command=self._export_diff)
        self.diff_export_btn.grid(row=0, column=5, padx=(0, 6))
        self.diff_clear_btn = ttk.Button(row, text='清空', style='Quiet.TButton', command=self._clear_diff)
        self.diff_clear_btn.grid(row=0, column=6)

        self.diff_summary_var = tk.StringVar(value='请输入两个母料号进行AB差异对比；多料号请使用下方“统一对比粘贴区”')
        ttk.Label(row, textvariable=self.diff_summary_var, style='Metric.TLabel').grid(
            row=1, column=0, columnspan=7, sticky='w', pady=(10, 4)
        )

        self.diff_progress = ttk.Progressbar(row, mode='determinate', length=180, maximum=100)
        self.diff_progress.grid(row=1, column=7, sticky='e', padx=(10, 0))

        batch = ttk.LabelFrame(f, text='统一对比粘贴区', style='Card.TLabelframe', padding=6)
        batch.grid(row=1, column=0, sticky='ew', padx=6, pady=(0, 3))
        batch.columnconfigure(1, weight=1)

        ttk.Label(batch, text='多料号', style='Section.TLabel').grid(row=0, column=0, sticky='nw', padx=(0, 6))
        self.diff_unified_text = tk.Text(batch, height=3, wrap='none', font=('Consolas', 9))
        self.diff_unified_text.grid(row=0, column=1, sticky='ew', padx=(0, 6))
        self.diff_unified_run_btn = ttk.Button(batch, text='开始统一对比', style='Accent.TButton', command=self._run_unified_diff_compare)
        self.diff_unified_run_btn.grid(
            row=0, column=2, sticky='n', padx=(0, 6)
        )
        ttk.Button(batch, text='清空粘贴', style='Quiet.TButton', command=lambda: self.diff_unified_text.delete('1.0', tk.END)).grid(
            row=0, column=3, sticky='n'
        )
        ttk.Label(
            batch,
            text='说明：直接粘贴多个母料号（可来自Excel单列/多列），系统会生成各母料号用量矩阵，并标出独有、部分共用、全部共用。',
            foreground='#64748B'
        ).grid(row=1, column=1, columnspan=3, sticky='w', pady=(3, 0))
        ttk.Label(batch, text='分类/前缀', style='Section.TLabel').grid(row=2, column=0, sticky='w', padx=(0, 6), pady=(5, 0))
        self.diff_prefix_var = tk.StringVar()
        ttk.Entry(batch, textvariable=self.diff_prefix_var, width=24).grid(row=2, column=1, sticky='w', pady=(5, 0))
        self.diff_prefix_run_btn = ttk.Button(batch, text='按前缀统一对比', command=self._run_prefix_diff_compare)
        self.diff_prefix_run_btn.grid(row=2, column=2, sticky='w', padx=(0, 6), pady=(5, 0))
        ttk.Label(batch, text='例如输入 040101，会自动对比所有 040101 开头的母料号。', foreground='#64748B').grid(
            row=2, column=3, sticky='w', pady=(5, 0)
        )
        ttk.Label(batch, text='外部BOM表', style='Section.TLabel').grid(row=3, column=0, sticky='w', padx=(0, 6), pady=(5, 0))
        self.diff_doc_compare_btn = ttk.Button(
            batch,
            text='上传量产BOM表对比',
            style='Accent.TButton',
            command=self._run_bom_doc_compare,
        )
        self.diff_doc_compare_btn.grid(row=3, column=1, sticky='w', pady=(5, 0))
        ttk.Label(
            batch,
            text='支持“左侧物料清单 + 右侧多机型数量矩阵”的量产BOM表，按04在制品号对比当前BOMMaster展开用量。',
            foreground='#64748B',
        ).grid(row=3, column=2, columnspan=2, sticky='w', pady=(5, 0))

        res_frame = ttk.LabelFrame(f, text='分列对比结果', style='Card.TLabelframe', padding=6)
        res_frame.grid(row=2, column=0, sticky='nsew', padx=6, pady=(0, 6))
        res_frame.columnconfigure(0, weight=1)
        res_frame.rowconfigure(0, weight=1)

        self.diff_tree = ttk.Treeview(res_frame, columns=HEADERS_DIFF_GRID, show='headings', selectmode='browse')
        self._configure_diff_tree(
            HEADERS_DIFF_GRID,
            COL_WIDTHS_D_GRID,
            numeric_columns={'A总用量', 'B总用量', '差值(A-B)'},
        )
        self.diff_tree.bind('<Double-Button-1>', self._copy_calc_cell)
        self.diff_tree.grid(row=0, column=0, sticky='nsew')

        diff_vsb = ttk.Scrollbar(res_frame, orient='vertical', command=self.diff_tree.yview)
        diff_hsb = ttk.Scrollbar(res_frame, orient='horizontal', command=self.diff_tree.xview)
        self.diff_tree.configure(yscrollcommand=diff_vsb.set, xscrollcommand=diff_hsb.set)
        diff_vsb.grid(row=0, column=1, sticky='ns')
        diff_hsb.grid(row=1, column=0, sticky='ew')

        self._diff_rows = []
        self._diff_export_rows = []
        self._diff_common_rows = []
        self._diff_input_codes = {'a': [], 'b': []}
        self._diff_unified_rows = []
        self._diff_code_label_map = {}
        self._diff_mode = 'pair'
        self._diff_unified_source = ''

    def _build_bom_clean_tab(self):
        f = self.tab_bom_clean
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='BOM 文件清洗参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(0, weight=3)
        control.columnconfigure(1, weight=2)

        self.clean_bom_var = tk.StringVar(value=self.current_file or '')
        self.clean_plan_var = tk.StringVar()
        self.clean_output_var = tk.StringVar()
        self.clean_summary_var = tk.StringVar(value='等待清洗')
        self.clean_autoload_var = tk.BooleanVar(value=True)
        self.clean_use_loaded_plan_var = tk.BooleanVar(value=True)

        form_panel = ttk.Frame(control)
        form_panel.grid(row=0, column=0, sticky='ew')
        form_panel.columnconfigure(1, weight=1)

        def _clean_field(row_index, label_text, variable, button_text, command):
            ttk.Label(form_panel, text=label_text, style='Section.TLabel').grid(
                row=row_index, column=0, sticky='w', padx=(0, 6), pady=3
            )
            label, _ = make_path_label(form_panel, variable)
            label.grid(row=row_index, column=1, sticky='ew', pady=3)
            ttk.Button(form_panel, text=button_text, command=command).grid(
                row=row_index, column=2, sticky='w', padx=(6, 0), pady=3
            )

        _clean_field(0, '原始 BOM 文件', self.clean_bom_var, '选择BOM', self._choose_clean_bom_file)
        _clean_field(1, '排产计划', self.clean_plan_var, '选择排产计划', self._choose_clean_plan_file)
        _clean_field(2, '输出文件', self.clean_output_var, '指定输出路径', self._choose_clean_output_file)

        side_panel = ttk.Frame(control)
        side_panel.grid(row=0, column=1, sticky='nsew', padx=(14, 0))
        side_panel.columnconfigure(0, weight=1)

        ttk.Checkbutton(
            side_panel,
            text='优先使用MRP计算已上传排产计划',
            variable=self.clean_use_loaded_plan_var,
        ).grid(row=0, column=0, sticky='w')
        ttk.Checkbutton(
            side_panel,
            text='生成后自动加载为当前BOM',
            variable=self.clean_autoload_var,
        ).grid(row=1, column=0, sticky='w', pady=(4, 0))
        ttk.Label(
            side_panel,
            text='如果MRP计算里已上传排产计划，这里可不重复选择；未上传时才使用左侧排产计划文件。清洗会递归保留下层BOM父项。',
            style='Subtle.TLabel',
            wraplength=360,
            justify='left',
        ).grid(row=2, column=0, sticky='w', pady=(6, 0))

        action_row = ttk.Frame(side_panel)
        action_row.grid(row=3, column=0, sticky='ew', pady=(8, 0))
        action_row.columnconfigure(2, weight=1)
        self.clean_run_btn = ttk.Button(action_row, text='开始清洗', style='Accent.TButton', command=self._run_bom_clean)
        self.clean_run_btn.grid(row=0, column=0, sticky='w')
        ttk.Button(action_row, text='清空结果', style='Quiet.TButton', command=self._clear_bom_clean_result).grid(
            row=0, column=1, sticky='w', padx=(8, 0)
        )
        ttk.Label(action_row, textvariable=self.clean_summary_var, style='Metric.TLabel', wraplength=240).grid(
            row=0, column=2, sticky='w', padx=(10, 0)
        )

        result = ttk.LabelFrame(f, text='清洗结果', style='Card.TLabelframe', padding=8)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(3, 6))
        result.columnconfigure(0, weight=1)
        result.rowconfigure(0, weight=1)

        columns = ['项目', '结果', '说明']
        widths = [180, 180, 760]
        self.clean_tree = ttk.Treeview(result, columns=columns, show='headings', selectmode='extended')
        for col, width in zip(columns, widths):
            self.clean_tree.heading(col, text=col)
            self.clean_tree.column(col, width=width, anchor='w')
        self.clean_tree.grid(row=0, column=0, sticky='nsew')
        clean_vsb = ttk.Scrollbar(result, orient='vertical', command=self.clean_tree.yview)
        clean_vsb.grid(row=0, column=1, sticky='ns')
        self.clean_tree.configure(yscrollcommand=clean_vsb.set)
        enable_treeview_copy(self.clean_tree, columns)

    def _default_clean_output_path(self):
        bom_path = self.clean_bom_var.get().strip() if hasattr(self, 'clean_bom_var') else ''
        if not bom_path:
            return ''
        folder = os.path.dirname(bom_path)
        stem = os.path.splitext(os.path.basename(bom_path))[0]
        return os.path.join(folder, f'{stem}_排产清洗后.xlsx')

    def _choose_clean_bom_file(self):
        path = filedialog.askopenfilename(
            title='选择需要清洗的 BOM 文件',
            filetypes=[('Excel/CSV 文件', '*.xlsx *.xls *.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.clean_bom_var.set(path)
        if not self.clean_output_var.get().strip():
            self.clean_output_var.set(self._default_clean_output_path())
        self.status_var.set(f'已选择 BOM: {os.path.basename(path)}')

    def _choose_clean_plan_file(self):
        path = filedialog.askopenfilename(
            title='选择排产计划文件',
            filetypes=[('Excel/CSV 文件', '*.xlsx *.xls *.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.clean_plan_var.set(path)
        if not self.clean_output_var.get().strip():
            self.clean_output_var.set(self._default_clean_output_path())
        self.status_var.set(f'已选择排产计划: {os.path.basename(path)}')

    def _choose_clean_output_file(self):
        default_path = self.clean_output_var.get().strip() or self._default_clean_output_path()
        initialdir = os.path.dirname(default_path) if default_path else os.getcwd()
        initialfile = os.path.basename(default_path) if default_path else 'BOM_排产清洗后.xlsx'
        path = filedialog.asksaveasfilename(
            title='指定清洗后 BOM 输出路径',
            defaultextension='.xlsx',
            initialdir=initialdir,
            initialfile=initialfile,
            filetypes=[('Excel 文件', '*.xlsx'), ('CSV 文件', '*.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.clean_output_var.set(path)
        self.status_var.set(f'已指定输出: {os.path.basename(path)}')

    def _run_bom_clean(self):
        if self._bom_clean_running:
            self.status_var.set('BOM 清洗正在执行中，请稍候')
            return
        bom_path = self.clean_bom_var.get().strip()
        plan_path = self.clean_plan_var.get().strip()
        output_path = self.clean_output_var.get().strip() or self._default_clean_output_path()
        use_loaded_plan = bool(
            getattr(self, 'clean_use_loaded_plan_var', None)
            and self.clean_use_loaded_plan_var.get()
            and self._po_data
        )
        if not bom_path or not os.path.exists(bom_path):
            messagebox.showwarning('缺少BOM', '请先选择原始 BOM 文件')
            return
        if not use_loaded_plan and (not plan_path or not os.path.exists(plan_path)):
            messagebox.showwarning('缺少排产计划', '请先在“MRP计算-外采物料管理”上传排产计划，或在本页选择排产计划文件')
            return
        if not output_path:
            messagebox.showwarning('缺少输出路径', '请指定清洗后 BOM 输出路径')
            return
        if not os.path.splitext(output_path)[1]:
            output_path += '.xlsx'
            self.clean_output_var.set(output_path)
        if os.path.normcase(os.path.abspath(output_path)) == os.path.normcase(os.path.abspath(bom_path)):
            messagebox.showwarning('输出路径无效', '输出文件不能覆盖原始 BOM，请换一个文件名')
            return
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        self._clear_bom_clean_result()
        self._bom_clean_running = True
        self.clean_run_btn.state(['disabled'])
        self.clean_summary_var.set('正在清洗BOM...')
        plan_source_text = f'MRP计算已上传排产计划（{len(self._po_data)}项）' if use_loaded_plan else os.path.basename(plan_path)
        self.status_var.set(f'正在按排产计划递归清洗 BOM... 来源：{plan_source_text}')
        plan_items = list(self._po_data) if use_loaded_plan else None
        self._bom_clean_thread = threading.Thread(
            target=self._run_bom_clean_worker,
            args=(bom_path, plan_path, output_path, bool(self.clean_autoload_var.get()), plan_items, plan_source_text),
            daemon=True,
        )
        self._bom_clean_thread.start()

    def _run_bom_clean_worker(self, bom_path, plan_path, output_path, autoload, plan_items=None, plan_source=None):
        try:
            result = clean_bom_by_production_plan(
                bom_path,
                plan_path,
                output_path,
                plan_items=plan_items,
                plan_source=plan_source,
            )
            self.root.after(0, lambda: self._finish_bom_clean(result, autoload))
        except Exception as exc:
            err_msg = str(exc)
            self.root.after(0, lambda err_msg=err_msg: self._handle_bom_clean_error(err_msg))

    def _finish_bom_clean(self, result, autoload):
        self._bom_clean_running = False
        self._bom_clean_last_result = result
        self.clean_run_btn.state(['!disabled'])
        self.clean_tree.delete(*self.clean_tree.get_children(''))

        rows = [
            ('原始BOM表头行', result.get('header_rows', 0), '已从原文件完整保留到输出文件顶部'),
            ('原始BOM数据行', result.get('total_rows', 0), '不含表头'),
            ('保留BOM数据行', result.get('kept_rows', 0), '排产母件及其递归下层父项对应的BOM行'),
            ('删除BOM数据行', result.get('deleted_rows', 0), '父项不在排产展开链路内的BOM行'),
            ('排产母件数', result.get('root_count', 0), '排产计划中数量大于0的母件'),
            ('排产来源', result.get('plan_source', ''), '清洗使用的排产计划来源'),
            ('匹配到BOM的排产母件数', result.get('matched_root_count', 0), '这些母件作为递归展开起点'),
            ('BOM中缺失的排产母件数', len(result.get('missing_roots', [])), '排产有但BOM中没有父项记录'),
            ('递归保留父项数', result.get('reachable_parent_count', 0), '包含排产母件和所有下层仍可继续展开的半成品父项'),
            ('列数不足跳过行数', result.get('skipped_short_rows', 0), '这些行列数小于当前工具识别BOM所需列数'),
            ('输出文件', result.get('output_path', ''), '已生成清洗后的BOM文件'),
        ]
        for row in rows:
            self.clean_tree.insert('', 'end', values=row)
        missing_rows = result.get('missing_root_rows', [])
        if missing_rows:
            self.clean_tree.insert('', 'end', values=(
                '缺失清单说明',
                f'{len(missing_rows)} 个',
                '完整清单已写入输出文件 sheet【排产有但BOM缺失】，下方也显示明细，可直接复制。',
            ))
            for item in missing_rows[:500]:
                self.clean_tree.insert('', 'end', values=(
                    '排产有但BOM缺失',
                    item.get('code', ''),
                    f'排产数量: {fmt_qty(item.get("qty", 0))}；{item.get("reason", "")}',
                ))
            if len(missing_rows) > 500:
                self.clean_tree.insert('', 'end', values=(
                    '缺失清单截断',
                    f'已显示500/{len(missing_rows)}',
                    '界面只显示前500条，完整清单请看输出Excel的【排产有但BOM缺失】sheet。',
                ))

        summary = f'清洗完成 | 保留 {result.get("kept_rows", 0)} 行，删除 {result.get("deleted_rows", 0)} 行'
        autoload_note = ''
        if autoload:
            try:
                cleaned_index, _ = load_bom_from_file(result['output_path'])
                self.bom_index = cleaned_index
                self._diff_meta_cache = None
                self._bom_edges_cache = None
                self.current_file = result['output_path']
                save_cache(cleaned_index, CACHE_FILE)
                self.file_var.set(os.path.basename(result['output_path']))
                if hasattr(self, 'clean_bom_var'):
                    self.clean_bom_var.set(result['output_path'])
                self._clear_tree()
                self._clear_calc()
                self._clear_diff()
                self._clear_balance_preview()
                self._clear_readiness_preview()
                self._update_balance_bom_badge()
                self._update_readiness_bom_badge()
                self._update_arrival_bom_badge()
                self._refresh_calc_hint()
                autoload_note = f' | 已加载为当前BOM：{len(cleaned_index)} 个母件'
            except Exception as exc:
                autoload_note = f' | 自动加载失败：{exc}'
        self.clean_summary_var.set(summary + autoload_note)
        self.status_var.set(summary + autoload_note)
        messagebox.showinfo('清洗完成', f'{summary}\n输出文件：\n{result.get("output_path", "")}{autoload_note}')

    def _handle_bom_clean_error(self, err_msg):
        self._bom_clean_running = False
        if hasattr(self, 'clean_run_btn'):
            self.clean_run_btn.state(['!disabled'])
        self.clean_summary_var.set('清洗失败')
        show_data_error(
            self.root,
            title='BOM清洗失败',
            summary='未能完成按排产计划清洗 BOM。',
            detail=err_msg,
            fix_hint=_suggest_fix_for_exception(Exception(err_msg)),
        )
        self.status_var.set(f'BOM清洗失败: {err_msg}')

    def _clear_bom_clean_result(self):
        if hasattr(self, 'clean_tree'):
            self.clean_tree.delete(*self.clean_tree.get_children(''))
        self._bom_clean_last_result = None
        if hasattr(self, 'clean_summary_var'):
            self.clean_summary_var.set('等待清洗')

    def _build_balance_tab(self):
        f = self.tab_balance
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='平衡表生成参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(0, weight=3)
        control.columnconfigure(1, weight=2)

        self.balance_bom_var = tk.StringVar(value='当前 BOM: 未加载')
        ttk.Label(control, textvariable=self.balance_bom_var, style='Metric.TLabel').grid(
            row=0, column=0, columnspan=2, sticky='w', pady=(0, 6)
        )

        default_input = ''
        default_usage_flow = ''
        default_arrival_record = ''
        default_reply = ''

        self.balance_input_var = tk.StringVar(value='')
        self.balance_template_var = tk.StringVar(value=default_balance_template_path())
        self.balance_usage_flow_var = tk.StringVar(value=default_usage_flow)
        self.balance_arrival_record_var = tk.StringVar(value=default_arrival_record)
        self.balance_output_var = tk.StringVar(value='')
        self.balance_reply_source_paths = []
        self.balance_reply_sources_var = tk.StringVar(
            value=self._format_selected_file_summary(self.balance_reply_source_paths) if self.balance_reply_source_paths else '未选择旧平衡表'
        )
        self.balance_summary_var = tk.StringVar(value='等待生成')
        self.balance_shortage_stat_var = tk.StringVar(value='缺料明细: 0')
        self.balance_material_stat_var = tk.StringVar(value='平衡料号: 0')
        self.balance_suggest_stat_var = tk.StringVar(value='建议行: 0')
        self.balance_suggest_pn_stat_var = tk.StringVar(value='建议料号: 0')
        self.balance_apply_suggestion_exclusions_var = tk.BooleanVar(value=True)

        form_panel = ttk.Frame(control)
        form_panel.grid(row=1, column=0, sticky='ew')
        form_panel.columnconfigure(1, weight=1)
        form_panel.columnconfigure(4, weight=1)

        def _balance_field(parent, row_index, label_text, variable, button_text, command, column_offset):
            ttk.Label(parent, text=label_text, style='Section.TLabel').grid(
                row=row_index, column=column_offset, sticky='w', padx=(0, 6), pady=3
            )
            label, _ = make_path_label(parent, variable)
            label.grid(row=row_index, column=column_offset + 1, sticky='ew', pady=3)
            ttk.Button(parent, text=button_text, command=command).grid(
                row=row_index, column=column_offset + 2, sticky='w', padx=(6, 0), pady=3
            )

        _balance_field(form_panel, 0, 'MRP 输入文件', self.balance_input_var, '选择输入文件', self._choose_balance_input_file, 0)
        _balance_field(form_panel, 0, '库存流水', self.balance_usage_flow_var, '选择流水', self._choose_balance_usage_flow_file, 3)
        _balance_field(form_panel, 1, '到货记录', self.balance_arrival_record_var, '选择到货', self._choose_balance_arrival_record_file, 0)
        _balance_field(form_panel, 1, '旧平衡表', self.balance_reply_sources_var, '选择旧表', self._choose_balance_reply_files, 3)
        _balance_field(form_panel, 2, '输出文件', self.balance_output_var, '指定输出路径', self._choose_balance_output_file, 0)
        form_panel.grid_columnconfigure(1, weight=1)
        form_panel.grid_columnconfigure(4, weight=1)

        side_panel = ttk.Frame(control)
        side_panel.grid(row=1, column=1, sticky='nsew', padx=(14, 0))
        side_panel.columnconfigure(0, weight=1)
        side_panel.columnconfigure(1, weight=1)

        ttk.Checkbutton(
            side_panel,
            text='建议排产应用排除清单',
            variable=self.balance_apply_suggestion_exclusions_var,
        ).grid(row=0, column=0, sticky='w', padx=(0, 8))

        note = ttk.Label(
            side_panel,
            text='复用当前 BOM；输入文件里的“物料编码排除清单”仍然有效。',
            style='Subtle.TLabel',
            wraplength=250,
            justify='left',
        )
        note.grid(row=0, column=1, sticky='w')

        action_row = ttk.Frame(side_panel)
        action_row.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(6, 0))
        action_row.columnconfigure(6, weight=1)
        self.balance_generate_btn = ttk.Button(action_row, text='开始生成', style='Accent.TButton', command=self._run_balance_generation)
        self.balance_generate_btn.grid(row=0, column=0, sticky='w')
        ttk.Button(action_row, text='清空预览', style='Quiet.TButton', command=self._clear_balance_preview).grid(row=0, column=1, sticky='w', padx=(8, 0))
        ttk.Button(action_row, text='查看日志', style='Quiet.TButton', command=lambda: self._show_log_window('balance')).grid(row=0, column=2, sticky='w', padx=(8, 0))
        ttk.Label(action_row, textvariable=self.balance_summary_var, style='Metric.TLabel', wraplength=240, justify='left').grid(
            row=0, column=3, sticky='w', padx=(10, 0)
        )
        self.balance_progress = ttk.Progressbar(action_row, mode='determinate', length=180, maximum=100)
        self.balance_progress.grid(row=1, column=0, columnspan=4, sticky='ew', pady=(6, 0))
        self.balance_progress.grid_remove()
        self._balance_stage_counter = 0

        result = ttk.LabelFrame(f, text='补排产建议预览', style='Card.TLabelframe', padding=6)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        result.columnconfigure(0, weight=1)
        result.rowconfigure(1, weight=1)

        stats = ttk.Frame(result)
        stats.grid(row=0, column=0, sticky='ew', pady=(0, 6))
        for idx in range(4):
            stats.columnconfigure(idx, weight=1)
        ttk.Label(stats, textvariable=self.balance_shortage_stat_var, style='Badge.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.balance_material_stat_var, style='Badge.TLabel').grid(row=0, column=1, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.balance_suggest_stat_var, style='Badge.TLabel').grid(row=0, column=2, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.balance_suggest_pn_stat_var, style='Badge.TLabel').grid(row=0, column=3, sticky='w')

        tree_frame = ttk.Frame(result)
        tree_frame.grid(row=1, column=0, sticky='nsew')
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        self.balance_tree = ttk.Treeview(tree_frame, columns=BALANCE_SUGGEST_HEADERS, show='headings', selectmode='extended')
        for h, w in zip(BALANCE_SUGGEST_HEADERS, COL_WIDTHS_BALANCE):
            self.balance_tree.column(h, width=w, anchor='w')
            self.balance_tree.heading(h, text=h)
        enable_treeview_sort(self.balance_tree, BALANCE_SUGGEST_HEADERS,
                             numeric_columns={'上线数量', '直接上层数量', '顶层来源数量'})
        enable_treeview_copy(self.balance_tree, BALANCE_SUGGEST_HEADERS)
        self.balance_tree.bind('<Double-Button-1>', self._copy_calc_cell)
        self.balance_tree.grid(row=0, column=0, sticky='nsew')

        balance_vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.balance_tree.yview)
        balance_hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.balance_tree.xview)
        self.balance_tree.configure(yscrollcommand=balance_vsb.set, xscrollcommand=balance_hsb.set)
        balance_vsb.grid(row=0, column=1, sticky='ns')
        balance_hsb.grid(row=1, column=0, sticky='ew')

        self._update_balance_bom_badge()

    def export_bom_dataframe(self):
        columns = [
            '状态', '母件料号', '母件品名', '母件规格', '版本号', '生产目的', '生产单位', '项目', '批量',
            '生效日期', '失效日期', 'BOM子项.子件料品.料号', 'BOM子项.子件料品.参考料号1',
            'BOM子项.子件料品.料品名称', 'BOM子项.子件料品.规格', 'BOM子项.子件用量'
        ]
        rows = []
        for records in self.bom_index.values():
            for row in records:
                rows.append({
                    '状态': row[0] if len(row) > 0 else '',
                    '母件料号': row[1] if len(row) > 1 else '',
                    '母件品名': row[2] if len(row) > 2 else '',
                    '母件规格': row[3] if len(row) > 3 else '',
                    '版本号': row[4] if len(row) > 4 else '',
                    '生产目的': row[5] if len(row) > 5 else '',
                    '生产单位': row[6] if len(row) > 6 else '',
                    '项目': row[7] if len(row) > 7 else '',
                    '批量': row[8] if len(row) > 8 else '',
                    '生效日期': row[9] if len(row) > 9 else '',
                    '失效日期': row[10] if len(row) > 10 else '',
                    'BOM子项.子件料品.料号': row[11] if len(row) > 11 else '',
                    'BOM子项.子件料品.参考料号1': row[12] if len(row) > 12 else '',
                    'BOM子项.子件料品.料品名称': row[13] if len(row) > 13 else '',
                    'BOM子项.子件料品.规格': row[14] if len(row) > 14 else '',
                    'BOM子项.子件用量': row[15] if len(row) > 15 else '',
                })
        return pd.DataFrame(rows, columns=columns)

    def _ensure_log_window(self):
        if self._log_window is not None and self._log_window.winfo_exists():
            return self._log_window

        win = tk.Toplevel(self.root)
        win.title('运行日志')
        win.geometry('860x480')
        win.minsize(720, 360)
        win.transient(self.root)
        win.columnconfigure(0, weight=1)
        win.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(win)
        notebook.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        self._log_window = win
        self._log_notebook = notebook
        self._log_text_widgets = {}

        for log_key, title in (
            ('balance', '平衡表生成日志'),
            ('readiness', '外购齐套分析日志'),
            ('arrival', '到货跟催分析日志'),
        ):
            frame = ttk.Frame(notebook)
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)
            text = tk.Text(
                frame,
                font=('Microsoft YaHei', 9),
                bg='#FFFFFF',
                fg='#1E293B',
                insertbackground='#1E293B',
                relief='flat',
                padx=10,
                pady=8,
            )
            text.grid(row=0, column=0, sticky='nsew')
            scroll = ttk.Scrollbar(frame, orient='vertical', command=text.yview)
            text.configure(yscrollcommand=scroll.set)
            scroll.grid(row=0, column=1, sticky='ns')
            notebook.add(frame, text=title)
            self._log_text_widgets[log_key] = text
            self._refresh_log_widget(log_key)

        def _on_close():
            if self._log_window is not None and self._log_window.winfo_exists():
                self._log_window.withdraw()

        win.protocol('WM_DELETE_WINDOW', _on_close)
        return win

    def _refresh_log_widget(self, log_key):
        widget = self._log_text_widgets.get(log_key)
        if widget is None or not widget.winfo_exists():
            return
        widget.delete('1.0', 'end')
        lines = self._log_buffers.get(log_key, [])
        if lines:
            widget.insert('1.0', '\n'.join(lines) + '\n')
            widget.see('end')

    def _append_named_log(self, log_key, message):
        self._log_buffers.setdefault(log_key, []).append(str(message))
        self._refresh_log_widget(log_key)

    def _clear_named_log(self, log_key):
        self._log_buffers[log_key] = []
        self._refresh_log_widget(log_key)

    def _show_log_window(self, log_key='balance'):
        win = self._ensure_log_window()
        if win.state() == 'withdrawn':
            win.deiconify()
        win.lift()
        try:
            win.focus_force()
        except Exception:
            pass
        if self._log_notebook is not None:
            tab_index_map = {'balance': 0, 'readiness': 1, 'arrival': 2}
            self._log_notebook.select(tab_index_map.get(log_key, 0))
        self._refresh_log_widget(log_key)

    def _update_balance_bom_badge(self):
        if not hasattr(self, 'balance_bom_var'):
            return
        if not self.bom_index:
            self.balance_bom_var.set('当前 BOM: 未加载')
            return
        source = os.path.basename(self.current_file) if self.current_file else '缓存'
        self.balance_bom_var.set(f'当前 BOM: {source} | 母件 {len(self.bom_index)} 个')

    def _append_balance_log(self, message):
        self._append_named_log('balance', message)

    def _suggest_balance_output_path(self, input_path):
        if not input_path:
            return ''
        folder = os.path.dirname(input_path)
        name, _ext = os.path.splitext(os.path.basename(input_path))
        return os.path.join(folder, f'{name}_平衡表.xlsx')

    def _choose_balance_input_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP 输入文件',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.balance_input_var.set(path)
        if not self.balance_output_var.get().strip():
            self.balance_output_var.set(self._suggest_balance_output_path(path))
        self._append_balance_log(f'已选择输入文件: {path}')

    def _choose_balance_template_file(self):
        path = filedialog.askopenfilename(
            title='选择平衡表模板',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.balance_template_var.set(path)
        self._append_balance_log(f'已选择模板: {path}')

    def _choose_balance_usage_flow_file(self):
        path = filedialog.askopenfilename(
            title='选择库存流水记录',
            initialdir=os.path.join(os.path.expanduser('~'), 'Downloads'),
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.balance_usage_flow_var.set(path)
        self._append_balance_log(f'已选择库存流水: {path}')

    def _choose_balance_arrival_record_file(self):
        path = filedialog.askopenfilename(
            title='选择到货记录',
            initialdir=os.path.join(os.path.expanduser('~'), 'Downloads'),
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.balance_arrival_record_var.set(path)
        self._append_balance_log(f'已选择到货记录: {path}')

    def _choose_balance_output_file(self):
        initial = self.balance_output_var.get().strip() or self._suggest_balance_output_path(self.balance_input_var.get().strip()) or WORKSPACE
        path = filedialog.asksaveasfilename(
            title='指定输出文件',
            defaultextension='.xlsx',
            initialfile=os.path.basename(initial) if initial else '',
            initialdir=os.path.dirname(initial) if initial else WORKSPACE,
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
        self.balance_output_var.set(path)
        self._append_balance_log(f'输出文件设置为: {path}')

    def _merge_selected_paths(self, existing_paths, new_paths):
        merged = []
        seen = set()
        for path in list(existing_paths or []) + list(new_paths or []):
            normalized = os.path.normcase(os.path.abspath(path))
            if normalized in seen:
                continue
            seen.add(normalized)
            merged.append(path)
        return merged

    def _format_selected_file_summary(self, paths, label='旧平衡表'):
        if not paths:
            return f'未选择{label}'
        return f'已选择 {len(paths)} 个{label}'

    def _run_input_precheck_dialog(self, input_path, external_bom_df, *, log_callback):
        if run_input_precheck is None:
            return True
        self.status_var.set('正在检查输入数据，请稍候...')
        self.root.update_idletasks()
        try:
            issues = run_input_precheck(Path(input_path), external_bom_df=external_bom_df)
        except Exception as exc:
            show_data_error(
                self.root,
                title='数据预检失败',
                summary='预检过程出现异常，无法判断输入数据是否可用。',
                detail=str(exc),
                fix_hint=_suggest_fix_for_exception(exc),
            )
            self.status_var.set('数据预检失败')
            return False

        if not issues:
            log_callback('数据预检通过，未发现明显源数据问题。')
            return True

        message_lines = ['检测到以下可能的源数据问题：', '']
        for index, issue in enumerate(issues, start=1):
            message_lines.append(f'{index}. {issue}')
            message_lines.append('')
            log_callback(f'数据预检提醒 {index}: {issue}')
        message_lines.append('是否继续执行？')
        return messagebox.askyesno('数据预检提醒', '\n'.join(message_lines))

    def _choose_balance_reply_files(self):
        paths = filedialog.askopenfilenames(
            title='\u9009\u62e9\u5df2\u586b\u5199\u7684\u65e7\u5e73\u8861\u8868',
            filetypes=[('Excel \u6587\u4ef6', '*.xlsx *.xls'), ('\u6240\u6709\u6587\u4ef6', '*.*')]
        )
        if not paths:
            return
        self.balance_reply_source_paths = self._merge_selected_paths(self.balance_reply_source_paths, paths)
        self.balance_reply_sources_var.set(self._format_selected_file_summary(self.balance_reply_source_paths))
        self._append_balance_log(f'旧平衡表累计已选择 {len(self.balance_reply_source_paths)} 个')

    def _clear_balance_preview(self):
        if hasattr(self, 'balance_tree'):
            self.balance_tree.delete(*self.balance_tree.get_children(''))
        self._clear_named_log('balance')
        if hasattr(self, 'balance_summary_var'):
            self.balance_summary_var.set('等待生成')
            self.balance_shortage_stat_var.set('缺料明细: 0')
            self.balance_material_stat_var.set('平衡料号: 0')
            self.balance_suggest_stat_var.set('建议行: 0')
            self.balance_suggest_pn_stat_var.set('建议料号: 0')

    def _run_balance_generation(self):
        if run_balance_pipeline is None:
            show_data_error(
                self.root,
                title='平衡表引擎不可用',
                summary='未能加载平衡表引擎（mrp_balance_tool），无法运行生成流程。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='运行源码请确认同级目录有 "静态平衡表/src"；打包后请确认 spec 文件正确包含 mrp_balance_tool 模块。',
            )
            return
        if not self.bom_index:
            messagebox.showwarning('缺少数据', '请先上传 BOM 文件')
            return
        if self._balance_running:
            self.status_var.set('平衡表正在生成中，请稍候')
            return

        input_path = self.balance_input_var.get().strip()
        template_path = self.balance_template_var.get().strip() or default_balance_template_path()
        usage_flow_path = self.balance_usage_flow_var.get().strip()
        arrival_record_path = self.balance_arrival_record_var.get().strip()
        output_path = self.balance_output_var.get().strip()
        reply_source_paths = list(self.balance_reply_source_paths)
        apply_suggestion_exclusions = bool(self.balance_apply_suggestion_exclusions_var.get())
        if not input_path:
            messagebox.showwarning('参数不完整', '请先选择 MRP 输入文件')
            return
        if not output_path:
            messagebox.showwarning('参数不完整', '请先指定输出文件路径')
            return
        if usage_flow_path and not os.path.exists(usage_flow_path):
            messagebox.showwarning('文件不存在', f'库存流水记录不存在：\n{usage_flow_path}')
            return
        if arrival_record_path and not os.path.exists(arrival_record_path):
            messagebox.showwarning('文件不存在', f'到货记录不存在：\n{arrival_record_path}')
            return

        external_bom_df = self.export_bom_dataframe()
        if not self._run_input_precheck_dialog(input_path, external_bom_df, log_callback=self._append_balance_log):
            self.balance_summary_var.set('已取消，请先检查源数据')
            self.status_var.set(self.balance_summary_var.get())
            return
        self._balance_running = True
        self.balance_output_path = output_path
        self.balance_generate_btn.state(['disabled'])
        self.balance_progress.grid()
        self._balance_stage_counter = 0
        self.balance_progress['value'] = 0
        self.balance_summary_var.set('正在生成平衡表，请稍候...')
        self.status_var.set(self.balance_summary_var.get())
        self._append_balance_log('开始生成平衡表...')
        self._balance_thread = threading.Thread(
            target=self._run_balance_generation_worker,
            args=(
                input_path,
                output_path,
                template_path,
                usage_flow_path,
                arrival_record_path,
                external_bom_df,
                reply_source_paths,
                apply_suggestion_exclusions,
            ),
            daemon=True,
        )
        self._balance_thread.start()

    def _run_balance_generation_worker(
        self,
        input_path,
        output_path,
        template_path,
        usage_flow_path,
        arrival_record_path,
        external_bom_df,
        reply_source_paths,
        apply_suggestion_exclusions,
    ):
        try:
            result = run_balance_pipeline(
                Path(input_path),
                Path(output_path),
                Path(template_path) if template_path else None,
                external_bom_df=external_bom_df,
                carry_forward_paths=[Path(path) for path in reply_source_paths],
                apply_suggestion_exclusions=apply_suggestion_exclusions,
                usage_flow_path=Path(usage_flow_path) if usage_flow_path else None,
                arrival_record_path=Path(arrival_record_path) if arrival_record_path else None,
                progress_callback=lambda message: self.root.after(
                    0,
                    lambda msg=message: self._report_balance_progress(msg),
                ),
            )
            self.root.after(0, lambda: self._finish_balance_generation(result))
        except Exception as exc:
            message = str(exc)
            self.root.after(0, lambda msg=message: self._handle_balance_error(msg))

    def _report_balance_progress(self, message):
        self._balance_stage_counter += 1
        # 典型流水线约 20 个阶段，渐进式推进进度条；保持在 95 以下，结束时补满
        progress = min(95, self._balance_stage_counter * 5)
        try:
            self.balance_progress['value'] = progress
        except Exception:
            pass
        stage_label = f'[阶段 {self._balance_stage_counter}] '
        self.balance_summary_var.set(stage_label + message)
        self.status_var.set(stage_label + message)
        self._append_balance_log(message)

    def _finish_balance_generation(self, result):
        self._balance_running = False
        self.balance_progress['value'] = 100
        self.balance_progress.grid_remove()
        self.balance_generate_btn.state(['!disabled'])

        self.balance_tree.delete(*self.balance_tree.get_children(''))
        suggestion_df = getattr(result, 'suggestion_df', pd.DataFrame())
        preview_source = '本次计算'
        if suggestion_df.empty and self.balance_output_path and os.path.exists(self.balance_output_path):
            try:
                preserved_df = pd.read_excel(self.balance_output_path, sheet_name='补排产建议')
                preserved_df = normalize_sheet_columns(preserved_df)
                if not preserved_df.empty:
                    suggestion_df = preserved_df
                    preview_source = '保留历史'
            except Exception as exc:
                self._append_balance_log(f'读取补排产建议预览失败: {exc}')
        for record in suggestion_df.to_dict('records'):
            values = [
                record.get('客户', ''),
                record.get('母件料号', ''),
                record.get('母件品名', ''),
                record.get('母件规格', ''),
                '' if pd.isna(record.get('上线日期', '')) else str(record.get('上线日期', '')),
                fmt_qty(record.get('上线数量', 0)),
                str(record.get('直接上层数量', '')),
                record.get('直接上层摘要', ''),
                str(record.get('顶层来源数量', '')),
                record.get('顶层来源摘要', ''),
                record.get('说明', ''),
            ]
            self.balance_tree.insert('', 'end', values=values)

        shortage_count = len(getattr(result, 'shortage_df', []))
        purchase_count = len(getattr(result, 'purchase_view_df', []))
        suggestion_count = len(suggestion_df)
        suggestion_material_count = suggestion_df['母件料号'].astype(str).str.strip().nunique() if not suggestion_df.empty else 0
        summary = f'生成完成 | 平衡料号 {purchase_count} 种 | 补排产建议 {suggestion_count} 行'
        if preview_source != '本次计算' and suggestion_count:
            summary += f'（{preview_source}）'
        self.balance_summary_var.set(summary)
        self.balance_shortage_stat_var.set(f'缺料明细: {shortage_count}')
        self.balance_material_stat_var.set(f'平衡料号: {purchase_count}')
        self.balance_suggest_stat_var.set(f'建议行: {suggestion_count}')
        self.balance_suggest_pn_stat_var.set(f'建议料号: {suggestion_material_count}')
        self.status_var.set(summary)
        self._append_balance_log(summary)
        carried_count = int(getattr(result, 'carried_reply_cell_count', 0) or 0)
        carried_material_count = int(getattr(result, 'carried_reply_material_count', 0) or 0)
        carried_file_count = int(getattr(result, 'carried_reply_file_count', 0) or 0)
        if carried_count:
            self._append_balance_log(f'\u5df2\u5e26\u5165\u91c7\u8d2d\u7b54\u590d {carried_count} \u683c\uff0c\u8986\u76d6\u6599\u53f7 {carried_material_count} \u4e2a\uff0c\u6765\u6e90\u6587\u4ef6 {carried_file_count} \u4e2a')
        if self.balance_output_path:
            self._append_balance_log(f'输出文件: {self.balance_output_path}')
            messagebox.showinfo('生成完成', f'已生成:\n{self.balance_output_path}')

    def _handle_balance_error(self, message):
        self._balance_running = False
        self.balance_progress['value'] = 0
        self.balance_progress.grid_remove()
        self.balance_generate_btn.state(['!disabled'])
        self.balance_summary_var.set('平衡表生成失败')
        self.status_var.set('平衡表生成失败')
        self._append_balance_log(f'生成失败: {message}')
        show_data_error(
            self.root,
            title='平衡表生成失败',
            summary='平衡表生成过程中出错，结果文件未生成。',
            detail=message,
            fix_hint=_suggest_fix_for_exception(Exception(message)),
        )

    def _build_readiness_tab(self):
        f = self.tab_readiness
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='外购齐套分析参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(0, weight=3)
        control.columnconfigure(1, weight=2)

        self.readiness_bom_var = tk.StringVar(value='当前 BOM: 未加载')
        ttk.Label(control, textvariable=self.readiness_bom_var, style='Metric.TLabel').grid(
            row=0, column=0, columnspan=2, sticky='w', pady=(0, 6)
        )

        self.readiness_input_var = tk.StringVar()
        self.readiness_reply_sources_var = tk.StringVar(value='未选择旧平衡表')
        self.readiness_summary_var = tk.StringVar(value='等待分析')
        self.readiness_root_stat_var = tk.StringVar(value='母料号: 0')
        self.readiness_material_stat_var = tk.StringVar(value='外购物料: 0')
        self.readiness_issue_stat_var = tk.StringVar(value='问题物料: 0')
        self.readiness_ready_stat_var = tk.StringVar(value='最晚齐套: -')

        left_panel = ttk.Frame(control)
        left_panel.grid(row=1, column=0, sticky='ew')
        left_panel.columnconfigure(1, weight=1)
        left_panel.columnconfigure(4, weight=1)

        ttk.Label(left_panel, text='MRP 输入文件', style='Section.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 6), pady=3)
        readiness_input_label, _ = make_path_label(left_panel, self.readiness_input_var)
        readiness_input_label.grid(row=0, column=1, sticky='ew', pady=3)
        ttk.Button(left_panel, text='选择输入文件', command=self._choose_readiness_input_file).grid(row=0, column=2, sticky='w', padx=(6, 0), pady=3)

        ttk.Label(left_panel, text='旧平衡表', style='Section.TLabel').grid(row=0, column=3, sticky='w', padx=(12, 6), pady=3)
        readiness_reply_label, _ = make_path_label(left_panel, self.readiness_reply_sources_var)
        readiness_reply_label.grid(row=0, column=4, sticky='ew', pady=3)
        ttk.Button(left_panel, text='选择旧表', command=self._choose_readiness_reply_files).grid(row=0, column=5, sticky='w', padx=(6, 0), pady=3)

        ttk.Label(
            left_panel,
            text='输入母料号和数量（双击单元格编辑；支持从 Excel 粘贴）',
            style='Subtle.TLabel',
        ).grid(row=1, column=0, columnspan=6, sticky='w', pady=(4, 3))

        input_frame = ttk.Frame(left_panel)
        input_frame.grid(row=2, column=0, columnspan=6, sticky='ew')
        input_frame.columnconfigure(0, weight=1)

        self.readiness_input_tree = ttk.Treeview(
            input_frame,
            columns=('母料号', '数量'),
            show='headings',
            selectmode='extended',
            height=4,
        )
        self.readiness_input_tree.heading('母料号', text='母料号')
        self.readiness_input_tree.heading('数量', text='数量')
        self.readiness_input_tree.column('母料号', width=200, anchor='w')
        self.readiness_input_tree.column('数量', width=100, anchor='e')
        self.readiness_input_tree.grid(row=0, column=0, sticky='ew')
        input_vsb = ttk.Scrollbar(input_frame, orient='vertical',
                                  command=self.readiness_input_tree.yview)
        self.readiness_input_tree.configure(yscrollcommand=input_vsb.set)
        input_vsb.grid(row=0, column=1, sticky='ns')

        self.readiness_input_tree.bind('<Double-Button-1>', self._readiness_edit_cell)
        self.readiness_input_tree.bind('<Control-v>', self._readiness_paste_clipboard)
        self.readiness_input_tree.bind('<Control-V>', self._readiness_paste_clipboard)

        btn_row = ttk.Frame(input_frame)
        btn_row.grid(row=1, column=0, columnspan=2, sticky='w', pady=(4, 0))
        ttk.Button(btn_row, text='＋ 添加行', command=self._readiness_add_row).pack(side='left')
        ttk.Button(btn_row, text='－ 删除选中', command=self._readiness_remove_selected).pack(side='left', padx=(6, 0))
        ttk.Button(btn_row, text='📋 从 Excel 粘贴', command=self._readiness_paste_clipboard).pack(side='left', padx=(6, 0))
        ttk.Button(btn_row, text='清空', style='Quiet.TButton', command=self._readiness_clear_input).pack(side='left', padx=(6, 0))

        # 初始给两行空位，降低用户学习成本
        for _ in range(2):
            self.readiness_input_tree.insert('', 'end', values=('', ''))

        side_panel = ttk.Frame(control)
        side_panel.grid(row=1, column=1, sticky='nsew', padx=(14, 0))
        side_panel.columnconfigure(0, weight=1)
        side_panel.columnconfigure(1, weight=1)

        note = ttk.Label(
            side_panel,
            text='复用当前 BOM；只分析外购物料，不计算半成品自制。',
            style='Subtle.TLabel',
            wraplength=250,
            justify='left',
        )
        note.grid(row=0, column=0, columnspan=2, sticky='w')

        action_row = ttk.Frame(side_panel)
        action_row.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(6, 0))
        action_row.columnconfigure(3, weight=1)
        self.readiness_run_btn = ttk.Button(action_row, text='开始分析', style='Accent.TButton', command=self._run_readiness_analysis)
        self.readiness_run_btn.grid(row=0, column=0, sticky='w')
        self.readiness_recommend_btn = ttk.Button(action_row, text='库存推荐', command=self._run_readiness_recommendation)
        self.readiness_recommend_btn.grid(row=0, column=1, sticky='w', padx=(8, 0))
        self.readiness_common_btn = ttk.Button(action_row, text='通用物料分析', command=self._run_common_material_analysis)
        self.readiness_common_btn.grid(row=0, column=2, sticky='w', padx=(8, 0))
        self.readiness_export_btn = ttk.Button(action_row, text='导出结果', command=self._export_readiness_result)
        self.readiness_export_btn.grid(row=0, column=3, sticky='w', padx=(8, 0))
        ttk.Button(action_row, text='清空结果', style='Quiet.TButton', command=self._clear_readiness_preview).grid(row=0, column=4, sticky='w', padx=(8, 0))
        ttk.Button(action_row, text='查看日志', style='Quiet.TButton', command=lambda: self._show_log_window('readiness')).grid(row=0, column=5, sticky='w', padx=(8, 0))
        ttk.Label(action_row, textvariable=self.readiness_summary_var, style='Metric.TLabel', wraplength=240, justify='left').grid(
            row=0, column=6, sticky='ew', padx=(10, 0)
        )
        self.readiness_progress = ttk.Progressbar(action_row, mode='determinate', length=180, maximum=100)
        self._readiness_stage_counter = 0
        self.readiness_progress.grid(row=1, column=0, columnspan=5, sticky='ew', pady=(6, 0))
        self.readiness_progress.grid_remove()

        result = ttk.LabelFrame(f, text='外购齐套结果', style='Card.TLabelframe', padding=6)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        result.columnconfigure(0, weight=1)
        result.rowconfigure(1, weight=1)

        stats = ttk.Frame(result)
        stats.grid(row=0, column=0, sticky='ew', pady=(0, 6))
        for idx in range(5):
            stats.columnconfigure(idx, weight=1)
        ttk.Label(stats, textvariable=self.readiness_root_stat_var, style='Badge.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.readiness_material_stat_var, style='Badge.TLabel').grid(row=0, column=1, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.readiness_issue_stat_var, style='Badge.TLabel').grid(row=0, column=2, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.readiness_ready_stat_var, style='Badge.TLabel').grid(row=0, column=3, sticky='w')
        ttk.Button(stats, text='导出齐套结果', command=self._export_readiness_result).grid(row=0, column=4, sticky='e')

        result_nb = ttk.Notebook(result)
        result_nb.grid(row=1, column=0, sticky='nsew')

        root_frame = ttk.Frame(result_nb)
        rolling_matrix_frame = ttk.Frame(result_nb)
        recommend_frame = ttk.Frame(result_nb)
        common_frame = ttk.Frame(result_nb)
        issue_frame = ttk.Frame(result_nb)
        material_frame = ttk.Frame(result_nb)
        unknown_frame = ttk.Frame(result_nb)
        result_nb.add(root_frame, text='母料号汇总')
        result_nb.add(rolling_matrix_frame, text='BOM横向滚动')
        result_nb.add(recommend_frame, text='库存推荐')
        result_nb.add(common_frame, text='通用物料排行')
        result_nb.add(issue_frame, text='问题物料')
        result_nb.add(material_frame, text='全量外购物料')
        result_nb.add(unknown_frame, text='未识别物料')

        self.readiness_root_tree = self._create_result_tree(root_frame, READINESS_ROOT_HEADERS, READINESS_ROOT_WIDTHS)
        self.readiness_rolling_matrix_tree = self._create_result_tree(rolling_matrix_frame, READINESS_ROLLING_MATRIX_HEADERS, READINESS_ROLLING_MATRIX_WIDTHS)
        self.readiness_recommend_tree = self._create_result_tree(recommend_frame, READINESS_RECOMMEND_HEADERS, READINESS_RECOMMEND_WIDTHS)
        self.readiness_common_tree = self._create_result_tree(common_frame, READINESS_COMMON_HEADERS, READINESS_COMMON_WIDTHS)
        self.readiness_issue_tree = self._create_result_tree(issue_frame, READINESS_ISSUE_HEADERS, READINESS_ISSUE_WIDTHS)
        self.readiness_material_tree = self._create_result_tree(material_frame, READINESS_MATERIAL_HEADERS, READINESS_MATERIAL_WIDTHS)
        self.readiness_unknown_tree = self._create_result_tree(unknown_frame, READINESS_UNKNOWN_HEADERS, READINESS_UNKNOWN_WIDTHS)

        self._update_readiness_bom_badge()

    def _create_result_tree(self, parent, headers, widths):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        tree = ttk.Treeview(parent, columns=headers, show='headings', selectmode='extended')
        numeric_headers = {h for h in headers
                           if any(keyword in h for keyword in ('数量', '需求', '可用', '可生产', '缺口', '物料数', '总计', '累计'))}
        for header, width in zip(headers, widths):
            anchor = 'e' if header in numeric_headers else 'w'
            tree.column(header, width=width, anchor=anchor)
            tree.heading(header, text=header)
        enable_treeview_sort(tree, list(headers), numeric_columns=numeric_headers)
        enable_treeview_copy(tree, list(headers))
        tree.bind('<Double-Button-1>', self._copy_calc_cell)
        tree.grid(row=0, column=0, sticky='nsew')

        vsb = ttk.Scrollbar(parent, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(parent, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        return tree

    def _configure_result_tree_columns(self, tree, headers, widths=None):
        widths = list(widths or [])
        tree.delete(*tree.get_children(''))
        tree['columns'] = list(headers)
        numeric_headers = {h for h in headers
                           if any(keyword in h for keyword in ('数量', '需求', '可用', '可生产', '缺口', '物料数', '总计', '累计', '用量', '库存', '扣减', '分配', '最大', '最小', '差值'))}
        for index, header in enumerate(headers):
            width = widths[index] if index < len(widths) else 110
            if header in ('品名', '规格', '替代料清单') or header.endswith('清单'):
                width = max(width, 220)
            if header.endswith(('需求', '分配前', '扣减', '缺口', '分配后')):
                width = max(width, 95)
            anchor = 'e' if header in numeric_headers else 'w'
            tree.column(header, width=width, anchor=anchor)
            tree.heading(header, text=header)
        enable_treeview_sort(tree, list(headers), numeric_columns=numeric_headers)
        enable_treeview_copy(tree, list(headers))

    def _update_readiness_bom_badge(self):
        if not hasattr(self, 'readiness_bom_var'):
            return
        if not self.bom_index:
            self.readiness_bom_var.set('当前 BOM: 未加载')
            return
        source = os.path.basename(self.current_file) if self.current_file else '缓存'
        self.readiness_bom_var.set(f'当前 BOM: {source} | 母件 {len(self.bom_index)} 个')

    def _append_readiness_log(self, message):
        self._append_named_log('readiness', message)

    def _choose_readiness_input_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP 输入文件',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.readiness_input_var.set(path)
        self._append_readiness_log(f'已选择输入文件: {path}')

    def _choose_readiness_reply_files(self):
        paths = filedialog.askopenfilenames(
            title='选择已填写的旧平衡表',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not paths:
            return
        self.readiness_reply_source_paths = self._merge_selected_paths(self.readiness_reply_source_paths, paths)
        self.readiness_reply_sources_var.set(self._format_selected_file_summary(self.readiness_reply_source_paths))
        self._append_readiness_log(f'旧平衡表累计已选择 {len(self.readiness_reply_source_paths)} 个')

    def _parse_readiness_root_items(self):
        if not hasattr(self, 'readiness_input_tree'):
            raise ValueError('输入表格未初始化')
        iids = self.readiness_input_tree.get_children('')
        if not iids:
            raise ValueError('请先在表格中添加母料号')
        merged: dict[str, float] = {}
        order: list[str] = []
        empty_rows = 0
        for row_index, iid in enumerate(iids, start=1):
            material_code = str(self.readiness_input_tree.set(iid, '母料号')).strip()
            qty_raw = str(self.readiness_input_tree.set(iid, '数量')).strip().replace(',', '')
            if not material_code and not qty_raw:
                empty_rows += 1
                continue
            if not material_code:
                raise ValueError(f'第 {row_index} 行缺少母料号')
            if not qty_raw:
                quantity = -1.0
            else:
                try:
                    quantity = float(qty_raw)
                except ValueError as exc:
                    raise ValueError(f'第 {row_index} 行数量无效：{qty_raw}') from exc
            if quantity == 0:
                continue
            if material_code not in merged:
                merged[material_code] = 0.0
                order.append(material_code)
            if quantity < 0:
                if merged[material_code] <= 0:
                    merged[material_code] = -1.0
            else:
                merged[material_code] = max(merged[material_code], 0.0) + quantity
        if not order:
            raise ValueError('未读取到有效的母料号，请检查表格')
        return [(material_code, merged[material_code]) for material_code in order]

    # ── 齐套分析：输入表格的交互行为 ───────────────────────
    def _readiness_add_row(self):
        iid = self.readiness_input_tree.insert('', 'end', values=('', ''))
        self.readiness_input_tree.selection_set(iid)
        self.readiness_input_tree.see(iid)

    def _readiness_remove_selected(self):
        sels = self.readiness_input_tree.selection()
        if not sels:
            return
        for iid in sels:
            self.readiness_input_tree.delete(iid)
        if not self.readiness_input_tree.get_children(''):
            # 保留至少一行空位便于用户继续输入
            self.readiness_input_tree.insert('', 'end', values=('', ''))

    def _readiness_clear_input(self):
        self.readiness_input_tree.delete(*self.readiness_input_tree.get_children(''))
        for _ in range(2):
            self.readiness_input_tree.insert('', 'end', values=('', ''))

    def _readiness_paste_clipboard(self, _event=None):
        try:
            data = self.root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning('粘贴失败', '剪贴板为空或无法读取')
            return 'break'
        if not data.strip():
            return 'break'
        # 删除表格末尾的空行，再把粘贴的每行追加
        for iid in list(self.readiness_input_tree.get_children('')):
            pn = str(self.readiness_input_tree.set(iid, '母料号')).strip()
            qty = str(self.readiness_input_tree.set(iid, '数量')).strip()
            if not pn and not qty:
                self.readiness_input_tree.delete(iid)

        added = 0
        skipped_header = False
        for raw_line in data.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            normalized = line.replace('，', ',').replace('\t', ',')
            parts = [p.strip() for p in normalized.split(',') if p.strip()]
            if len(parts) < 2:
                parts = line.split()
            if len(parts) < 2:
                continue
            pn, qty = parts[0], parts[1]
            # 第一行看起来像表头的话跳过
            if not skipped_header and added == 0 and ('料号' in pn or '数量' in qty):
                skipped_header = True
                continue
            self.readiness_input_tree.insert('', 'end', values=(pn, qty))
            added += 1
        if added == 0:
            messagebox.showinfo('粘贴', '未从剪贴板识别到有效数据（需要至少两列：母料号 数量）')
        else:
            self.status_var.set(f'已粘贴 {added} 行母料号到齐套输入表')
        return 'break'

    def _readiness_edit_cell(self, event):
        """双击单元格进入编辑模式，用 Entry 覆盖原位输入。"""
        tree = self.readiness_input_tree
        region = tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        column_id = tree.identify_column(event.x)
        iid = tree.identify_row(event.y)
        if not iid or not column_id:
            return
        col_index = int(column_id.replace('#', '')) - 1
        column_name = ('母料号', '数量')[col_index] if col_index in (0, 1) else None
        if column_name is None:
            return
        x, y, width, height = tree.bbox(iid, column_name)
        current = tree.set(iid, column_name)

        edit = tk.Entry(tree, font=('Microsoft YaHei', 10), borderwidth=1, relief='solid')
        edit.insert(0, current)
        edit.select_range(0, 'end')
        edit.focus_set()
        edit.place(x=x, y=y, width=width, height=height)

        def _commit(_e=None):
            new_value = edit.get().strip()
            if column_name == '数量' and new_value:
                try:
                    float(new_value.replace(',', ''))
                except ValueError:
                    messagebox.showwarning('数量无效', f'"{new_value}" 不是有效数字')
                    edit.focus_set()
                    return
            tree.set(iid, column_name, new_value)
            edit.destroy()

        def _cancel(_e=None):
            edit.destroy()

        edit.bind('<Return>', _commit)
        edit.bind('<Tab>', _commit)
        edit.bind('<Escape>', _cancel)
        edit.bind('<FocusOut>', _commit)

    def _clear_tree_rows(self, tree):
        if tree is not None:
            tree.delete(*tree.get_children(''))

    def _clear_readiness_preview(self):
        for tree_name in ('readiness_root_tree', 'readiness_producible_tree', 'readiness_rolling_matrix_tree', 'readiness_diff_capacity_tree', 'readiness_horizontal_tree', 'readiness_recommend_tree', 'readiness_common_tree', 'readiness_issue_tree', 'readiness_material_tree', 'readiness_unknown_tree'):
            tree = getattr(self, tree_name, None)
            if tree is not None:
                self._clear_tree_rows(tree)
        self._clear_named_log('readiness')
        if hasattr(self, 'readiness_summary_var'):
            self.readiness_summary_var.set('等待分析')
            self.readiness_root_stat_var.set('母料号: 0')
            self.readiness_material_stat_var.set('外购物料: 0')
            self.readiness_issue_stat_var.set('问题物料: 0')
            self.readiness_ready_stat_var.set('最晚齐套: -')

    def _export_readiness_result(self):
        tree_specs = [
            ('母料号汇总', getattr(self, 'readiness_root_tree', None)),
            ('BOM横向滚动', getattr(self, 'readiness_rolling_matrix_tree', None)),
            ('库存推荐', getattr(self, 'readiness_recommend_tree', None)),
            ('通用物料排行', getattr(self, 'readiness_common_tree', None)),
            ('问题物料', getattr(self, 'readiness_issue_tree', None)),
            ('全量外购物料', getattr(self, 'readiness_material_tree', None)),
            ('未识别物料', getattr(self, 'readiness_unknown_tree', None)),
        ]
        has_rows = any(tree is not None and tree.get_children('') for _name, tree in tree_specs)
        if not has_rows:
            messagebox.showwarning('无数据', '请先执行外购齐套分析或库存推荐')
            return
        path = filedialog.asksaveasfilename(
            title='导出外购齐套分析结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')],
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            header_fill = PatternFill('solid', fgColor='D9EAF7')
            header_font = Font(bold=True, color='000000')
            thin = Side(style='thin', color='D9E2EC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for sheet_name, tree in tree_specs:
                if tree is None or not tree.get_children(''):
                    continue
                ws = wb.create_sheet(sheet_name[:31])
                columns = list(tree['columns'])
                for ci, header in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=ci, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for ri, iid in enumerate(tree.get_children(''), 2):
                    values = tree.item(iid, 'values')
                    for ci, value in enumerate(values, 1):
                        cell = ws.cell(row=ri, column=ci, value=value)
                        cell.border = border
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                for ci, header in enumerate(columns, 1):
                    max_len = len(str(header))
                    for ri in range(2, min(ws.max_row, 200) + 1):
                        max_len = max(max_len, len(str(ws.cell(ri, ci).value or '')))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(8, min(60, max_len + 2))
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}{ws.max_row}"
            wb.save(path)
            self.status_var.set(f'已导出: {os.path.basename(path)}')
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
        except Exception as exc:
            messagebox.showerror('导出失败', str(exc))

    def _run_readiness_recommendation(self):
        if read_balance_workbook_tables is None:
            show_data_error(
                self.root,
                title='库存推荐引擎不可用',
                summary='未能加载 MRP 数据读取引擎，无法运行库存推荐。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='运行源码请确认同级目录有 "静态平衡表/src"；打包后请确认 spec 文件正确包含 mrp_balance_tool 模块。',
            )
            return
        if not self.bom_index:
            messagebox.showwarning('缺少数据', '请先上传 BOM 文件')
            return
        if self._readiness_running:
            self.status_var.set('外购齐套分析正在执行中，请稍候')
            return
        input_path = self.readiness_input_var.get().strip()
        if not input_path:
            messagebox.showwarning('参数不完整', '请先选择 MRP 输入文件')
            return

        external_bom_df = self.export_bom_dataframe()
        if not self._run_input_precheck_dialog(input_path, external_bom_df, log_callback=self._append_readiness_log):
            self.readiness_summary_var.set('已取消，请先检查源数据')
            self.status_var.set(self.readiness_summary_var.get())
            return

        self._readiness_running = True
        self.readiness_run_btn.state(['disabled'])
        self.readiness_recommend_btn.state(['disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['disabled'])
        self.readiness_progress.grid()
        self._readiness_stage_counter = 0
        self.readiness_progress['value'] = 0
        self.readiness_summary_var.set('正在按库存倒推可生产型号，请稍候...')
        self.status_var.set(self.readiness_summary_var.get())
        self._append_readiness_log('开始库存可生产型号推荐...')

        self._readiness_thread = threading.Thread(
            target=self._run_readiness_recommendation_worker,
            args=(input_path, external_bom_df),
            daemon=True,
        )
        self._readiness_thread.start()

    def _run_readiness_recommendation_worker(self, input_path, external_bom_df):
        try:
            rows = self._calculate_readiness_recommendation(
                input_path,
                external_bom_df,
                progress_callback=lambda message: self.root.after(
                    0,
                    lambda msg=message: self._report_readiness_progress(msg),
                ),
            )
            self.root.after(0, lambda: self._finish_readiness_recommendation(rows))
        except Exception as exc:
            self.root.after(0, lambda: self._handle_readiness_error(str(exc)))

    def _run_common_material_analysis(self):
        if read_balance_workbook_tables is None:
            show_data_error(
                self.root,
                title='通用物料分析引擎不可用',
                summary='未能加载 MRP 数据读取引擎，无法运行通用物料分析。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='运行源码请确认同级目录有 "静态平衡表/src"；打包后请确认 spec 文件正确包含 mrp_balance_tool 模块。',
            )
            return
        if not self.bom_index:
            messagebox.showwarning('缺少数据', '请先上传 BOM 文件')
            return
        if self._readiness_running:
            self.status_var.set('外购齐套分析正在执行中，请稍候')
            return
        input_path = self.readiness_input_var.get().strip()
        if not input_path:
            messagebox.showwarning('参数不完整', '请先选择 MRP 输入文件')
            return

        external_bom_df = self.export_bom_dataframe()
        if not self._run_input_precheck_dialog(input_path, external_bom_df, log_callback=self._append_readiness_log):
            self.readiness_summary_var.set('已取消，请先检查源数据')
            self.status_var.set(self.readiness_summary_var.get())
            return

        self._readiness_running = True
        self.readiness_run_btn.state(['disabled'])
        self.readiness_recommend_btn.state(['disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['disabled'])
        self.readiness_progress.grid()
        self._readiness_stage_counter = 0
        self.readiness_progress['value'] = 0
        self.readiness_summary_var.set('正在统计排产通用物料，请稍候...')
        self.status_var.set(self.readiness_summary_var.get())
        self._append_readiness_log('开始通用物料分析...')

        self._readiness_thread = threading.Thread(
            target=self._run_common_material_analysis_worker,
            args=(input_path, external_bom_df),
            daemon=True,
        )
        self._readiness_thread.start()

    def _run_common_material_analysis_worker(self, input_path, external_bom_df):
        try:
            rows = self._calculate_common_material_analysis(
                input_path,
                external_bom_df,
                progress_callback=lambda message: self.root.after(
                    0,
                    lambda msg=message: self._report_readiness_progress(msg),
                ),
            )
            self.root.after(0, lambda: self._finish_common_material_analysis(rows))
        except Exception as exc:
            self.root.after(0, lambda: self._handle_readiness_error(str(exc)))

    def _calculate_readiness_recommendation(self, input_path, external_bom_df, progress_callback=None):
        if read_balance_workbook_tables is None:
            raise RuntimeError('MRP 数据读取引擎不可用')

        if progress_callback:
            progress_callback('步骤 1/4：读取库存、未清PO、未转PR和外采清单...')
        frames = read_balance_workbook_tables(Path(input_path), external_bom_df=external_bom_df)

        def _sum_map(df, key_candidates, value_candidates, *, contains=False):
            key_col = find_first_matching_column(df, key_candidates)
            value_col = find_first_matching_column(df, value_candidates, contains=contains)
            temp = df.assign(
                _key=df[key_col].astype(str).str.strip(),
                _value=pd.to_numeric(df[value_col], errors='coerce').fillna(0),
            )
            return temp.groupby('_key', dropna=False)['_value'].sum().to_dict()

        purchase = frames['采购数据'].copy()
        purchase_item_col = find_first_matching_column(purchase, ['物料号'])
        purchase_name_col = find_first_matching_column(purchase, ['名称'], required=False)
        purchase_spec_col = find_first_matching_column(purchase, ['规格'], required=False)
        purchased_codes = {
            str(value).strip()
            for value in purchase[purchase_item_col].tolist()
            if str(value).strip()
        }
        purchase_name_map = {}
        purchase_spec_map = {}
        for record in purchase.to_dict('records'):
            code = str(record.get(purchase_item_col, '') or '').strip()
            if not code:
                continue
            if purchase_name_col and code not in purchase_name_map:
                purchase_name_map[code] = str(record.get(purchase_name_col, '') or '').strip()
            if purchase_spec_col and code not in purchase_spec_map:
                purchase_spec_map[code] = str(record.get(purchase_spec_col, '') or '').strip()

        inventory_map = _sum_map(
            frames['期初库存'],
            ['物料编码'],
            ['库存量'],
        )
        inbound_po_map = _sum_map(
            frames['在途采购'],
            ['料号'],
            ['欠交数量'],
            contains=True,
        )
        inbound_pr_map = _sum_map(
            frames['在途请购'],
            ['料号'],
            ['未转PO数量'],
            contains=True,
        )

        substitute_code_map = defaultdict(list)
        if read_balance_substitute_rules is not None:
            try:
                substitute_rules = read_balance_substitute_rules(Path(input_path))
                if substitute_rules is not None and not substitute_rules.empty:
                    for before_code, group_df in substitute_rules.groupby('before_code', sort=False):
                        before_code = str(before_code).strip()
                        for after_code in group_df.sort_values('sort_order')['after_code'].astype(str).str.strip():
                            if after_code and after_code not in substitute_code_map[before_code]:
                                substitute_code_map[before_code].append(after_code)
                            if len(substitute_code_map[before_code]) >= 2:
                                break
            except Exception as exc:
                self._append_readiness_log(f'替代料读取失败，库存推荐暂不计替代库存: {exc}')

        if progress_callback:
            progress_callback('步骤 2/4：整理当前 BOM 层级关系...')
        edges = defaultdict(list)
        material_name_map = {}
        material_spec_map = {}
        for parent_code, rows in self.bom_index.items():
            parent_code = normalize_material_code(parent_code)
            if not parent_code:
                continue
            for row in rows:
                if len(row) <= max(CHILD_PN_COL, CHILD_QTY_COL):
                    continue
                child_code = normalize_material_code(row[CHILD_PN_COL])
                if not child_code:
                    continue
                try:
                    usage = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] not in (None, '') else 1.0
                except (TypeError, ValueError):
                    usage = 1.0
                if usage <= 0:
                    continue
                parent_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                parent_spec = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                child_name = str(row[CHILD_NAME_COL]).strip() if len(row) > CHILD_NAME_COL and row[CHILD_NAME_COL] else ''
                child_spec = str(row[CHILD_SPEC_COL]).strip() if len(row) > CHILD_SPEC_COL and row[CHILD_SPEC_COL] else ''
                if parent_name:
                    material_name_map.setdefault(parent_code, parent_name)
                if parent_spec:
                    material_spec_map.setdefault(parent_code, parent_spec)
                if child_name:
                    material_name_map.setdefault(child_code, child_name)
                if child_spec:
                    material_spec_map.setdefault(child_code, child_spec)
                edges[parent_code].append((child_code, usage))

        for code, name in purchase_name_map.items():
            if name:
                material_name_map.setdefault(code, name)
        for code, spec in purchase_spec_map.items():
            if spec:
                material_spec_map.setdefault(code, spec)

        unit_cache = {}
        unknown_cache = {}
        unknown_detail_cache = {}

        def _explode_unit(root_code):
            root_code = normalize_material_code(root_code)
            if root_code in unit_cache:
                return unit_cache[root_code], unknown_cache.get(root_code, 0), unknown_detail_cache.get(root_code, '')
            totals = defaultdict(float)
            unknown_count = 0
            unknown_details = []

            def _add_unknown(code, parent_code, reason):
                nonlocal unknown_count
                code = normalize_material_code(code)
                if not code:
                    return
                unknown_count += 1
                name = material_name_map.get(code, '')
                spec = material_spec_map.get(code, '')
                parent_text = normalize_material_code(parent_code)
                label = code
                if name:
                    label += f' {name}'
                if spec:
                    label += f' {spec}'
                if parent_text:
                    label += f'；上层{parent_text}'
                if reason:
                    label += f'；{reason}'
                unknown_details.append(label)

            def _walk(code, qty, trail):
                children = edges.get(code, [])
                if not children:
                    if code in purchased_codes:
                        totals[code] += qty
                    else:
                        _add_unknown(code, '', '不在外采清单且无下层BOM')
                    return
                for child_code, usage in children:
                    child_qty = qty * float(usage or 0)
                    if child_qty <= 0:
                        continue
                    if child_code in purchased_codes:
                        totals[child_code] += child_qty
                    elif child_code in trail:
                        _add_unknown(child_code, code, 'BOM循环引用')
                    elif child_code in edges:
                        _walk(child_code, child_qty, trail | {child_code})
                    else:
                        _add_unknown(child_code, code, '不在外采清单且无下层BOM')

            _walk(root_code, 1.0, {root_code})
            unit_cache[root_code] = dict(totals)
            unknown_cache[root_code] = unknown_count
            unknown_detail_cache[root_code] = '；'.join(dict.fromkeys(unknown_details))
            return unit_cache[root_code], unknown_count, unknown_detail_cache[root_code]

        if progress_callback:
            progress_callback('步骤 3/4：按单台外购物料用量倒推可生产数量...')
        scheduled_roots = []
        try:
            production_orders = frames.get('生产订单')
            if production_orders is not None and not production_orders.empty:
                production_col = find_first_matching_column(production_orders, ['母件料号'])
                seen_roots = set()
                for value in production_orders[production_col].tolist():
                    code = normalize_material_code(value)
                    if code and code not in seen_roots:
                        seen_roots.add(code)
                        scheduled_roots.append(code)
        except Exception as exc:
            self._append_readiness_log(f'排产母料号读取失败，库存推荐改为扫描全部BOM母件: {exc}')

        recommend_rows = []
        if scheduled_roots:
            candidate_roots = [code for code in scheduled_roots if code in edges and code not in purchased_codes]
        else:
            candidate_roots = sorted(code for code in edges.keys() if code and code not in purchased_codes)
        for index, root_code in enumerate(candidate_roots, start=1):
            unit_totals, unknown_count, unknown_text = _explode_unit(root_code)
            if not unit_totals:
                continue
            stock_limits = []
            position_limits = []
            issue_count = 0
            gap_for_one = 0.0
            bottlenecks = []
            for material_code, unit_qty in unit_totals.items():
                if unit_qty <= 0:
                    continue
                substitute_stock = sum(float(inventory_map.get(sub_code, 0) or 0) for sub_code in substitute_code_map.get(material_code, [])[:2])
                stock_qty = float(inventory_map.get(material_code, 0) or 0) + substitute_stock
                position_qty = stock_qty + float(inbound_po_map.get(material_code, 0) or 0) + float(inbound_pr_map.get(material_code, 0) or 0)
                stock_can_make = math.floor(stock_qty / unit_qty) if unit_qty > 0 else 0
                position_can_make = math.floor(position_qty / unit_qty) if unit_qty > 0 else 0
                stock_limits.append(stock_can_make)
                position_limits.append(position_can_make)
                gap = max(unit_qty - stock_qty, 0.0)
                if gap > 1e-9:
                    issue_count += 1
                    gap_for_one += gap
                bottlenecks.append({
                    'code': material_code,
                    'name': material_name_map.get(material_code, ''),
                    'stock_can_make': stock_can_make,
                    'position_can_make': position_can_make,
                    'gap': gap,
                })
            if not stock_limits:
                continue
            stock_buildable = int(min(stock_limits))
            position_buildable = int(min(position_limits)) if position_limits else 0
            bottlenecks = sorted(
                bottlenecks,
                key=lambda item: (item['stock_can_make'], item['position_can_make'], -item['gap'], item['code']),
            )[:3]
            short_text = '；'.join(
                f"{item['code']} {item['name']} 现货可做{fmt_qty(item['stock_can_make'])}台"
                for item in bottlenecks
            )
            if unknown_count:
                conclusion = f'有{unknown_count}个未识别下层，需先查BOM/外采'
            elif stock_buildable >= 1:
                conclusion = f'现货可生产{fmt_qty(stock_buildable)}台'
            elif position_buildable >= 1:
                conclusion = f'现货不足，库存位置可生产{fmt_qty(position_buildable)}台'
            elif issue_count <= 3:
                conclusion = '缺料少，可优先协调'
            else:
                conclusion = '缺料较多'
            recommend_rows.append({
                '母料号': root_code,
                '料品名称': material_name_map.get(root_code, ''),
                '规格': material_spec_map.get(root_code, ''),
                '外购物料数': len(unit_totals),
                '现货可生产': stock_buildable,
                '库存位置可生产': position_buildable,
                '做1台缺料数': issue_count,
                '做1台缺口': gap_for_one,
                '短板物料': short_text,
                '未识别下层': unknown_text,
                '结论': conclusion,
            })
            if progress_callback and index % 800 == 0:
                progress_callback(f'已扫描 {index} 个母件...')

        if progress_callback:
            progress_callback('步骤 4/4：排序推荐结果...')
        recommend_rows.sort(
            key=lambda row: (
                -float(row.get('现货可生产', 0) or 0),
                int(row.get('做1台缺料数', 0) or 0),
                float(row.get('做1台缺口', 0) or 0),
                -float(row.get('库存位置可生产', 0) or 0),
                str(row.get('母料号', '')),
            )
        )
        return recommend_rows

    def _calculate_common_material_analysis(self, input_path, external_bom_df, progress_callback=None):
        if read_balance_workbook_tables is None:
            raise RuntimeError('MRP 数据读取引擎不可用')

        def _clean_text(value):
            if value is None:
                return ''
            try:
                if pd.isna(value):
                    return ''
            except (TypeError, ValueError):
                pass
            text = str(value).strip()
            return '' if text.lower() == 'nan' else text

        if progress_callback:
            progress_callback('步骤 1/4：读取MRP、库存、未清PO、未转PR和外采清单...')
        frames = read_balance_workbook_tables(Path(input_path), external_bom_df=external_bom_df)

        def _sum_map(df, key_candidates, value_candidates, *, contains=False):
            key_col = find_first_matching_column(df, key_candidates)
            value_col = find_first_matching_column(df, value_candidates, contains=contains)
            temp = df.assign(
                _key=df[key_col].astype(str).str.strip(),
                _value=pd.to_numeric(df[value_col], errors='coerce').fillna(0),
            )
            return temp.groupby('_key', dropna=False)['_value'].sum().to_dict()

        purchase = frames['采购数据'].copy()
        purchase_item_col = find_first_matching_column(purchase, ['物料号'])
        supplier_col = find_first_matching_column(purchase, ['供应商'], required=False)
        buyer_col = find_first_matching_column(purchase, ['采购'], required=False)
        purchase_name_col = find_first_matching_column(purchase, ['名称'], required=False)
        purchase_spec_col = find_first_matching_column(purchase, ['规格'], required=False)
        purchased_codes = {
            str(value).strip()
            for value in purchase[purchase_item_col].tolist()
            if str(value).strip()
        }
        purchase_meta = {}
        for record in purchase.to_dict('records'):
            code = str(record.get(purchase_item_col, '') or '').strip()
            if not code or code in purchase_meta:
                continue
            purchase_meta[code] = {
                'name': _clean_text(record.get(purchase_name_col, '')) if purchase_name_col else '',
                'spec': _clean_text(record.get(purchase_spec_col, '')) if purchase_spec_col else '',
                'supplier': _clean_text(record.get(supplier_col, '')) if supplier_col else '',
                'buyer': _clean_text(record.get(buyer_col, '')) if buyer_col else '',
            }

        inventory_map = _sum_map(frames['期初库存'], ['物料编码'], ['库存量'])
        inbound_po_map = _sum_map(frames['在途采购'], ['料号'], ['欠交数量'], contains=True)
        inbound_pr_map = _sum_map(frames['在途请购'], ['料号'], ['未转PO数量'], contains=True)

        if progress_callback:
            progress_callback('步骤 2/4：整理当前BOM并展开排产母料...')
        edges = defaultdict(list)
        material_name_map = {}
        material_spec_map = {}
        for parent_code, rows in self.bom_index.items():
            parent_code = normalize_material_code(parent_code)
            if not parent_code:
                continue
            for row in rows:
                if len(row) <= max(CHILD_PN_COL, CHILD_QTY_COL):
                    continue
                child_code = normalize_material_code(row[CHILD_PN_COL])
                if not child_code:
                    continue
                try:
                    usage = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] not in (None, '') else 1.0
                except (TypeError, ValueError):
                    usage = 1.0
                if usage <= 0:
                    continue
                parent_name = _clean_text(row[2]) if len(row) > 2 else ''
                parent_spec = _clean_text(row[3]) if len(row) > 3 else ''
                child_name = _clean_text(row[CHILD_NAME_COL]) if len(row) > CHILD_NAME_COL else ''
                child_spec = _clean_text(row[CHILD_SPEC_COL]) if len(row) > CHILD_SPEC_COL else ''
                if parent_name:
                    material_name_map.setdefault(parent_code, parent_name)
                if parent_spec:
                    material_spec_map.setdefault(parent_code, parent_spec)
                if child_name:
                    material_name_map.setdefault(child_code, child_name)
                if child_spec:
                    material_spec_map.setdefault(child_code, child_spec)
                edges[parent_code].append((child_code, usage))
        for code, meta in purchase_meta.items():
            if meta.get('name'):
                material_name_map.setdefault(code, meta['name'])
            if meta.get('spec'):
                material_spec_map.setdefault(code, meta['spec'])

        unit_cache = {}
        def _explode_unit(root_code):
            root_code = normalize_material_code(root_code)
            if root_code in unit_cache:
                return unit_cache[root_code]
            totals = defaultdict(float)
            def _walk(code, qty, trail):
                children = edges.get(code, [])
                if not children:
                    if code in purchased_codes:
                        totals[code] += qty
                    return
                for child_code, usage in children:
                    child_qty = qty * float(usage or 0)
                    if child_qty <= 0:
                        continue
                    if child_code in purchased_codes:
                        totals[child_code] += child_qty
                    elif child_code in trail:
                        continue
                    elif child_code in edges:
                        _walk(child_code, child_qty, trail | {child_code})
            _walk(root_code, 1.0, {root_code})
            unit_cache[root_code] = dict(totals)
            return unit_cache[root_code]

        production_orders = frames['生产订单'].copy()
        production_col = find_first_matching_column(production_orders, ['母件料号'])
        production_name_col = find_first_matching_column(production_orders, ['母件品名'], required=False)
        production_spec_col = find_first_matching_column(production_orders, ['母件规格'], required=False)
        production_qty_col = find_first_matching_column(production_orders, ['上线数量'])
        production_project_col = find_first_matching_column(production_orders, ['客户'], required=False)

        root_summary = {}
        for record in production_orders.to_dict('records'):
            root_code = normalize_material_code(record.get(production_col))
            if not root_code or root_code not in edges:
                continue
            try:
                qty = float(record.get(production_qty_col, 0) or 0)
            except (TypeError, ValueError):
                qty = 0.0
            if qty <= 0:
                continue
            item = root_summary.setdefault(
                root_code,
                {
                    'qty': 0.0,
                    'name': _clean_text(record.get(production_name_col, '')) if production_name_col else material_name_map.get(root_code, ''),
                    'spec': _clean_text(record.get(production_spec_col, '')) if production_spec_col else material_spec_map.get(root_code, ''),
                    'projects': set(),
                },
            )
            item['qty'] += qty
            if production_project_col:
                project = _clean_text(record.get(production_project_col, ''))
                if project:
                    item['projects'].add(project)

        if progress_callback:
            progress_callback(f'步骤 3/4：统计 {len(root_summary)} 个排产母料的外购通用次数...')
        material_usage = {}
        for index, (root_code, root_info) in enumerate(root_summary.items(), start=1):
            unit_totals = _explode_unit(root_code)
            if not unit_totals:
                continue
            root_qty = float(root_info.get('qty', 0) or 0)
            root_label = root_code
            root_name = str(root_info.get('name', '') or '').strip()
            root_spec = str(root_info.get('spec', '') or '').strip()
            if root_name:
                root_label += f' {root_name}'
            if root_spec:
                root_label += f' {root_spec}'
            for material_code, unit_qty in unit_totals.items():
                item = material_usage.setdefault(
                    material_code,
                    {
                        'roots': set(),
                        'root_labels': {},
                        'demand': 0.0,
                        'projects': set(),
                    },
                )
                item['roots'].add(root_code)
                item['root_labels'][root_code] = root_label
                item['demand'] += float(unit_qty or 0) * root_qty
                item['projects'].update(root_info.get('projects', set()))
            if progress_callback and index % 800 == 0:
                progress_callback(f'已展开 {index} 个排产母料...')

        rows = []
        for material_code, info in material_usage.items():
            meta = purchase_meta.get(material_code, {})
            demand = float(info.get('demand', 0) or 0)
            stock = float(inventory_map.get(material_code, 0) or 0)
            po = float(inbound_po_map.get(material_code, 0) or 0)
            pr = float(inbound_pr_map.get(material_code, 0) or 0)
            position = stock + po + pr
            root_count = len(info.get('roots', set()))
            if root_count >= 8 and position < demand:
                risk = '高通用高风险'
            elif root_count >= 8:
                risk = '高通用'
            elif root_count >= 3 and position < demand:
                risk = '中通用有缺口'
            elif position < demand:
                risk = '低通用有缺口'
            else:
                risk = '供给可覆盖'
            root_labels = [info['root_labels'][code] for code in sorted(info.get('roots', set())) if code in info['root_labels']]
            rows.append({
                '物料编码': material_code,
                '物料名称': meta.get('name') or material_name_map.get(material_code, ''),
                '规格': meta.get('spec') or material_spec_map.get(material_code, ''),
                '采购': meta.get('buyer', ''),
                '供应商': meta.get('supplier', ''),
                '使用母料数': root_count,
                '使用母料清单': '；'.join(root_labels[:30]) + ('；...' if len(root_labels) > 30 else ''),
                '总需求': demand,
                '当前库存': stock,
                '未清PO': po,
                '未转PR': pr,
                '库存位置': position,
                '通用风险等级': risk,
            })

        if progress_callback:
            progress_callback('步骤 4/4：排序通用物料排行...')
        rows.sort(
            key=lambda row: (
                -int(row.get('使用母料数', 0) or 0),
                -float(row.get('总需求', 0) or 0),
                str(row.get('物料编码', '')),
            )
        )
        for index, row in enumerate(rows, start=1):
            row['排名'] = index
        return rows

    def _finish_common_material_analysis(self, rows):
        self._readiness_running = False
        self.readiness_progress['value'] = 100
        self.readiness_progress.grid_remove()
        self.readiness_run_btn.state(['!disabled'])
        self.readiness_recommend_btn.state(['!disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['!disabled'])

        self._clear_tree_rows(self.readiness_common_tree)
        for record in rows:
            values = [
                record.get('排名', ''),
                record.get('物料编码', ''),
                record.get('物料名称', ''),
                record.get('规格', ''),
                record.get('采购', ''),
                record.get('供应商', ''),
                fmt_qty(record.get('使用母料数', 0)),
                record.get('使用母料清单', ''),
                fmt_qty(record.get('总需求', 0)),
                fmt_qty(record.get('当前库存', 0)),
                fmt_qty(record.get('未清PO', 0)),
                fmt_qty(record.get('未转PR', 0)),
                fmt_qty(record.get('库存位置', 0)),
                record.get('通用风险等级', ''),
            ]
            self.readiness_common_tree.insert('', 'end', values=values)
        top_text = ''
        if rows:
            top = rows[0]
            top_text = f" | TOP1 {top.get('物料编码', '')} 使用母料 {fmt_qty(top.get('使用母料数', 0))} 个"
        high_risk = sum(1 for row in rows if row.get('通用风险等级') == '高通用高风险')
        summary = f'通用物料分析完成 | 外购物料 {len(rows)} 个 | 高通用高风险 {high_risk} 个{top_text}'
        self.readiness_summary_var.set(summary)
        self.readiness_root_stat_var.set(f'通用料: {len(rows)}')
        self.readiness_material_stat_var.set(f'高风险: {high_risk}')
        self.readiness_issue_stat_var.set(f'TOP母料数: {fmt_qty(rows[0].get("使用母料数", 0)) if rows else 0}')
        self.readiness_ready_stat_var.set('排行完成')
        self.status_var.set(summary)
        self._append_readiness_log(summary)

    def _finish_readiness_recommendation(self, rows):
        self._readiness_running = False
        self.readiness_progress['value'] = 100
        self.readiness_progress.grid_remove()
        self.readiness_run_btn.state(['!disabled'])
        self.readiness_recommend_btn.state(['!disabled'])

        self._clear_tree_rows(self.readiness_recommend_tree)
        for record in rows:
            values = [
                record.get('母料号', ''),
                record.get('料品名称', ''),
                record.get('规格', ''),
                fmt_qty(record.get('外购物料数', 0)),
                fmt_qty(record.get('现货可生产', 0)),
                fmt_qty(record.get('库存位置可生产', 0)),
                fmt_qty(record.get('做1台缺料数', 0)),
                fmt_qty(record.get('做1台缺口', 0)),
                record.get('短板物料', ''),
                record.get('未识别下层', ''),
                record.get('结论', ''),
            ]
            self.readiness_recommend_tree.insert('', 'end', values=values)

        stock_ready = sum(1 for row in rows if float(row.get('现货可生产', 0) or 0) >= 1)
        position_ready = sum(1 for row in rows if float(row.get('库存位置可生产', 0) or 0) >= 1)
        low_gap = sum(
            1 for row in rows
            if float(row.get('现货可生产', 0) or 0) < 1 and int(row.get('做1台缺料数', 0) or 0) <= 3
        )
        summary = f'库存推荐完成 | 扫描型号 {len(rows)} 个 | 现货可做 {stock_ready} 个 | 库存位置可做 {position_ready} 个'
        self.readiness_summary_var.set(summary)
        self.readiness_root_stat_var.set(f'推荐型号: {len(rows)}')
        self.readiness_material_stat_var.set(f'现货可做: {stock_ready}')
        self.readiness_issue_stat_var.set(f'缺料少: {low_gap}')
        self.readiness_ready_stat_var.set(f'库存位置可做: {position_ready}')
        self.status_var.set(summary)
        self._append_readiness_log(summary)

    def _run_readiness_analysis(self):
        if run_purchase_readiness_analysis is None:
            show_data_error(
                self.root,
                title='齐套分析引擎不可用',
                summary='未能加载齐套分析引擎（mrp_balance_tool），无法运行分析。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='运行源码请确认同级目录有 "静态平衡表/src"；打包后请确认 spec 文件正确包含 mrp_balance_tool 模块。',
            )
            return
        if not self.bom_index:
            messagebox.showwarning('缺少数据', '请先上传 BOM 文件')
            return
        if self._readiness_running:
            self.status_var.set('外购齐套分析正在执行中，请稍候')
            return

        input_path = self.readiness_input_var.get().strip()
        reply_source_paths = list(self.readiness_reply_source_paths)
        if not input_path:
            messagebox.showwarning('参数不完整', '请先选择 MRP 输入文件')
            return
        try:
            root_items = self._parse_readiness_root_items()
        except Exception as exc:
            messagebox.showwarning('输入有误', str(exc))
            return

        external_bom_df = self.export_bom_dataframe()
        if not self._run_input_precheck_dialog(input_path, external_bom_df, log_callback=self._append_readiness_log):
            self.readiness_summary_var.set('已取消，请先检查源数据')
            self.status_var.set(self.readiness_summary_var.get())
            return
        self._readiness_running = True
        self.readiness_run_btn.state(['disabled'])
        if hasattr(self, 'readiness_recommend_btn'):
            self.readiness_recommend_btn.state(['disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['disabled'])
        self.readiness_progress.grid()
        self._readiness_stage_counter = 0
        self.readiness_progress['value'] = 0
        self.readiness_summary_var.set('正在分析外购物料齐套情况，请稍候...')
        self.status_var.set(self.readiness_summary_var.get())
        self._append_readiness_log('开始分析外购物料齐套情况...')

        self._readiness_thread = threading.Thread(
            target=self._run_readiness_analysis_worker,
            args=(input_path, external_bom_df, root_items, reply_source_paths),
            daemon=True,
        )
        self._readiness_thread.start()

    def _run_readiness_analysis_worker(self, input_path, external_bom_df, root_items, reply_source_paths):
        try:
            result = run_purchase_readiness_analysis(
                Path(input_path),
                root_items,
                external_bom_df=external_bom_df,
                carry_forward_paths=[Path(path) for path in reply_source_paths],
                progress_callback=lambda message: self.root.after(
                    0,
                    lambda msg=message: self._report_readiness_progress(msg),
                ),
            )
            self.root.after(0, lambda: self._finish_readiness_analysis(result))
        except Exception as exc:
            self.root.after(0, lambda: self._handle_readiness_error(str(exc)))

    def _report_readiness_progress(self, message):
        self._readiness_stage_counter += 1
        progress = min(95, self._readiness_stage_counter * 6)
        try:
            self.readiness_progress['value'] = progress
        except Exception:
            pass
        stage_label = f'[阶段 {self._readiness_stage_counter}] '
        self.readiness_summary_var.set(stage_label + message)
        self.status_var.set(stage_label + message)
        self._append_readiness_log(message)

    def _finish_readiness_analysis(self, result):
        self._readiness_running = False
        self.readiness_progress['value'] = 100
        self.readiness_progress.grid_remove()
        self.readiness_run_btn.state(['!disabled'])
        if hasattr(self, 'readiness_recommend_btn'):
            self.readiness_recommend_btn.state(['!disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['!disabled'])

        self._clear_tree_rows(self.readiness_root_tree)
        if hasattr(self, 'readiness_producible_tree'):
            self._clear_tree_rows(self.readiness_producible_tree)
        if hasattr(self, 'readiness_rolling_matrix_tree'):
            self._clear_tree_rows(self.readiness_rolling_matrix_tree)
        if hasattr(self, 'readiness_diff_capacity_tree'):
            self._clear_tree_rows(self.readiness_diff_capacity_tree)
        if hasattr(self, 'readiness_horizontal_tree'):
            self._clear_tree_rows(self.readiness_horizontal_tree)
        if hasattr(self, 'readiness_recommend_tree'):
            self._clear_tree_rows(self.readiness_recommend_tree)
        self._clear_tree_rows(self.readiness_issue_tree)
        self._clear_tree_rows(self.readiness_material_tree)
        self._clear_tree_rows(self.readiness_unknown_tree)

        for record in getattr(result, 'root_summary_df', pd.DataFrame()).to_dict('records'):
            values = [
                record.get('母料号', ''),
                record.get('料品名称', ''),
                fmt_qty(record.get('输入数量', 0)),
                fmt_qty(record.get('外购物料数', 0)),
                fmt_qty(record.get('问题物料数', 0)),
                fmt_qty(record.get('未识别物料数', 0)),
                '' if pd.isna(record.get('外购齐套日期')) else str(record.get('外购齐套日期')),
                record.get('结论', ''),
            ]
            self.readiness_root_tree.insert('', 'end', values=values)

        if hasattr(self, 'readiness_producible_tree'):
            for record in getattr(result, 'producible_df', pd.DataFrame()).to_dict('records'):
                values = [
                    record.get('母料号', ''),
                    record.get('料品名称', ''),
                    fmt_qty(record.get('输入数量', 0)),
                    record.get('排产顺序', ''),
                    record.get('排产日期', ''),
                    record.get('日期', ''),
                    fmt_qty(record.get('可生产数量', 0)),
                    fmt_qty(record.get('当前库存可生成', 0)),
                    fmt_qty(record.get('距离输入缺口', 0)),
                    record.get('瓶颈物料', ''),
                    record.get('瓶颈物料名称', ''),
                    record.get('瓶颈单台用量', ''),
                    record.get('瓶颈可用量', ''),
                    record.get('瓶颈原因', ''),
                ]
                self.readiness_producible_tree.insert('', 'end', values=values)

        if hasattr(self, 'readiness_rolling_matrix_tree'):
            rolling_df = getattr(result, 'rolling_matrix_df', pd.DataFrame())
            if isinstance(rolling_df, pd.DataFrame) and not rolling_df.empty:
                headers = list(rolling_df.columns)
                self._configure_result_tree_columns(self.readiness_rolling_matrix_tree, headers)
                for record in rolling_df.to_dict('records'):
                    values = [
                        fmt_qty(record.get(header, 0)) if any(keyword in header for keyword in ('数量', '需求', '可用', '可生产', '缺口', '物料数', '总计', '累计', '用量', '库存', '扣减', '分配', '最大', '最小', '差值')) and record.get(header, '') != '' else record.get(header, '')
                        for header in headers
                    ]
                    self.readiness_rolling_matrix_tree.insert('', 'end', values=values)
            else:
                self._configure_result_tree_columns(
                    self.readiness_rolling_matrix_tree,
                    READINESS_ROLLING_MATRIX_HEADERS,
                    READINESS_ROLLING_MATRIX_WIDTHS,
                )

        if hasattr(self, 'readiness_diff_capacity_tree'):
            for record in getattr(result, 'diff_capacity_df', pd.DataFrame()).to_dict('records'):
                values = [
                    record.get('排产顺序', ''),
                    record.get('排产日期', ''),
                    record.get('母料号', ''),
                    record.get('料品名称', ''),
                    fmt_qty(record.get('输入数量', 0)),
                    fmt_qty(record.get('差异物料数', 0)),
                    fmt_qty(record.get('通用物料数', 0)),
                    record.get('通用物料清单', ''),
                    fmt_qty(record.get('当前库存差异可生产', 0)),
                    fmt_qty(record.get('当前库存差异缺口', 0)),
                    fmt_qty(record.get('当前可用差异可生产', 0)),
                    fmt_qty(record.get('当前可用差异缺口', 0)),
                    record.get('瓶颈差异物料', ''),
                    record.get('瓶颈物料名称', ''),
                    record.get('瓶颈单台差异用量', ''),
                    record.get('瓶颈库存可用', ''),
                    record.get('瓶颈当前可用', ''),
                    record.get('BOM差异说明', ''),
                ]
                self.readiness_diff_capacity_tree.insert('', 'end', values=values)

        if hasattr(self, 'readiness_horizontal_tree'):
            for record in getattr(result, 'horizontal_shortage_df', pd.DataFrame()).to_dict('records'):
                values = [
                    record.get('排产顺序', ''),
                    record.get('排产日期', ''),
                    record.get('母料号', ''),
                    record.get('料品名称', ''),
                    fmt_qty(record.get('输入数量', 0)),
                    record.get('料号', ''),
                    record.get('物料名称', ''),
                    fmt_qty(record.get('单台用量', 0)),
                    fmt_qty(record.get('本行需求', 0)),
                    fmt_qty(record.get('库存分配前', 0)),
                    fmt_qty(record.get('库存扣减', 0)),
                    fmt_qty(record.get('库存缺口', 0)),
                    fmt_qty(record.get('库存分配后', 0)),
                    fmt_qty(record.get('当前可用分配前', 0)),
                    fmt_qty(record.get('当前可用扣减', 0)),
                    fmt_qty(record.get('当前可用缺口', 0)),
                    fmt_qty(record.get('当前可用分配后', 0)),
                    record.get('替代料清单', ''),
                    record.get('BOM差异标识', ''),
                    fmt_qty(record.get('共用母料数', 0)),
                    record.get('BOM差异说明', ''),
                ]
                self.readiness_horizontal_tree.insert('', 'end', values=values)

        for record in getattr(result, 'issue_df', pd.DataFrame()).to_dict('records'):
            values = [
                record.get('母料号', ''),
                fmt_qty(record.get('母件输入数量', 0)),
                record.get('料号', ''),
                fmt_qty(record.get('本母件需求', 0)),
                record.get('料品名称', ''),
                record.get('规格', ''),
                record.get('上层物料编码', ''),
                record.get('供应商', ''),
                record.get('采购', ''),
                fmt_qty(record.get('当前可用', 0)),
                fmt_qty(record.get('当前缺口', 0)),
                fmt_qty(record.get('当前可用缺口', 0)),
                record.get('替代料清单', ''),
                fmt_qty(record.get('采购答复累计', 0)),
                '' if pd.isna(record.get('最早齐套日期')) else str(record.get('最早齐套日期')),
                record.get('问题原因', ''),
            ]
            self.readiness_issue_tree.insert('', 'end', values=values)

        for record in getattr(result, 'material_df', pd.DataFrame()).to_dict('records'):
            values = [
                record.get('料号', ''),
                record.get('料品名称', ''),
                record.get('规格', ''),
                record.get('供应商', ''),
                record.get('采购', ''),
                fmt_qty(record.get('总需求', 0)),
                fmt_qty(record.get('当前库存', record.get('实时库存', 0))),
                fmt_qty(record.get('未清PO', 0)),
                fmt_qty(record.get('未转', 0)),
                fmt_qty(record.get('替代料库存', 0)),
                fmt_qty(record.get('替代料未清PO', 0)),
                fmt_qty(record.get('替代料未转', 0)),
                fmt_qty(record.get('当前库存+替代库存', 0)),
                record.get('替代料清单', ''),
                fmt_qty(record.get('当前可用', 0)),
                fmt_qty(record.get('当前缺口', 0)),
                fmt_qty(record.get('当前可用缺口', 0)),
                fmt_qty(record.get('采购答复累计', 0)),
                '' if pd.isna(record.get('最早齐套日期')) else str(record.get('最早齐套日期')),
                record.get('状态', ''),
                record.get('来源母料号', ''),
                record.get('上层物料编码', ''),
                record.get('问题原因', ''),
            ]
            self.readiness_material_tree.insert('', 'end', values=values)

        missing_root_df = getattr(result, 'missing_root_df', pd.DataFrame())
        for record in missing_root_df.to_dict('records'):
            values = [
                record.get('母料号', ''),
                fmt_qty(record.get('输入数量', 0)),
                '',
                record.get('料品名称', ''),
                '',
                '',
                '',
                record.get('问题原因', ''),
            ]
            self.readiness_unknown_tree.insert('', 'end', values=values)

        unknown_leaf_df = getattr(result, 'unknown_leaf_df', pd.DataFrame())
        for record in unknown_leaf_df.to_dict('records'):
            values = [
                record.get('母料号', ''),
                fmt_qty(record.get('母件输入数量', 0)),
                record.get('料号', ''),
                record.get('料品名称', ''),
                record.get('规格', ''),
                record.get('上层物料编码', ''),
                fmt_qty(record.get('需求数量', 0)),
                record.get('问题原因', ''),
            ]
            self.readiness_unknown_tree.insert('', 'end', values=values)

        batch_summary = getattr(result, 'batch_summary', {}) or {}
        root_count = int(batch_summary.get('root_count', 0) or 0)
        material_count = int(batch_summary.get('material_count', 0) or 0)
        issue_count = int(batch_summary.get('issue_count', 0) or 0)
        ready_date = batch_summary.get('ready_date')
        conclusion = batch_summary.get('conclusion', '分析完成')
        summary = f'分析完成 | 母料号 {root_count} 个 | 外购物料 {material_count} 个 | 结论：{conclusion}'
        self.readiness_summary_var.set(summary)
        self.readiness_root_stat_var.set(f'母料号: {root_count}')
        self.readiness_material_stat_var.set(f'外购物料: {material_count}')
        self.readiness_issue_stat_var.set(f'问题物料: {issue_count}')
        self.readiness_ready_stat_var.set(f'最晚齐套: {ready_date or "-"}')
        self.status_var.set(summary)
        self._append_readiness_log(summary)

        carried_count = int(getattr(result, 'carried_reply_cell_count', 0) or 0)
        carried_material_count = int(getattr(result, 'carried_reply_material_count', 0) or 0)
        carried_file_count = int(getattr(result, 'carried_reply_file_count', 0) or 0)
        if carried_count:
            self._append_readiness_log(f'已带入采购答复 {carried_count} 格，覆盖料号 {carried_material_count} 个，来源文件 {carried_file_count} 个')

    def _handle_readiness_error(self, message):
        self._readiness_running = False
        self.readiness_progress['value'] = 0
        self.readiness_progress.grid_remove()
        self.readiness_run_btn.state(['!disabled'])
        if hasattr(self, 'readiness_recommend_btn'):
            self.readiness_recommend_btn.state(['!disabled'])
        if hasattr(self, 'readiness_common_btn'):
            self.readiness_common_btn.state(['!disabled'])
        self.readiness_summary_var.set('外购齐套分析失败')
        self.status_var.set('外购齐套分析失败')
        self._append_readiness_log(f'分析失败: {message}')
        show_data_error(
            self.root,
            title='外购齐套分析失败',
            summary='外购物料齐套分析未能完成。',
            detail=message,
            fix_hint=_suggest_fix_for_exception(Exception(message)),
        )

    def _build_material_buy_tab(self):
        f = self.tab_material_buy
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='单料采购判断参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(1, weight=1)
        control.columnconfigure(4, weight=1)

        self.material_buy_issue_var = tk.StringVar()
        self.material_buy_stock_var = tk.StringVar()
        self.material_buy_po_var = tk.StringVar()
        self.material_buy_pr_var = tk.StringVar()
        self.material_buy_usage_var = tk.StringVar()
        self.material_buy_code_var = tk.StringVar()
        self.material_buy_doc_var = tk.StringVar()
        self.material_buy_month_var = tk.StringVar(value=date.today().strftime('%Y-%m'))
        self.material_buy_query_months_var = tk.IntVar(value=3)
        self.material_buy_summary_var = tk.StringVar(value='等待分析')
        self.material_buy_rows = []
        self.material_buy_usage_cache = {}
        self.material_buy_headers, self.material_buy_widths = self._material_buy_result_columns(
            self._material_buy_month_labels(date.today().year, date.today().month, 3)
        )

        def _file_row(row, label, var, button_text, command):
            ttk.Label(control, text=label, style='Section.TLabel').grid(row=row, column=0, sticky='w', padx=(0, 6), pady=3)
            path_label, _ = make_path_label(control, var)
            path_label.grid(row=row, column=1, columnspan=3, sticky='ew', pady=3)
            ttk.Button(control, text=button_text, command=command).grid(row=row, column=4, sticky='ew', padx=(8, 0), pady=3)

        _file_row(0, '收发记录', self.material_buy_issue_var, '选择收发记录', self._choose_material_buy_issue_file)
        _file_row(1, '库存明细', self.material_buy_stock_var, '选择库存明细', self._choose_material_buy_stock_file)
        _file_row(2, '采购订单PO', self.material_buy_po_var, '选择PO文件', self._choose_material_buy_po_file)
        _file_row(3, '请购单PR', self.material_buy_pr_var, '选择PR文件', self._choose_material_buy_pr_file)
        _file_row(4, '生产用耗材表', self.material_buy_usage_var, '选择耗材表', self._choose_material_buy_usage_file)

        ttk.Label(control, text='物料编码', style='Section.TLabel').grid(row=5, column=0, sticky='w', padx=(0, 6), pady=(8, 3))
        ttk.Entry(control, textvariable=self.material_buy_code_var).grid(row=5, column=1, sticky='ew', pady=(8, 3))
        ttk.Label(control, text='可多个；为空则按单据号或PR全量', style='Subtle.TLabel').grid(row=5, column=2, sticky='w', padx=(8, 0), pady=(8, 3))

        query_frame = ttk.Frame(control)
        query_frame.grid(row=5, column=3, columnspan=2, sticky='e', pady=(8, 3))
        ttk.Label(query_frame, text='截至月份', style='Subtle.TLabel').grid(row=0, column=0, sticky='w')
        ttk.Entry(query_frame, textvariable=self.material_buy_month_var, width=10).grid(row=0, column=1, sticky='w', padx=(4, 10))
        ttk.Label(query_frame, text='查询月数', style='Subtle.TLabel').grid(row=0, column=2, sticky='w')
        ttk.Spinbox(query_frame, from_=1, to=24, increment=1, textvariable=self.material_buy_query_months_var, width=7).grid(row=0, column=3, sticky='w', padx=(4, 0))

        ttk.Label(control, text='请购单号', style='Section.TLabel').grid(row=6, column=0, sticky='w', padx=(0, 6), pady=3)
        ttk.Entry(control, textvariable=self.material_buy_doc_var).grid(row=6, column=1, sticky='ew', pady=3)
        ttk.Label(control, text='例如 101PR2605110003；输入后只判断该单据里的料号', style='Subtle.TLabel').grid(
            row=6, column=2, columnspan=3, sticky='w', padx=(8, 0), pady=3
        )

        ttk.Label(control, text='批量粘贴', style='Section.TLabel').grid(row=7, column=0, sticky='nw', padx=(0, 6), pady=3)
        paste_frame = ttk.Frame(control)
        paste_frame.grid(row=7, column=1, columnspan=3, sticky='ew', pady=3)
        paste_frame.columnconfigure(0, weight=1)
        self.material_buy_paste_text = tk.Text(paste_frame, height=4, wrap='none', font=('Consolas', 10))
        self.material_buy_paste_text.grid(row=0, column=0, sticky='ew')
        paste_scroll = ttk.Scrollbar(paste_frame, orient='vertical', command=self.material_buy_paste_text.yview)
        paste_scroll.grid(row=0, column=1, sticky='ns')
        self.material_buy_paste_text.configure(yscrollcommand=paste_scroll.set)
        paste_buttons = ttk.Frame(control)
        paste_buttons.grid(row=7, column=4, sticky='nsew', padx=(8, 0), pady=3)
        ttk.Button(paste_buttons, text='粘贴剪贴板', command=self._paste_material_buy_items).grid(row=0, column=0, sticky='ew')
        ttk.Button(paste_buttons, text='清空粘贴', style='Quiet.TButton', command=self._clear_material_buy_paste).grid(row=1, column=0, sticky='ew', pady=(6, 0))
        ttk.Label(
            paste_buttons,
            text='支持从Excel复制两列：料号、数量',
            style='Subtle.TLabel',
            wraplength=150,
            justify='left',
        ).grid(row=2, column=0, sticky='w', pady=(6, 0))

        action = ttk.Frame(control)
        action.grid(row=8, column=0, columnspan=5, sticky='ew', pady=(8, 0))
        action.columnconfigure(4, weight=1)
        ttk.Button(action, text='开始判断', style='Accent.TButton', command=self._run_material_buy_analysis).grid(row=0, column=0, sticky='w')
        ttk.Button(action, text='导出结果', command=self._export_material_buy_result).grid(row=0, column=1, sticky='w', padx=(8, 0))
        ttk.Button(action, text='复制结果', style='Quiet.TButton', command=self._copy_material_buy_result).grid(row=0, column=2, sticky='w', padx=(8, 0))
        ttk.Button(action, text='清空', style='Quiet.TButton', command=self._clear_material_buy_result).grid(row=0, column=3, sticky='w', padx=(8, 0))
        ttk.Label(action, textvariable=self.material_buy_summary_var, style='Metric.TLabel', wraplength=520, justify='left').grid(row=0, column=4, sticky='w', padx=(12, 0))

        result = ttk.LabelFrame(f, text='采购判断结果', style='Card.TLabelframe', padding=6)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        self.material_buy_tree = self._create_result_tree(result, self.material_buy_headers, self.material_buy_widths)

    def _choose_material_buy_issue_file(self):
        path = filedialog.askopenfilename(title='选择收发记录', filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.material_buy_issue_var.set(path)

    def _choose_material_buy_stock_file(self):
        path = filedialog.askopenfilename(title='选择库存明细', filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.material_buy_stock_var.set(path)

    def _choose_material_buy_po_file(self):
        path = filedialog.askopenfilename(title='选择采购订单PO', filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.material_buy_po_var.set(path)

    def _choose_material_buy_pr_file(self):
        path = filedialog.askopenfilename(title='选择请购单PR', filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.material_buy_pr_var.set(path)

    def _choose_material_buy_usage_file(self):
        path = filedialog.askopenfilename(title='选择生产用耗材表', filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.material_buy_usage_var.set(path)

    def _parse_material_buy_month(self):
        text = str(self.material_buy_month_var.get() or '').strip()
        if not text:
            return date.today().year, date.today().month, date.today().strftime('%Y-%m')
        match = re.fullmatch(r'(\d{4})[-/年.]?(\d{1,2})月?', text)
        if match:
            year, month = int(match.group(1)), int(match.group(2))
        elif re.fullmatch(r'\d{6}', text):
            year, month = int(text[:4]), int(text[4:])
        elif re.fullmatch(r'\d{1,2}', text):
            year, month = date.today().year, int(text)
        else:
            raise ValueError('查询月份格式不正确，请输入 2026-05、202605 或 5')
        if month < 1 or month > 12:
            raise ValueError('查询月份必须在1-12之间')
        return year, month, f'{year:04d}-{month:02d}'

    def _material_buy_month_labels(self, year, month, month_count):
        labels = []
        current = date(year, month, 1)
        for offset in range(month_count - 1, -1, -1):
            total_month = current.year * 12 + current.month - 1 - offset
            y, m = divmod(total_month, 12)
            labels.append(f'{y:04d}-{m + 1:02d}')
        return labels

    def _material_buy_result_columns(self, month_labels):
        month_headers = [f'{label}领用量' for label in month_labels]
        headers = MATERIAL_BUY_PREFIX_HEADERS + month_headers + MATERIAL_BUY_SUFFIX_HEADERS
        widths = MATERIAL_BUY_PREFIX_WIDTHS + [92] * len(month_headers) + MATERIAL_BUY_SUFFIX_WIDTHS
        return headers, widths

    def _configure_material_buy_tree(self, month_labels):
        headers, widths = self._material_buy_result_columns(month_labels)
        self.material_buy_headers = headers
        self.material_buy_widths = widths
        if not hasattr(self, 'material_buy_tree'):
            return
        self.material_buy_tree.configure(columns=headers)
        numeric_headers = {
            h for h in headers
            if any(keyword in h for keyword in ('数量', '需求', '可用', '库存', '领用', '供给'))
        }
        for header, width in zip(headers, widths):
            anchor = 'e' if header in numeric_headers else 'w'
            self.material_buy_tree.column(header, width=width, anchor=anchor)
            self.material_buy_tree.heading(header, text=header)
        enable_treeview_sort(self.material_buy_tree, list(headers), numeric_columns=numeric_headers)
        enable_treeview_copy(self.material_buy_tree, list(headers))

    def _material_buy_split_codes(self, text):
        text = str(text or '').strip()
        if not text:
            return []
        for sep in ('，', '、', ';', '；', '\n', '\t'):
            text = text.replace(sep, ',')
        codes = []
        for part in text.split(','):
            for item in part.split():
                code = normalize_material_code(item)
                if code and code not in codes:
                    codes.append(code)
        return codes

    def _paste_material_buy_items(self):
        try:
            text = self.root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning('粘贴失败', '剪贴板为空或无法读取')
            return
        if not text.strip():
            return
        self.material_buy_paste_text.delete('1.0', 'end')
        self.material_buy_paste_text.insert('1.0', text)
        try:
            count = len(self._parse_material_buy_paste_items()[0])
        except Exception:
            count = 0
        self.status_var.set(f'已粘贴批量料号 {count} 行')

    def _clear_material_buy_paste(self):
        if hasattr(self, 'material_buy_paste_text'):
            self.material_buy_paste_text.delete('1.0', 'end')

    def _parse_material_buy_paste_items(self):
        if not hasattr(self, 'material_buy_paste_text'):
            return [], {}
        text = self.material_buy_paste_text.get('1.0', 'end').strip()
        if not text:
            return [], {}
        order = []
        qty_map = defaultdict(float)
        for row_index, raw_line in enumerate(text.splitlines(), start=1):
            line = raw_line.strip()
            if not line:
                continue
            normalized = line.replace('，', ',').replace('、', ',').replace(';', ',').replace('；', ',')
            if '\t' in normalized:
                parts = [part.strip() for part in normalized.split('\t')]
            elif ',' in normalized:
                parts = [part.strip() for part in normalized.split(',')]
            else:
                parts = normalized.split()
            parts = [part for part in parts if part]
            if not parts:
                continue
            if row_index == 1 and any(keyword in ''.join(parts[:2]) for keyword in ('料号', '物料', '编码', '数量')):
                continue
            code = normalize_material_code(parts[0])
            if not code:
                continue
            qty = 0.0
            if len(parts) >= 2:
                qty_text = str(parts[1]).strip().replace(',', '')
                try:
                    qty = float(qty_text)
                except ValueError as exc:
                    raise ValueError(f'批量粘贴第 {row_index} 行数量无效：{parts[1]}') from exc
            if code not in qty_map:
                order.append(code)
            qty_map[code] += qty
        return order, dict(qty_map)

    def _read_excel_table_auto(self, path, required_columns):
        preview = pd.read_excel(path, sheet_name=0, header=None, nrows=20, dtype=object)
        required_columns = [str(col).strip() for col in required_columns]
        header_row = 0
        best_score = -1
        for idx, row in preview.iterrows():
            values = {str(value).strip().replace('\n', '').replace('\r', '') for value in row.tolist() if pd.notna(value)}
            score = sum(1 for col in required_columns if col in values)
            if score > best_score:
                best_score = score
                header_row = int(idx)
        df = pd.read_excel(path, sheet_name=0, header=header_row, dtype=object)
        df = normalize_sheet_columns(df)
        df = df.loc[:, [column for column in df.columns if not str(column).startswith('Unnamed')]]
        return df.dropna(how='all')

    def _first_existing_col(self, df, candidates, required=True):
        normalized = {str(column).strip().replace('\n', '').replace('\r', ''): column for column in df.columns}
        for candidate in candidates:
            if candidate in normalized:
                return normalized[candidate]
        if required:
            raise ValueError(f'缺少字段: {", ".join(candidates)}')
        return None

    def _read_material_buy_usage_info(self, path):
        if not path:
            return {}
        cache_key = (os.path.abspath(path), os.path.getmtime(path))
        cache = getattr(self, 'material_buy_usage_cache', {})
        if cache_key in cache:
            return cache[cache_key]
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        preview_rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=20, values_only=True)]
        header_index = None
        for idx, row in enumerate(preview_rows):
            values = [str(value).strip() for value in row if value is not None]
            if 'U9料号' in values and '用途' in values:
                header_index = idx
                break
        if header_index is None:
            raise ValueError('耗材用途表未识别到“U9料号/用途”表头')

        header_row = preview_rows[header_index]
        subheader_row = preview_rows[header_index + 1] if header_index + 1 < len(preview_rows) else []

        def _find_header(name):
            for pos, value in enumerate(header_row):
                if str(value).strip() == name:
                    return pos
            return None

        code_idx = _find_header('U9料号')
        status_idx = _find_header('状态')
        name_idx = _find_header('子项名称')
        spec_idx = _find_header('型号')
        purpose_idx = _find_header('用途')
        planned_avg_idx = _find_header('平均月用量（按年总产量1万台推算）')
        dept_start_idx = _find_header('使用部门')
        if code_idx is None or purpose_idx is None:
            raise ValueError('耗材用途表缺少“U9料号”或“用途”字段')

        dept_columns = []
        if dept_start_idx is not None:
            for pos in range(dept_start_idx, len(header_row)):
                parent_header = str(header_row[pos]).strip() if pd.notna(header_row[pos]) else ''
                if pos > dept_start_idx and parent_header:
                    break
                subheader = str(subheader_row[pos]).strip() if pos < len(subheader_row) and pd.notna(subheader_row[pos]) else ''
                if subheader:
                    dept_columns.append((pos, subheader))

        result = {}
        blank_streak = 0
        for row in ws.iter_rows(min_row=header_index + 3, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                blank_streak += 1
                if blank_streak >= 100:
                    break
                continue
            blank_streak = 0
            if code_idx >= len(row):
                continue
            code = normalize_material_code(row[code_idx])
            if not code:
                continue
            item = result.setdefault(code, {
                'purposes': [],
                'departments': [],
                'status': '',
                'name': '',
                'spec': '',
                'planned_avg_usage': 0.0,
            })
            if status_idx is not None and status_idx < len(row) and not item['status'] and row[status_idx] is not None:
                item['status'] = str(row[status_idx]).strip()
            if name_idx is not None and name_idx < len(row) and not item['name'] and row[name_idx] is not None:
                item['name'] = str(row[name_idx]).strip()
            if spec_idx is not None and spec_idx < len(row) and not item['spec'] and row[spec_idx] is not None:
                item['spec'] = str(row[spec_idx]).strip()
            if planned_avg_idx is not None and planned_avg_idx < len(row):
                try:
                    item['planned_avg_usage'] = max(item['planned_avg_usage'], float(row[planned_avg_idx] or 0))
                except (TypeError, ValueError):
                    pass
            purpose = str(row[purpose_idx]).strip() if purpose_idx < len(row) and pd.notna(row[purpose_idx]) else ''
            if purpose and purpose not in item['purposes']:
                item['purposes'].append(purpose)
            for pos, dept in dept_columns:
                if pos >= len(row) or pd.isna(row[pos]):
                    continue
                value = row[pos]
                try:
                    used = float(value) != 0
                except (TypeError, ValueError):
                    used = bool(str(value).strip())
                if used and dept not in item['departments']:
                    item['departments'].append(dept)

        output = {
            code: {
                'purpose': '；'.join(item['purposes']),
                'departments': '、'.join(item['departments']),
                'status': item['status'],
                'name': item['name'],
                'spec': item['spec'],
                'planned_avg_usage': item['planned_avg_usage'],
            }
            for code, item in result.items()
        }
        wb.close()
        cache[cache_key] = output
        self.material_buy_usage_cache = cache
        return output

    def _run_material_buy_analysis(self):
        try:
            issue_path = self.material_buy_issue_var.get().strip()
            stock_path = self.material_buy_stock_var.get().strip()
            po_path = self.material_buy_po_var.get().strip()
            pr_path = self.material_buy_pr_var.get().strip()
            usage_path = self.material_buy_usage_var.get().strip()
            missing = [
                name for name, path in [
                    ('收发记录', issue_path), ('库存明细', stock_path), ('采购订单PO', po_path), ('请购单PR', pr_path)
                ] if not path
            ]
            if missing:
                messagebox.showwarning('参数不完整', '请先选择: ' + '、'.join(missing))
                return
            year, month, month_text = self._parse_material_buy_month()
            try:
                query_months = int(self.material_buy_query_months_var.get())
            except (tk.TclError, ValueError, TypeError):
                query_months = 3
            if query_months <= 0:
                raise ValueError('查询月数必须大于0')
            month_labels = self._material_buy_month_labels(year, month, query_months)
            self._configure_material_buy_tree(month_labels)
            self.status_var.set('正在分析单料采购判断...')
            self.material_buy_summary_var.set('正在读取数据，请稍候...')
            self.root.update_idletasks()
            rows = self._calculate_material_buy_analysis(
                issue_path, stock_path, po_path, pr_path, usage_path, month_labels
            )
            self.material_buy_rows = rows
            self._clear_tree_rows(self.material_buy_tree)
            for row in rows:
                self.material_buy_tree.insert('', 'end', values=[row.get(header, '') for header in self.material_buy_headers])
            need_count = sum(1 for row in rows if row.get('判定') == '建议购买')
            transfer_count = sum(1 for row in rows if row.get('判定') == '优先转单')
            expedite_count = sum(1 for row in rows if row.get('判定') == '优先催交')
            summary = f'分析完成 | 物料 {len(rows)} 个 | 建议购买 {need_count} 个 | 优先转单 {transfer_count} 个 | 优先催交 {expedite_count} 个'
            self.material_buy_summary_var.set(summary)
            self.status_var.set(summary)
        except Exception as exc:
            self.material_buy_summary_var.set('分析失败')
            self.status_var.set('单料采购判断失败')
            messagebox.showerror('分析失败', str(exc))

    def _calculate_material_buy_analysis(self, issue_path, stock_path, po_path, pr_path, usage_path, month_labels):
        issue_df = self._read_excel_table_auto(issue_path, ['物料编码', '业务方向', '收发日期', '执行数量'])
        stock_df = self._read_excel_table_auto(stock_path, ['物料编码', '库存量'])
        po_df = self._read_excel_table_auto(po_path, ['料号', '欠交数量'])
        pr_df = self._read_excel_table_auto(pr_path, ['料号', '未转PO数量'])

        issue_code_col = self._first_existing_col(issue_df, ['物料编码', '料号'])
        issue_name_col = self._first_existing_col(issue_df, ['物料', '物料名称', '料品名称'], required=False)
        issue_spec_col = self._first_existing_col(issue_df, ['规格', '料品规格'], required=False)
        issue_direction_col = self._first_existing_col(issue_df, ['业务方向'])
        issue_date_col = self._first_existing_col(issue_df, ['收发日期', '业务日期'])
        issue_qty_col = self._first_existing_col(issue_df, ['执行数量', '数量'])

        stock_code_col = self._first_existing_col(stock_df, ['物料编码', '料号'])
        stock_name_col = self._first_existing_col(stock_df, ['物料名称', '物料', '料品名称'], required=False)
        stock_spec_col = self._first_existing_col(stock_df, ['规格', '料品规格'], required=False)
        stock_qty_col = self._first_existing_col(stock_df, ['库存量', '库存数量'])

        po_code_col = self._first_existing_col(po_df, ['料号', '物料编码'])
        po_name_col = self._first_existing_col(po_df, ['料品名称', '物料名称', '物料'], required=False)
        po_spec_col = self._first_existing_col(po_df, ['料品规格', '规格'], required=False)
        po_supplier_col = self._first_existing_col(po_df, ['供应商'], required=False)
        po_buyer_col = self._first_existing_col(po_df, ['采购业务员', '采购员', '采购'], required=False)
        po_qty_col = self._first_existing_col(po_df, ['欠交数量', '未交付数量', '未到货数量'])

        pr_code_col = self._first_existing_col(pr_df, ['料号', '物料编码'])
        pr_doc_col = self._first_existing_col(pr_df, ['单据编号', '来源单据号', '请购单号'], required=False)
        pr_name_col = self._first_existing_col(pr_df, ['料品名称', '物料名称', '物料'], required=False)
        pr_spec_col = self._first_existing_col(pr_df, ['料品规格', '规格'], required=False)
        pr_buyer_col = self._first_existing_col(pr_df, ['采购员', '采购'], required=False)
        pr_qty_col = self._first_existing_col(pr_df, ['未转PO数量', '未转PO数', '未转单数量'])
        pr_approved_col = self._first_existing_col(pr_df, ['核准采购数量', '请购数量'], required=False)
        pr_status_col = self._first_existing_col(pr_df, ['行状态', '状态'], required=False)

        def _code_series(df, col):
            return df[col].map(normalize_material_code)

        issue_df = issue_df.copy()
        stock_df = stock_df.copy()
        po_df = po_df.copy()
        pr_df = pr_df.copy()
        issue_df['_code'] = _code_series(issue_df, issue_code_col)
        stock_df['_code'] = _code_series(stock_df, stock_code_col)
        po_df['_code'] = _code_series(po_df, po_code_col)
        pr_df['_code'] = _code_series(pr_df, pr_code_col)

        usage_info = self._read_material_buy_usage_info(usage_path)
        paste_codes, input_qty_map = self._parse_material_buy_paste_items()
        query_codes = list(paste_codes)
        typed_codes = self._material_buy_split_codes(self.material_buy_code_var.get())
        if not query_codes:
            query_codes = typed_codes
        doc_var = getattr(self, 'material_buy_doc_var', None)
        query_doc = str(doc_var.get() if doc_var else '').strip()
        if query_doc and not paste_codes:
            if not pr_doc_col:
                raise ValueError('PR表未识别到“单据编号”字段，无法按请购单号筛选')
            doc_pr_df = pr_df[pr_df[pr_doc_col].astype(str).str.strip().eq(query_doc)].copy()
            if doc_pr_df.empty:
                raise ValueError(f'PR表中未找到请购单号: {query_doc}')
            doc_codes = sorted({code for code in doc_pr_df['_code'].dropna().astype(str).tolist() if code})
            if query_codes:
                query_codes = [code for code in query_codes if code in set(doc_codes)]
                if not query_codes:
                    raise ValueError(f'输入的物料编码不在请购单 {query_doc} 中')
            else:
                query_codes = doc_codes
            if pr_approved_col:
                temp = doc_pr_df.copy()
                temp['_approved_qty'] = pd.to_numeric(temp[pr_approved_col], errors='coerce').fillna(0)
                input_qty_map = temp.groupby('_code')['_approved_qty'].sum().to_dict()
        if not query_codes and usage_info:
            query_codes = sorted(
                code for code, item in usage_info.items()
                if str(item.get('status', '')).strip() == '正常使用'
            )
        if not query_codes:
            query_codes = sorted({code for code in pr_df['_code'].dropna().astype(str).tolist() if code})
        if not query_codes:
            raise ValueError('未识别到要分析的物料编码。请输入物料编码，或确认PR表中存在料号。')

        issue_df['_date'] = pd.to_datetime(issue_df[issue_date_col], errors='coerce')
        issue_df['_qty'] = pd.to_numeric(issue_df[issue_qty_col], errors='coerce').fillna(0)
        issue_df['_month_text'] = issue_df['_date'].dt.strftime('%Y-%m')
        issue_month_df = issue_df[
            (issue_df['_month_text'].isin(month_labels))
            & (issue_df['_code'].isin(query_codes))
        ].copy()
        issue_month_df['_direction'] = issue_month_df[issue_direction_col].astype(str)

        out_df = issue_month_df[issue_month_df['_direction'].str.contains('出', na=False)]
        usage_month_map = out_df.groupby(['_code', '_month_text'])['_qty'].sum().to_dict()
        usage_total_map = out_df.groupby('_code')['_qty'].sum().to_dict()

        stock_df['_stock_qty'] = pd.to_numeric(stock_df[stock_qty_col], errors='coerce').fillna(0)
        stock_map = stock_df[stock_df['_code'].isin(query_codes)].groupby('_code')['_stock_qty'].sum().to_dict()

        po_df['_po_qty'] = pd.to_numeric(po_df[po_qty_col], errors='coerce').fillna(0)
        po_open_df = po_df[(po_df['_code'].isin(query_codes)) & (po_df['_po_qty'] > 0)].copy()
        po_map = po_open_df.groupby('_code')['_po_qty'].sum().to_dict()

        pr_df['_pr_qty'] = pd.to_numeric(pr_df[pr_qty_col], errors='coerce').fillna(0)
        if pr_status_col:
            pr_active_df = pr_df[~pr_df[pr_status_col].astype(str).str.contains('关闭|作废|取消', na=False)].copy()
        else:
            pr_active_df = pr_df.copy()
        if query_doc and pr_doc_col:
            pr_active_df = pr_active_df[pr_active_df[pr_doc_col].astype(str).str.strip().eq(query_doc)].copy()
        pr_open_df = pr_active_df[(pr_active_df['_code'].isin(query_codes)) & (pr_active_df['_pr_qty'] > 0)].copy()
        pr_map = pr_open_df.groupby('_code')['_pr_qty'].sum().to_dict()

        meta = {}

        def _record_meta(df, code_col='_code', name_col=None, spec_col=None, buyer_col=None, supplier_col=None):
            for rec in df.to_dict('records'):
                code = str(rec.get(code_col, '') or '').strip()
                if not code or code not in query_codes:
                    continue
                item = meta.setdefault(code, {'name': '', 'spec': '', 'buyer': '', 'supplier': ''})
                if name_col and not item['name'] and pd.notna(rec.get(name_col)):
                    item['name'] = str(rec.get(name_col, '') or '').strip()
                if spec_col and not item['spec'] and pd.notna(rec.get(spec_col)):
                    item['spec'] = str(rec.get(spec_col, '') or '').strip()
                if buyer_col and not item['buyer'] and pd.notna(rec.get(buyer_col)):
                    item['buyer'] = str(rec.get(buyer_col, '') or '').strip()
                if supplier_col and not item['supplier'] and pd.notna(rec.get(supplier_col)):
                    item['supplier'] = str(rec.get(supplier_col, '') or '').strip()

        _record_meta(pr_df, name_col=pr_name_col, spec_col=pr_spec_col, buyer_col=pr_buyer_col)
        _record_meta(po_df, name_col=po_name_col, spec_col=po_spec_col, buyer_col=po_buyer_col, supplier_col=po_supplier_col)
        _record_meta(stock_df, name_col=stock_name_col, spec_col=stock_spec_col)
        _record_meta(issue_df, name_col=issue_name_col, spec_col=issue_spec_col)
        rows = []
        month_range_text = month_labels[0] if len(month_labels) == 1 else f'{month_labels[0]}~{month_labels[-1]}'
        for code in query_codes:
            monthly_usage = {
                label: max(0.0, float(usage_month_map.get((code, label), 0) or 0))
                for label in month_labels
            }
            total_usage = max(0.0, float(usage_total_map.get(code, 0) or 0))
            avg_usage = total_usage / len(month_labels) if month_labels else 0.0
            usage_item = usage_info.get(code, {})
            planned_avg_usage = max(0.0, float(usage_item.get('planned_avg_usage', 0) or 0))
            stock_qty = max(0.0, float(stock_map.get(code, 0) or 0))
            pr_qty = max(0.0, float(pr_map.get(code, 0) or 0))
            po_qty = max(0.0, float(po_map.get(code, 0) or 0))
            input_qty = max(0.0, float(input_qty_map.get(code, 0) or 0))
            available_supply = stock_qty + pr_qty + po_qty
            target_qty = max(avg_usage, planned_avg_usage, input_qty)
            gap_qty = max(0.0, target_qty - available_supply)
            stock_gap = max(0.0, target_qty - stock_qty)
            if target_qty <= 0:
                judgement = '暂不购买'
                suggestion_qty = 0
                opinion = f'{month_range_text}无领用记录，且未输入本次需求数量，暂不建议新增购买；如有后续需求，请另行补充判断。'
            elif gap_qty > 1e-9:
                judgement = '建议购买'
                suggestion_qty = int(math.ceil(gap_qty))
                if input_qty >= avg_usage and input_qty >= planned_avg_usage:
                    basis_text = f'本次输入/请购数量{fmt_qty(input_qty)}为最高判断依据'
                elif planned_avg_usage >= avg_usage:
                    basis_text = f'表内月均用量{fmt_qty(planned_avg_usage)}高于近{len(month_labels)}个月实际月均领用{fmt_qty(avg_usage)}'
                else:
                    basis_text = f'近{len(month_labels)}个月实际月均领用{fmt_qty(avg_usage)}高于表内月均用量{fmt_qty(planned_avg_usage)}'
                opinion = (
                    f'{basis_text}，按两者取大作为判断需求{fmt_qty(target_qty)}；'
                    f'库存+PR未转+PO欠交合计{fmt_qty(available_supply)}，仍缺{fmt_qty(gap_qty)}，建议新增请购/补采{fmt_qty(suggestion_qty)}。'
                )
                if input_qty > 0:
                    if input_qty >= suggestion_qty:
                        opinion += f'本次输入数量{fmt_qty(input_qty)}可覆盖建议新增量。'
                    else:
                        opinion += f'本次输入数量{fmt_qty(input_qty)}低于建议新增量，还差{fmt_qty(suggestion_qty - input_qty)}。'
            elif stock_qty >= target_qty:
                judgement = '暂不购买'
                suggestion_qty = 0
                opinion = (
                    f'当前库存{fmt_qty(stock_qty)}已覆盖目标{fmt_qty(target_qty)}，请购记录无需新增购买；'
                    f'如已有PR未转{fmt_qty(pr_qty)}或PO欠交{fmt_qty(po_qty)}，可评估是否取消、延期或保留。'
                )
                if input_qty > 0:
                    opinion += f'本次输入数量{fmt_qty(input_qty)}建议暂缓/取消或改为备货审批。'
            elif pr_qty > 0:
                judgement = '优先转单'
                suggestion_qty = 0
                opinion = (
                    f'库存不足{fmt_qty(stock_gap)}，但PR未转PO{fmt_qty(pr_qty)}，加上PO欠交后可覆盖目标；'
                    f'处理建议：优先推动PR转PO，不新增请购。'
                )
                if input_qty > 0:
                    opinion += f'本次输入数量{fmt_qty(input_qty)}应先确认是否就是待转PO数量，避免重复请购。'
            elif po_qty > 0:
                judgement = '优先催交'
                suggestion_qty = 0
                opinion = (
                    f'库存不足{fmt_qty(stock_gap)}，但PO欠交{fmt_qty(po_qty)}可覆盖目标；'
                    f'处理建议：不新增请购，优先跟催供应商交付。'
                )
                if input_qty > 0:
                    opinion += f'本次输入数量{fmt_qty(input_qty)}建议先暂停新增，优先确认PO交期。'
            else:
                judgement = '暂不购买'
                suggestion_qty = 0
                opinion = f'可用供给{fmt_qty(available_supply)}可覆盖目标{fmt_qty(target_qty)}，暂不新增购买。'

            item = meta.get(code, {})
            row = {
                '判定': judgement,
                '物料编码': code,
                '物料名称': item.get('name', '') or usage_item.get('name', ''),
                '规格': item.get('spec', '') or usage_item.get('spec', ''),
                '状态': usage_item.get('status', ''),
                '用途': usage_item.get('purpose', ''),
                '使用部门': usage_item.get('departments', ''),
                '查询区间': month_range_text,
                '输入/请购数量': fmt_qty(input_qty) if input_qty > 0 else '',
                '查询期领用合计': fmt_qty(total_usage),
                '实际月均领用': fmt_qty(avg_usage),
                '表内月均用量': fmt_qty(planned_avg_usage),
                '当前库存': fmt_qty(stock_qty),
                'PR未转PO': fmt_qty(pr_qty),
                'PO欠交': fmt_qty(po_qty),
                '可用供给': fmt_qty(available_supply),
                '判断需求量': fmt_qty(target_qty),
                '建议新增购买量': fmt_qty(suggestion_qty),
                '采购员': item.get('buyer', ''),
                '供应商': item.get('supplier', ''),
                '处理意见': opinion,
            }
            for label, qty in monthly_usage.items():
                row[f'{label}领用量'] = fmt_qty(qty)
            rows.append(row)
        priority = {'建议购买': 0, '优先转单': 1, '优先催交': 2, '暂不购买': 3}
        rows.sort(key=lambda row: (priority.get(row.get('判定'), 9), str(row.get('物料编码', ''))))
        return rows

    def _clear_material_buy_result(self):
        self.material_buy_rows = []
        if hasattr(self, 'material_buy_tree'):
            self._clear_tree_rows(self.material_buy_tree)
        if hasattr(self, 'material_buy_summary_var'):
            self.material_buy_summary_var.set('等待分析')

    def _copy_material_buy_result(self):
        rows = getattr(self, 'material_buy_rows', []) or []
        if not rows:
            messagebox.showinfo('复制', '没有可复制的结果')
            return
        headers = getattr(self, 'material_buy_headers', []) or []
        lines = ['\t'.join(headers)]
        for row in rows:
            lines.append('\t'.join(str(row.get(header, '')) for header in headers))
        self.root.clipboard_clear()
        self.root.clipboard_append('\n'.join(lines))
        self.status_var.set(f'已复制 {len(rows)} 行采购判断结果')

    def _export_material_buy_result(self):
        rows = getattr(self, 'material_buy_rows', []) or []
        if not rows:
            messagebox.showwarning('无数据', '请先执行采购判断')
            return
        path = filedialog.asksaveasfilename(
            title='导出单料采购判断结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')],
        )
        if not path:
            return
        try:
            headers = getattr(self, 'material_buy_headers', []) or []
            widths = getattr(self, 'material_buy_widths', []) or []
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '单料采购判断'
            hdr_font = Font(bold=True, color='000000')
            header_fill = PatternFill('solid', fgColor='D9EAF7')
            need_fill = PatternFill('solid', fgColor='F8CBAD')
            action_fill = PatternFill('solid', fgColor='FFF2CC')
            ok_fill = PatternFill('solid', fgColor='D9EAD3')
            thin = Side(style='thin', color='AAAAAA')
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=header)
                cell.font = hdr_font
                cell.fill = header_fill
                cell.border = bdr
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for ri, row in enumerate(rows, 2):
                judgement = row.get('判定', '')
                fill = need_fill if judgement == '建议购买' else (action_fill if judgement in ('优先转单', '优先催交') else ok_fill)
                for ci, header in enumerate(headers, 1):
                    cell = ws.cell(row=ri, column=ci, value=row.get(header, ''))
                    cell.border = bdr
                    cell.fill = fill if ci == 1 else PatternFill()
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            for idx, width in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(8, min(70, width / 7))
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{ws.max_row}"
            ws.freeze_panes = 'A2'
            wb.save(path)
            self.status_var.set(f'已导出: {os.path.basename(path)}')
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
        except Exception as exc:
            messagebox.showerror('导出失败', str(exc))

    def _build_work_order_reply_tab(self):
        f = self.tab_work_order_reply
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='工单缺料回复参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(1, weight=1)
        control.columnconfigure(4, weight=1)

        self.work_order_shortage_var = tk.StringVar()
        self.work_order_balance_var = tk.StringVar()
        self.work_order_stock_var = tk.StringVar()
        self.work_order_arrival_var = tk.StringVar()
        self.work_order_output_var = tk.StringVar()
        self.work_order_summary_var = tk.StringVar(value='等待生成')
        self.work_order_preserve_var = tk.BooleanVar(value=True)
        self._work_order_reply_running = False
        self._work_order_reply_output_path = ''

        def _file_row(row_index, label_text, variable, button_text, command, column_offset):
            ttk.Label(control, text=label_text, style='Section.TLabel').grid(
                row=row_index, column=column_offset, sticky='w', padx=(0, 6), pady=3
            )
            label, _ = make_path_label(control, variable)
            label.grid(row=row_index, column=column_offset + 1, sticky='ew', pady=3)
            ttk.Button(control, text=button_text, command=command).grid(
                row=row_index, column=column_offset + 2, sticky='w', padx=(6, 0), pady=3
            )

        _file_row(0, '工单缺料清单', self.work_order_shortage_var, '选择清单', self._choose_work_order_shortage_file, 0)
        _file_row(0, '静态平衡表', self.work_order_balance_var, '选择平衡表', self._choose_work_order_balance_file, 3)
        _file_row(1, '库存明细', self.work_order_stock_var, '选择库存', self._choose_work_order_stock_file, 0)
        _file_row(1, '到货记录', self.work_order_arrival_var, '选择到货', self._choose_work_order_arrival_file, 3)
        _file_row(2, '输出文件', self.work_order_output_var, '指定输出路径', self._choose_work_order_output_file, 0)

        option_row = ttk.Frame(control)
        option_row.grid(row=2, column=3, columnspan=3, sticky='ew', padx=(12, 0))
        ttk.Checkbutton(
            option_row,
            text='保留已有最新交期，并占用对应回复数量',
            variable=self.work_order_preserve_var,
        ).grid(row=0, column=0, sticky='w')

        action_row = ttk.Frame(control)
        action_row.grid(row=3, column=0, columnspan=6, sticky='ew', pady=(8, 0))
        self.work_order_run_btn = ttk.Button(
            action_row,
            text='生成工单缺料回复',
            style='Accent.TButton',
            command=self._run_work_order_reply_fill,
        )
        self.work_order_run_btn.grid(row=0, column=0, sticky='w')
        self.work_order_open_btn = ttk.Button(
            action_row,
            text='打开结果',
            style='Quiet.TButton',
            command=self._open_work_order_reply_output,
        )
        self.work_order_open_btn.grid(row=0, column=1, sticky='w', padx=(8, 0))
        self.work_order_open_btn.state(['disabled'])
        ttk.Label(action_row, textvariable=self.work_order_summary_var, style='Metric.TLabel').grid(
            row=0, column=2, sticky='w', padx=(12, 0)
        )
        self.work_order_progress = ttk.Progressbar(action_row, mode='indeterminate', length=180)
        self.work_order_progress.grid(row=1, column=0, columnspan=3, sticky='ew', pady=(6, 0))
        self.work_order_progress.grid_remove()

        result = ttk.LabelFrame(f, text='处理说明', style='Card.TLabelframe', padding=6)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        result.columnconfigure(0, weight=1)
        result.rowconfigure(0, weight=1)
        self.work_order_log = tk.Text(result, height=12, wrap='word', font=('Consolas', 10))
        scroll = ttk.Scrollbar(result, orient='vertical', command=self.work_order_log.yview)
        self.work_order_log.configure(yscrollcommand=scroll.set)
        self.work_order_log.grid(row=0, column=0, sticky='nsew')
        scroll.grid(row=0, column=1, sticky='ns')
        self._append_work_order_reply_log('逻辑：先按库存扣减，再按到货记录中的接收/待检/在检/待入库扣减，最后按静态平衡表今天及以后的“采购答复”日期分配。')
        self._append_work_order_reply_log('同一料号重复出现时按工单清单行顺序扣减。')

    def _append_work_order_reply_log(self, message):
        if not hasattr(self, 'work_order_log'):
            return
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.work_order_log.insert('end', f'[{timestamp}] {message}\n')
        self.work_order_log.see('end')

    def _choose_work_order_shortage_file(self):
        path = filedialog.askopenfilename(
            title='选择工单缺料清单',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xltx *.xltm'), ('所有文件', '*.*')],
        )
        if path:
            self.work_order_shortage_var.set(path)
            if not self.work_order_output_var.get().strip():
                self.work_order_output_var.set(os.path.splitext(path)[0] + '_工单缺料回复.xlsx')

    def _choose_work_order_balance_file(self):
        path = filedialog.askopenfilename(
            title='选择静态平衡表',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xltx *.xltm'), ('所有文件', '*.*')],
        )
        if path:
            self.work_order_balance_var.set(path)

    def _choose_work_order_stock_file(self):
        path = filedialog.askopenfilename(
            title='选择库存明细',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xltx *.xltm'), ('所有文件', '*.*')],
        )
        if path:
            self.work_order_stock_var.set(path)

    def _choose_work_order_arrival_file(self):
        path = filedialog.askopenfilename(
            title='选择到货记录',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xltx *.xltm'), ('所有文件', '*.*')],
        )
        if path:
            self.work_order_arrival_var.set(path)

    def _choose_work_order_output_file(self):
        initial = self.work_order_output_var.get().strip()
        path = filedialog.asksaveasfilename(
            title='指定工单缺料回复输出文件',
            defaultextension='.xlsx',
            initialfile=os.path.basename(initial) if initial else '工单缺料回复.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')],
        )
        if path:
            self.work_order_output_var.set(path)

    def _run_work_order_reply_fill(self):
        if run_work_order_reply_fill is None:
            show_data_error(
                self.root,
                title='工单缺料回复功能不可用',
                summary='未能加载工单缺料回复引擎。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='请确认 mrp_balance_tool 模块已随工具一起打包。',
            )
            return
        if self._work_order_reply_running:
            self.status_var.set('工单缺料回复正在生成中，请稍候')
            return
        shortage_path = self.work_order_shortage_var.get().strip()
        balance_path = self.work_order_balance_var.get().strip()
        stock_path = self.work_order_stock_var.get().strip()
        arrival_path = self.work_order_arrival_var.get().strip()
        output_path = self.work_order_output_var.get().strip()
        if not shortage_path:
            messagebox.showwarning('参数不完整', '请先选择工单缺料清单')
            return
        if not balance_path:
            messagebox.showwarning('参数不完整', '请先选择静态平衡表')
            return

        self._work_order_reply_running = True
        self.work_order_run_btn.state(['disabled'])
        self.work_order_open_btn.state(['disabled'])
        self.work_order_progress.grid()
        self.work_order_progress.start(10)
        self.work_order_summary_var.set('正在生成...')
        self.status_var.set('正在生成工单缺料回复...')
        self._append_work_order_reply_log('开始读取工单缺料清单和静态平衡表采购答复...')

        thread = threading.Thread(
            target=self._run_work_order_reply_worker,
            args=(shortage_path, balance_path, stock_path, arrival_path, output_path, bool(self.work_order_preserve_var.get())),
            daemon=True,
        )
        thread.start()

    def _run_work_order_reply_worker(self, shortage_path, balance_path, stock_path, arrival_path, output_path, preserve_existing):
        try:
            def progress(message):
                self.root.after(0, lambda msg=message: self._append_work_order_reply_log(msg))

            result = run_work_order_reply_fill(
                Path(shortage_path),
                Path(balance_path),
                Path(output_path) if output_path else None,
                preserve_existing=preserve_existing,
                stock_path=Path(stock_path) if stock_path else None,
                arrival_path=Path(arrival_path) if arrival_path else None,
                progress_callback=progress,
            )
            self.root.after(0, lambda: self._finish_work_order_reply_fill(result))
        except Exception as exc:
            message = str(exc)
            self.root.after(0, lambda msg=message: self._handle_work_order_reply_error(msg))

    def _finish_work_order_reply_fill(self, result):
        self._work_order_reply_running = False
        self.work_order_progress.stop()
        self.work_order_progress.grid_remove()
        self.work_order_run_btn.state(['!disabled'])
        self.work_order_open_btn.state(['!disabled'])
        self._work_order_reply_output_path = str(result.get('output_path', '') or '')
        summary = (
            f"完成 | 缺料行 {result.get('shortage_rows', 0)} | 填写 {result.get('updated_rows', 0)} | "
            f"保留 {result.get('preserved_rows', 0)} | 不足 {result.get('insufficient_rows', 0)}"
        )
        self.work_order_summary_var.set(summary)
        self.status_var.set(summary)
        self._append_work_order_reply_log(summary)
        self._append_work_order_reply_log(f"结果文件：{self._work_order_reply_output_path}")
        self._append_work_order_reply_log(
            f"库存明细：识别料号 {result.get('inventory_material_count', 0)} 个，明细行 {result.get('inventory_row_count', 0)} 行"
        )
        self._append_work_order_reply_log(
            f"到货记录：识别料号 {result.get('arrival_material_count', 0)} 个，明细行 {result.get('arrival_row_count', 0)} 行"
        )
        self._append_work_order_reply_log(
            f"平衡表采购回复：从 {result.get('reply_start_date', '')} 起，料号 {result.get('reply_material_count', 0)} 个，单元格 {result.get('reply_cell_count', 0)} 个"
        )
        messagebox.showinfo('生成完成', f"{summary}\n\n输出文件：{self._work_order_reply_output_path}")

    def _handle_work_order_reply_error(self, message):
        self._work_order_reply_running = False
        self.work_order_progress.stop()
        self.work_order_progress.grid_remove()
        self.work_order_run_btn.state(['!disabled'])
        self.work_order_summary_var.set('生成失败')
        self.status_var.set('工单缺料回复生成失败')
        self._append_work_order_reply_log(f'生成失败：{message}')
        show_data_error(
            self.root,
            title='工单缺料回复生成失败',
            summary='未能完成工单缺料回复。',
            detail=message,
            fix_hint='请确认工单缺料清单包含“物料.编码、缺料数量、最新交期”，静态平衡表包含“供需平衡”页和采购答复行。',
        )

    def _open_work_order_reply_output(self):
        path = self._work_order_reply_output_path or self.work_order_output_var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning('文件不存在', '还没有可打开的结果文件')
            return
        try:
            open_local_path(path)
        except Exception as exc:
            messagebox.showerror('打开失败', str(exc))

    def _build_arrival_tab(self):
        f = self.tab_arrival
        f.columnconfigure(0, weight=1)
        f.rowconfigure(1, weight=1)

        control = ttk.LabelFrame(f, text='到货跟催分析参数', style='Card.TLabelframe', padding=8)
        control.grid(row=0, column=0, sticky='ew', padx=6, pady=(4, 3))
        control.columnconfigure(0, weight=3)
        control.columnconfigure(1, weight=2)

        self.arrival_bom_var = tk.StringVar(value='当前 BOM: 未加载，将使用输入文件里的 BOM')
        ttk.Label(control, textvariable=self.arrival_bom_var, style='Metric.TLabel').grid(
            row=0, column=0, columnspan=2, sticky='w', pady=(0, 6)
        )

        self.arrival_input_var = tk.StringVar()
        self.arrival_record_var = tk.StringVar()
        self.arrival_plan_var = tk.StringVar()
        self.arrival_summary_var = tk.StringVar(value='等待分析')
        self.arrival_quality_stat_var = tk.StringVar(value='在检: 0')
        self.arrival_warehouse_stat_var = tk.StringVar(value='待入库/待收货: 0')
        self.arrival_purchase_pending_stat_var = tk.StringVar(value='采购未到货: 0')
        self.arrival_all_stat_var = tk.StringVar(value='命中物料: 0')
        self.arrival_urgent_stat_var = tk.StringVar(value='紧急跟催: 0')
        self.arrival_analysis_date_var = tk.StringVar(value=date.today().strftime('%Y-%m-%d'))
        self.arrival_window_days_var = tk.IntVar(value=3)
        self.arrival_filter_var = tk.StringVar(value='urgent')

        left_panel = ttk.Frame(control)
        left_panel.grid(row=1, column=0, sticky='ew')
        left_panel.columnconfigure(1, weight=1)

        ttk.Label(left_panel, text='MRP 输入文件', style='Section.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 6), pady=3)
        arrival_input_label, _ = make_path_label(left_panel, self.arrival_input_var)
        arrival_input_label.grid(row=0, column=1, sticky='ew', pady=3)
        ttk.Button(left_panel, text='选择输入文件', command=self._choose_arrival_input_file).grid(row=0, column=2, sticky='w', padx=(6, 0), pady=3)

        ttk.Label(left_panel, text='到货记录', style='Section.TLabel').grid(row=1, column=0, sticky='w', padx=(0, 6), pady=3)
        arrival_record_label, _ = make_path_label(left_panel, self.arrival_record_var)
        arrival_record_label.grid(row=1, column=1, sticky='ew', pady=3)
        ttk.Button(left_panel, text='选择到货记录', command=self._choose_arrival_record_file).grid(row=1, column=2, sticky='w', padx=(6, 0), pady=3)

        ttk.Label(left_panel, text='到货计划', style='Section.TLabel').grid(row=2, column=0, sticky='w', padx=(0, 6), pady=3)
        arrival_plan_label, _ = make_path_label(left_panel, self.arrival_plan_var)
        arrival_plan_label.grid(row=2, column=1, sticky='ew', pady=3)
        ttk.Button(left_panel, text='选择到货计划', command=self._choose_arrival_plan_file).grid(row=2, column=2, sticky='w', padx=(6, 0), pady=3)

        ttk.Label(left_panel, text='分析日期/需求天数', style='Section.TLabel').grid(row=3, column=0, sticky='w', padx=(0, 6), pady=3)
        arrival_date_frame = ttk.Frame(left_panel)
        arrival_date_frame.grid(row=3, column=1, sticky='w', pady=3)
        ttk.Entry(arrival_date_frame, textvariable=self.arrival_analysis_date_var, width=12).grid(row=0, column=0, sticky='w')
        ttk.Label(arrival_date_frame, text='  需求天数', style='Subtle.TLabel').grid(row=0, column=1, sticky='w')
        ttk.Spinbox(arrival_date_frame, from_=1, to=15, textvariable=self.arrival_window_days_var, width=8).grid(
            row=0, column=2, sticky='w', padx=(4, 0)
        )

        side_panel = ttk.Frame(control)
        side_panel.grid(row=1, column=1, sticky='nsew', padx=(14, 0))
        side_panel.columnconfigure(0, weight=1)
        side_panel.columnconfigure(1, weight=1)

        note = ttk.Label(
            side_panel,
            text='到货记录用于拆分在检、待入库、待收货；到货计划用于筛选当天计划未到货的采购跟催。',
            style='Subtle.TLabel',
            wraplength=280,
            justify='left',
        )
        note.grid(row=0, column=0, columnspan=2, sticky='w')

        action_row = ttk.Frame(side_panel)
        action_row.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(6, 0))
        action_row.columnconfigure(3, weight=1)
        self.arrival_run_btn = ttk.Button(action_row, text='开始分析', style='Accent.TButton', command=self._run_arrival_analysis)
        self.arrival_run_btn.grid(row=0, column=0, sticky='w')
        ttk.Button(action_row, text='清空结果', style='Quiet.TButton', command=self._clear_arrival_preview).grid(row=0, column=1, sticky='w', padx=(8, 0))
        ttk.Button(action_row, text='查看日志', style='Quiet.TButton', command=lambda: self._show_log_window('arrival')).grid(row=0, column=2, sticky='w', padx=(8, 0))
        ttk.Label(action_row, textvariable=self.arrival_summary_var, style='Metric.TLabel', wraplength=260, justify='left').grid(
            row=0, column=3, sticky='w', padx=(10, 0)
        )
        self.arrival_progress = ttk.Progressbar(action_row, mode='determinate', length=180, maximum=100)
        self._arrival_stage_counter = 0
        self.arrival_progress.grid(row=1, column=0, columnspan=4, sticky='ew', pady=(6, 0))
        self.arrival_progress.grid_remove()

        result = ttk.LabelFrame(f, text='到货跟催结果', style='Card.TLabelframe', padding=6)
        result.grid(row=1, column=0, sticky='nsew', padx=6, pady=(0, 6))
        result.columnconfigure(0, weight=1)
        result.rowconfigure(2, weight=1)

        stats = ttk.Frame(result)
        stats.grid(row=0, column=0, sticky='ew', pady=(0, 6))
        for idx in range(5):
            stats.columnconfigure(idx, weight=1)
        ttk.Label(stats, textvariable=self.arrival_quality_stat_var, style='Badge.TLabel').grid(row=0, column=0, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.arrival_warehouse_stat_var, style='Badge.TLabel').grid(row=0, column=1, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.arrival_purchase_pending_stat_var, style='Badge.TLabel').grid(row=0, column=2, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.arrival_all_stat_var, style='Badge.TLabel').grid(row=0, column=3, sticky='w', padx=(0, 8))
        ttk.Label(stats, textvariable=self.arrival_urgent_stat_var, style='Badge.TLabel').grid(row=0, column=4, sticky='w')

        tool_row = ttk.Frame(result)
        tool_row.grid(row=1, column=0, sticky='ew', pady=(0, 6))
        tool_row.columnconfigure(6, weight=1)
        ttk.Button(tool_row, text='紧急跟催', style='Quiet.TButton', command=lambda: self._set_arrival_filter_mode('urgent')).grid(row=0, column=0, sticky='w')
        ttk.Button(tool_row, text='全部物料', style='Quiet.TButton', command=lambda: self._set_arrival_filter_mode('all')).grid(row=0, column=1, sticky='w', padx=(8, 0))
        ttk.Button(tool_row, text='复制品质催检', style='Quiet.TButton', command=lambda: self._copy_arrival_followup_text('quality')).grid(row=0, column=2, sticky='w', padx=(16, 0))
        ttk.Button(tool_row, text='复制仓库催办', style='Quiet.TButton', command=lambda: self._copy_arrival_followup_text('warehouse')).grid(row=0, column=3, sticky='w', padx=(8, 0))
        ttk.Button(tool_row, text='复制采购未到货', style='Quiet.TButton', command=lambda: self._copy_arrival_followup_text('purchase_pending')).grid(row=0, column=4, sticky='w', padx=(8, 0))
        ttk.Button(tool_row, text='复制全部结果', style='Quiet.TButton', command=lambda: self._copy_arrival_followup_text('all')).grid(row=0, column=5, sticky='w', padx=(8, 0))

        result_nb = ttk.Notebook(result)
        result_nb.grid(row=2, column=0, sticky='nsew')

        quality_frame = ttk.Frame(result_nb)
        warehouse_frame = ttk.Frame(result_nb)
        purchase_pending_frame = ttk.Frame(result_nb)
        all_frame = ttk.Frame(result_nb)
        result_nb.add(quality_frame, text='品质催检')
        result_nb.add(warehouse_frame, text='待入库/待收货')
        result_nb.add(purchase_pending_frame, text='采购未到货跟催')
        result_nb.add(all_frame, text='全部命中')

        self.arrival_quality_tree = self._create_result_tree(quality_frame, ARRIVAL_STATUS_HEADERS, ARRIVAL_STATUS_WIDTHS)
        self.arrival_warehouse_tree = self._create_result_tree(warehouse_frame, ARRIVAL_STATUS_HEADERS, ARRIVAL_STATUS_WIDTHS)
        self.arrival_purchase_pending_tree = self._create_result_tree(purchase_pending_frame, ARRIVAL_PURCHASE_PENDING_HEADERS, ARRIVAL_PURCHASE_PENDING_WIDTHS)
        self.arrival_all_tree = self._create_result_tree(all_frame, ARRIVAL_STATUS_HEADERS, ARRIVAL_STATUS_WIDTHS)

        self._update_arrival_bom_badge()

    def _update_arrival_bom_badge(self):
        if not hasattr(self, 'arrival_bom_var'):
            return
        if not self.bom_index:
            self.arrival_bom_var.set('当前 BOM: 未加载，将使用输入文件里的 BOM')
            return
        source = os.path.basename(self.current_file) if self.current_file else '缓存'
        self.arrival_bom_var.set(f'当前 BOM: {source} | 母件 {len(self.bom_index)} 个')

    def _append_arrival_log(self, message):
        self._append_named_log('arrival', message)

    def _choose_arrival_input_file(self):
        path = filedialog.askopenfilename(
            title='选择 MRP 输入文件',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.arrival_input_var.set(path)
        self._append_arrival_log(f'已选择输入文件: {path}')

    def _choose_arrival_record_file(self):
        path = filedialog.askopenfilename(
            title='选择到货记录文件',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.arrival_record_var.set(path)
        self._append_arrival_log(f'已选择到货记录: {path}')

    def _choose_arrival_plan_file(self):
        path = filedialog.askopenfilename(
            title='选择到货计划文件',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if not path:
            return
        self.arrival_plan_var.set(path)
        self._append_arrival_log(f'已选择到货计划: {path}')

    def _clear_arrival_preview(self):
        for tree_name in ('arrival_quality_tree', 'arrival_warehouse_tree', 'arrival_purchase_pending_tree', 'arrival_all_tree'):
            tree = getattr(self, tree_name, None)
            if tree is not None:
                self._clear_tree_rows(tree)
        self._clear_named_log('arrival')
        self._arrival_last_result = None
        if hasattr(self, 'arrival_summary_var'):
            self.arrival_summary_var.set('等待分析')
            self.arrival_quality_stat_var.set('在检: 0')
            self.arrival_warehouse_stat_var.set('待入库/待收货: 0')
            self.arrival_purchase_pending_stat_var.set('采购未到货: 0')
            self.arrival_all_stat_var.set('命中物料: 0')
            self.arrival_urgent_stat_var.set('紧急跟催: 0')

    def _set_arrival_filter_mode(self, mode):
        if mode not in {'urgent', 'all'}:
            return
        self.arrival_filter_var.set(mode)
        if self._arrival_last_result is not None:
            self._refresh_arrival_result_views()

    def _run_arrival_analysis(self):
        if run_arrival_status_analysis is None:
            show_data_error(
                self.root,
                title='到货跟催引擎不可用',
                summary='未能加载到货跟催分析引擎（mrp_balance_tool），无法运行分析。',
                detail=BALANCE_PIPELINE_IMPORT_ERROR or '(未提供详细错误)',
                fix_hint='运行源码请确认同级目录有 "静态平衡表/src"；打包后请确认 spec 文件正确包含 mrp_balance_tool 模块。',
            )
            return
        if self._arrival_running:
            self.status_var.set('到货跟催分析正在执行中，请稍候')
            return

        input_path = self.arrival_input_var.get().strip()
        record_path = self.arrival_record_var.get().strip()
        plan_path = self.arrival_plan_var.get().strip()
        if not input_path:
            messagebox.showwarning('参数不完整', '请先选择 MRP 输入文件')
            return
        if not record_path:
            messagebox.showwarning('参数不完整', '请先选择到货记录文件')
            return
        analysis_date_text = self.arrival_analysis_date_var.get().strip()
        try:
            analysis_date = datetime.strptime(analysis_date_text, '%Y-%m-%d').date()
        except ValueError:
            messagebox.showwarning('日期格式不正确', '分析日期请填写为 YYYY-MM-DD，例如 2026-04-25')
            return

        external_bom_df = self.export_bom_dataframe() if self.bom_index else None
        window_days = max(int(self.arrival_window_days_var.get() or 0), 1)
        if not self._run_input_precheck_dialog(
            input_path,
            external_bom_df,
            log_callback=self._append_arrival_log,
        ):
            self.arrival_summary_var.set('已取消，请先检查源数据')
            self.status_var.set(self.arrival_summary_var.get())
            return

        self._arrival_running = True
        self.arrival_run_btn.state(['disabled'])
        self.arrival_progress.grid()
        self._arrival_stage_counter = 0
        self.arrival_progress['value'] = 0
        self.arrival_summary_var.set('正在分析到货状态，请稍候...')
        self.status_var.set(self.arrival_summary_var.get())
        self._append_arrival_log('开始分析到货状态...')

        self._arrival_thread = threading.Thread(
            target=self._run_arrival_analysis_worker,
            args=(input_path, record_path, plan_path, analysis_date, external_bom_df, window_days),
            daemon=True,
        )
        self._arrival_thread.start()

    def _run_arrival_analysis_worker(self, input_path, record_path, plan_path, analysis_date, external_bom_df, window_days):
        try:
            result = run_arrival_status_analysis(
                Path(input_path),
                Path(record_path),
                external_bom_df=external_bom_df,
                arrival_plan_path=Path(plan_path) if plan_path else None,
                analysis_date=analysis_date,
                window_days=window_days,
                progress_callback=lambda message: self.root.after(
                    0,
                    lambda msg=message: self._report_arrival_progress(msg),
                ),
            )
            self.root.after(0, lambda: self._finish_arrival_analysis(result))
        except Exception as exc:
            self.root.after(0, lambda: self._handle_arrival_error(str(exc)))

    def _report_arrival_progress(self, message):
        self._arrival_stage_counter += 1
        progress = min(95, self._arrival_stage_counter * 8)
        try:
            self.arrival_progress['value'] = progress
        except Exception:
            pass
        stage_label = f'[阶段 {self._arrival_stage_counter}] '
        self.arrival_summary_var.set(stage_label + message)
        self.status_var.set(stage_label + message)
        self._append_arrival_log(message)

    def _format_arrival_display_value(self, value):
        if value is None:
            return ''
        if isinstance(value, float) and pd.isna(value):
            return ''
        if isinstance(value, pd.Timestamp):
            if pd.isna(value):
                return ''
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if hasattr(value, 'strftime') and not isinstance(value, str):
            try:
                return value.strftime('%Y-%m-%d')
            except Exception:
                pass
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return fmt_qty(value)
        text = str(value).strip()
        return '' if text.lower() == 'nan' else text

    def _fill_arrival_tree(self, tree, dataframe, headers=ARRIVAL_STATUS_HEADERS):
        self._clear_tree_rows(tree)
        for record in dataframe.to_dict('records'):
            values = [self._format_arrival_display_value(record.get(header, '')) for header in headers]
            tree.insert('', 'end', values=values)

    def _filter_arrival_dataframe(self, dataframe):
        if dataframe is None or dataframe.empty:
            return dataframe.copy() if dataframe is not None else pd.DataFrame(columns=ARRIVAL_STATUS_HEADERS)
        if self.arrival_filter_var.get() != 'urgent':
            return dataframe.copy().reset_index(drop=True)
        urgent_mask = (
            pd.to_numeric(dataframe['未来3天需求'], errors='coerce').fillna(0) > 0
        ) | (
            pd.to_numeric(dataframe['未来3天缺口'], errors='coerce').fillna(0) > 0
        )
        return dataframe.loc[urgent_mask].copy().reset_index(drop=True)

    def _refresh_arrival_result_views(self):
        if self._arrival_last_result is None:
            return

        base_quality_df = getattr(self._arrival_last_result, 'quality_df', pd.DataFrame())
        base_warehouse_df = getattr(self._arrival_last_result, 'warehouse_df', pd.DataFrame())
        base_purchase_pending_df = getattr(self._arrival_last_result, 'purchase_pending_df', pd.DataFrame())
        base_all_df = getattr(self._arrival_last_result, 'all_df', pd.DataFrame())
        summary_info = getattr(self._arrival_last_result, 'summary', {}) or {}

        quality_df = self._filter_arrival_dataframe(base_quality_df)
        warehouse_df = self._filter_arrival_dataframe(base_warehouse_df)
        purchase_pending_df = self._filter_arrival_dataframe(base_purchase_pending_df)
        all_df = self._filter_arrival_dataframe(base_all_df)

        self._fill_arrival_tree(self.arrival_quality_tree, quality_df)
        self._fill_arrival_tree(self.arrival_warehouse_tree, warehouse_df)
        self._fill_arrival_tree(self.arrival_purchase_pending_tree, purchase_pending_df, ARRIVAL_PURCHASE_PENDING_HEADERS)
        self._fill_arrival_tree(self.arrival_all_tree, all_df)

        urgent_count = int(summary_info.get('urgent_count', 0) or 0)
        window_label = summary_info.get('window_label', '')
        mode_label = '紧急跟催' if self.arrival_filter_var.get() == 'urgent' else '全部物料'
        summary = f'分析完成 | 当前显示：{mode_label} | 时间窗口 {window_label} | 命中物料 {len(all_df)} 个 | 采购未到货 {len(purchase_pending_df)} 条'
        self.arrival_summary_var.set(summary)
        self.arrival_quality_stat_var.set(f'在检: {len(quality_df)}')
        self.arrival_warehouse_stat_var.set(f'待入库/待收货: {len(warehouse_df)}')
        self.arrival_purchase_pending_stat_var.set(f'采购未到货: {len(purchase_pending_df)}')
        self.arrival_all_stat_var.set(f'命中物料: {len(all_df)}')
        self.arrival_urgent_stat_var.set(f'紧急跟催: {urgent_count}')
        self.status_var.set(summary)
        self._append_arrival_log(summary)

    def _finish_arrival_analysis(self, result):
        self._arrival_running = False
        self.arrival_progress['value'] = 100
        self.arrival_progress.grid_remove()
        self.arrival_run_btn.state(['!disabled'])
        self._arrival_last_result = result
        self._refresh_arrival_result_views()

    def _handle_arrival_error(self, message):
        self._arrival_running = False
        self.arrival_progress['value'] = 0
        self.arrival_progress.grid_remove()
        self.arrival_run_btn.state(['!disabled'])
        self.arrival_summary_var.set('到货跟催分析失败')
        self.status_var.set('到货跟催分析失败')
        self._append_arrival_log(f'分析失败: {message}')
        show_data_error(
            self.root,
            title='到货跟催分析失败',
            summary='到货跟催分析未能完成。',
            detail=message,
            fix_hint=_suggest_fix_for_exception(Exception(message)),
        )

    def _copy_arrival_followup_text(self, kind):
        if self._arrival_last_result is None:
            messagebox.showwarning('暂无结果', '请先完成到货跟催分析')
            return

        tree_map = {
            'quality': (self.arrival_quality_tree, '品质催检', ARRIVAL_STATUS_HEADERS),
            'warehouse': (self.arrival_warehouse_tree, '待入库/待收货', ARRIVAL_STATUS_HEADERS),
            'purchase_pending': (self.arrival_purchase_pending_tree, '采购未到货跟催', ARRIVAL_PURCHASE_PENDING_HEADERS),
            'all': (self.arrival_all_tree, '全部命中', ARRIVAL_STATUS_HEADERS),
        }
        tree, title, headers = tree_map.get(kind, (self.arrival_all_tree, '全部命中', ARRIVAL_STATUS_HEADERS))
        item_ids = tree.selection() or tree.get_children('')
        if not item_ids:
            messagebox.showinfo('没有可复制的数据', f'{title} 当前没有结果')
            return

        lines = ['\t'.join(headers)]
        for item_id in item_ids:
            values = tree.item(item_id, 'values')
            row_values = [str(values[idx]).strip() if idx < len(values) else '' for idx in range(len(headers))]
            lines.append('\t'.join(row_values))

        text = '\n'.join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_var.set(f'已复制{title}表格，共 {len(item_ids)} 条')
        self._append_arrival_log(f'已复制{title}表格，共 {len(item_ids)} 条')

    def _safe_float(self, val):
        try:
            if val is None:
                return 0.0
            s = str(val).strip().replace(',', '')
            if not s:
                return 0.0
            return float(s)
        except Exception:
            return 0.0

    def _get_bom_meta_cache(self):
        if self._diff_meta_cache is not None:
            return self._diff_meta_cache
        meta = {}
        for rows in self.bom_index.values():
            for r in rows:
                child = str(r[CHILD_PN_COL]).strip() if len(r) > CHILD_PN_COL and r[CHILD_PN_COL] else ''
                if not child or child in meta:
                    continue
                name = str(r[CHILD_NAME_COL]).strip() if len(r) > CHILD_NAME_COL and r[CHILD_NAME_COL] else ''
                spec = str(r[CHILD_SPEC_COL]).strip() if len(r) > CHILD_SPEC_COL and r[CHILD_SPEC_COL] else ''
                meta[child] = (name, spec)
        self._diff_meta_cache = meta
        return meta

    def _get_bom_root_meta(self, code):
        rows = self.bom_index.get(code) or []
        if not rows:
            return '', ''
        row = rows[0]
        name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        spec = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        return name, spec

    def _get_diff_code_label_map(self, codes):
        labels = {}
        for code in codes:
            name, _spec = self._get_bom_root_meta(code)
            labels[code] = f'{name}\n{code}' if name else code
        return labels

    def _get_bom_edges_cache(self):
        if self._bom_edges_cache is not None:
            return self._bom_edges_cache

        edges = {}
        for parent_pn, rows in self.bom_index.items():
            items = []
            for row in rows:
                child = str(row[CHILD_PN_COL]).strip() if len(row) > CHILD_PN_COL and row[CHILD_PN_COL] else ''
                if not child:
                    continue
                qty = self._safe_float(row[CHILD_QTY_COL] if len(row) > CHILD_QTY_COL else 0)
                if qty == 0:
                    continue
                name = str(row[CHILD_NAME_COL]).strip() if len(row) > CHILD_NAME_COL and row[CHILD_NAME_COL] else ''
                spec = str(row[CHILD_SPEC_COL]).strip() if len(row) > CHILD_SPEC_COL and row[CHILD_SPEC_COL] else ''
                items.append((child, qty, name, spec))
            edges[parent_pn] = items

        self._bom_edges_cache = edges
        return edges

    def _explode_bom(self, root_pn, max_depth=80):
        edges = self._get_bom_edges_cache()
        warnings = []
        memo = {}
        warned_cycles = set()

        def unit_totals(node_pn, depth_left, trail):
            if depth_left <= 0:
                warnings.append(f'超过最大递归层级，停止展开: {node_pn}')
                return {}

            key = (node_pn, depth_left)
            if key in memo:
                return memo[key]

            totals = defaultdict(float)
            for child, qty, _name, _spec in edges.get(node_pn, []):
                totals[child] += qty
                if child in trail:
                    cycle_key = (node_pn, child)
                    if cycle_key not in warned_cycles:
                        warnings.append(f'检测到循环引用，已跳过继续展开: {node_pn} -> {child}')
                        warned_cycles.add(cycle_key)
                    continue
                if child in edges:
                    sub_totals = unit_totals(child, depth_left - 1, trail | {child})
                    for sub_pn, sub_qty in sub_totals.items():
                        totals[sub_pn] += qty * sub_qty

            memo[key] = dict(totals)
            return memo[key]

        return unit_totals(root_pn, max_depth, {root_pn}), warnings

    def _parse_diff_code_list(self, raw_text):
        parts = re.split(r'[\s,，;；、]+', str(raw_text or '').strip())
        result = []
        seen = set()
        for part in parts:
            code = part.strip()
            if not code or code in seen:
                continue
            seen.add(code)
            result.append(code)
        return result

    def _format_diff_group_label(self, codes, limit=3):
        if not codes:
            return ''
        preview = '、'.join(codes[:limit])
        if len(codes) > limit:
            preview += f' 等{len(codes)}个'
        return preview

    def _explode_bom_group(self, root_codes):
        totals = defaultdict(float)
        warnings = []
        for root_code in root_codes:
            root_totals, root_warnings = self._explode_bom(root_code)
            for pn, qty in root_totals.items():
                totals[pn] += qty
            warnings.extend([f'{root_code}: {msg}' for msg in root_warnings])
        return dict(totals), warnings

    def _find_diff_root_codes_by_prefix(self, prefix):
        prefix = str(prefix or '').strip()
        if not prefix:
            return []
        return sorted(str(code).strip() for code in self.bom_index.keys() if str(code).strip().startswith(prefix))

    def _parse_bom_doc_matrix_file(self, path):
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        try:
            ws = wb['物料清单'] if '物料清单' in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            sheet_name = ws.title
        finally:
            wb.close()
        if not rows:
            raise ValueError('文件为空')

        def cell_text(row_index, col_index):
            if row_index is None or row_index < 0 or row_index >= len(rows):
                return ''
            row = rows[row_index] or []
            if col_index < 0 or col_index >= len(row):
                return ''
            value = row[col_index]
            return '' if value is None else str(value).strip()

        def find_row_by_labels(labels, max_row_exclusive=None):
            label_set = {str(label).strip() for label in labels}
            upper = len(rows) if max_row_exclusive is None else min(max_row_exclusive, len(rows))
            for ri in range(upper):
                for value in rows[ri] or []:
                    if str(value).strip() in label_set:
                        return ri
            return None

        def cell_obj(row_index, col_index):
            if row_index is None or row_index < 0:
                return None
            return ws.cell(row=row_index + 1, column=col_index + 1)

        def is_red_fill(cell):
            if cell is None or cell.fill is None or cell.fill.fill_type is None:
                return False
            color = cell.fill.fgColor
            rgb = ''
            if color is not None and color.type == 'rgb' and color.rgb:
                rgb = str(color.rgb).upper()
            return rgb in {'FFFF0000', 'FFFF6666', 'FFFFC7CE'} or rgb.endswith('FF0000')

        def is_struck(cell):
            return bool(cell is not None and cell.font is not None and cell.font.strike)

        header_row_index = None
        material_col = name_col = spec_col = None
        quantity_cols = []
        for ri, row in enumerate(rows):
            labels = [str(value).strip() if value is not None else '' for value in row or []]
            if '物料编号' not in labels:
                continue
            qty_candidates = [ci for ci, label in enumerate(labels) if label == '数量']
            if qty_candidates:
                header_row_index = ri
                material_col = labels.index('物料编号')
                name_col = labels.index('名称') if '名称' in labels else None
                spec_col = labels.index('型号') if '型号' in labels else None
                quantity_cols = qty_candidates
                break
        if header_row_index is None or material_col is None or not quantity_cols:
            raise ValueError('未识别到量产BOM表头，请确认表内包含“物料编号”和多个“数量”列')

        root_row = find_row_by_labels(['在制品号', '在制号'], header_row_index)
        product_row = find_row_by_labels(['成品号'], header_row_index)
        name_row = find_row_by_labels(['在制名称', '名称'], header_row_index)
        model_row = find_row_by_labels(['在制型号', '型号'], header_row_index)
        config_row = find_row_by_labels(['配置说明'], header_row_index)

        roots = []
        seen_roots = set()
        skipped_root_count = 0
        for col_index in quantity_cols:
            header_cells = [cell_obj(row_index, col_index) for row_index in range(0, min(header_row_index + 1, 8))]
            deleted_root = any(is_red_fill(cell) or is_struck(cell) for cell in header_cells)
            root_code = normalize_material_code(cell_text(root_row, col_index))
            product_code = normalize_material_code(cell_text(product_row, col_index))
            compare_code = root_code or product_code
            if deleted_root:
                if compare_code:
                    skipped_root_count += 1
                continue
            if not compare_code or compare_code in seen_roots:
                continue
            seen_roots.add(compare_code)
            roots.append({
                'code': compare_code,
                'product_code': product_code,
                'name': cell_text(name_row, col_index),
                'model': cell_text(model_row, col_index),
                'config': cell_text(config_row, col_index),
                'column': col_index,
                'column_label': cell_text(1, col_index),
            })

        if not roots:
            raise ValueError('未读取到可对比的04在制品号/成品号')

        totals_by_root = {root['code']: defaultdict(float) for root in roots}
        meta_by_code = {}
        root_by_col = {root['column']: root for root in roots}
        skipped_material_count = 0
        for row_index, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 1):
            if not row or material_col >= len(row):
                continue
            material_code = normalize_material_code(row[material_col])
            if not material_code or material_code in ('物料编号', '料号'):
                continue
            key_cells = [
                cell_obj(row_index, material_col),
                cell_obj(row_index, name_col) if name_col is not None else None,
                cell_obj(row_index, spec_col) if spec_col is not None else None,
            ]
            core_cells = [
                cell_obj(row_index, col_index)
                for col_index in range(0, min(12, ws.max_column))
                if cell_obj(row_index, col_index) is not None
            ]
            struck_core_count = sum(1 for cell in core_cells if is_struck(cell))
            if any(is_struck(cell) for cell in key_cells) or struck_core_count >= 3:
                skipped_material_count += 1
                continue
            material_name = ''
            material_spec = ''
            if name_col is not None and name_col < len(row) and row[name_col] is not None:
                material_name = str(row[name_col]).strip()
            if spec_col is not None and spec_col < len(row) and row[spec_col] is not None:
                material_spec = str(row[spec_col]).strip()
            if material_code not in meta_by_code:
                meta_by_code[material_code] = (material_name, material_spec)
            for col_index, root in root_by_col.items():
                qty_raw = row[col_index] if col_index < len(row) else None
                qty = self._safe_float(qty_raw)
                if abs(qty) <= 1e-12:
                    continue
                totals_by_root[root['code']][material_code] += qty

        return {
            'source': path,
            'source_label': os.path.basename(path),
            'sheet': sheet_name,
            'roots': roots,
            'totals': {code: dict(values) for code, values in totals_by_root.items()},
            'meta': meta_by_code,
            'skipped_root_count': skipped_root_count,
            'skipped_material_count': skipped_material_count,
        }

    def _compare_bom_doc_matrix(self, parsed):
        rows = []
        summary = []
        missing_roots = []
        source_label = parsed.get('source_label', os.path.basename(str(parsed.get('source', ''))) or '量产BOM')
        bom_meta = self._get_bom_meta_cache()
        type_priority = {
            '当前BOM缺母件': 0,
            'DOC有/BOM没有': 1,
            'BOM有/DOC没有': 2,
            '用量不同': 3,
            '名称规格不同': 4,
        }

        def norm_text(value):
            return re.sub(r'\s+', '', str(value or '').strip())

        for root in parsed.get('roots', []):
            root_code = root.get('code', '')
            doc_totals = parsed.get('totals', {}).get(root_code, {}) or {}
            if root_code not in self.bom_index:
                missing_roots.append({
                    '来源BOM': source_label,
                    '母件料号': root_code,
                    '成品号': root.get('product_code', ''),
                    '母件名称': root.get('name', ''),
                    '母件型号': root.get('model', ''),
                    '配置说明': root.get('config', ''),
                    'DOC物料数': len(doc_totals),
                    '问题说明': '当前BOMMaster没有这个母件，无法展开对比',
                })
                rows.append({
                    'source_label': source_label,
                    'root': root_code,
                    'root_name': root.get('name', ''),
                    'root_model': root.get('model', ''),
                    'type': '当前BOM缺母件',
                    'pn': '',
                    'doc_name': '',
                    'bom_name': '',
                    'doc_spec': '',
                    'bom_spec': '',
                    'doc_qty': 0,
                    'bom_qty': 0,
                    'diff': 0,
                    'note': '当前BOMMaster没有该04母件，整列未参与逐项对比',
                })
                summary.append({
                    '来源BOM': source_label,
                    '母件料号': root_code,
                    '成品号': root.get('product_code', ''),
                    '母件名称': root.get('name', ''),
                    '母件型号': root.get('model', ''),
                    '配置说明': root.get('config', ''),
                    'DOC物料数': len(doc_totals),
                    'BOM物料数': 0,
                    '差异行数': 1,
                    'DOC有BOM没有': 0,
                    'BOM有DOC没有': 0,
                    '用量不同': 0,
                    '名称规格不同': 0,
                    '警告': '当前BOM缺母件',
                })
                continue

            bom_totals, warnings = self._explode_bom(root_code)
            counts = defaultdict(int)
            diff_count = 0
            for material_code in sorted(set(doc_totals) | set(bom_totals)):
                doc_qty = float(doc_totals.get(material_code, 0) or 0)
                bom_qty = float(bom_totals.get(material_code, 0) or 0)
                diff_qty = doc_qty - bom_qty
                doc_name, doc_spec = parsed.get('meta', {}).get(material_code, ('', ''))
                bom_name, bom_spec = bom_meta.get(material_code, ('', ''))
                diff_type = ''
                note = ''
                if doc_qty and not bom_qty:
                    diff_type = 'DOC有/BOM没有'
                    note = '量产BOM表有此物料，当前BOM展开后没有'
                elif bom_qty and not doc_qty:
                    diff_type = 'BOM有/DOC没有'
                    note = '当前BOM展开后有此物料，量产BOM表没有'
                elif abs(diff_qty) > 1e-9:
                    diff_type = '用量不同'
                    note = '两边单台汇总用量不一致'
                elif (doc_name and bom_name and norm_text(doc_name) != norm_text(bom_name)) or (
                    doc_spec and bom_spec and norm_text(doc_spec) != norm_text(bom_spec)
                ):
                    diff_type = '名称规格不同'
                    note = '用量一致，但名称或规格文本不一致'

                if not diff_type:
                    continue
                counts[diff_type] += 1
                diff_count += 1
                rows.append({
                    'source_label': source_label,
                    'root': root_code,
                    'root_name': root.get('name', ''),
                    'root_model': root.get('model', ''),
                    'type': diff_type,
                    'pn': material_code,
                    'doc_name': doc_name,
                    'bom_name': bom_name,
                    'doc_spec': doc_spec,
                    'bom_spec': bom_spec,
                    'doc_qty': doc_qty,
                    'bom_qty': bom_qty,
                    'diff': diff_qty,
                    'note': note,
                })

            summary.append({
                '来源BOM': source_label,
                '母件料号': root_code,
                '成品号': root.get('product_code', ''),
                '母件名称': root.get('name', ''),
                '母件型号': root.get('model', ''),
                '配置说明': root.get('config', ''),
                'DOC物料数': len([code for code, qty in doc_totals.items() if abs(float(qty or 0)) > 1e-12]),
                'BOM物料数': len([code for code, qty in bom_totals.items() if abs(float(qty or 0)) > 1e-12]),
                '差异行数': diff_count,
                'DOC有BOM没有': counts['DOC有/BOM没有'],
                'BOM有DOC没有': counts['BOM有/DOC没有'],
                '用量不同': counts['用量不同'],
                '名称规格不同': counts['名称规格不同'],
                '警告': '；'.join(warnings[:5]),
            })

        rows.sort(key=lambda item: (item.get('root', ''), type_priority.get(item.get('type', ''), 9), item.get('pn', '')))
        return rows, summary, missing_roots

    def _configure_diff_tree(self, columns, widths, numeric_columns=None):
        if not hasattr(self, 'diff_tree'):
            return
        try:
            self.diff_tree.delete(*self.diff_tree.get_children(''))
        except tk.TclError:
            pass
        self.diff_tree['columns'] = list(columns)
        for h, w in zip(columns, widths):
            anchor = 'e' if h in set(numeric_columns or set()) or '用量' in h or h.startswith('差值') else 'w'
            self.diff_tree.column(h, width=w, anchor=anchor)
            self.diff_tree.heading(h, text=h)
        enable_treeview_sort(self.diff_tree, list(columns), numeric_columns=set(numeric_columns or set()))
        enable_treeview_copy(self.diff_tree, list(columns))

    def _run_diff_compare(self):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        if self._diff_compare_running:
            self.status_var.set('差异对比正在执行中，请稍候')
            return

        a = self.diff_a_var.get().strip()
        b = self.diff_b_var.get().strip()
        if not a or not b:
            messagebox.showwarning('参数不完整', '请同时输入物料编码 A / B')
            return
        if a not in self.bom_index:
            messagebox.showwarning('料号不存在', f'母料号不存在: {a}')
            return
        if b not in self.bom_index:
            messagebox.showwarning('料号不存在', f'母料号不存在: {b}')
            return

        self._diff_mode = 'pair'
        self._configure_diff_tree(
            HEADERS_DIFF_GRID,
            COL_WIDTHS_D_GRID,
            numeric_columns={'A总用量', 'B总用量', '差值(A-B)'},
        )
        self._diff_compare_running = True
        self.diff_run_btn.state(['disabled'])
        self.diff_export_btn.state(['disabled'])
        self.diff_progress['value'] = 5
        self.diff_summary_var.set(f'[阶段 1/4] 正在对比 {a} 与 {b}，请稍候...')
        self.status_var.set('差异对比计算中...')

        worker = threading.Thread(target=self._run_diff_compare_worker, args=([a], [b]), daemon=True)
        worker.start()

    def _run_bom_doc_compare(self):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        if self._diff_compare_running:
            self.status_var.set('差异对比正在执行中，请稍候')
            return
        paths = filedialog.askopenfilenames(
            title='选择量产BOM表（可多选）',
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')],
        )
        if not paths:
            return
        paths = list(paths)
        self._diff_mode = 'bom_doc'
        self._bom_doc_compare_rows = []
        self._bom_doc_compare_summary = []
        self._bom_doc_compare_missing_roots = []
        self._bom_doc_compare_source = '\n'.join(paths)
        self._bom_doc_compare_results = []
        self._configure_diff_tree(
            BOM_DOC_COMPARE_HEADERS,
            BOM_DOC_COMPARE_WIDTHS,
            numeric_columns={'DOC用量', 'BOM用量', '差值(DOC-BOM)'},
        )
        self._diff_compare_running = True
        for btn_name in ('diff_run_btn', 'diff_export_btn', 'diff_unified_run_btn', 'diff_prefix_run_btn', 'diff_doc_compare_btn'):
            if hasattr(self, btn_name):
                getattr(self, btn_name).state(['disabled'])
        self.diff_progress['value'] = 5
        self.diff_summary_var.set(f'[阶段 1/4] 正在读取量产BOM表：{len(paths)} 个文件')
        self.status_var.set('量产BOM表对比中...')
        worker = threading.Thread(target=self._run_bom_doc_compare_worker, args=(paths,), daemon=True)
        worker.start()

    def _set_diff_stage(self, stage: int, total: int, text: str):
        """在主线程里更新 diff 进度条与阶段文本。"""
        try:
            self.diff_progress['value'] = int(stage * 100 / total)
        except Exception:
            pass
        self.diff_summary_var.set(f'[阶段 {stage}/{total}] {text}')
        self.status_var.set(text)

    def _run_unified_diff_compare(self):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        if self._diff_compare_running:
            self.status_var.set('差异对比正在执行中，请稍候')
            return

        raw_text = self.diff_unified_text.get('1.0', tk.END) if hasattr(self, 'diff_unified_text') else ''
        codes = self._parse_diff_code_list(raw_text)
        if len(codes) < 2:
            messagebox.showwarning('参数不完整', '统一对比至少需要粘贴 2 个母料号')
            return
        missing = [code for code in codes if code not in self.bom_index]
        if missing:
            messagebox.showwarning('料号不存在', '以下母料号不存在：\n' + '\n'.join(missing[:30]))
            return
        self._start_unified_diff_compare(codes, source_label='粘贴料号')

    def _run_prefix_diff_compare(self):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        if self._diff_compare_running:
            self.status_var.set('差异对比正在执行中，请稍候')
            return
        prefix = self.diff_prefix_var.get().strip() if hasattr(self, 'diff_prefix_var') else ''
        if not prefix:
            messagebox.showwarning('参数不完整', '请输入分类/料号前缀，例如 040101')
            return
        codes = self._find_diff_root_codes_by_prefix(prefix)
        if len(codes) < 2:
            messagebox.showwarning('匹配不足', f'前缀 {prefix} 只匹配到 {len(codes)} 个母料号，至少需要 2 个才能对比')
            return
        self._start_unified_diff_compare(codes, source_label=f'前缀 {prefix}')

    def _start_unified_diff_compare(self, codes, source_label='统一对比'):
        self._diff_mode = 'unified'
        self._diff_code_label_map = self._get_diff_code_label_map(codes)
        code_columns = [self._diff_code_label_map.get(code, code) for code in codes]
        unified_columns = ['共用状态', '物料编码', '品名', '规格', '出现BOM数'] + code_columns + ['最大用量', '最小用量', '差值(最大-最小)']
        unified_widths = [90, 140, 180, 160, 80] + [180] * len(codes) + [90, 90, 120]
        numeric_columns = set(code_columns) | {'出现BOM数', '最大用量', '最小用量', '差值(最大-最小)'}
        self._configure_diff_tree(unified_columns, unified_widths, numeric_columns=numeric_columns)
        self._diff_unified_source = source_label

        self._diff_compare_running = True
        self.diff_run_btn.state(['disabled'])
        self.diff_export_btn.state(['disabled'])
        if hasattr(self, 'diff_unified_run_btn'):
            self.diff_unified_run_btn.state(['disabled'])
        if hasattr(self, 'diff_prefix_run_btn'):
            self.diff_prefix_run_btn.state(['disabled'])
        if hasattr(self, 'diff_doc_compare_btn'):
            self.diff_doc_compare_btn.state(['disabled'])
        self.diff_progress['value'] = 5
        self.diff_summary_var.set(f'[阶段 1/4] 正在统一对比 {len(codes)} 个母料号（{source_label}），请稍候...')
        self.status_var.set('统一对比计算中...')

        worker = threading.Thread(target=self._run_unified_diff_compare_worker, args=(codes,), daemon=True)
        worker.start()

    def _run_diff_compare_worker(self, a_codes, b_codes):
        try:
            self.root.after(0, lambda: self._set_diff_stage(1, 4, '读取 BOM 元信息...'))
            meta = self._get_bom_meta_cache()
            a_label = self._format_diff_group_label(a_codes)
            b_label = self._format_diff_group_label(b_codes)
            if len(a_codes) == 1 and len(b_codes) == 1:
                self.root.after(0, lambda: self._set_diff_stage(2, 4, f'展开 BOM：{a_label}'))
            else:
                self.root.after(0, lambda: self._set_diff_stage(2, 4, f'展开 A组 BOM：{a_label}'))
            qty_a, warn_a = self._explode_bom_group(a_codes)
            if len(a_codes) == 1 and len(b_codes) == 1:
                self.root.after(0, lambda: self._set_diff_stage(3, 4, f'展开 BOM：{b_label}'))
            else:
                self.root.after(0, lambda: self._set_diff_stage(3, 4, f'展开 B组 BOM：{b_label}'))
            qty_b, warn_b = self._explode_bom_group(b_codes)
            self.root.after(0, lambda: self._set_diff_stage(4, 4, '比对数量差异...'))
            keys = sorted(set(qty_a.keys()) | set(qty_b.keys()))

            count_only_a = 0
            count_only_b = 0
            count_diff = 0
            count_common = 0
            aligned_rows = []
            for pn in keys:
                qa = round(qty_a.get(pn, 0.0), 6)
                qb = round(qty_b.get(pn, 0.0), 6)
                name, spec = meta.get(pn, ('', ''))
                if qb == 0:
                    diff_type = '仅A有'
                    count_only_a += 1
                elif qa == 0:
                    diff_type = '仅B有'
                    count_only_b += 1
                elif abs(qa - qb) < 1e-12:
                    diff_type = '共用物料'
                    count_common += 1
                else:
                    diff_type = '用量不同'
                    count_diff += 1
                    count_common += 1
                aligned_rows.append({
                    'type': diff_type,
                    'pn': pn,
                    'name': name,
                    'spec': spec,
                    'qa': qa,
                    'qb': qb,
                    'diff': round(qa - qb, 6),
                })

            type_order = {'仅A有': 0, '仅B有': 1, '用量不同': 2, '共用物料': 3}
            aligned_rows.sort(key=lambda item: (type_order.get(item['type'], 99), item['pn']))
            self.root.after(
                0,
                lambda rows=aligned_rows, warns=warn_a + warn_b, ac=list(a_codes), bc=list(b_codes):
                    self._finish_diff_compare(rows, count_only_a, count_only_b, count_diff, count_common, warns, ac, bc)
            )
        except Exception as e:
            self.root.after(0, lambda: self._handle_diff_compare_error(str(e)))

    def _run_bom_doc_compare_worker(self, paths):
        try:
            paths = list(paths or [])
            all_rows = []
            all_summary = []
            all_missing_roots = []
            results = []
            total_files = len(paths)
            for file_index, path in enumerate(paths, start=1):
                file_label = os.path.basename(path)
                self.root.after(
                    0,
                    lambda idx=file_index, total=total_files, label=file_label: self._set_diff_stage(
                        1,
                        4,
                        f'读取量产BOM表 {idx}/{total}：{label}',
                    ),
                )
                parsed = self._parse_bom_doc_matrix_file(path)
                root_count = len(parsed.get('roots', []))
                self.root.after(
                    0,
                    lambda idx=file_index, total=total_files, count=root_count: self._set_diff_stage(
                        2,
                        4,
                        f'展开当前BOM {idx}/{total}：{count} 个04母件',
                    ),
                )
                rows, summary, missing_roots = self._compare_bom_doc_matrix(parsed)
                all_rows.extend(rows)
                all_summary.extend(summary)
                all_missing_roots.extend(missing_roots)
                results.append({
                    'source': path,
                    'source_label': parsed.get('source_label', file_label),
                    'sheet': parsed.get('sheet', ''),
                    'rows': rows,
                    'summary': summary,
                    'missing_roots': missing_roots,
                    'root_count': root_count,
                    'skipped_root_count': parsed.get('skipped_root_count', 0),
                    'skipped_material_count': parsed.get('skipped_material_count', 0),
                })
            self.root.after(0, lambda: self._set_diff_stage(3, 4, '汇总多文件差异明细...'))
            self.root.after(
                0,
                lambda: self._finish_bom_doc_compare(
                    all_rows,
                    all_summary,
                    all_missing_roots,
                    '\n'.join(paths),
                    f'{total_files}个文件',
                    results,
                ),
            )
        except Exception as e:
            self.root.after(0, lambda: self._handle_diff_compare_error(str(e)))

    def _run_unified_diff_compare_worker(self, codes):
        try:
            self.root.after(0, lambda: self._set_diff_stage(1, 4, '读取 BOM 元信息...'))
            meta = self._get_bom_meta_cache()
            exploded = {}
            warnings = []
            total = max(len(codes), 1)
            for idx, code in enumerate(codes, 1):
                self.root.after(0, lambda idx=idx, code=code: self._set_diff_stage(2, 4, f'展开统一对比 BOM：{idx}/{total} {code}'))
                qty_map, warn = self._explode_bom(code)
                exploded[code] = qty_map
                warnings.extend([f'{code}: {msg}' for msg in warn])

            self.root.after(0, lambda: self._set_diff_stage(3, 4, '汇总多料号用量矩阵...'))
            all_materials = sorted({pn for qty_map in exploded.values() for pn in qty_map.keys()})
            rows = []
            count_all_same = 0
            count_all_diff = 0
            count_partial_common = 0
            count_unique = 0
            for pn in all_materials:
                values = [round(float(exploded[code].get(pn, 0.0) or 0.0), 6) for code in codes]
                non_zero = [value for value in values if abs(value) >= 1e-12]
                present_count = len(non_zero)
                max_qty = max(values) if values else 0.0
                min_qty = min(values) if values else 0.0
                range_diff = round(max_qty - min_qty, 6)
                if present_count == len(codes):
                    if abs(range_diff) < 1e-12:
                        status = '全部共用'
                        count_all_same += 1
                    else:
                        status = '全部共用-用量不同'
                        count_all_diff += 1
                elif present_count >= 2:
                    status = '部分共用'
                    count_partial_common += 1
                else:
                    status = '独有'
                    count_unique += 1
                name, spec = meta.get(pn, ('', ''))
                rows.append({
                    'status': status,
                    'pn': pn,
                    'name': name,
                    'spec': spec,
                    'present_count': present_count,
                    'values': dict(zip(codes, values)),
                    'max_qty': round(max_qty, 6),
                    'min_qty': round(min_qty, 6),
                    'range_diff': range_diff,
                })

            order = {'全部共用-用量不同': 0, '部分共用': 1, '独有': 2, '全部共用': 3}
            rows.sort(key=lambda item: (order.get(item['status'], 99), item['pn']))
            self.root.after(0, lambda: self._set_diff_stage(4, 4, '渲染统一对比结果...'))
            self.root.after(
                0,
                lambda rows=rows, warns=warnings, codes=list(codes):
                    self._finish_unified_diff_compare(rows, codes, count_all_same, count_all_diff, count_partial_common, count_unique, warns)
            )
        except Exception as e:
            self.root.after(0, lambda: self._handle_diff_compare_error(str(e)))

    def _finish_diff_compare(self, aligned_rows, count_only_a, count_only_b, count_diff, count_common, warns, a_codes=None, b_codes=None):
        self.diff_tree.delete(*self.diff_tree.get_children(''))
        self._diff_rows = aligned_rows
        self._diff_common_rows = [row for row in aligned_rows if row.get('qa', 0) != 0 and row.get('qb', 0) != 0]
        self._diff_unified_rows = []
        self._diff_input_codes = {'a': list(a_codes or []), 'b': list(b_codes or [])}
        self._diff_export_rows = []
        self._diff_pending_rows = list(aligned_rows)
        self._diff_insert_index = 0
        self._diff_counts = (count_only_a, count_only_b, count_diff, count_common, warns)
        self._render_diff_rows_chunk()

    def _finish_unified_diff_compare(self, rows, codes, count_all_same, count_all_diff, count_partial_common, count_unique, warns):
        self.diff_tree.delete(*self.diff_tree.get_children(''))
        self._diff_mode = 'unified'
        self._diff_rows = []
        self._diff_common_rows = []
        self._diff_unified_rows = list(rows)
        code_labels = getattr(self, '_diff_code_label_map', {}) or self._get_diff_code_label_map(codes)
        self._diff_input_codes = {
            'unified': list(codes),
            'source': getattr(self, '_diff_unified_source', ''),
            'code_labels': dict(code_labels),
        }
        self._diff_pending_rows = list(rows)
        self._diff_insert_index = 0
        self._diff_counts = (count_all_same, count_all_diff, count_partial_common, count_unique, warns)
        self._render_unified_diff_rows_chunk()

    def _finish_bom_doc_compare(self, rows, summary, missing_roots, source_path, sheet_name, results=None):
        self.diff_tree.delete(*self.diff_tree.get_children(''))
        self._diff_mode = 'bom_doc'
        self._diff_rows = []
        self._diff_common_rows = []
        self._diff_unified_rows = []
        self._diff_export_rows = []
        self._bom_doc_compare_rows = list(rows)
        self._bom_doc_compare_summary = list(summary)
        self._bom_doc_compare_missing_roots = list(missing_roots)
        self._bom_doc_compare_source = source_path
        self._bom_doc_compare_results = list(results or [])
        self._diff_pending_rows = list(rows)
        self._diff_insert_index = 0
        self._diff_counts = (len(rows), len(summary), len(missing_roots), sheet_name)
        self._render_bom_doc_compare_rows_chunk()

    def _render_diff_rows_chunk(self, chunk_size=150):
        start = self._diff_insert_index
        end = min(start + chunk_size, len(self._diff_pending_rows))
        for idx in range(start, end):
            item = self._diff_pending_rows[idx]
            rid = f'diff_{idx}'
            row = [
                item['type'],
                item['pn'] if item['qa'] != 0 else '',
                item['name'] if item['qa'] != 0 else '',
                item['spec'] if item['qa'] != 0 else '',
                fmt_qty(item['qa']) if item['qa'] != 0 else '',
                item['pn'] if item['qb'] != 0 else '',
                item['name'] if item['qb'] != 0 else '',
                item['spec'] if item['qb'] != 0 else '',
                fmt_qty(item['qb']) if item['qb'] != 0 else '',
                fmt_qty(item['diff']),
            ]
            self.diff_tree.insert('', 'end', iid=rid, values=row)
            self._diff_export_rows.append([
                item['type'], item['pn'], item['name'], item['spec'],
                fmt_qty(item['qa']), fmt_qty(item['qb']), fmt_qty(item['diff']),
            ])

        self._diff_insert_index = end
        if end < len(self._diff_pending_rows):
            self.diff_summary_var.set(f'正在渲染结果... {end}/{len(self._diff_pending_rows)}')
            self.root.after(1, self._render_diff_rows_chunk)
            return

        count_only_a, count_only_b, count_diff, count_common, warns = self._diff_counts
        summary = f'对比完成 | 仅A有 {count_only_a} 条, 仅B有 {count_only_b} 条, 共用 {count_common} 条, 其中用量不同 {count_diff} 条'
        if warns:
            summary += f' | 警告 {len(warns)} 条'
        self.diff_progress['value'] = 100
        self.diff_run_btn.state(['!disabled'])
        self.diff_export_btn.state(['!disabled'])
        if hasattr(self, 'diff_unified_run_btn'):
            self.diff_unified_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_prefix_run_btn'):
            self.diff_prefix_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_doc_compare_btn'):
            self.diff_doc_compare_btn.state(['!disabled'])
        self._diff_compare_running = False
        self.diff_summary_var.set(summary)
        self.status_var.set(summary)

    def _render_unified_diff_rows_chunk(self, chunk_size=150):
        start = self._diff_insert_index
        end = min(start + chunk_size, len(self._diff_pending_rows))
        codes = getattr(self, '_diff_input_codes', {}).get('unified', [])
        for idx in range(start, end):
            item = self._diff_pending_rows[idx]
            row = [
                item['status'],
                item['pn'],
                item['name'],
                item['spec'],
                str(item['present_count']),
            ]
            row.extend(fmt_qty(item['values'].get(code, 0)) if item['values'].get(code, 0) else '' for code in codes)
            row.extend([
                fmt_qty(item['max_qty']),
                fmt_qty(item['min_qty']),
                fmt_qty(item['range_diff']),
            ])
            self.diff_tree.insert('', 'end', iid=f'unified_diff_{idx}', values=row)

        self._diff_insert_index = end
        if end < len(self._diff_pending_rows):
            self.diff_summary_var.set(f'正在渲染统一对比结果... {end}/{len(self._diff_pending_rows)}')
            self.root.after(1, self._render_unified_diff_rows_chunk)
            return

        count_all_same, count_all_diff, count_partial_common, count_unique, warns = self._diff_counts
        summary = (
            f'统一对比完成 | 全部共用 {count_all_same} 条, '
            f'全部共用但用量不同 {count_all_diff} 条, '
            f'部分共用 {count_partial_common} 条, 独有 {count_unique} 条'
        )
        if warns:
            summary += f' | 警告 {len(warns)} 条'
        self.diff_progress['value'] = 100
        self.diff_run_btn.state(['!disabled'])
        self.diff_export_btn.state(['!disabled'])
        if hasattr(self, 'diff_unified_run_btn'):
            self.diff_unified_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_prefix_run_btn'):
            self.diff_prefix_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_doc_compare_btn'):
            self.diff_doc_compare_btn.state(['!disabled'])
        self._diff_compare_running = False
        self.diff_summary_var.set(summary)
        self.status_var.set(summary)

    def _render_bom_doc_compare_rows_chunk(self, chunk_size=150):
        start = self._diff_insert_index
        end = min(start + chunk_size, len(self._diff_pending_rows))
        for idx in range(start, end):
            item = self._diff_pending_rows[idx]
            row = [
                item.get('source_label', ''),
                item.get('root', ''),
                item.get('root_name', ''),
                item.get('root_model', ''),
                item.get('type', ''),
                item.get('pn', ''),
                item.get('doc_name', ''),
                item.get('bom_name', ''),
                item.get('doc_spec', ''),
                item.get('bom_spec', ''),
                fmt_qty(item.get('doc_qty', 0)) if item.get('doc_qty', 0) else '',
                fmt_qty(item.get('bom_qty', 0)) if item.get('bom_qty', 0) else '',
                fmt_qty(item.get('diff', 0)) if item.get('diff', 0) else '',
                item.get('note', ''),
            ]
            self.diff_tree.insert('', 'end', iid=f'bom_doc_diff_{idx}', values=row)
        self._diff_insert_index = end
        if end < len(self._diff_pending_rows):
            self.diff_summary_var.set(f'正在渲染量产BOM差异... {end}/{len(self._diff_pending_rows)}')
            self.root.after(1, self._render_bom_doc_compare_rows_chunk)
            return

        diff_count, root_count, missing_count, sheet_name = self._diff_counts
        results = getattr(self, '_bom_doc_compare_results', []) or []
        skipped_roots = sum(int(result.get('skipped_root_count', 0) or 0) for result in results)
        skipped_materials = sum(int(result.get('skipped_material_count', 0) or 0) for result in results)
        summary = (
            f'量产BOM对比完成 | Sheet:{sheet_name or "-"} | 母件 {root_count} 个 | 差异 {diff_count} 行 | '
            f'当前BOM缺母件 {missing_count} 个 | 已跳过删除机型 {skipped_roots} 个、删除物料 {skipped_materials} 行'
        )
        self.diff_progress['value'] = 100
        for btn_name in ('diff_run_btn', 'diff_export_btn', 'diff_unified_run_btn', 'diff_prefix_run_btn', 'diff_doc_compare_btn'):
            if hasattr(self, btn_name):
                getattr(self, btn_name).state(['!disabled'])
        self._diff_compare_running = False
        self.diff_summary_var.set(summary)
        self.status_var.set(summary)

    def _handle_diff_compare_error(self, err_msg):
        self.diff_progress['value'] = 0
        self.diff_run_btn.state(['!disabled'])
        self.diff_export_btn.state(['!disabled'])
        if hasattr(self, 'diff_unified_run_btn'):
            self.diff_unified_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_prefix_run_btn'):
            self.diff_prefix_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_doc_compare_btn'):
            self.diff_doc_compare_btn.state(['!disabled'])
        self._diff_compare_running = False
        self.diff_summary_var.set('差异对比执行失败')
        show_data_error(
            self.root,
            title='差异对比失败',
            summary='BOM 差异对比未能完成。',
            detail=err_msg,
            fix_hint=_suggest_fix_for_exception(Exception(err_msg)),
        )

    def _clear_diff(self):
        if hasattr(self, 'diff_tree'):
            self.diff_tree.delete(*self.diff_tree.get_children(''))
        if hasattr(self, 'diff_progress'):
            self.diff_progress['value'] = 0
        if hasattr(self, 'diff_run_btn'):
            self.diff_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_export_btn'):
            self.diff_export_btn.state(['!disabled'])
        if hasattr(self, 'diff_unified_run_btn'):
            self.diff_unified_run_btn.state(['!disabled'])
        if hasattr(self, 'diff_prefix_run_btn'):
            self.diff_prefix_run_btn.state(['!disabled'])
        self._diff_rows = []
        self._diff_export_rows = []
        self._diff_common_rows = []
        self._diff_pending_rows = []
        self._diff_input_codes = {'a': [], 'b': []}
        self._diff_unified_rows = []
        self._diff_code_label_map = {}
        self._diff_unified_source = ''
        self._bom_doc_compare_rows = []
        self._bom_doc_compare_summary = []
        self._bom_doc_compare_missing_roots = []
        self._bom_doc_compare_source = ''
        self._bom_doc_compare_results = []
        self._diff_mode = 'pair'
        self._diff_compare_running = False
        self._configure_diff_tree(
            HEADERS_DIFF_GRID,
            COL_WIDTHS_D_GRID,
            numeric_columns={'A总用量', 'B总用量', '差值(A-B)'},
        )
        if hasattr(self, 'diff_summary_var'):
            self.diff_summary_var.set('请输入两个母料号进行AB差异对比；多料号请使用下方“统一对比粘贴区”')

    def _export_diff(self):
        if getattr(self, '_diff_mode', 'pair') == 'bom_doc':
            rows = getattr(self, '_bom_doc_compare_rows', [])
            summary_rows = getattr(self, '_bom_doc_compare_summary', [])
            missing_roots = getattr(self, '_bom_doc_compare_missing_roots', [])
            if not rows and not summary_rows:
                messagebox.showwarning('无数据', '请先上传量产BOM表并执行对比')
                return
            path = filedialog.asksaveasfilename(
                title='导出量产BOM对比当前BOM结果',
                defaultextension='.xlsx',
                filetypes=[('Excel 文件', '*.xlsx')],
                initialfile=f'量产BOM对比当前BOM_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            )
            if not path:
                return
            try:
                wb = openpyxl.Workbook()
                hdr_font = Font(bold=True, color='FFFFFF', size=10)
                hdr_fill = PatternFill('solid', fgColor='305496')
                thin = Side(style='thin', color='AAAAAA')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                even = PatternFill('solid', fgColor='EAF2FF')
                warn_fill = PatternFill('solid', fgColor='FFF2CC')
                bad_fill = PatternFill('solid', fgColor='FCE4D6')

                def write_table(ws, headers, data_rows, widths):
                    for ci, h in enumerate(headers, 1):
                        c = ws.cell(row=1, column=ci, value=h)
                        c.font = hdr_font
                        c.fill = hdr_fill
                        c.border = border
                        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.row_dimensions[1].height = 28
                    for ri, row_values in enumerate(data_rows, 2):
                        for ci, val in enumerate(row_values, 1):
                            c = ws.cell(row=ri, column=ci, value=val)
                            c.border = border
                            c.alignment = Alignment(vertical='center', wrap_text=True)
                            if ri % 2 == 0:
                                c.fill = even
                            if '差异类型' in headers:
                                diff_type_index = headers.index('差异类型')
                                diff_type = str(row_values[diff_type_index] if len(row_values) > diff_type_index else '')
                                if diff_type in ('当前BOM缺母件', 'DOC有/BOM没有', 'BOM有/DOC没有'):
                                    c.fill = bad_fill
                                elif diff_type in ('用量不同', '名称规格不同'):
                                    c.fill = warn_fill
                        ws.row_dimensions[ri].height = 20
                    for i, w in enumerate(widths, 1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{max(ws.max_row, 1)}"
                    ws.freeze_panes = 'A2'
                    ws.sheet_view.showGridLines = False

                def safe_sheet_name(name):
                    clean = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(name or '').strip())
                    return clean[:31] or 'Sheet'

                def unique_sheet_name(base_name):
                    base = safe_sheet_name(base_name)
                    if base not in wb.sheetnames:
                        return base
                    for idx in range(2, 1000):
                        suffix = f'_{idx}'
                        candidate = f'{base[:31 - len(suffix)]}{suffix}'
                        if candidate not in wb.sheetnames:
                            return candidate
                    return base[:28] + '_X'

                def build_diff_data(items):
                    return [[
                        item.get('source_label', ''),
                        item.get('root', ''),
                        item.get('root_name', ''),
                        item.get('root_model', ''),
                        item.get('type', ''),
                        item.get('pn', ''),
                        item.get('doc_name', ''),
                        item.get('bom_name', ''),
                        item.get('doc_spec', ''),
                        item.get('bom_spec', ''),
                        fmt_qty(item.get('doc_qty', 0)) if item.get('doc_qty', 0) else '',
                        fmt_qty(item.get('bom_qty', 0)) if item.get('bom_qty', 0) else '',
                        fmt_qty(item.get('diff', 0)) if item.get('diff', 0) else '',
                        item.get('note', ''),
                    ] for item in items]

                summary_headers = [
                    '来源BOM', '母件料号', '成品号', '母件名称', '母件型号', 'DOC物料数', 'BOM物料数',
                    '差异行数', 'DOC有BOM没有', 'BOM有DOC没有', '用量不同', '名称规格不同',
                    '警告', '配置说明'
                ]
                def build_summary_data(items):
                    return [[
                        row.get('来源BOM', ''),
                        row.get('母件料号', ''),
                        row.get('成品号', ''),
                        row.get('母件名称', ''),
                        row.get('母件型号', ''),
                        row.get('DOC物料数', 0),
                        row.get('BOM物料数', 0),
                        row.get('差异行数', 0),
                        row.get('DOC有BOM没有', 0),
                        row.get('BOM有DOC没有', 0),
                        row.get('用量不同', 0),
                        row.get('名称规格不同', 0),
                        row.get('警告', ''),
                        row.get('配置说明', ''),
                    ] for row in items]

                missing_headers = ['来源BOM', '母件料号', '成品号', '母件名称', '母件型号', 'DOC物料数', '问题说明', '配置说明']
                def build_missing_data(items):
                    return [[
                        row.get('来源BOM', ''),
                        row.get('母件料号', ''),
                        row.get('成品号', ''),
                        row.get('母件名称', ''),
                        row.get('母件型号', ''),
                        row.get('DOC物料数', 0),
                        row.get('问题说明', ''),
                        row.get('配置说明', ''),
                    ] for row in items]

                results = getattr(self, '_bom_doc_compare_results', []) or []
                if not results:
                    results = [{
                        'source': getattr(self, '_bom_doc_compare_source', ''),
                        'source_label': os.path.basename(str(getattr(self, '_bom_doc_compare_source', '') or '量产BOM')),
                        'sheet': '',
                        'rows': rows,
                        'summary': summary_rows,
                        'missing_roots': missing_roots,
                        'root_count': len(summary_rows),
                        'skipped_root_count': 0,
                        'skipped_material_count': 0,
                    }]

                overview_ws = wb.active
                overview_ws.title = '总览'
                overview_headers = [
                    '序号', '来源BOM', '原始Sheet', '04母件数', '差异行数',
                    '未匹配母件数', '跳过删除机型', '跳过删除物料行',
                    'DOC有BOM没有', 'BOM有DOC没有', '用量不同', '名称规格不同', '文件路径'
                ]
                overview_data = []
                for idx, result in enumerate(results, 1):
                    result_summary = result.get('summary', []) or []
                    overview_data.append([
                        idx,
                        result.get('source_label', ''),
                        result.get('sheet', ''),
                        result.get('root_count', len(result_summary)),
                        len(result.get('rows', []) or []),
                        len(result.get('missing_roots', []) or []),
                        result.get('skipped_root_count', 0),
                        result.get('skipped_material_count', 0),
                        sum(int(row.get('DOC有BOM没有', 0) or 0) for row in result_summary),
                        sum(int(row.get('BOM有DOC没有', 0) or 0) for row in result_summary),
                        sum(int(row.get('用量不同', 0) or 0) for row in result_summary),
                        sum(int(row.get('名称规格不同', 0) or 0) for row in result_summary),
                        result.get('source', ''),
                    ])
                write_table(overview_ws, overview_headers, overview_data, [8, 36, 16, 10, 10, 12, 12, 14, 13, 13, 10, 13, 80])

                for idx, result in enumerate(results, 1):
                    suffix = f'{int(idx):02d}'
                    label = result.get('source_label', f'BOM{idx}')
                    diff_ws = wb.create_sheet(unique_sheet_name(f'差异明细_{suffix}'))
                    diff_ws.cell(row=1, column=1, value=f'来源BOM：{label}')
                    write_table(diff_ws, BOM_DOC_COMPARE_HEADERS, build_diff_data(result.get('rows', []) or []), BOM_DOC_COMPARE_WIDTHS)

                    summary_ws = wb.create_sheet(unique_sheet_name(f'母件汇总_{suffix}'))
                    write_table(
                        summary_ws,
                        summary_headers,
                        build_summary_data(result.get('summary', []) or []),
                        [24, 16, 16, 18, 24, 10, 10, 10, 13, 13, 10, 13, 30, 50],
                    )

                    missing_ws = wb.create_sheet(unique_sheet_name(f'未匹配母件_{suffix}'))
                    write_table(
                        missing_ws,
                        missing_headers,
                        build_missing_data(result.get('missing_roots', []) or []),
                        [24, 16, 16, 18, 24, 10, 36, 50],
                    )

                info_ws = wb.create_sheet('说明')
                info_rows = [
                    ['项目', '内容'],
                    ['外部BOM表', getattr(self, '_bom_doc_compare_source', '')],
                    ['当前BOMMaster', self.current_file or ''],
                    ['对比口径', '外部量产BOM表每个04在制品号的数量列 vs 当前BOMMaster递归展开后的单台汇总用量'],
                    ['自动过滤', '优先读取“物料清单”sheet；红底/删除线机型列不参与核对；物料编号/名称/型号等关键格有删除线的物料行不参与核对'],
                    ['差值', 'DOC用量 - BOM用量'],
                ]
                write_table(info_ws, ['项目', '内容'], info_rows[1:], [18, 100])

                wb.save(path)
                messagebox.showinfo('导出成功', f'已保存至:\n{path}')
                self.status_var.set(f'已导出: {os.path.basename(path)}')
            except Exception as e:
                messagebox.showerror('导出失败', str(e))
            return

        if getattr(self, '_diff_mode', 'pair') == 'unified':
            rows = getattr(self, '_diff_unified_rows', [])
            codes = getattr(self, '_diff_input_codes', {}).get('unified', [])
            code_label_map = getattr(self, '_diff_input_codes', {}).get('code_labels', {}) or self._get_diff_code_label_map(codes)
            code_headers = [code_label_map.get(code, code) for code in codes]
            if not rows:
                messagebox.showwarning('无数据', '请先执行统一对比')
                return
            path = filedialog.asksaveasfilename(
                title='导出 BOM 统一对比结果',
                defaultextension='.xlsx',
                filetypes=[('Excel 文件', '*.xlsx')]
            )
            if not path:
                return
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '统一对比'
                hdr_font = Font(bold=True, color='FFFFFF', size=10)
                hdr_fill = PatternFill('solid', fgColor='305496')
                thin = Side(style='thin', color='AAAAAA')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                even = PatternFill('solid', fgColor='EAF2FF')

                headers = ['共用状态', '物料编码', '品名', '规格', '出现BOM数'] + code_headers + ['最大用量', '最小用量', '差值(最大-最小)']
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = hdr_font; c.fill = hdr_fill; c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 46
                for ri, item in enumerate(rows, 2):
                    row = [
                        item['status'],
                        item['pn'],
                        item['name'],
                        item['spec'],
                        item['present_count'],
                    ]
                    row.extend(fmt_qty(item['values'].get(code, 0)) if item['values'].get(code, 0) else '' for code in codes)
                    row.extend([fmt_qty(item['max_qty']), fmt_qty(item['min_qty']), fmt_qty(item['range_diff'])])
                    for ci, val in enumerate(row, 1):
                        c = ws.cell(row=ri, column=ci, value=val)
                        c.border = border
                        if ri % 2 == 0:
                            c.fill = even
                    ws.row_dimensions[ri].height = 20

                widths = [12, 18, 24, 20, 10] + [24] * len(codes) + [12, 12, 14]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{max(ws.max_row, 1)}"
                ws.freeze_panes = 'A2'

                common_ws = wb.create_sheet('共用物料')
                common_headers = ['共用状态', '物料编码', '品名', '规格', '出现BOM数'] + code_headers + ['最大用量', '最小用量', '差值(最大-最小)']
                for ci, h in enumerate(common_headers, 1):
                    c = common_ws.cell(row=1, column=ci, value=h)
                    c.font = hdr_font; c.fill = hdr_fill; c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                common_ws.row_dimensions[1].height = 46
                common_rows = [item for item in rows if item.get('present_count', 0) >= 2]
                for ri, item in enumerate(common_rows, 2):
                    row = [
                        item['status'],
                        item['pn'],
                        item['name'],
                        item['spec'],
                        item['present_count'],
                    ]
                    row.extend(fmt_qty(item['values'].get(code, 0)) if item['values'].get(code, 0) else '' for code in codes)
                    row.extend([fmt_qty(item['max_qty']), fmt_qty(item['min_qty']), fmt_qty(item['range_diff'])])
                    for ci, val in enumerate(row, 1):
                        c = common_ws.cell(row=ri, column=ci, value=val)
                        c.border = border
                        if ri % 2 == 0:
                            c.fill = even
                for i, w in enumerate(widths, 1):
                    common_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                common_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(common_headers))}{max(common_ws.max_row, 1)}"
                common_ws.freeze_panes = 'A2'

                diff_ws = wb.create_sheet('差异物料')
                diff_headers = ['共用状态', '物料编码', '品名', '规格', '出现BOM数'] + code_headers + ['最大用量', '最小用量', '差值(最大-最小)']
                for ci, h in enumerate(diff_headers, 1):
                    c = diff_ws.cell(row=1, column=ci, value=h)
                    c.font = hdr_font; c.fill = hdr_fill; c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                diff_ws.row_dimensions[1].height = 46
                diff_rows = [item for item in rows if item.get('status') != '全部共用']
                for ri, item in enumerate(diff_rows, 2):
                    row = [
                        item['status'],
                        item['pn'],
                        item['name'],
                        item['spec'],
                        item['present_count'],
                    ]
                    row.extend(fmt_qty(item['values'].get(code, 0)) if item['values'].get(code, 0) else '' for code in codes)
                    row.extend([fmt_qty(item['max_qty']), fmt_qty(item['min_qty']), fmt_qty(item['range_diff'])])
                    for ci, val in enumerate(row, 1):
                        c = diff_ws.cell(row=ri, column=ci, value=val)
                        c.border = border
                        if ri % 2 == 0:
                            c.fill = even
                for i, w in enumerate(widths, 1):
                    diff_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                diff_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(diff_headers))}{max(diff_ws.max_row, 1)}"
                diff_ws.freeze_panes = 'A2'

                input_ws = wb.create_sheet('输入料号')
                source_label = getattr(self, '_diff_input_codes', {}).get('source', '') or getattr(self, '_diff_unified_source', '')
                input_headers = ['来源', '序号', '母料号', '母件名称', '母件规格']
                for ci, h in enumerate(input_headers, 1):
                    c = input_ws.cell(row=1, column=ci, value=h)
                    c.font = hdr_font; c.fill = hdr_fill; c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center')
                for ri, code in enumerate(codes, 2):
                    root_name, root_spec = self._get_bom_root_meta(code)
                    for ci, val in enumerate([source_label, ri - 1, code, root_name, root_spec], 1):
                        c = input_ws.cell(row=ri, column=ci, value=val)
                        c.border = border
                        if ri % 2 == 0:
                            c.fill = even
                input_ws.column_dimensions['A'].width = 18
                input_ws.column_dimensions['B'].width = 8
                input_ws.column_dimensions['C'].width = 22
                input_ws.column_dimensions['D'].width = 34
                input_ws.column_dimensions['E'].width = 24
                input_ws.auto_filter.ref = f"A1:E{max(input_ws.max_row, 1)}"
                input_ws.freeze_panes = 'A2'

                wb.save(path)
                messagebox.showinfo('导出成功', f'已保存至:\n{path}')
                self.status_var.set(f'已导出: {os.path.basename(path)}')
            except Exception as e:
                messagebox.showerror('导出失败', str(e))
            return

        if not self._diff_rows:
            messagebox.showwarning('无数据', '请先执行差异对比')
            return
        path = filedialog.asksaveasfilename(
            title='导出 BOM 差异对比结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'BOM差异对比'
            hdr_font = Font(bold=True, color='FFFFFF', size=10)
            hdr_fill = PatternFill('solid', fgColor='305496')
            thin = Side(style='thin', color='AAAAAA')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            export_headers = ['左侧(A)结果', '左侧物料编码', '左侧品名', '左侧规格', '左侧总用量',
                              '右侧(B)结果', '右侧物料编码', '右侧品名', '右侧规格', '右侧总用量',
                              '差异类型', '差值(A-B)']
            for ci, h in enumerate(export_headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            even = PatternFill('solid', fgColor='EAF2FF')
            for ri, item in enumerate(self._diff_rows, 2):
                left_type = item['type'] if item['qa'] != 0 else ''
                right_type = item['type'] if item['qb'] != 0 else ''
                row = [
                    left_type,
                    item['pn'] if item['qa'] != 0 else '',
                    item['name'] if item['qa'] != 0 else '',
                    item['spec'] if item['qa'] != 0 else '',
                    fmt_qty(item['qa']) if item['qa'] != 0 else '',
                    right_type,
                    item['pn'] if item['qb'] != 0 else '',
                    item['name'] if item['qb'] != 0 else '',
                    item['spec'] if item['qb'] != 0 else '',
                    fmt_qty(item['qb']) if item['qb'] != 0 else '',
                    item['type'],
                    fmt_qty(item['diff']),
                ]
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = border
                    if ri % 2 == 0:
                        c.fill = even
                ws.row_dimensions[ri].height = 20

            for i, w in enumerate([12, 18, 24, 20, 10, 12, 18, 24, 20, 10, 11, 10], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(export_headers))}{max(ws.max_row, 1)}"
            ws.freeze_panes = 'A2'

            common_ws = wb.create_sheet('共用物料')
            common_headers = ['共用状态', '物料编码', '品名', '规格', 'A组总用量', 'B组总用量', '差值(A-B)']
            for ci, h in enumerate(common_headers, 1):
                c = common_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            for ri, item in enumerate(getattr(self, '_diff_common_rows', []), 2):
                status = '用量相同' if abs(float(item.get('diff', 0) or 0)) < 1e-12 else '用量不同'
                row = [
                    status,
                    item['pn'],
                    item['name'],
                    item['spec'],
                    fmt_qty(item['qa']),
                    fmt_qty(item['qb']),
                    fmt_qty(item['diff']),
                ]
                for ci, val in enumerate(row, 1):
                    c = common_ws.cell(row=ri, column=ci, value=val)
                    c.border = border
                    if ri % 2 == 0:
                        c.fill = even
                common_ws.row_dimensions[ri].height = 20
            for i, w in enumerate([12, 18, 24, 20, 12, 12, 12], 1):
                common_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            common_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(common_headers))}{max(common_ws.max_row, 1)}"
            common_ws.freeze_panes = 'A2'

            input_ws = wb.create_sheet('输入料号')
            input_headers = ['组别', '序号', '母料号']
            for ci, h in enumerate(input_headers, 1):
                c = input_ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            input_rows = []
            for group_name, key in [('A组', 'a'), ('B组', 'b')]:
                for idx, code in enumerate(getattr(self, '_diff_input_codes', {}).get(key, []), 1):
                    input_rows.append((group_name, idx, code))
            for ri, row in enumerate(input_rows, 2):
                for ci, val in enumerate(row, 1):
                    c = input_ws.cell(row=ri, column=ci, value=val)
                    c.border = border
                    if ri % 2 == 0:
                        c.fill = even
            for i, w in enumerate([10, 8, 22], 1):
                input_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            input_ws.auto_filter.ref = f"A1:C{max(input_ws.max_row, 1)}"
            input_ws.freeze_panes = 'A2'

            wb.save(path)
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
            self.status_var.set(f'已导出: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))

    # ── 切换选项卡时：先全展开树，再计算数量 ───────────────────
    def _on_tab_changed(self, event=None):
        if self._is_current_tab(getattr(self, 'tab_calc', None)):
            roots = self.tree.get_children('')
            if roots:
                self._expand_and_load_all(roots)
            self._update_material_summary()
            self._recalc_from_tree()
            self._refresh_calc_hint()

    def _expand_and_load_all(self, node_ids):
        """
        纯展开：只把界面里已有的树节点全部 open，不再调用 _rec_insert 插入新节点。
        说明：查询时 _query → _rec_insert 已一次性构建完整 BOM 树，
        这里如果再次调用 _rec_insert 会导致每次切换选项卡节点数翻倍（BUG）。
        """
        for nid in node_ids:
            self.tree.item(nid, open=True)
            for ch in self.tree.get_children(nid):
                self._expand_and_load_all([ch])

    def _count_tree_nodes(self):
        """遍历整棵树，返回节点总数（含主料）"""
        def walk(node_id):
            total = 1
            for ch in self.tree.get_children(node_id):
                total += walk(ch)
            return total
        total = 0
        for root_id in self.tree.get_children(''):
            total += walk(root_id)
        return total

    def _update_material_summary(self):
        total = self._count_tree_nodes()
        if total > 0:
            self.calc_summary_var.set(f'📌 当前BOM：{total}颗物料（含主料）')
        else:
            self.calc_summary_var.set('📌 当前BOM：0颗物料')
        return total

    # ── 核心：从展开树读取并计算数量 ─────────────────────────
    def _recalc_from_tree(self):
        """读取选项卡1的展开树，按层级递归计算各节点累计用量"""
        roots = self.tree.get_children('')
        if not roots:
            self.status_var.set('请先在「BOM结构展开」选项卡中查询并展开 BOM')
            return

        try:
            top_qty = float(self.calc_qty_var.get().strip())
            if top_qty <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showwarning('数量错误', '请输入正确的顶层母件数量（正数）')
            return

        self.status_var.set('正在从展开树计算 BOM 用量 …')
        self.root.update_idletasks()

        # 遍历树，用 (节点ID, 父件累计用量) 递归计算
        levels = {}   # {层级: [(料号, 品名, 规格, 单位用量, 累计用量), ...]}
        flat = {}     # {料号: {name, spec, total_qty, min_level}}

        def recurse(node_id, parent_qty, depth):
            """读取树节点，获取其子件，递归计算"""
            vals = self.tree.item(node_id, 'values')
            # vals: [层级, 母件料号, 母件品名, 母件规格, 子件料号, 子件品名, 子件规格, 用量]
            # 根节点：子件料号在 vals[4]，用量在 vals[7]
            # 子子节点：同样
            if len(vals) < 8:
                return

            child_pn  = str(vals[4]).strip()
            child_name = str(vals[5]).strip()
            child_spec = str(vals[6]).strip()
            child_qty_raw = str(vals[7]).strip()

            if not child_pn:
                # 根节点没有子件料号，是母件本身
                # 直接遍历其子节点
                for ch in self.tree.get_children(node_id):
                    recurse(ch, top_qty, depth + 1)
                return

            # 子件用量（单位用量）
            try:
                child_qty = float(child_qty_raw) if child_qty_raw else 1.0
            except ValueError:
                child_qty = 1.0

            total_qty = parent_qty * child_qty
            next_depth = depth + 1

            levels.setdefault(next_depth, []).append(
                (child_pn, child_name, child_spec, child_qty, total_qty)
            )
            if child_pn in flat:
                flat[child_pn]['total_qty'] += total_qty
                flat[child_pn]['min_level'] = min(flat[child_pn]['min_level'], next_depth)
                flat[child_pn]['paths'] += 1
            else:
                flat[child_pn] = {
                    'name': child_name,
                    'spec': child_spec,
                    'total_qty': total_qty,
                    'min_level': next_depth,
                    'paths': 1,
                }

            # 递归子节点（子件的子件）
            for ch in self.tree.get_children(node_id):
                recurse(ch, total_qty, next_depth)

        # 从每个根节点开始（根节点本身不计入，从其直接子件开始）
        for root_id in roots:
            root_vals = self.tree.item(root_id, 'values')
            root_pn = str(root_vals[1]).strip() if len(root_vals) > 1 else ''
            # 顶层 × 顶层用量 = 顶层累计
            root_qty_raw = str(root_vals[7]).strip() if len(root_vals) > 7 else '1'
            try:
                root_qty = float(root_qty_raw) if root_qty_raw else 1.0
            except ValueError:
                root_qty = 1.0
            top_total = top_qty * root_qty

            # 记录根节点（层级1 = 母件本身）
            if root_pn:
                levels.setdefault(1, []).append(
                    (root_pn,
                     str(root_vals[2]).strip() if len(root_vals) > 2 else '',
                     str(root_vals[3]).strip() if len(root_vals) > 3 else '',
                     root_qty, top_total)
                )
                if root_pn in flat:
                    flat[root_pn]['total_qty'] += top_total
                    flat[root_pn]['min_level'] = min(flat[root_pn]['min_level'], 1)
                    flat[root_pn]['paths'] += 1
                else:
                    flat[root_pn] = {
                        'name': str(root_vals[2]).strip() if len(root_vals) > 2 else '',
                        'spec': str(root_vals[3]).strip() if len(root_vals) > 3 else '',
                        'total_qty': top_total,
                        'min_level': 1,
                        'paths': 1,
                    }

            # 递归遍历根节点的直接子件
            for ch in self.tree.get_children(root_id):
                recurse(ch, top_total, 1)

        total_entries = sum(len(v) for v in levels.values())
        self._calc_result = {'levels': levels, 'flat': flat}
        self._refresh_calc_tree()

        max_lvl = max(levels.keys()) if levels else 0
        self._update_material_summary()
        self.status_var.set(
            f'📦 数量计算完成  |  参与计算 {total_entries}颗物料（含主料）|  独立料号 {len(flat)} 种  |  最大层级 {max_lvl}'
        )
        self._refresh_calc_hint()

    def _refresh_calc_tree(self):
        for item in self.calc_tree.get_children(''):
            self.calc_tree.delete(item)
        if not self._calc_result:
            return
        # 用 flat 展示：每个料号只出现一次，用量为所有路径累加后的汇总值
        flat = self._calc_result['flat']
        for pn in sorted(flat.keys(), key=lambda x: (flat[x]['min_level'], x)):
            info = flat[pn]
            self.calc_tree.insert('', 'end', text='', values=(
                str(info['min_level']),
                pn,
                info['name'],
                info['spec'],
                str(info.get('paths', '')),
                fmt_qty(info['total_qty'])
            ))


    # ── batch multi-root calc ────────────────────────────────
    def _import_root_qtys_file(self):
        path = filedialog.askopenfilename(
            title='\u9009\u62e9\u6bcd\u6599\u6570\u91cf\u6587\u4ef6',
            filetypes=[('Excel', '*.xlsx *.xls'),
                       ('CSV', '*.csv'),
                       ('\u6587\u672c', '*.txt'),
                       ('\u6240\u6709\u6587\u4ef6', '*.*')])
        if not path:
            return
        try:
            root_qtys = load_root_qtys_from_file(path)
            lines = '\n'.join(f'{pn},{qty}' for pn, qty in root_qtys)
            self._batch_text.delete('1.0', tk.END)
            self._batch_text.insert('1.0', lines)
            self.status_var.set(f'\u5df2\u52a0\u8f7d {len(root_qtys)} \u6761\u6bcd\u6599\u8bb0\u5f55')
        except Exception as e:
            messagebox.showerror('\u8bfb\u53d6\u5931\u8d25', str(e))

    def _parse_batch_input(self):
        raw = self._batch_text.get('1.0', tk.END).strip()
        if not raw:
            return []
        entries = []
        skipped_header = 0
        for line in raw.replace('\r\n', '\n').replace('\r', '\n').split('\n'):
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # Excel copies cells as Tab-separated; also support comma/Chinese comma
            # Split by tab first (Excel), then by comma
            if '\t' in line:
                parts = [p.strip() for p in line.split('\t') if p.strip()]
            else:
                parts = [p.strip() for p in
                         line.replace('\uff0c', ',').split(',') if p.strip()]
            if not parts:
                continue
            pn = parts[0]
            # skip pure-text header rows (e.g. col header '\u6bcd\u6599\u53f7')
            if not any(c.isdigit() or c == '-' or c == '_' or c == '.' for c in pn):
                skipped_header += 1
                continue
            qty = 1.0
            if len(parts) >= 2:
                qty_str = parts[1].replace(',', '').strip()
                try:
                    qty = float(qty_str)
                except ValueError:
                    # non-numeric in qty column -> treat as header row, skip
                    skipped_header += 1
                    continue
            if qty <= 0:
                messagebox.showwarning('\u6570\u91cf\u9519\u8bef',
                    f'\u6570\u91cf\u5fc5\u987b\u4e3a\u6b63\u6570\uff1a{line}')
                return []
            entries.append((pn, qty))
        return entries

    def _run_batch_calc(self):
        if not self.bom_index:
            messagebox.showwarning('\u672a\u52a0\u8f7d BOM',
                '\u8bf7\u5148\u4e0a\u4f20 BOM \u6587\u4ef6')
            return
        entries = self._parse_batch_input()
        if not entries:
            messagebox.showwarning('\u65e0\u6570\u636e',
                '\u8bf7\u8f93\u5165\u81f3\u5c11\u4e00\u7ec4\u300c\u6bcd\u6599\u53f7,\u6570\u91cf\u300d')
            return
        missing = [pn for pn, _ in entries if pn not in self.bom_index]
        if missing:
            messagebox.showwarning(
                '\u6bcd\u6599\u53f7\u4e0d\u5b58\u5728',
                '\u4ee5\u4e0b\u6bcd\u6599\u53f7\u4e0d\u5728\u5f53\u524d BOM \u4e2d\uff1a\n' +
                '\n'.join(missing[:10]) +
                ('\n...' if len(missing) > 10 else ''))
            return

        self.status_var.set(
            f'\u6b63\u5728\u6279\u91cf\u8ba1\u7b97 {len(entries)} \u4e2a\u6bcd\u6599\u53f7 \u2026')
        self.root.update_idletasks()

        flat = {}
        batch_cycle_hits: set[tuple[str, str]] = set()

        def _recurse(pn, qty, depth, trail: frozenset[str] = frozenset()):
            if depth > MAX_DEPTH or pn not in self.bom_index:
                return
            for row in self.bom_index[pn]:
                cp = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
                if not cp:
                    continue
                if cp in trail or cp == pn:
                    batch_cycle_hits.add((pn, cp))
                    continue
                cname = str(row[CHILD_NAME_COL]).strip() if row[CHILD_NAME_COL] else ''
                cspec = str(row[CHILD_SPEC_COL]).strip() if row[CHILD_SPEC_COL] else ''
                try:
                    cqty = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else 1.0
                except (ValueError, TypeError):
                    cqty = 1.0
                tqty = qty * cqty
                if cp in flat:
                    flat[cp]['total_qty'] += tqty
                    flat[cp]['paths'] += 1
                    flat[cp]['min_level'] = min(flat[cp]['min_level'], depth + 1)
                else:
                    flat[cp] = {'name': cname, 'spec': cspec,
                                'total_qty': tqty, 'paths': 1,
                                'min_level': depth + 1}
                _recurse(cp, tqty, depth + 1, trail=trail | {pn})

        for mother_pn, mother_qty in entries:
            r0 = self.bom_index[mother_pn][0]
            rname = str(r0[2]).strip() if len(r0) > 2 and r0[2] else ''
            rspec = str(r0[3]).strip() if len(r0) > 3 and r0[3] else ''
            if mother_pn in flat:
                flat[mother_pn]['total_qty'] += mother_qty
                flat[mother_pn]['paths'] += 1
                flat[mother_pn]['min_level'] = min(flat[mother_pn]['min_level'], 1)
            else:
                flat[mother_pn] = {'name': rname, 'spec': rspec,
                                   'total_qty': mother_qty, 'paths': 1,
                                   'min_level': 1}
            _recurse(mother_pn, mother_qty, 1, trail=frozenset({mother_pn}))

        if batch_cycle_hits:
            sample = '、'.join(f'{a}→{b}' for a, b in list(batch_cycle_hits)[:3])
            logging.getLogger(__name__).warning('BOM 存在环路：%s', sample)

        self._batch_result = flat
        self._refresh_calc_hint()

        for item in self.calc_tree.get_children(''):
            self.calc_tree.delete(item)

        for pn in sorted(flat.keys(), key=lambda x: (flat[x]['min_level'], x)):
            info = flat[pn]
            self.calc_tree.insert('', 'end', values=(
                str(info['min_level']), pn, info['name'], info['spec'],
                str(info['paths']), fmt_qty(info['total_qty'])))

        self.calc_summary_var.set(
            f'\U0001f4cb \u6279\u91cf\u6c47\u603b\uff1a{len(flat)} \u79cd\u7269\u6599\uff08{len(entries)} \u4e2a\u6bcd\u6599\u53f7\uff09')
        self.status_var.set(
            f'\u2705 \u6279\u91cf\u8ba1\u7b97\u5b8c\u6210  |  \u6bcd\u6599\u53f7 {len(entries)} \u4e2a  '
            f'|  \u7269\u6599\u79cd\u7c7b {len(flat)} \u79cd')

    def _export_batch_calc(self):
        if not getattr(self, '_batch_result', None):
            messagebox.showwarning('\u65e0\u6570\u636e',
                '\u8bf7\u5148\u6267\u884c\u6279\u91cf\u6c47\u603b\u8ba1\u7b97')
            return
        path = filedialog.asksaveasfilename(
            title='\u5bfc\u51fa\u6279\u91cf\u6c47\u603b\u7ed3\u679c',
            defaultextension='.xlsx',
            filetypes=[('Excel \u6587\u4ef6', '*.xlsx')])
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '\u6279\u91cf\u6c47\u603b'
            hf = Font(bold=True, color='FFFFFF', size=10)
            hfill = PatternFill('solid', fgColor='1F4E79')
            thin = Side(style='thin', color='AAAAAA')
            bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdrs = ['\u6700\u6d45\u5c42\u7ea7', '\u6599\u53f7', '\u54c1\u540d',
                    '\u89c4\u683c', '\u6c47\u603b\u7528\u91cf', '\u6765\u6e90\u8def\u5f84\u6570']
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hf; c.fill = hfill; c.border = bdr
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22
            even = PatternFill('solid', fgColor='DCE6F1')
            flat = self._batch_result
            for ri, pn in enumerate(
                    sorted(flat.keys(), key=lambda x: (flat[x]['min_level'], x)), 2):
                info = flat[pn]
                row_data = [str(info['min_level']), pn, info['name'],
                            info['spec'],
                            round(info['total_qty'], 6),
                            info['paths']]
                for ci, val in enumerate(row_data, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = bdr
                    if ri % 2 == 0:
                        c.fill = even
                ws.row_dimensions[ri].height = 20
            for i, w in enumerate([6, 18, 28, 22, 12, 10], 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo('\u5bfc\u51fa\u6210\u529f',
                f'\u5df2\u4fdd\u5b58\u81f3\uff1a\n{path}')
            self.status_var.set(f'\u5df2\u5bfc\u51fa: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('\u5bfc\u51fa\u5931\u8d25', str(e))

    def _clear_batch_calc(self):
        self._batch_result = None
        self._batch_text.delete('1.0', tk.END)
        placeholder = ('\u652f\u6301\u76f4\u63a5\u4ece Excel \u590d\u5236\u4e24\u5217\uff08\u6bcd\u6599\u53f7 + \u6570\u91cf\uff09\u7c98\u8d34\n'
                       '\u4e5f\u53ef\u624b\u8f93\uff1a\u6bcf\u884c\u4e00\u6761\uff0c\u683c\u5f0f  \u6bcd\u6599\u53f7<Tab>\u6570\u91cf  \u6216  \u6bcd\u6599\u53f7,\u6570\u91cf\n'
                       '\uff08\u4e0d\u586b\u6570\u91cf\u9ed8\u8ba4=1\uff1b\u542b\u8868\u5934\u884c\u4f1a\u81ea\u52a8\u8df3\u8fc7\uff09')
        self._batch_text.insert('1.0', placeholder)
        self.calc_summary_var.set('\U0001f4cc \u5f53\u524dBOM\uff1a0\u9897\u7269\u6599')
        self.status_var.set(f'\u5c31\u7eea  |  \u6bcd\u5668\u4ef6: {len(self.bom_index)} \u4e2a')

    # ── 树形展开 ──────────────────────────────────────────────
    def _match_bom_roots(self, query_text, limit=20):
        query_text = str(query_text or '').strip()
        if not query_text:
            return []
        if query_text in self.bom_index:
            return [query_text]

        q_lower = query_text.lower()
        exact_spec_matches = []
        exact_name_matches = []
        fuzzy_matches = []
        seen = set()

        def add_unique(bucket, code):
            code = str(code or '').strip()
            if code and code not in seen:
                seen.add(code)
                bucket.append(code)

        for parent_pn, rows in sorted(self.bom_index.items()):
            if not rows:
                continue
            first = rows[0]
            parent_code = str(parent_pn).strip()
            parent_name = str(first[2]).strip() if len(first) > 2 and first[2] else ''
            parent_spec = str(first[3]).strip() if len(first) > 3 and first[3] else ''
            if parent_spec and parent_spec.lower() == q_lower:
                add_unique(exact_spec_matches, parent_code)
            elif parent_name and parent_name.lower() == q_lower:
                add_unique(exact_name_matches, parent_code)
            elif (
                q_lower in parent_code.lower()
                or (parent_spec and q_lower in parent_spec.lower())
                or (parent_name and q_lower in parent_name.lower())
            ):
                add_unique(fuzzy_matches, parent_code)

        return [*exact_spec_matches, *exact_name_matches, *fuzzy_matches][:limit]

    def _query(self):
        raw = self.pn_entry.get().strip()
        if not raw:
            messagebox.showwarning('请输入料号', '请输入料号')
            return
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return

        pns = [s.strip() for seg in raw.replace('\n', ',').replace('，', ',').split(',')
               if (s := seg.strip())]

        self._clear_tree()
        max_depth = self.depth_var.get()
        total = [0, False]  # [count, truncated]
        resolved_pns = []
        unresolved_terms = []
        expanded_terms = []

        for query_term in pns:
            matches = self._match_bom_roots(query_term, limit=20)
            if not matches:
                unresolved_terms.append(query_term)
                continue
            if len(matches) > 1 or matches[0] != query_term:
                expanded_terms.append(f'{query_term}→{len(matches)}个')
            for root_pn in matches:
                if root_pn not in resolved_pns:
                    resolved_pns.append(root_pn)

        for root_pn in resolved_pns:
            if total[1]:
                break
            if root_pn not in self.bom_index:
                self.status_var.set(f'⚠ [{root_pn}] 不在 BOM 中')
                continue
            r0 = self.bom_index[root_pn][0]
            root_values = [
                '1',
                str(r0[MOTHER_PN_COL]) if r0[MOTHER_PN_COL] else '',
                str(r0[2]) if len(r0) > 2 and r0[2] else '',
                str(r0[3]) if len(r0) > 3 and r0[3] else '',
                '', '', '', ''
            ]
            root_id = self.tree.insert('', 'end', text=root_pn,
                                       values=root_values, open=False)
            total[0] += 1
            self._rec_insert(root_id, root_pn, 1, max_depth, total, trail=frozenset({root_pn}))

        total_nodes = self._update_material_summary()
        if total[1]:
            self.status_var.set(
                f'⚠ BOM 展开已截断（达到 {MAX_ROWS} 行上限，可能有环路或 BOM 过大） |  当前 {total_nodes} 颗物料')
        elif unresolved_terms:
            missing_text = '、'.join(unresolved_terms[:5])
            suffix = ' 等' if len(unresolved_terms) > 5 else ''
            self.status_var.set(f'查询完成  |  当前BOM {total_nodes}颗物料 | 未匹配: {missing_text}{suffix}')
        elif expanded_terms:
            expand_text = '；'.join(expanded_terms[:3])
            suffix = ' …' if len(expanded_terms) > 3 else ''
            self.status_var.set(f'查询完成  |  当前BOM {total_nodes}颗物料 | 规格/名称匹配: {expand_text}{suffix}')
        else:
            self.status_var.set(f'查询完成  |  当前BOM {total_nodes}颗物料（含主料）')
        for item in self.tree.get_children(''):
            self.tree.item(item, open=True)

    def _rec_insert(self, parent_id, mother_pn, depth, max_depth, counter, trail=None):
        """递归展开节点。counter 形如 [count, truncated]；trail 是当前祖先路径用于环路检测。"""
        if counter[1]:
            return
        if depth >= max_depth or mother_pn not in self.bom_index:
            return
        if trail is None:
            trail = frozenset({mother_pn})
        for row in self.bom_index[mother_pn]:
            if counter[1]:
                return
            child_pn = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
            if not child_pn:
                continue
            next_d = depth + 1
            values = [
                str(next_d), '', '', '',
                child_pn,
                str(row[CHILD_NAME_COL]) if row[CHILD_NAME_COL] else '',
                str(row[CHILD_SPEC_COL]) if row[CHILD_SPEC_COL] else '',
                str(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else '',
            ]
            # 环路：child 已在祖先路径上，展开成标注节点但不再下钻
            is_cycle = child_pn in trail or child_pn == mother_pn
            if is_cycle:
                values[5] = (values[5] + ' [环路]').strip()
            has_kids = (not is_cycle) and bool(child_pn in self.bom_index)
            nid = self.tree.insert(parent_id, 'end', text=child_pn,
                                   values=values, open=False)
            counter[0] += 1
            if counter[0] > MAX_ROWS:
                counter[1] = True  # 触发整棵树停止
                return
            if has_kids:
                self._rec_insert(nid, child_pn, next_d, max_depth, counter, trail | {mother_pn})

    def _on_node_open(self, event=None):
        focused = self.tree.focus()
        if not focused or focused in self._opened:
            return
        self._opened.add(focused)
        mother_pn = self.tree.item(focused, 'text')
        if not mother_pn or mother_pn not in self.bom_index:
            return
        existing = self.tree.get_children(focused)
        if existing:
            return
        try:
            depth = int(self.tree.item(focused, 'values')[0])
        except (ValueError, IndexError):
            depth = 1
        counter = [0, False]
        # 懒加载时用当前节点及其所有祖先作为环路检测路径
        trail = set()
        cur = focused
        while cur:
            cur_pn = self.tree.item(cur, 'text')
            if cur_pn:
                trail.add(cur_pn)
            cur = self.tree.parent(cur)
        self._rec_insert(focused, mother_pn, depth, self.depth_var.get(), counter, frozenset(trail))
        suffix = '（已截断）' if counter[1] else ''
        self.status_var.set(f'+ 展开 [{mother_pn}]  |  +{counter[0]} 节点{suffix}')

    def _clear_tree(self):
        for item in self.tree.get_children(''):
            self.tree.delete(item)
        self._opened.clear()
        self._calc_result = None
        self.status_var.set(f'就绪  |  母器件: {len(self.bom_index)} 个')

    def _clear_calc(self):
        for item in self.calc_tree.get_children(''):
            self.calc_tree.delete(item)
        self._calc_result = None
        self._batch_result = None
        self.status_var.set(f'就绪  |  母器件: {len(self.bom_index)} 个')
        self._refresh_calc_hint()

    def _expand_all(self):
        for item in self.tree.get_children(''):
            self._open_rec(item)
        self.status_var.set('已展开全部节点')

    def _open_rec(self, node):
        self.tree.item(node, open=True)
        for ch in self.tree.get_children(node):
            self._open_rec(ch)

    def _collapse_all(self):
        for item in self.tree.get_children(''):
            self.tree.item(item, open=False)

    # ── 从选中节点计算 ───────────────────────────────────────
    def _calc_from_selected(self):
        """右键菜单：计算选中节点的用量（以该节点为顶层）"""
        sel = self.tree.selection()
        if not sel:
            return
        # 切换到数量计算选项卡
        self._switch_to_tab(self.tab_calc)
        # 把顶层数量设为1，从选中节点重新计算
        self.calc_qty_var.set('1')
        self._recalc_from_tree()

    def _reverse_lookup_from_entry(self):
        raw = self.reverse_pn_entry.get().strip() if hasattr(self, 'reverse_pn_entry') else ''
        if not raw:
            messagebox.showwarning('请输入子料号', '请输入要反查的子料号')
            return
        self._show_reverse_parent_lookup(raw)

    def _reverse_lookup_selected_tree_item(self):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], 'values')
        code = ''
        if len(vals) > 4 and str(vals[4]).strip():
            code = str(vals[4]).strip()
        else:
            code = str(self.tree.item(sel[0], 'text') or '').strip()
        if not code:
            return
        if hasattr(self, 'reverse_pn_entry'):
            self.reverse_pn_entry.delete(0, tk.END)
            self.reverse_pn_entry.insert(0, code)
        self._show_reverse_parent_lookup(code)

    def _build_reverse_parent_index(self):
        parent_map = defaultdict(list)
        for parent_pn, rows in self.bom_index.items():
            parent_name = ''
            parent_spec = ''
            if rows:
                first = rows[0]
                parent_name = str(first[2]).strip() if len(first) > 2 and first[2] else ''
                parent_spec = str(first[3]).strip() if len(first) > 3 and first[3] else ''
            for row in rows:
                child_pn = str(row[CHILD_PN_COL]).strip() if len(row) > CHILD_PN_COL and row[CHILD_PN_COL] else ''
                if not child_pn:
                    continue
                child_name = str(row[CHILD_NAME_COL]).strip() if len(row) > CHILD_NAME_COL and row[CHILD_NAME_COL] else ''
                child_spec = str(row[CHILD_SPEC_COL]).strip() if len(row) > CHILD_SPEC_COL and row[CHILD_SPEC_COL] else ''
                qty = row[CHILD_QTY_COL] if len(row) > CHILD_QTY_COL else ''
                parent_map[child_pn].append({
                    'parent': str(parent_pn).strip(),
                    'parent_name': parent_name,
                    'parent_spec': parent_spec,
                    'child': child_pn,
                    'child_name': child_name,
                    'child_spec': child_spec,
                    'qty': qty,
                })
        return parent_map

    def _collect_reverse_parent_rows(self, child_pn, max_depth=None):
        if max_depth is None:
            try:
                max_depth = int(self.depth_var.get())
            except Exception:
                max_depth = MAX_DEPTH
        parent_map = self._build_reverse_parent_index()
        normalized = normalize_material_code(child_pn) or str(child_pn).strip()
        targets = []
        if normalized:
            targets.append(normalized)
        raw = str(child_pn).strip()
        if raw and raw not in targets:
            targets.append(raw)

        direct_edges = []
        for target in targets:
            if target in parent_map:
                direct_edges = parent_map[target]
                child_pn = target
                break
        if not direct_edges:
            return str(child_pn).strip(), []

        rows = []
        queue = deque()
        seen_paths = set()
        for edge in direct_edges:
            parent = edge['parent']
            path = (parent, child_pn)
            queue.append((1, edge, path))

        while queue and len(rows) < MAX_ROWS:
            level, edge, path = queue.popleft()
            parent = edge['parent']
            path_text = ' ← '.join(path)
            path_key = (level, parent, edge['child'], path_text)
            if path_key in seen_paths:
                continue
            seen_paths.add(path_key)
            rows.append({
                '反查层级': level,
                '母料号': parent,
                '母件品名': edge.get('parent_name', ''),
                '母件规格': edge.get('parent_spec', ''),
                '直接下层料号': edge.get('child', ''),
                '直接下层品名': edge.get('child_name', ''),
                '直接下层规格': edge.get('child_spec', ''),
                '单位用量': edge.get('qty', ''),
                '路径': path_text,
            })
            if level >= max_depth:
                continue
            for upper_edge in parent_map.get(parent, []):
                upper_parent = upper_edge['parent']
                if upper_parent in path:
                    continue
                queue.append((level + 1, upper_edge, (upper_parent,) + path))
        return child_pn, rows

    def _show_reverse_parent_lookup(self, child_pn):
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return
        resolved_child, rows = self._collect_reverse_parent_rows(child_pn)

        win = tk.Toplevel(self.root)
        win.title(f'子物料反查上层母料号 ← {resolved_child}')
        win.geometry('1180x560')
        win.resizable(True, True)

        direct_count = sum(1 for row in rows if row.get('反查层级') == 1)
        unique_parent_count = len({row.get('母料号') for row in rows if row.get('母料号')})
        ttk.Label(
            win,
            text=f'子料号：{resolved_child}    直接上层 {direct_count} 个    全部上层 {unique_parent_count} 个    路径 {len(rows)} 条',
            font=('Microsoft YaHei', 10, 'bold'),
            foreground='#2F5496'
        ).pack(anchor='w', padx=10, pady=(8, 4))
        ttk.Label(
            win,
            text='反查层级=1 表示直接父项；层级越大表示继续往上追溯。路径格式为：最上层母料号 ← ... ← 当前子料号。',
            foreground='#6B7A8F',
            font=('Microsoft YaHei', 9),
            wraplength=1120,
        ).pack(anchor='w', padx=10, pady=(0, 6))

        cols = ('反查层级', '母料号', '母件品名', '母件规格', '直接下层料号', '直接下层品名', '直接下层规格', '单位用量', '路径')
        widths = (80, 145, 220, 180, 145, 220, 180, 90, 430)
        fr = ttk.Frame(win)
        fr.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        fr.columnconfigure(0, weight=1)
        fr.rowconfigure(0, weight=1)
        tv = ttk.Treeview(fr, columns=cols, show='headings', selectmode='extended')
        numeric_cols = {'反查层级', '单位用量'}
        for col, width in zip(cols, widths):
            anchor = 'e' if col in numeric_cols else 'w'
            tv.column(col, width=width, anchor=anchor)
            tv.heading(col, text=col)
        enable_treeview_sort(tv, list(cols), numeric_columns=numeric_cols)
        enable_treeview_copy(tv, list(cols))
        vsb = ttk.Scrollbar(fr, orient='vertical', command=tv.yview)
        hsb = ttk.Scrollbar(fr, orient='horizontal', command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        if rows:
            for row in sorted(rows, key=lambda r: (r.get('反查层级', 0), str(r.get('母料号', '')), str(r.get('路径', '')))):
                tv.insert('', 'end', values=[row.get(col, '') for col in cols])
        else:
            tv.insert('', 'end', values=('—', '未找到上层母料号', '', '', resolved_child, '', '', '', ''))

        def _copy_selected_or_all():
            item_ids = tv.selection() or tv.get_children('')
            lines = ['\t'.join(cols)]
            for item in item_ids:
                vals = tv.item(item, 'values')
                lines.append('\t'.join(str(v) for v in vals))
            win.clipboard_clear()
            win.clipboard_append('\n'.join(lines))
            self.status_var.set(f'已复制子物料反查结果 {len(item_ids)} 行')

        def _query_parent_bom():
            sel = tv.selection()
            if not sel:
                return
            vals = tv.item(sel[0], 'values')
            if len(vals) < 2:
                return
            parent = str(vals[1]).strip()
            if not parent or parent == '未找到上层母料号':
                return
            win.destroy()
            self.pn_entry.delete(0, tk.END)
            self.pn_entry.insert(0, parent)
            self._switch_to_tab(self.tab_tree)
            self._query()

        def _export_reverse_result():
            if not rows:
                messagebox.showinfo('无数据', '没有可导出的反查结果')
                return
            path = filedialog.asksaveasfilename(
                title='导出子物料反查结果',
                defaultextension='.xlsx',
                filetypes=[('Excel 文件', '*.xlsx')],
            )
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '子物料反查'
            hdr_font = Font(bold=True, color='000000')
            hdr_fill = PatternFill('solid', fgColor='D9EAF7')
            thin = Side(style='thin', color='AAAAAA')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, header in enumerate(cols, 1):
                cell = ws.cell(1, ci, header)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for ri, row in enumerate(rows, 2):
                for ci, header in enumerate(cols, 1):
                    cell = ws.cell(ri, ci, row.get(header, ''))
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            for ci, width in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(8, min(70, width / 7))
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}{ws.max_row}"
            ws.freeze_panes = 'A2'
            wb.save(path)
            self.status_var.set(f'已导出子物料反查结果: {path}')

        ctx = tk.Menu(tv, tearoff=0)
        ctx.add_command(label='复制选中/全部', command=_copy_selected_or_all)
        ctx.add_command(label='展开此母件 BOM', command=_query_parent_bom)
        ctx.add_separator()
        ctx.add_command(label='导出 Excel', command=_export_reverse_result)
        tv.bind('<Button-3>', lambda e: ctx.tk_popup(e.x_root, e.y_root))

        btn_frame = ttk.Frame(win)
        btn_frame.pack(side='bottom', fill='x', padx=8, pady=(0, 8))
        ttk.Button(btn_frame, text='复制选中/全部', command=_copy_selected_or_all).pack(side='left', padx=(0, 6))
        ttk.Button(btn_frame, text='导出 Excel', command=_export_reverse_result).pack(side='left', padx=(0, 6))
        ttk.Button(btn_frame, text='展开此母件 BOM', command=_query_parent_bom).pack(side='left', padx=(0, 6))
        ttk.Button(btn_frame, text='关闭', command=win.destroy).pack(side='right')
        self.status_var.set(f'子物料反查完成: {resolved_child} | 直接上层 {direct_count} 个 | 全部上层 {unique_parent_count} 个')

    # ── 复制工具 ──────────────────────────────────────────────
    def _copy_tree_cell(self, event=None):
        region = self.tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        col = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id or not col:
            return
        vals = self.tree.item(item_id, 'values')
        col_idx = int(col[1:]) - 1
        if col_idx < len(vals):
            val = str(vals[col_idx]).strip()
            if val:
                self.root.clipboard_clear()
                self.root.clipboard_append(val)
                self.status_var.set(f'已复制: {val}')

    def _copy_tree_row(self):
        sel = self.tree.selection()
        if not sel:
            return
        row = self.tree.item(sel[0], 'values')
        text = ' | '.join(str(v).strip() for v in row if str(v).strip())
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_var.set(f'已复制: {text[:80]}')

    def _copy_tree_pn(self):
        sel = self.tree.selection()
        if sel:
            pn = self.tree.item(sel[0], 'text')
            if pn:
                self.root.clipboard_clear()
                self.root.clipboard_append(str(pn))
                self.status_var.set(f'已复制料号: {pn}')

    def _copy_calc_cell(self, event=None):
        tree = event.widget if event is not None and hasattr(event, 'widget') else self.calc_tree
        if event is None:
            return
        region = tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        col = tree.identify_column(event.x)
        item_id = tree.identify_row(event.y)
        if not item_id or not col:
            return
        vals = tree.item(item_id, 'values')
        col_idx = int(col[1:]) - 1
        if col_idx < len(vals):
            val = str(vals[col_idx]).strip()
            if val:
                self.root.clipboard_clear()
                self.root.clipboard_append(val)
                self.status_var.set(f'已复制: {val}')

    def _copy_calc_row(self):
        sel = self.calc_tree.selection()
        if not sel:
            return
        row = self.calc_tree.item(sel[0], 'values')
        text = ' | '.join(str(v).strip() for v in row if str(v).strip())
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_var.set(f'已复制: {text[:80]}')

    def _copy_calc_pn(self):
        sel = self.calc_tree.selection()
        if sel:
            vals = self.calc_tree.item(sel[0], 'values')
            if len(vals) > 1:
                self.root.clipboard_clear()
                self.root.clipboard_append(str(vals[1]).strip())
                self.status_var.set(f'已复制料号: {vals[1]}')

    def _show_tree_menu(self, event=None):
        self._ctx_menu.tk_popup(event.x_root, event.y_root)

    def _show_calc_menu(self, event=None):
        self._calc_ctx.tk_popup(event.x_root, event.y_root)


    def _show_parent_usage(self):
        """查询选中料号的上一级物料及用量"""
        sel = self.calc_tree.selection()
        if not sel:
            return
        vals = self.calc_tree.item(sel[0], 'values')
        if len(vals) < 2:
            return
        child_pn = str(vals[1]).strip()
        if not child_pn:
            return
        if not self.bom_index:
            messagebox.showwarning('未加载 BOM', '请先上传 BOM 文件')
            return

        # 获取子件的汇总信息
        child_total = None
        child_paths = None
        if self._calc_result and child_pn in self._calc_result.get('flat', {}):
            info = self._calc_result['flat'][child_pn]
            child_total = info.get('total_qty')
            child_paths = info.get('paths', '')

        # 反向查找：哪些母件使用了这个子件
        parents = []
        for mother_pn, rows in self.bom_index.items():
            for row in rows:
                cpn = str(row[CHILD_PN_COL]).strip() if row[CHILD_PN_COL] else ''
                if cpn == child_pn:
                    mname = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    mspec = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                    try:
                        qty = float(row[CHILD_QTY_COL]) if row[CHILD_QTY_COL] else 1.0
                    except (ValueError, TypeError):
                        qty = 1.0
                    parents.append((mother_pn, mname, mspec, qty))
                    break

        # 创建弹窗
        win = tk.Toplevel(self.root)
        win.title(f'上一级物料用量 ← {child_pn}')
        win.geometry('920x440')
        win.resizable(True, True)

        # 标题
        if child_total is not None:
            title_text = (f'子件：{child_pn}    共 {len(parents)} 个上级母件    '
                          f'|    子件汇总用量：{fmt_qty(child_total)}（路径数：{child_paths}）')
        else:
            title_text = f'子件：{child_pn}    共 {len(parents)} 个上级母件'
        ttk.Label(win, text=title_text, font=('Microsoft YaHei', 10, 'bold'),
                  foreground='#2F5496').pack(anchor='w', padx=10, pady=(8, 4))

        ttk.Label(win, text='"单位用量" = 该母件 BOM 中对子件的直接用量；"母件汇总用量" = 该母件在当前计算结果中的累计总需求量',
                  foreground='#6B7A8F', font=('Microsoft YaHei', 9),
                  wraplength=880).pack(anchor='w', padx=10, pady=(0, 6))

        # 表格
        cols = ('母件料号', '母件品名', '母件规格', '单位用量', '母件汇总用量')
        widths = (150, 200, 170, 80, 100)

        fr = ttk.Frame(win)
        fr.pack(fill='both', expand=True, padx=8, pady=(0, 8))
        fr.columnconfigure(0, weight=1)
        fr.rowconfigure(0, weight=1)

        tv = ttk.Treeview(fr, columns=cols, show='headings', selectmode='extended')
        numeric_cols = {c for c in cols if any(k in c for k in ('数量', '用量', '层级'))}
        for col, w in zip(cols, widths):
            anchor = 'e' if col in numeric_cols else 'w'
            tv.column(col, width=w, anchor=anchor)
            tv.heading(col, text=col)
        enable_treeview_sort(tv, list(cols), numeric_columns=numeric_cols)
        enable_treeview_copy(tv, list(cols))

        vsb = ttk.Scrollbar(fr, orient='vertical', command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')

        # 填充数据
        if parents:
            for mpn, mname, mspec, qty in sorted(parents, key=lambda x: x[0]):
                m_total = None
                # 先尝试从 _calc_result 读取
                if self._calc_result and mpn in self._calc_result.get('flat', {}):
                    m_total = self._calc_result['flat'][mpn].get('total_qty')
                # 如果没有，尝试从 _batch_result 读取
                elif hasattr(self, '_batch_result') and self._batch_result and mpn in self._batch_result:
                    m_total = self._batch_result[mpn].get('total_qty')
                tv.insert('', 'end', values=(
                    mpn, mname, mspec, fmt_qty(qty),
                    fmt_qty(m_total) if m_total is not None else '—'
                ))
        else:
            tv.insert('', 'end', values=('— 未找到上级母件 —', '', '', '', '—'))

        # 双击复制
        def _on_dbl(e):
            region = tv.identify('region', e.x, e.y)
            if region != 'cell':
                return
            col_id = tv.identify_column(e.x)
            item_id = tv.identify_row(e.y)
            if not item_id or not col_id:
                return
            row_vals = tv.item(item_id, 'values')
            ci = int(col_id[1:]) - 1
            if ci < len(row_vals):
                v = str(row_vals[ci]).strip()
                if v and v != '— 未找到上级母件 —':
                    win.clipboard_clear()
                    win.clipboard_append(v)
                    self.status_var.set(f'已复制: {v}')

        tv.bind('<Double-Button-1>', _on_dbl)

        # 右键菜单
        ctx = tk.Menu(tv, tearoff=0)
        def _copy_row():
            sel2 = tv.selection()
            if sel2:
                r = tv.item(sel2[0], 'values')
                txt = '\t'.join(str(v) for v in r)
                win.clipboard_clear()
                win.clipboard_append(txt)
        def _query_parent():
            sel2 = tv.selection()
            if sel2:
                mpn = str(tv.item(sel2[0], 'values')[0]).strip()
                if mpn and mpn != '— 未找到上级母件 —':
                    win.destroy()
                    self.pn_entry.delete(0, tk.END)
                    self.pn_entry.insert(0, mpn)
                    self._switch_to_tab(self.tab_tree)
                    self._query()
        def _select_all():
            for item in tv.get_children():
                tv.selection_add(item)
        
        def _copy_all():
            lines = []
            for item in tv.get_children():
                vals = tv.item(item, 'values')
                lines.append('\t'.join(str(v) for v in vals))
            if lines:
                win.clipboard_clear()
                win.clipboard_append('\n'.join(lines))
                self.status_var.set(f'已复制 {len(lines)} 行')
        
        def _locate_in_calc():
            """定位到数量计算表格中的对应物料"""
            sel2 = tv.selection()
            if not sel2:
                return
            mpn = str(tv.item(sel2[0], 'values')[0]).strip()
            if mpn and mpn != '— 未找到上级母件 —':
                # 切换到数量计算页
                self._switch_to_tab(self.tab_calc)
                # 在数量计算表格中查找该料号
                for item in self.calc_tree.get_children():
                    vals = self.calc_tree.item(item, 'values')
                    if len(vals) > 1 and str(vals[1]).strip() == mpn:
                        # 清除之前的选择
                        self.calc_tree.selection_remove(self.calc_tree.selection())
                        # 选中并滚动到该项
                        self.calc_tree.selection_set(item)
                        self.calc_tree.see(item)
                        self.calc_tree.focus(item)
                        self.status_var.set(f'已定位到: {mpn}')
                        break
        
        ctx.add_command(label='复制整行', command=_copy_row)
        ctx.add_separator()
        ctx.add_command(label='📋 全选', command=_select_all)
        ctx.add_command(label='📄 复制全部', command=_copy_all)
        ctx.add_separator()
        ctx.add_command(label='📍 定位到对应物料', command=_locate_in_calc)
        ctx.add_command(label='🌲 展开此母件 BOM', command=_query_parent)
        tv.bind('<Button-3>', lambda e: ctx.tk_popup(e.x_root, e.y_root))

        # 底部按钮
        btn_frame2 = ttk.Frame(win)
        btn_frame2.pack(side='bottom', fill='x', padx=8, pady=(0, 8))
        ttk.Button(btn_frame2, text='📋 全选', command=_select_all).pack(side='left', padx=(0, 4))
        ttk.Button(btn_frame2, text='📄 复制全部', command=_copy_all).pack(side='left', padx=(0, 4))
        ttk.Button(btn_frame2, text='关闭', command=win.destroy).pack(side='right')

    # ── 导出 ──────────────────────────────────────────────────
    def _export_tree(self):
        items = self.tree.get_children('')
        if not items:
            messagebox.showwarning('无数据', '请先查询 BOM')
            return
        path = filedialog.asksaveasfilename(
            title='导出 BOM 树形展开结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'BOM 展开结果'
            hdr_font = Font(bold=True, color='FFFFFF', size=10)
            hdr_fill = PatternFill('solid', fgColor='2F5496')
            thin = Side(style='thin', color='AAAAAA')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ci, h in enumerate(HEADERS_TREE, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            def collect(node):
                rows = []
                for ch in self.tree.get_children(node):
                    rows.append(list(self.tree.item(ch, 'values')))
                    rows.extend(collect(ch))
                return rows

            all_rows = []
            for rid in items:
                all_rows.append(list(self.tree.item(rid, 'values')))
                all_rows.extend(collect(rid))

            even = PatternFill('solid', fgColor='DCE6F1')
            for ri, rd in enumerate(all_rows, 2):
                for ci, val in enumerate(rd, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = border
                    if ri % 2 == 0:
                        c.fill = even
                ws.row_dimensions[ri].height = 20

            for i, w in enumerate([8, 18, 28, 22, 18, 28, 22, 8], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
            self.status_var.set(f'已导出: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))

    def _export_calc(self):
        if not self._calc_result:
            messagebox.showwarning('无数据', '请先计算 BOM 用量')
            return
        items = self.calc_tree.get_children('')
        if not items:
            messagebox.showwarning('无数据', '无数据可导出')
            return
        path = filedialog.asksaveasfilename(
            title='导出 BOM 数量计算结果',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')]
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'BOM 数量计算'
            hdr_font = Font(bold=True, color='FFFFFF', size=10)
            hdr_fill = PatternFill('solid', fgColor='1F7A3C')
            thin = Side(style='thin', color='AAAAAA')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ci, h in enumerate(HEADERS_CALC, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font; c.fill = hdr_fill; c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            even = PatternFill('solid', fgColor='D9F0E2')
            for ri, item in enumerate(self.calc_tree.get_children(''), 2):
                vals = self.calc_tree.item(item, 'values')
                for ci, val in enumerate(vals, 1):
                    # 层级(1)、路径数(5)、汇总用量(6) 写数值，其余写字符串
                    if ci in (1, 5, 6):
                        try:
                            val = float(val) if val != '' else ''
                            if ci in (1, 5) and val != '':
                                val = int(val)
                        except (ValueError, TypeError):
                            pass
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = border
                    if ri % 2 == 0:
                        c.fill = even
                ws.row_dimensions[ri].height = 20

            for i, w in enumerate([6, 18, 28, 22, 10, 10], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo('导出成功', f'已保存至:\n{path}')
            self.status_var.set(f'已导出: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))


    def _do_quick_search(self):
        """快速搜索：在当前表格中查找匹配项"""
        query = self.search_var.get().strip()
        if not query:
            return

        # 根据当前 tab 选择合适的 Treeview 和搜索列
        current = self._get_current_tab()
        if current is getattr(self, 'tab_calc', None):
            tree = self.calc_tree
            search_cols = [1, 2, 3]
        elif current is getattr(self, 'tab_diff', None):
            tree = self.diff_tree
            search_cols = [1, 2, 3]
        else:
            tree = self.tree
            search_cols = [4, 5, 6]
        
        # 清空之前的搜索结果
        self._search_matches = []
        self._search_current_index = -1
        
        # 查找所有匹配项
        query_lower = query.lower()
        
        def search_tree(parent=''):
            items = tree.get_children(parent)
            for item in items:
                vals = tree.item(item, 'values')
                matched = False
                for col_idx in search_cols:
                    if col_idx < len(vals):
                        val_str = str(vals[col_idx]).lower()
                        if query_lower in val_str:
                            matched = True
                            break
                if matched:
                    self._search_matches.append(item)
                # 递归搜索子节点
                search_tree(item)
        
        search_tree()
        
        if self._search_matches:
            self._search_current_index = 0
            self._highlight_search_match()
            self.status_var.set(f'找到 {len(self._search_matches)} 个匹配项')
        else:
            self.status_var.set(f'未找到匹配项: {query}')
    
    def _search_next_match(self):
        """跳转到下一个匹配项"""
        if not self._search_matches:
            self.status_var.set('请先执行搜索')
            return
        
        self._search_current_index = (self._search_current_index + 1) % len(self._search_matches)
        self._highlight_search_match()
    
    def _search_prev_match(self):
        """跳转到上一个匹配项"""
        if not self._search_matches:
            self.status_var.set('请先执行搜索')
            return
        
        self._search_current_index = (self._search_current_index - 1) % len(self._search_matches)
        self._highlight_search_match()
    
    def _highlight_search_match(self):
        """高亮显示当前匹配项"""
        if not self._search_matches or self._search_current_index < 0:
            return

        current = self._get_current_tab()
        if current is getattr(self, 'tab_calc', None):
            tree = self.calc_tree
        elif current is getattr(self, 'tab_diff', None):
            tree = self.diff_tree
        else:
            tree = self.tree
        
        item = self._search_matches[self._search_current_index]
        
        # 清除之前的选择
        tree.selection_remove(tree.selection())
        
        # 选中并滚动到当前项
        tree.selection_set(item)
        tree.see(item)
        tree.focus(item)
        
        # 更新状态栏
        self.status_var.set(
            f'匹配项 {self._search_current_index + 1}/{len(self._search_matches)}'
        )

    # ── 全部母器件列表 ───────────────────────────────────────
    def _list_all_pns(self):
        if not self.bom_index:
            messagebox.showwarning('未加载', '请先上传 BOM 文件')
            return
        win = tk.Toplevel(self.root)
        win.title('所有母器件料号')
        win.geometry('760x580')

        fr = ttk.Frame(win)
        fr.pack(fill='both', expand=True, padx=6, pady=6)
        fr.columnconfigure(0, weight=1)
        fr.rowconfigure(1, weight=1)

        ttk.Label(fr, text='快速过滤:').grid(row=0, column=0, sticky='w', pady=(0, 4))
        fe = ttk.Entry(fr)
        fe.grid(row=0, column=0, sticky='we', padx=(65, 0), pady=(0, 4))

        tv = ttk.Treeview(fr, columns=('品名', '规格', '子件数'), show='tree', height=25)
        tv.column('#0', width=240, anchor='w')
        tv.heading('#0', text='母器件料号')
        tv.column('品名', width=260, anchor='w')
        tv.column('规格', width=180, anchor='w')
        tv.column('子件数', width=70, anchor='center')
        tv.grid(row=1, column=0, sticky='nsew')
        ttk.Scrollbar(fr, orient='vertical', command=tv.yview).grid(
            row=1, column=1, sticky='ns')

        def populate(ft=''):
            tv.delete(*tv.get_children(''))
            ft_l = ft.lower()
            cnt = 0
            for pn, rows in sorted(self.bom_index.items()):
                r = rows[0]
                pn_disp = str(r[MOTHER_PN_COL])[:40] if r[MOTHER_PN_COL] else ''
                name = str(r[2])[:40] if len(r) > 2 and r[2] else ''
                spec = str(r[3])[:30] if len(r) > 3 and r[3] else ''
                searchable = f'{pn_disp} {name} {spec}'.lower()
                if ft_l and ft_l not in searchable:
                    continue
                tv.insert('', 'end', text=pn_disp, values=(name, spec, str(len(rows))))
                cnt += 1
                if cnt > 500:
                    break

        fe.bind('<KeyRelease>', lambda e: populate(fe.get()))
        populate()

        def on_dbl(e=None):
            sel = tv.selection()
            if sel:
                win.destroy()
                self.pn_entry.delete(0, tk.END)
                self.pn_entry.insert(0, tv.item(sel[0], 'text'))
                self._query()
        tv.bind('<Double-Button-1>', on_dbl)

    def _import_pns_file(self):
        path = filedialog.askopenfilename(
            title='导入料号文件',
            filetypes=[('文本文件', '*.txt *.csv'), ('所有文件', '*.*')]
        )
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                content = f.read()
            text = content.replace('\n', ',').replace('，', ',')
            self.pn_entry.delete(0, tk.END)
            self.pn_entry.insert(0, text.strip(','))
            self.status_var.set(f'已导入: {os.path.basename(path)}')
        except Exception as e:
            messagebox.showerror('读取失败', str(e))

    # ── 上传 BOM ─────────────────────────────────────────────
    def _upload_bom(self):
        path = filedialog.askopenfilename(
            title='选择 BOM 文件',
            filetypes=[
                ('Excel 文件', '*.xlsx *.xls'),
                ('CSV 文件', '*.csv'),
                ('所有文件', '*.*')
            ]
        )
        if not path:
            return
        self.status_var.set(f'正在加载: {os.path.basename(path)} …')
        self.root.update_idletasks()
        try:
            bom_index, _ = load_bom_from_file(path)
            self.bom_index = bom_index
            self._diff_meta_cache = None
            self._bom_edges_cache = None
            self.current_file = path
            save_cache(bom_index, CACHE_FILE)
            self.file_var.set(os.path.basename(path))
            if hasattr(self, 'clean_bom_var'):
                self.clean_bom_var.set(path)
                if not self.clean_output_var.get().strip():
                    self.clean_output_var.set(self._default_clean_output_path())
            self.status_var.set(f'加载成功  |  母器件: {len(bom_index)} 个')
            self._clear_tree()
            self._clear_calc()
            self._clear_diff()
            self._clear_bom_supply()
            self._clear_balance_preview()
            self._clear_readiness_preview()
            self._update_balance_bom_badge()
            self._update_readiness_bom_badge()
            self._update_arrival_bom_badge()
            self._refresh_calc_hint()
        except Exception as e:
            messagebox.showerror('加载失败', str(e))
            self.status_var.set(f'加载失败: {e}')

    # ── 模糊搜索 ─────────────────────────────────────────────

    def _fuzzy_go(self, lb):
        sel = lb.curselection()
        if sel:
            item = lb.get(sel[0])
            # 提取料号（格式：料号 | 品名）
            pn = item.split(' | ')[0].strip()
            self.search_var.set(pn)
            self._close_search_popup()
            # 直接填入查询框并执行查询
            self.pn_entry.delete(0, tk.END)
            self.pn_entry.insert(0, pn)
            self._query()

    def _do_fuzzy_select(self, event=None):
        if self._search_listbox and self._search_listbox.winfo_viewable():
            sel = self._search_listbox.curselection()
            if sel:
                pn = self._search_listbox.get(sel[0])
                self.search_var.set(pn)
                self._close_search_popup()
                # 直接填入查询框并执行查询
                self.pn_entry.delete(0, tk.END)
                self.pn_entry.insert(0, pn)
                self._query()
            else:
                # 如果没有选中项，选中第一个
                if self._search_listbox.size() > 0:
                    self._search_listbox.selection_set(0)
                    self._do_fuzzy_select()


    def _search_next(self):
        """选择下一个搜索结果"""
        if not self._search_listbox:
            return
        sel = self._search_listbox.curselection()
        if sel:
            current = sel[0]
            if current < self._search_listbox.size() - 1:
                self._search_listbox.selection_clear(current)
                self._search_listbox.selection_set(current + 1)
                self._search_listbox.see(current + 1)

    def _search_prev(self):
        """选择上一个搜索结果"""
        if not self._search_listbox:
            return
        sel = self._search_listbox.curselection()
        if sel:
            current = sel[0]
            if current > 0:
                self._search_listbox.selection_clear(current)
                self._search_listbox.selection_set(current - 1)
                self._search_listbox.see(current - 1)

    def _close_search_popup(self):
        if hasattr(self, '_search_popup_frame') and self._search_popup_frame:
            try:
                self._search_popup_frame.destroy()
            except:
                pass
            self._search_popup_frame = None
        if self._search_listbox:
            try:
                self._search_listbox.destroy()
            except:
                pass
            self._search_listbox = None

    def run(self):
        self.root.mainloop()


def run():
    app = BOMApp()
    app.run()
    return 0


if __name__ == '__main__':
    raise SystemExit(run())


