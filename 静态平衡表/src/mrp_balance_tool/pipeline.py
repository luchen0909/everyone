from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict

# 浮点比较的容差：Excel 导入的小数常有 1e-12 级误差，避免齐套/缺料边界误判
FLOAT_ABS_TOL = 1e-6


def _fge(a: float, b: float) -> bool:
    """a >= b，使用绝对容差。"""
    return a > b or math.isclose(a, b, abs_tol=FLOAT_ABS_TOL)


def _fle(a: float, b: float) -> bool:
    """a <= b，使用绝对容差。"""
    return a < b or math.isclose(a, b, abs_tol=FLOAT_ABS_TOL)
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


SHORTAGE_COLUMNS = [
    "辅助列2",
    "NO.",
    "客户",
    "母件料号",
    "母件品名",
    "上层来源品名",
    "母件规格",
    "版本号",
    "上线日期",
    "上线数量",
    "BOM校核",
    "BOM2.BOM子项.子件料品.料号",
    "BOM2.BOM子项.子件料品.参考料号1",
    "BOM2.BOM子项.子件料品.料品名称",
    "BOM2.BOM子项.子件料品.规格",
    "BOM2.BOM子项.子件用量",
    "需求",
    "累计需求",
    "本体库存",
    "替代料库存",
    "替代料清单",
    "库存",
    "累计到货",
    "累计缺料",
    "累计缺料2",
    "到货数量",
    "到货日期",
    "请购日期",
    "L/T",
    "流水号",
    "辅助列",
    "工单缺料",
]

PURCHASE_VIEW_FIXED_COLUMNS = [
    "辅助列",
    "序号",
    "料号",
    "料品名称",
    "规格",
    "供应商",
    "采购",
    "未清PO",
    "未转",
    None,
    "替代1",
    "替代2",
    "项目",
    "项目需求",
    "项目短缺",
    "单套用量",
    "项目风险",
    "供应商库存",
    "实时库存",
    "总需求",
    "工单缺料",
    "4月需求",
    "5月需求",
    "6月需求",
    "备注",
    "停线预警",
    "父项分类2",
]

BALANCE_ROW_LABELS = ["需求数量", "采购答复", "到货计划", "到货数量", "领用数量", "差异"]
BALANCE_BLOCK_SIZE = len(BALANCE_ROW_LABELS)
BALANCE_DEMAND_OFFSET = 0
BALANCE_REPLY_OFFSET = 1
BALANCE_ARRIVAL_PLAN_OFFSET = 2
BALANCE_ARRIVAL_QTY_OFFSET = 3
BALANCE_USAGE_OFFSET = 4
BALANCE_DIFF_OFFSET = 5
LEGACY_BALANCE_BLOCK_SIZE = 5

MANUAL_ARRIVALS_HEADERS = [
    "ID",
    "外部单据编号",
    "编码",
    "采购订单",
    "业务类型",
    "供应商",
    "到货日期",
    "物料",
    "物料编码",
    "规格",
    "计划数量",
    "到货数量",
    "拒收退回数量",
    "入库数量",
    "单位",
    "计划状态",
    "委外",
    "执行组织",
    "货主",
    "VMI",
    "关闭",
    "备注",
]

USAGE_HEADERS = [
    "ID",
    "记账日期",
    "流水业务类型",
    "记账方向",
    "仓库",
    "物料",
    "物料编码",
    "规格",
    "单位",
    "作业",
    "货主",
    "业务数量",
    "库存状态",
    "VMI",
    "供应商",
    "来源实体编码",
    "来源实体名称",
    "来源实体ID",
    "同步时间",
    "同步状态",
    "同步异常信息",
]
USAGE_START_DATETIME_CELL = "$L$2"
USAGE_START_DATETIME_COL = 12
INVENTORY_CUTOFF_DATETIME_CELL = "$M$2"
INVENTORY_CUTOFF_DATETIME_COL = 13
RISK_CONFIG_LABEL_COL = 12
RISK_CONFIG_VALUE_COL = 13
RISK_EXTREME_DAYS_CELL = "$M$5"
RISK_HIGH_DAYS_CELL = "$M$6"
RISK_MEDIUM_DAYS_CELL = "$M$7"
RISK_LOW_BUFFER_DAYS_CELL = "$M$8"

PROJECT_DETAIL_SHEET_NAME = "项目明细"
BALANCE_HIERARCHY_FILTER_SHEET_NAME = "BOM层级筛选"
PURCHASE_REPLY_SUMMARY_SHEET_NAME = "采购待回复统计"
CALCULATION_EXPLANATION_SHEET_NAME = "计算说明"
PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME = "母料用量排除清单"
UPPER_EXPANSION_CHECK_SHEET_NAME = "上层展开校验"
NEAR_TERM_SHORTAGE_SHEET_NAME = "近三天排产缺料"
RECEIVING_STATUS_SHEET_NAME = "接收待检待入库"
INDUSTRIAL_CONFIG_SHEET_NAME = "工业配置表"
INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME = "工业发货配件"
CONFIG_SUPPLEMENT_PLAN_SHEET_NAME = "配置补排产"
ENABLE_CONFIG_SUPPLEMENT_PLAN = False
USAGE_SHEET_NAME = "领用"
ARRIVAL_RECORD_SHEET_NAME = "到货记录"
USAGE_START_DATETIME_SHEET_NAME = "生产计划"
MANAGEMENT_DASHBOARD_SHEET_NAME = "管理驾驶舱"
PROJECT_RISK_VIEW_SHEET_NAME = "项目风险视图"
PURCHASE_ACTION_VIEW_SHEET_NAME = "采购待办视图"
MATERIAL_CONTROL_VIEW_SHEET_NAME = "物控缺口视图"
PRODUCTION_SHORTAGE_VIEW_SHEET_NAME = "生产T+3视图"
DATA_QUALITY_VIEW_SHEET_NAME = "数据质量视图"
INFO_AUDIT_VIEW_SHEET_NAME = "信息化审计"
ROLE_QUALITY_SHEET_NAMES = [
    MANAGEMENT_DASHBOARD_SHEET_NAME,
    PROJECT_RISK_VIEW_SHEET_NAME,
    PURCHASE_ACTION_VIEW_SHEET_NAME,
    MATERIAL_CONTROL_VIEW_SHEET_NAME,
    PRODUCTION_SHORTAGE_VIEW_SHEET_NAME,
    DATA_QUALITY_VIEW_SHEET_NAME,
    INFO_AUDIT_VIEW_SHEET_NAME,
]


class WorkbookInputError(ValueError):
    """Raised when the workbook structure does not match the expected format."""


@dataclass
class PipelineResult:
    shortage_df: pd.DataFrame
    exported_shortage_df: pd.DataFrame
    purchase_view_df: pd.DataFrame
    production_plan_df: pd.DataFrame
    upper_expansion_df: pd.DataFrame
    hierarchy_filter_df: pd.DataFrame
    near_term_shortage_df: pd.DataFrame
    config_supplement_plan_df: pd.DataFrame
    usage_df: pd.DataFrame
    arrival_record_df: pd.DataFrame
    date_headers: list[date]
    suggestion_df: pd.DataFrame
    carried_reply_cell_count: int = 0
    carried_reply_material_count: int = 0
    carried_reply_file_count: int = 0
    usage_start_datetime: datetime | None = None


@dataclass
class PurchaseReadinessResult:
    root_summary_df: pd.DataFrame
    issue_df: pd.DataFrame
    material_df: pd.DataFrame
    producible_df: pd.DataFrame
    horizontal_shortage_df: pd.DataFrame
    diff_capacity_df: pd.DataFrame
    rolling_matrix_df: pd.DataFrame
    missing_root_df: pd.DataFrame
    unknown_leaf_df: pd.DataFrame
    batch_summary: dict[str, object]
    carried_reply_cell_count: int = 0
    carried_reply_material_count: int = 0
    carried_reply_file_count: int = 0


@dataclass
class ArrivalStatusResult:
    quality_df: pd.DataFrame
    warehouse_df: pd.DataFrame
    all_df: pd.DataFrame
    purchase_pending_df: pd.DataFrame
    summary: dict[str, object]


MATERIAL_CODE_PATTERN = re.compile(r"^[A-Za-z0-9]+(?:-[A-Za-z0-9]+)?$")

ProgressCallback = Callable[[str], None]


def emit_progress(progress_callback: ProgressCallback | None, message: str) -> None:
    if progress_callback is not None:
        progress_callback(message)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for column in df.columns:
        if isinstance(column, str):
            renamed[column] = column.strip().replace("\n", "").replace("\r", "")
    return df.rename(columns=renamed)


def first_matching_column(
    df: pd.DataFrame,
    candidates: Iterable[str],
    *,
    contains: bool = False,
    required: bool = True,
) -> str | None:
    columns = [str(column).strip() for column in df.columns]
    for candidate in candidates:
        if contains:
            for original, current in zip(df.columns, columns):
                if candidate in current:
                    return str(original)
        else:
            for original, current in zip(df.columns, columns):
                if current == candidate:
                    return str(original)
    if required:
        raise WorkbookInputError(f"缺少字段: {', '.join(candidates)}")
    return None


def coerce_date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.date


def coerce_number_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def extract_material_code(value) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    match = re.search(r"[A-Za-z0-9]+-\d+", text)
    return match.group(0) if match else text


BALANCE_SHEET_NAME = "\u4f9b\u9700\u5e73\u8861"
PURCHASE_REPLY_ROW_LABEL = "\u91c7\u8d2d\u7b54\u590d"


def coerce_excel_date(value) -> date | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def has_manual_reply_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        text = value.strip()
        return bool(text) and text != "0"
    if isinstance(value, (int, float)):
        return float(value) != 0
    return True


def load_carried_purchase_replies(
    source_paths: Iterable[Path],
    *,
    cutoff_date: date,
) -> tuple[dict[tuple[str, date], object], dict[str, int]]:
    reply_map: dict[tuple[str, date], object] = {}
    source_file_count = 0
    carried_cell_count = 0

    for raw_path in source_paths:
        path = Path(raw_path).expanduser().resolve()
        if not path.exists():
            raise WorkbookInputError(f"\u65e7\u5e73\u8861\u8868\u4e0d\u5b58\u5728: {path}")

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if BALANCE_SHEET_NAME in workbook.sheetnames:
                worksheet = workbook[BALANCE_SHEET_NAME]
            else:
                worksheet = workbook[workbook.sheetnames[0]]

            header_dates: dict[int, date] = {}
            reply_label_col: int | None = None
            current_material_code = ""

            def row_value(row_values: tuple[object, ...], index_1_based: int):
                index_0_based = index_1_based - 1
                if index_0_based < 0 or index_0_based >= len(row_values):
                    return None
                return row_values[index_0_based]

            for row_index, row_values in enumerate(
                worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True),
                start=2,
            ):
                if row_index == 2:
                    all_date_columns: list[int] = []
                    for column_index in range(1, len(row_values) + 1):
                        header_date = coerce_excel_date(row_value(row_values, column_index))
                        if not header_date:
                            continue
                        all_date_columns.append(column_index)
                        if header_date >= cutoff_date:
                            header_dates[column_index] = header_date
                    if all_date_columns:
                        reply_label_col = min(all_date_columns) - 1
                    if header_dates:
                        source_file_count += 1
                    continue

                if not header_dates:
                    continue

                row_material_code = extract_material_code(row_value(row_values, 3))
                if row_material_code:
                    current_material_code = row_material_code

                row_label = str(row_value(row_values, reply_label_col or 0) or "").strip()
                if row_label != PURCHASE_REPLY_ROW_LABEL or not current_material_code:
                    continue

                for column_index, header_date in header_dates.items():
                    reply_value = row_value(row_values, column_index)
                    if not has_manual_reply_value(reply_value):
                        continue
                    key = (current_material_code, header_date)
                    if key not in reply_map:
                        carried_cell_count += 1
                    reply_map[key] = reply_value
        finally:
            workbook.close()

    stats = {
        "file_count": source_file_count,
        "cell_count": carried_cell_count,
        "material_count": len({material_code for material_code, _header_date in reply_map}),
    }
    return reply_map, stats


def load_carried_balance_remarks(source_paths: Iterable[Path]) -> tuple[dict[str, str], dict[str, int]]:
    remark_map: dict[str, str] = {}
    source_file_count = 0
    carried_remark_count = 0

    def clean_text(value: object) -> str:
        if value is None:
            return ""
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            pass
        return str(value).strip()

    for raw_path in source_paths:
        path = Path(raw_path).expanduser().resolve()
        if not path.exists():
            raise WorkbookInputError(f"旧平衡表不存在: {path}")

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            worksheet = workbook[BALANCE_SHEET_NAME] if BALANCE_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            header_values = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if not header_values:
                continue

            remark_col_index_0 = None
            for index_0, value in enumerate(header_values):
                header_text = clean_text(value)
                if header_text in {"备注", "说明"}:
                    remark_col_index_0 = index_0
                    break
            if remark_col_index_0 is None:
                continue
            source_file_count += 1

            current_material_code = ""
            for row_index, row_values in enumerate(
                worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True),
                start=3,
            ):
                block_offset = (row_index - 3) % BALANCE_BLOCK_SIZE
                if block_offset == 0:
                    current_material_code = extract_material_code(row_values[2] if len(row_values) > 2 else "")
                    remark_value = row_values[remark_col_index_0] if remark_col_index_0 < len(row_values) else None
                    remark_text = clean_text(remark_value)
                    if current_material_code and remark_text:
                        previous = remark_map.get(current_material_code, "")
                        merged = _merge_note_texts_v2(previous, remark_text)
                        if not previous:
                            carried_remark_count += 1
                        remark_map[current_material_code] = merged
        finally:
            workbook.close()

    stats = {
        "file_count": source_file_count,
        "remark_count": carried_remark_count,
        "material_count": len(remark_map),
    }
    return remark_map, stats


def maybe_material_code(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if MATERIAL_CODE_PATTERN.fullmatch(text):
        return text
    match = re.search(r"[A-Za-z0-9]+-\d+", text)
    if match:
        return match.group(0)
    if text.isdigit():
        return text
    return None


def coerce_reply_quantity(value) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_quantity_text(value: float) -> str:
    if math.isclose(float(value), int(float(value)), abs_tol=FLOAT_ABS_TOL):
        return str(int(round(float(value))))
    return f"{float(value):.6f}".rstrip("0").rstrip(".")


def format_short_date(value: date) -> str:
    return f"{value.month}/{value.day}"


def load_balance_purchase_reply_schedule(
    balance_path: Path,
    *,
    start_date: date | None = None,
) -> tuple[dict[str, list[list[object]]], dict[str, int]]:
    path = Path(balance_path).expanduser().resolve()
    if not path.exists():
        raise WorkbookInputError(f"静态平衡表不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    schedules: dict[str, list[list[object]]] = defaultdict(list)
    reply_cell_count = 0
    try:
        worksheet = workbook[BALANCE_SHEET_NAME] if BALANCE_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header_values = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if not header_values:
            raise WorkbookInputError("静态平衡表缺少表头行")
        header_texts = [str(value or "").strip() for value in header_values]
        row_label_index = header_texts.index("父项分类2") if "父项分类2" in header_texts else 24
        date_columns: list[tuple[int, date]] = []
        has_any_date_header = False
        for index_0, value in enumerate(header_values):
            header_date = coerce_excel_date(value)
            if header_date:
                has_any_date_header = True
            if start_date and header_date and header_date < start_date:
                continue
            if header_date:
                date_columns.append((index_0, header_date))
        if not has_any_date_header:
            raise WorkbookInputError("静态平衡表未识别到横向日期列")

        for row_values in worksheet.iter_rows(min_row=3, values_only=True):
            row_label = str(
                row_values[row_label_index]
                if len(row_values) > row_label_index and row_values[row_label_index] is not None
                else ""
            ).strip()
            if row_label != PURCHASE_REPLY_ROW_LABEL:
                continue
            material_code = extract_material_code(row_values[2] if len(row_values) > 2 else "")
            if not material_code:
                continue
            for index_0, header_date in date_columns:
                reply_value = row_values[index_0] if index_0 < len(row_values) else None
                reply_qty = coerce_reply_quantity(reply_value)
                if reply_qty is None or reply_qty <= FLOAT_ABS_TOL:
                    continue
                schedules[material_code].append([header_date, float(reply_qty)])
                reply_cell_count += 1
    finally:
        workbook.close()

    for material_code in list(schedules):
        schedules[material_code].sort(key=lambda item: item[0])

    stats = {
        "material_count": len(schedules),
        "reply_cell_count": reply_cell_count,
        "reply_total_qty": int(round(sum(float(qty) for values in schedules.values() for _reply_date, qty in values))),
    }
    return schedules, stats


def _workbook_first_sheet_dataframes(path: Path) -> list[pd.DataFrame]:
    file_path = Path(path).expanduser().resolve()
    if not file_path.exists():
        raise WorkbookInputError(f"文件不存在: {file_path}")
    frames: list[pd.DataFrame] = []
    try:
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            frame = normalize_columns(pd.read_excel(file_path, sheet_name=sheet_name))
            if not frame.empty or len(frame.columns) > 0:
                frames.append(frame)
    except Exception as exc:
        raise WorkbookInputError(f"读取文件失败: {file_path}；{exc}") from exc
    return frames


def _optional_column(df: pd.DataFrame, candidates: Iterable[str], *, contains: bool = False) -> str | None:
    try:
        return first_matching_column(df, candidates, contains=contains, required=False)
    except Exception:
        return None


def build_usage_df_from_inventory_flow(flow_path: Path | None, *, start_datetime: datetime | None = None) -> pd.DataFrame:
    if not flow_path:
        return pd.DataFrame(columns=USAGE_HEADERS)
    source_path = Path(flow_path).expanduser().resolve()
    if not source_path.exists():
        raise WorkbookInputError(f"库存流水记录不存在: {source_path}")

    rows: list[pd.DataFrame] = []
    for frame in _workbook_first_sheet_dataframes(source_path):
        date_col = _optional_column(frame, ["记账日期", "业务日期", "日期"], contains=True)
        type_col = _optional_column(frame, ["流水业务类型", "业务类型", "单据类型"], contains=True)
        direction_col = _optional_column(frame, ["记账方向", "出入库方向", "方向"], contains=True)
        material_col = _optional_column(frame, ["物料编码", "物料.编码", "料号", "物料号", "物料代码"])
        qty_col = _optional_column(frame, ["业务数量", "数量", "出库数量"], contains=True)
        if not date_col or not type_col or not direction_col or not material_col or not qty_col:
            continue
        work = frame.copy()
        for header in USAGE_HEADERS:
            if header not in work.columns:
                work[header] = ""
        work["记账日期"] = pd.to_datetime(work[date_col], errors="coerce")
        work["物料编码"] = work[material_col].astype(str).str.strip()
        work["流水业务类型"] = work[type_col].astype(str).str.strip()
        work["记账方向"] = work[direction_col].astype(str).str.strip()
        work["业务数量"] = coerce_number_series(work[qty_col])
        work = work[
            (work["物料编码"] != "")
            & (work["记账日期"].notna())
            & (work["业务数量"] > 0)
            & (work["记账方向"].str.contains("入库|出库", na=False, regex=True))
        ].copy()
        if start_datetime is not None and not work.empty:
            work = work.loc[pd.to_datetime(work["记账日期"], errors="coerce") >= start_datetime].copy()
        if not work.empty:
            rows.append(work[USAGE_HEADERS])

    if not rows:
        return pd.DataFrame(columns=USAGE_HEADERS)
    return pd.concat(rows, ignore_index=True)[USAGE_HEADERS]


def load_work_order_inventory_pool(stock_path: Path | None) -> tuple[dict[str, float], dict[str, int]]:
    if not stock_path:
        return {}, {"material_count": 0, "row_count": 0}
    inventory: defaultdict[str, float] = defaultdict(float)
    row_count = 0
    for frame in _workbook_first_sheet_dataframes(Path(stock_path)):
        material_col = _optional_column(frame, ["物料编码", "物料.编码", "料号", "物料号", "物料代码", "存货编码"])
        qty_col = _optional_column(
            frame,
            ["可用量", "库存量", "库存数量", "现存量", "结存数量", "结余库存", "数量"],
            contains=True,
        )
        if not material_col or not qty_col:
            continue
        work = pd.DataFrame(
            {
                "料号": frame[material_col].astype(str).str.strip(),
                "数量": coerce_number_series(frame[qty_col]),
            }
        )
        work = work[(work["料号"] != "") & (work["数量"] > 0)].copy()
        row_count += len(work)
        for row in work.itertuples(index=False):
            inventory[str(row.料号).strip()] += float(row.数量 or 0)
    return dict(inventory), {"material_count": len(inventory), "row_count": row_count}


ARRIVAL_STATUS_ORDER = ["接收", "待检", "在检", "待入库"]
ARRIVAL_RECORD_DEFAULT_HEADERS = [
    "ID",
    "编码",
    "到货计划",
    "到货日期",
    "供应商",
    "物料",
    "物料编码",
    "规格",
    "批号",
    "批次入库时间",
    "货主",
    "到货数量",
    "拒收数量",
    "破坏数量",
    "单位",
    "业务类型",
    "检验级别",
    "收货状态",
    "仓库",
    "委外",
    "VMI",
    "备注",
]


def load_work_order_arrival_pool(arrival_path: Path | None) -> tuple[dict[str, dict[str, float]], dict[str, int]]:
    if not arrival_path:
        return {}, {"material_count": 0, "row_count": 0}
    arrivals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    row_count = 0
    for frame in _workbook_first_sheet_dataframes(Path(arrival_path)):
        material_col = _optional_column(frame, ["物料编码", "物料.编码", "料号", "物料号", "物料代码"])
        if not material_col:
            material_col = _optional_column(frame, ["编码"])
        if not material_col:
            continue

        status_qty_columns: dict[str, str] = {}
        for status in ARRIVAL_STATUS_ORDER:
            column = _optional_column(frame, [f"{status}数量", status], contains=True)
            if column:
                status_qty_columns[status] = column
        if status_qty_columns:
            for status, qty_col in status_qty_columns.items():
                work = pd.DataFrame(
                    {
                        "料号": frame[material_col].astype(str).str.strip(),
                        "数量": coerce_number_series(frame[qty_col]),
                    }
                )
                work = work[(work["料号"] != "") & (work["数量"] > 0)].copy()
                row_count += len(work)
                for row in work.itertuples(index=False):
                    arrivals[str(row.料号).strip()][status] += float(row.数量 or 0)
            continue

        status_col = _optional_column(frame, ["收货状态", "状态", "计划状态", "检验状态", "入库状态", "业务类型"], contains=True)
        qty_col = _optional_column(frame, ["可用数量", "可催数量", "未到货数量", "到货数量", "计划数量", "数量"], contains=True)
        reject_col = _optional_column(frame, ["拒收数量", "拒收退回数量"], contains=True)
        damage_col = _optional_column(frame, ["破坏数量"], contains=True)
        if not status_col or not qty_col:
            continue
        for row in frame.to_dict("records"):
            material_code = str(row.get(material_col, "") or "").strip()
            status_text = str(row.get(status_col, "") or "").strip()
            if not material_code:
                continue
            matched_status = next((status for status in ARRIVAL_STATUS_ORDER if status in status_text), "")
            if not matched_status:
                continue
            qty = coerce_reply_quantity(row.get(qty_col)) or 0
            if reject_col:
                qty -= coerce_reply_quantity(row.get(reject_col)) or 0
            if damage_col:
                qty -= coerce_reply_quantity(row.get(damage_col)) or 0
            if qty <= FLOAT_ABS_TOL:
                continue
            arrivals[material_code][matched_status] += float(qty)
            row_count += 1
    return {code: dict(values) for code, values in arrivals.items()}, {"material_count": len(arrivals), "row_count": row_count}


def read_arrival_record_dataframe(arrival_record_path: Path | None) -> pd.DataFrame:
    if not arrival_record_path:
        return pd.DataFrame(columns=ARRIVAL_RECORD_DEFAULT_HEADERS)
    file_path = Path(arrival_record_path).expanduser().resolve()
    if not file_path.exists():
        raise WorkbookInputError(f"到货记录文件不存在: {file_path}")
    try:
        frame = pd.read_excel(file_path, sheet_name=0)
    except Exception as exc:
        raise WorkbookInputError(f"读取到货记录失败: {file_path}；{exc}") from exc
    frame = frame.dropna(how="all").copy()
    frame.columns = [str(column).strip() for column in frame.columns]
    if frame.empty and not len(frame.columns):
        return pd.DataFrame(columns=ARRIVAL_RECORD_DEFAULT_HEADERS)
    return frame


def format_arrival_record_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    max_col = max(len(dataframe.columns), 1)
    max_row = max(worksheet.max_row, 1)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1][:max_col]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 34,
        "B": 18,
        "C": 18,
        "D": 20,
        "E": 24,
        "F": 22,
        "G": 18,
        "H": 24,
        "I": 18,
        "J": 20,
        "K": 26,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 8,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 12,
        "T": 8,
        "U": 8,
        "V": 24,
    }
    for column_letter, width in widths.items():
        if worksheet.max_column >= column_index_from_string(column_letter):
            worksheet.column_dimensions[column_letter].width = width


def _allocate_supply_quantity(schedule: list[list[object]], required_qty: float) -> tuple[list[tuple[object, float]], float]:
    remaining_required = max(float(required_qty or 0), 0.0)
    allocations: list[tuple[object, float]] = []
    if remaining_required <= FLOAT_ABS_TOL:
        return allocations, 0.0

    for item in schedule:
        if remaining_required <= FLOAT_ABS_TOL:
            break
        available_qty = float(item[1] or 0)
        if available_qty <= FLOAT_ABS_TOL:
            continue
        used_qty = min(available_qty, remaining_required)
        item[1] = available_qty - used_qty
        remaining_required -= used_qty
        allocations.append((item[0], used_qty))

    if remaining_required <= FLOAT_ABS_TOL:
        remaining_required = 0.0
    return allocations, remaining_required


def _format_work_order_reply(allocations: list[tuple[object, float]], remaining_qty: float) -> str:
    if not allocations and remaining_qty > FLOAT_ABS_TOL:
        return f"无可用供给；缺{format_quantity_text(remaining_qty)}"
    parts = []
    for source_label, reply_qty in allocations:
        if reply_qty <= FLOAT_ABS_TOL:
            continue
        qty_text = format_quantity_text(reply_qty)
        if isinstance(source_label, date):
            parts.append(f"{format_short_date(source_label)}到货{qty_text}")
        elif str(source_label) == "库存":
            parts.append(f"库存满足{qty_text}")
        else:
            parts.append(f"{source_label}{qty_text}")
    if remaining_qty > FLOAT_ABS_TOL:
        parts.append(f"剩余缺口{format_quantity_text(remaining_qty)}")
    return "；".join(parts)


def fill_work_order_shortage_replies(
    shortage_path: Path,
    balance_path: Path,
    output_path: Path | None = None,
    *,
    preserve_existing: bool = True,
    stock_path: Path | None = None,
    arrival_path: Path | None = None,
    reply_start_date: date | None = None,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, object]:
    shortage_file = Path(shortage_path).expanduser().resolve()
    if not shortage_file.exists():
        raise WorkbookInputError(f"工单缺料清单不存在: {shortage_file}")
    if output_path is None:
        output_file = shortage_file.with_name(f"{shortage_file.stem}_工单缺料回复.xlsx")
    else:
        output_file = Path(output_path).expanduser().resolve()
    output_file.parent.mkdir(parents=True, exist_ok=True)

    reply_start = reply_start_date or date.today()
    emit_progress(progress_callback, "步骤 1/5：读取静态平衡表采购答复...")
    reply_schedules, reply_stats = load_balance_purchase_reply_schedule(Path(balance_path), start_date=reply_start)
    emit_progress(
        progress_callback,
        f"步骤 1/5：采购答复读取完成，料号 {reply_stats['material_count']} 个，单元格 {reply_stats['reply_cell_count']} 个",
    )
    emit_progress(progress_callback, "步骤 2/5：读取库存明细...")
    inventory_pool, inventory_stats = load_work_order_inventory_pool(stock_path)
    emit_progress(
        progress_callback,
        f"步骤 2/5：库存读取完成，料号 {inventory_stats['material_count']} 个，明细行 {inventory_stats['row_count']} 行",
    )
    emit_progress(progress_callback, "步骤 3/5：读取到货记录...")
    arrival_pool, arrival_stats = load_work_order_arrival_pool(arrival_path)
    emit_progress(
        progress_callback,
        f"步骤 3/5：到货记录读取完成，料号 {arrival_stats['material_count']} 个，明细行 {arrival_stats['row_count']} 行",
    )

    material_schedules: dict[str, list[list[object]]] = {}
    material_codes = set(reply_schedules) | set(inventory_pool) | set(arrival_pool)
    for material_code in material_codes:
        schedule: list[list[object]] = []
        inventory_qty = float(inventory_pool.get(material_code, 0) or 0)
        if inventory_qty > FLOAT_ABS_TOL:
            schedule.append(["库存", inventory_qty])
        for status in ARRIVAL_STATUS_ORDER:
            qty = float(arrival_pool.get(material_code, {}).get(status, 0) or 0)
            if qty > FLOAT_ABS_TOL:
                schedule.append([status, qty])
        schedule.extend(reply_schedules.get(material_code, []))
        material_schedules[material_code] = schedule

    emit_progress(progress_callback, "步骤 4/5：读取工单缺料清单并分配库存/到货/采购答复...")
    workbook = load_workbook(shortage_file, read_only=False, data_only=False)
    value_workbook = load_workbook(shortage_file, read_only=True, data_only=True)
    try:
        sheet_name = "配送计划" if "配送计划" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        value_worksheet = value_workbook[sheet_name]
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        header_map = {header: index + 1 for index, header in enumerate(headers) if header}
        material_col = header_map.get("物料.编码") or header_map.get("物料编码") or header_map.get("料号")
        shortage_col = header_map.get("缺料数量") or header_map.get("缺料")
        plan_col = header_map.get("计划数量")
        picked_col = header_map.get("分拣数量") or header_map.get("配送数量")
        reply_col = header_map.get("最新交期")
        if not material_col:
            raise WorkbookInputError("工单缺料清单缺少“物料.编码/物料编码/料号”列")
        if not reply_col:
            reply_col = len(headers) + 1
            worksheet.cell(1, reply_col).value = "最新交期"
            emit_progress(progress_callback, "步骤 4/5：工单缺料清单缺少“最新交期”列，已自动新增")
        if not shortage_col and not (plan_col and picked_col):
            raise WorkbookInputError("工单缺料清单缺少“缺料数量”列，且无法用计划数量-分拣数量补算")

        updated_rows = 0
        preserved_rows = 0
        insufficient_rows = 0
        no_reply_rows = 0
        shortage_rows = 0
        affected_materials: set[str] = set()

        for row_index in range(2, worksheet.max_row + 1):
            if row_index > 2 and row_index % 100 == 0:
                emit_progress(progress_callback, f"步骤 4/5：已处理工单缺料清单 {row_index - 1} 行...")
            material_code = extract_material_code(worksheet.cell(row_index, material_col).value)
            if not material_code:
                continue
            shortage_value = value_worksheet.cell(row_index, shortage_col).value if shortage_col else None
            shortage_qty = coerce_reply_quantity(shortage_value)
            if shortage_qty is None and plan_col and picked_col:
                plan_qty = coerce_reply_quantity(value_worksheet.cell(row_index, plan_col).value) or 0
                picked_qty = coerce_reply_quantity(value_worksheet.cell(row_index, picked_col).value) or 0
                shortage_qty = max(plan_qty - picked_qty, 0)
            shortage_qty = float(shortage_qty or 0)
            if shortage_qty <= FLOAT_ABS_TOL:
                continue

            shortage_rows += 1
            affected_materials.add(material_code)
            schedule = material_schedules.get(material_code, [])
            allocations, remaining_qty = _allocate_supply_quantity(schedule, shortage_qty)
            if not allocations:
                no_reply_rows += 1
            if remaining_qty > FLOAT_ABS_TOL:
                insufficient_rows += 1

            existing_reply = str(worksheet.cell(row_index, reply_col).value or "").strip()
            if preserve_existing and existing_reply:
                preserved_rows += 1
                continue

            reply_text = _format_work_order_reply(allocations, remaining_qty)
            if not reply_text:
                reply_text = f"缺料数量{format_quantity_text(shortage_qty)}"
            worksheet.cell(row_index, reply_col).value = reply_text
            updated_rows += 1

        emit_progress(progress_callback, f"步骤 5/5：保存工单缺料回复结果到 {output_file}...")
        workbook.save(output_file)
    finally:
        workbook.close()
        value_workbook.close()

    return {
        "output_path": str(output_file),
        "sheet_name": sheet_name,
        "shortage_rows": shortage_rows,
        "updated_rows": updated_rows,
        "preserved_rows": preserved_rows,
        "insufficient_rows": insufficient_rows,
        "no_reply_rows": no_reply_rows,
        "affected_material_count": len(affected_materials),
        "reply_material_count": reply_stats["material_count"],
        "reply_cell_count": reply_stats["reply_cell_count"],
        "reply_start_date": str(reply_start),
        "inventory_material_count": inventory_stats["material_count"],
        "inventory_row_count": inventory_stats["row_count"],
        "arrival_material_count": arrival_stats["material_count"],
        "arrival_row_count": arrival_stats["row_count"],
    }


def unique_join(values: Iterable[object], separator: str = "、") -> str:
    ordered: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = "" if value is None or pd.isna(value) else str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
    return separator.join(ordered)


def limited_unique_join(
    values: Iterable[object],
    *,
    limit: int = 5,
    separator: str = "、",
) -> str:
    ordered: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = "" if value is None or pd.isna(value) else str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
        if len(ordered) >= limit:
            break
    return separator.join(ordered)


def analyze_arrival_status_records(
    input_path: Path,
    arrival_record_path: Path,
    *,
    external_bom_df: pd.DataFrame | None = None,
    arrival_plan_path: Path | None = None,
    analysis_date: date | None = None,
    window_days: int = 3,
    progress_callback: ProgressCallback | None = None,
) -> ArrivalStatusResult:
    total_steps = 5 if arrival_plan_path else 4
    emit_progress(progress_callback, f"步骤 1/{total_steps}：读取 MRP 输入数据...")
    frames = read_workbook_tables(
        input_path,
        external_bom_df=external_bom_df,
        progress_callback=progress_callback,
        progress_prefix=f"步骤 1/{total_steps}",
    )

    emit_progress(progress_callback, f"步骤 2/{total_steps}：计算缺料明细并汇总未来需求...")
    shortage_df = build_shortage_df(frames)
    current_inventory_map = aggregate_map(
        frames["期初库存"],
        first_matching_column(frames["期初库存"], ["物料编码"]),
        first_matching_column(frames["期初库存"], ["库存量"], contains=True),
    )

    emit_progress(progress_callback, f"步骤 3/{total_steps}：读取到货记录并匹配收货状态...")
    arrival_path = Path(arrival_record_path).expanduser().resolve()
    if not arrival_path.exists():
        raise WorkbookInputError(f"到货记录文件不存在: {arrival_path}")

    arrival_raw = normalize_columns(pd.read_excel(arrival_path, sheet_name=0))
    code_col = first_matching_column(arrival_raw, ["物料编码"])
    status_col = first_matching_column(arrival_raw, ["收货状态"])
    qty_col = first_matching_column(arrival_raw, ["到货数量"])
    reject_col = first_matching_column(arrival_raw, ["拒收数量"], required=False)
    broken_col = first_matching_column(arrival_raw, ["破坏数量"], required=False)
    name_col = first_matching_column(arrival_raw, ["物料"], required=False)
    supplier_col = first_matching_column(arrival_raw, ["供应商"], required=False)
    warehouse_col = first_matching_column(arrival_raw, ["仓库"], required=False)
    arrival_date_col = first_matching_column(arrival_raw, ["到货日期"], required=False)
    inbound_time_col = first_matching_column(arrival_raw, ["批次入库时间"], required=False)
    note_col = first_matching_column(arrival_raw, ["备注"], required=False)

    arrival_df = pd.DataFrame(
        {
            "收货状态": arrival_raw[status_col].astype(str).str.strip(),
            "物料编码": arrival_raw[code_col].astype(str).str.strip(),
            "物料": arrival_raw[name_col].astype(str).str.strip() if name_col else "",
            "供应商": arrival_raw[supplier_col].astype(str).str.strip() if supplier_col else "",
            "仓库": arrival_raw[warehouse_col].astype(str).str.strip() if warehouse_col else "",
            "到货数量": coerce_number_series(arrival_raw[qty_col]),
            "拒收数量": coerce_number_series(arrival_raw[reject_col]) if reject_col else 0,
            "破坏数量": coerce_number_series(arrival_raw[broken_col]) if broken_col else 0,
            "到货日期": coerce_date_series(arrival_raw[arrival_date_col]) if arrival_date_col else pd.NaT,
            "批次入库时间": pd.to_datetime(arrival_raw[inbound_time_col], errors="coerce")
            if inbound_time_col
            else pd.NaT,
            "备注": arrival_raw[note_col].astype(str).str.strip() if note_col else "",
        }
    )
    arrival_df = arrival_df[arrival_df["物料编码"] != ""].copy()
    arrival_df["可催数量"] = (
        arrival_df["到货数量"] - arrival_df["拒收数量"] - arrival_df["破坏数量"]
    ).clip(lower=0)
    arrival_df = arrival_df[
        (arrival_df["到货数量"].fillna(0) != 0)
        | (arrival_df["可催数量"].fillna(0) != 0)
    ].copy()

    result_columns = [
        "收货状态",
        "物料编码",
        "物料",
        "供应商",
        "仓库",
        "到货数量",
        "可催数量",
        "当前库存",
        "未来3天需求",
        "未来3天缺口",
        "未来总缺口",
        "最早需求日期",
        "是否缺料",
        "影响母料号",
        "影响项目",
        "到货日期",
        "批次入库时间",
        "备注",
        "跟催建议",
    ]

    purchase_pending_columns = [
        "供应商",
        "物料编码",
        "物料",
        "规格",
        "计划到货日期",
        "计划数量",
        "到货数量",
        "未到货数量",
        "外部单据编号",
        "编码",
        "采购订单",
        "计划状态",
        "当前库存",
        "未来3天需求",
        "未来3天缺口",
        "未来总缺口",
        "最早需求日期",
        "是否缺料",
        "备注",
        "跟催建议",
    ]

    today = analysis_date or date.today()
    window_days = max(int(window_days or 0), 1)
    window_end = today + timedelta(days=window_days - 1)
    empty_df = pd.DataFrame(columns=result_columns)
    empty_purchase_pending_df = pd.DataFrame(columns=purchase_pending_columns)

    shortage_work = shortage_df.copy()
    material_code_col = "BOM2.BOM子项.子件料品.料号"
    shortage_work["物料编码"] = shortage_work[material_code_col].astype(str).str.strip()
    shortage_work["需求"] = coerce_number_series(shortage_work["需求"])
    shortage_work["到货数量"] = coerce_number_series(shortage_work["到货数量"])

    near_mask = shortage_work["到货日期"].apply(
        lambda value: isinstance(value, date) and today <= value <= window_end
    )
    future_mask = shortage_work["到货日期"].apply(
        lambda value: isinstance(value, date) and value >= today
    )

    near_summary = (
        shortage_work.loc[near_mask]
        .groupby("物料编码", sort=True)
        .agg(
            未来3天需求=("需求", "sum"),
            未来3天缺口=("到货数量", "sum"),
            最早需求日期=("到货日期", "min"),
            影响母料号=("母件料号", lambda series: limited_unique_join(series, limit=6)),
            影响项目=("客户", lambda series: limited_unique_join(series, limit=6)),
        )
        .reset_index()
    )
    future_summary = (
        shortage_work.loc[future_mask]
        .groupby("物料编码", sort=True)
        .agg(未来总缺口=("到货数量", "sum"))
        .reset_index()
    )

    def build_follow_up_note(row: pd.Series) -> str:
        status = str(row.get("收货状态", "")).strip()
        future_gap = float(row.get("未来3天缺口", 0) or 0)
        future_demand = float(row.get("未来3天需求", 0) or 0)
        total_gap = float(row.get("未来总缺口", 0) or 0)
        if status == "在检":
            if future_gap > 0:
                return "优先催检验，未来3天已影响缺料"
            if future_demand > 0:
                return "建议优先催检验，未来3天有需求"
            if total_gap > 0:
                return "建议跟进检验，存在后续缺口"
            return "暂无近3天风险"
        if status in {"待入库", "待收货"}:
            if future_gap > 0:
                return "优先催入库/收货，未来3天已影响缺料"
            if future_demand > 0:
                return "建议优先入库/收货，未来3天有需求"
            if total_gap > 0:
                return "建议跟进入库/收货，存在后续缺口"
            return "暂无近3天风险"
        if future_gap > 0:
            return "建议优先处理，未来3天已影响缺料"
        if future_demand > 0:
            return "建议优先处理，未来3天有需求"
        if total_gap > 0:
            return "建议关注后续缺口"
        return "暂无近3天风险"

    def enrich_demand_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
        if dataframe.empty:
            return dataframe.copy()
        result = dataframe.merge(near_summary, on="物料编码", how="left")
        result = result.merge(future_summary, on="物料编码", how="left")
        result["当前库存"] = result["物料编码"].map(current_inventory_map).fillna(0)
        for column_name in ("当前库存", "未来3天需求", "未来3天缺口", "未来总缺口"):
            result[column_name] = result[column_name].fillna(0)
        for column_name in ("影响母料号", "影响项目"):
            if column_name in result.columns:
                result[column_name] = result[column_name].fillna("")
        result["是否缺料"] = result["未来总缺口"].apply(
            lambda value: "是" if float(value) > 0 else "否"
        )
        result["_priority"] = result.apply(
            lambda row: 0
            if float(row.get("未来3天缺口", 0) or 0) > 0
            else 1
            if float(row.get("未来3天需求", 0) or 0) > 0
            else 2
            if float(row.get("未来总缺口", 0) or 0) > 0
            else 3,
            axis=1,
        )
        return result

    if arrival_df.empty:
        grouped_arrival = pd.DataFrame(columns=[*result_columns, "_priority"])
    else:
        grouped_arrival = (
            arrival_df.groupby(["收货状态", "物料编码"], sort=True)
            .agg(
                物料=("物料", "first"),
                供应商=("供应商", unique_join),
                仓库=("仓库", unique_join),
                到货数量=("到货数量", "sum"),
                可催数量=("可催数量", "sum"),
                到货日期=("到货日期", "min"),
                批次入库时间=("批次入库时间", "min"),
                备注=("备注", unique_join),
            )
            .reset_index()
        )
        grouped_arrival = enrich_demand_columns(grouped_arrival)
        grouped_arrival["跟催建议"] = grouped_arrival.apply(build_follow_up_note, axis=1)
        grouped_arrival = grouped_arrival.sort_values(
            by=["_priority", "未来3天缺口", "未来3天需求", "未来总缺口", "可催数量", "物料编码"],
            ascending=[True, False, False, False, False, True],
            kind="stable",
        ).reset_index(drop=True)

    all_df = grouped_arrival[result_columns].copy() if not grouped_arrival.empty else empty_df.copy()
    quality_df = all_df[all_df["收货状态"] == "在检"].reset_index(drop=True)
    warehouse_df = all_df[all_df["收货状态"].isin({"待入库", "待收货"})].reset_index(drop=True)

    purchase_pending_df = empty_purchase_pending_df.copy()
    if arrival_plan_path:
        emit_progress(progress_callback, f"步骤 4/{total_steps}：读取到货计划并筛选采购未到货...")
        plan_path = Path(arrival_plan_path).expanduser().resolve()
        if not plan_path.exists():
            raise WorkbookInputError(f"到货计划文件不存在: {plan_path}")

        plan_raw = normalize_columns(pd.read_excel(plan_path, sheet_name=0))
        plan_code_col = first_matching_column(plan_raw, ["物料编码"])
        plan_date_col = first_matching_column(plan_raw, ["到货日期"])
        plan_qty_col = first_matching_column(plan_raw, ["计划数量"])
        plan_arrived_qty_col = first_matching_column(plan_raw, ["到货数量"], required=False)
        plan_supplier_col = first_matching_column(plan_raw, ["供应商"], required=False)
        plan_name_col = first_matching_column(plan_raw, ["物料"], required=False)
        plan_spec_col = first_matching_column(plan_raw, ["规格"], required=False)
        plan_external_doc_col = first_matching_column(plan_raw, ["外部单据编号"], required=False)
        plan_doc_col = first_matching_column(plan_raw, ["编码"], required=False)
        plan_po_col = first_matching_column(plan_raw, ["采购订单"], required=False)
        plan_status_col = first_matching_column(plan_raw, ["计划状态"], required=False)
        plan_note_col = first_matching_column(plan_raw, ["备注"], required=False)

        purchase_plan_df = pd.DataFrame(
            {
                "供应商": plan_raw[plan_supplier_col].astype(str).str.strip() if plan_supplier_col else "",
                "物料编码": plan_raw[plan_code_col].astype(str).str.strip(),
                "物料": plan_raw[plan_name_col].astype(str).str.strip() if plan_name_col else "",
                "规格": plan_raw[plan_spec_col].astype(str).str.strip() if plan_spec_col else "",
                "计划到货日期": coerce_date_series(plan_raw[plan_date_col]),
                "计划数量": coerce_number_series(plan_raw[plan_qty_col]),
                "到货数量": coerce_number_series(plan_raw[plan_arrived_qty_col]) if plan_arrived_qty_col else 0,
                "外部单据编号": plan_raw[plan_external_doc_col].astype(str).str.strip() if plan_external_doc_col else "",
                "编码": plan_raw[plan_doc_col].astype(str).str.strip() if plan_doc_col else "",
                "采购订单": plan_raw[plan_po_col].astype(str).str.strip() if plan_po_col else "",
                "计划状态": plan_raw[plan_status_col].astype(str).str.strip() if plan_status_col else "",
                "备注": plan_raw[plan_note_col].astype(str).str.strip() if plan_note_col else "",
            }
        )
        purchase_plan_df = purchase_plan_df[
            (purchase_plan_df["物料编码"] != "")
            & (purchase_plan_df["计划到货日期"] == today)
            & (purchase_plan_df["计划数量"] > 0)
            & (purchase_plan_df["到货数量"].fillna(0) <= 0)
        ].copy()
        if not purchase_plan_df.empty:
            purchase_plan_df["未到货数量"] = (
                purchase_plan_df["计划数量"] - purchase_plan_df["到货数量"].fillna(0)
            ).clip(lower=0)
            purchase_plan_df = enrich_demand_columns(purchase_plan_df)

            def build_purchase_pending_note(row: pd.Series) -> str:
                future_gap = float(row.get("未来3天缺口", 0) or 0)
                future_demand = float(row.get("未来3天需求", 0) or 0)
                total_gap = float(row.get("未来总缺口", 0) or 0)
                if future_gap > 0:
                    return "今日计划未到，已影响未来3天缺料，优先催采购到货"
                if future_demand > 0:
                    return "今日计划未到，未来3天有需求，建议催采购确认"
                if total_gap > 0:
                    return "今日计划未到，存在后续缺口，建议跟进"
                return "今日计划未到，建议确认供应商到货"

            purchase_plan_df["跟催建议"] = purchase_plan_df.apply(build_purchase_pending_note, axis=1)
            purchase_plan_df = purchase_plan_df.sort_values(
                by=["_priority", "未来3天缺口", "未来3天需求", "未来总缺口", "未到货数量", "物料编码"],
                ascending=[True, False, False, False, False, True],
                kind="stable",
            ).reset_index(drop=True)
            purchase_pending_df = purchase_plan_df[purchase_pending_columns].copy()

    urgent_count = int((grouped_arrival["_priority"] <= 1).sum()) if not grouped_arrival.empty else 0
    emit_progress(progress_callback, f"步骤 {total_steps}/{total_steps}：到货跟催分析完成")
    return ArrivalStatusResult(
        quality_df=quality_df,
        warehouse_df=warehouse_df,
        all_df=all_df,
        purchase_pending_df=purchase_pending_df,
        summary={
            "window_days": window_days,
            "window_label": f"{today} ~ {window_end}",
            "analysis_date": today,
            "quality_count": len(quality_df),
            "warehouse_count": len(warehouse_df),
            "all_count": len(all_df),
            "purchase_pending_count": len(purchase_pending_df),
            "urgent_count": urgent_count,
        },
    )


def analyze_external_purchase_readiness(
    input_path: Path,
    root_items: Iterable[tuple[str, float]],
    *,
    external_bom_df: pd.DataFrame | None = None,
    carry_forward_paths: Iterable[Path] | None = None,
    reply_cutoff_date: date | None = None,
    progress_callback: ProgressCallback | None = None,
) -> PurchaseReadinessResult:
    emit_progress(progress_callback, "步骤 1/4：读取输入数据...")
    frames = read_workbook_tables(input_path, external_bom_df=external_bom_df)
    substitute_rules = read_substitute_rules(input_path)

    normalized_roots: list[tuple[str, float]] = []
    root_order: list[str] = []
    root_qty_map: dict[str, float] = {}
    inferred_quantity_roots: set[str] = set()
    for raw_code, raw_qty in root_items:
        material_code = str(raw_code or "").strip()
        try:
            quantity = float(raw_qty)
        except (TypeError, ValueError):
            quantity = 0
        infer_quantity = quantity < 0
        if infer_quantity:
            quantity = 1.0
        if not material_code or quantity <= 0:
            continue
        if material_code not in root_qty_map:
            root_order.append(material_code)
            root_qty_map[material_code] = 0.0
        if infer_quantity:
            inferred_quantity_roots.add(material_code)
        root_qty_map[material_code] += quantity
    normalized_roots = [(material_code, root_qty_map[material_code]) for material_code in root_order]
    if not normalized_roots:
        raise WorkbookInputError("未提供有效的母料号和数量")

    purchase = frames["采购数据"].copy()
    purchase_item_col = first_matching_column(purchase, ["物料号"])
    supplier_col = first_matching_column(purchase, ["供应商"], required=False)
    buyer_col = first_matching_column(purchase, ["采购"], required=False)
    purchase_name_col = first_matching_column(purchase, ["名称"], required=False)
    purchase_spec_col = first_matching_column(purchase, ["规格"], required=False)
    purchase_info = purchase.assign(
        料号=purchase[purchase_item_col].astype(str).str.strip(),
        供应商=purchase[supplier_col] if supplier_col else "",
        采购=purchase[buyer_col] if buyer_col else "",
        料品名称=purchase[purchase_name_col] if purchase_name_col else "",
        规格=purchase[purchase_spec_col] if purchase_spec_col else "",
    )
    purchase_info = purchase_info[purchase_info["料号"] != ""].drop_duplicates(subset=["料号"], keep="first")
    purchased_codes = set(purchase_info["料号"].astype(str).str.strip())
    purchase_info_map = {
        row.料号: {
            "供应商": "" if pd.isna(row.供应商) else str(row.供应商).strip(),
            "采购": "" if pd.isna(row.采购) else str(row.采购).strip(),
            "料品名称": "" if pd.isna(row.料品名称) else str(row.料品名称).strip(),
            "规格": "" if pd.isna(row.规格) else str(row.规格).strip(),
        }
        for row in purchase_info.itertuples(index=False)
    }

    inventory_map = aggregate_map(
        frames["期初库存"],
        first_matching_column(frames["期初库存"], ["物料编码"]),
        first_matching_column(frames["期初库存"], ["库存量"]),
    )
    inbound_po_map = aggregate_map(
        frames["在途采购"],
        first_matching_column(frames["在途采购"], ["料号"]),
        first_matching_column(frames["在途采购"], ["欠交数量"], contains=True),
    )
    inbound_pr_map = aggregate_map(
        frames["在途请购"],
        first_matching_column(frames["在途请购"], ["料号"]),
        first_matching_column(frames["在途请购"], ["未转PO数量"], contains=True),
    )

    bom = frames["BOM"].copy()
    bom_parent_col = first_matching_column(bom, ["母件料号"])
    bom_parent_name_col = first_matching_column(bom, ["母件品名"], required=False)
    bom_parent_spec_col = first_matching_column(bom, ["母件规格"], required=False)
    child_item_col = first_matching_column(bom, ["BOM子项.子件料品.料号"])
    child_name_col = first_matching_column(bom, ["BOM子项.子件料品.料品名称"], required=False)
    child_spec_col = first_matching_column(bom, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom, ["BOM子项.子件用量"])

    bom_graph = pd.DataFrame(
        {
            "parent_code": bom[bom_parent_col].astype(str).str.strip(),
            "parent_name": bom[bom_parent_name_col] if bom_parent_name_col else "",
            "parent_spec": bom[bom_parent_spec_col] if bom_parent_spec_col else "",
            "child_code": bom[child_item_col].astype(str).str.strip(),
            "child_name": bom[child_name_col] if child_name_col else "",
            "child_spec": bom[child_spec_col] if child_spec_col else "",
            "usage": coerce_number_series(bom[child_usage_col]),
        }
    )
    bom_graph = bom_graph[
        (bom_graph["parent_code"] != "") & (bom_graph["child_code"] != "") & (bom_graph["usage"] != 0)
    ].copy()

    edges: dict[str, list[dict[str, object]]] = defaultdict(list)
    material_name_map: dict[str, str] = {}
    material_spec_map: dict[str, str] = {}
    for row in bom_graph.itertuples(index=False):
        parent_code = str(row.parent_code).strip()
        child_code = str(row.child_code).strip()
        parent_name = "" if pd.isna(row.parent_name) else str(row.parent_name).strip()
        parent_spec = "" if pd.isna(row.parent_spec) else str(row.parent_spec).strip()
        child_name = "" if pd.isna(row.child_name) else str(row.child_name).strip()
        child_spec = "" if pd.isna(row.child_spec) else str(row.child_spec).strip()
        usage = float(row.usage)
        if not parent_code or not child_code or usage == 0:
            continue
        edges[parent_code].append(
            {
                "child_code": child_code,
                "child_name": child_name,
                "child_spec": child_spec,
                "usage": usage,
            }
        )
        if parent_name:
            material_name_map[parent_code] = parent_name
        if parent_spec:
            material_spec_map[parent_code] = parent_spec
        if child_name and child_code not in material_name_map:
            material_name_map[child_code] = child_name
        if child_spec and child_code not in material_spec_map:
            material_spec_map[child_code] = child_spec

    for material_code, info in purchase_info_map.items():
        if info["料品名称"] and material_code not in material_name_map:
            material_name_map[material_code] = info["料品名称"]
        if info["规格"] and material_code not in material_spec_map:
            material_spec_map[material_code] = info["规格"]

    substitute_code_map: dict[str, list[str]] = {}
    if substitute_rules is not None and not substitute_rules.empty:
        for before_code, group_df in substitute_rules.groupby("before_code", sort=False):
            matched_codes: list[str] = []
            for after_code in group_df.sort_values("sort_order")["after_code"].astype(str).str.strip():
                if not after_code or after_code in matched_codes:
                    continue
                matched_codes.append(after_code)
            if matched_codes:
                substitute_code_map[str(before_code).strip()] = matched_codes

    carry_stats = {"file_count": 0, "cell_count": 0, "material_count": 0}
    reply_numeric_map: dict[str, dict[date, float]] = defaultdict(lambda: defaultdict(float))
    reply_text_map: dict[str, list[str]] = defaultdict(list)
    if carry_forward_paths:
        emit_progress(progress_callback, "步骤 2/4：读取旧平衡表采购答复...")
        reply_overrides, carry_stats = load_carried_purchase_replies(
            carry_forward_paths,
            cutoff_date=reply_cutoff_date or (date.today() - timedelta(days=1)),
        )
        for (material_code, header_date), reply_value in reply_overrides.items():
            numeric_value = coerce_reply_quantity(reply_value)
            if numeric_value is None:
                text_value = str(reply_value).strip()
                if text_value:
                    reply_text_map[material_code].append(text_value)
                continue
            reply_numeric_map[material_code][header_date] += numeric_value
    else:
        emit_progress(progress_callback, "步骤 2/4：未选择旧平衡表，按现有库存与在途分析")

    emit_progress(progress_callback, "步骤 3/4：展开母料号并识别外购物料...")
    demand_rows: list[dict[str, object]] = []
    missing_root_rows: list[dict[str, object]] = []
    unknown_leaf_rows: list[dict[str, object]] = []

    def is_purchased_or_has_substitute(material_code: str) -> bool:
        material_code = str(material_code or "").strip()
        return material_code in purchased_codes or bool(substitute_code_map.get(material_code))

    def add_purchase_demand_row(
        *,
        material_code: str,
        required_qty: float,
        root_code: str,
        root_qty: float,
        parent_code: str,
        material_name: str = "",
        material_spec: str = "",
        level: int,
    ) -> None:
        purchase_meta = purchase_info_map.get(material_code, {})
        demand_rows.append(
            {
                "母料号": root_code,
                "母件输入数量": root_qty,
                "上层物料编码": parent_code,
                "上层物料名称": material_name_map.get(parent_code, ""),
                "料号": material_code,
                "料品名称": material_name or material_name_map.get(material_code, purchase_meta.get("料品名称", "")),
                "规格": material_spec or material_spec_map.get(material_code, purchase_meta.get("规格", "")),
                "需求数量": required_qty,
                "层级": level,
            }
        )

    def walk_to_purchased(
        current_code: str,
        required_qty: float,
        root_code: str,
        root_qty: float,
        level: int,
        trail: set[str],
    ) -> None:
        children = edges.get(current_code, [])
        if not children:
            if is_purchased_or_has_substitute(current_code):
                add_purchase_demand_row(
                    material_code=current_code,
                    required_qty=required_qty,
                    root_code=root_code,
                    root_qty=root_qty,
                    parent_code=root_code,
                    level=level,
                )
            else:
                unknown_leaf_rows.append(
                    {
                        "母料号": root_code,
                        "母件输入数量": root_qty,
                        "上层物料编码": root_code,
                        "上层物料名称": material_name_map.get(root_code, ""),
                        "料号": current_code,
                        "料品名称": material_name_map.get(current_code, ""),
                        "规格": material_spec_map.get(current_code, ""),
                        "需求数量": required_qty,
                        "问题原因": "物料未在采购数据中识别，且无下阶BOM",
                    }
                )
            return

        for child in children:
            child_code = str(child["child_code"]).strip()
            child_qty = required_qty * float(child["usage"])
            child_name = str(child.get("child_name", "") or "").strip()
            child_spec = str(child.get("child_spec", "") or "").strip()
            if not child_code or child_qty == 0:
                continue
            if child_code in purchased_codes:
                add_purchase_demand_row(
                    material_code=child_code,
                    required_qty=child_qty,
                    root_code=root_code,
                    root_qty=root_qty,
                    parent_code=current_code,
                    material_name=child_name,
                    material_spec=child_spec,
                    level=level + 1,
                )
                continue
            if child_code in trail:
                unknown_leaf_rows.append(
                    {
                        "母料号": root_code,
                        "母件输入数量": root_qty,
                        "上层物料编码": current_code,
                        "上层物料名称": material_name_map.get(current_code, ""),
                        "料号": child_code,
                        "料品名称": child_name,
                        "规格": child_spec,
                        "需求数量": child_qty,
                        "问题原因": "BOM存在循环引用，无法继续展开",
                    }
                )
                continue
            if child_code in edges:
                walk_to_purchased(
                    child_code,
                    child_qty,
                    root_code,
                    root_qty,
                    level + 1,
                    trail | {child_code},
                )
                continue
            if substitute_code_map.get(child_code):
                add_purchase_demand_row(
                    material_code=child_code,
                    required_qty=child_qty,
                    root_code=root_code,
                    root_qty=root_qty,
                    parent_code=current_code,
                    material_name=child_name,
                    material_spec=child_spec,
                    level=level + 1,
                )
                continue
            unknown_leaf_rows.append(
                {
                    "母料号": root_code,
                    "母件输入数量": root_qty,
                    "上层物料编码": current_code,
                    "上层物料名称": material_name_map.get(current_code, ""),
                    "料号": child_code,
                    "料品名称": child_name,
                    "规格": child_spec,
                    "需求数量": child_qty,
                    "问题原因": "物料未在采购数据中识别，且无下阶BOM",
                }
            )

    for root_code, root_qty in normalized_roots:
        if is_purchased_or_has_substitute(root_code) and root_code not in edges:
            demand_rows.append(
                {
                    "母料号": root_code,
                    "母件输入数量": root_qty,
                    "上层物料编码": root_code,
                    "上层物料名称": material_name_map.get(root_code, ""),
                    "料号": root_code,
                    "料品名称": material_name_map.get(root_code, purchase_info_map.get(root_code, {}).get("料品名称", "")),
                    "规格": material_spec_map.get(root_code, purchase_info_map.get(root_code, {}).get("规格", "")),
                    "需求数量": root_qty,
                    "层级": 0,
                }
            )
            continue
        if root_code not in edges:
            missing_root_rows.append(
                {
                    "母料号": root_code,
                    "输入数量": root_qty,
                    "料品名称": material_name_map.get(root_code, ""),
                    "问题原因": "母料号未在当前BOM中识别",
                }
            )
            continue
        walk_to_purchased(root_code, root_qty, root_code, root_qty, 0, {root_code})

    def build_unrecognized_material_rows(
        unknown_df: pd.DataFrame,
        missing_df: pd.DataFrame | None = None,
    ) -> list[dict[str, object]]:
        display_rows: list[dict[str, object]] = []

        def supply_snapshot(material_code: str) -> dict[str, object]:
            material_code = str(material_code or "").strip()
            substitutes = substitute_code_map.get(material_code, [])
            real_inventory = float(inventory_map.get(material_code, 0) or 0)
            inbound_po = float(inbound_po_map.get(material_code, 0) or 0)
            inbound_pr = float(inbound_pr_map.get(material_code, 0) or 0)
            substitute_stock = sum(float(inventory_map.get(code, 0) or 0) for code in substitutes)
            substitute_open_po = sum(float(inbound_po_map.get(code, 0) or 0) for code in substitutes)
            substitute_unconverted = sum(float(inbound_pr_map.get(code, 0) or 0) for code in substitutes)
            stock_pool = real_inventory + substitute_stock
            available_pool = stock_pool + inbound_po + inbound_pr + substitute_open_po + substitute_unconverted
            return {
                "当前库存": real_inventory,
                "实时库存": real_inventory,
                "未清PO": inbound_po,
                "未转": inbound_pr,
                "替代料库存": substitute_stock,
                "替代料未清PO": substitute_open_po,
                "替代料未转": substitute_unconverted,
                "当前库存+替代库存": stock_pool,
                "当前可用": available_pool,
                "替代料清单": _summarize_codes_v2(substitutes, limit=20),
            }

        if unknown_df is not None and not unknown_df.empty:
            for record in unknown_df.to_dict("records"):
                material_code = str(record.get("料号", "") or "").strip()
                demand_qty = float(record.get("需求数量", 0) or 0)
                snapshot = supply_snapshot(material_code)
                display_rows.append(
                    {
                        "料号": material_code,
                        "料品名称": record.get("料品名称", ""),
                        "规格": record.get("规格", ""),
                        "供应商": "",
                        "采购": "",
                        "总需求": demand_qty,
                        **snapshot,
                        "当前缺口": max(demand_qty - float(snapshot["当前库存+替代库存"]), 0),
                        "当前可用缺口": max(demand_qty - float(snapshot["当前可用"]), 0),
                        "采购答复累计": 0,
                        "最早齐套日期": None,
                        "状态": "未识别",
                        "问题原因": record.get("问题原因", "物料未识别"),
                        "来源母料号": record.get("母料号", ""),
                        "上层物料编码": record.get("上层物料编码", ""),
                    }
                )

        if missing_df is not None and not missing_df.empty:
            for record in missing_df.to_dict("records"):
                material_code = str(record.get("母料号", "") or "").strip()
                demand_qty = float(record.get("输入数量", 0) or 0)
                snapshot = supply_snapshot(material_code)
                display_rows.append(
                    {
                        "料号": material_code,
                        "料品名称": record.get("料品名称", ""),
                        "规格": material_spec_map.get(material_code, ""),
                        "供应商": "",
                        "采购": "",
                        "总需求": demand_qty,
                        **snapshot,
                        "当前缺口": max(demand_qty - float(snapshot["当前库存+替代库存"]), 0),
                        "当前可用缺口": max(demand_qty - float(snapshot["当前可用"]), 0),
                        "采购答复累计": 0,
                        "最早齐套日期": None,
                        "状态": "母料号未识别",
                        "问题原因": record.get("问题原因", "母料号未识别"),
                        "来源母料号": material_code,
                        "上层物料编码": "",
                    }
                )

        return display_rows

    demand_df = pd.DataFrame(demand_rows)
    if demand_df.empty:
        material_df = pd.DataFrame(
            columns=[
                "料号",
                "料品名称",
                "规格",
                "供应商",
                "采购",
                "总需求",
                "当前库存",
                "未清PO",
                "未转",
                "替代料库存",
                "替代料未清PO",
                "替代料未转",
                "当前库存+替代库存",
                "替代料清单",
                "实时库存",
                "替代1",
                "替代1库存",
                "替代2",
                "替代2库存",
                "当前可用",
                "当前缺口",
                "当前可用缺口",
                "采购答复累计",
                "最早齐套日期",
                "状态",
                "问题原因",
                "来源母料号",
                "上层物料编码",
            ]
        )
        issue_df = material_df.copy()
        producible_df = pd.DataFrame(
            columns=[
                "母料号",
                "料品名称",
                "输入数量",
                "排产顺序",
                "排产日期",
                "日期",
                "可生产数量",
                "当前库存可生成",
                "距离输入缺口",
                "瓶颈物料",
                "瓶颈物料名称",
                "瓶颈单台用量",
                "瓶颈可用量",
                "瓶颈原因",
            ]
        )
        horizontal_shortage_df = pd.DataFrame(
            columns=[
                "排产顺序",
                "排产日期",
                "母料号",
                "料品名称",
                "输入数量",
                "料号",
                "物料名称",
                "单台用量",
                "本行需求",
                "库存分配前",
                "库存扣减",
                "库存缺口",
                "库存分配后",
                "当前可用分配前",
                "当前可用扣减",
                "当前可用缺口",
                "当前可用分配后",
                "替代料清单",
                "BOM差异标识",
                "共用母料数",
                "BOM差异说明",
            ]
        )
        diff_capacity_df = pd.DataFrame(
            columns=[
                "排产顺序",
                "排产日期",
                "母料号",
                "料品名称",
                "输入数量",
                "差异物料数",
                "通用物料数",
                "通用物料清单",
                "当前库存差异可生产",
                "当前库存差异缺口",
                "当前可用差异可生产",
                "当前可用差异缺口",
                "瓶颈差异物料",
                "瓶颈物料名称",
                "瓶颈单台差异用量",
                "瓶颈库存可用",
                "瓶颈当前可用",
                "BOM差异说明",
            ]
        )
        rolling_matrix_df = pd.DataFrame()
        root_summary_df = pd.DataFrame(
            [
                {
                    "母料号": root_code,
                    "料品名称": material_name_map.get(root_code, ""),
                    "输入数量": root_qty,
                    "外购物料数": 0,
                    "问题物料数": 0,
                    "未识别物料数": 0,
                    "外购齐套日期": None,
                    "结论": "母料号未识别" if any(row["母料号"] == root_code for row in missing_root_rows) else "未识别到外购物料",
                }
                for root_code, root_qty in normalized_roots
            ]
        )
        missing_root_df = pd.DataFrame(missing_root_rows)
        unknown_leaf_df = pd.DataFrame(unknown_leaf_rows)
        if not unknown_leaf_df.empty:
            unknown_leaf_df = (
                unknown_leaf_df.groupby(["母料号", "母件输入数量", "料号"], sort=True)
                .agg(
                    料品名称=("料品名称", "first"),
                    规格=("规格", "first"),
                    上层物料编码=("上层物料编码", unique_join),
                    上层物料名称=("上层物料名称", unique_join),
                    需求数量=("需求数量", "sum"),
                    问题原因=("问题原因", "first"),
                )
                .reset_index()
            )
        else:
            unknown_leaf_df = pd.DataFrame(
                columns=["母料号", "母件输入数量", "料号", "料品名称", "规格", "上层物料编码", "上层物料名称", "需求数量", "问题原因"]
            )
        if missing_root_df.empty:
            missing_root_df = pd.DataFrame(columns=["母料号", "输入数量", "料品名称", "问题原因"])
        if not root_summary_df.empty:
            unknown_counts = (
                unknown_leaf_df.groupby("母料号")["料号"].nunique().to_dict()
                if not unknown_leaf_df.empty
                else {}
            )
            missing_roots = set(missing_root_df["母料号"].astype(str).str.strip()) if not missing_root_df.empty else set()
            for idx, row in root_summary_df.iterrows():
                root_code_text = str(row.get("母料号", "") or "").strip()
                root_summary_df.at[idx, "未识别物料数"] = int(unknown_counts.get(root_code_text, 0) or 0)
                if root_code_text in missing_roots:
                    root_summary_df.at[idx, "结论"] = "母料号未识别"
                elif int(unknown_counts.get(root_code_text, 0) or 0) > 0:
                    root_summary_df.at[idx, "结论"] = "存在未识别外购物料"
        unrecognized_material_rows = build_unrecognized_material_rows(unknown_leaf_df, missing_root_df)
        if unrecognized_material_rows:
            material_df = pd.DataFrame(unrecognized_material_rows)
        batch_summary = {
            "root_count": len(normalized_roots),
            "material_count": int(material_df["料号"].astype(str).str.strip().nunique()) if not material_df.empty else 0,
            "issue_count": int((material_df["问题原因"].fillna("") != "").sum()) if not material_df.empty else 0,
            "unknown_count": int(unknown_leaf_df["料号"].astype(str).str.strip().nunique()) if not unknown_leaf_df.empty else 0,
            "missing_root_count": len(missing_root_df),
            "ready_date": None,
            "conclusion": "未识别到外购物料",
        }
        emit_progress(progress_callback, "步骤 4/4：分析完成")
        return PurchaseReadinessResult(
            root_summary_df=root_summary_df,
            issue_df=issue_df,
            material_df=material_df,
            producible_df=producible_df,
            horizontal_shortage_df=horizontal_shortage_df,
            diff_capacity_df=diff_capacity_df,
            rolling_matrix_df=rolling_matrix_df,
            missing_root_df=missing_root_df,
            unknown_leaf_df=unknown_leaf_df,
            batch_summary=batch_summary,
            carried_reply_cell_count=carry_stats["cell_count"],
            carried_reply_material_count=carry_stats["material_count"],
            carried_reply_file_count=carry_stats["file_count"],
        )

    demand_df["料号"] = demand_df["料号"].astype(str).str.strip()
    root_material_df = (
        demand_df.groupby(["母料号", "母件输入数量", "料号"], sort=True)
        .agg(
            本次需求=("需求数量", "sum"),
            料品名称=("料品名称", "first"),
            规格=("规格", "first"),
            上层物料编码=("上层物料编码", unique_join),
            上层物料名称=("上层物料名称", unique_join),
        )
        .reset_index()
    )

    material_df = (
        demand_df.groupby("料号", sort=True)
        .agg(
            总需求=("需求数量", "sum"),
            料品名称=("料品名称", "first"),
            规格=("规格", "first"),
            来源母料号=("母料号", unique_join),
            上层物料编码=("上层物料编码", unique_join),
        )
        .reset_index()
    )

    assessment_rows: list[dict[str, object]] = []
    today = date.today()
    for record in material_df.to_dict("records"):
        material_code = str(record["料号"]).strip()
        substitutes = substitute_code_map.get(material_code, [])
        substitute_summary = _summarize_codes_v2(substitutes, limit=20)
        substitute_one = substitutes[0] if len(substitutes) >= 1 else ""
        substitute_two = substitutes[1] if len(substitutes) >= 2 else ""

        def substitute_stock_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inventory_map.get(substitute_code, 0) or 0)

        def substitute_open_po_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inbound_po_map.get(substitute_code, 0) or 0)

        def substitute_unconverted_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inbound_pr_map.get(substitute_code, 0) or 0)

        def substitute_available_qty(substitute_code: str) -> float:
            return (
                substitute_stock_qty(substitute_code)
                + substitute_open_po_qty(substitute_code)
                + substitute_unconverted_qty(substitute_code)
            )

        substitute_one_inventory = substitute_stock_qty(substitute_one)
        substitute_two_inventory = sum(substitute_stock_qty(code) for code in substitutes[1:])
        substitute_one_available = substitute_available_qty(substitute_one)
        substitute_two_available = sum(substitute_available_qty(code) for code in substitutes[1:])
        substitute_open_po_total = sum(substitute_open_po_qty(code) for code in substitutes)
        substitute_unconverted_total = sum(substitute_unconverted_qty(code) for code in substitutes)
        real_inventory = float(inventory_map.get(material_code, 0))
        substitute_stock_inventory = sum(substitute_stock_qty(substitute_code) for substitute_code in substitutes)
        stock_available_now = real_inventory + substitute_stock_inventory
        inbound_po = float(inbound_po_map.get(material_code, 0))
        inbound_pr = float(inbound_pr_map.get(material_code, 0))
        available_now = real_inventory + inbound_po + inbound_pr + substitute_one_available + substitute_two_available
        demand_qty = float(record["总需求"])

        reply_schedule = sorted(reply_numeric_map.get(material_code, {}).items())
        reply_total = sum(quantity for _reply_date, quantity in reply_schedule)
        earliest_ready_date: date | None = None
        if _fle(demand_qty, available_now):
            earliest_ready_date = today
        else:
            running_qty = available_now
            for reply_date, reply_qty in reply_schedule:
                running_qty += reply_qty
                if _fge(running_qty, demand_qty):
                    earliest_ready_date = reply_date
                    break

        non_numeric_reply = unique_join(reply_text_map.get(material_code, []), separator="；")
        current_gap = max(demand_qty - stock_available_now, 0)
        available_gap = max(demand_qty - available_now, 0)
        final_gap = max(demand_qty - (available_now + reply_total), 0)
        if _fle(demand_qty, available_now):
            status = "现有量满足"
            reason = ""
        elif not reply_schedule:
            if non_numeric_reply:
                status = "采购回复非数量"
                reason = "采购回复存在非数量内容，暂无法判断齐套"
            else:
                status = "无采购回复"
                reason = "当前可用量不足，且没有可累计的采购答复"
        elif earliest_ready_date is None:
            status = "采购回复不足"
            reason = "采购答复累计后仍无法满足需求"
        else:
            status = "依赖采购回复"
            reason = f"依赖采购回复，最早 {earliest_ready_date} 齐套"

        purchase_meta = purchase_info_map.get(material_code, {})
        assessment_rows.append(
            {
                "料号": material_code,
                "料品名称": record["料品名称"] or purchase_meta.get("料品名称", ""),
                "规格": record["规格"] or purchase_meta.get("规格", ""),
                "供应商": purchase_meta.get("供应商", ""),
                "采购": purchase_meta.get("采购", ""),
                "总需求": demand_qty,
                "当前库存": real_inventory,
                "未清PO": inbound_po,
                "未转": inbound_pr,
                "替代料库存": substitute_stock_inventory,
                "替代料未清PO": substitute_open_po_total,
                "替代料未转": substitute_unconverted_total,
                "当前库存+替代库存": stock_available_now,
                "替代料清单": substitute_summary,
                "实时库存": real_inventory,
                "替代1": substitute_one,
                "替代1库存": substitute_one_inventory,
                "替代2": substitute_two,
                "替代2库存": substitute_two_inventory,
                "当前可用": available_now,
                "当前缺口": current_gap,
                "当前可用缺口": available_gap,
                "采购答复累计": reply_total,
                "采购答复说明": non_numeric_reply,
                "最早齐套日期": earliest_ready_date,
                "状态": status,
                "问题原因": reason,
                "最终缺口": final_gap,
                "来源母料号": record["来源母料号"],
                "上层物料编码": record["上层物料编码"],
            }
        )

    material_df = pd.DataFrame(assessment_rows).sort_values(
        by=["状态", "最早齐套日期", "料号"],
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    material_assessment_map = {
        str(record.get("料号", "")).strip(): record
        for record in material_df.to_dict("records")
        if str(record.get("料号", "")).strip()
    }

    production_priority_rows: list[dict[str, object]] = []
    production_orders = frames["生产订单"].copy()
    production_parent_col = first_matching_column(production_orders, ["母件料号"], required=False)
    production_date_col = first_matching_column(production_orders, ["上线日期"], required=False)
    production_qty_col = first_matching_column(production_orders, ["上线数量"], required=False)
    if production_parent_col:
        production_dates = (
            coerce_date_series(production_orders[production_date_col])
            if production_date_col
            else pd.Series([pd.NaT] * len(production_orders))
        )
        production_qty = (
            coerce_number_series(production_orders[production_qty_col])
            if production_qty_col
            else pd.Series([0.0] * len(production_orders))
        )
        order_rows_by_code: dict[str, list[dict[str, object]]] = defaultdict(list)
        for position, row in production_orders.iterrows():
            root_code_text = str(row.get(production_parent_col, "") or "").strip()
            if not root_code_text or root_code_text not in root_qty_map:
                continue
            row_qty = float(production_qty.loc[position] or 0)
            if row_qty <= 0:
                continue
            row_date = production_dates.loc[position]
            if pd.isna(row_date):
                row_date = None
            order_rows_by_code[root_code_text].append(
                {
                    "root_code": root_code_text,
                    "qty": row_qty,
                    "plan_date": row_date,
                    "source_order": int(position) + 1,
                }
            )
        for root_code, root_qty in normalized_roots:
            remaining_qty = float(root_qty or 0)
            matched_orders = sorted(
                order_rows_by_code.get(root_code, []),
                key=lambda item: (
                    item.get("plan_date") or date.max,
                    int(item.get("source_order", 10**9)),
                ),
            )
            for matched in matched_orders:
                if remaining_qty <= FLOAT_ABS_TOL:
                    break
                allocated_qty = min(remaining_qty, float(matched.get("qty", 0) or 0))
                if allocated_qty <= FLOAT_ABS_TOL:
                    continue
                production_priority_rows.append(
                    {
                        "root_code": root_code,
                        "qty": allocated_qty,
                        "plan_date": matched.get("plan_date"),
                        "source_order": matched.get("source_order"),
                        "input_order": root_order.index(root_code) + 1,
                    }
                )
                remaining_qty -= allocated_qty
            if remaining_qty > FLOAT_ABS_TOL:
                production_priority_rows.append(
                    {
                        "root_code": root_code,
                        "qty": remaining_qty,
                        "plan_date": None,
                        "source_order": 10**9,
                        "input_order": root_order.index(root_code) + 1,
                    }
                )
    if not production_priority_rows:
        production_priority_rows = [
            {
                "root_code": root_code,
                "qty": root_qty,
                "plan_date": None,
                "source_order": 10**9,
                "input_order": index,
            }
            for index, (root_code, root_qty) in enumerate(normalized_roots, start=1)
        ]
    production_priority_rows.sort(
        key=lambda item: (
            item.get("plan_date") or date.max,
            int(item.get("source_order", 10**9)),
            int(item.get("input_order", 10**9)),
            str(item.get("root_code", "")),
        )
    )
    for priority, item in enumerate(production_priority_rows, start=1):
        item["priority"] = priority

    root_material_profiles: dict[str, list[dict[str, object]]] = {}
    for root_code, _root_qty in normalized_roots:
        root_code_text = str(root_code).strip()
        profile_rows: list[dict[str, object]] = []
        root_materials = root_material_df[root_material_df["母料号"] == root_code_text].copy()
        for rm_record in root_materials.to_dict("records"):
            material_code = str(rm_record.get("料号", "")).strip()
            root_input_qty = float(rm_record.get("母件输入数量", root_qty_map.get(root_code_text, 0)) or 0)
            demand_qty = float(rm_record.get("本次需求", 0) or 0)
            unit_usage = demand_qty / root_input_qty if root_input_qty else demand_qty
            if not material_code or unit_usage <= 0:
                continue
            assessed = material_assessment_map.get(material_code, {})
            profile_rows.append(
                {
                    "料号": material_code,
                    "料品名称": rm_record.get("料品名称", "") or assessed.get("料品名称", ""),
                    "单台用量": unit_usage,
                }
            )
        root_material_profiles[root_code_text] = profile_rows

    all_relevant_codes = {
        str(profile.get("料号", "")).strip()
        for profiles in root_material_profiles.values()
        for profile in profiles
        if str(profile.get("料号", "")).strip()
    }
    all_timeline_dates = {today}
    for material_code in all_relevant_codes:
        all_timeline_dates.update(reply_numeric_map.get(material_code, {}).keys())
    sorted_timeline_dates = sorted(d for d in all_timeline_dates if isinstance(d, date))

    unknown_leaf_df_for_capacity = pd.DataFrame(unknown_leaf_rows)
    stock_only_pool: dict[str, float] = {}
    for material_code in all_relevant_codes:
        substitutes = substitute_code_map.get(material_code, [])
        substitute_stock = sum(float(inventory_map.get(sub_code, 0) or 0) for sub_code in substitutes)
        stock_only_pool[material_code] = float(inventory_map.get(material_code, 0) or 0) + substitute_stock

    def recommend_qty_from_unique_materials(root_code_text: str) -> tuple[int, bool]:
        profiles = root_material_profiles.get(root_code_text, [])
        unique_profiles: list[dict[str, object]] = []
        fallback_profiles: list[dict[str, object]] = []
        for profile in profiles:
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            if not material_code or unit_usage <= 0:
                continue
            cap = int(math.floor(max(float(stock_only_pool.get(material_code, 0) or 0), 0) / unit_usage))
            item = {**profile, "可生产数量": max(cap, 0)}
            fallback_profiles.append(item)
            appears_in = sum(
                1
                for other_profiles in root_material_profiles.values()
                if any(str(other.get("料号", "")).strip() == material_code for other in other_profiles)
            )
            if appears_in == 1:
                unique_profiles.append(item)
        candidate_profiles = unique_profiles or fallback_profiles
        positive_caps = sorted(
            int(item.get("可生产数量", 0) or 0)
            for item in candidate_profiles
            if int(item.get("可生产数量", 0) or 0) > 0
        )
        return (positive_caps[0] if positive_caps else 0, bool(unique_profiles))

    recommended_qty_by_root: dict[str, int] = {}
    recommended_uses_unique: dict[str, bool] = {}
    for root_code_text in root_order:
        recommended_qty, used_unique = recommend_qty_from_unique_materials(root_code_text)
        recommended_qty_by_root[root_code_text] = recommended_qty
        recommended_uses_unique[root_code_text] = used_unique

    inferred_quantity_notes: dict[str, str] = {}
    if inferred_quantity_roots:
        for root_code_text in list(inferred_quantity_roots):
            inferred_qty = recommended_qty_by_root.get(root_code_text, 0)
            root_qty_map[root_code_text] = float(inferred_qty)
            inferred_quantity_notes[root_code_text] = (
                f"数量空白，按{'独用物料' if recommended_uses_unique.get(root_code_text) else '全量物料'}库存建议排产 {inferred_qty} 套"
            )
        normalized_roots = [(material_code, root_qty_map[material_code]) for material_code in root_order]
        production_priority_rows = [
            item for item in production_priority_rows
            if str(item.get("root_code", "")).strip() not in inferred_quantity_roots
        ]
        for root_code_text in root_order:
            if root_code_text not in inferred_quantity_roots:
                continue
            production_priority_rows.append(
                {
                    "root_code": root_code_text,
                    "qty": root_qty_map.get(root_code_text, 0),
                    "plan_date": None,
                    "source_order": 10**9,
                    "input_order": root_order.index(root_code_text) + 1,
                }
            )
        production_priority_rows.sort(
            key=lambda item: (
                item.get("plan_date") or date.max,
                int(item.get("source_order", 10**9)),
                int(item.get("input_order", 10**9)),
                str(item.get("root_code", "")),
            )
        )
        for priority, item in enumerate(production_priority_rows, start=1):
            item["priority"] = priority

    stock_capacity_by_entry: dict[int, int] = {}
    stock_remaining_pool = dict(stock_only_pool)
    for entry_index, entry in enumerate(production_priority_rows):
        root_code_text = str(entry.get("root_code", "")).strip()
        root_qty = float(entry.get("qty", 0) or 0)
        profiles = root_material_profiles.get(root_code_text, [])
        if not profiles:
            stock_capacity_by_entry[entry_index] = 0
            continue
        material_caps: list[int] = []
        for profile in profiles:
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            available_stock = float(stock_remaining_pool.get(material_code, 0) or 0)
            capacity = int(math.floor(max(available_stock, 0) / unit_usage)) if unit_usage > 0 else 0
            material_caps.append(max(capacity, 0))
        stock_capacity_qty = min(root_qty, min(material_caps)) if material_caps else 0
        stock_capacity_qty = int(math.floor(max(stock_capacity_qty, 0)))
        stock_capacity_by_entry[entry_index] = stock_capacity_qty
        for profile in profiles:
            material_code = str(profile.get("料号", "")).strip()
            stock_remaining_pool[material_code] = (
                float(stock_remaining_pool.get(material_code, 0) or 0)
                - stock_capacity_qty * float(profile.get("单台用量", 0) or 0)
            )

    material_usage_by_root: dict[str, dict[str, float]] = defaultdict(dict)
    for root_code_text, profiles in root_material_profiles.items():
        for profile in profiles:
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            if material_code and unit_usage > 0:
                material_usage_by_root[material_code][root_code_text] = unit_usage

    def horizontal_diff_info(material_code: str) -> tuple[str, int, str]:
        usage_map = material_usage_by_root.get(material_code, {})
        usage_values = list(usage_map.values())
        shared_root_count = len(usage_map)
        root_count = max(len(root_order), 1)
        if not usage_values:
            return "", 0, ""
        max_usage = max(usage_values)
        min_usage = min(usage_values)
        if shared_root_count == 1:
            diff_label = "本排产专用"
        elif shared_root_count < root_count:
            diff_label = "部分共用"
        elif abs(max_usage - min_usage) > FLOAT_ABS_TOL:
            diff_label = "共用但用量不同"
        else:
            diff_label = "全部共用"
        max_roots = [code for code, usage in usage_map.items() if abs(float(usage) - max_usage) <= FLOAT_ABS_TOL]
        min_roots = [code for code, usage in usage_map.items() if abs(float(usage) - min_usage) <= FLOAT_ABS_TOL]
        detail = (
            f"出现{shared_root_count}/{root_count}个母料；"
            f"最大单台{format_project_display_value(max_usage)}({unique_join(max_roots[:3])})；"
            f"最小单台{format_project_display_value(min_usage)}({unique_join(min_roots[:3])})"
        )
        return diff_label, shared_root_count, detail

    def differential_unit_usage(root_code_text: str, material_code: str, unit_usage: float) -> tuple[float, str]:
        usage_map = material_usage_by_root.get(material_code, {})
        usage_values = list(usage_map.values())
        if not usage_values:
            return 0.0, ""
        root_count = max(len(root_order), 1)
        shared_root_count = len(usage_map)
        min_usage = min(usage_values)
        if shared_root_count == 1:
            return max(unit_usage, 0), "本排产专用"
        if shared_root_count < root_count:
            return max(unit_usage, 0), "部分共用"
        diff_unit = max(float(unit_usage or 0) - float(min_usage or 0), 0)
        if diff_unit > FLOAT_ABS_TOL:
            return diff_unit, "共用但用量更高"
        return 0.0, "共用基准用量"

    horizontal_rows: list[dict[str, object]] = []
    horizontal_stock_pool = dict(stock_only_pool)
    horizontal_available_pool = {
        material_code: float(material_assessment_map.get(material_code, {}).get("当前可用", 0) or 0)
        for material_code in all_relevant_codes
    }
    for entry in production_priority_rows:
        root_code_text = str(entry.get("root_code", "")).strip()
        root_qty = float(entry.get("qty", 0) or 0)
        for profile in root_material_profiles.get(root_code_text, []):
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            if not material_code or unit_usage <= 0:
                continue
            required_qty = root_qty * unit_usage
            stock_before = float(horizontal_stock_pool.get(material_code, 0) or 0)
            available_before = float(horizontal_available_pool.get(material_code, 0) or 0)
            stock_deduct = min(max(stock_before, 0), required_qty)
            available_deduct = min(max(available_before, 0), required_qty)
            stock_after = stock_before - required_qty
            available_after = available_before - required_qty
            assessed = material_assessment_map.get(material_code, {})
            diff_label, shared_root_count, diff_detail = horizontal_diff_info(material_code)
            horizontal_rows.append(
                {
                    "排产顺序": entry.get("priority", ""),
                    "排产日期": entry.get("plan_date") or "",
                    "母料号": root_code_text,
                    "料品名称": material_name_map.get(root_code_text, ""),
                    "输入数量": root_qty,
                    "料号": material_code,
                    "物料名称": profile.get("料品名称", "") or assessed.get("料品名称", ""),
                    "单台用量": unit_usage,
                    "本行需求": required_qty,
                    "库存分配前": stock_before,
                    "库存扣减": stock_deduct,
                    "库存缺口": max(required_qty - stock_before, 0),
                    "库存分配后": stock_after,
                    "当前可用分配前": available_before,
                    "当前可用扣减": available_deduct,
                    "当前可用缺口": max(required_qty - available_before, 0),
                    "当前可用分配后": available_after,
                    "替代料清单": assessed.get("替代料清单", ""),
                    "BOM差异标识": diff_label,
                    "共用母料数": shared_root_count,
                    "BOM差异说明": diff_detail,
                }
            )
            horizontal_stock_pool[material_code] = stock_after
            horizontal_available_pool[material_code] = available_after

    horizontal_shortage_df = pd.DataFrame(horizontal_rows)
    horizontal_columns = [
        "排产顺序",
        "排产日期",
        "母料号",
        "料品名称",
        "输入数量",
        "料号",
        "物料名称",
        "单台用量",
        "本行需求",
        "库存分配前",
        "库存扣减",
        "库存缺口",
        "库存分配后",
        "当前可用分配前",
        "当前可用扣减",
        "当前可用缺口",
        "当前可用分配后",
        "替代料清单",
        "BOM差异标识",
        "共用母料数",
        "BOM差异说明",
    ]
    if horizontal_shortage_df.empty:
        horizontal_shortage_df = pd.DataFrame(columns=horizontal_columns)
    else:
        horizontal_shortage_df = horizontal_shortage_df[horizontal_columns].sort_values(
            by=["排产顺序", "料号"], kind="stable"
        ).reset_index(drop=True)

    def root_display_label(root_code_text: str) -> str:
        root_spec = str(material_spec_map.get(root_code_text, "") or "").strip()
        return f"{root_code_text}-{root_spec}" if root_spec else root_code_text

    root_usage_columns = [f"{root_display_label(code)}单台用量" for code in root_order]
    root_recommend_columns = [f"{root_display_label(code)}推荐排产数量" for code in root_order]
    rolling_entry_columns: list[str] = []
    for entry in production_priority_rows:
        entry_label = f"{entry.get('priority')}-{root_display_label(str(entry.get('root_code', '')).strip())}"
        rolling_entry_columns.extend(
            [
                f"{entry_label}需求",
                f"{entry_label}缺口",
            ]
        )

    rolling_matrix_rows: list[dict[str, object]] = []
    rolling_matrix_pool = dict(stock_only_pool)
    for material_code in sorted(all_relevant_codes):
        usage_map = material_usage_by_root.get(material_code, {})
        usage_values = [float(value or 0) for value in usage_map.values()]
        present_count = len([value for value in usage_values if abs(value) >= FLOAT_ABS_TOL])
        max_qty = max(usage_values) if usage_values else 0.0
        min_qty = min(usage_values) if usage_values else 0.0
        range_diff = round(max_qty - min_qty, 6)
        if present_count == len(root_order):
            status = "全部共用" if abs(range_diff) < FLOAT_ABS_TOL else "全部共用-用量不同"
        elif present_count >= 2:
            status = "部分共用"
        else:
            status = "独有"
        assessed = material_assessment_map.get(material_code, {})
        row = {
            "共用状态": status,
            "物料编码": material_code,
            "品名": assessed.get("料品名称", "") or material_name_map.get(material_code, ""),
            "规格": assessed.get("规格", "") or material_spec_map.get(material_code, ""),
            "出现BOM数": present_count,
        }
        for root_code_text in root_order:
            usage_value = float(usage_map.get(root_code_text, 0) or 0)
            row[f"{root_display_label(root_code_text)}单台用量"] = usage_value if abs(usage_value) >= FLOAT_ABS_TOL else ""
        for root_code_text in root_order:
            row[f"{root_display_label(root_code_text)}推荐排产数量"] = recommended_qty_by_root.get(root_code_text, 0)
        row.update(
            {
                "供应商": assessed.get("供应商", ""),
                "采购": assessed.get("采购", ""),
                "当前库存": float(inventory_map.get(material_code, 0) or 0),
                "替代料库存": float(assessed.get("替代料库存", 0) or 0),
                "当前库存+替代库存": float(stock_only_pool.get(material_code, 0) or 0),
                "未转PR": float(inbound_pr_map.get(material_code, 0) or 0),
                "未清PO": float(inbound_po_map.get(material_code, 0) or 0),
                "当前可用": float(assessed.get("当前可用", 0) or 0),
                "替代料清单": assessed.get("替代料清单", ""),
            }
        )
        for entry in production_priority_rows:
            entry_label = f"{entry.get('priority')}-{root_display_label(str(entry.get('root_code', '')).strip())}"
            root_code_text = str(entry.get("root_code", "")).strip()
            root_qty = float(entry.get("qty", 0) or 0)
            unit_usage = float(usage_map.get(root_code_text, 0) or 0)
            required_qty = root_qty * unit_usage
            before_qty = float(rolling_matrix_pool.get(material_code, 0) or 0)
            shortage_qty = max(required_qty - before_qty, 0)
            after_qty = before_qty - required_qty
            row[f"{entry_label}需求"] = required_qty if abs(required_qty) >= FLOAT_ABS_TOL else ""
            row[f"{entry_label}缺口"] = shortage_qty if abs(required_qty) >= FLOAT_ABS_TOL else ""
            rolling_matrix_pool[material_code] = after_qty
        rolling_matrix_rows.append(row)

    rolling_matrix_columns = (
        ["共用状态", "物料编码", "品名", "规格", "出现BOM数"]
        + root_usage_columns
        + root_recommend_columns
        + [
            "供应商",
            "采购",
            "当前库存",
            "替代料库存",
            "当前库存+替代库存",
            "未转PR",
            "未清PO",
            "当前可用",
            "替代料清单",
        ]
        + rolling_entry_columns
    )
    rolling_matrix_df = pd.DataFrame(rolling_matrix_rows)
    if rolling_matrix_df.empty:
        rolling_matrix_df = pd.DataFrame(columns=rolling_matrix_columns)
    else:
        order_map = {"全部共用-用量不同": 0, "部分共用": 1, "独有": 2, "全部共用": 3}
        rolling_matrix_df = rolling_matrix_df[rolling_matrix_columns].sort_values(
            by=["共用状态", "物料编码"],
            key=lambda series: series.map(order_map).fillna(99) if series.name == "共用状态" else series,
            kind="stable",
        ).reset_index(drop=True)

    diff_stock_pool = dict(stock_only_pool)
    diff_available_pool = dict(horizontal_available_pool)
    diff_capacity_rows: list[dict[str, object]] = []
    for entry in production_priority_rows:
        root_code_text = str(entry.get("root_code", "")).strip()
        root_qty = float(entry.get("qty", 0) or 0)
        diff_profiles: list[dict[str, object]] = []
        for profile in root_material_profiles.get(root_code_text, []):
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            diff_unit, diff_label = differential_unit_usage(root_code_text, material_code, unit_usage)
            if not material_code:
                continue
            assessed = material_assessment_map.get(material_code, {})
            if diff_unit <= FLOAT_ABS_TOL:
                continue
            profile_copy = dict(profile)
            profile_copy["差异单台用量"] = diff_unit
            profile_copy["差异标识"] = diff_label
            profile_copy["料品名称"] = profile_copy.get("料品名称", "") or assessed.get("料品名称", "")
            diff_profiles.append(profile_copy)

        if not diff_profiles:
            stock_capacity_qty = int(math.floor(max(root_qty, 0)))
            available_capacity_qty = stock_capacity_qty
            stock_gap_qty = 0
            available_gap_qty = 0
            bottleneck_stock_items: list[dict[str, object]] = []
            bottleneck_available_items: list[dict[str, object]] = []
            diff_note = "无专用/多用差异物料，按差异物料口径不限制本段排产"
        else:
            stock_caps: list[dict[str, object]] = []
            available_caps: list[dict[str, object]] = []
            for profile in diff_profiles:
                material_code = str(profile.get("料号", "")).strip()
                diff_unit = float(profile.get("差异单台用量", 0) or 0)
                stock_before = float(diff_stock_pool.get(material_code, 0) or 0)
                available_before = float(diff_available_pool.get(material_code, 0) or 0)
                stock_cap = int(math.floor(max(stock_before, 0) / diff_unit)) if diff_unit > 0 else 0
                available_cap = int(math.floor(max(available_before, 0) / diff_unit)) if diff_unit > 0 else 0
                stock_caps.append({**profile, "可用量": stock_before, "可生产数量": max(stock_cap, 0)})
                available_caps.append({**profile, "可用量": available_before, "可生产数量": max(available_cap, 0)})
            stock_capacity_qty = min(root_qty, min(int(item["可生产数量"]) for item in stock_caps)) if stock_caps else root_qty
            available_capacity_qty = min(root_qty, min(int(item["可生产数量"]) for item in available_caps)) if available_caps else root_qty
            stock_capacity_qty = int(math.floor(max(stock_capacity_qty, 0)))
            available_capacity_qty = int(math.floor(max(available_capacity_qty, 0)))
            stock_gap_qty = max(root_qty - stock_capacity_qty, 0)
            available_gap_qty = max(root_qty - available_capacity_qty, 0)
            min_stock_cap = min(int(item["可生产数量"]) for item in stock_caps) if stock_caps else stock_capacity_qty
            min_available_cap = min(int(item["可生产数量"]) for item in available_caps) if available_caps else available_capacity_qty
            bottleneck_stock_items = [item for item in stock_caps if int(item["可生产数量"]) == min_stock_cap][:5]
            bottleneck_available_items = [item for item in available_caps if int(item["可生产数量"]) == min_available_cap][:5]
            diff_note = unique_join(
                [
                    f"{item.get('料号', '')}:{item.get('差异标识', '')}/单台差异{format_project_display_value(item.get('差异单台用量', 0))}"
                    for item in diff_profiles[:6]
                ],
                separator="；",
            )
            for profile in diff_profiles:
                material_code = str(profile.get("料号", "")).strip()
                diff_unit = float(profile.get("差异单台用量", 0) or 0)
                diff_stock_pool[material_code] = float(diff_stock_pool.get(material_code, 0) or 0) - stock_capacity_qty * diff_unit
                diff_available_pool[material_code] = float(diff_available_pool.get(material_code, 0) or 0) - available_capacity_qty * diff_unit

        common_profiles: list[dict[str, object]] = []
        for profile in root_material_profiles.get(root_code_text, []):
            material_code = str(profile.get("料号", "")).strip()
            unit_usage = float(profile.get("单台用量", 0) or 0)
            diff_unit, diff_label = differential_unit_usage(root_code_text, material_code, unit_usage)
            if material_code and diff_unit <= FLOAT_ABS_TOL:
                common_profiles.append({**profile, "差异标识": diff_label})
        common_note = unique_join(
            [
                f"{item.get('料号', '')}:{item.get('差异标识', '')}/单台{format_project_display_value(item.get('单台用量', 0))}"
                for item in common_profiles[:10]
            ],
            separator="；",
        )

        stock_bottleneck_source = bottleneck_stock_items or bottleneck_available_items
        available_bottleneck_source = bottleneck_available_items or bottleneck_stock_items
        diff_capacity_rows.append(
            {
                "排产顺序": entry.get("priority", ""),
                "排产日期": entry.get("plan_date") or "",
                "母料号": root_code_text,
                "料品名称": material_name_map.get(root_code_text, ""),
                "输入数量": root_qty,
                "差异物料数": len(diff_profiles),
                "通用物料数": len(common_profiles),
                "通用物料清单": common_note,
                "当前库存差异可生产": stock_capacity_qty,
                "当前库存差异缺口": stock_gap_qty,
                "当前可用差异可生产": available_capacity_qty,
                "当前可用差异缺口": available_gap_qty,
                "瓶颈差异物料": unique_join([item.get("料号", "") for item in stock_bottleneck_source[:5]]),
                "瓶颈物料名称": unique_join([item.get("料品名称", "") for item in stock_bottleneck_source[:5]]),
                "瓶颈单台差异用量": unique_join(
                    [format_project_display_value(item.get("差异单台用量", 0)) for item in stock_bottleneck_source[:5]]
                ),
                "瓶颈库存可用": unique_join(
                    [format_project_display_value(item.get("可用量", 0)) for item in stock_bottleneck_source[:5]]
                ),
                "瓶颈当前可用": unique_join(
                    [format_project_display_value(item.get("可用量", 0)) for item in available_bottleneck_source[:5]]
                ),
                "BOM差异说明": diff_note,
            }
        )

    diff_capacity_columns = [
        "排产顺序",
        "排产日期",
        "母料号",
        "料品名称",
        "输入数量",
        "差异物料数",
        "通用物料数",
        "通用物料清单",
        "当前库存差异可生产",
        "当前库存差异缺口",
        "当前可用差异可生产",
        "当前可用差异缺口",
        "瓶颈差异物料",
        "瓶颈物料名称",
        "瓶颈单台差异用量",
        "瓶颈库存可用",
        "瓶颈当前可用",
        "BOM差异说明",
    ]
    diff_capacity_df = pd.DataFrame(diff_capacity_rows)
    if diff_capacity_df.empty:
        diff_capacity_df = pd.DataFrame(columns=diff_capacity_columns)
    else:
        diff_capacity_df = diff_capacity_df[diff_capacity_columns].sort_values(
            by=["排产顺序", "母料号"], kind="stable"
        ).reset_index(drop=True)

    last_capacity_by_entry: dict[int, int] = {}
    producible_rows: list[dict[str, object]] = []
    for timeline_date in sorted_timeline_dates:
        remaining_pool: dict[str, float] = {}
        for material_code in all_relevant_codes:
            assessed = material_assessment_map.get(material_code, {})
            available_now = float(assessed.get("当前可用", 0) or 0)
            reply_before_date = sum(
                qty
                for reply_date, qty in reply_numeric_map.get(material_code, {}).items()
                if isinstance(reply_date, date) and reply_date <= timeline_date
            )
            remaining_pool[material_code] = available_now + float(reply_before_date or 0)

        for entry_index, entry in enumerate(production_priority_rows):
            root_code_text = str(entry.get("root_code", "")).strip()
            root_qty = float(entry.get("qty", 0) or 0)
            profiles = root_material_profiles.get(root_code_text, [])
            root_unknown = pd.DataFrame()
            if not unknown_leaf_df_for_capacity.empty and "母料号" in unknown_leaf_df_for_capacity.columns:
                root_unknown = unknown_leaf_df_for_capacity[
                    unknown_leaf_df_for_capacity["母料号"].astype(str).str.strip() == root_code_text
                ].copy()

            if not profiles:
                capacity_qty = 0
                bottleneck_items: list[dict[str, object]] = []
                reason = "母料号未识别或未识别到外购物料"
                if not root_unknown.empty:
                    reason = f"存在 {root_unknown['料号'].astype(str).str.strip().nunique()} 个未识别下层物料"
            else:
                material_caps: list[dict[str, object]] = []
                for profile in profiles:
                    material_code = str(profile.get("料号", "")).strip()
                    unit_usage = float(profile.get("单台用量", 0) or 0)
                    available_at_date = float(remaining_pool.get(material_code, 0) or 0)
                    capacity = int(math.floor(max(available_at_date, 0) / unit_usage)) if unit_usage > 0 else 0
                    material_caps.append(
                        {
                            "料号": material_code,
                            "料品名称": profile.get("料品名称", ""),
                            "单台用量": unit_usage,
                            "可用量": available_at_date,
                            "可生产数量": max(capacity, 0),
                        }
                    )
                capacity_qty = min(root_qty, min(int(item["可生产数量"]) for item in material_caps)) if material_caps else 0
                capacity_qty = int(math.floor(max(capacity_qty, 0)))
                bottleneck_items = [
                    item for item in material_caps
                    if int(item["可生产数量"]) == min(int(cap_item["可生产数量"]) for cap_item in material_caps)
                ][:5] if material_caps else []
                for profile in profiles:
                    material_code = str(profile.get("料号", "")).strip()
                    remaining_pool[material_code] = float(remaining_pool.get(material_code, 0) or 0) - capacity_qty * float(profile.get("单台用量", 0) or 0)

                if not root_unknown.empty:
                    reason = f"存在 {root_unknown['料号'].astype(str).str.strip().nunique()} 个未识别下层物料，产能仅按已识别外购物料测算"
                elif _fge(capacity_qty, root_qty):
                    reason = "按MRP排产优先级分配后，可满足本段排产/输入数量"
                else:
                    reason_parts = []
                    for item in bottleneck_items[:3]:
                        gap_to_target = max(root_qty * float(item.get("单台用量", 0) or 0) - float(item.get("可用量", 0) or 0), 0)
                        reason_parts.append(
                            f"优先级{entry.get('priority')}分配后 {item.get('料号', '')} 缺 {format_project_display_value(gap_to_target)}"
                        )
                    reason = "；".join(reason_parts) if reason_parts else "无可用外购物料数据"

            if (
                last_capacity_by_entry.get(entry_index) == capacity_qty
                and timeline_date != sorted_timeline_dates[-1]
            ):
                continue
            last_capacity_by_entry[entry_index] = capacity_qty
            bottleneck_display = unique_join([item.get("料号", "") for item in bottleneck_items[:5]])
            bottleneck_name_display = unique_join([item.get("料品名称", "") for item in bottleneck_items[:5]])
            bottleneck_usage_display = unique_join(
                [format_project_display_value(item.get("单台用量", 0)) for item in bottleneck_items[:5]]
            )
            bottleneck_available_display = unique_join(
                [format_project_display_value(item.get("可用量", 0)) for item in bottleneck_items[:5]]
            )
            producible_rows.append(
                {
                    "母料号": root_code_text,
                    "料品名称": material_name_map.get(root_code_text, ""),
                    "输入数量": root_qty,
                    "排产顺序": entry.get("priority", ""),
                    "排产日期": entry.get("plan_date") or "",
                    "日期": timeline_date,
                    "可生产数量": capacity_qty,
                    "当前库存可生成": stock_capacity_by_entry.get(entry_index, 0),
                    "距离输入缺口": max(root_qty - capacity_qty, 0),
                    "瓶颈物料": bottleneck_display,
                    "瓶颈物料名称": bottleneck_name_display,
                    "瓶颈单台用量": bottleneck_usage_display,
                    "瓶颈可用量": bottleneck_available_display,
                    "瓶颈原因": reason,
                }
            )

    producible_df = pd.DataFrame(producible_rows)
    if producible_df.empty:
        producible_df = pd.DataFrame(
            columns=[
                "母料号",
                "料品名称",
                "输入数量",
                "排产顺序",
                "排产日期",
                "日期",
                "可生产数量",
                "当前库存可生成",
                "距离输入缺口",
                "瓶颈物料",
                "瓶颈物料名称",
                "瓶颈单台用量",
                "瓶颈可用量",
                "瓶颈原因",
            ]
        )
    else:
        producible_df = producible_df.sort_values(
            by=["日期", "排产顺序", "母料号"],
            kind="stable",
        ).reset_index(drop=True)

    issue_df = root_material_df.merge(
        material_df[
            [
                "料号",
                "供应商",
                "采购",
                "总需求",
                "当前可用",
                "当前缺口",
                "当前可用缺口",
                "替代料清单",
                "采购答复累计",
                "采购答复说明",
                "最早齐套日期",
                "状态",
                "问题原因",
            ]
        ],
        on="料号",
        how="left",
    )
    issue_df = issue_df[issue_df["问题原因"].fillna("") != ""].copy()
    issue_df = issue_df.rename(columns={"本次需求": "本母件需求"})

    unknown_leaf_df = pd.DataFrame(unknown_leaf_rows)
    if not unknown_leaf_df.empty:
        unknown_leaf_df = (
            unknown_leaf_df.groupby(["母料号", "母件输入数量", "料号"], sort=True)
            .agg(
                料品名称=("料品名称", "first"),
                规格=("规格", "first"),
                上层物料编码=("上层物料编码", unique_join),
                上层物料名称=("上层物料名称", unique_join),
                需求数量=("需求数量", "sum"),
                问题原因=("问题原因", "first"),
            )
            .reset_index()
        )
    else:
        unknown_leaf_df = pd.DataFrame(
            columns=["母料号", "母件输入数量", "料号", "料品名称", "规格", "上层物料编码", "上层物料名称", "需求数量", "问题原因"]
        )

    missing_root_df = pd.DataFrame(missing_root_rows)
    if missing_root_df.empty:
        missing_root_df = pd.DataFrame(columns=["母料号", "输入数量", "料品名称", "问题原因"])

    root_summary_rows: list[dict[str, object]] = []
    material_ready_map = {
        str(record["料号"]).strip(): record.get("最早齐套日期")
        for record in material_df.to_dict("records")
    }
    for root_code, root_qty in normalized_roots:
        root_code_text = str(root_code).strip()
        root_materials = root_material_df[root_material_df["母料号"] == root_code_text].copy()
        root_issues = issue_df[issue_df["母料号"] == root_code_text].copy()
        root_unknown = unknown_leaf_df[unknown_leaf_df["母料号"] == root_code_text].copy()
        ready_dates = [
            material_ready_map.get(material_code)
            for material_code in root_materials["料号"].astype(str).str.strip().tolist()
            if material_ready_map.get(material_code)
        ]
        all_root_materials_ready = (
            not root_materials.empty
            and len(ready_dates) == len(root_materials["料号"].astype(str).str.strip().unique())
            and root_unknown.empty
        )
        root_ready_date = max(ready_dates) if ready_dates and all_root_materials_ready else None

        if not missing_root_df.empty and root_code_text in set(missing_root_df["母料号"].astype(str).str.strip()):
            conclusion = "母料号未识别"
        elif not root_unknown.empty:
            conclusion = "存在未识别外购物料"
        elif root_materials.empty:
            conclusion = "未识别到外购物料"
        elif root_issues.empty:
            conclusion = "现有量可开工"
        elif root_ready_date:
            conclusion = f"{root_ready_date} 可开工"
        else:
            conclusion = "待采购回复"
        if root_code_text in inferred_quantity_notes:
            conclusion = f"{inferred_quantity_notes[root_code_text]}；{conclusion}"

        root_summary_rows.append(
            {
                "母料号": root_code_text,
                "料品名称": material_name_map.get(root_code_text, ""),
                "输入数量": root_qty,
                "外购物料数": int(root_materials["料号"].astype(str).str.strip().nunique()) if not root_materials.empty else 0,
                "问题物料数": int(root_issues["料号"].astype(str).str.strip().nunique()) if not root_issues.empty else 0,
                "未识别物料数": int(root_unknown["料号"].astype(str).str.strip().nunique()) if not root_unknown.empty else 0,
                "外购齐套日期": root_ready_date,
                "结论": conclusion,
            }
        )

    root_summary_df = pd.DataFrame(root_summary_rows).sort_values(by=["母料号"], kind="stable").reset_index(drop=True)

    unrecognized_material_rows = build_unrecognized_material_rows(unknown_leaf_df, missing_root_df)
    if unrecognized_material_rows:
        material_df = pd.concat([material_df, pd.DataFrame(unrecognized_material_rows)], ignore_index=True)

    emit_progress(progress_callback, "步骤 4/4：汇总齐套结果...")
    unresolved_issue_count = int((material_df["问题原因"].fillna("") != "").sum()) if not material_df.empty else 0
    batch_ready_date = None
    if not root_summary_df.empty and root_summary_df["外购齐套日期"].notna().all() and unknown_leaf_df.empty and missing_root_df.empty:
        batch_ready_date = max(root_summary_df["外购齐套日期"].tolist())
    if not missing_root_df.empty:
        batch_conclusion = "存在未识别母料号"
    elif not unknown_leaf_df.empty:
        batch_conclusion = "存在未识别外购物料"
    elif unresolved_issue_count == 0:
        batch_conclusion = "现有量可开工"
    elif batch_ready_date:
        batch_conclusion = f"{batch_ready_date} 可开工"
    else:
        batch_conclusion = "待采购回复"

    batch_summary = {
        "root_count": len(normalized_roots),
        "material_count": int(material_df["料号"].astype(str).str.strip().nunique()) if not material_df.empty else 0,
        "issue_count": unresolved_issue_count,
        "unknown_count": int(unknown_leaf_df["料号"].astype(str).str.strip().nunique()) if not unknown_leaf_df.empty else 0,
        "missing_root_count": len(missing_root_df),
        "ready_date": batch_ready_date,
        "conclusion": batch_conclusion,
    }
    emit_progress(progress_callback, "步骤 4/4：分析完成")
    return PurchaseReadinessResult(
        root_summary_df=root_summary_df,
        issue_df=issue_df.sort_values(by=["母料号", "问题原因", "料号"], kind="stable").reset_index(drop=True),
        material_df=material_df,
        producible_df=producible_df,
        horizontal_shortage_df=horizontal_shortage_df,
        diff_capacity_df=diff_capacity_df,
        rolling_matrix_df=rolling_matrix_df,
        missing_root_df=missing_root_df,
        unknown_leaf_df=unknown_leaf_df,
        batch_summary=batch_summary,
        carried_reply_cell_count=carry_stats["cell_count"],
        carried_reply_material_count=carry_stats["material_count"],
        carried_reply_file_count=carry_stats["file_count"],
    )


def read_workbook_tables(
    workbook_path: Path,
    *,
    external_bom_df: pd.DataFrame | None = None,
    progress_callback: ProgressCallback | None = None,
    progress_prefix: str | None = None,
) -> dict[str, pd.DataFrame]:
    with pd.ExcelFile(workbook_path) as excel_file:
        available_sheet_names = set(excel_file.sheet_names)

    sheet_names = [
        "生产订单",
        "采购数据",
        "期初库存",
        "供应商库存",
        "在途请购",
        "在途采购",
        "期初工单缺料",
    ]
    if external_bom_df is None:
        sheet_names.insert(1, "BOM")

    frames: dict[str, pd.DataFrame] = {}
    for sheet_name in sheet_names:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：读取 {sheet_name}...")
            else:
                emit_progress(progress_callback, f"读取 {sheet_name}...")
        try:
            frames[sheet_name] = normalize_columns(pd.read_excel(workbook_path, sheet_name=sheet_name))
        except ValueError as exc:
            raise WorkbookInputError(f"缺少工作表: {sheet_name}") from exc

    if external_bom_df is not None:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：使用当前已加载 BOM...")
            else:
                emit_progress(progress_callback, "使用当前已加载 BOM...")
        frames["BOM"] = normalize_columns(external_bom_df.copy())
    if PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME in available_sheet_names:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：读取 {PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME}...")
            else:
                emit_progress(progress_callback, f"读取 {PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME}...")
        frames[PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME] = normalize_columns(
            pd.read_excel(workbook_path, sheet_name=PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME)
        )
    if RECEIVING_STATUS_SHEET_NAME in available_sheet_names:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：读取 {RECEIVING_STATUS_SHEET_NAME}...")
            else:
                emit_progress(progress_callback, f"读取 {RECEIVING_STATUS_SHEET_NAME}...")
        frames[RECEIVING_STATUS_SHEET_NAME] = normalize_columns(
            pd.read_excel(workbook_path, sheet_name=RECEIVING_STATUS_SHEET_NAME)
        )
    if ENABLE_CONFIG_SUPPLEMENT_PLAN and INDUSTRIAL_CONFIG_SHEET_NAME in available_sheet_names:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：读取 {INDUSTRIAL_CONFIG_SHEET_NAME}...")
            else:
                emit_progress(progress_callback, f"读取 {INDUSTRIAL_CONFIG_SHEET_NAME}...")
        frames[INDUSTRIAL_CONFIG_SHEET_NAME] = normalize_columns(
            pd.read_excel(workbook_path, sheet_name=INDUSTRIAL_CONFIG_SHEET_NAME)
        )
    if INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME in available_sheet_names:
        if progress_callback is not None:
            if progress_prefix:
                emit_progress(progress_callback, f"{progress_prefix}：读取 {INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME}...")
            else:
                emit_progress(progress_callback, f"读取 {INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME}...")
        frames[INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME] = normalize_columns(
            pd.read_excel(workbook_path, sheet_name=INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME)
        )
    return frames


def validate_input_workbook(
    workbook_path: Path,
    *,
    external_bom_df: pd.DataFrame | None = None,
) -> list[str]:
    frames = read_workbook_tables(workbook_path, external_bom_df=external_bom_df)
    issues: list[str] = []

    production_orders = frames["生产订单"].copy()
    bom_df = frames["BOM"].copy()
    parent_item_col = first_matching_column(production_orders, ["母件料号"])
    bom_parent_col = first_matching_column(bom_df, ["母件料号"])
    scheduled_codes = {
        str(value).strip()
        for value in production_orders[parent_item_col].tolist()
        if str(value).strip() and str(value).strip().lower() != "nan"
    }
    bom_parent_codes = {
        str(value).strip()
        for value in bom_df[bom_parent_col].tolist()
        if str(value).strip() and str(value).strip().lower() != "nan"
    }
    missing_bom_codes = sorted(code for code in scheduled_codes if code not in bom_parent_codes)
    if missing_bom_codes:
        sample = "、".join(missing_bom_codes[:8])
        suffix = " 等" if len(missing_bom_codes) > 8 else ""
        issues.append(
            f"生产订单里有 {len(missing_bom_codes)} 个母料号未在当前 BOM 中找到，可能导致展开漏项。示例：{sample}{suffix}"
        )

    usage_exclusions = normalize_parent_child_usage_exclusions(
        frames.get(PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME)
    )
    if not usage_exclusions.empty:
        bom_child_col = first_matching_column(bom_df, ["BOM子项.子件料品.料号"])
        bom_pairs = set(
            zip(
                bom_df[bom_parent_col].astype(str).str.strip(),
                bom_df[bom_child_col].astype(str).str.strip(),
            )
        )
        missed_pairs = [
            (row.parent_code, row.child_code)
            for row in usage_exclusions.itertuples(index=False)
            if (row.parent_code, row.child_code) not in bom_pairs
        ]
        if missed_pairs:
            sample = "、".join(f"{parent}->{child}" for parent, child in missed_pairs[:8])
            suffix = " 等" if len(missed_pairs) > 8 else ""
            issues.append(
                f"{PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME} 有 {len(missed_pairs)} 条规则未在当前 BOM 中命中，可能是料号写错或 BOM 已变化。示例：{sample}{suffix}"
            )

    work_shortage = frames["期初工单缺料"].copy()
    work_item_col = first_matching_column(work_shortage, ["物料.编码"], required=False)
    work_shortage_col = first_matching_column(work_shortage, ["缺料"], required=False)
    work_plan_col = first_matching_column(work_shortage, ["计划数量"], required=False)
    work_delivery_col = first_matching_column(work_shortage, ["配送数量"], required=False)
    if work_item_col and work_shortage_col and work_plan_col and work_delivery_col:
        item_codes = work_shortage[work_item_col].astype(str).str.strip()
        plan_qty = coerce_number_series(work_shortage[work_plan_col])
        delivery_qty = coerce_number_series(work_shortage[work_delivery_col])
        shortage_qty = pd.to_numeric(work_shortage[work_shortage_col], errors="coerce")
        missing_shortage_mask = (
            (item_codes != "")
            & (item_codes.str.lower() != "nan")
            & (plan_qty > delivery_qty)
            & shortage_qty.isna()
        )
        if missing_shortage_mask.any():
            sample_codes = unique_join(item_codes[missing_shortage_mask].head(8).tolist())
            issues.append(
                f"期初工单缺料里有 {int(missing_shortage_mask.sum())} 行“计划数量 > 配送数量”但“缺料”列为空，程序会自动按“计划数量 - 配送数量”补算。示例料号：{sample_codes}"
            )

    purchase_df = frames["采购数据"].copy()
    purchase_item_col = first_matching_column(purchase_df, ["物料号"], required=False)
    if purchase_item_col:
        purchase_codes = purchase_df[purchase_item_col].astype(str).str.strip()
        duplicate_codes = sorted(
            {
                code
                for code in purchase_codes[purchase_codes.duplicated(keep=False)].tolist()
                if code and code.lower() != "nan"
            }
        )
        if duplicate_codes:
            sample = "、".join(duplicate_codes[:8])
            suffix = " 等" if len(duplicate_codes) > 8 else ""
            issues.append(
                f"采购数据里有 {len(duplicate_codes)} 个料号重复，供应商/采购员可能会取到第一条记录。示例：{sample}{suffix}"
            )

    return issues


def read_substitute_rules(workbook_path: Path) -> pd.DataFrame:
    try:
        raw = pd.read_excel(workbook_path, sheet_name="替代料", header=None)
    except ValueError:
        return pd.DataFrame(columns=["before_code", "after_code", "contexts", "sort_order"])

    rules = []
    for idx in range(2, len(raw)):
        before_code = extract_material_code(raw.iat[idx, 1] if raw.shape[1] > 1 else None)
        after_code = extract_material_code(raw.iat[idx, 6] if raw.shape[1] > 6 else None)
        if not before_code or not after_code:
            continue

        contexts = []
        for customer_col, parent_col in [(11, 12), (14, 15)]:
            customer = raw.iat[idx, customer_col] if raw.shape[1] > customer_col else None
            parent_code = raw.iat[idx, parent_col] if raw.shape[1] > parent_col else None
            customer_text = "" if customer is None or pd.isna(customer) else str(customer).strip()
            parent_text = extract_material_code(parent_code)
            if customer_text or parent_text:
                contexts.append((customer_text, parent_text))

        rules.append(
            {
                "before_code": before_code,
                "after_code": after_code,
                "contexts": contexts,
                "sort_order": idx,
            }
        )

    return pd.DataFrame(rules)


def read_excluded_material_codes(workbook_path: Path) -> set[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "物料编码排除清单" not in workbook.sheetnames:
            return set()
        worksheet = workbook["物料编码排除清单"]
        codes = set()
        for row in worksheet.iter_rows(values_only=True):
            for value in row:
                material_code = maybe_material_code(value)
                if material_code:
                    codes.add(material_code)
        return codes
    finally:
        workbook.close()


def normalize_parent_child_usage_exclusions(raw_df: pd.DataFrame | None) -> pd.DataFrame:
    result_columns = ["parent_code", "child_code", "remark"]
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=result_columns)

    parent_col = first_matching_column(
        raw_df,
        ["母件料号", "母料号", "父项料号", "父项物料编码", "上层物料编码"],
        required=False,
    )
    child_col = first_matching_column(
        raw_df,
        ["子件料号", "子料号", "子项料号", "子项物料编码", "下层物料编码", "物料编码"],
        required=False,
    )
    remark_col = first_matching_column(raw_df, ["备注", "说明"], required=False)

    non_empty_rows = raw_df.dropna(how="all")
    if not parent_col or not child_col:
        if non_empty_rows.empty:
            return pd.DataFrame(columns=result_columns)
        raise WorkbookInputError(
            f"{PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME} 缺少字段：需要“母件料号”和“子件料号”两列"
        )

    exclusions = pd.DataFrame(
        {
            "parent_code": raw_df[parent_col].apply(extract_material_code),
            "child_code": raw_df[child_col].apply(extract_material_code),
            "remark": raw_df[remark_col].fillna("").astype(str).str.strip() if remark_col else "",
        }
    )
    exclusions = exclusions[
        (exclusions["parent_code"] != "")
        & (exclusions["child_code"] != "")
    ].copy()
    if exclusions.empty:
        return pd.DataFrame(columns=result_columns)
    return exclusions.drop_duplicates(subset=["parent_code", "child_code"], keep="first").reset_index(drop=True)


def parent_child_exclusion_pairs(raw_df: pd.DataFrame | None) -> set[tuple[str, str]]:
    exclusions = normalize_parent_child_usage_exclusions(raw_df)
    if exclusions.empty:
        return set()
    return set(zip(exclusions["parent_code"], exclusions["child_code"]))


def _pair_label(code: str, name: str) -> str:
    code = str(code or "").strip()
    name = str(name or "").strip()
    return f"{code} {name}".strip()


def _summarize_pairs(pairs: list[tuple[str, str]], limit: int = 5) -> str:
    unique_pairs: list[tuple[str, str]] = []
    seen = set()
    for code, name in pairs:
        key = (str(code or "").strip(), str(name or "").strip())
        if not key[0] or key in seen:
            continue
        seen.add(key)
        unique_pairs.append(key)
    labels = [_pair_label(code, name) for code, name in unique_pairs[:limit]]
    return "；".join(label for label in labels if label)


def build_subassembly_suggestions(
    production_plan_df: pd.DataFrame,
    bom_df: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, str]]:
    empty_columns = [
        "客户",
        "母件料号",
        "母件品名",
        "母件规格",
        "上线日期",
        "上线数量",
        "直接上层数量",
        "直接上层摘要",
        "顶层来源数量",
        "顶层来源摘要",
        "说明",
    ]
    if production_plan_df.empty or bom_df.empty:
        return pd.DataFrame(columns=empty_columns), {}

    bom_parent_col = first_matching_column(bom_df, ["母件料号"])
    parent_name_col = first_matching_column(bom_df, ["母件品名"], required=False)
    parent_spec_col = first_matching_column(bom_df, ["母件规格"], required=False)
    child_item_col = first_matching_column(bom_df, ["BOM子项.子件料品.料号"])
    child_name_col = first_matching_column(bom_df, ["BOM子项.子件料品.料品名称"])
    child_spec_col = first_matching_column(bom_df, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom_df, ["BOM子项.子件用量"])

    bom_graph = pd.DataFrame(
        {
            "parent_code": bom_df[bom_parent_col].astype(str).str.strip(),
            "parent_name": bom_df[parent_name_col] if parent_name_col else "",
            "parent_spec": bom_df[parent_spec_col] if parent_spec_col else "",
            "child_code": bom_df[child_item_col].astype(str).str.strip(),
            "child_name": bom_df[child_name_col],
            "child_spec": bom_df[child_spec_col] if child_spec_col else "",
            "usage": coerce_number_series(bom_df[child_usage_col]),
        }
    )
    bom_graph = bom_graph[(bom_graph["parent_code"] != "") & (bom_graph["child_code"] != "") & (bom_graph["usage"] != 0)].copy()
    if bom_graph.empty:
        return pd.DataFrame(columns=empty_columns), {}

    assembly_codes = set(bom_graph["parent_code"])
    scheduled_codes = set(production_plan_df["母件料号"].astype(str).str.strip())
    edges: dict[str, list[dict[str, object]]] = defaultdict(list)
    material_names: dict[str, str] = {}
    material_specs: dict[str, str] = {}

    for row in bom_graph.itertuples(index=False):
        parent_code = str(row.parent_code).strip()
        child_code = str(row.child_code).strip()
        parent_name = "" if pd.isna(row.parent_name) else str(row.parent_name).strip()
        parent_spec = "" if pd.isna(row.parent_spec) else str(row.parent_spec).strip()
        child_name = "" if pd.isna(row.child_name) else str(row.child_name).strip()
        child_spec = "" if pd.isna(row.child_spec) else str(row.child_spec).strip()
        usage = float(row.usage)
        if not parent_code or not child_code or usage == 0:
            continue
        edges[parent_code].append(
            {
                "child_code": child_code,
                "child_name": child_name,
                "child_spec": child_spec,
                "usage": usage,
            }
        )
        if parent_name:
            material_names[parent_code] = parent_name
        if parent_spec:
            material_specs[parent_code] = parent_spec
        if child_name and child_code not in material_names:
            material_names[child_code] = child_name
        if child_spec and child_code not in material_specs:
            material_specs[child_code] = child_spec

    suggestion_rows: list[dict[str, object]] = []

    def walk(
        root_code: str,
        root_name: str,
        customer: str,
        online_date: date | None,
        parent_code: str,
        demand_qty: float,
        trail: set[str],
    ) -> None:
        for edge in edges.get(parent_code, []):
            child_code = edge["child_code"]
            child_qty = demand_qty * float(edge["usage"])
            if child_code in assembly_codes and child_code not in scheduled_codes:
                suggestion_rows.append(
                    {
                        "客户": customer,
                        "建议母件料号": child_code,
                        "建议母件品名": material_names.get(child_code, ""),
                        "建议母件规格": material_specs.get(child_code, ""),
                        "建议上线日期": online_date,
                        "建议上线数量": child_qty,
                        "直接上层料号": parent_code,
                        "直接上层品名": material_names.get(parent_code, ""),
                        "顶层来源料号": root_code,
                        "顶层来源品名": root_name,
                    }
                )
            if child_code in assembly_codes and child_code not in trail:
                walk(
                    root_code,
                    root_name,
                    customer,
                    online_date,
                    child_code,
                    child_qty,
                    trail | {child_code},
                )

    for record in production_plan_df.to_dict("records"):
        root_code = str(record.get("母件料号", "")).strip()
        if not root_code:
            continue
        qty = float(record.get("上线数量", 0) or 0)
        if qty == 0:
            continue
        online_date = record.get("上线日期")
        if pd.isna(online_date):
            online_date = None
        root_name = str(record.get("母件品名", "") or "").strip()
        customer = str(record.get("客户", "") or "").strip()
        walk(root_code, root_name, customer, online_date, root_code, qty, {root_code})

    if not suggestion_rows:
        return pd.DataFrame(columns=empty_columns), {}

    details_df = pd.DataFrame(suggestion_rows)
    aggregated_rows = []
    for group_key, group_df in details_df.groupby(
        ["客户", "建议母件料号", "建议母件品名", "建议母件规格", "建议上线日期"],
        sort=True,
        dropna=False,
    ):
        customer, material_code, material_name, material_spec, online_date = group_key
        direct_pairs = list(
            zip(group_df["直接上层料号"].astype(str), group_df["直接上层品名"].astype(str), strict=False)
        )
        root_pairs = list(
            zip(group_df["顶层来源料号"].astype(str), group_df["顶层来源品名"].astype(str), strict=False)
        )
        aggregated_rows.append(
            {
                "客户": customer,
                "母件料号": material_code,
                "母件品名": material_name,
                "母件规格": material_spec,
                "上线日期": online_date,
                "上线数量": float(group_df["建议上线数量"].sum()),
                "直接上层数量": len({pair for pair in direct_pairs if pair[0].strip()}),
                "直接上层摘要": _summarize_pairs([(code, name) for code, name in direct_pairs if str(code).strip()]),
                "顶层来源数量": len({pair for pair in root_pairs if pair[0].strip()}),
                "顶层来源摘要": _summarize_pairs([(code, name) for code, name in root_pairs if str(code).strip()]),
                "说明": "建议补充排产",
            }
        )

    suggestion_df = pd.DataFrame(aggregated_rows).sort_values(
        by=["上线日期", "母件料号", "客户"],
        kind="stable",
    ).reset_index(drop=True)

    note_map: dict[str, str] = {}
    for material_code, group_df in details_df.groupby("建议母件料号", sort=True):
        direct_pairs = list(
            zip(group_df["直接上层料号"].astype(str), group_df["直接上层品名"].astype(str), strict=False)
        )
        unique_parents = []
        seen = set()
        for code, name in direct_pairs:
            key = (str(code).strip(), str(name).strip())
            if not key[0] or key in seen:
                continue
            seen.add(key)
            unique_parents.append(key)
        summary_text = _summarize_pairs(unique_parents)
        if summary_text:
            note_map[str(material_code).strip()] = f"建议补充排产；直接上层 {len(unique_parents)} 项：{summary_text}"

    return suggestion_df, note_map


def _unique_codes_v2(values: Iterable[str]) -> list[str]:
    unique_codes: list[str] = []
    seen = set()
    for value in values:
        code = str(value or "").strip()
        if not code or code in seen:
            continue
        seen.add(code)
        unique_codes.append(code)
    return unique_codes


def _serialize_codes_v2(values: Iterable[str]) -> str:
    return "|".join(_unique_codes_v2(values))


def _deserialize_codes_v2(value: object) -> list[str]:
    if value is None or pd.isna(value):
        return []
    return _unique_codes_v2(str(value).split("|"))


def _summarize_codes_v2(values: Iterable[str], limit: int = 5) -> str:
    unique_codes = _unique_codes_v2(values)
    if not unique_codes:
        return ""
    if len(unique_codes) <= limit:
        return "、".join(unique_codes)
    return f"{'、'.join(unique_codes[:limit])} 等{len(unique_codes)}项"


def _merge_note_texts_v2(*values: object) -> str:
    merged: list[str] = []
    seen = set()
    for value in values:
        text = "" if value is None or pd.isna(value) else str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        merged.append(text)
    return "；".join(merged)


def build_subassembly_suggestions_v2(
    production_plan_df: pd.DataFrame,
    bom_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    suggestion_df, _legacy_note_map = build_subassembly_suggestions(production_plan_df, bom_df)
    if production_plan_df.empty or bom_df.empty:
        return suggestion_df, pd.DataFrame(
            columns=[
                "客户",
                "建议母件料号",
                "建议母件品名",
                "建议母件规格",
                "建议上线日期",
                "建议上线数量",
                "直接上层料号",
                "直接上层品名",
                "顶层来源料号",
                "顶层来源品名",
            ]
        )

    bom_parent_col = first_matching_column(bom_df, ["母件料号"])
    parent_name_col = first_matching_column(bom_df, ["母件品名"], required=False)
    parent_spec_col = first_matching_column(bom_df, ["母件规格"], required=False)
    child_item_col = first_matching_column(bom_df, ["BOM子项.子件料品.料号"])
    child_name_col = first_matching_column(bom_df, ["BOM子项.子件料品.料品名称"])
    child_spec_col = first_matching_column(bom_df, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom_df, ["BOM子项.子件用量"])

    bom_graph = pd.DataFrame(
        {
            "parent_code": bom_df[bom_parent_col].astype(str).str.strip(),
            "parent_name": bom_df[parent_name_col] if parent_name_col else "",
            "parent_spec": bom_df[parent_spec_col] if parent_spec_col else "",
            "child_code": bom_df[child_item_col].astype(str).str.strip(),
            "child_name": bom_df[child_name_col],
            "child_spec": bom_df[child_spec_col] if child_spec_col else "",
            "usage": coerce_number_series(bom_df[child_usage_col]),
        }
    )
    bom_graph = bom_graph[(bom_graph["parent_code"] != "") & (bom_graph["child_code"] != "") & (bom_graph["usage"] != 0)].copy()
    if bom_graph.empty:
        return suggestion_df, pd.DataFrame(
            columns=[
                "客户",
                "建议母件料号",
                "建议母件品名",
                "建议母件规格",
                "建议上线日期",
                "建议上线数量",
                "直接上层料号",
                "直接上层品名",
                "顶层来源料号",
                "顶层来源品名",
            ]
        )

    assembly_codes = set(bom_graph["parent_code"])
    scheduled_codes = set(production_plan_df["母件料号"].astype(str).str.strip())
    edges: dict[str, list[dict[str, object]]] = defaultdict(list)
    material_names: dict[str, str] = {}
    material_specs: dict[str, str] = {}

    for row in bom_graph.itertuples(index=False):
        parent_code = str(row.parent_code).strip()
        child_code = str(row.child_code).strip()
        parent_name = "" if pd.isna(row.parent_name) else str(row.parent_name).strip()
        parent_spec = "" if pd.isna(row.parent_spec) else str(row.parent_spec).strip()
        child_name = "" if pd.isna(row.child_name) else str(row.child_name).strip()
        child_spec = "" if pd.isna(row.child_spec) else str(row.child_spec).strip()
        usage = float(row.usage)
        if not parent_code or not child_code or usage == 0:
            continue
        edges[parent_code].append(
            {
                "child_code": child_code,
                "child_name": child_name,
                "child_spec": child_spec,
                "usage": usage,
            }
        )
        if parent_name:
            material_names[parent_code] = parent_name
        if parent_spec:
            material_specs[parent_code] = parent_spec
        if child_name and child_code not in material_names:
            material_names[child_code] = child_name
        if child_spec and child_code not in material_specs:
            material_specs[child_code] = child_spec

    detail_rows: list[dict[str, object]] = []

    def walk(
        root_code: str,
        root_name: str,
        customer: str,
        online_date: date | None,
        parent_code: str,
        demand_qty: float,
        trail: set[str],
    ) -> None:
        for edge in edges.get(parent_code, []):
            child_code = edge["child_code"]
            child_qty = demand_qty * float(edge["usage"])
            if child_code in assembly_codes and child_code not in scheduled_codes:
                detail_rows.append(
                    {
                        "客户": customer,
                        "建议母件料号": child_code,
                        "建议母件品名": material_names.get(child_code, ""),
                        "建议母件规格": material_specs.get(child_code, ""),
                        "建议上线日期": online_date,
                        "建议上线数量": child_qty,
                        "直接上层料号": parent_code,
                        "直接上层品名": material_names.get(parent_code, ""),
                        "顶层来源料号": root_code,
                        "顶层来源品名": root_name,
                    }
                )
            if child_code in assembly_codes and child_code not in trail:
                walk(
                    root_code,
                    root_name,
                    customer,
                    online_date,
                    child_code,
                    child_qty,
                    trail | {child_code},
                )

    for record in production_plan_df.to_dict("records"):
        root_code = str(record.get("母件料号", "")).strip()
        if not root_code:
            continue
        qty = float(record.get("上线数量", 0) or 0)
        if qty == 0:
            continue
        online_date = record.get("上线日期")
        if pd.isna(online_date):
            online_date = None
        root_name = str(record.get("母件品名", "") or "").strip()
        customer = str(record.get("客户", "") or "").strip()
        walk(root_code, root_name, customer, online_date, root_code, qty, {root_code})

    return suggestion_df, pd.DataFrame(detail_rows)


def build_suggestion_metadata_v2(details_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    metadata: dict[str, dict[str, str]] = {
        "parent_note_map": {},
        "child_note_map": {},
        "child_parent_codes_map": {},
    }
    if details_df.empty:
        return metadata

    parent_children_map: dict[str, set[str]] = defaultdict(set)
    child_parent_codes: dict[str, list[str]] = defaultdict(list)

    for record in details_df.to_dict("records"):
        child_code = str(record.get("建议母件料号", "")).strip()
        parent_code = str(record.get("直接上层料号", "")).strip()
        if not child_code or not parent_code:
            continue
        parent_children_map[parent_code].add(child_code)
        child_parent_codes[child_code].append(parent_code)

    metadata["parent_note_map"] = {
        parent_code: f"建议补排产，下层{len(child_codes)}颗物料未在表内"
        for parent_code, child_codes in parent_children_map.items()
        if parent_code and child_codes
    }
    metadata["child_note_map"] = {
        child_code: f"上层物料编码:{_summarize_codes_v2(parent_codes)}"
        for child_code, parent_codes in child_parent_codes.items()
        if child_code and _unique_codes_v2(parent_codes)
    }
    metadata["child_parent_codes_map"] = {
        child_code: _serialize_codes_v2(parent_codes)
        for child_code, parent_codes in child_parent_codes.items()
        if child_code and _unique_codes_v2(parent_codes)
    }
    return metadata


def build_production_plan_df(production_orders: pd.DataFrame, bom: pd.DataFrame) -> pd.DataFrame:
    customer_col = first_matching_column(production_orders, ["客户"], required=False)
    parent_item_col = first_matching_column(production_orders, ["母件料号"])
    parent_name_col = first_matching_column(production_orders, ["母件品名"])
    parent_spec_col = first_matching_column(production_orders, ["母件规格"], required=False)
    version_col = first_matching_column(production_orders, ["版本号"], required=False)
    online_date_col = first_matching_column(production_orders, ["上线日期"])
    quantity_col = first_matching_column(production_orders, ["上线数量"])

    result = pd.DataFrame(
        {
            "客户": production_orders[customer_col] if customer_col else "",
            "母件料号": production_orders[parent_item_col].astype(str).str.strip(),
            "母件品名": production_orders[parent_name_col],
            "母件规格": production_orders[parent_spec_col] if parent_spec_col else "",
            "版本号": production_orders[version_col] if version_col else "",
            "上线日期": coerce_date_series(production_orders[online_date_col]),
            "上线数量": coerce_number_series(production_orders[quantity_col]),
        }
    )
    result = result[result["母件料号"] != ""].copy()
    result["NO."] = range(1, len(result) + 1)

    bom_parent_col = first_matching_column(bom, ["母件料号"])
    bom_check = (
        bom[bom_parent_col]
        .astype(str)
        .str.strip()
        .value_counts(dropna=False)
        .rename_axis("母件料号")
        .reset_index(name="BOM校核")
    )
    result = result.merge(bom_check, on="母件料号", how="left")
    result["BOM校核"] = result["BOM校核"].fillna(0).astype(int)
    result = result[
        [
            "NO.",
            "客户",
            "母件料号",
            "母件品名",
            "母件规格",
            "版本号",
            "上线日期",
            "上线数量",
            "BOM校核",
        ]
    ]
    return result


def planning_month_key(value) -> tuple[int, int] | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return int(value.year), int(value.month)
    if isinstance(value, date):
        return int(value.year), int(value.month)
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return int(parsed.year), int(parsed.month)


def build_upper_expansion_check_df(
    production_plan_df: pd.DataFrame,
    bom: pd.DataFrame,
    usage_exclusion_pairs: set[tuple[str, str]] | None = None,
) -> pd.DataFrame:
    """Build a read-only check table for planned subassemblies covered by upper-level BOM."""
    columns = [
        "月份",
        "母件料号",
        "母件品名",
        "母件规格",
        "排产数量",
        "上层展开需求",
        "差异(排产-上层展开)",
        "状态",
        "上层来源",
    ]
    if production_plan_df.empty or bom.empty:
        return pd.DataFrame(columns=columns)

    bom_parent_col = first_matching_column(bom, ["母件料号"])
    child_item_col = first_matching_column(bom, ["BOM子项.子件料品.料号"])
    child_name_col = first_matching_column(bom, ["BOM子项.子件料品.料品名称"], required=False)
    child_spec_col = first_matching_column(bom, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom, ["BOM子项.子件用量"])

    bom_slim = pd.DataFrame(
        {
            "母件料号": bom[bom_parent_col].astype(str).str.strip(),
            "子件料号": bom[child_item_col].astype(str).str.strip(),
            "子件名称": bom[child_name_col] if child_name_col else "",
            "子件规格": bom[child_spec_col] if child_spec_col else "",
            "子件用量": coerce_number_series(bom[child_usage_col]),
        }
    )
    if usage_exclusion_pairs:
        pair_index = list(zip(bom_slim["母件料号"], bom_slim["子件料号"]))
        bom_slim = bom_slim.loc[[pair not in usage_exclusion_pairs for pair in pair_index]].copy()

    adjacency: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    child_info: dict[str, dict[str, object]] = {}
    for parent_code, child_code, child_name, child_spec, usage in bom_slim.itertuples(index=False, name=None):
        parent_code = str(parent_code).strip()
        child_code = str(child_code).strip()
        if not parent_code or not child_code:
            continue
        try:
            usage_qty = float(usage or 0)
        except (TypeError, ValueError):
            usage_qty = 0.0
        if usage_qty:
            adjacency[parent_code][child_code] += usage_qty
        if child_code and child_code not in child_info:
            child_info[child_code] = {
                "母件品名": child_name,
                "母件规格": child_spec,
            }

    if not adjacency:
        return pd.DataFrame(columns=columns)

    unit_cache: dict[str, dict[str, float]] = {}
    visiting: set[str] = set()

    def unit_totals(root_code: str) -> dict[str, float]:
        root_code = str(root_code).strip()
        if root_code in unit_cache:
            return unit_cache[root_code]
        if root_code in visiting:
            return {}
        visiting.add(root_code)
        totals: dict[str, float] = defaultdict(float)
        for child_code, usage_qty in adjacency.get(root_code, {}).items():
            totals[child_code] += usage_qty
            for grandchild_code, grandchild_usage in unit_totals(child_code).items():
                totals[grandchild_code] += usage_qty * grandchild_usage
        visiting.remove(root_code)
        unit_cache[root_code] = dict(totals)
        return unit_cache[root_code]

    grouped: dict[tuple[int, int] | None, list[dict[str, object]]] = defaultdict(list)
    for position, (index, row) in enumerate(production_plan_df.iterrows()):
        code = str(row["母件料号"]).strip()
        grouped[planning_month_key(row["上线日期"])].append(
            {
                "position": position,
                "index": index,
                "code": code,
                "qty": float(row["上线数量"] or 0),
                "name": row.get("母件品名", ""),
                "spec": row.get("母件规格", ""),
            }
        )

    result_rows: list[dict[str, object]] = []
    for _month, items in grouped.items():
        if _month is None:
            month_label = ""
        else:
            month_label = f"{int(_month[0])}-{int(_month[1]):02d}"
        planned_codes = {str(item["code"]) for item in items if item["code"]}
        plan_qty_by_code: dict[str, float] = defaultdict(float)
        first_item_by_code: dict[str, dict[str, object]] = {}
        source_by_code: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
        for item in items:
            code = str(item["code"])
            plan_qty_by_code[code] += float(item["qty"] or 0)
            first_item_by_code.setdefault(code, item)

        for item in items:
            parent_code = str(item["code"])
            if not parent_code:
                continue
            parent_qty = float(item["qty"] or 0)
            parent_name = str(item.get("name", "") or "")
            for child_code, unit_qty in unit_totals(parent_code).items():
                if child_code in planned_codes and child_code != parent_code:
                    source_label = f"{parent_code} {parent_name}".strip()
                    source_by_code[child_code][source_label] += parent_qty * float(unit_qty or 0)

        for code in sorted(planned_codes, key=lambda value: int(first_item_by_code[value]["position"])):
            plan_qty = float(plan_qty_by_code.get(code, 0.0))
            source_items = source_by_code.get(code, {})
            upper_qty = float(sum(source_items.values()))
            if upper_qty <= FLOAT_ABS_TOL:
                continue
            diff_qty = plan_qty - upper_qty
            if math.isclose(diff_qty, 0, abs_tol=FLOAT_ABS_TOL):
                status = "排产等于上层展开"
            elif diff_qty > 0:
                status = "排产大于上层展开"
            else:
                status = "排产小于上层展开"
            first_item = first_item_by_code[code]
            info = child_info.get(code, {})
            source_summary = "；".join(
                f"{label}:{qty:g}"
                for label, qty in sorted(source_items.items(), key=lambda item: item[1], reverse=True)[:20]
            )
            result_rows.append(
                {
                    "月份": month_label,
                    "母件料号": code,
                    "母件品名": first_item.get("name") or info.get("母件品名", ""),
                    "母件规格": first_item.get("spec") or info.get("母件规格", ""),
                    "排产数量": plan_qty,
                    "上层展开需求": upper_qty,
                    "差异(排产-上层展开)": diff_qty,
                    "状态": status,
                    "上层来源": source_summary,
                }
            )

    return pd.DataFrame(result_rows, columns=columns)


def build_balance_hierarchy_filter_df(
    production_plan_df: pd.DataFrame,
    bom: pd.DataFrame,
    usage_exclusion_pairs: set[tuple[str, str]] | None = None,
) -> pd.DataFrame:
    columns = ["料号", "匹配文本", "来源料号", "来源名称", "来源规格"]
    if production_plan_df.empty or bom.empty:
        return pd.DataFrame(columns=columns)

    bom_parent_col = first_matching_column(bom, ["母件料号"])
    bom_parent_name_col = first_matching_column(bom, ["母件品名"], required=False)
    bom_parent_spec_col = first_matching_column(bom, ["母件规格"], required=False)
    child_item_col = first_matching_column(bom, ["BOM子项.子件料品.料号"])
    child_name_col = first_matching_column(bom, ["BOM子项.子件料品.料品名称"], required=False)
    child_spec_col = first_matching_column(bom, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom, ["BOM子项.子件用量"])

    bom_slim = pd.DataFrame(
        {
            "母件料号": bom[bom_parent_col].astype(str).str.strip(),
            "母件品名": bom[bom_parent_name_col] if bom_parent_name_col else "",
            "母件规格": bom[bom_parent_spec_col] if bom_parent_spec_col else "",
            "子件料号": bom[child_item_col].astype(str).str.strip(),
            "子件名称": bom[child_name_col] if child_name_col else "",
            "子件规格": bom[child_spec_col] if child_spec_col else "",
            "子件用量": coerce_number_series(bom[child_usage_col]),
        }
    )
    bom_slim = bom_slim[
        (bom_slim["母件料号"] != "") & (bom_slim["子件料号"] != "") & (bom_slim["子件用量"] != 0)
    ].copy()
    if usage_exclusion_pairs:
        pair_index = list(zip(bom_slim["母件料号"], bom_slim["子件料号"]))
        bom_slim = bom_slim.loc[[pair not in usage_exclusion_pairs for pair in pair_index]].copy()

    def clean_text(value: object) -> str:
        if value is None or pd.isna(value):
            return ""
        text = str(value).strip()
        return "" if text.lower() == "nan" else text

    adjacency: dict[str, list[str]] = defaultdict(list)
    item_info: dict[str, dict[str, str]] = {}
    for row in bom_slim.to_dict("records"):
        parent_code = clean_text(row.get("母件料号"))
        child_code = clean_text(row.get("子件料号"))
        if not parent_code or not child_code:
            continue
        if child_code not in adjacency[parent_code]:
            adjacency[parent_code].append(child_code)
        parent_info = item_info.setdefault(parent_code, {"名称": "", "规格": ""})
        if not parent_info["名称"]:
            parent_info["名称"] = clean_text(row.get("母件品名"))
        if not parent_info["规格"]:
            parent_info["规格"] = clean_text(row.get("母件规格"))
        child_info = item_info.setdefault(child_code, {"名称": "", "规格": ""})
        if not child_info["名称"]:
            child_info["名称"] = clean_text(row.get("子件名称"))
        if not child_info["规格"]:
            child_info["规格"] = clean_text(row.get("子件规格"))

    root_rows = production_plan_df[["母件料号", "母件品名", "母件规格"]].drop_duplicates()
    for row in root_rows.to_dict("records"):
        root_code = clean_text(row.get("母件料号"))
        if not root_code:
            continue
        root_info = item_info.setdefault(root_code, {"名称": "", "规格": ""})
        if not root_info["名称"]:
            root_info["名称"] = clean_text(row.get("母件品名"))
        if not root_info["规格"]:
            root_info["规格"] = clean_text(row.get("母件规格"))

    def token_for(code: str) -> str:
        info = item_info.get(code, {})
        return " ".join(part for part in [clean_text(code), clean_text(info.get("名称")), clean_text(info.get("规格"))] if part)

    rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()

    def add_match(material_code: str, source_code: str) -> None:
        material_code = clean_text(material_code)
        source_code = clean_text(source_code)
        match_text = token_for(source_code)
        if not material_code or not match_text:
            return
        key = (material_code, source_code, match_text)
        if key in seen:
            return
        seen.add(key)
        source_info = item_info.get(source_code, {})
        rows.append(
            {
                "料号": material_code,
                "匹配文本": match_text,
                "来源料号": source_code,
                "来源名称": clean_text(source_info.get("名称")),
                "来源规格": clean_text(source_info.get("规格")),
            }
        )

    def walk(current_code: str, active_sources: list[str], trail: set[str]) -> None:
        current_code = clean_text(current_code)
        if not current_code or current_code in trail:
            return
        next_trail = set(trail)
        next_trail.add(current_code)
        sources = [*active_sources, current_code]
        for source_code in sources:
            add_match(current_code, source_code)
        for child_code in adjacency.get(current_code, []):
            walk(child_code, sources, next_trail)

    for root_code in root_rows["母件料号"].astype(str).str.strip().tolist():
        walk(root_code, [], set())

    return pd.DataFrame(rows, columns=columns)


def build_shortage_df(
    frames: dict[str, pd.DataFrame],
    substitute_rules: pd.DataFrame | None = None,
) -> pd.DataFrame:
    production_plan_df = build_production_plan_df(frames["生产订单"], frames["BOM"])
    bom = frames["BOM"].copy()

    bom_parent_col = first_matching_column(bom, ["母件料号"])
    child_item_col = first_matching_column(bom, ["BOM子项.子件料品.料号"])
    child_ref_col = first_matching_column(bom, ["BOM子项.子件料品.参考料号1"], required=False)
    child_name_col = first_matching_column(bom, ["BOM子项.子件料品.料品名称"])
    child_spec_col = first_matching_column(bom, ["BOM子项.子件料品.规格"], required=False)
    child_usage_col = first_matching_column(bom, ["BOM子项.子件用量"])

    bom_slim = pd.DataFrame(
        {
            "母件料号": bom[bom_parent_col].astype(str).str.strip(),
            "BOM2.BOM子项.子件料品.料号": bom[child_item_col].astype(str).str.strip(),
            "BOM2.BOM子项.子件料品.参考料号1": bom[child_ref_col] if child_ref_col else "",
            "BOM2.BOM子项.子件料品.料品名称": bom[child_name_col],
            "BOM2.BOM子项.子件料品.规格": bom[child_spec_col] if child_spec_col else "",
            "BOM2.BOM子项.子件用量": coerce_number_series(bom[child_usage_col]),
        }
    )
    usage_exclusion_pairs = parent_child_exclusion_pairs(
        frames.get(PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME)
    )
    if usage_exclusion_pairs:
        pair_index = list(
            zip(
                bom_slim["母件料号"],
                bom_slim["BOM2.BOM子项.子件料品.料号"],
            )
        )
        keep_mask = [pair not in usage_exclusion_pairs for pair in pair_index]
        bom_slim = bom_slim.loc[keep_mask].copy()

    source_name_by_planned_subassembly: dict[str, str] = {}
    planned_code_names = {
        str(row.get("母件料号", "") or "").strip(): str(row.get("母件品名", "") or "").strip()
        for row in production_plan_df.to_dict("records")
        if str(row.get("母件料号", "") or "").strip()
    }
    planned_codes = set(planned_code_names)
    if planned_codes:
        adjacency: dict[str, list[str]] = defaultdict(list)
        for parent_code, child_code in zip(
            bom_slim["母件料号"].astype(str).str.strip(),
            bom_slim["BOM2.BOM子项.子件料品.料号"].astype(str).str.strip(),
            strict=False,
        ):
            if parent_code and child_code and child_code not in adjacency[parent_code]:
                adjacency[parent_code].append(child_code)

        descendants_cache: dict[str, set[str]] = {}

        def descendants(root_code: str, trail: set[str] | None = None) -> set[str]:
            root_code = str(root_code).strip()
            if root_code in descendants_cache:
                return descendants_cache[root_code]
            trail = set(trail or set())
            if root_code in trail:
                return set()
            trail.add(root_code)
            result_codes: set[str] = set()
            for child_code in adjacency.get(root_code, []):
                result_codes.add(child_code)
                result_codes.update(descendants(child_code, trail))
            descendants_cache[root_code] = result_codes
            return result_codes

        source_names_by_child: dict[str, list[str]] = defaultdict(list)
        for root_code, root_name in planned_code_names.items():
            if not root_code.startswith("04"):
                continue
            for child_code in descendants(root_code):
                if child_code in planned_codes and child_code != root_code and not child_code.startswith("04"):
                    display_name = root_name or root_code
                    if display_name and display_name not in source_names_by_child[child_code]:
                        source_names_by_child[child_code].append(display_name)
        source_name_by_planned_subassembly = {
            child_code: "、".join(names[:8])
            for child_code, names in source_names_by_child.items()
            if names
        }

    shortage = production_plan_df.merge(bom_slim, on="母件料号", how="left")
    shortage = shortage[shortage["BOM2.BOM子项.子件料品.料号"].notna()].copy()
    shortage["上层来源品名"] = shortage["母件料号"].astype(str).str.strip().map(source_name_by_planned_subassembly).fillna("")
    shortage["需求"] = shortage["上线数量"] * shortage["BOM2.BOM子项.子件用量"]
    shortage["到货日期"] = shortage["上线日期"].apply(
        lambda value: value - timedelta(days=2) if isinstance(value, date) else pd.NaT
    )
    shipping_parts = frames.get(INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME)
    if shipping_parts is not None and not shipping_parts.empty:
        code_col = first_matching_column(shipping_parts, ["U9物料编号", "物料编号", "物料编码", "料号"], required=False)
        safety_col = first_matching_column(shipping_parts, ["安全库存数量", "安全库存"], contains=True, required=False)
        name_col = first_matching_column(shipping_parts, ["名称", "料品名称", "物料名称"], required=False)
        spec_col = first_matching_column(shipping_parts, ["型号", "规格"], required=False)
        remark_col = first_matching_column(shipping_parts, ["备注"], required=False)
        if code_col and safety_col:
            safety_work = pd.DataFrame(
                {
                    "物料编码": shipping_parts[code_col].astype(str).str.strip(),
                    "料品名称": shipping_parts[name_col] if name_col else "",
                    "规格": shipping_parts[spec_col] if spec_col else "",
                    "备注": shipping_parts[remark_col] if remark_col else "",
                    "安全库存": coerce_number_series(shipping_parts[safety_col]),
                }
            )
            safety_work = safety_work[
                (safety_work["物料编码"].astype(str).str.strip() != "")
                & (safety_work["安全库存"] > 0)
            ].copy()
            if not safety_work.empty:
                safety_work = (
                    safety_work.groupby("物料编码", dropna=False, sort=False)
                    .agg(
                        料品名称=("料品名称", "first"),
                        规格=("规格", "first"),
                        备注=("备注", lambda values: "；".join(dict.fromkeys(str(value).strip() for value in values if str(value).strip()))),
                        安全库存=("安全库存", "sum"),
                    )
                    .reset_index()
                )
                max_no = pd.to_numeric(shortage["NO."], errors="coerce").max()
                next_no = int(max_no) + 1 if not pd.isna(max_no) else 1
                safety_due_date = date.today()
                safety_rows: list[dict[str, object]] = []
                for offset, record in enumerate(safety_work.to_dict("records"), start=0):
                    material_code = str(record.get("物料编码", "") or "").strip()
                    safety_qty = float(record.get("安全库存", 0) or 0)
                    if not material_code or safety_qty <= 0:
                        continue
                    safety_rows.append(
                        {
                            "NO.": next_no + offset,
                            "客户": INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME,
                            "母件料号": "",
                            "母件品名": INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME,
                            "上层来源品名": "",
                            "母件规格": str(record.get("备注", "") or "").strip(),
                            "版本号": "",
                            "上线日期": safety_due_date,
                            "上线数量": 0,
                            "BOM校核": "安全库存",
                            "BOM2.BOM子项.子件料品.料号": material_code,
                            "BOM2.BOM子项.子件料品.参考料号1": "",
                            "BOM2.BOM子项.子件料品.料品名称": str(record.get("料品名称", "") or "").strip(),
                            "BOM2.BOM子项.子件料品.规格": str(record.get("规格", "") or "").strip(),
                            "BOM2.BOM子项.子件用量": 0,
                            "需求": safety_qty,
                            "到货日期": safety_due_date,
                        }
                    )
                if safety_rows:
                    shortage = pd.concat([shortage, pd.DataFrame(safety_rows)], ignore_index=True, sort=False)
    shortage = shortage.sort_values(
        by=["到货日期", "上线日期", "NO.", "BOM2.BOM子项.子件料品.料号"],
        kind="stable",
    ).reset_index(drop=True)
    shortage["累计需求"] = shortage.groupby("BOM2.BOM子项.子件料品.料号")["需求"].cumsum()

    inventory = frames["期初库存"].copy()
    inv_item_col = first_matching_column(inventory, ["物料编码"])
    inv_qty_col = first_matching_column(inventory, ["库存量"])
    inventory_map = (
        inventory.assign(
            物料编码=inventory[inv_item_col].astype(str).str.strip(),
            库存量=coerce_number_series(inventory[inv_qty_col]),
        )
        .groupby("物料编码", dropna=False)["库存量"]
        .sum()
    )
    substitute_code_map: dict[str, list[str]] = {}
    if substitute_rules is not None and not substitute_rules.empty:
        for before_code, group_df in substitute_rules.groupby("before_code", sort=False):
            matched_codes: list[str] = []
            for after_code in group_df.sort_values("sort_order")["after_code"].astype(str).str.strip():
                if not after_code or after_code in matched_codes:
                    continue
                matched_codes.append(after_code)
            before_code_text = str(before_code).strip()
            if before_code_text and matched_codes:
                substitute_code_map[before_code_text] = matched_codes

    def substitute_current_stock_qty(material_code: str) -> float:
        return float(
            sum(float(inventory_map.get(code, 0) or 0) for code in substitute_code_map.get(material_code, []))
        )

    def substitute_current_stock_detail(material_code: str) -> str:
        parts: list[str] = []
        for code in substitute_code_map.get(material_code, []):
            qty = float(inventory_map.get(code, 0) or 0)
            if qty:
                parts.append(f"{code}:{qty:g}")
        return "；".join(parts)

    purchase = frames["采购数据"].copy()
    purchase_item_col = first_matching_column(purchase, ["物料号"])
    supplier_col = first_matching_column(purchase, ["供应商"], required=False)
    buyer_col = first_matching_column(purchase, ["采购"], required=False)
    lt_col = first_matching_column(purchase, ["提前期"], contains=True, required=False)
    purchase_name_col = first_matching_column(purchase, ["名称"], required=False)
    purchase_spec_col = first_matching_column(purchase, ["规格"], required=False)
    purchase_info = purchase.assign(
        物料号=purchase[purchase_item_col].astype(str).str.strip(),
        供应商=purchase[supplier_col] if supplier_col else "",
        采购=purchase[buyer_col] if buyer_col else "",
        LT=coerce_number_series(purchase[lt_col]) if lt_col else 0,
        名称=purchase[purchase_name_col] if purchase_name_col else "",
        规格=purchase[purchase_spec_col] if purchase_spec_col else "",
    ).drop_duplicates(subset=["物料号"], keep="first")

    work_shortage = frames["期初工单缺料"].copy()
    work_item_col = first_matching_column(work_shortage, ["物料.编码"], required=False)
    work_qty_col = first_matching_column(work_shortage, ["缺料"], required=False)
    work_plan_col = first_matching_column(work_shortage, ["计划数量"], required=False)
    work_delivery_col = first_matching_column(work_shortage, ["配送数量"], required=False)
    work_parent_col = first_matching_column(
        work_shortage,
        ["在制品.编码", "生产订单.物料编码", "母件料号", "母料号"],
        required=False,
    )
    work_parent_name_col = first_matching_column(work_shortage, ["在制品.名称", "生产订单.物料名称", "母件品名", "母件名称"], required=False)
    work_parent_spec_col = first_matching_column(work_shortage, ["在制品.规格", "生产订单.物料规格", "母件规格"], required=False)
    work_item_name_col = first_matching_column(work_shortage, ["物料.名称", "物料名称", "名称"], required=False)
    work_item_spec_col = first_matching_column(work_shortage, ["物料.规格", "物料规格", "规格"], required=False)
    work_shortage_detail = pd.DataFrame()
    if work_item_col and work_qty_col:
        work_codes = work_shortage[work_item_col].astype(str).str.strip()
        explicit_shortage = pd.to_numeric(work_shortage[work_qty_col], errors="coerce")
        fallback_shortage = None
        if work_plan_col and work_delivery_col:
            fallback_shortage = (
                coerce_number_series(work_shortage[work_plan_col])
                - coerce_number_series(work_shortage[work_delivery_col])
            ).clip(lower=0)
        if fallback_shortage is not None:
            final_shortage = explicit_shortage.where(~explicit_shortage.isna(), fallback_shortage)
        else:
            final_shortage = explicit_shortage.fillna(0)
        work_shortage_map = (
            pd.DataFrame(
                {
                    "物料编码": work_codes,
                    "缺料": final_shortage.fillna(0),
                }
            )
            .groupby("物料编码", dropna=False)["缺料"]
            .sum()
        )
        work_shortage_detail = pd.DataFrame(
            {
                "物料编码": work_codes,
                "缺料": final_shortage.fillna(0),
                "母件料号": work_shortage[work_parent_col].astype(str).str.strip() if work_parent_col else "",
                "母件品名": work_shortage[work_parent_name_col] if work_parent_name_col else "",
                "母件规格": work_shortage[work_parent_spec_col] if work_parent_spec_col else "",
                "料品名称": work_shortage[work_item_name_col] if work_item_name_col else "",
                "规格": work_shortage[work_item_spec_col] if work_item_spec_col else "",
            }
        )
        work_shortage_detail = work_shortage_detail[
            (work_shortage_detail["物料编码"].astype(str).str.strip() != "")
            & (coerce_number_series(work_shortage_detail["缺料"]) > 0)
        ].copy()
        if not work_shortage_detail.empty:
            work_shortage_detail["缺料"] = coerce_number_series(work_shortage_detail["缺料"])
            work_shortage_detail = (
                work_shortage_detail.groupby("物料编码", dropna=False, sort=False)
                .agg(
                    缺料=("缺料", "sum"),
                    母件料号=("母件料号", lambda values: "、".join(dict.fromkeys(str(value).strip() for value in values if str(value).strip()))),
                    母件品名=("母件品名", "first"),
                    母件规格=("母件规格", "first"),
                    料品名称=("料品名称", "first"),
                    规格=("规格", "first"),
                )
                .reset_index()
            )
    else:
        work_shortage_map = pd.Series(dtype=float)

    shortage = shortage.merge(
        purchase_info[["物料号", "LT"]].rename(columns={"物料号": "BOM2.BOM子项.子件料品.料号"}),
        on="BOM2.BOM子项.子件料品.料号",
        how="left",
    )
    shortage["本体库存"] = shortage["BOM2.BOM子项.子件料品.料号"].map(inventory_map).fillna(0)
    shortage["替代料库存"] = shortage["BOM2.BOM子项.子件料品.料号"].map(substitute_current_stock_qty).fillna(0)
    shortage["替代料清单"] = shortage["BOM2.BOM子项.子件料品.料号"].map(substitute_current_stock_detail).fillna("")
    shortage["库存"] = shortage["本体库存"] + shortage["替代料库存"]
    # 累计需求超出库存的缺口（非负），库存含本体库存和替代料当前库存，用于决定每行应下单数量
    shortage["累计缺料2"] = (shortage["累计需求"] - shortage["库存"]).clip(lower=0)
    # 每行到货数量 = min(本行需求, 当前累计缺口)；按料号依 到货日期 排序后，逐行累加刚好等于 max(0, 累计需求-库存)
    shortage["到货数量"] = shortage[["需求", "累计缺料2"]].min(axis=1).clip(lower=0)
    # 累计到货：按料号累计的到货数量（用于报表显示，替换原先恒为 0 的占位）
    shortage["累计到货"] = shortage.groupby("BOM2.BOM子项.子件料品.料号")["到货数量"].cumsum()
    # 累计缺料：库存 + 已规划到货 - 累计需求（>=0 为盈余，<0 表示仍有缺口）
    shortage["累计缺料"] = shortage["库存"] + shortage["累计到货"] - shortage["累计需求"]
    shortage["L/T"] = shortage["LT"].fillna(0)

    def _safe_lt_days(value) -> int:
        try:
            return max(0, int(float(value)))
        except (TypeError, ValueError):
            return 0

    if not work_shortage_detail.empty:
        existing_codes = set(shortage["BOM2.BOM子项.子件料品.料号"].astype(str).str.strip())
        work_only_shortage = work_shortage_detail[
            ~work_shortage_detail["物料编码"].astype(str).str.strip().isin(existing_codes)
        ].copy()
        if not work_only_shortage.empty:
            valid_due_dates = [value for value in shortage["到货日期"].dropna().tolist() if isinstance(value, date)]
            first_due_date = min(valid_due_dates) if valid_due_dates else date.today()
            max_no = pd.to_numeric(shortage["NO."], errors="coerce").max()
            next_no = int(max_no) + 1 if not pd.isna(max_no) else 1
            purchase_lookup = purchase_info.set_index("物料号", drop=False)
            work_only_rows: list[dict[str, object]] = []
            def _clean_work_text(value: object) -> str:
                if value is None or pd.isna(value):
                    return ""
                text = str(value).strip()
                return "" if text.lower() == "nan" else text

            for offset, record in enumerate(work_only_shortage.to_dict("records"), start=0):
                material_code = _clean_work_text(record.get("物料编码", ""))
                if not material_code:
                    continue
                shortage_qty = float(record.get("缺料", 0) or 0)
                purchase_row = purchase_lookup.loc[material_code] if material_code in purchase_lookup.index else None
                lt_value = purchase_row["LT"] if purchase_row is not None else 0
                material_name = _clean_work_text(record.get("料品名称", ""))
                material_spec = _clean_work_text(record.get("规格", ""))
                if purchase_row is not None:
                    if not material_name:
                        material_name = str(purchase_row.get("名称", "") or "").strip()
                    if not material_spec:
                        material_spec = str(purchase_row.get("规格", "") or "").strip()
                own_inventory_qty = float(inventory_map.get(material_code, 0) or 0)
                substitute_inventory_qty = substitute_current_stock_qty(material_code)
                inventory_qty = own_inventory_qty + substitute_inventory_qty
                work_only_rows.append(
                    {
                        "NO.": next_no + offset,
                        "客户": "期初工单缺料",
                        "母件料号": _clean_work_text(record.get("母件料号", "")),
                        "母件品名": _clean_work_text(record.get("母件品名", "")),
                        "上层来源品名": "",
                        "母件规格": _clean_work_text(record.get("母件规格", "")),
                        "版本号": "",
                        "上线日期": first_due_date,
                        "上线数量": 0,
                        "BOM校核": "期初工单缺料",
                        "BOM2.BOM子项.子件料品.料号": material_code,
                        "BOM2.BOM子项.子件料品.参考料号1": "",
                        "BOM2.BOM子项.子件料品.料品名称": material_name,
                        "BOM2.BOM子项.子件料品.规格": material_spec,
                        "BOM2.BOM子项.子件用量": 0,
                        "需求": shortage_qty,
                        "累计需求": shortage_qty,
                        "本体库存": own_inventory_qty,
                        "替代料库存": substitute_inventory_qty,
                        "替代料清单": substitute_current_stock_detail(material_code),
                        "库存": inventory_qty,
                        "累计到货": 0,
                        "累计缺料": inventory_qty - shortage_qty,
                        "累计缺料2": max(shortage_qty - inventory_qty, 0),
                        "到货数量": 0,
                        "到货日期": first_due_date,
                        "请购日期": first_due_date - timedelta(days=_safe_lt_days(lt_value)),
                        "L/T": lt_value,
                        "流水号": 0,
                        "辅助列": 1,
                        "工单缺料": shortage_qty,
                    }
                )
            if work_only_rows:
                shortage = pd.concat([shortage, pd.DataFrame(work_only_rows)], ignore_index=True, sort=False)

    shortage["请购日期"] = shortage.apply(
        lambda row: row["到货日期"] - timedelta(days=_safe_lt_days(row["L/T"]))
        if isinstance(row["到货日期"], date)
        else pd.NaT,
        axis=1,
    )
    shortage["流水号"] = range(1, len(shortage) + 1)
    shortage["辅助列"] = shortage.groupby("NO.").cumcount() + 1
    shortage["工单缺料"] = shortage["BOM2.BOM子项.子件料品.料号"].map(work_shortage_map).fillna(0)
    shortage["辅助列2"] = shortage["NO."].astype(int).astype(str) + "-" + shortage["辅助列"].astype(str)

    shortage = shortage[SHORTAGE_COLUMNS].copy()
    shortage = shortage.rename(columns={"LT": "L/T"})
    return shortage


def date_window_from_shortage(shortage_df: pd.DataFrame) -> list[date]:
    valid_dates = sorted({value for value in shortage_df["到货日期"].dropna().tolist() if isinstance(value, date)})
    return valid_dates or [date.today()]


def planning_month_headers(date_headers: list[date]) -> list[date]:
    if not date_headers:
        today = date.today()
        return [date(today.year, today.month, 1)]
    month_headers = sorted({date(value.year, value.month, 1) for value in date_headers})
    return month_headers


def aggregate_map(df: pd.DataFrame, key_col: str, value_col: str) -> pd.Series:
    return (
        df.assign(
            _key=df[key_col].astype(str).str.strip(),
            _value=coerce_number_series(df[value_col]),
        )
        .groupby("_key", dropna=False)["_value"]
        .sum()
    )


def build_receiving_status_maps(frames: dict[str, pd.DataFrame]) -> dict[str, pd.Series]:
    receiving = frames.get(RECEIVING_STATUS_SHEET_NAME)
    empty = pd.Series(dtype=float)
    if receiving is None or receiving.empty:
        return {"接收": empty, "待检": empty, "待入库": empty}

    code_col = first_matching_column(receiving, ["物料编码"], required=False)
    status_col = first_matching_column(receiving, ["收货状态"], required=False)
    qty_col = first_matching_column(receiving, ["到货数量"], required=False)
    if not code_col or not status_col or not qty_col:
        return {"接收": empty, "待检": empty, "待入库": empty}

    reject_col = first_matching_column(receiving, ["拒收数量"], required=False)
    broken_col = first_matching_column(receiving, ["破坏数量"], required=False)
    work = pd.DataFrame(
        {
            "物料编码": receiving[code_col].astype(str).str.strip(),
            "收货状态": receiving[status_col].astype(str).str.strip(),
            "数量": coerce_number_series(receiving[qty_col]),
        }
    )
    if reject_col:
        work["数量"] -= coerce_number_series(receiving[reject_col])
    if broken_col:
        work["数量"] -= coerce_number_series(receiving[broken_col])
    work["数量"] = work["数量"].clip(lower=0)
    work = work[(work["物料编码"] != "") & (work["数量"] > 0)].copy()
    if work.empty:
        return {"接收": empty, "待检": empty, "待入库": empty}

    def bucket_status(status: object) -> str:
        text = str(status or "").strip()
        if "待入库" in text:
            return "待入库"
        if "待检" in text or "在检" in text:
            return "待检"
        if "接收" in text:
            return "接收"
        return ""

    work["状态分类"] = work["收货状态"].apply(bucket_status)
    work = work[work["状态分类"] != ""].copy()
    result: dict[str, pd.Series] = {}
    for bucket in ("接收", "待检", "待入库"):
        result[bucket] = (
            work[work["状态分类"] == bucket]
            .groupby("物料编码", dropna=False)["数量"]
            .sum()
        )
    return result


def build_near_term_production_shortage_df(
    shortage_df: pd.DataFrame,
    frames: dict[str, pd.DataFrame],
    *,
    analysis_date: date | None = None,
    window_days: int = 3,
) -> pd.DataFrame:
    today = analysis_date or date.today()
    window_days = max(int(window_days or 0), 1)
    window_dates = [today + timedelta(days=offset) for offset in range(window_days)]
    demand_columns = [f"{day.strftime('%Y-%m-%d')}需求" for day in window_dates]
    columns = [
        "物料编码",
        "料品名称",
        "规格",
        "供应商",
        "采购",
        "当前库存",
        "接收",
        "待检",
        "待入库",
        "可用合计",
        "近三天需求",
        "近三天缺口",
        *demand_columns,
        "最早上线日期",
        "影响母件料号",
        "影响母件规格",
        "状态",
    ]
    if shortage_df.empty:
        return pd.DataFrame(columns=columns)

    work = shortage_df.copy()
    work["上线日期"] = work["上线日期"].apply(coerce_excel_date)
    near_mask = work["上线日期"].apply(lambda value: isinstance(value, date) and value in set(window_dates))
    work = work.loc[near_mask].copy()
    if work.empty:
        return pd.DataFrame(columns=columns)

    material_col = "BOM2.BOM子项.子件料品.料号"
    name_col = "BOM2.BOM子项.子件料品.料品名称"
    spec_col = "BOM2.BOM子项.子件料品.规格"
    work["物料编码"] = work[material_col].astype(str).str.strip()
    work["需求"] = coerce_number_series(work["需求"])
    work = work[(work["物料编码"] != "") & (work["需求"] > 0)].copy()
    if work.empty:
        return pd.DataFrame(columns=columns)

    inventory_map = aggregate_map(
        frames["期初库存"],
        first_matching_column(frames["期初库存"], ["物料编码"]),
        first_matching_column(frames["期初库存"], ["库存量"]),
    )
    receiving_maps = build_receiving_status_maps(frames)

    purchase = frames["采购数据"].copy()
    purchase_item_col = first_matching_column(purchase, ["物料号"])
    supplier_col = first_matching_column(purchase, ["供应商"], required=False)
    buyer_col = first_matching_column(purchase, ["采购"], required=False)
    purchase_info = purchase.assign(
        物料编码=purchase[purchase_item_col].astype(str).str.strip(),
        供应商=purchase[supplier_col].astype(str).str.strip() if supplier_col else "",
        采购=purchase[buyer_col].astype(str).str.strip() if buyer_col else "",
    ).drop_duplicates(subset=["物料编码"], keep="first")
    supplier_map = purchase_info.set_index("物料编码")["供应商"] if not purchase_info.empty else pd.Series(dtype=str)
    buyer_map = purchase_info.set_index("物料编码")["采购"] if not purchase_info.empty else pd.Series(dtype=str)

    grouped = (
        work.groupby("物料编码", sort=True)
        .agg(
            料品名称=(name_col, "first"),
            规格=(spec_col, "first"),
            近三天需求=("需求", "sum"),
            最早上线日期=("上线日期", "min"),
            影响母件料号=("母件料号", lambda series: limited_unique_join(series, limit=8)),
            影响母件规格=("母件规格", lambda series: limited_unique_join(series, limit=8)),
        )
        .reset_index()
    )

    daily_summary = (
        work.groupby(["物料编码", "上线日期"], sort=False)["需求"]
        .sum()
        .reset_index()
    )
    daily_map = {
        (str(row["物料编码"]).strip(), row["上线日期"]): float(row["需求"] or 0)
        for row in daily_summary.to_dict("records")
    }
    for day, column_name in zip(window_dates, demand_columns, strict=False):
        grouped[column_name] = grouped["物料编码"].apply(
            lambda code, _day=day: daily_map.get((str(code).strip(), _day), 0.0)
        )

    grouped["供应商"] = grouped["物料编码"].map(supplier_map).fillna("")
    grouped["采购"] = grouped["物料编码"].map(buyer_map).fillna("")
    grouped["当前库存"] = grouped["物料编码"].map(inventory_map).fillna(0)
    grouped["接收"] = grouped["物料编码"].map(receiving_maps["接收"]).fillna(0)
    grouped["待检"] = grouped["物料编码"].map(receiving_maps["待检"]).fillna(0)
    grouped["待入库"] = grouped["物料编码"].map(receiving_maps["待入库"]).fillna(0)
    grouped["可用合计"] = grouped["当前库存"] + grouped["接收"] + grouped["待检"] + grouped["待入库"]
    grouped["近三天缺口"] = (grouped["近三天需求"] - grouped["可用合计"]).clip(lower=0)
    grouped["状态"] = grouped["近三天缺口"].apply(lambda value: "缺料" if float(value or 0) > 0 else "满足")
    grouped = grouped[grouped["近三天缺口"] > 0].copy()
    if grouped.empty:
        return pd.DataFrame(columns=columns)

    grouped = grouped.sort_values(
        by=["近三天缺口", "最早上线日期", "物料编码"],
        ascending=[False, True, True],
        kind="stable",
    ).reset_index(drop=True)
    return grouped[columns]


def build_config_supplement_plan_df(
    production_plan_df: pd.DataFrame,
    frames: dict[str, pd.DataFrame],
    *,
    analysis_date: date | None = None,
) -> pd.DataFrame:
    columns = [
        "状态",
        "补排类型",
        "补排母件料号",
        "补排母件品名",
        "补排母件规格",
        "建议上线日期",
        "建议数量",
        "来源本体料号",
        "来源本体规格",
        "配置表产品线",
        "配置项",
        "匹配说明",
    ]
    config = frames.get(INDUSTRIAL_CONFIG_SHEET_NAME)
    bom = frames.get("BOM")
    if config is None or config.empty or bom is None or bom.empty or production_plan_df.empty:
        return pd.DataFrame(columns=columns)

    def clean_text(value: object) -> str:
        if value is None or pd.isna(value):
            return ""
        text = str(value).strip()
        return "" if text.lower() == "nan" else text

    def norm_model(value: object) -> str:
        text = clean_text(value)
        text = text.replace("（", "(").replace("）", ")")
        text = re.sub(r"\s+", "", text)
        text = text.strip("，,。;；:：()（）")
        return text.lower()

    def option_number_text(value: object) -> set[str]:
        text = clean_text(value)
        return set(re.findall(r"配置\s*([0-9]+)", text, flags=re.IGNORECASE))

    model_pattern = re.compile(r"[A-Za-z0-9][A-Za-z0-9/._-]*[A-Za-z0-9]")
    target_pattern = re.compile(r"\b(?:XBC|MCP|TP)[A-Za-z0-9/._-]*\b", flags=re.IGNORECASE)

    def split_config_options(text: str) -> list[tuple[str, str]]:
        text = clean_text(text)
        if not text:
            return []
        matches = list(re.finditer(r"[（(]\s*([0-9]+)\s*[）)]", text))
        if not matches:
            return [("", text)]
        result = []
        for index, match in enumerate(matches):
            start = match.end()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
            result.append((match.group(1), text[start:end].strip()))
        return result

    def is_target_token(token: str) -> bool:
        token_norm = norm_model(token)
        return token_norm.startswith(("xbc", "mcp", "tp"))

    def supplement_type(target_spec: str) -> str:
        target_norm = norm_model(target_spec)
        if target_norm.startswith(("xbc", "mcp")) or "控制柜" in target_spec:
            return "控制柜"
        if target_norm.startswith("tp") or "示教" in target_spec:
            return "示教器"
        return "配置件"

    def parse_pairs_from_text(text: str) -> list[dict[str, str]]:
        text = clean_text(text)
        if not text:
            return []
        pairs: list[dict[str, str]] = []
        for target_match in target_pattern.finditer(text):
            target_spec = target_match.group(0).strip("，,。;；:：()（）")
            prefix = text[:target_match.start()]
            body_tokens = [
                token.strip("，,。;；:：()（）")
                for token in model_pattern.findall(prefix)
                if "-" in token or "/" in token
            ]
            body_tokens = [token for token in body_tokens if not is_target_token(token)]
            body_spec = body_tokens[-1] if body_tokens else ""
            pairs.append(
                {
                    "body_spec": body_spec,
                    "target_spec": target_spec,
                    "config_item": text,
                }
            )
        return pairs

    default_month_label = f"{(analysis_date or date.today()).month}月"
    config_columns = list(config.columns)
    line_col = first_matching_column(config, ["工业产品线"], required=False)
    detail_col = first_matching_column(config, ["具体配置"], required=False)
    if detail_col is None:
        return pd.DataFrame(columns=columns)

    month_labels = [
        clean_text(column)
        for column in config_columns
        if re.fullmatch(r"[0-9]{1,2}月", clean_text(column))
    ]
    if default_month_label not in month_labels:
        month_labels.append(default_month_label)

    def parse_config_pairs_for_month(month_label: str) -> list[dict[str, str]]:
        month_col = first_matching_column(config, [month_label], required=False)
        month_pairs: list[dict[str, str]] = []
        for record in config.to_dict("records"):
            product_line = clean_text(record.get(line_col, "")) if line_col else ""
            detail_text = clean_text(record.get(detail_col, ""))
            month_text = clean_text(record.get(month_col, "")) if month_col else ""
            selected_numbers = option_number_text(month_text)
            option_texts = split_config_options(detail_text)
            selected_option_texts = [
                option_text
                for option_number, option_text in option_texts
                if not selected_numbers or option_number in selected_numbers
            ]
            parse_sources = selected_option_texts if selected_option_texts else [detail_text]
            if month_text and not selected_numbers:
                parse_sources.append(month_text)
            for source_text in parse_sources:
                for pair in parse_pairs_from_text(source_text):
                    if not pair["body_spec"] or not pair["target_spec"]:
                        continue
                    pair["product_line"] = product_line
                    pair["month_label"] = month_label
                    month_pairs.append(pair)
        return month_pairs

    config_pairs_by_month = {
        month_label: parse_config_pairs_for_month(month_label)
        for month_label in month_labels
    }
    config_pairs = [
        pair
        for month_pairs in config_pairs_by_month.values()
        for pair in month_pairs
    ]
    if not config_pairs:
        return pd.DataFrame(columns=columns)

    bom_parent_col = first_matching_column(bom, ["母件料号"])
    bom_parent_name_col = first_matching_column(bom, ["母件品名"], required=False)
    bom_parent_spec_col = first_matching_column(bom, ["母件规格"], required=False)
    bom_parent_rows = (
        pd.DataFrame(
            {
                "母件料号": bom[bom_parent_col].astype(str).str.strip(),
                "母件品名": bom[bom_parent_name_col] if bom_parent_name_col else "",
                "母件规格": bom[bom_parent_spec_col] if bom_parent_spec_col else "",
            }
        )
        .drop_duplicates(subset=["母件料号"], keep="first")
    )
    bom_by_spec: dict[str, list[dict[str, str]]] = defaultdict(list)
    for record in bom_parent_rows.to_dict("records"):
        spec = clean_text(record.get("母件规格", ""))
        code = clean_text(record.get("母件料号", ""))
        if not spec or not code:
            continue
        bom_by_spec[norm_model(spec)].append(
            {
                "code": code,
                "name": clean_text(record.get("母件品名", "")),
                "spec": spec,
            }
        )

    def pick_preferred_bom_parent(matches: list[dict[str, str]]) -> tuple[dict[str, str] | None, str]:
        if not matches:
            return None, ""
        preferred_04 = [match for match in matches if match["code"].startswith("04")]
        if len(preferred_04) == 1:
            return preferred_04[0], "优先取04层级母件"
        preferred_02 = [match for match in matches if match["code"].startswith("02")]
        if len(preferred_02) == 1 and not preferred_04:
            return preferred_02[0], "优先取02层级母件"
        if len(matches) == 1:
            return matches[0], ""
        return None, ""

    def find_bom_parent_by_spec(target_spec: str) -> tuple[dict[str, str] | None, str]:
        target_norm = norm_model(target_spec)
        if not target_norm:
            return None, "配置件规格为空"
        exact_matches = bom_by_spec.get(target_norm, [])
        preferred_parent, preferred_note = pick_preferred_bom_parent(exact_matches)
        if preferred_parent is not None:
            return preferred_parent, f"BOM母件规格精确匹配{('；' + preferred_note) if preferred_note else ''}"
        if exact_matches:
            sample = "、".join(match["code"] for match in exact_matches[:5])
            return None, f"BOM母件规格匹配多个料号: {sample}"
        contains_matches = [
            match
            for spec_norm, matches in bom_by_spec.items()
            if target_norm in spec_norm or spec_norm in target_norm
            for match in matches
        ]
        preferred_parent, preferred_note = pick_preferred_bom_parent(contains_matches)
        if preferred_parent is not None:
            return preferred_parent, f"BOM母件规格模糊匹配{('；' + preferred_note) if preferred_note else ''}"
        if contains_matches:
            sample = "、".join(match["code"] for match in contains_matches[:5])
            return None, f"BOM母件规格模糊匹配多个料号: {sample}"
        return None, "BOM中未找到该配置件规格"

    pair_by_body_spec: dict[str, list[dict[str, str]]] = defaultdict(list)
    for pair in config_pairs:
        pair_by_body_spec[norm_model(pair["body_spec"])].append(pair)
    pair_by_month_body_spec: dict[str, dict[str, list[dict[str, str]]]] = {}
    for month_label, month_pairs in config_pairs_by_month.items():
        month_map: dict[str, list[dict[str, str]]] = defaultdict(list)
        for pair in month_pairs:
            month_map[norm_model(pair["body_spec"])].append(pair)
        pair_by_month_body_spec[month_label] = month_map

    def month_label_from_value(value: object) -> str:
        parsed = coerce_excel_date(value)
        if parsed is None:
            return default_month_label
        return f"{parsed.month}月"

    bom_child_col = first_matching_column(bom, ["BOM子项.子件料品.料号"])
    bom_child_name_col = first_matching_column(bom, ["BOM子项.子件料品.料品名称"], required=False)
    bom_child_spec_col = first_matching_column(bom, ["BOM子项.子件料品.规格"], required=False)
    bom_child_usage_col = first_matching_column(bom, ["BOM子项.子件用量"], required=False)
    adjacency: dict[str, list[dict[str, object]]] = defaultdict(list)
    parents_by_child: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in bom.to_dict("records"):
        parent_code = clean_text(record.get(bom_parent_col))
        parent_spec = clean_text(record.get(bom_parent_spec_col)) if bom_parent_spec_col else ""
        child_code = clean_text(record.get(bom_child_col))
        if not parent_code or not child_code:
            continue
        try:
            usage_qty = float(record.get(bom_child_usage_col, 1) or 1) if bom_child_usage_col else 1.0
        except (TypeError, ValueError):
            usage_qty = 1.0
        adjacency[parent_code].append(
            {
                "code": child_code,
                "name": clean_text(record.get(bom_child_name_col)) if bom_child_name_col else "",
                "spec": clean_text(record.get(bom_child_spec_col)) if bom_child_spec_col else "",
                "usage": usage_qty,
            }
        )
        parents_by_child[child_code].append(
            {
                "code": parent_code,
                "spec": parent_spec,
                "usage": usage_qty,
            }
        )

    def find_body_matches(root_code: str, root_spec: str) -> list[dict[str, object]]:
        matches: list[dict[str, object]] = []
        seen: set[tuple[str, str]] = set()
        seen_specs: set[str] = set()

        def add_match(code: str, spec: str, multiplier: float, note: str) -> None:
            spec_norm = norm_model(spec)
            if not spec_norm or spec_norm not in pair_by_body_spec:
                return
            if spec_norm in seen_specs:
                return
            key = (clean_text(code), spec_norm)
            if key in seen:
                return
            seen.add(key)
            seen_specs.add(spec_norm)
            matches.append(
                {
                    "code": clean_text(code),
                    "spec": clean_text(spec),
                    "multiplier": float(multiplier or 0),
                    "note": note,
                }
            )

        add_match(root_code, root_spec, 1.0, "生产订单母件规格匹配配置表")

        visiting: set[str] = set()

        def walk(parent_code: str, multiplier: float, depth: int) -> None:
            parent_code = clean_text(parent_code)
            if not parent_code or depth > 8 or parent_code in visiting:
                return
            visiting.add(parent_code)
            for child in adjacency.get(parent_code, []):
                child_code = clean_text(child.get("code"))
                child_spec = clean_text(child.get("spec"))
                child_multiplier = multiplier * float(child.get("usage", 1) or 0)
                add_match(child_code, child_spec, child_multiplier, "BOM下层本体规格匹配配置表")
                if child_code in adjacency:
                    walk(child_code, child_multiplier, depth + 1)
            visiting.remove(parent_code)

        walk(root_code, 1.0, 1)
        if clean_text(root_code).startswith("04"):
            ancestor_visiting: set[str] = set()

            def walk_ancestors(child_code: str, multiplier: float, depth: int) -> None:
                child_code = clean_text(child_code)
                if not child_code or depth > 4 or child_code in ancestor_visiting:
                    return
                ancestor_visiting.add(child_code)
                for parent in parents_by_child.get(child_code, []):
                    parent_code = clean_text(parent.get("code"))
                    parent_spec = clean_text(parent.get("spec"))
                    usage = float(parent.get("usage", 1) or 1)
                    parent_multiplier = multiplier / usage if usage else 0
                    add_match(parent_code, parent_spec, parent_multiplier, "BOM上层本体规格匹配配置表")
                    walk_ancestors(parent_code, parent_multiplier, depth + 1)
                ancestor_visiting.remove(child_code)

            walk_ancestors(root_code, 1.0, 1)
        return matches

    rows: list[dict[str, object]] = []
    grouped_orders = (
        production_plan_df.copy()
        .assign(
            _body_spec=production_plan_df["母件规格"].apply(norm_model),
            _qty=coerce_number_series(production_plan_df["上线数量"]),
        )
        .groupby(["母件料号", "母件规格", "上线日期", "_body_spec"], dropna=False, sort=False)["_qty"]
        .sum()
        .reset_index()
    )
    for order in grouped_orders.to_dict("records"):
        order_month_label = month_label_from_value(order.get("上线日期"))
        order_pair_by_body_spec = (
            pair_by_month_body_spec.get(order_month_label)
            or pair_by_month_body_spec.get(default_month_label)
            or pair_by_body_spec
        )
        body_matches = find_body_matches(
            clean_text(order.get("母件料号")),
            clean_text(order.get("母件规格")),
        )
        for body_match in body_matches:
            body_spec_norm = norm_model(body_match.get("spec"))
            for pair in order_pair_by_body_spec.get(body_spec_norm, []):
                parent, match_note = find_bom_parent_by_spec(pair["target_spec"])
                status = "可生成" if parent else "待确认"
                order_qty = float(order.get("_qty", 0) or 0)
                multiplier = float(body_match.get("multiplier", 0) or 0)
                rows.append(
                    {
                        "状态": status,
                        "补排类型": supplement_type(pair["target_spec"]),
                        "补排母件料号": parent["code"] if parent else "",
                        "补排母件品名": parent["name"] if parent else "",
                        "补排母件规格": parent["spec"] if parent else pair["target_spec"],
                        "建议上线日期": order.get("上线日期"),
                        "建议数量": order_qty * multiplier,
                        "来源本体料号": clean_text(body_match.get("code")),
                        "来源本体规格": clean_text(body_match.get("spec")),
                        "配置表产品线": clean_text(pair.get("product_line")),
                        "配置项": clean_text(pair.get("config_item")),
                        "匹配说明": f"{clean_text(body_match.get('note'))}；{match_note}",
                    }
                )

    if not rows:
        return pd.DataFrame(columns=columns)
    result = pd.DataFrame(rows, columns=columns)
    result = (
        result.groupby(
            [
                "状态",
                "补排类型",
                "补排母件料号",
                "补排母件品名",
                "补排母件规格",
                "建议上线日期",
                "配置表产品线",
                "配置项",
                "匹配说明",
            ],
            dropna=False,
            sort=False,
        )
        .agg(
            建议数量=("建议数量", "sum"),
            来源本体料号=("来源本体料号", lambda series: limited_unique_join(series, limit=10)),
            来源本体规格=("来源本体规格", lambda series: limited_unique_join(series, limit=10)),
        )
        .reset_index()
    )
    result = result[
        [
            "状态",
            "补排类型",
            "补排母件料号",
            "补排母件品名",
            "补排母件规格",
            "建议上线日期",
            "建议数量",
            "来源本体料号",
            "来源本体规格",
            "配置表产品线",
            "配置项",
            "匹配说明",
        ]
    ]
    return result.sort_values(
        by=["状态", "建议上线日期", "补排类型", "补排母件规格"],
        ascending=[True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)


def build_purchase_view_df(
    shortage_df: pd.DataFrame,
    frames: dict[str, pd.DataFrame],
    substitute_rules: pd.DataFrame | None = None,
    note_map: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, list[date]]:
    purchase = frames["采购数据"].copy()
    purchase_item_col = first_matching_column(purchase, ["物料号"])
    supplier_col = first_matching_column(purchase, ["供应商"], required=False)
    buyer_col = first_matching_column(purchase, ["采购"], required=False)
    purchase_name_col = first_matching_column(purchase, ["名称"], required=False)
    purchase_spec_col = first_matching_column(purchase, ["规格"], required=False)
    purchase_info = purchase.assign(
        料号=purchase[purchase_item_col].astype(str).str.strip(),
        供应商=purchase[supplier_col] if supplier_col else "",
        采购=purchase[buyer_col] if buyer_col else "",
        料品名称=purchase[purchase_name_col] if purchase_name_col else "",
        规格=purchase[purchase_spec_col] if purchase_spec_col else "",
    ).drop_duplicates(subset=["料号"], keep="first")

    inbound_po_map = aggregate_map(
        frames["在途采购"],
        first_matching_column(frames["在途采购"], ["料号"]),
        first_matching_column(frames["在途采购"], ["欠交数量"], contains=True),
    )
    inbound_pr_map = aggregate_map(
        frames["在途请购"],
        first_matching_column(frames["在途请购"], ["料号"]),
        first_matching_column(frames["在途请购"], ["未转PO数量"], contains=True),
    )
    inventory_map = aggregate_map(
        frames["期初库存"],
        first_matching_column(frames["期初库存"], ["物料编码"]),
        first_matching_column(frames["期初库存"], ["库存量"]),
    )
    supplier_inventory_key_col = first_matching_column(frames["供应商库存"], ["物料码"], required=False)
    if not supplier_inventory_key_col:
        supplier_inventory_key_col = first_matching_column(frames["供应商库存"], ["物料编码"], required=False)
    supplier_inventory_value_col = first_matching_column(frames["供应商库存"], ["结余库存"], contains=True, required=False)
    if supplier_inventory_key_col and supplier_inventory_value_col:
        supplier_inventory_map = aggregate_map(
            frames["供应商库存"],
            supplier_inventory_key_col,
            supplier_inventory_value_col,
        )
    else:
        supplier_inventory_map = pd.Series(dtype=float)

    grouped = shortage_df.groupby("BOM2.BOM子项.子件料品.料号", sort=True)
    summary = grouped.agg(
        料品名称=("BOM2.BOM子项.子件料品.料品名称", "first"),
        规格=("BOM2.BOM子项.子件料品.规格", "first"),
        总需求=("到货数量", "sum"),
        工单缺料=("工单缺料", "max"),
    ).reset_index()
    summary = summary.rename(columns={"BOM2.BOM子项.子件料品.料号": "料号"})
    summary = summary.merge(purchase_info, on="料号", how="left", suffixes=("", "_采购"))
    summary["料品名称"] = summary["料品名称"].where(summary["料品名称"].notna(), summary["料品名称_采购"])
    summary["规格"] = summary["规格"].where(summary["规格"].notna(), summary["规格_采购"])
    summary = summary.drop(columns=[column for column in summary.columns if column.endswith("_采购")])

    summary["未清PO"] = summary["料号"].map(inbound_po_map).fillna(0)
    summary["未转"] = summary["料号"].map(inbound_pr_map).fillna(0)
    summary["供应商库存"] = summary["料号"].map(supplier_inventory_map).fillna(0)
    summary["实时库存"] = summary["料号"].map(inventory_map).fillna(0)
    summary["替代1"] = ""
    summary["替代1库存"] = 0
    summary["替代2"] = ""
    summary["替代2库存"] = 0
    summary["_替代1库存批注"] = ""
    summary["_替代2库存批注"] = ""
    summary["_替代2批注"] = ""
    summary["_替代未清未转合计"] = 0
    summary["_替代料清单"] = ""
    summary["项目"] = ""
    summary["项目需求"] = 0
    summary["项目短缺"] = 0
    summary["单套用量"] = 0
    summary["项目风险"] = ""
    summary["备注"] = ""

    note_map = {str(code).strip(): str(note).strip() for code, note in (note_map or {}).items() if str(code).strip()}
    if note_map:
        summary["备注"] = summary["料号"].map(note_map).fillna(summary["备注"])

    if substitute_rules is not None and not substitute_rules.empty:
        substitute_rules = substitute_rules.copy()

        def substitute_stock_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inventory_map.get(substitute_code, 0) or 0)

        def substitute_open_po_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inbound_po_map.get(substitute_code, 0) or 0)

        def substitute_unconverted_qty(substitute_code: str) -> float:
            if not substitute_code:
                return 0.0
            return float(inbound_pr_map.get(substitute_code, 0) or 0)

        def substitute_available_qty(substitute_code: str) -> float:
            return (
                substitute_stock_qty(substitute_code)
                + substitute_open_po_qty(substitute_code)
                + substitute_unconverted_qty(substitute_code)
            )

        def format_qty(value: float) -> str:
            return f"{value:g}"

        def build_substitute_inventory_comment(codes: list[str]) -> str:
            if not codes:
                return ""
            total_stock = sum(substitute_stock_qty(code) for code in codes)
            total_open_po = sum(substitute_open_po_qty(code) for code in codes)
            total_unconverted = sum(substitute_unconverted_qty(code) for code in codes)
            if not any([total_stock, total_open_po, total_unconverted]) and len(codes) == 1:
                return ""
            lines = [
                f"当前库存：{format_qty(total_stock)}",
                f"未清PO：{format_qty(total_open_po)}",
                f"未转：{format_qty(total_unconverted)}",
                f"可用合计：{format_qty(total_stock + total_open_po + total_unconverted)}",
            ]
            if len(codes) > 1:
                lines.append("分项：")
                for code in codes:
                    stock_qty = substitute_stock_qty(code)
                    open_po_qty = substitute_open_po_qty(code)
                    unconverted_qty = substitute_unconverted_qty(code)
                    total_qty = stock_qty + open_po_qty + unconverted_qty
                    lines.append(
                        f"{code} 库存{format_qty(stock_qty)} 未清PO{format_qty(open_po_qty)} "
                        f"未转{format_qty(unconverted_qty)} 合计{format_qty(total_qty)}"
                    )
            return "\n".join(lines)

        def resolve_substitutes(material_code: str) -> tuple[str, float, str, float, str, str, str, float, str]:
            matched_codes = []
            candidate_rows = substitute_rules[substitute_rules["before_code"] == material_code].sort_values("sort_order")

            for row in candidate_rows.itertuples(index=False):
                if row.after_code not in matched_codes:
                    matched_codes.append(row.after_code)

            first_code = matched_codes[0] if len(matched_codes) >= 1 else ""
            second_code = matched_codes[1] if len(matched_codes) >= 2 else ""
            first_inventory = substitute_stock_qty(first_code)
            remaining_codes = matched_codes[1:]
            second_inventory = sum(substitute_stock_qty(code) for code in remaining_codes)
            first_comment = build_substitute_inventory_comment([first_code] if first_code else [])
            second_inventory_comment = build_substitute_inventory_comment(remaining_codes)
            substitute_inbound_total = sum(
                substitute_open_po_qty(code) + substitute_unconverted_qty(code)
                for code in matched_codes
            )
            second_comment = ""
            if len(matched_codes) > 2:
                extra_codes = matched_codes[2:]
                second_comment = (
                    f"替代料共{len(matched_codes)}款；替代2库存已包含第2款及后续替代料当前库存合计。\n"
                    f"第3款及以后：{_summarize_codes_v2(extra_codes, limit=20)}"
                )
            return (
                first_code,
                first_inventory,
                second_code,
                second_inventory,
                second_comment,
                first_comment,
                second_inventory_comment,
                substitute_inbound_total,
                "|".join(matched_codes),
            )

        substitute_values = summary["料号"].apply(resolve_substitutes)
        summary["替代1"] = substitute_values.apply(lambda item: item[0])
        summary["替代1库存"] = substitute_values.apply(lambda item: item[1])
        summary["替代2"] = substitute_values.apply(lambda item: item[2])
        summary["替代2库存"] = substitute_values.apply(lambda item: item[3])
        summary["_替代2批注"] = substitute_values.apply(lambda item: item[4])
        summary["_替代1库存批注"] = substitute_values.apply(lambda item: item[5])
        summary["_替代2库存批注"] = substitute_values.apply(lambda item: item[6])
        summary["_替代未清未转合计"] = substitute_values.apply(lambda item: item[7])
        summary["_替代料清单"] = substitute_values.apply(lambda item: item[8])

    date_headers = date_window_from_shortage(shortage_df)
    daily_summary = (
        shortage_df.groupby(["BOM2.BOM子项.子件料品.料号", "到货日期"], dropna=False)["到货数量"]
        .sum()
        .reset_index()
    )
    for header_date in date_headers:
        column_name = header_date
        values = daily_summary[daily_summary["到货日期"] == header_date][
            ["BOM2.BOM子项.子件料品.料号", "到货数量"]
        ].rename(columns={"BOM2.BOM子项.子件料品.料号": "料号", "到货数量": column_name})
        summary = summary.merge(values, on="料号", how="left")
        summary[column_name] = summary[column_name].fillna(0)

    month_headers = planning_month_headers(date_headers)
    month_columns: list[str] = []
    for month_index, month_header in enumerate(month_headers):
        matching_dates = [
            header
            for header in date_headers
            if (header.year, header.month) == (month_header.year, month_header.month)
        ]
        month_label = f"{month_header.month}月需求"
        month_values = summary[matching_dates].sum(axis=1) if matching_dates else pd.Series(0, index=summary.index)
        if month_index == 0:
            month_values = month_values + summary["工单缺料"]
        summary[month_label] = month_values
        month_columns.append(month_label)
    summary["总需求"] = summary[month_columns].sum(axis=1) if month_columns else 0
    summary["停线预警"] = (
        summary["总需求"]
        > summary["实时库存"] + summary["未清PO"] + summary["未转"] + summary["替代1库存"] + summary["替代2库存"]
        + summary["_替代未清未转合计"]
    ).astype(int)
    summary = summary.sort_values(by=["料号"], kind="stable").reset_index(drop=True)
    summary["序号"] = range(1, len(summary) + 1)
    summary["辅助列"] = ""
    summary["父项分类2"] = ""

    result_columns = [
        "辅助列",
        "序号",
        "料号",
        "料品名称",
        "规格",
        "供应商",
        "采购",
        "未清PO",
        "未转",
        "替代1",
        "替代1库存",
        "替代2",
        "替代2库存",
        "项目",
        "项目需求",
        "项目短缺",
        "单套用量",
        "项目风险",
        "供应商库存",
        "实时库存",
        "总需求",
        "工单缺料",
        *month_columns,
        "备注",
        "停线预警",
        "父项分类2",
        "_替代1库存批注",
        "_替代2库存批注",
        "_替代2批注",
        "_替代未清未转合计",
        "_替代料清单",
        *date_headers,
    ]
    result = summary[result_columns].copy()
    return result, date_headers


def build_purchase_view_df_v2(
    shortage_df: pd.DataFrame,
    frames: dict[str, pd.DataFrame],
    substitute_rules: pd.DataFrame | None = None,
    suggestion_metadata: dict[str, dict[str, str]] | None = None,
) -> tuple[pd.DataFrame, list[date]]:
    purchase_view_df, date_headers = build_purchase_view_df(
        shortage_df,
        frames,
        substitute_rules,
        note_map=None,
    )

    code_col = "料号"
    note_col = "备注"
    suggestion_metadata = suggestion_metadata or {}
    parent_note_map = {
        str(code).strip(): str(note).strip()
        for code, note in suggestion_metadata.get("parent_note_map", {}).items()
        if str(code).strip()
    }
    child_note_map = {
        str(code).strip(): str(note).strip()
        for code, note in suggestion_metadata.get("child_note_map", {}).items()
        if str(code).strip()
    }
    child_parent_codes_map = {
        str(code).strip(): str(value).strip()
        for code, value in suggestion_metadata.get("child_parent_codes_map", {}).items()
        if str(code).strip()
    }

    purchase_view_df[note_col] = purchase_view_df[note_col].fillna("")
    if parent_note_map:
        parent_notes = purchase_view_df[code_col].astype(str).str.strip().map(parent_note_map).fillna("")
        purchase_view_df[note_col] = [
            _merge_note_texts_v2(base_note, parent_note)
            for base_note, parent_note in zip(purchase_view_df[note_col], parent_notes, strict=False)
        ]
    if child_note_map:
        child_notes = purchase_view_df[code_col].astype(str).str.strip().map(child_note_map).fillna("")
        purchase_view_df[note_col] = [
            _merge_note_texts_v2(base_note, child_note)
            for base_note, child_note in zip(purchase_view_df[note_col], child_notes, strict=False)
        ]

    purchase_view_df["_suggest_parent_codes"] = (
        purchase_view_df[code_col].astype(str).str.strip().map(child_parent_codes_map).fillna("")
    )
    purchase_view_df["_base_rank"] = range(len(purchase_view_df))
    return purchase_view_df, date_headers


def apply_carried_balance_remarks(purchase_view_df: pd.DataFrame, remark_map: dict[str, str]) -> pd.DataFrame:
    if purchase_view_df.empty or not remark_map:
        return purchase_view_df
    result = purchase_view_df.copy()
    if "备注" not in result.columns:
        result["备注"] = ""
    carried_notes = result["料号"].astype(str).str.strip().map(remark_map).fillna("")
    result["备注"] = [
        _merge_note_texts_v2(base_note, carried_note)
        for base_note, carried_note in zip(result["备注"].fillna(""), carried_notes, strict=False)
    ]
    return result


def reorder_purchase_rows_by_suggestion_v2(purchase_view_df: pd.DataFrame) -> pd.DataFrame:
    if purchase_view_df.empty or "_suggest_parent_codes" not in purchase_view_df.columns:
        return purchase_view_df.copy()

    result = purchase_view_df.copy()
    code_col = result.columns[2]
    code_list = result[code_col].astype(str).str.strip().tolist()
    base_rank_map = {code: idx for idx, code in enumerate(code_list)}
    code_set = set(code_list)
    children_by_parent: dict[str, list[str]] = defaultdict(list)
    anchored_codes: set[str] = set()

    for record in result.to_dict("records"):
        child_code = str(record.get(code_col, "")).strip()
        parent_codes = [
            code
            for code in _deserialize_codes_v2(record.get("_suggest_parent_codes", ""))
            if code in code_set and code != child_code
        ]
        if not child_code or not parent_codes:
            continue
        anchor_code = min(parent_codes, key=lambda code: base_rank_map.get(code, len(base_rank_map)))
        children_by_parent[anchor_code].append(child_code)
        anchored_codes.add(child_code)

    for parent_code, child_codes in children_by_parent.items():
        child_codes.sort(key=lambda code: base_rank_map.get(code, len(base_rank_map)))

    ordered_codes: list[str] = []
    visited: set[str] = set()

    def append_code(code: str) -> None:
        if not code or code in visited:
            return
        visited.add(code)
        ordered_codes.append(code)
        for child_code in children_by_parent.get(code, []):
            append_code(child_code)

    for code in code_list:
        if code not in anchored_codes:
            append_code(code)
    for code in code_list:
        append_code(code)

    display_rank_map = {code: idx for idx, code in enumerate(ordered_codes)}
    result["_display_rank"] = result[code_col].astype(str).str.strip().map(display_rank_map).fillna(len(display_rank_map))
    result = result.sort_values(by=["_display_rank", "_base_rank"], kind="stable").reset_index(drop=True)
    return result.drop(columns=["_display_rank"], errors="ignore")

def filter_suggestion_detail_data_v2(
    detail_df: pd.DataFrame,
    excluded_codes: set[str] | None = None,
) -> pd.DataFrame:
    if detail_df.empty:
        return detail_df.copy()

    excluded_codes = {str(code).strip() for code in (excluded_codes or set()) if str(code).strip()}
    material_codes = detail_df["建议母件料号"].astype(str).str.strip()
    keep_mask = ~material_codes.str.startswith("70")
    if excluded_codes:
        keep_mask &= ~material_codes.isin(excluded_codes)
    return detail_df.loc[keep_mask].copy().reset_index(drop=True)


def filter_balance_data(
    purchase_view_df: pd.DataFrame,
    shortage_df: pd.DataFrame,
    excluded_codes: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    excluded_codes = {str(code).strip() for code in (excluded_codes or set()) if str(code).strip()}
    material_codes = purchase_view_df["料号"].astype(str).str.strip()
    keep_mask = purchase_view_df["总需求"].fillna(0) != 0
    keep_mask &= ~material_codes.str.startswith("70")
    if excluded_codes:
        keep_mask &= ~material_codes.isin(excluded_codes)

    filtered_purchase = purchase_view_df.loc[keep_mask].copy().reset_index(drop=True)
    filtered_purchase = reorder_purchase_rows_by_suggestion_v2(filtered_purchase)
    filtered_purchase["序号"] = range(1, len(filtered_purchase) + 1)

    keep_codes = set(filtered_purchase["料号"].astype(str).str.strip())
    filtered_shortage = shortage_df[
        shortage_df["BOM2.BOM子项.子件料品.料号"].astype(str).str.strip().isin(keep_codes)
    ].copy()
    return filtered_purchase, filtered_shortage


def filter_suggestion_data(
    suggestion_df: pd.DataFrame,
    excluded_codes: set[str] | None = None,
) -> pd.DataFrame:
    if suggestion_df.empty:
        return suggestion_df.copy()

    excluded_codes = {str(code).strip() for code in (excluded_codes or set()) if str(code).strip()}
    material_codes = suggestion_df["母件料号"].astype(str).str.strip()
    keep_mask = ~material_codes.str.startswith("70")
    if excluded_codes:
        keep_mask &= ~material_codes.isin(excluded_codes)
    return suggestion_df.loc[keep_mask].copy().reset_index(drop=True)


def build_project_detail_df(shortage_df: pd.DataFrame) -> pd.DataFrame:
    if shortage_df.empty:
        return pd.DataFrame(columns=["料号", "项目序号", "项目", "母件料号", "母件品名", "使用上层", "项目需求", "项目短缺", "单套用量", "月份需求"])

    detail_source = shortage_df.copy()
    customer_text = detail_source["客户"].fillna("").astype(str).str.strip()
    parent_name_text = detail_source["母件品名"].fillna("").astype(str).str.strip()
    detail_source["项目显示"] = customer_text.where(customer_text != "", parent_name_text)
    if "上层来源品名" in detail_source.columns:
        source_name_text = detail_source["上层来源品名"].fillna("").astype(str).str.strip()
        parent_code_text = detail_source["母件料号"].fillna("").astype(str).str.strip()
        detail_source["项目显示"] = detail_source["项目显示"].where(
            parent_code_text.str.startswith("04") | (source_name_text == ""),
            source_name_text,
        )
    parent_code_text = detail_source["母件料号"].fillna("").astype(str).str.strip()
    detail_source["批注母件料号"] = detail_source["母件料号"].where(~parent_code_text.str.startswith("04"), "")
    detail_source["批注母件品名"] = detail_source["母件品名"].where(~parent_code_text.str.startswith("04"), "")
    parent_name_for_comment = detail_source["母件品名"].fillna("").astype(str).str.strip()
    parent_spec_for_comment = detail_source["母件规格"].fillna("").astype(str).str.strip()
    parent_display_for_comment = [
        " / ".join(dict.fromkeys(part for part in (name, spec) if part))
        for name, spec in zip(parent_name_for_comment, parent_spec_for_comment, strict=False)
    ]
    parent_display_for_comment = pd.Series(parent_display_for_comment, index=detail_source.index)
    parent_display_for_comment = parent_display_for_comment.where(parent_display_for_comment != "", parent_code_text)
    detail_source["批注使用上层"] = parent_display_for_comment.where(~parent_code_text.str.startswith("04"), "")

    def _month_label(value: object) -> str:
        if isinstance(value, datetime):
            return f"{value.month}月"
        if isinstance(value, date):
            return f"{value.month}月"
        return "未指定月份"

    valid_months = sorted(
        {
            label
            for label in detail_source["到货日期"].apply(_month_label).tolist()
            if label != "未指定月份"
        },
        key=lambda text: int(re.match(r"(\d+)月", text).group(1)) if re.match(r"(\d+)月", text) else 99,
    )
    detail_source["_月份"] = detail_source["到货日期"].apply(_month_label)
    month_summary = (
        detail_source.groupby(["BOM2.BOM子项.子件料品.料号", "项目显示", "批注母件料号", "批注母件品名", "批注使用上层", "_月份"], dropna=False, sort=False)["需求"]
        .sum()
        .reset_index()
    )

    def _month_sort_key(month_text: object) -> tuple[int, str]:
        text = str(month_text or "").strip()
        match = re.match(r"(\d+)月", text)
        if match:
            return (int(match.group(1)), text)
        return (99, text)

    month_demand_map: dict[tuple[str, str, str, str, str], str] = {}
    for (material_code, project_name, parent_code, parent_name, used_on), month_rows in month_summary.groupby(
        ["BOM2.BOM子项.子件料品.料号", "项目显示", "批注母件料号", "批注母件品名", "批注使用上层"],
        sort=False,
    ):
        month_qty = {
            str(row.get("_月份", "") or "").strip(): float(row.get("需求", 0) or 0)
            for row in month_rows.to_dict("records")
        }
        month_labels = valid_months or sorted(month_qty.keys(), key=_month_sort_key)
        parts = []
        for month_label in month_labels:
            qty_text = format_project_display_value(month_qty.get(month_label, 0))
            parts.append(f"{month_label}: {qty_text or '0'}")
        month_demand_map[(
            str(material_code).strip(),
            str(project_name).strip(),
            str(parent_code).strip(),
            str(parent_name).strip(),
            str(used_on).strip(),
        )] = "，".join(parts)

    project_summary = (
        detail_source.groupby(["BOM2.BOM子项.子件料品.料号", "项目显示", "批注母件料号", "批注母件品名", "批注使用上层"], dropna=False, sort=False)
        .agg(
            项目需求=("需求", "sum"),
            项目短缺=("到货数量", "sum"),
            单套用量=("BOM2.BOM子项.子件用量", "first"),
        )
        .reset_index()
        .rename(columns={
            "BOM2.BOM子项.子件料品.料号": "料号",
            "项目显示": "项目",
            "批注母件料号": "母件料号",
            "批注母件品名": "母件品名",
            "批注使用上层": "使用上层",
        })
    )
    project_summary["项目"] = project_summary["项目"].fillna("").astype(str).str.strip()
    project_summary["母件料号"] = project_summary["母件料号"].fillna("").astype(str).str.strip()
    project_summary["母件品名"] = project_summary["母件品名"].fillna("").astype(str).str.strip()
    project_summary["使用上层"] = project_summary["使用上层"].fillna("").astype(str).str.strip()
    project_summary["月份需求"] = [
        month_demand_map.get((
            str(material_code).strip(),
            str(project_name).strip(),
            str(parent_code).strip(),
            str(parent_name).strip(),
            str(used_on).strip(),
        ), "")
        for material_code, project_name, parent_code, parent_name, used_on in zip(
            project_summary["料号"],
            project_summary["项目"],
            project_summary["母件料号"],
            project_summary["母件品名"],
            project_summary["使用上层"],
            strict=False,
        )
    ]
    project_summary = project_summary.sort_values(by=["料号", "项目需求"], ascending=[True, False], kind="stable").reset_index(drop=True)
    project_summary["项目序号"] = project_summary.groupby("料号", sort=False).cumcount() + 1
    return project_summary[["料号", "项目序号", "项目", "母件料号", "母件品名", "使用上层", "项目需求", "项目短缺", "单套用量", "月份需求"]]


def format_project_display_value(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if number.is_integer():
        return str(int(number))
    return f"{number:.4f}".rstrip("0").rstrip(".")


def _role_safe_float(value: object) -> float:
    try:
        if value is None or pd.isna(value):
            return 0.0
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _role_clean_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _role_format_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return _role_clean_text(value)


def _role_top_values(values: Iterable[object], limit: int = 5) -> str:
    items = []
    seen = set()
    for value in values:
        text = _role_clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        items.append(text)
        if len(items) >= limit:
            break
    return "、".join(items)


def evaluate_material_risk_for_views(
    record: dict[str, object],
    date_headers: list[date],
    reply_overrides: dict[tuple[str, date], object] | None = None,
) -> dict[str, object]:
    material_code = _role_clean_text(record.get("料号"))
    total_demand = _role_safe_float(record.get("总需求", 0))
    current_inventory = _role_safe_float(record.get("实时库存", 0))
    substitute_inventory = _role_safe_float(record.get("替代1库存", 0)) + _role_safe_float(record.get("替代2库存", 0))
    current_available = current_inventory + substitute_inventory
    cumulative_demand = 0.0
    cumulative_reply = 0.0
    earliest_gap_date: date | None = None
    earliest_gap_qty = 0.0
    reply_after_today = 0.0
    demand_after_today = 0.0
    today = date.today()

    for header_date in date_headers:
        daily_demand = _role_safe_float(record.get(header_date, 0))
        cumulative_demand += daily_demand
        reply_value = 0.0
        if reply_overrides is not None:
            reply_value = _role_safe_float(reply_overrides.get((material_code, header_date), 0))
        cumulative_reply += reply_value
        if header_date >= today:
            reply_after_today += reply_value
            demand_after_today += daily_demand
        position = current_available + cumulative_reply - cumulative_demand
        if position < 0 and earliest_gap_date is None:
            earliest_gap_date = header_date
            earliest_gap_qty = abs(position)

    if total_demand <= current_available:
        risk = "低"
    elif earliest_gap_date is None:
        risk = "中"
    elif current_available + cumulative_reply >= total_demand:
        risk = "较高"
    else:
        risk = "高"

    return {
        "风险程度": risk,
        "实时库存": current_inventory,
        "替代库存": substitute_inventory,
        "当前可用": current_available,
        "总需求": total_demand,
        "当前缺口": max(total_demand - current_available, 0.0),
        "采购答复累计": cumulative_reply,
        "今日后需求": demand_after_today,
        "今日后答复": reply_after_today,
        "最早缺口日期": earliest_gap_date,
        "最早缺口数量": earliest_gap_qty,
    }


def build_material_role_records(
    purchase_view_df: pd.DataFrame,
    project_detail_df: pd.DataFrame,
    date_headers: list[date],
    reply_overrides: dict[tuple[str, date], object] | None = None,
) -> list[dict[str, object]]:
    project_map: dict[str, dict[str, object]] = {}
    if not project_detail_df.empty:
        for material_code, material_rows in project_detail_df.groupby("料号", sort=False):
            rows = material_rows.to_dict("records")
            project_map[_role_clean_text(material_code)] = {
                "项目摘要": _role_top_values((row.get("项目") for row in rows), limit=4),
                "项目数": len({_role_clean_text(row.get("项目")) for row in rows if _role_clean_text(row.get("项目"))}),
                "项目短缺": sum(_role_safe_float(row.get("项目短缺", 0)) for row in rows),
            }

    role_records: list[dict[str, object]] = []
    for record in purchase_view_df.to_dict("records"):
        material_code = _role_clean_text(record.get("料号"))
        risk_info = evaluate_material_risk_for_views(record, date_headers, reply_overrides)
        project_info = project_map.get(material_code, {})
        open_po = _role_safe_float(record.get("未清PO", 0))
        unconverted = _role_safe_float(record.get("未转", 0))
        substitute_inbound = _role_safe_float(record.get("_替代未清未转合计", 0))
        balance_gap = open_po + unconverted + substitute_inbound - risk_info["总需求"]
        role_records.append({
            "料号": material_code,
            "料品名称": _role_clean_text(record.get("料品名称")),
            "规格": _role_clean_text(record.get("规格")),
            "供应商": _role_clean_text(record.get("供应商")),
            "采购": _role_clean_text(record.get("采购")),
            "未清PO": open_po,
            "未转": unconverted,
            "替代1": _role_clean_text(record.get("替代1")),
            "替代1库存": _role_safe_float(record.get("替代1库存", 0)),
            "替代2": _role_clean_text(record.get("替代2")),
            "替代2库存": _role_safe_float(record.get("替代2库存", 0)),
            "替代未清未转": substitute_inbound,
            "工单缺料": _role_safe_float(record.get("工单缺料", 0)),
            "补": balance_gap,
            **risk_info,
            "项目摘要": project_info.get("项目摘要", ""),
            "项目数": project_info.get("项目数", 0),
            "项目短缺": project_info.get("项目短缺", 0),
        })
    return role_records


def build_role_quality_views(
    result: PipelineResult,
    project_detail_df: pd.DataFrame,
    output_path: Path,
    reply_overrides: dict[tuple[str, date], object] | None = None,
) -> dict[str, pd.DataFrame]:
    material_records = build_material_role_records(
        result.purchase_view_df,
        project_detail_df,
        result.date_headers,
        reply_overrides,
    )
    material_df = pd.DataFrame(material_records)
    risk_order = {"高": 0, "较高": 1, "中": 2, "低": 3}

    if not material_df.empty:
        material_df["_风险排序"] = material_df["风险程度"].map(risk_order).fillna(9)
        material_df["_缺口排序"] = material_df["当前缺口"].fillna(0)

    purchase_action_columns = [
        "采购", "供应商", "料号", "料品名称", "规格", "风险程度", "总需求", "当前可用", "当前缺口",
        "未清PO", "未转", "替代未清未转", "补", "最早缺口日期", "最早缺口数量", "今日后需求", "今日后答复", "项目摘要", "建议动作",
    ]
    if material_df.empty:
        purchase_action_df = pd.DataFrame(columns=purchase_action_columns)
    else:
        purchase_action_df = material_df[
            (material_df["风险程度"].isin(["高", "较高", "中"]))
            | (material_df["补"] < 0)
            | ((material_df["今日后需求"] > 0) & (material_df["今日后答复"] <= 0))
        ].copy()
        purchase_action_df["建议动作"] = purchase_action_df["风险程度"].map({
            "高": "优先确认到货日期、替代料或调整排产",
            "较高": "日期有缺口，补采购答复并确认可覆盖",
            "中": "总量可覆盖，继续跟进日期",
        }).fillna("核对采购答复")
        purchase_action_df = purchase_action_df.sort_values(
            by=["_风险排序", "_缺口排序", "采购", "料号"],
            ascending=[True, False, True, True],
            kind="stable",
        )[purchase_action_columns]

    material_control_columns = [
        "料号", "料品名称", "规格", "供应商", "采购", "风险程度", "实时库存", "替代库存", "当前可用",
        "总需求", "当前缺口", "工单缺料", "未清PO", "未转", "替代未清未转", "项目摘要",
    ]
    if material_df.empty:
        material_control_df = pd.DataFrame(columns=material_control_columns)
    else:
        material_control_df = material_df[
            (material_df["当前缺口"] > 0)
            | (material_df["工单缺料"] > 0)
            | (material_df["风险程度"].isin(["高", "较高"]))
        ].copy()
        material_control_df = material_control_df.sort_values(
            by=["_风险排序", "_缺口排序", "料号"],
            ascending=[True, False, True],
            kind="stable",
        )[material_control_columns]

    project_risk_columns = ["项目", "影响料号数", "高风险料号数", "项目需求", "项目短缺", "主要缺口物料", "建议关注"]
    if project_detail_df.empty:
        project_risk_df = pd.DataFrame(columns=project_risk_columns)
    else:
        risk_map = {row["料号"]: row["风险程度"] for row in material_records}
        project_rows = []
        for project_name, rows in project_detail_df.groupby("项目", sort=False):
            records = rows.to_dict("records")
            material_codes = [_role_clean_text(row.get("料号")) for row in records if _role_clean_text(row.get("料号"))]
            shortage_rows = sorted(records, key=lambda row: _role_safe_float(row.get("项目短缺", 0)), reverse=True)
            high_risk_count = sum(1 for code in set(material_codes) if risk_map.get(code) in {"高", "较高"})
            project_shortage = sum(_role_safe_float(row.get("项目短缺", 0)) for row in records)
            project_rows.append({
                "项目": _role_clean_text(project_name) or "未识别项目",
                "影响料号数": len(set(material_codes)),
                "高风险料号数": high_risk_count,
                "项目需求": sum(_role_safe_float(row.get("项目需求", 0)) for row in records),
                "项目短缺": project_shortage,
                "主要缺口物料": _role_top_values((row.get("料号") for row in shortage_rows if _role_safe_float(row.get("项目短缺", 0)) > 0), limit=8),
                "建议关注": "优先确认高风险缺口物料" if high_risk_count or project_shortage > 0 else "当前按项目追溯无明显缺口",
            })
        project_risk_df = pd.DataFrame(project_rows, columns=project_risk_columns).sort_values(
            by=["高风险料号数", "项目短缺"],
            ascending=[False, False],
            kind="stable",
        )

    production_df = result.near_term_shortage_df.copy()
    if production_df.empty:
        production_df = pd.DataFrame(columns=list(result.near_term_shortage_df.columns) + ["建议动作"])
    else:
        production_df["建议动作"] = "确认当前库存、替代料、接收待检待入库；仍不足时协调采购交期或排产顺序"

    issue_rows: list[dict[str, object]] = []

    def append_issue(issue_type: str, mask: pd.Series, sample_column: str, suggestion: str) -> None:
        count = int(mask.fillna(False).sum()) if not material_df.empty else 0
        sample = ""
        if count:
            sample = _role_top_values(material_df.loc[mask, sample_column].tolist(), limit=6)
        issue_rows.append({"问题类型": issue_type, "数量": count, "示例": sample, "建议处理": suggestion})

    if material_df.empty:
        data_quality_df = pd.DataFrame(columns=["问题类型", "数量", "示例", "建议处理"])
    else:
        append_issue("缺供应商", material_df["供应商"].astype(str).str.strip() == "", "料号", "补齐采购基础资料，便于采购经理按供应商筛选")
        append_issue("缺采购负责人", material_df["采购"].astype(str).str.strip() == "", "料号", "补齐采购负责人，避免待回复统计漏人")
        append_issue("当前库存缺口", material_df["当前缺口"] > 0, "料号", "物控优先核对库存、替代料和安全库存")
        append_issue("有替代料但替代当前库存为0", (
            ((material_df["替代1"].astype(str).str.strip() != "") & (material_df["替代1库存"] <= 0))
            | ((material_df["替代2"].astype(str).str.strip() != "") & (material_df["替代2库存"] <= 0))
        ), "料号", "核对替代料库存、未清PO、未转是否需要补录或转换")
        append_issue("今日后有需求但采购答复为0", (material_df["今日后需求"] > 0) & (material_df["今日后答复"] <= 0), "料号", "采购在供需平衡日期行补交期/数量")
        issue_rows.append({
            "问题类型": "上层展开校验",
            "数量": len(result.upper_expansion_df),
            "示例": _role_top_values(result.upper_expansion_df.iloc[:, 0].tolist(), limit=6) if not result.upper_expansion_df.empty else "",
            "建议处理": "信息化/计划核对BOM上层展开与项目追溯",
        })
        issue_rows.append({
            "问题类型": "近三天排产缺料",
            "数量": len(result.near_term_shortage_df),
            "示例": _role_top_values(result.near_term_shortage_df.iloc[:, 0].tolist(), limit=6) if not result.near_term_shortage_df.empty else "",
            "建议处理": "生产、物控、采购按T+3视图逐项关闭",
        })
        data_quality_df = pd.DataFrame(issue_rows, columns=["问题类型", "数量", "示例", "建议处理"])

    risk_counts = material_df["风险程度"].value_counts().to_dict() if not material_df.empty else {}
    shipping_material_count = 0
    if not project_detail_df.empty and "项目" in project_detail_df.columns:
        shipping_material_count = project_detail_df.loc[
            project_detail_df["项目"].astype(str).str.strip() == INDUSTRIAL_SHIPPING_PARTS_SHEET_NAME,
            "料号",
        ].astype(str).str.strip().nunique()

    dashboard_rows = [
        {"角色": "公司高层", "指标": "高风险物料", "数值": risk_counts.get("高", 0), "说明": "总量加采购答复仍无法覆盖需求的物料"},
        {"角色": "计划经理", "指标": "较高风险物料", "数值": risk_counts.get("较高", 0), "说明": "总量可覆盖但日期滚动存在缺口"},
        {"角色": "采购经理", "指标": "采购待办物料", "数值": len(purchase_action_df), "说明": "需要补采购答复、确认交期或处理负补料的物料"},
        {"角色": "物控经理", "指标": "当前库存缺口物料", "数值": int((material_df["当前缺口"] > 0).sum()) if not material_df.empty else 0, "说明": "仅按实时库存+替代当前库存判断仍不足"},
        {"角色": "生产经理", "指标": "近三天排产缺料", "数值": len(result.near_term_shortage_df), "说明": "按T+3排产计划识别的缺料行"},
        {"角色": "项目经理", "指标": "项目风险项", "数值": len(project_risk_df[project_risk_df["高风险料号数"] > 0]) if not project_risk_df.empty else 0, "说明": "项目下存在高/较高风险物料"},
        {"角色": "信息化经理", "指标": "数据质量问题数", "数值": int(data_quality_df["数量"].sum()) if not data_quality_df.empty else 0, "说明": "供应商、采购、替代料、采购答复、BOM展开等数据问题合计"},
        {"角色": "物控经理", "指标": "工业发货配件安全库存料号", "数值": shipping_material_count, "说明": "来自MRP计算表“工业发货配件”页的安全库存需求"},
    ]
    dashboard_df = pd.DataFrame(dashboard_rows, columns=["角色", "指标", "数值", "说明"])

    audit_rows = [
        {"项目": "生成时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "说明": "本次平衡表写入时间"},
        {"项目": "输出文件", "值": str(output_path), "说明": "当前生成的工作簿路径"},
        {"项目": "物料行数", "值": len(result.purchase_view_df), "说明": "供需平衡物料数"},
        {"项目": "缺料明细行数", "值": len(result.exported_shortage_df), "说明": "缺料表行数"},
        {"项目": "项目明细行数", "值": len(project_detail_df), "说明": "项目追溯明细行数"},
        {"项目": "日期范围", "值": f"{_role_format_date(min(result.date_headers))} 至 {_role_format_date(max(result.date_headers))}" if result.date_headers else "", "说明": "供需平衡横向日期区间"},
        {"项目": "带入采购回复单元格", "值": result.carried_reply_cell_count, "说明": "从旧平衡表继承的采购答复单元格数量"},
        {"项目": "带入采购回复物料", "值": result.carried_reply_material_count, "说明": "从旧平衡表继承采购答复的料号数量"},
        {"项目": "带入采购回复文件", "值": result.carried_reply_file_count, "说明": "参与继承采购答复的旧文件数量"},
    ]
    audit_df = pd.DataFrame(audit_rows, columns=["项目", "值", "说明"])

    return {
        MANAGEMENT_DASHBOARD_SHEET_NAME: dashboard_df,
        PROJECT_RISK_VIEW_SHEET_NAME: project_risk_df,
        PURCHASE_ACTION_VIEW_SHEET_NAME: purchase_action_df,
        MATERIAL_CONTROL_VIEW_SHEET_NAME: material_control_df,
        PRODUCTION_SHORTAGE_VIEW_SHEET_NAME: production_df,
        DATA_QUALITY_VIEW_SHEET_NAME: data_quality_df,
        INFO_AUDIT_VIEW_SHEET_NAME: audit_df,
    }


def split_project_buckets(project_rows: list[dict[str, object]], bucket_count: int = 5) -> list[list[dict[str, object]]]:
    if not project_rows:
        return []
    if len(project_rows) <= bucket_count:
        return [[row] for row in project_rows]

    base_size, remainder = divmod(len(project_rows), bucket_count)
    bucket_sizes = [base_size + (1 if index < remainder else 0) for index in range(bucket_count)]
    buckets: list[list[dict[str, object]]] = []
    cursor = 0
    for bucket_size in bucket_sizes:
        next_cursor = cursor + bucket_size
        buckets.append(project_rows[cursor:next_cursor])
        cursor = next_cursor
    return buckets


def build_project_rows_map(project_detail_df: pd.DataFrame) -> dict[str, list[dict[str, object]]]:
    if project_detail_df.empty:
        return {}

    project_rows_map: dict[str, list[dict[str, object]]] = {}
    for material_code, material_rows in project_detail_df.groupby("料号", sort=False):
        row_records = material_rows.to_dict("records")
        display_rows = []
        for bucket in split_project_buckets(row_records, bucket_count=5):
            display_rows.append(
                {
                    "项目": "\n".join(str(row["项目"]).strip() for row in bucket if str(row["项目"]).strip()),
                    "项目需求": "\n".join(format_project_display_value(row["项目需求"]) for row in bucket),
                    "项目短缺": "\n".join(format_project_display_value(row["项目短缺"]) for row in bucket),
                    "单套用量": "\n".join(format_project_display_value(row["单套用量"]) for row in bucket),
                    "line_count": max(1, len(bucket)),
                }
            )
        project_rows_map[str(material_code).strip()] = display_rows
    return project_rows_map


def build_project_detail_link_map(project_detail_df: pd.DataFrame, header_row: int = 1) -> dict[str, dict[str, object]]:
    if project_detail_df.empty:
        return {}

    result: dict[str, dict[str, object]] = {}
    for excel_row, record in enumerate(project_detail_df.to_dict("records"), start=header_row + 1):
        material_code = str(record.get("料号", "") or "").strip()
        if not material_code:
            continue
        project_name = str(record.get("项目", "") or "").strip()
        display_name = project_name
        info = result.setdefault(
            material_code,
            {
                "first_row": excel_row,
                "projects": [],
            },
        )
        if display_name and display_name not in info["projects"]:
            info["projects"].append(display_name)
        month_demand_text = str(record.get("月份需求", "") or "").strip()
        used_on_text = str(record.get("使用上层", "") or "").strip()
        if month_demand_text:
            qty_parts = []
            for part in re.split(r"[，,]", month_demand_text):
                if ":" in part:
                    qty_parts.append(part.split(":", 1)[1].strip())
                elif "：" in part:
                    qty_parts.append(part.split("：", 1)[1].strip())
                elif part.strip():
                    qty_parts.append(part.strip())
            qty_text = " ".join(qty for qty in qty_parts if qty)
            line_text = f"{display_name} {qty_text}".strip()
            if used_on_text:
                line_text = f"{line_text}；用在：{used_on_text}" if line_text else f"用在：{used_on_text}"
            info.setdefault("project_lines", []).append(line_text)
        elif display_name:
            line_text = display_name
            if used_on_text:
                line_text = f"{line_text}；用在：{used_on_text}"
            info.setdefault("project_lines", []).append(line_text)

    for info in result.values():
        projects = list(info.get("projects", []))
        if not projects:
            info["summary"] = ""
            info["comment"] = ""
            continue
        preview = "、".join(projects[:3])
        if len(projects) > 3:
            preview = f"{preview} 等{len(projects)}个项目"
        info["summary"] = preview
        project_lines = list(dict.fromkeys(str(line).strip() for line in info.get("project_lines", []) if str(line).strip()))
        comment_parts = ["全部项目：", *(project_lines or projects)[:120]]
        info["comment"] = "\n".join(comment_parts)
    return result


def format_project_detail_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_state = "visible"
    if not dataframe.empty:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(dataframe.columns))}{worksheet.max_row}"
    else:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max(len(dataframe.columns), 1))}1"

    widths = {
        "A": 16,
        "B": 10,
        "C": 34,
        "D": 18,
        "E": 28,
        "F": 36,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 34,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def format_near_term_shortage_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_state = "visible"
    max_col = max(len(dataframe.columns), 1)
    if not dataframe.empty:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{worksheet.max_row}"
    else:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    widths = {
        "A": 16,
        "B": 22,
        "C": 28,
        "D": 24,
        "E": 10,
        "F": 12,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 14,
        "Q": 34,
        "R": 34,
        "S": 10,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    shortage_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if worksheet.max_row >= 2:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            row[11].fill = shortage_fill


def format_config_supplement_plan_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_state = "visible"
    max_col = max(len(dataframe.columns), 1)
    if not dataframe.empty:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{worksheet.max_row}"
    else:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

    widths = {
        "A": 10,
        "B": 10,
        "C": 16,
        "D": 24,
        "E": 24,
        "F": 14,
        "G": 12,
        "H": 20,
        "I": 30,
        "J": 24,
        "K": 60,
        "L": 30,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if worksheet.max_row >= 2:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row[0].value == "待确认":
                for cell in row:
                    cell.fill = pending_fill


def copy_block_style(worksheet, start_row: int, end_row: int, max_col: int):
    block = []
    for row in range(start_row, end_row + 1):
        row_styles = []
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row, col)
            row_styles.append(
                {
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                }
            )
        block.append(row_styles)
    return block


def adapt_balance_body_styles(style_rows: list) -> list:
    """Convert the legacy 5-row balance style block to the current 6-row block.

    Old rows: demand, reply, arrival plan, arrival qty, diff.
    New rows: demand, reply, arrival plan, arrival qty, usage qty, diff.
    The new usage row uses the old arrival-qty row style; the diff row keeps the old bottom border.
    """
    rows = list(style_rows or [])
    if not rows:
        return rows
    if len(rows) >= BALANCE_BLOCK_SIZE:
        return rows[:BALANCE_BLOCK_SIZE]
    if len(rows) == LEGACY_BALANCE_BLOCK_SIZE and BALANCE_BLOCK_SIZE == 6:
        return [
            rows[BALANCE_DEMAND_OFFSET],
            rows[BALANCE_REPLY_OFFSET],
            rows[BALANCE_ARRIVAL_PLAN_OFFSET],
            rows[BALANCE_ARRIVAL_QTY_OFFSET],
            rows[BALANCE_ARRIVAL_QTY_OFFSET],
            rows[LEGACY_BALANCE_BLOCK_SIZE - 1],
        ]
    while len(rows) < BALANCE_BLOCK_SIZE:
        rows.append(rows[-1])
    return rows[:BALANCE_BLOCK_SIZE]


def capture_cell_style(cell):
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def apply_cell_style(cell, style) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def first_date_header_col(worksheet, header_row: int = 2) -> int | None:
    for column_index in range(1, worksheet.max_column + 1):
        if coerce_excel_date(worksheet.cell(header_row, column_index).value):
            return column_index
    return None


def capture_column_format(worksheet, column_index: int, sample_rows: int = 7) -> dict[str, object]:
    column_letter = get_column_letter(column_index)
    dimension = worksheet.column_dimensions[column_letter]
    return {
        "width": dimension.width,
        "hidden": dimension.hidden,
        "outline_level": dimension.outlineLevel,
        "collapsed": dimension.collapsed,
        "best_fit": dimension.bestFit,
        "styles": [
            capture_cell_style(worksheet.cell(row_index, column_index))
            for row_index in range(1, sample_rows + 1)
        ],
    }


def apply_repeating_column_format(
    worksheet,
    target_column_index: int,
    column_format: dict[str, object],
    *,
    max_row: int,
) -> None:
    target_letter = get_column_letter(target_column_index)
    target_dimension = worksheet.column_dimensions[target_letter]
    target_dimension.width = column_format["width"]
    target_dimension.hidden = column_format["hidden"]
    target_dimension.outlineLevel = column_format["outline_level"]
    target_dimension.collapsed = column_format["collapsed"]
    target_dimension.bestFit = column_format["best_fit"]

    styles = column_format["styles"]
    header_styles = styles[:2]
    body_styles = adapt_balance_body_styles(styles[2:]) if len(styles) > 2 else styles
    for row_index in range(1, max_row + 1):
        if row_index <= len(header_styles):
            style = header_styles[row_index - 1]
        elif body_styles:
            style = body_styles[(row_index - 3) % len(body_styles)]
        else:
            style = styles[-1]
        apply_cell_style(worksheet.cell(row_index, target_column_index), style)


def apply_date_column_formats(
    worksheet,
    *,
    date_start_col: int,
    last_date_col: int,
    max_row: int,
    template_column_format: dict[str, object] | None,
) -> None:
    if last_date_col < date_start_col:
        return

    if template_column_format:
        for column_index in range(date_start_col, last_date_col + 1):
            apply_repeating_column_format(
                worksheet,
                column_index,
                template_column_format,
                max_row=max_row,
            )
        return

    for column_index in range(date_start_col + 1, last_date_col + 1):
        source_format = capture_column_format(worksheet, column_index - 1)
        apply_repeating_column_format(
            worksheet,
            column_index,
            source_format,
            max_row=max_row,
        )


def apply_block_style(worksheet, row_index: int, style_block) -> None:
    template_row_index = (row_index - 3) % len(style_block)
    row_styles = style_block[template_row_index]
    for col, style in enumerate(row_styles, start=1):
        cell = worksheet.cell(row_index, col)
        apply_cell_style(cell, style)


def clear_sheet_values(worksheet, start_row: int) -> None:
    if worksheet.max_row >= start_row:
        worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)


def write_dataframe_sheet(worksheet, dataframe: pd.DataFrame, header_row: int = 1) -> None:
    clear_sheet_values(worksheet, header_row + 1)
    for index, header in enumerate(dataframe.columns, start=1):
        worksheet.cell(header_row, index, header)
    for record in dataframe.itertuples(index=False, name=None):
        worksheet.append([_excel_value(value) for value in record])


def reset_generated_sheet(worksheet) -> None:
    if worksheet.merged_cells.ranges:
        for merged_range in list(worksheet.merged_cells.ranges):
            worksheet.unmerge_cells(str(merged_range))
    if hasattr(worksheet.conditional_formatting, "_cf_rules"):
        worksheet.conditional_formatting._cf_rules.clear()
    worksheet.auto_filter.ref = None
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)


def format_role_quality_sheet(worksheet, dataframe: pd.DataFrame, sheet_name: str) -> None:
    worksheet.sheet_state = "visible"
    worksheet.freeze_panes = "A2"
    max_col = max(len(dataframe.columns), 1)
    max_row = max(worksheet.max_row, 1)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    fill_by_sheet = {
        MANAGEMENT_DASHBOARD_SHEET_NAME: "D9EAF7",
        PROJECT_RISK_VIEW_SHEET_NAME: "FCE4D6",
        PURCHASE_ACTION_VIEW_SHEET_NAME: "FFF2CC",
        MATERIAL_CONTROL_VIEW_SHEET_NAME: "E2F0D9",
        PRODUCTION_SHORTAGE_VIEW_SHEET_NAME: "D9EAD3",
        DATA_QUALITY_VIEW_SHEET_NAME: "EADCF8",
        INFO_AUDIT_VIEW_SHEET_NAME: "D9D9D9",
    }
    header_fill = PatternFill("solid", fgColor=fill_by_sheet.get(sheet_name, "D9EAF7"))
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(dataframe.columns, start=1):
        header_text = str(header)
        if header_text in {"说明", "建议处理", "建议动作", "项目摘要", "主要缺口物料", "规格"}:
            width = 34
        elif header_text in {"料品名称", "供应商"}:
            width = 24
        elif header_text in {"最早缺口日期", "生成时间", "日期范围"}:
            width = 16
        else:
            width = min(max(len(header_text) + 4, 12), 20)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    date_headers = {"最早缺口日期", "到货日期", "上线日期"}
    numeric_keywords = ("数量", "需求", "缺口", "库存", "未清", "未转", "答复", "补", "物料数", "行数", "数值")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=max_col):
        for cell in row:
            header_text = str(worksheet.cell(1, cell.column).value or "")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if header_text in date_headers:
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, (int, float)) and any(keyword in header_text for keyword in numeric_keywords):
                cell.number_format = "#,##0.00"


def write_role_quality_sheets(
    workbook: Workbook,
    role_views: dict[str, pd.DataFrame],
    progress_callback: ProgressCallback | None = None,
) -> None:
    for sheet_name in ROLE_QUALITY_SHEET_NAMES:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        worksheet = workbook[sheet_name]
        dataframe = role_views.get(sheet_name, pd.DataFrame())
        emit_progress(progress_callback, f"步骤 6/6：写入{sheet_name} {len(dataframe)} 行...")
        reset_generated_sheet(worksheet)
        write_dataframe_sheet(worksheet, dataframe, header_row=1)
        format_role_quality_sheet(worksheet, dataframe, sheet_name)


def sheet_has_body_values(worksheet, start_row: int = 2) -> bool:
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        if any(value not in (None, "") for value in row):
            return True
    return False


def write_suggestion_sheet_preserving_history(worksheet, dataframe: pd.DataFrame) -> None:
    if not dataframe.empty or not sheet_has_body_values(worksheet):
        write_dataframe_sheet(worksheet, dataframe, header_row=1)


def _excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _format_explanation_value(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def build_calculation_explanation_sheet(
    worksheet,
    *,
    shortage_df: pd.DataFrame,
    purchase_view_df: pd.DataFrame,
    project_detail_df: pd.DataFrame,
) -> None:
    if worksheet.merged_cells.ranges:
        for merged_range in list(worksheet.merged_cells.ranges):
            worksheet.unmerge_cells(str(merged_range))
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)

    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="FFF2CC")

    shortage_example = shortage_df.to_dict("records")[0] if not shortage_df.empty else {}
    balance_example = purchase_view_df.to_dict("records")[0] if not purchase_view_df.empty else {}
    project_example = project_detail_df.to_dict("records")[0] if not project_detail_df.empty else {}

    month_columns = [column for column in purchase_view_df.columns if isinstance(column, str) and column.endswith("月需求")]
    first_month_col = month_columns[0] if month_columns else ""

    rows: list[list[object]] = [
        ["计算说明", "", "", "", "", ""],
        ["范围", "字段", "计算方式", "解释", "示例料号", "示例值"],
        ["缺料表", "", "", "", "", ""],
        [
            "缺料表",
            "需求",
            "上线数量 × 子件用量",
            "把生产订单母件展开到子件后，得到每一行子件需求。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            f"上线数量={_format_explanation_value(shortage_example.get('上线数量'))}；子件用量={_format_explanation_value(shortage_example.get('BOM2.BOM子项.子件用量'))}；需求={_format_explanation_value(shortage_example.get('需求'))}",
        ],
        [
            "缺料表",
            "累计需求",
            "同料号按到货日期/上线日期顺序累计需求",
            "用于判断同一个子件到当前行为止的累计消耗。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            _format_explanation_value(shortage_example.get("累计需求")),
        ],
        [
            "缺料表",
            "库存",
            "按子件料号汇总期初库存",
            "从期初库存表取该料号实时库存。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            _format_explanation_value(shortage_example.get("库存")),
        ],
        [
            "缺料表",
            "到货数量",
            "按同料号累计缺口逐行分摊，单行不超过本行需求",
            "用于生成真正需要补进来的缺口数量，不是简单复制需求。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            _format_explanation_value(shortage_example.get("到货数量")),
        ],
        [
            "缺料表",
            "到货日期 / 请购日期",
            "到货日期 = 上线日期 - 2天；请购日期 = 到货日期 - L/T",
            "把缺口往前推到采购动作日期。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            f"到货日期={_format_explanation_value(shortage_example.get('到货日期'))}；请购日期={_format_explanation_value(shortage_example.get('请购日期'))}",
        ],
        [
            "缺料表",
            "工单缺料",
            "按料号汇总期初工单缺料；缺料为空时自动按计划数量-配送数量补算",
            "这是对历史工单缺口的兜底，不依赖手工必须填缺料列。",
            shortage_example.get("BOM2.BOM子项.子件料品.料号", ""),
            _format_explanation_value(shortage_example.get("工单缺料")),
        ],
        ["供需平衡", "", "", "", "", ""],
        [
            "供需平衡",
            "未清PO",
            "在途采购表按料号汇总欠交数量",
            "表示已经下单但尚未完全到货的数量。",
            balance_example.get("料号", ""),
            _format_explanation_value(balance_example.get("未清PO")),
        ],
        [
            "供需平衡",
            "未转",
            "在途请购表按料号汇总未转PO数量",
            "表示请购已提但尚未转采购订单的数量。",
            balance_example.get("料号", ""),
            _format_explanation_value(balance_example.get("未转")),
        ],
        [
            "供需平衡",
            "供应商库存 / 实时库存",
            "供应商库存来自供应商库存表；实时库存来自期初库存表",
            "供应商库存只展示，不参与短缺运算；实时库存参与短缺运算。",
            balance_example.get("料号", ""),
            f"供应商库存={_format_explanation_value(balance_example.get('供应商库存'))}；实时库存={_format_explanation_value(balance_example.get('实时库存'))}",
        ],
        [
            "供需平衡",
            "替代1 / 替代2",
            "替代库存列只显示期初库存；未清PO和未转写入库存单元格批注",
            "补料=未清PO+未转+替代料未清未转-总需求；总需求已按当前库存和替代料库存扣减，不再重复加实时库存/替代库存。",
            balance_example.get("料号", ""),
            f"替代1库存={_format_explanation_value(balance_example.get('替代1库存'))}；替代未清未转={_format_explanation_value(balance_example.get('_替代未清未转合计'))}",
        ],
        [
            "供需平衡",
            first_month_col or "月需求",
            "按到货日期落到对应月份汇总到货数量；首月额外加上工单缺料",
            "月需求列是动态扩展的，有新月份会自动新增对应列。",
            balance_example.get("料号", ""),
            _format_explanation_value(balance_example.get(first_month_col)) if first_month_col else "",
        ],
        [
            "供需平衡",
            "总需求",
            "各月需求列合计",
            "是平衡表筛选和排序的核心需求口径。",
            balance_example.get("料号", ""),
            _format_explanation_value(balance_example.get("总需求")),
        ],
        [
            "供需平衡",
            "停线预警",
            "按表头目标月份汇总差异行数值",
            "用于看指定月份是否存在停线风险，实际显示由模板公式驱动。",
            balance_example.get("料号", ""),
            _format_explanation_value(balance_example.get("停线预警")),
        ],
        [
            "供需平衡",
            "未回复标记",
            "未来有需求且采购答复整行仍全为空/0，则记 1；否则记 0",
            "采购待回复统计就是按这个标记汇总。",
            balance_example.get("料号", ""),
            "1=未回复，0=已回复或未来无需求",
        ],
        ["项目明细", "", "", "", "", ""],
        [
            "项目明细",
            "项目",
            "按料号 + 客户分组",
            "这里的项目来源于缺料明细里的客户字段。",
            project_example.get("料号", ""),
            _format_explanation_value(project_example.get("项目")),
        ],
        [
            "项目明细",
            "项目需求 / 项目短缺 / 单套用量",
            "同料号同项目汇总需求、到货数量，并保留该项目单套用量",
            "供需平衡 A 列项目筛选会直接读取这个 sheet。",
            project_example.get("料号", ""),
            f"项目需求={_format_explanation_value(project_example.get('项目需求'))}；项目短缺={_format_explanation_value(project_example.get('项目短缺'))}；单套用量={_format_explanation_value(project_example.get('单套用量'))}",
        ],
    ]

    for row_index, row_values in enumerate(rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row_index, column_index, value)

    worksheet["A1"].font = title_font
    for column_index in range(1, 7):
        worksheet.cell(2, column_index).font = header_font
        worksheet.cell(2, column_index).fill = header_fill

    for row_index in range(3, len(rows) + 1):
        if worksheet.cell(row_index, 2).value == "":
            for column_index in range(1, 7):
                worksheet.cell(row_index, column_index).fill = section_fill
                worksheet.cell(row_index, column_index).font = header_font

    worksheet.freeze_panes = "A3"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 36
    worksheet.column_dimensions["D"].width = 42
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["F"].width = 48

def apply_balance_header(
    worksheet,
    date_headers: list[date],
    *,
    month_start_col: int = 20,
    date_start_col: int = 26,
    freeze_panes: str = "Z3",
) -> None:
    month_headers = planning_month_headers(date_headers)
    for col in range(month_start_col, worksheet.max_column + 1):
        worksheet.cell(2, col).value = None
    for offset, month_header in enumerate(month_headers, start=month_start_col):
        worksheet.cell(2, offset, f"{month_header.month}月需求")
    for offset, header_date in enumerate(date_headers, start=date_start_col):
        worksheet.cell(2, offset, datetime.combine(header_date, datetime.min.time()))
    worksheet.freeze_panes = freeze_panes


def template_formula_text(cell) -> str | None:
    value = cell.value
    if value is None:
        return None
    if hasattr(value, "text"):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def translated_formula(formula_text: str | None, origin: str, target: str) -> str | None:
    if not formula_text:
        return None
    return Translator(formula_text, origin=origin).translate_formula(target)


def _build_monthly_coverage_terms(
    *,
    row_index: int,
    date_start_col: int,
    date_headers: list[date],
    month_start_col: int,
    compare_op: str,
) -> list[str]:
    """生成"未来每个月份是否满足/不满足"的单项表达式列表。

    compare_op = "<"  → 每项 1 表示"该未来月份不满足（月份需求 > 月采购答复累计）"
    compare_op = ">=" → 每项 1 表示"该未来月份满足（月份需求 <= 月采购答复累计）"
    """
    demand_row = row_index - 1
    reply_row = row_index
    cutoff_date = f"INT('{USAGE_START_DATETIME_SHEET_NAME}'!{USAGE_START_DATETIME_CELL})"
    month_headers = planning_month_headers(date_headers)
    terms: list[str] = []
    for month_idx, month_header in enumerate(month_headers):
        in_month_indices = [
            idx for idx, d in enumerate(date_headers)
            if d.year == month_header.year and d.month == month_header.month
        ]
        if not in_month_indices:
            continue
        month_col_letter = get_column_letter(month_start_col + month_idx)
        month_demand_cell = f"${month_col_letter}${demand_row}"
        first_col_letter = get_column_letter(date_start_col + in_month_indices[0])
        last_col_letter = get_column_letter(date_start_col + in_month_indices[-1])
        month_reply_range = f"{first_col_letter}{reply_row}:{last_col_letter}{reply_row}"
        month_last_header = max(date_headers[i] for i in in_month_indices)
        date_literal = f"DATE({month_last_header.year},{month_last_header.month},{month_last_header.day})"
        # 该月份需在"未来或当月"并且月份需求 > 0 时才纳入统计
        if compare_op == "<":
            cond = f"({month_demand_cell}>SUM({month_reply_range}))"
        else:
            cond = f"({month_demand_cell}<=SUM({month_reply_range}))"
        terms.append(
            f"((({date_literal})>={cutoff_date})*({month_demand_cell}>0)*{cond})"
        )
    return terms


def build_purchase_reply_flag_formula(
    *,
    row_index: int,
    row_label_col: int,
    date_start_col: int,
    last_date_col: int,
    date_headers: list[date] | None = None,
    month_start_col: int | None = None,
) -> str:
    """需求满足标记：返回该物料"未来月份中满足的月份数"（整数，0 表示全都不满足）。

    月份满足口径：月份需求 > 0，并且该月"采购答复"累计 >= 月份需求。
    仅统计月末 >= 今天 的月份（未来月份或当月）。
    如果该物料未来完全没有月份需求 -> 返回 ""（不计数）。
    """
    row_label_ref = f"${get_column_letter(row_label_col)}{row_index}"
    if last_date_col < date_start_col:
        return f'=IF({row_label_ref}<>"采购答复","",0)'

    header_range = f"${get_column_letter(date_start_col)}$2:${get_column_letter(last_date_col)}$2"
    need_range = f"{get_column_letter(date_start_col)}{row_index - 1}:{get_column_letter(last_date_col)}{row_index - 1}"
    cutoff_date = f"INT('{USAGE_START_DATETIME_SHEET_NAME}'!{USAGE_START_DATETIME_CELL})"

    if not date_headers or month_start_col is None:
        diff_range = f"{get_column_letter(date_start_col)}{row_index + 3}:{get_column_letter(last_date_col)}{row_index + 3}"
        return (
            f'=IF({row_label_ref}<>"采购答复","",'
            f'IF(SUMPRODUCT(--({header_range}>={cutoff_date}),--({need_range}>0))=0,"",'
            f'IF(SUMPRODUCT(--({header_range}>={cutoff_date}),--({need_range}>0),--({diff_range}<0))>0,0,1)))'
        )

    terms = _build_monthly_coverage_terms(
        row_index=row_index,
        date_start_col=date_start_col,
        date_headers=date_headers,
        month_start_col=month_start_col,
        compare_op=">=",
    )
    if not terms:
        return f'=IF({row_label_ref}<>"采购答复","",0)'
    return (
        f'=IF({row_label_ref}<>"采购答复","",'
        f'IF(SUMPRODUCT(--({header_range}>={cutoff_date}),--({need_range}>0))=0,"",'
        f'{"+".join(terms)}))'
    )


def build_purchase_reply_unsatisfied_formula(
    *,
    row_index: int,
    row_label_col: int,
    date_start_col: int,
    last_date_col: int,
    date_headers: list[date] | None = None,
    month_start_col: int | None = None,
) -> str:
    """需求不满足月份数：返回该物料未来月份中"不满足"的月份个数（月份需求 > 采购答复累计）。"""
    row_label_ref = f"${get_column_letter(row_label_col)}{row_index}"
    if last_date_col < date_start_col or not date_headers or month_start_col is None:
        return f'=IF({row_label_ref}<>"采购答复","",0)'

    header_range = f"${get_column_letter(date_start_col)}$2:${get_column_letter(last_date_col)}$2"
    need_range = f"{get_column_letter(date_start_col)}{row_index - 1}:{get_column_letter(last_date_col)}{row_index - 1}"
    cutoff_date = f"INT('{USAGE_START_DATETIME_SHEET_NAME}'!{USAGE_START_DATETIME_CELL})"

    terms = _build_monthly_coverage_terms(
        row_index=row_index,
        date_start_col=date_start_col,
        date_headers=date_headers,
        month_start_col=month_start_col,
        compare_op="<",
    )
    if not terms:
        return f'=IF({row_label_ref}<>"采购答复","",0)'
    return (
        f'=IF({row_label_ref}<>"采购答复","",'
        f'IF(SUMPRODUCT(--({header_range}>={cutoff_date}),--({need_range}>0))=0,"",'
        f'{"+".join(terms)}))'
    )


def build_purchase_reply_missing_flag_formula(
    *,
    row_index: int,
    row_label_col: int,
    date_start_col: int,
    last_date_col: int,
) -> str:
    """未回复标记：只要这一行"采购答复"整行求和 > 0，就视为已回复（标记 0）；全为空/0 才标 1。

    口径：只看是否填了，不看是否按日期逐格对应需求，也不和月份需求比较。
    """
    row_label_ref = f"${get_column_letter(row_label_col)}{row_index}"
    if last_date_col < date_start_col:
        return f'=IF({row_label_ref}<>"采购答复","",0)'

    header_range = f"${get_column_letter(date_start_col)}$2:${get_column_letter(last_date_col)}$2"
    need_range = f"{get_column_letter(date_start_col)}{row_index - 1}:{get_column_letter(last_date_col)}{row_index - 1}"
    reply_range = f"{get_column_letter(date_start_col)}{row_index}:{get_column_letter(last_date_col)}{row_index}"
    cutoff_date = f"INT('{USAGE_START_DATETIME_SHEET_NAME}'!{USAGE_START_DATETIME_CELL})"
    return (
        f'=IF({row_label_ref}<>"采购答复","",'
        f'IF(SUMPRODUCT(--({header_range}>={cutoff_date}),--({need_range}>0))=0,"",'
        f'IF(SUM({reply_range})>0,0,1)))'
    )


def rebuild_purchase_reply_summary_sheet(
    worksheet,
    purchase_view_df: pd.DataFrame,
    *,
    balance_sheet_name: str,
    balance_buyer_col: int,
    balance_missing_count_col: int,
) -> None:
    if worksheet.merged_cells.ranges:
        for merged_range in list(worksheet.merged_cells.ranges):
            worksheet.unmerge_cells(str(merged_range))
    if hasattr(worksheet.conditional_formatting, "_cf_rules"):
        worksheet.conditional_formatting._cf_rules.clear()
    worksheet.auto_filter.ref = None
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)

    title_font = Font(size=13, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    positive_fill = PatternFill("solid", fgColor="FFF4CCCC")
    positive_font = Font(color="FF9C0006")

    worksheet.sheet_view.showGridLines = True
    worksheet.freeze_panes = "A5"
    worksheet.cell(1, 1, "采购待回复统计").font = title_font
    worksheet.cell(2, 1, "统计口径：未来有需求，且采购答复整行为空或 0 的物料，按 1 条未回复物料统计。")
    worksheet.merge_cells("A2:B2")
    worksheet.cell(4, 1, "采购").font = header_font
    worksheet.cell(4, 2, "未回复物料数").font = header_font
    worksheet.cell(4, 3, "采购辅助").font = header_font
    for column in range(1, 4):
        worksheet.cell(4, column).fill = header_fill

    buyer_series = (
        purchase_view_df["采购"].fillna("").astype(str).str.strip()
        if "采购" in purchase_view_df.columns
        else pd.Series(dtype=str)
    )
    buyers: list[str] = []
    seen: set[str] = set()
    for buyer in buyer_series.tolist():
        if buyer in seen:
            continue
        seen.add(buyer)
        buyers.append(buyer)

    balance_buyer_letter = get_column_letter(balance_buyer_col)
    balance_missing_count_letter = get_column_letter(balance_missing_count_col)
    start_row = 5
    for row_index, buyer in enumerate(buyers, start=start_row):
        worksheet.cell(row_index, 1, buyer or "未填写采购")
        worksheet.cell(row_index, 3, buyer)
        worksheet.cell(
            row_index,
            2,
            (
                f"=COUNTIFS({balance_sheet_name}!${balance_buyer_letter}:${balance_buyer_letter},"
                f"$C{row_index},{balance_sheet_name}!${balance_missing_count_letter}:${balance_missing_count_letter},1)"
            ),
        )

    worksheet.cell(1, 1, "未回复物料合计").font = header_font
    if buyers:
        worksheet.cell(1, 2, f"=SUM(B{start_row}:B{start_row + len(buyers) - 1})").fill = warning_fill
        worksheet.auto_filter.ref = f"A4:B{start_row + len(buyers) - 1}"
        worksheet.conditional_formatting.add(
            f"B{start_row}:B{start_row + len(buyers) - 1}",
            FormulaRule(
                formula=[f"B{start_row}>0"],
                stopIfTrue=False,
                fill=positive_fill,
                font=positive_font,
            ),
        )
    else:
        worksheet.cell(1, 2, 0).fill = warning_fill

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].hidden = True


def build_balance_sheet(
    worksheet,
    purchase_view_df: pd.DataFrame,
    shortage_df: pd.DataFrame,
    date_headers: list[date],
    project_detail_df: pd.DataFrame | None = None,
    reply_overrides: dict[tuple[str, date], object] | None = None,
) -> int:
    del shortage_df

    month_headers = planning_month_headers(date_headers)
    month_labels = [f"{month_header.month}月需求" for month_header in month_headers]
    arrival_status_col = 13
    project_risk_col = arrival_status_col
    supplier_inventory_col = 14
    real_inventory_col = 15
    total_demand_col = 16
    work_order_shortage_col = 17
    month_start_col = 18
    remark_col = month_start_col + len(month_labels)
    stopline_col = remark_col + 1
    row_label_col = stopline_col + 1
    date_start_col = row_label_col + 1
    last_fixed_col = row_label_col
    last_date_col = last_fixed_col + len(date_headers)
    missing_count_col = last_date_col + 1
    substitute_inbound_col = missing_count_col + 1
    project_detail_link_map = build_project_detail_link_map(project_detail_df) if project_detail_df is not None else {}

    if worksheet.cell(2, 15).value == "项目":
        worksheet.delete_cols(15, 4)

    def header_col(header_text: str) -> int | None:
        for column_index in range(1, worksheet.max_column + 1):
            if worksheet.cell(2, column_index).value == header_text:
                return column_index
        return None

    template_remark_col = header_col("备注")
    template_stopline_col = header_col("停线预警")
    template_row_label_col = header_col("父项分类2")
    template_block_size = LEGACY_BALANCE_BLOCK_SIZE
    if template_row_label_col:
        template_diff_label = str(worksheet.cell(3 + BALANCE_DIFF_OFFSET, template_row_label_col).value or "").strip()
        if template_diff_label == BALANCE_ROW_LABELS[BALANCE_DIFF_OFFSET]:
            template_block_size = BALANCE_BLOCK_SIZE
    template_header_styles = {}
    template_column_styles: dict[str, list[dict[str, object]]] = {}
    template_j_formulas = [
        template_formula_text(worksheet.cell(3 + offset, 10))
        for offset in range(template_block_size)
    ]
    template_stopline_formulas: list[str | None] = []
    if template_remark_col and template_stopline_col and template_row_label_col:
        template_header_styles = {
            "remark": capture_cell_style(worksheet.cell(2, template_remark_col)),
            "stopline": capture_cell_style(worksheet.cell(2, template_stopline_col)),
            "row_label": capture_cell_style(worksheet.cell(2, template_row_label_col)),
        }
        template_column_styles = {
            "remark": adapt_balance_body_styles([
                capture_cell_style(worksheet.cell(3 + offset, template_remark_col))
                for offset in range(template_block_size)
            ]),
            "stopline": adapt_balance_body_styles([
                capture_cell_style(worksheet.cell(3 + offset, template_stopline_col))
                for offset in range(template_block_size)
            ]),
            "row_label": adapt_balance_body_styles([
                capture_cell_style(worksheet.cell(3 + offset, template_row_label_col))
                for offset in range(template_block_size)
            ]),
        }
        template_stopline_formulas = [
            template_formula_text(worksheet.cell(3 + offset, template_stopline_col))
            for offset in range(template_block_size)
        ]
    template_date_col = first_date_header_col(worksheet)
    template_date_column_format = (
        capture_column_format(worksheet, template_date_col, sample_rows=2 + template_block_size)
        if template_date_col is not None
        else None
    )

    max_data_col = max(worksheet.max_column, substitute_inbound_col)
    style_block = adapt_balance_body_styles(
        copy_block_style(worksheet, 3, 2 + template_block_size, max_data_col)
    )
    existing_max_row = worksheet.max_row
    target_max_row = 2 + len(purchase_view_df) * BALANCE_BLOCK_SIZE
    if existing_max_row > target_max_row:
        worksheet.delete_rows(target_max_row + 1, existing_max_row - target_max_row)
    if hasattr(worksheet.conditional_formatting, "_cf_rules"):
        worksheet.conditional_formatting._cf_rules.clear()
    apply_balance_header(
        worksheet,
        date_headers,
        month_start_col=month_start_col,
        date_start_col=date_start_col,
        freeze_panes=f"{get_column_letter(date_start_col)}3",
    )
    worksheet.cell(2, 11, "替代1")
    worksheet.cell(2, 12, "替代2")
    worksheet.cell(2, arrival_status_col, "风险程度")
    worksheet.cell(2, supplier_inventory_col, "供应商库存")
    worksheet.cell(2, real_inventory_col, "实时库存")
    worksheet.cell(2, total_demand_col, "总需求")
    worksheet.cell(2, work_order_shortage_col, "工单缺料")
    worksheet.cell(2, remark_col, "备注")
    worksheet.cell(2, stopline_col, "停线预警")
    worksheet.cell(2, row_label_col, "父项分类2")
    worksheet.cell(2, missing_count_col, "未回复标记")
    worksheet.cell(2, substitute_inbound_col, "替代未清未转")
    worksheet.column_dimensions[get_column_letter(row_label_col)].width = 8
    worksheet.column_dimensions[get_column_letter(missing_count_col)].hidden = True
    worksheet.column_dimensions[get_column_letter(substitute_inbound_col)].hidden = True
    for column_letter in ["H", "I", "J", "K", "L", "M", "N"]:
        worksheet.column_dimensions[column_letter].outlineLevel = 0
        worksheet.column_dimensions[column_letter].collapsed = False
    worksheet.column_dimensions.group("H", "L", outline_level=1, hidden=False)
    if template_header_styles:
        apply_cell_style(worksheet.cell(2, remark_col), template_header_styles["remark"])
        apply_cell_style(worksheet.cell(2, stopline_col), template_header_styles["stopline"])
        apply_cell_style(worksheet.cell(2, row_label_col), template_header_styles["row_label"])

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    diff_negative_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
    diff_negative_font = Font(color="FF9C0006")
    risk_filter_cell = worksheet.cell(1, 4)
    risk_filter_cell.fill = yellow_fill
    risk_filter_cell.comment = Comment("风险筛选：输入 极高 / 高 / 较高 / 中 / 低 后，将A列辅助列筛选为1，可显示对应风险物料的6行明细。", "MRP")

    def risk_level(record: dict[str, object]) -> str:
        def safe_float(value: object) -> float:
            try:
                if value is None or pd.isna(value):
                    return 0.0
                return float(value)
            except (TypeError, ValueError):
                return 0.0

        material_code = str(record.get("料号", "") or "").strip()
        total_demand = safe_float(record.get("总需求", 0))
        current_inventory = safe_float(record.get("实时库存", 0))
        substitute_inventory = safe_float(record.get("替代1库存", 0)) + safe_float(record.get("替代2库存", 0))
        current_available = current_inventory + substitute_inventory
        if total_demand <= current_available:
            return "低"

        cumulative_demand = 0.0
        cumulative_reply = 0.0
        min_position = current_available
        for header_date in date_headers:
            cumulative_demand += safe_float(record.get(header_date, 0))
            reply_value = 0
            if reply_overrides is not None:
                reply_value = reply_overrides.get((material_code, header_date), 0)
            cumulative_reply += safe_float(reply_value)
            min_position = min(min_position, current_available + cumulative_reply - cumulative_demand)

        if min_position >= 0:
            return "中"
        if current_available + cumulative_reply >= total_demand:
            return "较高"
        return "高"

    def risk_comment(record: dict[str, object]) -> str:
        del record
        return "\n".join(
            [
                "按本表库存、采购答复、需求、领用滚动判断。",
                "极高：无库存且近期无采购答复。",
                "高/较高：观察天数内有缺口。",
                "中/低：暂可覆盖，天数在生产计划页调整。",
            ]
        )

    def build_dynamic_risk_formula(base_row: int) -> str:
        real_col = get_column_letter(real_inventory_col)
        total_col = get_column_letter(total_demand_col)
        current_available = f"SUM({real_col}{base_row}:{real_col}{base_row + BALANCE_ARRIVAL_PLAN_OFFSET})"
        total_demand = f"{total_col}{base_row}"
        if not date_headers:
            return (
                f'=IF({total_demand}<={current_available},"低",'
                f'IF({current_available}<=0,"极高","高"))'
            )

        start_col = get_column_letter(date_start_col)
        last_col = get_column_letter(last_date_col)
        reply_row = base_row + BALANCE_REPLY_OFFSET
        usage_row = base_row + BALANCE_USAGE_OFFSET
        diff_row = base_row + BALANCE_DIFF_OFFSET
        all_replies = f"{start_col}{reply_row}:{last_col}{reply_row}"
        all_usages = f"{start_col}{usage_row}:{last_col}{usage_row}"
        all_diff = f"{start_col}{diff_row}:{last_col}{diff_row}"
        config_sheet = f"'{USAGE_START_DATETIME_SHEET_NAME}'"

        def window_expr(function_name: str, row_index: int, days_cell: str) -> str:
            full_range = f"{start_col}{row_index}:{last_col}{row_index}"
            width = f"MIN(MAX(1,{config_sheet}!{days_cell}),COLUMNS({full_range}))"
            return f"{function_name}(OFFSET({start_col}{row_index},0,0,1,{width}))"

        extreme_replies = window_expr("SUM", reply_row, RISK_EXTREME_DAYS_CELL)
        low_demand = window_expr("SUM", base_row, RISK_LOW_BUFFER_DAYS_CELL)
        high_diff = window_expr("MIN", diff_row, RISK_HIGH_DAYS_CELL)
        medium_diff = window_expr("MIN", diff_row, RISK_MEDIUM_DAYS_CELL)
        return (
            f'=IF(OR({total_demand}<={current_available},{current_available}>={low_demand}),"低",'
            f'IF(AND({current_available}<=0,{extreme_replies}<=0),"极高",'
            f'IF({current_available}+SUM({all_replies})-SUM({all_usages})<{total_demand},"高",'
            f'IF({high_diff}<0,"高",'
            f'IF({medium_diff}<0,"较高",'
            f'IF(MIN({all_diff})<0,"较高","中"))))))'
        )

    def build_dynamic_inventory_formula(code_ref: str, initial_inventory: object) -> str:
        try:
            if initial_inventory is None or pd.isna(initial_inventory):
                initial_inventory = 0.0
            initial_value = float(initial_inventory)
        except (TypeError, ValueError):
            initial_value = 0.0
        if float(initial_value).is_integer():
            initial_text = str(int(initial_value))
        else:
            initial_text = str(initial_value)
        return (
            f'=IF(TRIM({code_ref})="",0,'
            f'{initial_text}'
            f'+IFERROR(SUMIFS(\'{USAGE_SHEET_NAME}\'!$L:$L,\'{USAGE_SHEET_NAME}\'!$G:$G,{code_ref},'
            f'\'{USAGE_SHEET_NAME}\'!$D:$D,"*入库*",'
            f'\'{USAGE_SHEET_NAME}\'!$B:$B,">="&\'{USAGE_START_DATETIME_SHEET_NAME}\'!{USAGE_START_DATETIME_CELL},'
            f'\'{USAGE_SHEET_NAME}\'!$B:$B,"<"&TODAY()+1),0)'
            f'-IFERROR(SUMIFS(\'{USAGE_SHEET_NAME}\'!$L:$L,\'{USAGE_SHEET_NAME}\'!$G:$G,{code_ref},'
            f'\'{USAGE_SHEET_NAME}\'!$D:$D,"*出库*",'
            f'\'{USAGE_SHEET_NAME}\'!$B:$B,">="&\'{USAGE_START_DATETIME_SHEET_NAME}\'!{USAGE_START_DATETIME_CELL},'
            f'\'{USAGE_SHEET_NAME}\'!$B:$B,"<"&TODAY()+1),0))'
        )

    def numeric_formula_constant(value: object) -> str:
        try:
            if value is None or pd.isna(value):
                number = 0.0
            else:
                number = float(value)
        except (TypeError, ValueError):
            number = 0.0
        if float(number).is_integer():
            return str(int(number))
        return str(number)

    def build_arrival_status_formula(row_index: int, label: str, statuses: list[str]) -> str:
        material_ref = f"$C{row_index}"
        qty_col = f"'{ARRIVAL_RECORD_SHEET_NAME}'!$L:$L"
        reject_col = f"'{ARRIVAL_RECORD_SHEET_NAME}'!$M:$M"
        broken_col = f"'{ARRIVAL_RECORD_SHEET_NAME}'!$N:$N"
        material_col = f"'{ARRIVAL_RECORD_SHEET_NAME}'!$G:$G"
        status_col = f"'{ARRIVAL_RECORD_SHEET_NAME}'!$R:$R"
        parts = []
        for status in statuses:
            criteria = f'"*{status}*"'
            parts.append(
                f"SUMIFS({qty_col},{material_col},{material_ref},{status_col},{criteria})"
                f"-SUMIFS({reject_col},{material_col},{material_ref},{status_col},{criteria})"
                f"-SUMIFS({broken_col},{material_col},{material_ref},{status_col},{criteria})"
            )
        qty_expr = "+".join(parts) or "0"
        return f'="{label}："&IFERROR({qty_expr},0)'

    for material_index, record in enumerate(purchase_view_df.to_dict("records"), start=0):
        base_row = 3 + material_index * BALANCE_BLOCK_SIZE
        material_code = record["料号"]
        for row_offset in range(BALANCE_BLOCK_SIZE):
            current_row = base_row + row_offset
            apply_block_style(worksheet, current_row, style_block)
            worksheet.row_dimensions[current_row].height = 13.5
            if template_column_styles:
                apply_cell_style(worksheet.cell(current_row, remark_col), template_column_styles["remark"][row_offset])
                apply_cell_style(worksheet.cell(current_row, stopline_col), template_column_styles["stopline"][row_offset])
                apply_cell_style(worksheet.cell(current_row, row_label_col), template_column_styles["row_label"][row_offset])
            worksheet.cell(
                current_row,
                1,
                (
                    f'=IF(AND('
                    f'IF($C$1="",1,IF(OR('
                    f'COUNTIFS(项目明细!$A:$A,$C{current_row},项目明细!$C:$C,"*"&$C$1&"*")>0,'
                    f'COUNTIFS({BALANCE_HIERARCHY_FILTER_SHEET_NAME}!$A:$A,$C{current_row},'
                    f'{BALANCE_HIERARCHY_FILTER_SHEET_NAME}!$B:$B,"*"&$C$1&"*")>0),1,0))=1,'
                    f'IF($D$1="",1,IF(${get_column_letter(project_risk_col)}${base_row}=$D$1,1,0))=1'
                    f'),1,0)'
                ),
            )
            worksheet.cell(current_row, 2, record["序号"])
            worksheet.cell(current_row, 3, material_code)
            worksheet.cell(current_row, 4, record["料品名称"])
            worksheet.cell(current_row, 5, record["规格"])
            worksheet.cell(current_row, 6, record["供应商"])
            worksheet.cell(current_row, 7, record["采购"])
            worksheet.cell(current_row, 8, record["未清PO"])
            worksheet.cell(current_row, 9, record["未转"])
            worksheet.cell(current_row, 11, record["替代1"])
            substitute_one_inventory_comment = str(record.get("_替代1库存批注", "") or "").strip()
            substitute_two_cell = worksheet.cell(current_row, 12, record["替代2"])
            substitute_two_comment = str(record.get("_替代2批注", "") or "").strip()
            substitute_two_cell.comment = (
                Comment(substitute_two_comment, "供需协同工具") if substitute_two_comment and row_offset == 0 else None
            )
            substitute_two_inventory_comment = str(record.get("_替代2库存批注", "") or "").strip()
            project_cell = worksheet.cell(current_row, project_risk_col, None)
            project_cell.hyperlink = None
            project_cell.comment = None
            project_cell.font = copy(worksheet.cell(current_row, 3).font)
            if row_offset == BALANCE_ARRIVAL_PLAN_OFFSET:
                worksheet.cell(current_row, project_risk_col, build_arrival_status_formula(current_row, "接收", ["接收"]))
            elif row_offset == BALANCE_ARRIVAL_QTY_OFFSET:
                worksheet.cell(current_row, project_risk_col, build_arrival_status_formula(current_row, "在检", ["待检", "在检"]))
            elif row_offset == BALANCE_USAGE_OFFSET:
                worksheet.cell(current_row, project_risk_col, build_arrival_status_formula(current_row, "待入库", ["待入库"]))
            worksheet.cell(current_row, supplier_inventory_col, record["供应商库存"])
            if row_offset == 0:
                inventory_code_ref = f"C{base_row}"
                real_inventory_value = build_dynamic_inventory_formula(inventory_code_ref, record["实时库存"])
            elif row_offset == 1:
                substitute_one_code = str(record.get("替代1", "") or "").strip()
                if substitute_one_code:
                    inventory_code_ref = f"K{base_row}"
                    real_inventory_value = build_dynamic_inventory_formula(inventory_code_ref, record["替代1库存"])
                else:
                    real_inventory_value = None
            elif row_offset == 2:
                substitute_two_code = str(record.get("替代2", "") or "").strip()
                if substitute_two_code:
                    inventory_code_ref = f"L{base_row}"
                    real_inventory_value = build_dynamic_inventory_formula(inventory_code_ref, record["替代2库存"])
                else:
                    real_inventory_value = None
            else:
                real_inventory_value = None
            real_inventory_cell = worksheet.cell(current_row, real_inventory_col, real_inventory_value)
            real_inventory_cell.number_format = "0"
            if row_offset == 1 and substitute_one_inventory_comment:
                real_inventory_cell.comment = Comment(substitute_one_inventory_comment, "供需协同工具")
            elif row_offset == 2 and substitute_two_inventory_comment:
                real_inventory_cell.comment = Comment(substitute_two_inventory_comment, "供需协同工具")
            worksheet.cell(current_row, total_demand_col, record["总需求"])
            worksheet.cell(current_row, work_order_shortage_col, record["工单缺料"])
            for month_offset, month_label in enumerate(month_labels):
                worksheet.cell(current_row, month_start_col + month_offset, record.get(month_label, 0))
            worksheet.cell(current_row, remark_col, record["备注"])
            worksheet.cell(current_row, row_label_col, BALANCE_ROW_LABELS[row_offset])
            worksheet.cell(current_row, missing_count_col, None)
            if current_row <= existing_max_row:
                for clear_col in range(date_start_col, max_data_col + 1):
                    worksheet.cell(current_row, clear_col).value = None
            worksheet.cell(current_row, substitute_inbound_col, record.get("_替代未清未转合计", 0))

        project_info = project_detail_link_map.get(str(material_code).strip())
        risk_note = risk_comment(record)
        dynamic_risk_formula = build_dynamic_risk_formula(base_row)
        if project_info and project_info.get("summary"):
            project_cell = worksheet.cell(base_row, project_risk_col, dynamic_risk_formula)
            project_cell.hyperlink = None
            project_cell.font = copy(worksheet.cell(base_row, 3).font)
            project_cell.number_format = "General"
            project_cell.alignment = Alignment(wrap_text=True, vertical="center")
            comment_text = str(project_info.get("comment", ""))
            if comment_text:
                project_cell.comment = Comment(comment_text, "MRP")
        else:
            project_cell = worksheet.cell(base_row, project_risk_col, dynamic_risk_formula)
            project_cell.hyperlink = None
            project_cell.font = copy(worksheet.cell(base_row, 3).font)
            project_cell.number_format = "General"
        if risk_note:
            worksheet.cell(base_row + 1, project_risk_col).comment = Comment(risk_note, "MRP")

        worksheet.cell(
            base_row,
            10,
            (
                f"=H{base_row}+I{base_row}"
                f"+{get_column_letter(substitute_inbound_col)}{base_row}"
                f"-{get_column_letter(total_demand_col)}{base_row}"
            ),
        )
        material_codes_for_supply = [str(material_code).strip(), *_deserialize_codes_v2(record.get("_替代料清单", ""))]
        material_codes_for_supply = [code for code in dict.fromkeys(material_codes_for_supply) if code]

        # 停线预警：根据 row 1 的目标日期，在对应日期列返回"差异"行的数值。
        # 不使用 template 翻译后的公式——当月份数（stopline_col 的位置）与模板不同时，
        # openpyxl 的 Translator 无法正确迁移 $M$1 这类绝对引用，会导致公式指错列。
        date_row_abs = f"${get_column_letter(date_start_col)}$2:${get_column_letter(last_date_col)}$2"
        stopline_criterion_abs = f"${get_column_letter(stopline_col)}$1"
        diff_row_range = (
            f"${get_column_letter(date_start_col)}{base_row + BALANCE_DIFF_OFFSET}"
            f":${get_column_letter(last_date_col)}{base_row + BALANCE_DIFF_OFFSET}"
        )
        worksheet.cell(
            base_row,
            stopline_col,
            f"=IFERROR(SUMIF({date_row_abs},{stopline_criterion_abs},{diff_row_range}),\"\")",
        )
        for row_offset in range(1, BALANCE_BLOCK_SIZE):
            current_row = base_row + row_offset
            worksheet.cell(current_row, 10, f"=J{base_row}")

            # 行块内其余行复用物料首行的停线预警值，避免使用模板翻译导致的列漂移
            worksheet.cell(
                current_row,
                stopline_col,
                f"={get_column_letter(stopline_col)}{base_row}",
            )

        worksheet.cell(
            base_row + 1,
            missing_count_col,
            build_purchase_reply_missing_flag_formula(
                row_index=base_row + 1,
                row_label_col=row_label_col,
                date_start_col=date_start_col,
                last_date_col=last_date_col,
            ),
        )
        for date_offset, header_date in enumerate(date_headers, start=date_start_col):
            need_value = record.get(header_date, 0)
            worksheet.cell(base_row, date_offset, need_value)
            reply_value = 0
            if reply_overrides is not None:
                reply_value = reply_overrides.get((material_code, header_date), 0)
            purchase_reply_cell = worksheet.cell(base_row + 1, date_offset, reply_value)
            purchase_reply_cell.fill = yellow_fill
            header_ref = f"{get_column_letter(date_offset)}$2"
            arrival_plan_formula = "+".join(
                f'SUMIFS(到货!$K:$K,到货!$I:$I,"{code}",到货!$G:$G,{header_ref})'
                for code in material_codes_for_supply
            ) or "0"
            arrival_qty_formula = "+".join(
                f'SUMIFS(到货!$L:$L,到货!$I:$I,"{code}",到货!$G:$G,{header_ref})'
                for code in material_codes_for_supply
            ) or "0"
            usage_qty_formula = "+".join(
                (
                    f'SUMIFS(\'{USAGE_SHEET_NAME}\'!$L:$L,\'{USAGE_SHEET_NAME}\'!$G:$G,"{code}",'
                    f'\'{USAGE_SHEET_NAME}\'!$D:$D,"*出库*",'
                    f'\'{USAGE_SHEET_NAME}\'!$B:$B,">="&MAX({header_ref},\'{USAGE_START_DATETIME_SHEET_NAME}\'!{USAGE_START_DATETIME_CELL}),'
                    f'\'{USAGE_SHEET_NAME}\'!$B:$B,"<"&{header_ref}+1)'
                )
                for code in material_codes_for_supply
            ) or "0"
            worksheet.cell(
                base_row + BALANCE_ARRIVAL_PLAN_OFFSET,
                date_offset,
                f"={arrival_plan_formula}",
            )
            worksheet.cell(
                base_row + BALANCE_ARRIVAL_QTY_OFFSET,
                date_offset,
                f"={arrival_qty_formula}",
            )
            worksheet.cell(
                base_row + BALANCE_USAGE_OFFSET,
                date_offset,
                f"={usage_qty_formula}",
            )
            if date_offset == date_start_col:
                generated_inventory_terms = [numeric_formula_constant(record.get("实时库存", 0))]
                if str(record.get("替代1", "") or "").strip():
                    generated_inventory_terms.append(numeric_formula_constant(record.get("替代1库存", 0)))
                if str(record.get("替代2", "") or "").strip():
                    generated_inventory_terms.append(numeric_formula_constant(record.get("替代2库存", 0)))
                generated_inventory_ref = (
                    generated_inventory_terms[0]
                    if len(generated_inventory_terms) == 1
                    else f"SUM({','.join(generated_inventory_terms)})"
                )
                current_col = get_column_letter(date_offset)
                worksheet.cell(
                    base_row + BALANCE_DIFF_OFFSET,
                    date_offset,
                    (
                        f"={generated_inventory_ref}"
                        f"+{current_col}{base_row + BALANCE_REPLY_OFFSET}"
                        f"-{current_col}{base_row}"
                        f"-{current_col}{base_row + BALANCE_USAGE_OFFSET}"
                    ),
                )
            else:
                prev_col = get_column_letter(date_offset - 1)
                current_col = get_column_letter(date_offset)
                worksheet.cell(
                    base_row + BALANCE_DIFF_OFFSET,
                    date_offset,
                    (
                        f"={prev_col}{base_row + BALANCE_DIFF_OFFSET}"
                        f"+{current_col}{base_row + BALANCE_REPLY_OFFSET}"
                        f"-{current_col}{base_row}"
                        f"-{current_col}{base_row + BALANCE_USAGE_OFFSET}"
                    ),
                )

    if date_headers:
        apply_date_column_formats(
            worksheet,
            date_start_col=date_start_col,
            last_date_col=last_date_col,
            max_row=max(target_max_row, worksheet.max_row),
            template_column_format=template_date_column_format,
        )
        for material_index in range(len(purchase_view_df)):
            reply_row = 3 + material_index * BALANCE_BLOCK_SIZE + BALANCE_REPLY_OFFSET
            for date_offset in range(date_start_col, last_date_col + 1):
                worksheet.cell(reply_row, date_offset).fill = yellow_fill

    # Risk formula must stay in column O; do not reuse this column for arrival-status text.
    for material_index in range(len(purchase_view_df)):
        base_row = 3 + material_index * BALANCE_BLOCK_SIZE
        risk_formula_cell = worksheet.cell(base_row, project_risk_col, build_dynamic_risk_formula(base_row))
        risk_formula_cell.number_format = "General"

    worksheet.auto_filter.ref = f"A2:{get_column_letter(missing_count_col)}{worksheet.max_row}"
    try:
        worksheet.auto_filter.filterColumn = []
        worksheet.auto_filter.add_filter_column(0, ["1"])
    except Exception:
        pass
    # 注意：调用方解构使用了 reply_flag_col 和 missing_count_col；reply_unsatisfied_col 作为中间列由下方一起返回
    if date_headers and worksheet.max_row >= 3:
        diff_range = f"{get_column_letter(date_start_col)}3:{get_column_letter(last_date_col)}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            diff_range,
            FormulaRule(
                formula=[f'AND(${get_column_letter(row_label_col)}3="差异",{get_column_letter(date_start_col)}3<0)'],
                stopIfTrue=False,
                fill=diff_negative_fill,
                font=diff_negative_font,
            ),
        )
    return missing_count_col


def ensure_workbook(template_path: Path | None) -> Workbook:
    if template_path and template_path.exists():
        return load_workbook(template_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "供需平衡"
    workbook.create_sheet("缺料表")
    workbook.create_sheet("到货")
    workbook.create_sheet(ARRIVAL_RECORD_SHEET_NAME)
    workbook.create_sheet(USAGE_SHEET_NAME)
    workbook.create_sheet("生产计划")
    workbook.create_sheet("补排产建议")
    workbook.create_sheet(UPPER_EXPANSION_CHECK_SHEET_NAME)
    workbook.create_sheet(PROJECT_DETAIL_SHEET_NAME)
    workbook.create_sheet(BALANCE_HIERARCHY_FILTER_SHEET_NAME)
    workbook.create_sheet(NEAR_TERM_SHORTAGE_SHEET_NAME)
    if ENABLE_CONFIG_SUPPLEMENT_PLAN:
        workbook.create_sheet(CONFIG_SUPPLEMENT_PLAN_SHEET_NAME)
    workbook.create_sheet(PURCHASE_REPLY_SUMMARY_SHEET_NAME)
    sheet.freeze_panes = "AD3"
    return workbook


def prepare_manual_arrivals_sheet(worksheet) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for index, header in enumerate(MANUAL_ARRIVALS_HEADERS, start=1):
        worksheet.cell(1, index, header)


def prepare_usage_sheet(worksheet, usage_df: pd.DataFrame, usage_start_datetime: datetime | None = None) -> None:
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for index, header in enumerate(USAGE_HEADERS, start=1):
        worksheet.cell(1, index, header)
    if not usage_df.empty:
        write_dataframe_sheet(worksheet, usage_df[USAGE_HEADERS], header_row=1)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(USAGE_HEADERS))}{max(worksheet.max_row, 1)}"
    widths = {
        "A": 34,
        "B": 20,
        "C": 24,
        "D": 10,
        "E": 14,
        "F": 28,
        "G": 18,
        "H": 26,
        "I": 8,
        "J": 12,
        "K": 24,
        "L": 12,
        "M": 12,
        "N": 8,
        "O": 18,
        "P": 18,
        "Q": 40,
        "R": 34,
        "S": 18,
        "T": 12,
        "U": 22,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1][: len(USAGE_HEADERS)]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        row[1].number_format = "yyyy-mm-dd hh:mm:ss"
        row[11].number_format = "#,##0.00"
        row[18].number_format = "yyyy-mm-dd hh:mm:ss"


def write_usage_start_datetime(worksheet, usage_start_datetime: datetime | None = None) -> None:
    header_cell = worksheet.cell(1, USAGE_START_DATETIME_COL)
    value_cell = worksheet.cell(2, USAGE_START_DATETIME_COL)
    header_cell.value = "领用起算时间"
    start_datetime = usage_start_datetime or datetime.now()
    value_cell.value = start_datetime
    value_cell.number_format = "yyyy-mm-dd hh:mm:ss"
    worksheet.column_dimensions[get_column_letter(USAGE_START_DATETIME_COL)].width = 22
    header_cell.fill = PatternFill("solid", fgColor="E2F0D9")
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    cutoff_header_cell = worksheet.cell(1, INVENTORY_CUTOFF_DATETIME_COL)
    cutoff_value_cell = worksheet.cell(2, INVENTORY_CUTOFF_DATETIME_COL)
    cutoff_header_cell.value = "库存截止时间"
    cutoff_value_cell.value = datetime.combine(start_datetime.date() + timedelta(days=1), datetime.min.time())
    cutoff_value_cell.number_format = "yyyy-mm-dd hh:mm:ss"
    worksheet.column_dimensions[get_column_letter(INVENTORY_CUTOFF_DATETIME_COL)].width = 22
    cutoff_header_cell.fill = PatternFill("solid", fgColor="E2F0D9")
    cutoff_header_cell.font = Font(bold=True)
    cutoff_header_cell.alignment = Alignment(horizontal="center", vertical="center")


def write_risk_config(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(bold=True)
    configs = [
        (4, "风险参数", "天数"),
        (5, "极高：无采购答复观察天数", 3),
        (6, "高：滚动缺口观察天数", 7),
        (7, "较高：滚动缺口观察天数", 14),
        (8, "低：库存覆盖天数", 14),
    ]
    for row_index, label, value in configs:
        label_cell = worksheet.cell(row_index, RISK_CONFIG_LABEL_COL, label)
        value_cell = worksheet.cell(row_index, RISK_CONFIG_VALUE_COL, value)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_index == 4:
            label_cell.fill = header_fill
            value_cell.fill = header_fill
            label_cell.font = header_font
            value_cell.font = header_font
        elif isinstance(value, int):
            value_cell.number_format = "0"
    worksheet.column_dimensions[get_column_letter(RISK_CONFIG_LABEL_COL)].width = 24
    worksheet.column_dimensions[get_column_letter(RISK_CONFIG_VALUE_COL)].width = 10



def export_workbook(
    result: PipelineResult,
    output_path: Path,
    template_path: Path | None,
    *,
    reply_overrides: dict[tuple[str, date], object] | None = None,
    progress_callback: ProgressCallback | None = None,
) -> None:
    emit_progress(progress_callback, "步骤 6/6：写入 Excel 文件...")
    if output_path.exists():
        try:
            with output_path.open("a+b"):
                pass
        except OSError as exc:
            raise WorkbookInputError(f"输出文件可能正被 WPS/Excel 占用，请先关闭后重试: {output_path}") from exc
    emit_progress(progress_callback, "步骤 6/6：打开输出模板...")
    workbook = ensure_workbook(template_path)
    workbook.calculation.calcId = 0
    workbook.calculation.calcMode = "auto"
    workbook.calculation.calcOnSave = True
    workbook.calculation.calcCompleted = False
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    if CALCULATION_EXPLANATION_SHEET_NAME in workbook.sheetnames:
        workbook.remove(workbook[CALCULATION_EXPLANATION_SHEET_NAME])
    if "供需平衡" not in workbook.sheetnames:
        workbook.create_sheet("供需平衡", 0)
    if "缺料表" not in workbook.sheetnames:
        workbook.create_sheet("缺料表")
    if "到货" not in workbook.sheetnames:
        workbook.create_sheet("到货")
    if ARRIVAL_RECORD_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(ARRIVAL_RECORD_SHEET_NAME)
    if USAGE_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(USAGE_SHEET_NAME)
    if "生产计划" not in workbook.sheetnames:
        workbook.create_sheet("生产计划")
    if "补排产建议" not in workbook.sheetnames:
        workbook.create_sheet("补排产建议")
    if UPPER_EXPANSION_CHECK_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(UPPER_EXPANSION_CHECK_SHEET_NAME)
    if PROJECT_DETAIL_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(PROJECT_DETAIL_SHEET_NAME)
    if BALANCE_HIERARCHY_FILTER_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(BALANCE_HIERARCHY_FILTER_SHEET_NAME)
    if NEAR_TERM_SHORTAGE_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(NEAR_TERM_SHORTAGE_SHEET_NAME)
    if ENABLE_CONFIG_SUPPLEMENT_PLAN and CONFIG_SUPPLEMENT_PLAN_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(CONFIG_SUPPLEMENT_PLAN_SHEET_NAME)
    if PURCHASE_REPLY_SUMMARY_SHEET_NAME not in workbook.sheetnames:
        workbook.create_sheet(PURCHASE_REPLY_SUMMARY_SHEET_NAME)
    for sheet_name in ROLE_QUALITY_SHEET_NAMES:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    shortage_sheet = workbook["缺料表"]
    arrivals_sheet = workbook["到货"]
    arrival_record_sheet = workbook[ARRIVAL_RECORD_SHEET_NAME]
    usage_sheet = workbook[USAGE_SHEET_NAME]
    production_sheet = workbook["生产计划"]
    balance_sheet = workbook["供需平衡"]
    suggestion_sheet = workbook["补排产建议"]
    upper_expansion_sheet = workbook[UPPER_EXPANSION_CHECK_SHEET_NAME]
    project_detail_sheet = workbook[PROJECT_DETAIL_SHEET_NAME]
    hierarchy_filter_sheet = workbook[BALANCE_HIERARCHY_FILTER_SHEET_NAME]
    near_term_shortage_sheet = workbook[NEAR_TERM_SHORTAGE_SHEET_NAME]
    config_supplement_plan_sheet = workbook[CONFIG_SUPPLEMENT_PLAN_SHEET_NAME] if ENABLE_CONFIG_SUPPLEMENT_PLAN else None
    purchase_reply_summary_sheet = workbook[PURCHASE_REPLY_SUMMARY_SHEET_NAME]
    emit_progress(progress_callback, "步骤 6/6：整理项目明细...")
    project_detail_df = build_project_detail_df(result.shortage_df)

    emit_progress(progress_callback, f"步骤 6/6：写入缺料表 {len(result.exported_shortage_df)} 行...")
    write_dataframe_sheet(shortage_sheet, result.exported_shortage_df, header_row=1)
    emit_progress(progress_callback, f"步骤 6/6：写入生产计划 {len(result.production_plan_df)} 行...")
    write_dataframe_sheet(production_sheet, result.production_plan_df, header_row=1)
    write_usage_start_datetime(production_sheet, result.usage_start_datetime)
    write_risk_config(production_sheet)
    emit_progress(progress_callback, f"步骤 6/6：写入补排产建议 {len(result.suggestion_df)} 行...")
    write_suggestion_sheet_preserving_history(suggestion_sheet, result.suggestion_df)
    emit_progress(progress_callback, f"步骤 6/6：写入上层展开校验 {len(result.upper_expansion_df)} 行...")
    write_dataframe_sheet(upper_expansion_sheet, result.upper_expansion_df, header_row=1)
    emit_progress(progress_callback, f"步骤 6/6：写入项目明细 {len(project_detail_df)} 行...")
    write_dataframe_sheet(project_detail_sheet, project_detail_df, header_row=1)
    emit_progress(progress_callback, f"步骤 6/6：写入 BOM 层级筛选 {len(result.hierarchy_filter_df)} 行...")
    write_dataframe_sheet(hierarchy_filter_sheet, result.hierarchy_filter_df, header_row=1)
    hierarchy_filter_sheet.sheet_state = "hidden"
    emit_progress(progress_callback, f"步骤 6/6：写入近三天排产缺料 {len(result.near_term_shortage_df)} 行...")
    write_dataframe_sheet(near_term_shortage_sheet, result.near_term_shortage_df, header_row=1)
    format_near_term_shortage_sheet(near_term_shortage_sheet, result.near_term_shortage_df)
    if ENABLE_CONFIG_SUPPLEMENT_PLAN and config_supplement_plan_sheet is not None:
        emit_progress(progress_callback, f"步骤 6/6：写入配置补排产 {len(result.config_supplement_plan_df)} 行...")
        write_dataframe_sheet(config_supplement_plan_sheet, result.config_supplement_plan_df, header_row=1)
        format_config_supplement_plan_sheet(config_supplement_plan_sheet, result.config_supplement_plan_df)
    format_project_detail_sheet(project_detail_sheet, project_detail_df)
    prepare_manual_arrivals_sheet(arrivals_sheet)
    emit_progress(progress_callback, f"步骤 6/6：写入到货记录 {len(result.arrival_record_df)} 行...")
    reset_generated_sheet(arrival_record_sheet)
    write_dataframe_sheet(arrival_record_sheet, result.arrival_record_df, header_row=1)
    format_arrival_record_sheet(arrival_record_sheet, result.arrival_record_df)
    prepare_usage_sheet(usage_sheet, result.usage_df)
    emit_progress(progress_callback, f"步骤 6/6：重建供需平衡 {len(result.purchase_view_df)} 项...")
    missing_count_col = build_balance_sheet(
        balance_sheet,
        result.purchase_view_df,
        result.shortage_df,
        result.date_headers,
        project_detail_df=project_detail_df,
        reply_overrides=reply_overrides,
    )
    rebuild_purchase_reply_summary_sheet(
        purchase_reply_summary_sheet,
        result.purchase_view_df,
        balance_sheet_name="供需平衡",
        balance_buyer_col=7,
        balance_missing_count_col=missing_count_col,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    emit_progress(progress_callback, "步骤 6/6：保存 Excel 文件...")
    workbook.save(output_path)
    emit_progress(progress_callback, "步骤 6/6：Excel 文件写入完成")


def run_pipeline(
    input_path: Path,
    output_path: Path,
    template_path: Path | None = None,
    *,
    external_bom_df: pd.DataFrame | None = None,
    carry_forward_paths: Iterable[Path] | None = None,
    reply_cutoff_date: date | None = None,
    apply_suggestion_exclusions: bool = True,
    usage_flow_path: Path | None = None,
    arrival_record_path: Path | None = None,
    progress_callback: ProgressCallback | None = None,
) -> PipelineResult:
    try:
        usage_start_datetime = datetime.fromtimestamp(input_path.stat().st_mtime)
    except OSError:
        usage_start_datetime = datetime.now()
    emit_progress(progress_callback, "步骤 1/6：读取输入数据...")
    frames = read_workbook_tables(input_path, external_bom_df=external_bom_df)
    substitute_rules = read_substitute_rules(input_path)
    excluded_codes = read_excluded_material_codes(input_path)
    usage_exclusion_pairs = parent_child_exclusion_pairs(
        frames.get(PARENT_CHILD_USAGE_EXCLUSION_SHEET_NAME)
    )
    suggestion_excluded_codes = excluded_codes if apply_suggestion_exclusions else set()

    emit_progress(progress_callback, "步骤 2/6：计算缺料明细...")
    if usage_exclusion_pairs:
        emit_progress(
            progress_callback,
            f"步骤 2/6：已应用母料用量排除规则 {len(usage_exclusion_pairs)} 条",
        )
    emit_progress(progress_callback, "步骤 2/6：展开BOM并计算缺料明细...")
    raw_shortage_df = build_shortage_df(frames, substitute_rules)
    production_orders_key = next(iter(frames))
    emit_progress(progress_callback, "步骤 2/6：整理生产计划...")
    production_plan_df = build_production_plan_df(frames[production_orders_key], frames["BOM"])
    emit_progress(progress_callback, "步骤 2/6：生成上层展开校验...")
    upper_expansion_df = build_upper_expansion_check_df(
        production_plan_df,
        frames["BOM"],
        usage_exclusion_pairs,
    )
    emit_progress(progress_callback, "步骤 2/6：生成BOM层级筛选...")
    hierarchy_filter_df = build_balance_hierarchy_filter_df(
        production_plan_df,
        frames["BOM"],
        usage_exclusion_pairs,
    )

    emit_progress(progress_callback, "步骤 3/6：生成补排产建议...")
    suggestion_df, suggestion_detail_df = build_subassembly_suggestions_v2(production_plan_df, frames["BOM"])
    suggestion_df = filter_suggestion_data(suggestion_df, suggestion_excluded_codes)
    suggestion_detail_df = filter_suggestion_detail_data_v2(suggestion_detail_df, suggestion_excluded_codes)
    suggestion_metadata = build_suggestion_metadata_v2(suggestion_detail_df)

    emit_progress(progress_callback, "步骤 4/6：汇总供需平衡数据...")
    purchase_view_df, date_headers = build_purchase_view_df_v2(
        raw_shortage_df,
        frames,
        substitute_rules,
        suggestion_metadata,
    )
    purchase_view_df, filtered_shortage_df = filter_balance_data(
        purchase_view_df,
        raw_shortage_df,
        excluded_codes,
    )
    near_term_shortage_df = build_near_term_production_shortage_df(
        filtered_shortage_df,
        frames,
    )
    if ENABLE_CONFIG_SUPPLEMENT_PLAN:
        config_supplement_plan_df = build_config_supplement_plan_df(
            production_plan_df,
            frames,
        )
    else:
        config_supplement_plan_df = pd.DataFrame()
    if usage_flow_path:
        emit_progress(progress_callback, "步骤 4/6：读取库存流水领用记录...")
    usage_df = build_usage_df_from_inventory_flow(usage_flow_path, start_datetime=usage_start_datetime)
    if usage_flow_path:
        emit_progress(progress_callback, f"步骤 4/6：库存流水领用记录 {len(usage_df)} 行")
    if arrival_record_path:
        emit_progress(progress_callback, "步骤 4/6：读取到货记录...")
    arrival_record_df = read_arrival_record_dataframe(arrival_record_path)
    if arrival_record_path:
        emit_progress(progress_callback, f"步骤 4/6：到货记录 {len(arrival_record_df)} 行")
    date_headers = date_window_from_shortage(filtered_shortage_df)
    reply_overrides: dict[tuple[str, date], object] = {}
    carry_stats = {"file_count": 0, "cell_count": 0, "material_count": 0}
    carried_remark_stats = {"file_count": 0, "remark_count": 0, "material_count": 0}
    if carry_forward_paths:
        emit_progress(progress_callback, "步骤 5/6：读取旧平衡表采购答复和备注...")
        reply_overrides, carry_stats = load_carried_purchase_replies(
            carry_forward_paths,
            cutoff_date=reply_cutoff_date or date.today(),
        )
        carried_remark_map, carried_remark_stats = load_carried_balance_remarks(carry_forward_paths)
        purchase_view_df = apply_carried_balance_remarks(purchase_view_df, carried_remark_map)
        emit_progress(
            progress_callback,
            (
                f"步骤 5/6：旧平衡表读取完成，带入采购答复 {carry_stats['cell_count']} 格，"
                f"备注 {carried_remark_stats['remark_count']} 条"
            ),
        )
    else:
        emit_progress(progress_callback, "步骤 5/6：未选择旧平衡表，跳过采购答复回填")
    result = PipelineResult(
        shortage_df=filtered_shortage_df,
        exported_shortage_df=raw_shortage_df,
        purchase_view_df=purchase_view_df,
        production_plan_df=production_plan_df,
        upper_expansion_df=upper_expansion_df,
        hierarchy_filter_df=hierarchy_filter_df,
        near_term_shortage_df=near_term_shortage_df,
        config_supplement_plan_df=config_supplement_plan_df,
        usage_df=usage_df,
        arrival_record_df=arrival_record_df,
        date_headers=date_headers,
        suggestion_df=suggestion_df,
        carried_reply_cell_count=carry_stats["cell_count"],
        carried_reply_material_count=carry_stats["material_count"],
        carried_reply_file_count=carry_stats["file_count"],
        usage_start_datetime=usage_start_datetime,
    )
    export_workbook(
        result,
        output_path,
        template_path,
        reply_overrides=reply_overrides,
        progress_callback=progress_callback,
    )
    emit_progress(progress_callback, "平衡表生成完成")
    return result

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据标准输入工作簿生成 MRP 缺料结果和平衡表。")
    parser.add_argument("--input", required=True, help="标准输入工作簿路径")
    parser.add_argument("--output", required=True, help="输出工作簿路径")
    parser.add_argument("--template", help="供需平衡模板路径，可选")
    parser.add_argument("--carry-forward", nargs="*", help="历史答复结转文件路径，可传多个")
    parser.add_argument("--usage-flow", help="库存流水记录路径，可选；只统计生成时间之后的出库作为领用")
    parser.add_argument("--arrival-record", help="到货记录路径，可选；会原样写入输出平衡表的到货记录页")
    parser.add_argument("--reply-after", help="只保留该日期之后的答复记录，格式 YYYY-MM-DD")
    parser.add_argument("--ignore-suggestion-exclusions", action="store_true", help="忽略采购建议排除规则，输出全部建议")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve() if args.template else None
    usage_flow_path = Path(args.usage_flow).expanduser().resolve() if args.usage_flow else None
    arrival_record_path = Path(args.arrival_record).expanduser().resolve() if args.arrival_record else None
    carry_forward_paths = [Path(path).expanduser().resolve() for path in (args.carry_forward or [])]
    reply_cutoff_date = None
    if args.reply_after:
        reply_cutoff_date = datetime.strptime(args.reply_after, "%Y-%m-%d").date()
    run_pipeline(
        input_path,
        output_path,
        template_path,
        carry_forward_paths=carry_forward_paths,
        reply_cutoff_date=reply_cutoff_date,
        apply_suggestion_exclusions=not args.ignore_suggestion_exclusions,
        usage_flow_path=usage_flow_path,
        arrival_record_path=arrival_record_path,
    )
    print(f"生成完成: {output_path}")
    return 0
